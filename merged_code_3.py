# Auto-merged batch 3/4
# Total files in this batch: 59



#==============================================================================
# File: ch4_01_rules_for_outliers.py
#==============================================================================

# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
sys.path.append("../")

import pandas as pd
from utils import data_utils

def rule_evaluate(selected_df, total_df, target, rate=0.15, amount=10000):
    """
    :param selected_df: 子特征列表
    :param total_df: 特征宽表
    :param target: 目标变量
    :param rate: 息费（%）
    :param amount: 平均每笔借款金额
    :return:
    """
    # 命中规则的子群体指标统计
    hit_size = selected_df.shape[0]
    hit_bad_size = selected_df[target].sum()
    hit_bad_rate = selected_df[target].mean()
    # 总体指标统计
    total_size = total_df.shape[0]
    total_bad_size = total_df[target].sum()
    total_bad_rate = total_df[target].mean()
    # 命中率
    hit_rate = hit_size / total_size
    # 提升度
    lift = hit_bad_rate / total_bad_rate
    # 收益
    profit = hit_bad_size * amount - (hit_size - hit_bad_size) * rate * amount
    res = [total_size, total_bad_size, total_bad_rate,
           hit_rate, hit_size, hit_bad_size, hit_bad_rate, lift, profit]
    return res


def rule_discover(data_df, var, target, rule_term, rate=0.15, amount=10000):
    """
    :param data_df: 特征宽表
    :param var: 特征名称
    :param target: 目标变量
    :param rule_term: 分位数列表或规则条件
    :param rate: 息费（%）
    :param amount: 平均每笔借款金额
    :return:
    """
    res_list = []
    if rule_term is None:
        rule_term = [0.005, 0.01, 0.02, 0.05, 0.95, 0.98, 0.99, 0.995]
    if isinstance(rule_term, list):
        for q in rule_term:
            threshold = data_df[var].quantile(q).round(2)
            if q < 0.5:
                temp = data_df.query("`{0}` <= @threshold".format(var))
                rule = "<= {0}".format(threshold)
            else:
                temp = data_df.query("`{0}` >= @threshold".format(var))
                rule = ">= {0}".format(threshold)
            res = rule_evaluate(temp, data_df, target, rate, amount)
            res_list.append([var, rule] + res)
    else:
        temp = data_df.query("`{0}` {1}".format(var, rule_term))
        rule = rule_term
        res = rule_evaluate(temp, data_df, target, rate, amount)
        res_list.append([var, rule] + res)
    columns = ['var', 'rule', 'total_size', 'total_bad_size', 'total_bad_rate',
               'hit_rate', 'hit_size', 'hit_bad_size', 'hit_bad_rate', 'lift',
               'profit']
    result_df = pd.DataFrame(res_list, columns=columns)
    return result_df


if __name__ == '__main__':
    # 数据读入
    german_credit_data = data_utils.get_data()
    german_credit_data.loc[german_credit_data.sample(
        frac=0.2, random_state=0).index, 'sample_set'] = 'Train'
    german_credit_data['sample_set'].fillna('OOT', inplace=True)
    # 使用分位数列表构建规则集
    rule_table = rule_discover(data_df=german_credit_data, var='credit.amount',
                               target='creditability',
                               rule_term=[0.005, 0.01, 0.02, 0.05, 0.95, 0.98, 0.99, 0.995])
    print(rule_table)
    # 规则效果评估
    rule_analyze = german_credit_data.groupby('sample_set').apply(
        lambda x: rule_discover(data_df=x, var='credit.amount',
                                target='creditability', rule_term='>12366.0'))
    print(rule_analyze)



#==============================================================================
# File: ch4_02_rules_for_decisiontree.py
#==============================================================================

# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
sys.path.append("../")

import sklearn.tree as st
import graphviz
from utils import data_utils


def decision_tree_resolve(train_x, train_y, class_names=None, max_depth=3, fig_path=''):
    """
    基于决策树可视化
    :param train_x: data of train
    :param train_y: data of y
    :param class_names:  标签名称
    :param max_depth: 树最大深度
    :param fig_path: 图片路径和名称
    :return:
    """
    if class_names is None:
        class_names = ['good', 'bad']
    clf = st.DecisionTreeClassifier(max_depth=max_depth,
                                    min_samples_leaf=0.01,
                                    min_samples_split=0.01,
                                    criterion='gini',
                                    splitter='best',
                                    max_features=None)
    clf = clf.fit(train_x, train_y)

    # 比例图
    dot_data = st.export_graphviz(clf, out_file=None,
                                  feature_names=train_x.columns.tolist(),
                                  class_names=class_names,
                                  filled=True,
                                  rounded=True,
                                  node_ids=True,
                                  special_characters=True,
                                  proportion=True,
                                  leaves_parallel=True)
    graph = graphviz.Source(dot_data, filename=fig_path)
    return graph


# 加载数据
german_credit_data = data_utils.get_data()

# 构造数据集
X = german_credit_data[data_utils.numeric_cols].copy()
y = german_credit_data['creditability']

graph = decision_tree_resolve(X, y, fig_path='data/tree')
graph.view()

# 转化为规则
X['node_5'] = X.apply(lambda x: 1 if x['duration.in.month'] <= 34.5 and x['credit.amount'] > 8630.5 else 0, axis=1)
X['node_9'] = X.apply(
    lambda x: 1 if x['duration.in.month'] > 34.5 and x['age.in.years'] <= 29.5 and x['credit.amount'] > 4100.0 else 0,
    axis=1)
X['node_12'] = X.apply(lambda x: 1 if x['duration.in.month'] > 34.5 and x['age.in.years'] > 56.5 else 0, axis=1)



#==============================================================================
# File: ch4_03_rules_for_isolationforest.py
#==============================================================================

# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
sys.path.append("../")

import numpy as np
from chapter4.ch4_01_rules_for_outliers import rule_discover
from pyod.models.iforest import IForest
from utils import data_utils

# 加载数据
german_credit_data = data_utils.get_data()

# 构造数据集
X = german_credit_data[data_utils.numeric_cols]
y = german_credit_data['creditability']

# 初始化模型
clf = IForest(behaviour='new', bootstrap=False, contamination=0.1, max_features=1.0, max_samples='auto', n_estimators=500, random_state=20, verbose=0)

# 训练模型  
clf.fit(X)

# 预测结果  
german_credit_data['out_pred'] = clf.predict_proba(X)[:, 1]
# 将预测概率大于0.7以上的设为异常值  
german_credit_data['iforest_rule'] = np.where(german_credit_data['out_pred'] > 0.7, 1, 0)

# 效果评估  
rule_iforest = rule_discover(data_df=german_credit_data, var='iforest_rule', target='creditability', rule_term='==1')
print("孤立森林评估结果: \n", rule_iforest.T)



#==============================================================================
# File: ch4_04_modelstrategy_for_optimization.py
#==============================================================================

# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
sys.path.append("../")

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from numpy import polyfit, poly1d
from sklearn.metrics import r2_score
from scipy.optimize import minimize


def calculate_pass_loss_decile(score_series, y_series):
    """
    模型分取值变化时通过率与坏账率关系
    :param score_series: 模型分
    :param y_series: Y标签
    :return:  
    """
    decile_df = pd.crosstab(score_series, y_series).rename(columns={0: 'N_nonEvent', 1: 'N_Event'})
    decile_df.loc[:, 'N_sample'] = score_series.value_counts()

    decile_df.loc[:, 'EventRate'] = decile_df.N_Event * 1.0 / decile_df.N_sample
    decile_df.loc[:, 'BadPct'] = decile_df.N_Event * 1.0 / sum(decile_df.N_Event)
    decile_df.loc[:, 'GoodPct'] = decile_df.N_nonEvent * 1.0 / sum(decile_df.N_nonEvent)
    decile_df.loc[:, 'CumBadPct'] = decile_df.BadPct.cumsum()
    decile_df.loc[:, 'CumGoodPct'] = decile_df.GoodPct.cumsum()

    decile_df = decile_df.sort_index(ascending=False)
    decile_df.loc[:, 'ApprovalRate'] = decile_df.N_sample.cumsum() / decile_df.N_sample.sum()
    decile_df.loc[:, 'ApprovedEventRate'] = decile_df.N_Event.cumsum() / decile_df.N_sample.cumsum()
    decile_df = decile_df.sort_index(ascending=True)
    return decile_df


def poly_regression(x_series, y_series, degree, plot=True):
    """
    多项式回归拟合
    :param x_series: x数据
    :param y_series: y数据
    :param degree: 指定多项式次数
    :param plot: 是否作图
    :return:
    """
    coeff = polyfit(x_series, y_series, degree)
    f = poly1d(coeff)
    R2 = r2_score(y_series.values, f(x_series))

    print(f'coef:{coeff},R2: {R2}')

    if plot:
        # 用来正常显示中文标签
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False

        plt.figure(figsize=(10, 5))
        plt.plot(x_series, y_series, 'rx')
        plt.plot(x_series, f(x_series))
        plt.xlabel('通过率', {'size': 15})
        plt.ylabel('坏账率', {'size': 15})
        plt.show()
    return coeff


german_score = pd.read_csv('data/german_score.csv')
german_score.head()

decile_df = calculate_pass_loss_decile(german_score['score'],
                                       german_score['creditability'])
print(decile_df.head())

# 数据准备
x = decile_df['ApprovalRate']
# 逾期率折算为坏账率
y = decile_df['ApprovedEventRate'] / 2.5

poly_coef = poly_regression(x, y, 2, plot=True)
# 坏账率L(x)与通过率x的关系
l_x = poly1d(poly_coef)
print(l_x)


def find_best_approval_rate(x_to_loss_func, score_df):
    """
    定义最优化函数
    坏账率L(x)与通过率x的关系函数
    :param x_to_loss_func: 坏账率与通过率的函数关系
    :param score_df: 模型分与通过率的对应关系，index为模型分，"ApprovalRate"列为对应的通过率
    :return:
    """

    # 定义目标函数，求解最大值即为负的最小值
    def fun(x_array):
        # 其中x_list[0]为通过率x，x_array[1]为对应的坏账率L(x)
        return -10000 * (0.16 * (1 - x_array[1]) - x_array[1]
                         - 30 / (x_array[0] * 0.6) / 10000)

    # eq表示 函数结果等于0 ； ineq 表示 表达式大于等于0， 下面式子1e-6项确保相应变量不等于0或1
    cons = ({'type': 'eq', 'fun': lambda x_array: x_to_loss_func(x_array[0]) - x_array[1]},
            {'type': 'ineq', 'fun': lambda x_array: x_array[0] - 1e-6},
            {'type': 'ineq', 'fun': lambda x_array: x_array[1] - 1e-6},
            {'type': 'ineq', 'fun': lambda x_array: 1 - x_array[0] - 1e-6},
            {'type': 'ineq', 'fun': lambda x_array: 1 - x_array[0] - 1e-6}
            )

    # 设置初始值
    x_base = np.array((0.10, 0.10))
    # 采用SLSQP进行最优化求解
    res = minimize(fun, x_base, method='SLSQP', constraints=cons)
    print('利润最优：', "{:.2f}".format(-res.fun))
    print('最优解对应通过率：', "{:.2%}".format(res.x[0]), '坏账率：', "{:.2%}".format(res.x[1]))
    print("模型分阈值：", score_df[score_df['ApprovalRate'] >= res.x[0]].index.max())
    print('迭代终止是否成功：', res.success)
    print('迭代终止原因：', res.message)


find_best_approval_rate(l_x, decile_df)



#==============================================================================
# File: cli.py
#==============================================================================

"""Console script for autotreemodel."""
import sys
import click


@click.command()
def main(args=None):
    """Console script for autotreemodel."""
    click.echo("Replace this message by putting your code into "
               "autotreemodel.cli.main")
    click.echo("See click documentation at https://click.palletsprojects.com/")
    return 0


if __name__ == "__main__":
    sys.exit(main())  # pragma: no cover



#==============================================================================
# File: code00_Proc00_init_library.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc00_init_library
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## init_library
## 导入必要的分析库

####################################################################
# 导入必要的分析库

####################################################################
import os
import re
import sys
import csv
import json
import time
import pytz
import copy
import random
import datetime

strptime = datetime.datetime.strptime
strftime = datetime.datetime.strftime
from dateutil.relativedelta import relativedelta
from dateutil import rrule

from collections import OrderedDict
from itertools import product
import pickle

import gc
import multiprocessing

import math
import pandas as pd
import numpy as np
import scipy
from scipy.stats import pearsonr, spearmanr, kendalltau

import matplotlib.pyplot as plt
# %matplotlib inline
import seaborn as sns

pd.pandas.set_option("display.max_columns", 100)
pd.pandas.set_option("display.max_rows", 300)

from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import train_test_split

# import xgboost as xgb
# from xgboost.sklearn import XGBClassifier
# from xgboost import plot_tree
# from sklearn.ensemble import GradientBoostingClassifier

# pd.set_option('max_colwidth', 200)
# pd.set_option('display.max_rows', 40)
# pd.set_option('display.max_columns', None)

# import tensorflow as tf
# import tensorflow.keras as keras
# import tensorflow.experimental.numpy as tnp
# import tensorflow_probability as tfp

# import lifelines
# from lifelines import KaplanMeierFitter
# from lifelines import CoxPHFitter
# from lifelines import CoxTimeVaryingFitter

####################################################################
inf = np.inf

####################################################################
plt.style.use({"figure.figsize": [s0*2 for s0 in (5, 3)]})
plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False

sns.set(
    style="whitegrid",
    rc={"figure.figsize": [s0*2 for s0 in (5, 3)]},
    font="SimHei",
)

####################################################################
package_path = r"C:\huangting\长期A卡"
# package_path = r"/media/sf_Desktop"
# package_path = r"C:\Users\luzhidong-alienware\Desktop"
# package_path = r"C:\Users\luzd\Desktop"
# package_path = r"/tf/luzd"
os.chdir(package_path)

####################################################################
from ModelingToolkit.LogisticModel import *
from ModelingToolkit.FeatureEngineering import *
# from ModelingToolkit.BoostrapSampling import *
# from ModelingToolkit.TensorProcessing import *










#==============================================================================
# File: code01_Proc00_DataPreperation.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc00_DataPreperation
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc00_DataPreperation
## 将数据样本进行预处理，整理成样本宽表形式
## 根据建模场景不一样，预处理的程度也不完全一致

####################################################################
# Proc00_DataPreperation
# 将数据样本进行预处理，整理成样本宽表形式

####################################################################
# 数据导入
data_path = r"C:\huangting\海尔项目\建模_lr_dpd3"
# data_path = r"{}/ModelingToolkit/test_data".format(package_path)
# data_path = r"C:/Users/luzhidong-alienware/Desktop/vzoom_data/data"
# data_path = r"C:/Users/luzd/Desktop/data"



####################################################################
# 设置结果输出路径
result_path = r"{}/res_output".format(data_path)


if not os.path.isdir("{}".format(result_path)):
    os.mkdir("{}".format(result_path))

####################################################################
# Proc01_SamplingSplit
if not os.path.isdir("{}/Proc01_SamplingSplit".format(result_path)):
    os.mkdir("{}/Proc01_SamplingSplit".format(result_path))

####################################################################
# Proc02_FeatDesc
if not os.path.isdir("{}/Proc02_FeatDesc".format(result_path)):
    os.mkdir("{}/Proc02_FeatDesc".format(result_path))

####################################################################
# Proc03_FeatFineBinning
if not os.path.isdir("{}/Proc03_FeatFineBinning".format(result_path)):
    os.mkdir("{}/Proc03_FeatFineBinning".format(result_path))

####################################################################
# Proc04_WoeMonotonic
if not os.path.isdir("{}/Proc04_WoeMonotonic".format(result_path)):
    os.mkdir("{}/Proc04_WoeMonotonic".format(result_path))

####################################################################
# Proc05_FeatCoarseBinning_batch
if not os.path.isdir("{}/Proc05_FeatCoarseBinning_batch".format(result_path)):
    os.mkdir("{}/Proc05_FeatCoarseBinning_batch".format(result_path))

####################################################################
# Proc06_Correlation
if not os.path.isdir("{}/Proc06_Correlation".format(result_path)):
    os.mkdir("{}/Proc06_Correlation".format(result_path))

####################################################################
# Proc07_Model_LR_stepwise
if not os.path.isdir("{}/Proc07_Model_LR_stepwise".format(result_path)):
    os.mkdir("{}/Proc07_Model_LR_stepwise".format(result_path))

####################################################################
# Proc07_Model_LR_adj
if not os.path.isdir("{}/Proc07_Model_LR_adj".format(result_path)):
    os.mkdir("{}/Proc07_Model_LR_adj".format(result_path))

####################################################################
# Proc08_Model_Evaluation
if not os.path.isdir("{}/Proc08_Model_Evaluation".format(result_path)):
    os.mkdir("{}/Proc08_Model_Evaluation".format(result_path))

####################################################################
# Proc09_Scorecard_Trans
if not os.path.isdir("{}/Proc09_Scorecard_Trans".format(result_path)):
    os.mkdir("{}/Proc09_Scorecard_Trans".format(result_path))

####################################################################
# Proc90_Feature_Selection
if not os.path.isdir("{}/Proc90_Feature_Selection".format(result_path)):
    os.mkdir("{}/Proc90_Feature_Selection".format(result_path))

####################################################################
# Proc99_Modeling_Report
if not os.path.isdir("{}/Proc99_Modeling_Report".format(result_path)):
    os.mkdir("{}/Proc99_Modeling_Report".format(result_path))




df_dpd1 = pd.read_csv(r"C:\huangting\海尔项目\数据提取\海尔dpd1_mob_2_2023-04-11_2023-09-12_202309141344.csv")
df_dpd1 = df_dpd1.rename(columns = {'y': 'dpd1'})
# (402153, 21)
print(df_dpd1.shape)


# 这里面dpd3有问题，有些是dpd1是1的，在dpd3变成0了，只看下面的df_yx就好，df_yx已修正
df_dpd3 = pd.read_csv(r"C:\huangting\海尔项目\数据提取\海尔dpd3_mob_2_2023-04-11_2023-09-12_202309151426.csv")
df_dpd3 = df_dpd3.rename(columns = {'y': 'dpd3'})
# (402153, 21)
print(df_dpd3.shape)


print(len(set(df_dpd1.auth_order_no)),len(set(df_dpd3.auth_order_no)))



df_dpd = pd.merge(df_dpd1, df_dpd3[['auth_order_no','dpd3']])
# (402153, 22)
print(df_dpd.shape)


print(len(data_all), len(set(data_all.order_no)))
print(len(df_dpd), len(set(df_dpd.auth_order_no)))


df_yx = pd.merge(df_dpd
         , data_all
         , left_on = 'auth_order_no'
         , right_on = 'order_no'
         , how = 'inner'
         )
# (392062, 100)
print(df_yx.shape)

# -1.0    9498
#  0.0    8649
#  1.0    1987
print(df_yx.dpd1.value_counts())
# -1.0    10734
#  0.0     8834
#  1.0      566
print(df_yx.dpd3.value_counts())


# df_yx.to_csv(r"C:\huangting\海尔项目\数据提取\海尔三方数据_带Y.csv", index = False)

df_yx = pd.read_csv(r"C:\huangting\海尔项目\数据提取\海尔三方数据_带Y.csv")
print(df_yx.shape)

# df_yx.loc[(df_yx['dpd1'] == 1)&(df_yx['dpd3'] != 1), 'dpd3'] = -1


df_wb = df_yx.copy()
df_wb = df_wb[df_wb['dpd3'].isin([0,1])].reset_index(drop = True)
# 9215
print(df_wb.shape)


####################################################################
# df_wb = pd.read_csv(r"{}/test_data_for_modeling.csv".format(data_path))
# with open(r"{}/df_wb.pkl".format(data_path), mode="rb") as fr:
#     df_wb = pickle.load(file=fr) \
#         .rename(columns={"nsrsbh": "uid"}) \
#         .reset_index(drop=True)

####################################################################
df_wb["flag"] = df_wb["dpd3"]
df_wb["target_label"] = df_wb["flag"].apply(lambda s0: ("1_bad" if s0==1 else "0_good"))
df_wb["observation_dt_YM"] = df_wb["auth_date"].apply(lambda s0: s0[0:7])
df_wb["observation_dt_YQ"] = df_wb["auth_date"].apply(lambda s0: "{}Q{:02d}".format(s0[0:4], int((float(s0[5:7])-1)//3)+1))

df_wb["auth_month"] = df_wb["auth_date"].apply(lambda x: x[:7])


# 剔除无用字段
for i in ['Unnamed: 0', 'user_id_x', 'isna_cnt']:
    del df_wb[i]


############################################################
# 异常值处理
# -9999999置换为np.NaN
# def _func_trans_nan(data):
#     try:
#         rt = (np.NaN if abs(data)==99999999 else data)
#     except:
#         rt = data
#     return rt
# # df_wb[cols_v] = \
# #     df_wb[cols_v].applymap(lambda s0: _func_trans_nan(data=s0))
# df_wb[[s0 for s0 in df_wb.columns if re.search("^cw_", s0)!=None]] = \
#     df_wb[[s0 for s0 in df_wb.columns if re.search("^cw_", s0)!=None]].applymap(lambda s0: _func_trans_nan(data=s0))


# 异常值处理
# 小于0的，替换为空    
for i in ["三方数据源_baihang_1",
"三方数据源_bairong_14",
"三方数据源_bairong_15",
"三方数据源_duxiaoman_1",
"三方数据源_hangliezhi_1",
"三方数据源_hengpu_4",
"三方数据源_hengpu_5",
"三方数据源_HL_A1",
"三方数据源_HL_A2",
"三方数据源_HL_A3",
"三方数据源_HL_A4",
"三方数据源_rong360_4",
"三方数据源_ruizhi_6",
"三方数据源_xinyongsuanli_1",
"is_cheDai",
"education",
"pboc_djk_edsyl",
"cc_overdue_last_months_l3m",
"loan_overdue_last_months_l9m",
"quae_xcva_p5",
"query_tfau_mg",
"pboc_dkcpxzhzz_zxzl",
"haier_pfk_v1",
"overdue_m1_cnt_l3m_all",
"Loan0020",
"coaa_zbva_xavh_bbvf_n9",
"debts_sgac_mc",
"pboc_education_new",
"haier_pfk_v2",
"rebo_zcva_bfvd_n8",
"pboc_zhzt_zxjl",
"pboc_sycpsyed",
"pboc_overdue_cnt_48mr",
"pboc_quota_amount",
"cc_overdue_last_months_l9m",
"CARD0004",
"pboc_1m_cxcs_dksp",
"pboc_dkywzx_12m_overdue_num",
"pboc_base_education",
"repay_dcal_tmd",
"loan_overdue_last_months_l12m",
"loan_overdue_last_months_l6m",
"pboc_50000_account_cnt",
"INQUIRY0047",
"CARD0009",
"cc_overdue_last_months_l12m",
"pboc_wjqfyjggrxfdkbs",
"debts_sgad_mc",
"pboc_quota_rate_6mr",
"pboc_szjdf",
"pboc_zx_age",
"pboc_mob_mx",
"pboc_3m_xdspcxyy_num",
"repay_dhaa_md",
"pboc_szjdf_xdwz",
"pboc_marital_new",
"is_fangDai",
"result",
"loan_overdue_last_months_l3m",
"pboc_dkcpxzhzz_pjzl",
"overdue_m1_cnt_l6m_all",
"overdue_m2_cnt_l6m_all",
"cur_overdue_amount",
"TYPEaSUMHOUS",
"pboc_is_bh",
"coaa_zbvg_xawd_bbvd_n3",
"cur_overdue_account_cnt",
"AGE",
"pboc_overdue_months",
"pboc_is_house_loan",
"pboc_dkwjfl_new",
"pboc_all_account_cnt",
"pboc_account_cnt",
"CARD0065",
"pboc_12m_ffbjq_num",
"cc_overdue_last_months_l6m",
"Inquiry0023",
"overdue_m2_cnt_l12m_all",
"deaa_zbvg_bbvb_o4",
"quad_p1",
"INQUIRY0064",
"pboc_dkywzlwgrxfdk_num",
"income_izai_tmf",
"digitalScore",
"pboc_is_car_loan",
"pboc_sex_new",
"pboc_rate_12mr",
]:
    df_wb.loc[df_wb[i]<0, i] = np.nan



############################################################
# df_wb = df_wb \
#     .sort_values(by=["uid", "observation_dt"], ascending=[True, True]) \
#     .reset_index(drop=True)
# df_wb = df_wb.reset_index(drop = True)


############################################################
# _cols = df_wb.apply(lambda s0: (s0.name if re.search("(float|int)", s0.dtypes.name)==None else np.NaN), axis=0).dropna().tolist()
# df_wb[_cols] = df_wb[_cols].applymap(lambda s0: (np.NaN if (not pd.isna(s0)) and s0.replace(" ", "")=="" else s0))


############################################################
# 定义字段列
cols_base = [
'auth_order_no',
'auth_date',
'auth_status',
'dpd1',
'min_lending_date',
'dpd3',
'user_id_y',
'order_no',
'create_time',
'dt',
'flag',
'target_label',
'observation_dt_YM',
'observation_dt_YQ',
'auth_month',
# 'data_role'
]

############################################################
cols_v0 = [s0 for s0 in df_wb.columns if s0 not in cols_base]


cols_v = cols_v0
   

df_wb = df_wb[
        cols_base+cols_v0
        # cols_base+cols_v
    ] \
    .reset_index(drop=True)



print(df_wb[cols_v].shape)


















#==============================================================================
# File: code02_Proc01_SamplingSplit.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc01_SamplingSplit
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## 样本抽样与样本集划分方案
## Proc01_SamplingSplit

####################################################################
# Proc01_SamplingSplit
# 样本抽样与样本集划分方案

#################################################################################
# 分布：observation_dt_YM
df_t_stat_obs_dt_YM = df_wb.groupby(by=["observation_dt_YM"]).apply(
    lambda s0: pd.Series({
        "cnt": s0.shape[0],
        "bad_cnt": s0.query("flag==1").shape[0],
        "bad_rate": s0.query("flag==1").shape[0]/s0.shape[0],
    })
).reset_index(drop=False)

df_t_stat_obs_dt_YM.to_clipboard()
df_t_stat_obs_dt_YM

#################################################################################
# 分布：observation_dt_YQ
# df_t_stat_obs_dt_YQ = df_wb.groupby(by=["observation_dt_YQ"]).apply(
#     lambda s0: pd.Series({
#         "cnt": s0.shape[0],
#         "bad_cnt": s0.query("flag==1").shape[0],
#         "bad_rate": s0.query("flag==1").shape[0]/s0.shape[0],
#     })
# ).reset_index(drop=False)
# df_t_stat_obs_dt_YQ.to_clipboard()
# df_t_stat_obs_dt_YQ



#################################################################################
# _df_uid = df_wb[["user_id", "auth_time"]].groupby(by=["user_id"]).apply(
#     lambda s0: pd.Series(dict({
#         "observation_dt_min": s0["auth_time"].min(),
#     }))
# ).reset_index(drop=False)

# #################################################################################
# sample_split_random_seed = 1234567890

# #################################################################################




# _oot_cp_observation_dt = "2022-10-31"
# _cond_oot = (
#     (_df_uid["observation_dt_min"]>=_oot_cp_observation_dt)
# )

# _df_uid_idx_oot = _df_uid \
#     [_cond_oot] \
#     [["user_id", "observation_dt_min"]] \
#     .reset_index(drop=True)

# _df_uid_idx = _df_uid \
#     [-_cond_oot] \
#     [["user_id", "observation_dt_min"]] \
#     .reset_index(drop=True)

# #################################################################################
# # _df_uid_idx_dev, _df_uid_idx_oos = \
# #     train_test_split(
# #         _df_uid_idx,
# #         train_size=0.8,
# #         random_state=sample_split_random_seed,
# #         shuffle=True,
# #     )

# #################################################################################
# _df_uid_idx_train, _df_uid_idx_test = \
#     train_test_split(
#         # _df_uid_idx_dev,
#         _df_uid_idx,
#         train_size=0.75,
#         random_state=sample_split_random_seed,
#         shuffle=True,
#     )

# #################################################################################
# _df_uid_idx_train["data_role"] = "01_train"
# _df_uid_idx_test["data_role"] = "02_test"
# # _df_uid_idx_oos["data_role"] = "03_oos"
# _df_uid_idx_oot["data_role"] = "04_oot"

# df_wb = pd.merge(
#     left=pd.concat(
#         [
#             _df_uid_idx_train,
#             _df_uid_idx_test,
#             # _df_uid_idx_oos,
#             _df_uid_idx_oot,
#         ],
#         ignore_index=True,
#     )[["user_id", "data_role"]],
#     right=df_wb[[s0 for s0 in df_wb.columns if s0!="data_role"]],
#     how="left", left_on=["user_id"], right_on=["user_id"],
# ) \
#     .sort_values(by=["user_id", "auth_time"], ascending=[True, True]) \
#     .reset_index(drop=True)
# cols_base = [s0 for s0 in cols_base if s0 not in ["data_role"]]+["data_role"]

# # #################################################################################
# _df_wb_dev = df_wb.query("data_role in ['01_train', '02_test']").reset_index(drop=True)
# _df_wb_train = df_wb.query("data_role in ['01_train']").reset_index(drop=True)
# _df_wb_test = df_wb.query("data_role in ['02_test']").reset_index(drop=True)
# _df_wb_oos = df_wb.query("data_role in ['03_oos']").reset_index(drop=True)
# _df_wb_oot = df_wb.query("data_role in ['04_oot']").reset_index(drop=True)

df_wb = df_wb.reset_index()

# df_wb['data_role'] = df_wb['auth_time'].apply(lambda x: '04_oot' if x[:7]>='2022-11' and x[:7]<='2022-12'
#                          else '01_train')




# 划分样本
# OOT
df_wb_oot = df_wb[(df_wb['auth_date'] >= '2023-08-01')]
# 965 103
print(df_wb_oot.shape)

oot_index = df_wb_oot.index



# del df_wb['data_role']

df_wb.loc[df_wb.index.isin(oot_index), 'data_role'] = '04_oot'
df_wb['data_role'] = df_wb['data_role'].fillna('01_train')
df_wb['data_role'].value_counts()


_df_wb_dev = df_wb.query("data_role in ['01_train', '02_test']").reset_index(drop=True)
_df_wb_train = df_wb.query("data_role in ['01_train']").reset_index(drop=True)
_df_wb_oot = df_wb.query("data_role in ['04_oot']").reset_index(drop=True)


#################################################################################
df_t_stat_data_role = df_wb.groupby(by=["data_role"]).apply(
    lambda s0: pd.Series({
        "cnt": s0.shape[0],
        "bad_cnt": s0.query("flag==1").shape[0],
        "bad_rate": s0.query("flag==1").shape[0]/s0.shape[0],
    })
).reset_index(drop=False)
# df_t_stat_data_role["comment"] = df_t_stat_data_role["data_role"].apply(
#     lambda s0: (
#         'observation_dt>="{}"'.format(_oot_cp_observation_dt) if s0=="04_oot" else ""
#     )
# )
df_t_stat_data_role.to_clipboard()
df_t_stat_data_role




### （结果导出）

# ####################################################################################
# # to_clipboard
# df_t_stat_data_role.to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YM.xlsx".format(result_path), mode="wb") as fw:
    df_t_stat_obs_dt_YM \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
# with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YQ.xlsx".format(result_path), mode="wb") as fw:
#     df_t_stat_obs_dt_YQ \
#         .to_excel(
#             fw,
#             index=False,
#             sheet_name="data",
#         )
with open(file="{}/Proc01_SamplingSplit/df_t_stat_data_role.xlsx".format(result_path), mode="wb") as fw:
    df_t_stat_data_role \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YM.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_stat_obs_dt_YM, file=fw)
# with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YQ.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_t_stat_obs_dt_YQ, file=fw)
with open(file="{}/Proc01_SamplingSplit/df_t_stat_data_role.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_stat_data_role, file=fw)

####################################################################################
with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YM.pkl".format(result_path), mode="rb") as fr:
    df_t_stat_obs_dt_YM = pickle.load(file=fr)
# with open(file="{}/Proc01_SamplingSplit/df_t_stat_obs_dt_YQ.pkl".format(result_path), mode="rb") as fr:
#     df_t_stat_obs_dt_YQ = pickle.load(file=fr)
with open(file="{}/Proc01_SamplingSplit/df_t_stat_data_role.pkl".format(result_path), mode="rb") as fr:
    df_t_stat_data_role = pickle.load(file=fr)

df_t_stat_data_role


# # df_wb_0 = df_wb.reset_index(drop=True)
# df_wb = df_wb_0.sample(5000, random_state=12345678).reset_index(drop=True)
# _df_wb_dev = df_wb.query("data_role in ['01_train', '02_test']").reset_index(drop=True)






























# #################################################################################
# sample_split_random_seed = 1234567890

# #################################################################################
# _oot_cp_observation_dt = "2020-01-01"
# _cond_oot = (
#     (df_wb["observation_dt"]>=_oot_cp_observation_dt)
# )

# _df_idx_oot = df_wb \
#     [_cond_oot] \
#     [["uid", "observation_dt"]] \
#     .reset_index(drop=True)

# _df_idx = df_wb \
#     [-_cond_oot] \
#     [["uid", "observation_dt"]] \
#     .reset_index(drop=True)

# #################################################################################
# _df_idx_dev, _df_idx_oos = \
#     train_test_split(
#         _df_idx,
#         train_size=0.8,
#         random_state=sample_split_random_seed,
#         shuffle=True,
#     )

# #################################################################################
# _df_idx_train, _df_idx_test = \
#     train_test_split(
#         _df_idx_dev,
#         train_size=0.75,
#         random_state=sample_split_random_seed,
#         shuffle=True,
#     )

# #################################################################################
# _df_idx_train["data_role"] = "01_train"
# _df_idx_test["data_role"] = "02_test"
# _df_idx_oos["data_role"] = "03_oos"
# _df_idx_oot["data_role"] = "04_oot"

# df_wb = pd.merge(
#     left=pd.concat(
#         [
#             _df_idx_train,
#             _df_idx_test,
#             _df_idx_oos,
#             _df_idx_oot,
#         ],
#         ignore_index=True,
#     )[["uid", "observation_dt", "data_role"]],
#     right=df_wb[[s0 for s0 in df_wb.columns if s0!="data_role"]],
#     how="left", left_on=["uid", "observation_dt"], right_on=["uid", "observation_dt"],
# )
# cols_base = [s0 for s0 in cols_base if s0 not in ["data_role"]]+["data_role"]

# # #################################################################################
# _df_wb_dev = df_wb.query("data_role in ['01_train', '02_test']").reset_index(drop=True)
# # _df_wb_train = df_wb.query("data_role in ['01_train']").reset_index(drop=True)
# # _df_wb_test = df_wb.query("data_role in ['02_test']").reset_index(drop=True)
# # _df_wb_oos = df_wb.query("data_role in ['03_oos']").reset_index(drop=True)
# # _df_wb_oot = df_wb.query("data_role in ['04_oot']").reset_index(drop=True)








# 数据导出
_df_wb_train_output = _df_wb_train[['user_id_y', 
                                      'auth_date', 
                                      'flag', 
                                      'data_role',
                                    "三方数据源_hengpu_4",
                                    "INQUIRY0047",
                                    "三方数据源_hangliezhi_1",
                                    "digitalScore",
                                    "CARD0065",
                                    "三方数据源_ruizhi_6",
                                    "query_tfau_mg",
                                    "pboc_education_new",
                                    "income_izai_tmf",
                                    "quad_p1",
                                    "三方数据源_bairong_14",
                                    "loan_overdue_last_months_l12m",
                                    "三方数据源_xinyongsuanli_1",
                                    "debts_sgad_mc",
                                    "repay_dcal_tmd",
                                    "repay_dhaa_md"
                                      ]]
_df_wb_train_output['loan_type'] = '海尔'
_df_wb_train_output['data_role'] = 'tr'
_df_wb_train_output = _df_wb_train_output.rename(columns = {'user_id_y': 'user_id',
                                                            'auth_date': 'date',
                                                            'flag': 'y',
                                                            'data_role': 'flagtrte',
                                                            })

_df_wb_oot_output = _df_wb_oot[['user_id_y', 
                                      'auth_date', 
                                      'flag', 
                                      'data_role',
                                      "三方数据源_hengpu_4",
                                    "INQUIRY0047",
                                    "三方数据源_hangliezhi_1",
                                    "digitalScore",
                                    "CARD0065",
                                    "三方数据源_ruizhi_6",
                                    "query_tfau_mg",
                                    "pboc_education_new",
                                    "income_izai_tmf",
                                    "quad_p1",
                                    "三方数据源_bairong_14",
                                    "loan_overdue_last_months_l12m",
                                    "三方数据源_xinyongsuanli_1",
                                    "debts_sgad_mc",
                                    "repay_dcal_tmd",
                                    "repay_dhaa_md"
                                      ]]
_df_wb_oot_output['loan_type'] = '海尔'
_df_wb_oot_output['data_role'] = 'oot'
_df_wb_oot_output = _df_wb_oot_output.rename(columns = {'user_id_y': 'user_id',
                                                            'auth_date': 'date',
                                                            'flag': 'y',
                                                            'data_role': 'flagtrte',
                                                            })


_df_wb_rej_output = df_yx[df_yx.auth_status == '拒绝'][['user_id_y', 
                                      'auth_date', 
                                      "三方数据源_hengpu_4",
                                    "INQUIRY0047",
                                    "三方数据源_hangliezhi_1",
                                    "digitalScore",
                                    "CARD0065",
                                    "三方数据源_ruizhi_6",
                                    "query_tfau_mg",
                                    "pboc_education_new",
                                    "income_izai_tmf",
                                    "quad_p1",
                                    "三方数据源_bairong_14",
                                    "loan_overdue_last_months_l12m",
                                    "三方数据源_xinyongsuanli_1",
                                    "debts_sgad_mc",
                                    "repay_dcal_tmd",
                                    "repay_dhaa_md"
                                      ]]
_df_wb_rej_output['loan_type'] = '海尔'
_df_wb_rej_output['y'] = -1
_df_wb_rej_output['data_role'] = 'rej'
_df_wb_rej_output = _df_wb_rej_output.rename(columns = {'user_id_y': 'user_id',
                                                            'auth_date': 'date',
                                                            'data_role': 'flagtrte',
                                                            })


print(_df_wb_train_output.shape)
print(_df_wb_oot_output.shape)
print(_df_wb_rej_output.shape)


_df_wb_output = pd.concat([_df_wb_train_output[['user_id','y','date','loan_type','flagtrte',"三方数据源_hengpu_4","INQUIRY0047","三方数据源_hangliezhi_1","digitalScore","CARD0065","三方数据源_ruizhi_6","query_tfau_mg","pboc_education_new","income_izai_tmf","quad_p1","三方数据源_bairong_14","loan_overdue_last_months_l12m","三方数据源_xinyongsuanli_1","debts_sgad_mc","repay_dcal_tmd","repay_dhaa_md"]]
,_df_wb_oot_output[['user_id','y','date','loan_type','flagtrte',"三方数据源_hengpu_4","INQUIRY0047","三方数据源_hangliezhi_1","digitalScore","CARD0065","三方数据源_ruizhi_6","query_tfau_mg","pboc_education_new","income_izai_tmf","quad_p1","三方数据源_bairong_14","loan_overdue_last_months_l12m","三方数据源_xinyongsuanli_1","debts_sgad_mc","repay_dcal_tmd","repay_dhaa_md"]]
,_df_wb_rej_output[['user_id','y','date','loan_type','flagtrte',"三方数据源_hengpu_4","INQUIRY0047","三方数据源_hangliezhi_1","digitalScore","CARD0065","三方数据源_ruizhi_6","query_tfau_mg","pboc_education_new","income_izai_tmf","quad_p1","三方数据源_bairong_14","loan_overdue_last_months_l12m","三方数据源_xinyongsuanli_1","debts_sgad_mc","repay_dcal_tmd","repay_dhaa_md"]]
], axis = 0)
print(_df_wb_output.shape)


for i in ["三方数据源_hengpu_4","INQUIRY0047","三方数据源_hangliezhi_1","digitalScore","CARD0065","三方数据源_ruizhi_6","query_tfau_mg","pboc_education_new","income_izai_tmf","quad_p1","三方数据源_bairong_14","loan_overdue_last_months_l12m","三方数据源_xinyongsuanli_1","debts_sgad_mc","repay_dcal_tmd","repay_dhaa_md"
]:
    _df_wb_output.loc[_df_wb_output[i]<0, i] = np.nan



_df_wb_output.to_csv(r"C:\huangting\海尔项目\建模_lr_dpd3\输出\模型报告入参2_XY.csv", index = False)





#==============================================================================
# File: code03_Proc02_FeatDesc.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc02_FeatDesc
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc02_FeatDesc
## 特征描述性统计


####################################################################
# Proc02_FeatDesc
# 特征描述性统计


###########################################################################
df_t_describe_summary = func_dataframe_describe(
    in_df=df_wb,
    var_names=cols_v,
).reset_index(drop=False)

# df_t_describe_summary.to_clipboard(index=False)


###########################################################################
cols_v_cate = [
    s0 for s0 in cols_v
    if s0 in df_t_describe_summary.query("data_type=='Categorical'")["column_name"].tolist()
]
cols_v_num = [
    s0 for s0 in cols_v
    if s0 in df_t_describe_summary.query("data_type=='Numerical'")["column_name"].tolist()
]





### （结果导出）

# ####################################################################################
# # to_clipboard
# df_t_describe_summary.to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc02_FeatDesc/df_t_describe_summary.xlsx".format(result_path), mode="wb") as fw:
    df_t_describe_summary \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc02_FeatDesc/df_t_describe_summary.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_describe_summary, file=fw)


####################################################################################
with open(file="{}/Proc02_FeatDesc/df_t_describe_summary.pkl".format(result_path), mode="rb") as fr:
    df_t_describe_summary = pickle.load(file=fr)

####################################################################################
cols_v_cate = [
    s0 for s0 in cols_v
    if s0 in df_t_describe_summary.query("data_type=='Categorical'")["column_name"].tolist()
]
cols_v_num = [
    s0 for s0 in cols_v
    if s0 in df_t_describe_summary.query("data_type=='Numerical'")["column_name"].tolist()
]

####################################################################################
print(df_t_describe_summary.shape[0])








# in_df=df_wb
# var_names=cols_v[:10]+["sb_24m_zzs_ge07_adv_cnt_dup"]
# drop_labels=None





# _df = (in_df if var_names==None else in_df[var_names]).drop(labels=([] if drop_labels==None else drop_labels), axis=1)


# describe_info =  _df.groupby(axis=1, level=0, sort=False).apply(
#     lambda s0: OrderedDict({
#         "data_type": ("Numerical" if re.search("(float|int)", s0.dtypes[0].name)!=None else "Categorical"),
#         "count": s0.shape[0],
#         "count_missing": s0[s0.iloc[:, 0].isna()].shape[0],
#         "count_nomissing": s0[-s0.iloc[:, 0].isna()].shape[0],
#         "pct_missing": s0[s0.iloc[:, 0].isna()].shape[0]/s0.shape[0],
#         "pct_nomissing": s0[-s0.iloc[:, 0].isna()].shape[0]/s0.shape[0],
#         "unique_count": s0.iloc[:, 0].unique().shape[0],
#         "unique_pct": s0.iloc[:, 0].unique().shape[0]/s0.shape[0],
        
# #             "min": (s0.iloc[:, 0].dropna().min() if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "min": (s0.iloc[:, 0].dropna().min() if (s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "mean": (s0.iloc[:, 0].dropna().mean() if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
# #             "max": (s0.iloc[:, 0].dropna().max() if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "max": (s0.iloc[:, 0].dropna().max() if (s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
        
#         "percentile_05": (np.percentile(a=s0.iloc[:, 0].dropna(), q=0.05, interpolation="midpoint") if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "percentile_25": (np.percentile(a=s0.iloc[:, 0].dropna(), q=0.25, interpolation="midpoint") if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "percentile_50": (np.percentile(a=s0.iloc[:, 0].dropna(), q=0.50, interpolation="midpoint") if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "percentile_75": (np.percentile(a=s0.iloc[:, 0].dropna(), q=0.75, interpolation="midpoint") if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),
#         "percentile_95": (np.percentile(a=s0.iloc[:, 0].dropna(), q=0.95, interpolation="midpoint") if (re.search("(float|int)", s0.dtypes[0].name)!=None and s0.iloc[:, 0].dropna().shape[0]>0) else np.NaN),

#         "top5_value": str(dict(s0.iloc[:, 0].dropna().value_counts().reset_index().values[:5, :])),
        
#         # "entropy": np.sum([-t*np.log2(t+1e-2) for t in s0.iloc[:, 0].value_counts()/s0.shape[0]]),
#         # "entropy_ratio": np.sum([-t*np.log2(t+1e-2) for t in s0.iloc[:, 0].value_counts()/s0.shape[0]])/(np.log(s0.iloc[:, 0].unique().shape[0]+1e-2)+1e-2),
        
#         "entropy": np.sum([-t*np.log(t) for t in s0.iloc[:, 0].value_counts(normalize=True)]),
#         "entropy_ratio": np.sum([-t*np.log(t) for t in s0.iloc[:, 0].value_counts(normalize=True)])/np.log(s0.iloc[:, 0].unique().shape[0]),
        
        
#     })
# )


# _df["sb_24m_zzs_ge07_adv_cnt_dup"] = np.NaN

# _df["sb_24m_zzs_ge07_adv_cnt_dup"].value_counts()


# np.sum([-t*np.log(t) for t in _df["sb_24m_zzs_ge07_adv_cnt_dup"].value_counts(normalize=True)])


# (np.log(_df["sb_24m_zzs_ge07_adv_cnt_dup"].unique().shape[0]+1e-10)+1e-10)




# np.sum([-t*np.log(t+1e-10) for t in _df["sb_24m_zzs_ge07_adv_cnt_dup"].value_counts()/_df["sb_24m_zzs_ge07_adv_cnt_dup"].shape[0]]) \
#         /(np.log(_df["sb_24m_zzs_ge07_adv_cnt_dup"].unique().shape[0]+1e-10)+1e-10)



# describe_info["sb_24m_zzs_ge07_adv_cnt_dup"]














#==============================================================================
# File: code04_Proc03_FeatFineBinning.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc03_FeatFineBinning
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc03_FeatFineBinning
## 特征细分箱

####################################################################
# Proc03_FeatFineBinning
# 特征细分箱


###########################################################################
# 计算变量自动分箱后WOE、IV

###########################################################################
def _func_calc_fine_iv(
        in_df,
        var_name,
        target_label,
        data_type,
        dict_fine_binning_methods,
    ):
    ############################################################
    rt = []
    for binning_method, calc_params in list(dict_fine_binning_methods.items())[:]:
        
        # print(binning_method)
        
        if dict_fine_binning_methods[binning_method]["data_type"]!=data_type:
            continue
        
        # print(binning_method, calc_params)
        
        _out = OrderedDict({
            "binning_method": binning_method,
            "column_name": var_name,
            "data_type": data_type,
            "IV": None,
            "IV_dropna": None,
            "boundary": None,
            "mapping_gb_class": None,
            "crosstab": None,
            "bin_cnt": None,
            "_calc_params": calc_params,
        })

        ############################################################
        if data_type=="Categorical":
            _, _out["crosstab"], _out["mapping_gb_class"] = func_auto_combining_discrete_v2(
                in_df=in_df, var_name=var_name, target_label=target_label,
                min_pct=calc_params.get("min_pct"),
                max_bins_cnt=calc_params.get("max_bins_cnt"),
                method=calc_params.get("method"),
                with_lift_ks=calc_params.get("with_lift_ks"),
                lift_calc_ascending=calc_params.get("lift_calc_ascending"),
                with_wls_adj_woe=calc_params.get("with_wls_adj_woe"),
                woe_with_nan=calc_params.get("woe_with_nan"),
            )

        ############################################################
        elif data_type=="Numerical" and calc_params["method"] in ["01_equal_freq", "02_decision_tree", "03_chi2_comb"]:
            _, _out["crosstab"], _out["boundary"] = func_auto_binning_continuous_v2(
                in_df=in_df, var_name=var_name, target_label=target_label,
                min_pct=calc_params.get("min_pct"),
                max_bins_cnt=calc_params.get("max_bins_cnt"),
                method=calc_params.get("method"),
                chi2_min_pvalue=calc_params.get("chi2_min_pvalue"),
                chi2_min_cnt=calc_params.get("chi2_min_cnt"),
                with_lift_ks=calc_params.get("with_lift_ks"),
                lift_calc_ascending=calc_params.get("lift_calc_ascending"),
                with_wls_adj_woe=calc_params.get("with_wls_adj_woe"),
                woe_with_nan=calc_params.get("woe_with_nan"),
            )

        ############################################################
        elif data_type=="Numerical" and calc_params["method"]=="10_quantile":
            _data_converted, _boundary, _df_qcut_mapping_dups = func_binning_continuous_quantile_v1(
                in_data=in_df[var_name],
                bins_q=calc_params["bins_q"],
                out_type="01_info_cp",
                right_border=True, include_lowest=False,
                output_cp=True,
            )
            _out["boundary"] = _boundary
            _out["crosstab"] = func_woe_report_v1(
                in_var=_data_converted,
                in_target=in_df[target_label],
                with_total=True,
                with_lift_ks=calc_params.get("with_lift_ks"),
                lift_calc_ascending=calc_params.get("lift_calc_ascending"),
                with_wls_adj_woe=calc_params.get("with_wls_adj_woe"),
                woe_with_nan=calc_params.get("woe_with_nan"),
            )

        _crosstab = _out["crosstab"]
        _out["IV"] = _crosstab.loc["total", "IV"]
        _out["IV_dropna"] = _crosstab.loc[[s0 for s0 in _crosstab.index if re.search("(_NaN$)|(NaN, NaN)|(^total$)", s0)==None], "IV"].sum()
        _out["bin_cnt"] = _crosstab[_crosstab.index!="total"].shape[0]
        
        rt.append(_out)
    
    ############################################################
    rt = pd.DataFrame(rt)
    return rt

# ###########################################################################
# def _func_calc_fine_iv_BoostrapSampling(
#         in_df,
#         var_name,
#         target_label,
#         data_type,
#         dict_fine_binning_methods,
#         random_seed=9999,
#         sampling_n=3000,
#         sampling_epochs=100,
#         sample_replace=False,
#         verbose=True,
#         with_target_balance=False,
#     ):
    
#     ################################################
#     if verbose:
#         # print("-"*60)
#         print("\n>>>>  BoostrapSampling Processing: {}".format(var_name))
#         _time = time.time()
    
#     df = in_df[[var_name, target_label]].reset_index(drop=True)
#     sample_replace = (True if sampling_n>df.shape[0] else sample_replace)
#     # sample_replace = (True if int(sampling_n/2)>df.shape[0] else sample_replace)
    
#     ################################################
#     df_iv_info_bs = pd.DataFrame([])
#     for idx in range(sampling_epochs):
#         _time_0 = time.time()
#         _random_seed = random_seed+idx
        
#         if with_target_balance:
#             _df = pd.concat(
#                 [
#                     df.query("{}=='0_good'".format(target_label)).sample(n=int(sampling_n/2), replace=True, random_state=_random_seed, ignore_index=True),
#                     df.query("{}=='1_bad'".format(target_label)).sample(n=int(sampling_n/2), replace=True, random_state=_random_seed, ignore_index=True),
#                 ],
#                 ignore_index=True,
#             )
#         else:
#             _df = df.sample(n=sampling_n, replace=sample_replace, random_state=_random_seed, ignore_index=True)
        
#         ################################################
#         _df_iv_info = _func_calc_fine_iv(
#             in_df=_df,
#             var_name=var_name,
#             target_label=target_label,
#             data_type=data_type,
#             dict_fine_binning_methods=dict_fine_binning_methods,
#         )
#         _df_iv_info.insert(loc=0, column="bs_random_seed", value=_random_seed)
#         _df_iv_info.insert(loc=0, column="bs_idx", value=idx)
        
#         df_iv_info_bs = df_iv_info_bs.append(_df_iv_info, ignore_index=True)
        
#         ################################################
#         if verbose:
#             print("    [ {:.2f}, {:.2f} ] {:7.2f}%".format(time.time()-_time, time.time()-_time_0, idx/sampling_epochs*100))
    
#     ################################################
#     if verbose:
#         # print("\n")
#         print("\n    cost time: {} sec.\n".format(time.time()-_time))
    
#     return df_iv_info_bs


###########################################################################
_dict_fine_binning_methods = {
    
    # 始值统计（cate01_ori）
    "cate01_ori": {
        "data_type": "Categorical",
        "min_pct": 0.000001,
        "max_bins_cnt": None,
        "method": "01_ori",
        "with_lift_ks": False,
        "lift_calc_ascending": False,
        "with_wls_adj_woe": False,
        "woe_with_nan": True,
        # "woe_with_nan": False,
    },
    
    # best_ks算法（cate02_bestks）
    "cate02_bestks": {
        "data_type": "Categorical",
        "min_pct": 0.05-0.0001,
        "max_bins_cnt": 5,
        "method": "09_best_ks",
        "with_lift_ks": False,
        "lift_calc_ascending": False,
        "with_wls_adj_woe": False,
        "woe_with_nan": True,
        # "woe_with_nan": False,
    },
    
    # 等频分箱，最大20箱（num01_eq20）
    "num01_eq20": {
        "data_type": "Numerical",
        "min_pct": 0.05-0.0001,
        "max_bins_cnt": 20,
        "method": "01_equal_freq",
        "with_lift_ks": True,
        "lift_calc_ascending": True,
        "with_wls_adj_woe": True,
        "woe_with_nan": True,
        # "woe_with_nan": False,
    },
    
    # # 等频分箱，最大6箱（num02_eq06）
    # "num02_eq06": {
    #     "data_type": "Numerical",
    #     "min_pct": 0.05-0.0001,
    #     "max_bins_cnt": 6,
    #     "method": "01_equal_freq",
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
    # 决策树分箱，最大9箱（num03_dt09）
    "num03_dt09": {
        "data_type": "Numerical",
        "min_pct": 0.1-0.0001,
        "max_bins_cnt": 9,
        "method": "02_decision_tree",
        "with_lift_ks": True,
        "lift_calc_ascending": True,
        "with_wls_adj_woe": True,
        "woe_with_nan": True,
        # "woe_with_nan": False,
    },
    
    # # 决策树分箱，最大5箱（num04_dt05）
    # "num04_dt05": {
    #     "data_type": "Numerical",
    #     "min_pct": 0.1-0.0001,
    #     "max_bins_cnt": 5,
    #     "method": "02_decision_tree",
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
    # # 卡方分箱（合并）（num05_chi2comb）
    # "num05_chi2comb": {
    #     "data_type": "Numerical",
    #     "min_pct": 0.1-0.0001,
    #     "max_bins_cnt": 50,
    #     "method": "03_chi2_comb",
    #     "chi2_min_pvalue": 0.1,
    #     "chi2_min_cnt": 2,
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
    # # 分位点分箱（num06_quantile_1）
    # "num06_quantile_1": {
    #     "data_type": "Numerical",
    #     "method": "10_quantile",
    #     "bins_q": [0.5, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99, 99.5],
    #     # "bins_q": [1, 5, 10, 25, 50, 75, 90, 95, 99],
    #     # "bins_q": [5, 10, 25, 50, 75, 90, 95],
    #     # "bins_q": [10, 25, 50, 75, 90],
    #     # "bins_q": [25, 50, 75],
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
    # # 分位点分箱（num07_quantile_2）
    # "num07_quantile_2": {
    #     "data_type": "Numerical",
    #     "method": "10_quantile",
    #     # "bins_q": [0.5, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99, 99.5],
    #     "bins_q": [1, 5, 10, 25, 50, 75, 90, 95, 99],
    #     # "bins_q": [5, 10, 25, 50, 75, 90, 95],
    #     # "bins_q": [10, 25, 50, 75, 90],
    #     # "bins_q": [25, 50, 75],
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
    # # 分位点分箱（num08_quantile_3）
    # "num08_quantile_3": {
    #     "data_type": "Numerical",
    #     "method": "10_quantile",
    #     # "bins_q": [0.5, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99, 99.5],
    #     # "bins_q": [1, 5, 10, 25, 50, 75, 90, 95, 99],
    #     # "bins_q": [5, 10, 25, 50, 75, 90, 95],
    #     "bins_q": [10, 25, 50, 75, 90],
    #     # "bins_q": [25, 50, 75],
    #     "with_lift_ks": True,
    #     "lift_calc_ascending": True,
    #     "with_wls_adj_woe": True,
    #     "woe_with_nan": True,
    #     # "woe_with_nan": False,
    # },
    
}




##########################################################################
# 计算IV
_time = time.time()
_idx = 1

_df_t_features_fine_iv = pd.DataFrame([])

_df = df_t_describe_summary \
    [df_t_describe_summary["column_name"].isin(cols_v[:])] \
    [["column_name", "data_type"]] \
    .reset_index(drop=True)
for _column_name, _data_type in _df[["column_name", "data_type"]].values[:]:
    print("[ {:.2f} ] {:7.2f}%: {}".format(time.time()-_time, _idx/_df.shape[0]*100, _column_name))
    
    ##########################################################################
    if _df_wb_dev[_column_name].notna().sum()==0:
        continue
    
    ##########################################################################
    _df_iv_info = _func_calc_fine_iv(
        in_df=_df_wb_dev,
        var_name=_column_name,
        target_label="target_label",
        data_type=_data_type,
        dict_fine_binning_methods=_dict_fine_binning_methods,
    )
    
    ##########################################################################
    with_boostrap_sampling = False
    # with_boostrap_sampling = True
    ##########################################################################
    # if with_boostrap_sampling:
    #     _df_iv_info_bs = _func_calc_fine_iv_BoostrapSampling(
    #         in_df=df_wb,
    #         var_name=_column_name,
    #         target_label="target_label",
    #         data_type=_data_type,
    #         dict_fine_binning_methods=_dict_fine_binning_methods,
    #         random_seed=9999,
    #         # random_seed=99999,
    #         # sampling_n=300,
    #         # sampling_n=3000,
    #         # sampling_n=10000,
    #         # sampling_n=int(df_wb.shape[0]*0.6),
    #         sampling_n=int(df_wb.shape[0]*0.1),
    #         # sampling_epochs=100,
    #         sampling_epochs=30,
    #         # sampling_epochs=20,
    #         sample_replace=False,
    #         verbose=True,
    #         # verbose=False,
    #         with_target_balance=False,
    #         # with_target_balance=True,
    #     )
    #     _df_iv_info_bs["crosstab"] = _df_iv_info_bs["crosstab"].apply(lambda s0: s0.to_dict(orient="records", into=dict))
    #     _df_iv_info = _df_iv_info.merge(
    #         right=_df_iv_info_bs.groupby(by=["column_name", "binning_method"]).apply(
    #             lambda s0: pd.Series(dict({
    #                 # "bs_epochs": s0.shape[0],
    #                 "bs_result": s0[[
    #                                     "bs_idx", "bs_random_seed",
    #                                     "IV", "IV_dropna", "boundary", "mapping_gb_class", "crosstab", "_calc_params",
    #                                 ]].to_dict(orient="records", into=dict),

    #             }))
    #         ).reset_index(drop=False),
    #         how="left", left_on=["column_name", "binning_method"], right_on=["column_name", "binning_method"],
    #     )
    
    ##########################################################################
    _df_t_features_fine_iv = _df_t_features_fine_iv.append(_df_iv_info, ignore_index=True)
    _idx = _idx+1

##########################################################################
_df_t_features_fine_iv["output_cutpoint"] = _df_t_features_fine_iv.apply(
    lambda s0: (
        dict(
            [
                (_idx, int(_v))
                for _idx, _v in
                np.concatenate([
                    list(product(_v.split("/"), [_idx]))
                    for _idx, _v in
                    [
                        (s0.split("_")[0], "_".join(s0.split("_")[1:]))
                        for s0 in s0["crosstab"].index if re.search("(^total$)", s0)==None
                    ]
                ])
            ]
        )
        if s0["data_type"]=="Categorical" else
        [-np.inf]+pd.Series([
            float(re.sub("[ \]]", "", s0.split("_")[1].split(",")[1]))
            for s0 in s0["crosstab"].index
            if re.search("(NaN, NaN)|(, inf)|(^total$)", s0)==None
        ]).drop_duplicates().tolist()+[np.inf]
        if s0["data_type"]=="Numerical" else
        None
    ),
    axis=1,
)

_df_t_features_fine_iv["_output_object"] = _df_t_features_fine_iv.apply(
    lambda s0: {
        "feature_name": s0["column_name"],
        "data_type": s0["data_type"],
        "binning_method": s0["binning_method"],
        "boundary": s0["boundary"],
        "mapping_gb_class": s0["mapping_gb_class"],
        "crosstab_index": [s0 for s0 in s0["crosstab"].index],
        "crosstab_cnt": s0["crosstab"] \
            [["0_good_#", "1_bad_#"]] \
            .to_dict(orient="list"),
        "crosstab_details": s0["crosstab"] \
            [(
                ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate"]
                if s0["data_type"]=="Categorical" else
                ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks"]
                if s0["data_type"]=="Numerical" and "_wls_adj_woe_details" not in s0["crosstab"].columns else
                ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
                if s0["data_type"]=="Numerical" and "_wls_adj_woe_details" in s0["crosstab"].columns else
                []
            )] \
            .to_dict(orient="list"),
        "IV": s0["IV"],
        "IV_dropna": s0["IV_dropna"],
        "output_cutpoint": s0["output_cutpoint"],
        "_calc_params": s0["_calc_params"],
        "bs_result": (s0["bs_result"] if "bs_result" in _df_t_features_fine_iv.columns else None),
    },
    axis=1,
)

# _df_t_features_fine_iv = _df_t_features_fine_iv \
#     .reset_index(drop=False) \
#     .sort_values(by=["data_type", "binning_method", "index"], ascending=[True, True, True]) \
#     .drop(labels=["index"], axis=1) \
#     .reset_index(drop=True)
_df_t_features_fine_iv = _df_t_features_fine_iv.sort_values(by=["data_type", "binning_method", "IV_dropna"], ascending=[True, True, False]) \
    .reset_index(drop=True)


print("\n")
print("cost time: {} sec.".format(time.time()-_time))



###########################################################################
# 计算开发样本的WOE报告
_t_features_fine_woe_report = []
for _binning_method, _column_name, _data_type, _crosstab, _boundary, _mapping_gb_class in \
            _df_t_features_fine_iv[["binning_method", "column_name", "data_type", "crosstab", "boundary", "mapping_gb_class"]].values[:]:
    _crosstab = _crosstab.reset_index().rename(columns={"index": "gb_idx"}).query("gb_idx!='total'")
    _crosstab.insert(loc=0, column="binning_method", value=_binning_method)
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx"), column="feature_name", value=_column_name)
    _crosstab.insert(loc=_crosstab.columns.tolist().index("feature_name"), column="data_type", value=_data_type)
    # if _data_type=="Categorical":
    #     _mapping_gb_label = dict(
    #         pd.Series(_mapping_gb_class).reset_index().rename(columns={"index": "value_label", 0: "gb_idx"}) \
    #             .groupby(by=["gb_idx"])["value_label"].apply(lambda s0: "/".join(s0.sort_values(ascending=True).tolist()))
    #     )
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: _mapping_gb_label.get(s0, ""))
    # elif _data_type=="Numerical":
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx")+1, column="label", value=_data_label)
    _t_features_fine_woe_report.append(_crosstab)
_df_t_features_fine_woe_report = pd.concat(_t_features_fine_woe_report, ignore_index=True).reset_index(drop=True)
_df_t_features_fine_woe_report = _df_t_features_fine_woe_report[[
    s0 for s0 in _df_t_features_fine_woe_report.columns
    if s0 in ["binning_method", "data_type", "feature_name", "gb_idx", "label", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
]].reset_index(drop=True)


###########################################################################
# OUTPUT:
#     df_t_features_fine_iv
#     df_t_features_fine_woe_report

df_t_features_fine_iv = _df_t_features_fine_iv.reset_index(drop=True)
df_t_features_fine_woe_report = _df_t_features_fine_woe_report.reset_index(drop=True)

###########################################################################
print(df_t_features_fine_iv.shape)
print(df_t_features_fine_woe_report.shape)





### （分箱结果导出）

# ####################################################################################
# df_t_features_fine_iv.drop(labels=["crosstab", "_output_object"], axis=1).to_clipboard(index=False)
# # df_t_features_fine_woe_report.to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc03_FeatFineBinning/df_t_features_fine_iv.xlsx".format(result_path), mode="wb") as fw:
    df_t_features_fine_iv \
        [[s0 for s0 in df_t_features_fine_iv.columns if s0 not in ["crosstab", "_output_object", "bs_result"]]] \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )


###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc03_FeatFineBinning/df_t_features_fine_iv.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_features_fine_iv, file=fw)

####################################################################################
with open(file="{}/Proc03_FeatFineBinning/df_t_features_fine_iv.pkl".format(result_path), mode="rb") as fr:
    df_t_features_fine_iv = pickle.load(file=fr)



### （分箱结果保存输出json）

###########################################################################
for _binning_method, _df in df_t_features_fine_iv.groupby(by=["binning_method"]):
    
    print(_binning_method)
    with open(file="{}/Proc03_FeatFineBinning/obj_res_binning_{}.json".format(result_path, _binning_method), mode="w", encoding="utf-8") as fw:
        json.dump(
            obj=_df["_output_object"].tolist(),
            fp=fw,
            indent=4,
            ensure_ascii=False,
        )



### （分箱结果导入json）

###########################################################################
_obj_cp = []
for _fn in pd.Series(data=[
                        s0 for s0 in os.listdir(path="{}/Proc03_FeatFineBinning".format(result_path))
                        if re.search("obj_res_binning_", s0)
                    ]).sort_values(ascending=True):
    
    print(_fn)
    with open(file="{}/Proc03_FeatFineBinning/{}".format(result_path, _fn), mode="r", encoding="utf-8") as fr:
        _obj_cp = _obj_cp+json.load(fp=fr)

###########################################################################
# IV
df_t_features_fine_iv = pd.DataFrame([
    {
        "binning_method": s0["binning_method"],
        "column_name": s0["feature_name"],
        "data_type": s0["data_type"],
        "IV": s0["IV"],
        "IV_dropna": s0["IV_dropna"],
        "boundary": s0["boundary"],
        "mapping_gb_class": s0["mapping_gb_class"],
        "crosstab": pd.DataFrame(index=s0["crosstab_index"], data=dict(list(s0["crosstab_cnt"].items())+list(s0["crosstab_details"].items()))),
        "bin_cnt": len([s0 for s0 in s0["crosstab_index"] if s0!="total"]),
        "output_cutpoint": s0["output_cutpoint"],
        "_output_object": s0,
        # "_output_object": dict([(_k, _v) for _k, _v in s0.items() if _k!="bs_result"]),
        "_calc_params": s0["_calc_params"],
        "bs_result": (s0["bs_result"] if "bs_result" in s0.keys() else None),
        # "bs_result": None,
    }
    for s0 in _obj_cp
])

###########################################################################
# WOE report
_t_features_fine_woe_report = []
for _binning_method, _column_name, _data_type, _crosstab, _boundary, _mapping_gb_class in \
            df_t_features_fine_iv[["binning_method", "column_name", "data_type", "crosstab", "boundary", "mapping_gb_class"]].values[:]:
    _crosstab = _crosstab.reset_index().rename(columns={"index": "gb_idx"}).query("gb_idx!='total'")
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx"), column="feature_name", value=_column_name)
    _crosstab.insert(loc=_crosstab.columns.tolist().index("feature_name"), column="data_type", value=_data_type)
    # if _data_type=="Categorical":
    #     _mapping_gb_label = dict(
    #         pd.Series(_mapping_gb_class).reset_index().rename(columns={"index": "value_label", 0: "gb_idx"}) \
    #             .groupby(by=["gb_idx"])["value_label"].apply(lambda s0: "/".join(s0.sort_values(ascending=True).tolist()))
    #     )
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: _mapping_gb_label.get(s0, ""))
    # elif _data_type=="Numerical":
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx")+1, column="label", value=_data_label)
    _crosstab.insert(loc=0, column="binning_method", value=_binning_method)
    _t_features_fine_woe_report.append(_crosstab)
df_t_features_fine_woe_report = pd.concat(_t_features_fine_woe_report, ignore_index=True).reset_index(drop=True)
df_t_features_fine_woe_report = df_t_features_fine_woe_report[[
    s0 for s0 in df_t_features_fine_woe_report.columns
    if s0 in ["binning_method", "data_type", "feature_name", "gb_idx", "label", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
]].reset_index(drop=True)

###########################################################################
_dict_fine_binning_methods = dict(
    df_t_features_fine_iv[["binning_method", "_calc_params"]] \
        .drop_duplicates(subset=["binning_method"], keep="first").values
)

###########################################################################
print("\n")
print(df_t_features_fine_iv.shape)
print(df_t_features_fine_woe_report.shape)
print("\n")


























# ###########################################################################
# _dict_fine_binning_methods = {
    
#     # 始值统计（cate01_ori）
#     "cate01_ori": {
#         "data_type": "Categorical",
#         "min_pct": 0.000001,
#         "max_bins_cnt": None,
#         "method": "01_ori",
#         "with_lift_ks": False,
#         "lift_calc_ascending": False,
#         "with_wls_adj_woe": False,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
#     # best_ks算法（cate02_bestks）
#     "cate02_bestks": {
#         "data_type": "Categorical",
#         "min_pct": 0.05-0.0001,
#         "max_bins_cnt": 5,
#         "method": "09_best_ks",
#         "with_lift_ks": False,
#         "lift_calc_ascending": False,
#         "with_wls_adj_woe": False,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
#     # 等频分箱，最大20箱（num01_eq20）
#     "num01_eq20": {
#         "data_type": "Numerical",
#         "min_pct": 0.05-0.0001,
#         "max_bins_cnt": 20,
#         "method": "01_equal_freq",
#         "with_lift_ks": True,
#         "lift_calc_ascending": True,
#         "with_wls_adj_woe": True,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
#     # 决策树分箱，最大9箱（num02_dt09）
#     "num02_dt09": {
#         "data_type": "Numerical",
#         "min_pct": 0.1-0.0001,
#         "max_bins_cnt": 9,
#         "method": "02_decision_tree",
#         "with_lift_ks": True,
#         "lift_calc_ascending": True,
#         "with_wls_adj_woe": True,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
#     # 卡方分箱（合并）（num03_chi2comb）
#     "num03_chi2comb": {
#         "data_type": "Numerical",
#         "min_pct": 0.1-0.0001,
#         "max_bins_cnt": 50,
#         "method": "03_chi2_comb",
#         "chi2_min_pvalue": 0.1,
#         "chi2_min_cnt": 2,
#         "with_lift_ks": True,
#         "lift_calc_ascending": True,
#         "with_wls_adj_woe": True,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
#     # 分位点分箱（num04_quantile）
#     "num04_quantile": {
#         "data_type": "Numerical",
#         "method": "10_quantile",
#         "bins_q": [0.5, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99, 99.5],
#         # "bins_q": [1, 5, 10, 25, 50, 75, 90, 95, 99],
#         # "bins_q": [5, 10, 25, 50, 75, 90, 95],
#         # "bins_q": [10, 25, 50, 75, 90],
#         # "bins_q": [25, 50, 75],
#         "with_lift_ks": True,
#         "lift_calc_ascending": True,
#         "with_wls_adj_woe": True,
#         "woe_with_nan": True,
#         # "woe_with_nan": False,
#     },
    
# }







# _df_iv_info_bs = _func_calc_fine_iv_BoostrapSampling(
#     in_df=df_wb_0,
#     # var_name="sb_12m_zzs_qbxse_amount_sum",
#     var_name="zs_12m_zzs_ge03_adv_cnt_dup",
#     target_label="target_label",
#     data_type="Numerical",
#     dict_fine_binning_methods=_dict_fine_binning_methods,
#     random_seed=9999,
#     # random_seed=99999,
#     sampling_n=1000,
#     # sampling_n=3000,
#     # sampling_n=10000,
#     # sampling_n=int(df_wb_0.shape[0]*0.6),
#     sampling_epochs=30,
#     sample_replace=False,
#     verbose=True,
#     # verbose=False,
#     with_target_balance=False,
#     # with_target_balance=True,
# )

# # plt.xticks(rotation=30)
# # sns.violinplot(
# #     x=_df_iv_info_bs["binning_method"],
# # #     y=_df_iv_info_bs["crosstab"] \
# # #         .apply(lambda s0: abs(s0["WOE"]-s0["wls_adj_WOE"]).mean()),
# #     y=_df_iv_info_bs["IV_dropna"],
# # #     y=_df_iv_info_bs["IV"],
# # #     y=_df_iv_info_bs["crosstab"].apply(lambda s0: s0["_wls_adj_woe_details"].dropna()[0]["_rsquared"]),
# # )

# # _df_iv_info_bs.groupby(by=["column_name", "binning_method"]).apply(
# #     lambda s0: pd.Series(dict({
# #         # "bs_epochs": s0.shape[0],
# #         "bs_result": s0[["bs_idx", "bs_random_seed", "IV", "IV_dropna", "boundary", "mapping_gb_class", "crosstab", "_calc_params"]] \
# #                 .to_dict(orient="records", into=dict),
        
# #     }))
# # ).reset_index(drop=False)













### woe编码转换

# ###########################################################################
# # woe编码
# def _func_woe_mapping(in_df, df_mapping_info, cols_keep=None, print_log=False, with_wls_adj_woe=False):
    
#     ################################################
#     df_mapping_info = df_mapping_info \
#         .sort_values(by=["column_name", "IV_dropna"], ascending=[True, False]) \
#         [["binning_method", "column_name", "data_type", "crosstab", "output_cutpoint"]] \
#         .drop_duplicates(subset=["column_name"], keep="first") \
#         .sort_index(ascending=True) \
#         .reset_index(drop=True)
#     if cols_keep is None:
#         cols_keep = [s0 for s0 in in_df.columns if s0 not in df_mapping_info["column_name"].tolist()]
    
#     ################################################
#     df_woe = in_df[cols_keep].reset_index(drop=True)
#     _idx = 1
#     for _binning_method, _column_name, _data_type, _crosstab, _output_cutpoint in df_mapping_info.values[:]:
        
#         if print_log:
#             print("{} {:7.2f}%: {}".format(_binning_method, _idx/df_mapping_info.shape[0]*100, _column_name))
#             _idx = _idx+1
        
#         ################################################
#         if _data_type=="Numerical":
#             _mapping = dict([
#                 (re.sub("[, \]]", "", _k.split("_")[1].split(",")[1]), _v)
#                 # (_k, _v)
#                 for _k, _v in (_crosstab["wls_adj_WOE"].to_dict().items() if with_wls_adj_woe else _crosstab["WOE"].to_dict().items())
#                 # for _k, _v in _crosstab["WOE"].to_dict().items()
#                 # for _k, _v in _crosstab["wls_adj_WOE"].to_dict().items()
#                 # if re.search("(NaN, NaN)|(^total$)", _k)==None
#                 if re.search("(^total$)", _k)==None
#             ])
#             df_woe["BIN_{}".format(_column_name)] = func_binning_continuous_v1(
#                 in_data=in_df[_column_name],
#                 bins=_output_cutpoint,
#                 out_type="01_info", right_border=True, include_lowest=False,
#             )
#             df_woe["WOE_{}".format(_column_name)] = df_woe["BIN_{}".format(_column_name)] \
#                 .apply(lambda s0: _mapping.get(re.sub("[, \]]", "", s0.split("_")[1].split(",")[1])))
#         elif _data_type=="Categorical":
#             _mapping = dict([
#                 (_k.split("_")[1], _v)
#                 # (_k, _v)
#                 for _k, _v in _crosstab["WOE"].to_dict().items()
#                 # for _k, _v in _crosstab["wls_adj_WOE"].to_dict().items()
#                 # if re.search("(NaN, NaN)|(^total$)", _k)==None
#                 if re.search("(^total$)", _k)==None
#             ])
#             df_woe["BIN_{}".format(_column_name)] = func_combining_discrete_v1(
#                 in_data=in_df[_column_name],
#                 mapping_gb_class=_output_cutpoint,
#                 fillna_value="NaN", cvt_fillna_value=_output_cutpoint.get("NaN", 0),
#             )
#             df_woe["WOE_{}".format(_column_name)] = df_woe["BIN_{}".format(_column_name)] \
#                 .apply(lambda s0: _mapping.get(s0.split("_")[1]))
#         else:
#             pass
    
#     return df_woe

# ###########################################################################
# df_wb_fine_woe_cate01_ori = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("cate01_ori")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )
# df_wb_fine_woe_cate02_bestks = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("cate02_bestks")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )
# df_wb_fine_woe_num01_eq20 = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("num01_eq20")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )
# df_wb_fine_woe_num02_dt09 = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("num02_dt09")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )
# df_wb_fine_woe_num03_chi2comb = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("num03_chi2comb")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )
# df_wb_fine_woe_num04_quantile = \
#     _func_woe_mapping(
#         in_df=df_wb,
#         df_mapping_info=df_t_features_fine_iv.query("binning_method=='{}'".format("num04_quantile")),
#         cols_keep=cols_base,
#         print_log=True,
#         with_wls_adj_woe=False,
#     )














#==============================================================================
# File: code05_Proc04_WoeMonotonic.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc03_FeatFineBinning
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc04_WoeMonotonic
## woe单调性分析（针对数值型）


###########################################################################
# Proc04_WoeMonotonic
# woe单调性分析（针对数值型）


###########################################################################
def _func_calc_binning_woe_stats_compare(
        in_df, var_name, target_label,
        # list_bins_cnt=[20, 6, 5, 4, 3],
        list_bins_cnt=[100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 9, 8, 7, 6, 5, 4, 3],
    ):
    _result = []
    for _max_bins_cnt in list_bins_cnt[:]:
        _, _crosstab, _bin = func_auto_binning_continuous_v1(
            in_var=in_df[var_name],
            in_target=in_df[target_label],
            min_pct=0.05, max_bins_cnt=_max_bins_cnt,
            # min_pct=0.05-0.0001, max_bins_cnt=20,
            # min_pct=0.1-0.0001, max_bins_cnt=9,
            # min_pct=0.01-0.0001, max_bins_cnt=50,
            method="01_equal_freq",
            # method="02_decision_tree",
            # method="03_chi2_comb",
            # method="09_best_ks",
            # chi2_min_pvalue=0.1, chi2_min_cnt=2,
            with_lift_ks=True, lift_calc_ascending=True, with_wls_adj_woe=True,
            right_border=True, include_lowest=True,
            woe_with_nan=True,
        )
        _res = func_calc_binning_woe_stats(
            in_crosstab=_crosstab,
            plot=False,
        )
        _res = OrderedDict(
            [
                ("column_name", var_name),
                ("max_bins_cnt", _max_bins_cnt),
                ("bin", _bin),
                ("output_cutpoint", [-np.inf]+pd.Series([
                                        float(re.sub("[ \]]", "", s0.split("_")[1].split(",")[1]))
                                        for s0 in _crosstab.index
                                        if re.search("(NaN, NaN)|(, inf)|(^total$)", s0)==None
                                    ]).drop_duplicates().tolist()+[np.inf]
                ),
                ("crosstab", _crosstab),
                ("_wls_adj_stats_info", 
                     (
                        {
                            "_rsquared_p1": _crosstab.iloc[1]["_wls_adj_woe_details"]["_rsquared_p1"],
                            "_rsquared_p2": _crosstab.iloc[1]["_wls_adj_woe_details"]["_rsquared_p2"],
                            "_WOE_MSE": (
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), ["WOE", "wls_adj_WOE"]].apply(lambda s0: (s0["wls_adj_WOE"]-s0["WOE"])**2, axis=1) * \
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum() / \
                                    (_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum()).sum()
                                ).sum(),
                            "_WOE_RMSE": (
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), ["WOE", "wls_adj_WOE"]].apply(lambda s0: (s0["wls_adj_WOE"]-s0["WOE"])**2, axis=1) * \
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum() / \
                                    (_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum()).sum()
                                ).sum()**(0.5),
                            "_WOE_MAE": (
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), ["WOE", "wls_adj_WOE"]].apply(lambda s0: abs(s0["wls_adj_WOE"]-s0["WOE"]), axis=1) * \
                                    _crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum() / \
                                    (_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"]/_crosstab.loc[_crosstab["wls_adj_WOE"].notna(), "total"].sum()).sum()
                                ).sum(),
                        }
                        if "_wls_adj_woe_details" in _crosstab.columns else
                        dict()
                    )
                ),
                ("_bin_str", "[{}]".format(", ".join([str(s0) for s0 in _bin]))),
            ]+
            list(_res.to_dict(into=OrderedDict).items())
        )
        _result.append(_res)
    
    rt = pd.DataFrame(_result)
    
    rt = pd.merge(
        left=rt[["column_name", "max_bins_cnt", "bin", "output_cutpoint", "_bin_str"]],
        right=rt.drop_duplicates(subset=["_bin_str"], keep="first").drop(labels=["bin", "output_cutpoint"], axis=1),
        how="left",
        left_on=["column_name", "max_bins_cnt", "_bin_str"],
        right_on=["column_name", "max_bins_cnt", "_bin_str"],
    )[[
        "column_name",
        # "binning_method",
        "max_bins_cnt", "bin", "output_cutpoint",
        "crosstab", "_wls_adj_stats_info",
        "binning_cnt", "binning_cnt_notna", "IV_notna", "IV_pre_bin_notna", "total_pct_cv", "total_pct_max",
        "corr_pearson", "corr_spearman", "corr_kendall",
        "zp_cnt",
        # "check_type", "keep",
    ]]
    
    return rt

###########################################################################
def _func_calc_binning_woe_stats_compare_BoostrapSampling(
        in_df, var_name, target_label,
        # list_bins_cnt=[20, 6, 5, 4, 3],
        list_bins_cnt=[100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 9, 8, 7, 6, 5, 4, 3],
        random_seed=9999,
        sampling_n=3000,
        sampling_epochs=100,
        sample_replace=False,
        verbose=True,
        with_target_balance=False,
    ):
    
    ################################################
    if verbose:
        # print("-"*60)
        print("\n>>>>  BoostrapSampling Processing: {}".format(var_name))
        _time = time.time()
    
    df = in_df[[var_name, target_label]].reset_index(drop=True)
    sample_replace = (True if sampling_n>df.shape[0] else sample_replace)
    # sample_replace = (True if int(sampling_n/2)>df.shape[0] else sample_replace)
    
    ################################################
    df_t_auto_binning_compare_result_bs = pd.DataFrame([])
    for idx in range(sampling_epochs):
        _time_0 = time.time()
        _random_seed = random_seed+idx
        
        if with_target_balance:
            _df = pd.concat(
                [
                    df.query("{}=='0_good'".format(target_label)).sample(n=int(sampling_n/2), replace=True, random_state=_random_seed, ignore_index=True),
                    df.query("{}=='1_bad'".format(target_label)).sample(n=int(sampling_n/2), replace=True, random_state=_random_seed, ignore_index=True),
                ],
                ignore_index=True,
            )
        else:
            _df = df.sample(n=sampling_n, replace=sample_replace, random_state=_random_seed, ignore_index=True)
        
        ################################################
        _df_t_auto_binning_compare_result = _func_calc_binning_woe_stats_compare(
            in_df=_df,
            var_name=var_name,
            target_label=target_label,
            list_bins_cnt=list_bins_cnt,
        )
        _df_t_auto_binning_compare_result = _df_t_auto_binning_compare_result[
            _df_t_auto_binning_compare_result["IV_notna"].notna()
        ].reset_index(drop=True)
        _df_t_auto_binning_compare_result.insert(loc=0, column="bs_random_seed", value=_random_seed)
        _df_t_auto_binning_compare_result.insert(loc=0, column="bs_idx", value=idx)
        
        df_t_auto_binning_compare_result_bs = df_t_auto_binning_compare_result_bs.append(
            _df_t_auto_binning_compare_result,
            ignore_index=True,
        )
        
        ################################################
        if verbose:
            print("    [ {:.2f}, {:.2f} ] {:7.2f}%".format(time.time()-_time, time.time()-_time_0, idx/sampling_epochs*100))
    
    ################################################
    if verbose:
        # print("\n")
        print("\n    cost time: {} sec.\n".format(time.time()-_time))
    
    return df_t_auto_binning_compare_result_bs


###########################################################################
# 筛选需要分析的数值型特征清单
cols_v_mid = cols_v_num[:]

# cols_v_mid = [
#     s0 for s0 in cols_v
#     if s0 in df_t_describe_summary.query("data_type=='Numerical'")["column_name"].tolist()
# ]

###########################################################################
# 计算
_time = time.time()
df_t_auto_binning_compare_result = pd.DataFrame([])
for _idx, _column_name in enumerate(cols_v_mid[:]):
    print("[ {:.2f} ] {:7.2f}%: {}".format(time.time()-_time, _idx/len(cols_v_num[:])*100, _column_name))
    
    ##########################################################################
    if _df_wb_dev[_column_name].notna().sum()==0:
        continue
    
    ##########################################################################
    _t_auto_binning_compare_result = _func_calc_binning_woe_stats_compare(
        in_df=_df_wb_dev,
        var_name=_column_name,
        target_label="target_label",
        # list_bins_cnt=[20, 6, 5, 4, 3],
        list_bins_cnt=[20, 10, 9, 8, 7, 6, 5, 4, 3],
        # list_bins_cnt=[100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 9, 8, 7, 6, 5, 4, 3],
        # list_bins_cnt=list(np.arange(3, 100+1)),
    )
    _t_auto_binning_compare_result.insert(
        loc=list(_t_auto_binning_compare_result.columns).index("column_name")+1,
        column="binning_method", value="num01_eq"
    )
    
    ##########################################################################
    with_boostrap_sampling = False
    # with_boostrap_sampling = True
    ##########################################################################
    if with_boostrap_sampling:
        _t_auto_binning_compare_result_bs = _func_calc_binning_woe_stats_compare_BoostrapSampling(
            in_df=df_wb,
            var_name=_column_name,
            target_label="target_label",
            # list_bins_cnt=[20, 6, 5, 4, 3],
            list_bins_cnt=[20, 10, 9, 8, 7, 6, 5, 4, 3],
            # list_bins_cnt=[100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 9, 8, 7, 6, 5, 4, 3],
            # list_bins_cnt=list(np.arange(3, 100+1)),
            random_seed=9999,
            # random_seed=99999,
            # sampling_n=300,
            # sampling_n=3000,
            # sampling_n=10000,
            # sampling_n=int(df_wb.shape[0]*0.6),
            sampling_n=int(df_wb.shape[0]*0.1),
            # sampling_epochs=100,
            sampling_epochs=30,
            # sampling_epochs=20,
            sample_replace=False,
            verbose=True,
            # verbose=False,
            with_target_balance=False,
            # with_target_balance=True,
        )
        _t_auto_binning_compare_result_bs["crosstab"] = _t_auto_binning_compare_result_bs["crosstab"].apply(
            lambda s0: (None if s0 is None else s0.to_dict(orient="records", into=dict))
        )
        _t_auto_binning_compare_result = _t_auto_binning_compare_result.merge(
            right=_t_auto_binning_compare_result_bs.groupby(by=["column_name", "max_bins_cnt"]).apply(
                lambda s0: pd.Series(dict({
                    # "bs_epochs": s0.shape[0],
                    "bs_result": s0[[
                                        "bs_idx", "bs_random_seed",
                                        "bin", "output_cutpoint", "crosstab", "_wls_adj_stats_info",
                                        "binning_cnt_notna", "IV_notna", "IV_pre_bin_notna", "total_pct_cv", "total_pct_max",
                                        "corr_pearson", "corr_spearman", "corr_kendall", "zp_cnt",
                                    ]].to_dict(orient="records", into=dict),

                }))
            ).reset_index(drop=False),
            how="left", left_on=["column_name", "max_bins_cnt"], right_on=["column_name", "max_bins_cnt"],
        )
    
    ##########################################################################
    df_t_auto_binning_compare_result = df_t_auto_binning_compare_result.append(
        _t_auto_binning_compare_result,
        ignore_index=True,
    )

##########################################################################
print("\n")
print("cost time: {} sec.".format(time.time()-_time))




####################################################################################
# 针对上述分箱结果，进行严格单调性判断
_t_auto_binning_compare_result_chk = []

_df_0 = df_t_auto_binning_compare_result \
    [df_t_auto_binning_compare_result["zp_cnt"].notna()] \
    .reset_index(drop=True)

######################################################
# （分箱数在3～6之间选择）
# 判断corr_kendall in [1, -1]
# 确定存在严格单调的分箱处理，即严格单调判断
# 并且优先保留IV_pre_bin_notna最大、binning_cnt_notna最小的分箱
_df = _df_0 \
    .sort_values(by=["column_name", "binning_cnt_notna", "total_pct_cv"], ascending=[True, True, True]) \
    .drop_duplicates(subset=["column_name", "binning_cnt_notna"]) \
    [_df_0["corr_kendall"].apply(lambda s0: round(s0, 2) in [-1, 1])] \
    .query("binning_cnt_notna>={} and binning_cnt_notna<={}".format(3, 6)) \
    .sort_values(by=["column_name", "IV_pre_bin_notna", "binning_cnt_notna"], ascending=[True, False, True]) \
    .drop_duplicates(subset=["column_name"], keep="first") \
    .reset_index(drop=True)
_df["check_type"] = _df["corr_kendall"].apply(
    lambda s0: (
        "1_严格单调递增" if round(s0, 2)==1 else
        "2_严格单调递减" if round(s0, 2)==-1 else
        ""
    )
)

_t_auto_binning_compare_result_chk.append(_df)

######################################################
# （分箱数在4～6之间选择）
# 判断corr_kendall not in [1, -1]，并且存在一个零点(zp_cnt==1)
# 确定存在非严格单调、有且仅有一个零点（拐点）的分箱处理，即U型趋势判断
# 并且优先保留IV_pre_bin_notna最大、binning_cnt_notna最小的分箱
_df = _df_0 \
    .sort_values(by=["column_name", "binning_cnt_notna", "total_pct_cv"], ascending=[True, True, True]) \
    .drop_duplicates(subset=["column_name", "binning_cnt_notna"]) \
    [_df_0["corr_kendall"].apply(lambda s0: round(s0, 2) not in [-1, 1])] \
    .query("zp_cnt==1 and binning_cnt_notna>={} and binning_cnt_notna<={}".format(4, 6)) \
    .sort_values(by=["column_name", "IV_pre_bin_notna", "binning_cnt_notna"], ascending=[True, False, True]) \
    .drop_duplicates(subset=["column_name"], keep="first") \
    .reset_index(drop=True)
_df["check_type"] = "3_U型"

_t_auto_binning_compare_result_chk.append(_df)

######################################################
# 结果汇总
# 考虑筛选IV_notna大于等于特定阈值的变量分箱结果
_t_auto_binning_compare_result_chk = pd.concat(
    _t_auto_binning_compare_result_chk,
    ignore_index=True,
) \
    .query("IV_notna>={}".format(0.02)) \
    .sort_values(by=["column_name", "zp_cnt", "IV_pre_bin_notna"], ascending=[True, True, False]) \
    .drop_duplicates(subset=["column_name"], keep="first") \
    .reset_index(drop=True)

_t_auto_binning_compare_result_chk["keep"] = "Y"

df_t_auto_binning_compare_result_chk = pd.merge(
    left=_df_0,
    right=_t_auto_binning_compare_result_chk[["column_name", "max_bins_cnt", "check_type", "keep"]],
    how="left", left_on=["column_name", "max_bins_cnt"], right_on=["column_name", "max_bins_cnt"],
) \
    .sort_values(by=["keep", "check_type", "IV_notna"], ascending=[False, True, False]) \
    .reset_index(drop=True)

# df_t_auto_binning_compare_result_chk["output_cutpoint"] = df_t_auto_binning_compare_result_chk.apply(
#     lambda s0: (
#         [-np.inf]+pd.Series([
#             float(re.sub("[ \]]", "", s0.split("_")[1].split(",")[1]))
#             for s0 in s0["crosstab"].index
#             if re.search("(NaN, NaN)|(, inf)|(^total$)", s0)==None
#         ]).drop_duplicates().tolist()+[np.inf]
#     ),
#     axis=1,
# )

df_t_auto_binning_compare_result_chk[["check_type", "keep"]] = \
    df_t_auto_binning_compare_result_chk[["check_type", "keep"]].fillna("")

df_t_auto_binning_compare_result_chk = df_t_auto_binning_compare_result_chk \
    [[
        "column_name", "binning_method", "max_bins_cnt", "bin", "crosstab", "output_cutpoint",
        "binning_cnt", "binning_cnt_notna", "IV_notna", "IV_pre_bin_notna", "total_pct_cv", "total_pct_max",
        "corr_pearson", "corr_spearman", "corr_kendall", "zp_cnt", "check_type", "keep",
    ]].reset_index(drop=True)


####################################################################################
print(df_t_auto_binning_compare_result_chk.shape[0], df_t_auto_binning_compare_result_chk.query("keep=='Y'").shape[0])
print()
print(df_t_auto_binning_compare_result_chk["check_type"].value_counts().sort_index())








### （结果导出）

# ####################################################################################
# # to_clipboard
# df_t_auto_binning_compare_result_chk.to_clipboard(index=False)
# df_t_auto_binning_compare_result_chk.drop(labels=["crosstab"], axis=1).to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result.xlsx".format(result_path), mode="wb") as fw:
    df_t_auto_binning_compare_result \
        .drop(labels=["crosstab"], axis=1) \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result_chk.xlsx".format(result_path), mode="wb") as fw:
    df_t_auto_binning_compare_result_chk \
        .drop(labels=["crosstab"], axis=1) \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_auto_binning_compare_result, file=fw)
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result_chk.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_auto_binning_compare_result_chk, file=fw)


####################################################################################
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result.pkl".format(result_path), mode="rb") as fr:
    df_t_auto_binning_compare_result = pickle.load(file=fr)
with open(file="{}/Proc04_WoeMonotonic/df_t_auto_binning_compare_result_chk.pkl".format(result_path), mode="rb") as fr:
    df_t_auto_binning_compare_result_chk = pickle.load(file=fr)

print(df_t_auto_binning_compare_result_chk.shape[0], df_t_auto_binning_compare_result_chk.query("keep=='Y'").shape[0])
print()
print(df_t_auto_binning_compare_result_chk["check_type"].value_counts().sort_index())


































# df_t_auto_binning_compare_result_bs = _func_calc_binning_woe_stats_compare_BoostrapSampling(
#     in_df=df_wb,
#     var_name="jcxx_reg_capital",
#     target_label="target_label",
#     list_bins_cnt=[20, 10, 9, 8, 7, 6, 5, 4, 3],
#     random_seed=9999,
#     sampling_n=3000,
#     sampling_epochs=30,
#     sample_replace=False,
#     verbose=True,
#     with_target_balance=False,
# )





# _col = "jcxx_reg_capital"

# sns.scatterplot(
#     x=df_t_auto_binning_compare_result.query("zp_cnt!='' and column_name=='{}'".format(_col))["max_bins_cnt"],
#     y=df_t_auto_binning_compare_result.query("zp_cnt!='' and column_name=='{}'".format(_col))["_wls_adj_stats_info"].apply(lambda s0: s0["_WOE_RMSE"]),
#     label="_WOE_RMSE",
# )
# sns.scatterplot(
#     x=df_t_auto_binning_compare_result.query("zp_cnt!='' and column_name=='{}'".format(_col))["max_bins_cnt"],
#     y=df_t_auto_binning_compare_result.query("zp_cnt!='' and column_name=='{}'".format(_col))["_wls_adj_stats_info"].apply(lambda s0: s0["_WOE_MAE"]),
#     label="_WOE_MAE",
# )











#==============================================================================
# File: code06_Proc05_FeatCoarseBinning_ind.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc05_FeatCoarseBinning_ind
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc05_FeatCoarseBinning_ind
## 特征粗分箱（手动调整，单指标处理）


####################################################################
# Proc05_FeatCoarseBinning_ind
# 特征粗分箱（手动调整，单指标处理）


### 分箱配置（Categorical）

###########################################################################
# # 分箱配置（categorical）
# dict_coarse_bin_mapping_categorical = {
    
#     "jcxx_tax_credit_rating": {
#         "NaN": 0,
#         "暂无": 0,
#         "A": 1,
#         "B": 2,
#         "C": 3,
#         "D": 3,
#         "M": 3,
#     },
#     "jcxx_nsr_types": {
#         "NaN": 0,
#         "一般纳税人": 1,
#         "小规模纳税人": 2,
#     },
#     "jcxx_tax_auth_prov": {
#         "NaN": 0,
#         "福建省": 1,
#         "广东省": 1,
#         "浙江省": 2,
#         "江苏省": 2,
#     },
# #     "jcxx_nsr_status": {
# #         "NaN": 0,
# #         "注销": 1,
# #         "简易注销无异议": 1,
# #         "非正常": 1,
# #         "停业": 2,
# #         "正常": 2,
# #         "清算": 2,
# #     },
    
# }


# dict_en_cn['income_izai_tmf']

### 分箱配置（Numerical）

###########################################################################
# 分箱配置（Numerical）
inf = np.inf
dict_coarse_bin_mapping_numerical = {
    
    # "三方数据源_baihang_1":[-inf, 600, 700, 800, inf],
    # "三方数据源_bairong_14":[-inf, 640, 700, 780, inf],  # 
    "三方数据源_bairong_14":[-inf, 600, 660, 700, 770, inf],  # 
    "三方数据源_bairong_15":[-inf, 580, 640, 680, 730, inf],  # 
    # "三方数据源_duxiaoman_1":[-inf, 520, 530, 540, 550, 560, 570, 580, inf],
    # "三方数据源_hangliezhi_1":[-inf, 480, 495, 515.0, inf],    # 
    # "三方数据源_hangliezhi_1":[-inf, 495, 515, 535, inf],    # 
    "三方数据源_hangliezhi_1":[-inf, 495, 505, 515, 535, inf],    # 
    
    # "三方数据源_hengpu_4":[-inf, 25, 30, 35, 40, inf],  # 
    "三方数据源_hengpu_4":[-inf, 25, 30, 35, 40, 45, inf],  # 
    # 缺失高
    # "三方数据源_hengpu_5":[-inf, 0.4, 0.5, 0.6, inf],
    # "三方数据源_HL_A1":[-inf, 652.0, 663.0, 677.0, inf],    
    # "三方数据源_HL_A2":[-inf, 597.0, 607.0, inf],    
    # "三方数据源_HL_A3":[-inf, 597.0, 607.0, inf],   
    # "三方数据源_HL_A4":[-inf, 597.0, 607.0, inf],  
    # "三方数据源_rong360_4":[-inf, 0.03, 0.04, 0.05, 0.06, inf],  
    
    "三方数据源_ruizhi_6":[-inf, 720, 765, 800, 840, inf],  #
    "三方数据源_xinyongsuanli_1":[-inf, 620, 640, 670, 700, inf],  # 
    # "三方数据源_xinyongsuanli_1":[-inf, 640, 670, 700, inf],  # 

    "is_cheDai":[-inf, 0, inf],  
    "education":[-inf, 20,  inf],  # 

    # 贷款近9个月连续逾期月份数
    "loan_overdue_last_months_l9m":[-inf, 0, inf],  # 
    # 近12个月内由商业银行发放查询次数
    "quae_xcva_p5":[-inf, 5, 10, 20, inf],
    
    
    # 近6个月贷款审批最近一次查询距今时长
    # 中融信映射
    # 0 --> 0
    # (0, 1] --> 1
    # (1, 2] --> 2
    # (2, 4] --> 3
    # (4, 6] --> 4
    # (6, +)--> 5
    # 中融信
    # 中融信是月？实际数字是天？？？
    # 海尔
    # "query_tfau_mg":[-inf, 1, 10, 30, inf],  # 
    "query_tfau_mg":[-inf, 0, 10, 30, inf],  # 
    # "query_tfau_mg":[-inf, 0, 1, 2, 3, 5, 10, 20, 30, 60, inf],  # 
    
    
    # 贷款发放月份数最大值
    "Loan0020":[-inf, 24, 36, 48, 60, 72, 96, 120, inf],  
    # 所有非循环贷贷款金额合计
    "coaa_zbva_xavh_bbvf_n9": [-inf, 1e5, 10e5, inf],   # 
    # 所有信贷当前应还款总金额
    "debts_sgac_mc": [-inf, 2e4, 4e4, 6e4, inf],  # 
    # "pboc_education_new": [-inf, 5, inf],  #  
    "pboc_education_new": [-inf, 3, 5, inf],  #  
    
    # 发放时间在60个月内人民币账户销户状态信用卡机构数
    "rebo_zcva_bfvd_n8": [-inf, 0, inf],  
    # 贷记卡近9个月连续逾期月份数
    "cc_overdue_last_months_l9m": [-inf, 0, inf], 
    # 正常贷记卡发放月份数最大值
    "CARD0004": [-inf, 84, 120, 144, inf],
    
    # 贷款首次逾期距今时长(最长月份) (报告日期)间隔月份
    # 中融信映射
    # 0 --> 0
    # (0, 3] --> 1
    # (3, 6] --> 2
    # (6, 24] --> 3
    # (24, 36] --> 4
    # (36, 48] --> 5
    # (48, 60] --> 6
    # (60, 90] --> 7
    # (90, 120] --> 8
    # (120, +) --> 9    
    # 中融信
    # "repay_dcal_tmd": [-inf, 0, 3, 6, 24, 36, 48, 60, 90, 120, inf],  # 
    # 海尔
    # "repay_dcal_tmd": [-inf, 30, 50, inf],  # 
    "repay_dcal_tmd": [-inf, 20, 35, 50, inf],  # 
    
    # 贷款近12个月连续逾期月份数
    "loan_overdue_last_months_l12m": [-inf, 0, inf],  # 
    "loan_overdue_last_months_l6m": [-inf, 0, inf],
    "loan_overdue_last_months_l3m": [-inf, 0, inf],
    

    # '最近3个月审批查询次数/最近12个月审批查询次数'
    "INQUIRY0047": [-inf, 0.35, 0.5, inf],  # 
    # 正常贷记卡余额比共享额度均值
    "CARD0009": [-inf, 0.35, 0.6, 0.99, inf],
    # 贷记卡近12个月连续逾期月份数
    "cc_overdue_last_months_l12m": [-inf, 0, inf],  # 
    "cc_overdue_last_months_l6m": [-inf, 0, inf],  # 
    
    
    
    # 所有信贷当前实际还款总金额占应还款总金
    # 中融信映射
    # [0, 0.25) --> 0
    # [0.25, 0.5) --> 1
    # [0.5, 0.75) --> 2
    # [0.75, 1) --> 3
    # [1, 2) --> 4
    # [2, 4) --> 5
    # [4, +) --> 6
    # 中融信
    # "debts_sgad_mc": [-inf, 0.25, 0.5, 0.75, 1, 2, 4, inf],  # 
    # 海尔
    # "debts_sgad_mc": [-inf, 1, inf],  # 
    "debts_sgad_mc": [-inf, 1, 1.2, 1.8, inf],  # 
    # "debts_sgad_mc": [-inf, 0.25, 0.50, 0.75, 1, 1.2, 1.5, 1.8, 2.5, 4, inf],  # 
    
    # 贷款剩余总还款期数
    # 中融信映射
    # 0 --> 0
    # (0, 3] --> 1
    # (3, 6] --> 2
    # (6, 12] --> 3
    # (12, 24] --> 4
    # (24, 48] --> 5
    # (48, 60] --> 6
    # (60, 90] --> 7
    # (90, 120] --> 8
    # (120, +) --> 9
    # 中融信
    # "repay_dhaa_md":[-inf, 0, 3, 6, 12, 24, 48, 60, 90, 120, inf], # 
    # 海尔
    "repay_dhaa_md":[-inf, 40,  150, inf], # 
    
    "pboc_marital_new":[-inf, 1, 2, 3, inf],
    "is_fangDai":[-inf, 0, inf],  
    # "result":[-inf, 0, inf],  
    # 近6个月出现逾期M1次数
    "overdue_m1_cnt_l6m_all":[-inf, 0, inf],  # 
    "overdue_m2_cnt_l6m_all":[-inf, 0, inf],  
    "overdue_m2_cnt_l12m_all":[-inf, 0, inf],  
    
    # 房屋贷款总金额（余额>0）
    "TYPEaSUMHOUS":[-inf, 0, 5e5, inf],  
    # 发放时间在6个月内贷款由所有机构发放的个人消费贷款已结清贷款金额合计
    "coaa_zbvg_xawd_bbvd_n3":[-inf, 0, 3e4, inf],  # 
    "cur_overdue_account_cnt":[-inf, 0, 1, inf],
    "AGE":[-inf, 25, 30, 35, 40, 45, 50, inf],
    # '正常贷记卡最近24个月逾期期数最大值'
    "CARD0065":[-inf, 0, inf],  # 
    # 贷款查询最近12个月查询机构数
    "Inquiry0023":[-inf, 2,10,15,20,inf],  
    # 上报时间在12个月内贷款正常状态贷款总余额
    "deaa_zbvg_bbvb_o4":[-inf, 1e4, 5e4, 10e4, 20e4,  50e4, inf],
    
    # 近1个月内查询机构类型数
    # 中融信映射
    # 0 --> 0
    # [1, 2) --> 1
    # [2, 3) --> 2
    # [3, 5) --> 3
    # [5, 8) --> 4
    # [8, 10) --> 5
    # [10, 15) --> 6
    # [15, 20) --> 7
    # [20, +) --> 8    
    # 中融信
    # "quad_p1":[-inf, 0, 2, 3, 5, 8, 10, 15, 20, inf],  # 
    # 海尔
    "quad_p1":[-inf, 1, 2, 3, inf],  # 
    
    
    # 担保资格审查最近2年查询次数
    "INQUIRY0064":[-inf, 1, 3, 6, inf],
    
    
    # 工资
    # 中融信映射
    # [0, 1) --> 0
    # [1, 1000) --> 1
    # [1000, 3000) --> 2
    # [3000, 5000) --> 3
    # [5000, 8000) --> 4
    # [8000, 10000) --> 5
    # [10000, 30000) --> 6
    # [30000, 50000) --> 7
    # [50000, 100000) --> 8
    # [100000, +) --> 9    
    # 中融信
    # "income_izai_tmf":[-inf, 0.99, 999, 2999, 4999, 7999, 9999, 29999, 49999, 99999, inf],  # 
    # 海尔
    # "income_izai_tmf":[-inf, 2000, 4000, 6000, inf],  # 
    "income_izai_tmf":[-inf, 2000, 4000, 8000, inf],  # 
    # "income_izai_tmf":[-inf, 1000, 2000, 3000, 4000, 5000, 7000, 10000, 15000, 20000, inf],  # 
    
    # "digitalScore":[-inf, 800, 825, 870, inf],  # 
    "digitalScore":[-inf, 775, 800, 825, 860, inf],  # 

    "pboc_sex_new":[-inf, 0, 1, inf],



}






### 单指标调整（Numerical）
### 查看woe趋势图

################################################################################################
# 单指标调整（Numerical）
# 查看woe趋势图

_col = "query_tfau_mg"
try:
    print(dict_en_cn[_col])
except:
    pass
# _col = random.choice(list(dict_coarse_bin_mapping_numerical.keys()))
# _col = random.choice(cols_v_num)
# _bin = [-inf, 0, 0.3, 1, inf]
_bin = dict_coarse_bin_mapping_numerical.get(_col)
# _col, _bin = list(dict_coarse_bin_mapping_numerical.items())[-1]

_df_wb_oot


################################################################################################
# _f_auto = True
_f_auto = False
if _f_auto:
    ######################################################
    _, _crosstab, _bin = func_auto_binning_continuous_v1(
        in_var=_df_wb_dev[_col],
        in_target=_df_wb_dev["target_label"],
        
        min_pct=0.05-0.0001, max_bins_cnt=20,
        # min_pct=0.05-0.0001, max_bins_cnt=6,
        # min_pct=0.05-0.0001, max_bins_cnt=5,
        # min_pct=0.05-0.0001, max_bins_cnt=4,
        # min_pct=0.05-0.0001, max_bins_cnt=3,
        # min_pct=0.1-0.0001, max_bins_cnt=9,
        # min_pct=0.01-0.0001, max_bins_cnt=50,
        method="01_equal_freq",
        # method="02_decision_tree",
        # method="03_chi2_comb",
        # method="09_best_ks",
        chi2_min_pvalue=0.1,
        chi2_min_cnt=2,
        with_lift_ks=False,
        # with_lift_ks=True,
        lift_calc_ascending=True,
        # with_wls_adj_woe=True,
        woe_with_nan=True,
        # woe_with_nan=False,
    )
 
else:
    ######################################################
    _crosstab = func_woe_report_v1(
        in_var=func_binning_continuous_v1(
            in_data=_df_wb_dev[_col],
            # in_data=_df_wb_oot[_col],
            
            bins=_bin,
            right_border=True, include_lowest=True,
        ),
        in_target=_df_wb_dev["target_label"],
        # in_target=_df_wb_oot["target_label"],
        
        with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.0001,
        # with_lift_ks=True,
        with_lift_ks=False,
        lift_calc_ascending=True,
        # with_wls_adj_woe=True,
        woe_with_nan=True,
        # woe_with_nan=False,
    )


print(_col)
print(_bin)
print()
# func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=False
#               # , with_wls_adj_woe=True
#               )
func_plot_woe_1(_crosstab, _col, plot_badrate=False, with_nan_info=False
              # , with_wls_adj_woe=True
              )
# func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=True, with_wls_adj_woe=True)
# func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=False, with_wls_adj_woe=False)

_crosstab.to_clipboard(index=True)
_crosstab










#==============================================================================
# File: code07_Proc05_FeatCoarseBinning_batch.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc05_FeatCoarseBinning_batch
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc05_FeatCoarseBinning_batch
## 特征粗分箱（手动调整，批量指标处理）


####################################################################
# Proc05_FeatCoarseBinning_batch
# 特征粗分箱（手动调整，批量指标处理）


### 分箱配置（Categorical）

###########################################################################
# 分箱配置（categorical）
# dict_coarse_bin_mapping_categorical = {
    
#     "jcxx_tax_credit_rating": {
#         "NaN": 0,
#         "暂无": 0,
#         "A": 1,
#         "B": 2,
#         "C": 3,
#         "D": 3,
#         "M": 3,
#     },
#     "jcxx_nsr_types": {
#         "NaN": 0,
#         "一般纳税人": 1,
#         "小规模纳税人": 2,
#     },
#     "jcxx_tax_auth_prov": {
#         "NaN": 0,
#         "福建省": 1,
#         "广东省": 1,
#         "浙江省": 2,
#         "江苏省": 2,
#     },
# #     "jcxx_nsr_status": {
# #         "NaN": 0,
# #         "注销": 1,
# #         "简易注销无异议": 1,
# #         "非正常": 1,
# #         "停业": 2,
# #         "正常": 2,
# #         "清算": 2,
# #     },
    
# }


### 分箱配置（Numerical）

###########################################################################
# 分箱配置（Numerical）
inf = np.inf
dict_coarse_bin_mapping_numerical = {
 
    # "三方数据源_baihang_1":[-inf, 600, 700, 800, inf],
    # "三方数据源_bairong_14":[-inf, 640, 700, 780, inf],  # 
    "三方数据源_bairong_14":[-inf, 600, 660, 700, 770, inf],  # 
    "三方数据源_bairong_15":[-inf, 580, 640, 680, 730, inf],  # 
    # "三方数据源_duxiaoman_1":[-inf, 520, 530, 540, 550, 560, 570, 580, inf],
    # "三方数据源_hangliezhi_1":[-inf, 480, 495, 515.0, inf],    # 
    # "三方数据源_hangliezhi_1":[-inf, 495, 515, 535, inf],    # 
    "三方数据源_hangliezhi_1":[-inf, 495, 505, 515, 535, inf],    # 
    
    # "三方数据源_hengpu_4":[-inf, 25, 30, 35, 40, inf],  # 
    "三方数据源_hengpu_4":[-inf, 25, 30, 35, 40, 45, inf],  # 
    # 缺失高
    # "三方数据源_hengpu_5":[-inf, 0.4, 0.5, 0.6, inf],
    # "三方数据源_HL_A1":[-inf, 652.0, 663.0, 677.0, inf],    
    # "三方数据源_HL_A2":[-inf, 597.0, 607.0, inf],    
    # "三方数据源_HL_A3":[-inf, 597.0, 607.0, inf],   
    # "三方数据源_HL_A4":[-inf, 597.0, 607.0, inf],  
    # "三方数据源_rong360_4":[-inf, 0.03, 0.04, 0.05, 0.06, inf],  
    
    "三方数据源_ruizhi_6":[-inf, 720, 765, 800, 840, inf],  #
    "三方数据源_xinyongsuanli_1":[-inf, 620, 640, 670, 700, inf],  # 
    # "三方数据源_xinyongsuanli_1":[-inf, 640, 670, 700, inf],  # 

    "is_cheDai":[-inf, 0, inf],  
    "education":[-inf, 20,  inf],  # 

    # 贷款近9个月连续逾期月份数
    "loan_overdue_last_months_l9m":[-inf, 0, inf],  # 
    # 近12个月内由商业银行发放查询次数
    "quae_xcva_p5":[-inf, 5, 10, 20, inf],
    # 近6个月贷款审批最近一次查询距今时长
    # "query_tfau_mg":[-inf, 1, 10, 30, inf],  # 
    "query_tfau_mg":[-inf, 0, 10, 30, inf],  # 
    # 贷款发放月份数最大值
    "Loan0020":[-inf, 24, 36, 48, 60, 72, 96, 120, inf],  
    # 所有非循环贷贷款金额合计
    "coaa_zbva_xavh_bbvf_n9": [-inf, 1e5, 10e5, inf],   # 
    # 所有信贷当前应还款总金额
    "debts_sgac_mc": [-inf, 2e4, 4e4, 6e4, inf],  # 
    # "pboc_education_new": [-inf, 5, inf],  #  
    "pboc_education_new": [-inf, 3, 5, inf],  #  
    
    # 发放时间在60个月内人民币账户销户状态信用卡机构数
    "rebo_zcva_bfvd_n8": [-inf, 0, inf],  
    # 贷记卡近9个月连续逾期月份数
    "cc_overdue_last_months_l9m": [-inf, 0, inf], 
    # 正常贷记卡发放月份数最大值
    "CARD0004": [-inf, 84, 120, 144, inf],
    # 贷款首次逾期距今时长(最长月份) (报告日期)间隔月份
    # "repay_dcal_tmd": [-inf, 30, 50, inf],  # 
    "repay_dcal_tmd": [-inf, 20, 35, 50, inf],  # 
    
    # 贷款近12个月连续逾期月份数
    "loan_overdue_last_months_l12m": [-inf, 0, inf],  # 
    "loan_overdue_last_months_l6m": [-inf, 0, inf],
    "loan_overdue_last_months_l3m": [-inf, 0, inf],
    

    # '最近3个月审批查询次数/最近12个月审批查询次数'
    "INQUIRY0047": [-inf, 0.35, 0.5, inf],  # 
    # 正常贷记卡余额比共享额度均值
    "CARD0009": [-inf, 0.35, 0.6, 0.99, inf],
    # 贷记卡近12个月连续逾期月份数
    "cc_overdue_last_months_l12m": [-inf, 0, inf],  # 
    "cc_overdue_last_months_l6m": [-inf, 0, inf],  # 
    
    
    # 所有信贷当前实际还款总金额占应还款总金
    # "debts_sgad_mc": [-inf, 1, inf],  # 
    "debts_sgad_mc": [-inf, 1, 1.2, 1.8, inf],  # 
    
    # 贷款剩余总还款期数
    "repay_dhaa_md":[-inf, 40,  150, inf], # 
    # "repay_dhaa_md":[-inf, 40, 120, inf], # 
    
    "pboc_marital_new":[-inf, 1, 2, 3, inf],
    "is_fangDai":[-inf, 0, inf],  
    # "result":[-inf, 0, inf],  
    # 近6个月出现逾期M1次数
    "overdue_m1_cnt_l6m_all":[-inf, 0, inf],  # 
    "overdue_m2_cnt_l6m_all":[-inf, 0, inf],  
    "overdue_m2_cnt_l12m_all":[-inf, 0, inf],  
    
    # 房屋贷款总金额（余额>0）
    "TYPEaSUMHOUS":[-inf, 0, 5e5, inf],  
    # 发放时间在6个月内贷款由所有机构发放的个人消费贷款已结清贷款金额合计
    "coaa_zbvg_xawd_bbvd_n3":[-inf, 0, 3e4, inf],  # 
    "cur_overdue_account_cnt":[-inf, 0, 1, inf],
    "AGE":[-inf, 25, 30, 35, 40, 45, 50, inf],
    # '正常贷记卡最近24个月逾期期数最大值'
    "CARD0065":[-inf, 0, inf],  # 
    # 贷款查询最近12个月查询机构数
    "Inquiry0023":[-inf, 2,10,15,20,inf],  
    # 上报时间在12个月内贷款正常状态贷款总余额
    "deaa_zbvg_bbvb_o4":[-inf, 1e4, 5e4, 10e4, 20e4,  50e4, inf],
    # 近1个月内查询机构类型数
    "quad_p1":[-inf, 1, 2, 3, inf],  # 
    
    
    # 担保资格审查最近2年查询次数
    "INQUIRY0064":[-inf, 1, 3, 6, inf],
    # "income_izai_tmf":[-inf, 2000, 4000, 6000, inf],  # 
    "income_izai_tmf":[-inf, 2000, 4000, 8000, inf],  # 
    
    # "digitalScore":[-inf, 800, 825, 870, inf],  # 
    "digitalScore":[-inf, 775, 800, 825, 860, inf],  # 

    "pboc_sex_new":[-inf, 0, 1, inf],

}


### 离散型特征（Categorical）
### 粗分箱处理：根据分箱配置dict_coarse_bin_mapping_categorical进行woe计算

################################################################################################
# 粗分箱处理：根据分箱配置dict_coarse_bin_mapping_categorical进行woe计算

def _func_calc_coarse_iv_categorical(in_df, var_name, target_label, mapping):
    rt = OrderedDict({
        "crosstab": None,
        "IV": None,
        "IV_dropna": None,
        "boundary": None,
        "mapping_gb_class": mapping,
    })
    
    _data_converted = func_combining_discrete_v1(
        in_data=in_df[var_name],
        mapping_gb_class=mapping,
        fillna_value="NaN", cvt_fillna_value=mapping.get("NaN", 0),
    )
    rt["crosstab"] = func_woe_report_v1(
        in_var=_data_converted,
        in_target=_df_wb_dev[target_label],
        with_total=True,
        with_lift_ks=False,
        lift_calc_ascending=False,
        with_wls_adj_woe=False,
        woe_with_nan=True,
    )
    
    _crosstab = rt["crosstab"]
    rt["IV"] = _crosstab.loc["total", "IV"]
    rt["IV_dropna"] = _crosstab.loc[[s0 for s0 in _crosstab.index if re.search("(_NaN$)|(NaN, NaN)|(^total$)", s0)==None], "IV"].sum()
    # _d = _crosstab["WOE"].reset_index(drop=True).dropna().reset_index().values
    # _X = _d[:, 0]
    # _y = _d[:, 1]
    return rt

# _df_t_features_coarse_iv_cate = df_t_describe_summary \
#     .query("data_type=='Categorical'") \
#     [df_t_describe_summary["column_name"].isin((list(dict_coarse_bin_mapping_categorical.keys()))[:])] \
#     [["column_name", "data_type"]].reset_index(drop=True)
# _df_t_features_coarse_iv_cate["binning_method"] = "coarse_binning"
# _df_t_features_coarse_iv_cate = _df_t_features_coarse_iv_cate[["binning_method", "column_name", "data_type"]]


##########################################################################
# 计算IV
_time = time.time()
_idx = 1

# _coarse_binning_result = []
# for _column_name in _df_t_features_coarse_iv_cate["column_name"].tolist()[:]:
#     _mapping = dict_coarse_bin_mapping_categorical.get(_column_name)
#     _coarse_binning_result.append(
#         _func_calc_coarse_iv_categorical(
#             in_df=_df_wb_dev,
#             var_name=_column_name,
#             target_label="target_label",
#             mapping=_mapping,
#         )
#     )
#     print("{:7.2f}%: {}".format(_idx/_df_t_features_coarse_iv_cate.shape[0]*100, _column_name))
#     _idx = _idx+1

# print("\n")
# print("cost time: {} sec.".format(time.time()-_time))

# ##########################################################################
# _df_t_features_coarse_iv_cate["IV"] = [s0["IV"] for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_cate["IV_dropna"] = [s0["IV_dropna"] for s0 in _coarse_binning_result]
# # _df_t_features_coarse_iv_cate["boundary"] = [([-np.inf, np.inf] if s0["boundary"]==None else [round(t, 4) for t in s0["boundary"]]) for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_cate["boundary"] = [([round(t, 4) for t in s0["boundary"]] if s0["boundary"]!=None else None) for s0 in _coarse_binning_result]
# # _df_t_features_coarse_iv_cate["boundary"] = [([t for t in s0["boundary"]] if s0["boundary"]!=None else None) for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_cate["mapping_gb_class"] = [s0["mapping_gb_class"] for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_cate["crosstab"] = [s0["crosstab"] for s0 in _coarse_binning_result]
# # _df_t_features_coarse_iv_cate["bin_cnt"] = _df_t_features_coarse_iv_cate["crosstab"].apply(lambda s0: s0[s0.index!="total"].shape[0])
# _df_t_features_coarse_iv_cate["output_cutpoint"] = _df_t_features_coarse_iv_cate.apply(
#     lambda s0: (
#         dict(
#             [
#                 (_idx, int(_v))
#                 for _idx, _v in
#                 np.concatenate([
#                     list(product(_v.split("/"), [_idx]))
#                     for _idx, _v in
#                     [s0.split("_") for s0 in s0["crosstab"].index if re.search("(^total$)", s0)==None]
#                 ])
#             ]
#         )
#         if s0["data_type"]=="Categorical" else
#         [-np.inf]+pd.Series([
#             float(re.sub("[ \]]", "", s0.split("_")[1].split(",")[1]))
#             for s0 in s0["crosstab"].index
#             if re.search("(NaN, NaN)|(, inf)|(^total$)", s0)==None
#         ]).drop_duplicates().tolist()+[np.inf]
#         if s0["data_type"]=="Numerical" else
#         None
#     ),
#     axis=1,
# )
# _df_t_features_coarse_iv_cate["crosstab"] = _df_t_features_coarse_iv_cate.apply(
#     lambda s0: (
#         func_woe_report_v1(
#             in_var=func_combining_discrete_v1(
#                 in_data=_df_wb_dev[s0["column_name"]],
#                 mapping_gb_class=s0["output_cutpoint"],
#                 fillna_value="NaN", cvt_fillna_value=s0["output_cutpoint"].get("NaN", 0),
#             ),
#             in_target=_df_wb_dev["target_label"],
#             with_total=True,
#             with_lift_ks=True,
#             lift_calc_ascending=True,
#             with_wls_adj_woe=False,
#             woe_with_nan=True,
#         )
#         if s0["data_type"]=="Categorical" else
#         func_woe_report_v1(
#             in_var=func_binning_continuous_v1(
#                 in_data=_df_wb_dev[s0["column_name"]],
#                 bins=s0["output_cutpoint"],
#                 right_border=True, include_lowest=True,
#             ),
#             in_target=_df_wb_dev["target_label"],
#             with_total=True,
#             with_lift_ks=True,
#             lift_calc_ascending=True,
#             with_wls_adj_woe=True,
#             woe_with_nan=True,
#         )
#         if s0["data_type"]=="Numerical" else
#         None
#     ),
#     axis=1,
# )
# _df_t_features_coarse_iv_cate["bin_cnt"] = _df_t_features_coarse_iv_cate["crosstab"].apply(lambda s0: s0[s0.index!="total"].shape[0])

# _df_t_features_coarse_iv_cate["_output_object"] = _df_t_features_coarse_iv_cate.apply(
#     lambda s0: {
#         "feature_name": s0["column_name"],
#         "data_type": s0["data_type"],
#         "binning_method": s0["binning_method"],
#         "boundary": s0["boundary"],
#         "mapping_gb_class": s0["mapping_gb_class"],
#         "crosstab_index": [s0 for s0 in s0["crosstab"].index],
#         "crosstab_cnt": s0["crosstab"] \
#             [["0_good_#", "1_bad_#"]] \
#             .to_dict(orient="list"),
#         "crosstab_details": s0["crosstab"] \
#             [(
#                 ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate"]
#                 if s0["data_type"]=="Categorical" else
#                 ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
#                 if s0["data_type"]=="Numerical" else
#                 []
#             )] \
#             .to_dict(orient="list"),
#         "IV": s0["IV"],
#         "IV_dropna": s0["IV_dropna"],
#         "output_cutpoint": s0["output_cutpoint"],
#     },
#     axis=1,
# )

# _df_t_features_coarse_iv_cate = _df_t_features_coarse_iv_cate.sort_values(by=["IV"], ascending=False).reset_index(drop=True)


###########################################################################
# 计算开发样本的WOE报告
# _t_features_coarse_woe_report_cate = []
# for _column_name, _data_type, _crosstab, _boundary, _mapping_gb_class in \
#             _df_t_features_coarse_iv_cate[["column_name", "data_type", "crosstab", "boundary", "mapping_gb_class"]].values[:]:
#     _crosstab = _crosstab.reset_index().rename(columns={"index": "gb_idx"}).query("gb_idx!='total'")
#     _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx"), column="feature_name", value=_column_name)
#     _crosstab.insert(loc=_crosstab.columns.tolist().index("feature_name"), column="data_type", value=_data_type)
#     # if _data_type=="Categorical":
#     #     _mapping_gb_label = dict(
#     #         pd.Series(_mapping_gb_class).reset_index().rename(columns={"index": "value_label", 0: "gb_idx"}) \
#     #             .groupby(by=["gb_idx"])["value_label"].apply(lambda s0: "/".join(s0.sort_values(ascending=True).tolist()))
#     #     )
#     #     _data_label = _crosstab["gb_idx"].apply(lambda s0: _mapping_gb_label.get(s0, ""))
#     # elif _data_type=="Numerical":
#     #     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
#     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
#     _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx")+1, column="label", value=_data_label)
#     _t_features_coarse_woe_report_cate.append(_crosstab)
# _df_t_features_coarse_woe_report_cate = pd.concat(_t_features_coarse_woe_report_cate, ignore_index=True).reset_index(drop=True)
# _df_t_features_coarse_woe_report_cate.insert(loc=0, column="binning_method", value="coarse_binning")
# _df_t_features_coarse_woe_report_cate = _df_t_features_coarse_woe_report_cate[[
#     s0 for s0 in _df_t_features_coarse_woe_report_cate.columns
#     if s0 in ["binning_method", "data_type", "feature_name", "gb_idx", "label", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
# ]].reset_index(drop=True)



### 数值型特征（Numerical）
### 粗分箱处理：根据分箱配置dict_coarse_bin_mapping_numerical进行woe计算

################################################################################################
# 粗分箱处理：根据分箱配置dict_coarse_bin_mapping_numerical进行woe计算

def _func_calc_coarse_iv_numerical(in_df, var_name, target_label, mapping):
    rt = OrderedDict({
        "crosstab": None,
        "IV": None,
        "IV_dropna": None,
        "boundary": mapping,
        "mapping_gb_class": None,
    })
    
    _data_converted = func_binning_continuous_v1(
        in_data=in_df[var_name],
        bins=mapping,
        right_border=True, include_lowest=True,
    )
    rt["crosstab"] = func_woe_report_v1(
        in_var=_data_converted,
        in_target=_df_wb_dev[target_label],
        with_total=True,
        with_lift_ks=True,
        lift_calc_ascending=True,
        with_wls_adj_woe=True,
        woe_with_nan=True,
    )
    
    _crosstab = rt["crosstab"]
    rt["IV"] = _crosstab.loc["total", "IV"]
    rt["IV_dropna"] = _crosstab.loc[[s0 for s0 in _crosstab.index if re.search("(_NaN$)|(NaN, NaN)|(^total$)", s0)==None], "IV"].sum()
    # _d = _crosstab["WOE"].reset_index(drop=True).dropna().reset_index().values
    # _X = _d[:, 0]
    # _y = _d[:, 1]
    return rt

_df_t_features_coarse_iv_num = df_t_describe_summary \
    .query("data_type=='Numerical'") \
    [df_t_describe_summary["column_name"].isin((list(dict_coarse_bin_mapping_numerical.keys()))[:])] \
    [["column_name", "data_type"]].reset_index(drop=True)
_df_t_features_coarse_iv_num["binning_method"] = "coarse_binning"
_df_t_features_coarse_iv_num = _df_t_features_coarse_iv_num[["binning_method", "column_name", "data_type"]]


##########################################################################
# 计算IV
_time = time.time()
_idx = 1

_coarse_binning_result = []
for _column_name in _df_t_features_coarse_iv_num["column_name"].tolist()[:]:
    _mapping = dict_coarse_bin_mapping_numerical.get(_column_name)
    _coarse_binning_result.append(
        _func_calc_coarse_iv_numerical(
            in_df=_df_wb_dev,
            var_name=_column_name,
            target_label="target_label",
            mapping=_mapping,
        )
    )
    print("{:7.2f}%: {}".format(_idx/_df_t_features_coarse_iv_num.shape[0]*100, _column_name))
    _idx = _idx+1

print("\n")
print("cost time: {} sec.".format(time.time()-_time))

##########################################################################
_df_t_features_coarse_iv_num["IV"] = [s0["IV"] for s0 in _coarse_binning_result]
_df_t_features_coarse_iv_num["IV_dropna"] = [s0["IV_dropna"] for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_num["boundary"] = [([-np.inf, np.inf] if s0["boundary"]==None else [round(t, 4) for t in s0["boundary"]]) for s0 in _coarse_binning_result]
_df_t_features_coarse_iv_num["boundary"] = [([round(t, 4) for t in s0["boundary"]] if s0["boundary"]!=None else None) for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_num["boundary"] = [([t for t in s0["boundary"]] if s0["boundary"]!=None else None) for s0 in _coarse_binning_result]
_df_t_features_coarse_iv_num["mapping_gb_class"] = [s0["mapping_gb_class"] for s0 in _coarse_binning_result]
_df_t_features_coarse_iv_num["crosstab"] = [s0["crosstab"] for s0 in _coarse_binning_result]
# _df_t_features_coarse_iv_num["bin_cnt"] = _df_t_features_coarse_iv_num["crosstab"].apply(lambda s0: s0[s0.index!="total"].shape[0])
_df_t_features_coarse_iv_num["output_cutpoint"] = _df_t_features_coarse_iv_num.apply(
    lambda s0: (
        dict(
            [
                (_idx, int(_v))
                for _idx, _v in
                np.concatenate([
                    list(product(_v.split("/"), [_idx]))
                    for _idx, _v in
                    [s0.split("_") for s0 in s0["crosstab"].index if re.search("(^total$)", s0)==None]
                ])
            ]
        )
        if s0["data_type"]=="Categorical" else
        [-np.inf]+pd.Series([
            float(re.sub("[ \]]", "", s0.split("_")[1].split(",")[1]))
            for s0 in s0["crosstab"].index
            if re.search("(NaN, NaN)|(, inf)|(^total$)", s0)==None
        ]).drop_duplicates().tolist()+[np.inf]
        if s0["data_type"]=="Numerical" else
        None
    ),
    axis=1,
)
# _df_t_features_coarse_iv_num["crosstab"] = _df_t_features_coarse_iv_num.apply(
#     lambda s0: (
#         func_woe_report_v1(
#             in_var=func_combining_discrete_v1(
#                 in_data=_df_wb_dev[s0["column_name"]],
#                 mapping_gb_class=s0["output_cutpoint"],
#                 fillna_value="NaN", cvt_fillna_value=s0["output_cutpoint"].get("NaN", 0),
#             ),
#             in_target=_df_wb_dev["target_label"],
#             with_total=True,
#             with_lift_ks=True,
#             lift_calc_ascending=True,
#             with_wls_adj_woe=False,
#             woe_with_nan=True,
#         )
#         if s0["data_type"]=="Categorical" else
#         func_woe_report_v1(
#             in_var=func_binning_continuous_v1(
#                 in_data=_df_wb_dev[s0["column_name"]],
#                 bins=s0["output_cutpoint"],
#                 right_border=True, include_lowest=True,
#             ),
#             in_target=_df_wb_dev["target_label"],
#             with_total=True,
#             with_lift_ks=True,
#             lift_calc_ascending=True,
#             with_wls_adj_woe=True,
#             woe_with_nan=True,
#         )
#         if s0["data_type"]=="Numerical" else
#         None
#     ),
#     axis=1,
# )
_df_t_features_coarse_iv_num["bin_cnt"] = _df_t_features_coarse_iv_num["crosstab"].apply(lambda s0: s0[s0.index!="total"].shape[0])

_df_t_features_coarse_iv_num["_output_object"] = _df_t_features_coarse_iv_num.apply(
    lambda s0: {
        "feature_name": s0["column_name"],
        "data_type": s0["data_type"],
        "binning_method": s0["binning_method"],
        "boundary": s0["boundary"],
        "mapping_gb_class": s0["mapping_gb_class"],
        "crosstab_index": [s0 for s0 in s0["crosstab"].index],
        "crosstab_cnt": s0["crosstab"] \
            [["0_good_#", "1_bad_#"]] \
            .to_dict(orient="list"),
        "crosstab_details": s0["crosstab"] \
            [(
                ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate"]
                if s0["data_type"]=="Categorical" else
                ["0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
                if s0["data_type"]=="Numerical" else
                []
            )] \
            .to_dict(orient="list"),
        "IV": s0["IV"],
        "IV_dropna": s0["IV_dropna"],
        "output_cutpoint": s0["output_cutpoint"],
    },
    axis=1,
)

_df_t_features_coarse_iv_num = _df_t_features_coarse_iv_num.sort_values(by=["IV"], ascending=False).reset_index(drop=True)


###########################################################################
# 计算开发样本的WOE报告
_t_features_coarse_woe_report_num = []
for _column_name, _data_type, _crosstab, _boundary, _mapping_gb_class in \
            _df_t_features_coarse_iv_num[["column_name", "data_type", "crosstab", "boundary", "mapping_gb_class"]].values[:]:
    _crosstab = _crosstab.reset_index().rename(columns={"index": "gb_idx"}).query("gb_idx!='total'")
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx"), column="feature_name", value=_column_name)
    _crosstab.insert(loc=_crosstab.columns.tolist().index("feature_name"), column="data_type", value=_data_type)
    # if _data_type=="Categorical":
    #     _mapping_gb_label = dict(
    #         pd.Series(_mapping_gb_class).reset_index().rename(columns={"index": "value_label", 0: "gb_idx"}) \
    #             .groupby(by=["gb_idx"])["value_label"].apply(lambda s0: "/".join(s0.sort_values(ascending=True).tolist()))
    #     )
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: _mapping_gb_label.get(s0, ""))
    # elif _data_type=="Numerical":
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx")+1, column="label", value=_data_label)
    _t_features_coarse_woe_report_num.append(_crosstab)
_df_t_features_coarse_woe_report_num = pd.concat(_t_features_coarse_woe_report_num, ignore_index=True).reset_index(drop=True)
_df_t_features_coarse_woe_report_num.insert(loc=0, column="binning_method", value="coarse_binning")
_df_t_features_coarse_woe_report_num = _df_t_features_coarse_woe_report_num[[
    s0 for s0 in _df_t_features_coarse_woe_report_num.columns
    if s0 in ["binning_method", "data_type", "feature_name", "gb_idx", "label", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
]].reset_index(drop=True)




### （粗分箱）结果汇总输出

###########################################################################
# （粗分箱）结果汇总输出

df_t_features_coarse_iv = pd.concat(
    objs=[
        # _df_t_features_coarse_iv_cate,
        _df_t_features_coarse_iv_num,
    ],
    axis=0,
).reset_index(drop=True)

df_t_features_coarse_woe_report = pd.concat(
    objs=[
        # _df_t_features_coarse_woe_report_cate,
        _df_t_features_coarse_woe_report_num,
    ],
    axis=0,
).reset_index(drop=True)

###########################################################################
print(df_t_features_coarse_iv.shape)
print(df_t_features_coarse_woe_report.shape)



# ####################################################################################
# df_t_features_coarse_iv.drop(labels=["crosstab", "_output_object"], axis=1).to_clipboard(index=False)
# # df_t_features_coarse_woe_report.to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc05_FeatCoarseBinning_batch/df_t_features_coarse_iv.xlsx".format(result_path), mode="wb") as fw:
    df_t_features_coarse_iv \
        .drop(labels=["crosstab", "_output_object"], axis=1) \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc05_FeatCoarseBinning_batch/df_t_features_coarse_iv.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_features_coarse_iv, file=fw)

####################################################################################
with open(file="{}/Proc05_FeatCoarseBinning_batch/df_t_features_coarse_iv.pkl".format(result_path), mode="rb") as fr:
    df_t_features_coarse_iv = pickle.load(file=fr)



### （分箱结果保存输出json）

###########################################################################
# coarse_cate
with open(file="{}/Proc05_FeatCoarseBinning_batch/obj_res_coarse_binning_{}.json".format(result_path, "cate"), mode="w", encoding="utf-8") as fw:
    json.dump(
        obj=df_t_features_coarse_iv.query("data_type=='{}'".format("Categorical"))["_output_object"].tolist(),
        fp=fw,
        indent=4,
        ensure_ascii=False,
    )

# coarse_num
with open(file="{}/Proc05_FeatCoarseBinning_batch/obj_res_coarse_binning_{}.json".format(result_path, "num"), mode="w", encoding="utf-8") as fw:
    json.dump(
        obj=df_t_features_coarse_iv.query("data_type=='{}'".format("Numerical"))["_output_object"].tolist(),
        fp=fw,
        indent=4,
        ensure_ascii=False,
    )



### （分箱结果导入json）

###########################################################################
_obj_cp = []

# coarse_cate
with open(file="{}/Proc05_FeatCoarseBinning_batch/obj_res_coarse_binning_{}.json".format(result_path, "cate"), mode="r", encoding="utf-8") as fr:
    _obj_cp = _obj_cp+json.load(fp=fr)

# coarse_num
with open(file="{}/Proc05_FeatCoarseBinning_batch/obj_res_coarse_binning_{}.json".format(result_path, "num"), mode="r", encoding="utf-8") as fr:
    _obj_cp = _obj_cp+json.load(fp=fr)

###########################################################################
# IV
df_t_features_coarse_iv = pd.DataFrame([
    {
        "binning_method": s0["binning_method"],
        "column_name": s0["feature_name"],
        "data_type": s0["data_type"],
        "IV": s0["IV"],
        "IV_dropna": s0["IV_dropna"],
        "boundary": s0["boundary"],
        "mapping_gb_class": s0["mapping_gb_class"],
        "crosstab": pd.DataFrame(index=s0["crosstab_index"], data=dict(list(s0["crosstab_cnt"].items())+list(s0["crosstab_details"].items()))),
        "bin_cnt": len([s0 for s0 in s0["crosstab_index"] if s0!="total"]),
        "output_cutpoint": s0["output_cutpoint"],
        "_output_object": s0,
    }
    for s0 in _obj_cp
])

###########################################################################
# WOE report
_t_features_coar_woe_report = []
for _binning_method, _column_name, _data_type, _crosstab, _boundary, _mapping_gb_class in \
            df_t_features_coarse_iv[["binning_method", "column_name", "data_type", "crosstab", "boundary", "mapping_gb_class"]].values[:]:
    _crosstab = _crosstab.reset_index().rename(columns={"index": "gb_idx"}).query("gb_idx!='total'")
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx"), column="feature_name", value=_column_name)
    _crosstab.insert(loc=_crosstab.columns.tolist().index("feature_name"), column="data_type", value=_data_type)
    # if _data_type=="Categorical":
    #     _mapping_gb_label = dict(
    #         pd.Series(_mapping_gb_class).reset_index().rename(columns={"index": "value_label", 0: "gb_idx"}) \
    #             .groupby(by=["gb_idx"])["value_label"].apply(lambda s0: "/".join(s0.sort_values(ascending=True).tolist()))
    #     )
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: _mapping_gb_label.get(s0, ""))
    # elif _data_type=="Numerical":
    #     _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _data_label = _crosstab["gb_idx"].apply(lambda s0: (s0.split("_")[-1] if s0!="total" else ""))
    _crosstab.insert(loc=_crosstab.columns.tolist().index("gb_idx")+1, column="label", value=_data_label)
    _crosstab.insert(loc=0, column="binning_method", value=_binning_method)
    _t_features_coar_woe_report.append(_crosstab)
df_t_features_coarse_woe_report = pd.concat(_t_features_coar_woe_report, ignore_index=True).reset_index(drop=True)
df_t_features_coarse_woe_report = df_t_features_coarse_woe_report[[
    s0 for s0 in df_t_features_coarse_woe_report.columns
    if s0 in ["binning_method", "data_type", "feature_name", "gb_idx", "label", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%", "WOE", "IV", "total", "total_pct", "bad_rate", "ks", "wls_adj_WOE", "_wls_adj_woe_details"]
]].reset_index(drop=True)

###########################################################################
print(df_t_features_coarse_iv.shape)
print(df_t_features_coarse_woe_report.shape)



### woe编码转换

###########################################################################
# woe编码
def _func_woe_mapping(in_df, df_mapping_info, cols_keep=None, print_log=False, with_wls_adj_woe=False):
    
    ################################################
    df_mapping_info = df_mapping_info \
        .sort_values(by=["column_name", "IV_dropna"], ascending=[True, False]) \
        [["binning_method", "column_name", "data_type", "crosstab", "output_cutpoint"]] \
        .drop_duplicates(subset=["column_name"], keep="first") \
        .sort_index(ascending=True) \
        .reset_index(drop=True)
    if cols_keep is None:
        cols_keep = [s0 for s0 in in_df.columns if s0 not in df_mapping_info["column_name"].tolist()]
    
    ################################################
    df_woe = in_df[cols_keep].reset_index(drop=True)
    _idx = 1
    for _binning_method, _column_name, _data_type, _crosstab, _output_cutpoint in df_mapping_info.values[:]:
        
 
        # 20230919添加
        # 把下述的这些变量空值用woe=0进行填充
        if _column_name in [
                            '三方数据源_hengpu_4', 
                            '三方数据源_ruizhi_6', 
                            '三方数据源_bairong_14', 
                            '三方数据源_hangliezhi_1',
                            # 'INQUIRY0047',
                            'digitalScore', 
                            'CARD0065',
                           # 'query_tfau_mg', 
                           'pboc_education_new', 
                           'income_izai_tmf', 
                           'quad_p1',
                           'loan_overdue_last_months_l12m', 
                           'debts_sgad_mc', 
                           'repay_dcal_tmd',
                           'repay_dhaa_md'
                            ]:
            _index = [re.search("NaN", s0.split("_")[-1])!=None for s0 in _crosstab.index]
            _crosstab.loc[_index, "WOE"] = 0
        # 20230919添加       
        
        
        
        if print_log:
            print("{} {:7.2f}%: {}".format(_binning_method, _idx/df_mapping_info.shape[0]*100, _column_name))
            _idx = _idx+1
        
        ################################################
        if _data_type=="Numerical":
            _mapping = dict([
                (re.sub("[, \]]", "", _k.split("_")[1].split(",")[1]), _v)
                # (_k, _v)
                for _k, _v in (_crosstab["wls_adj_WOE"].to_dict().items() if with_wls_adj_woe else _crosstab["WOE"].to_dict().items())
                # for _k, _v in _crosstab["WOE"].to_dict().items()
                # for _k, _v in _crosstab["wls_adj_WOE"].to_dict().items()
                # if re.search("(NaN, NaN)|(^total$)", _k)==None
                if re.search("(^total$)", _k)==None
            ])
            df_woe["BIN_{}".format(_column_name)] = func_binning_continuous_v1(
                in_data=in_df[_column_name],
                bins=_output_cutpoint,
                out_type="01_info", right_border=True, include_lowest=False,
            )
            df_woe["WOE_{}".format(_column_name)] = df_woe["BIN_{}".format(_column_name)] \
                .apply(lambda s0: _mapping.get(re.sub("[, \]]", "", s0.split("_")[1].split(",")[1])))
        elif _data_type=="Categorical":
            _mapping = dict([
                (_k.split("_")[1], _v)
                # (_k, _v)
                for _k, _v in _crosstab["WOE"].to_dict().items()
                # for _k, _v in _crosstab["wls_adj_WOE"].to_dict().items()
                # if re.search("(NaN, NaN)|(^total$)", _k)==None
                if re.search("(^total$)", _k)==None
            ])
            df_woe["BIN_{}".format(_column_name)] = func_combining_discrete_v1(
                in_data=in_df[_column_name],
                mapping_gb_class=_output_cutpoint,
                fillna_value="NaN", cvt_fillna_value=_output_cutpoint.get("NaN", 0),
            )
            df_woe["WOE_{}".format(_column_name)] = df_woe["BIN_{}".format(_column_name)] \
                .apply(lambda s0: _mapping.get(s0.split("_")[1]))
        else:
            pass
    
    return df_woe

###########################################################################
df_wb_coarse_woe = \
    _func_woe_mapping(
        in_df=df_wb,
        df_mapping_info=df_t_features_coarse_iv,
        cols_keep=cols_base,
        print_log=True,
        with_wls_adj_woe=False,
    )


print(df_wb_coarse_woe.shape)
# df_wb_coarse_woe.sample(2)

df_wb_coarse_woe['data_role'] = df_wb['data_role']
# cols_base = cols_base+['data_role']

# cols_base = cols_base[:-1]




# 计算PSI
df_psi = pd.DataFrame()
for i in ["三方数据源_hengpu_4",
        "三方数据源_hangliezhi_1",
        "三方数据源_ruizhi_6",
        "三方数据源_bairong_14",
        "INQUIRY0047",
        "digitalScore",
        "CARD0065",
        "query_tfau_mg",
        "debts_sgad_mc",
        "repay_dcal_tmd",
        "repay_dhaa_md",    
        "pboc_education_new",
        "income_izai_tmf",
        "quad_p1",
        ]:
    _t = 'BIN_{}'.format(i)
    _t_df = df_wb_coarse_woe.groupby\
        ([_t,'observation_dt_YM'])\
            ['order_no'].count()
    # _t_df.loc[:,'columns'] = i    
    _t_df_stack = _t_df.unstack()
    _t_df_stack.loc[:,'columns'] = i    
    # _t_df_stack = _t_df_stack.append(pd.DataFrame(_t_df_stack.sum()).rename(columns = {0: 'total'}).T)
    
    df_psi = df_psi.append(_t_df_stack)
df_psi = df_psi.append(pd.DataFrame(_t_df_stack.sum()).rename(columns = {0: 'total'}).T)



df_psi_1 = df_psi.iloc[:,:-1]/df_psi.iloc[-1,:-1]
df_psi_1['columns'] = df_psi['columns']
df_psi_1 = df_psi_1[~df_psi_1.index.isin(['0_(NaN, NaN]'])]
df_psi_1 = df_psi_1.iloc[:-1, :]
df_psi_1

for i in list(df_psi_1.columns[1:-1]):
    new_col = i+'_psi'
    df_psi_1[new_col] = df_psi_1.apply(lambda x: (x[i]-x['2023-05'])*math.log(x[i]/x['2023-05']), axis = 1)
    print(df_psi_1.groupby('columns')[new_col].sum())


df_psi_1.to_clipboard()


#==============================================================================
# File: code08_Proc06_Correlation.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc06_Correlation
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc06_Correlation
## 特征相关性分析


####################################################################
# Proc06_Correlation
# 特征相关性分析


####################################################################
# 对粗分箱结果进行排序（IV升序），相关性剔除时，确保优先保留IV值更高的特征
_cols_coarse_binning_iv_sorted = df_t_features_coarse_iv \
    [["column_name", "IV"]] \
    .sort_values(by=["IV"], ascending=[True]) \
    ["column_name"].tolist()

_df_features = df_wb_coarse_woe \
    .query("data_role in ['01_train', '02_test']") \
    [["WOE_{}".format(s0) for s0 in _cols_coarse_binning_iv_sorted]] \
    .reset_index(drop=True)


####################################################################
# 过滤变量之间相关系数大于阈值
_corr_threshold = 0.99
cols_non_collin_0, df_corr_table = func_colinearity_rt_col(
    in_df_features=_df_features,
    corr_threshold=_corr_threshold,
    corr_method="pearson",
    # corr_method="kendall",
    # corr_method="spearman"
)

cols_non_collin_0 = [re.findall("^WOE_(.+)$", s0)[0] for s0 in cols_non_collin_0 if "WOE_" in s0]

####################################################################
print("remove cols num: {}".format(len([s0 for s0 in _cols_coarse_binning_iv_sorted if s0 not in cols_non_collin_0])))
print("keep cols num: {}".format(len(cols_non_collin_0)))

# df_corr_table.to_clipboard(index=True, header=True)

####################################################################
_df_features = _df_features \
    [["WOE_{}".format(s0) for s0 in cols_non_collin_0]] \
    .reset_index(drop=True)


####################################################################
# （相关性剔除前）
# 计算相关性矩阵，WOE编码后的pearson相关系数

# ####################################################################
# _, df_corr_table = func_colinearity_rt_col(
#     in_df_features=_df_features,
#     corr_threshold=1,
#     corr_method="pearson",
#     # corr_method="kendall",
#     # corr_method="spearman"
# )

####################################################################
func_correlation_matrix_hotmap(
    df=_df_features,
    figsize=6,
    plot_output_fn=None,
)

####################################################################
df_vif_table = func_multicolinearity_vif(
    in_df_features=_df_features,
    with_constant=True,
)
df_vif_table.insert(
    loc=0, column="column_name",
    value=df_vif_table["feature_name"].apply(lambda s0: s0.replace("WOE_", ""))
)
df_vif_table = df_vif_table \
    .sort_values(by=["VIF"], ascending=[False]) \
    .reset_index(drop=True)
df_vif_table

####################################################################
# # df_corr_table.to_clipboard(index=True, header=True)
# df_vif_table.to_clipboard(index=False)


####################################################################
# 过滤变量之间相关系数大于阈值
# _corr_threshold = 0.6
# _corr_threshold = 0.7
_corr_threshold = 0.8
# _corr_threshold = 0.9
cols_non_collin, df_corr_table = func_colinearity_rt_col(
    in_df_features=_df_features,
    corr_threshold=_corr_threshold,
    corr_method="pearson",
    # corr_method="kendall",
    # corr_method="spearman"
)

cols_non_collin = [re.findall("^WOE_(.+)$", s0)[0] for s0 in cols_non_collin if "WOE_" in s0]

####################################################################
print("remove cols num: {}".format(len([s0 for s0 in _cols_coarse_binning_iv_sorted if s0 not in cols_non_collin])))
print("keep cols num: {}".format(len(cols_non_collin)))

# df_corr_table.to_clipboard(index=True, header=True)


####################################################################
# （相关性剔除后）
# 计算相关性矩阵，WOE编码后的pearson相关系数

####################################################################
_, df_corr_table_non_collin = func_colinearity_rt_col(
    in_df_features=_df_features[["WOE_{}".format(s0) for s0 in cols_non_collin]],
    corr_threshold=1,
    corr_method="pearson",
    # corr_method="kendall",
    # corr_method="spearman"
)

####################################################################
func_correlation_matrix_hotmap(
    df=_df_features[["WOE_{}".format(s0) for s0 in cols_non_collin]],
    figsize=6,
    plot_output_fn=None,
)

####################################################################
df_vif_table_non_collin = func_multicolinearity_vif(
    in_df_features=_df_features[["WOE_{}".format(s0) for s0 in cols_non_collin]],
    with_constant=True,
)
df_vif_table_non_collin.insert(
    loc=0, column="column_name",
    value=df_vif_table_non_collin["feature_name"].apply(lambda s0: s0.replace("WOE_", ""))
)
df_vif_table_non_collin = df_vif_table_non_collin \
    .sort_values(by=["VIF"], ascending=[False]) \
    .reset_index(drop=True)

####################################################################
print(len(cols_non_collin))
print(df_corr_table_non_collin.shape)
# df_corr_table_non_collin.to_clipboard(index=True, header=True)
# df_vif_table_non_collin.to_clipboard(index=False)

df_vif_table_non_collin




### （结果导出）

####################################################################################
df_corr_table.to_clipboard(index=True, header=True)
df_vif_table.to_clipboard(index=False)

####################################################################################
df_corr_table_non_collin.to_clipboard(index=True, header=True)
df_vif_table_non_collin.to_clipboard(index=False)

####################################################################################
pd.Series(len(_cols_coarse_binning_iv_sorted), name="column_name").to_clipboard(index=False)
pd.Series(cols_non_collin_0, name="column_name").to_clipboard(index=False)
pd.Series(cols_non_collin, name="column_name").to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc06_Correlation/df_corr_table.xlsx".format(result_path), mode="wb") as fw:
    df_corr_table \
        .to_excel(
            fw,
            index=True,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/df_vif_table.xlsx".format(result_path), mode="wb") as fw:
    df_vif_table \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/df_corr_table_non_collin.xlsx".format(result_path), mode="wb") as fw:
    df_corr_table_non_collin \
        .to_excel(
            fw,
            index=True,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/df_vif_table_non_collin.xlsx".format(result_path), mode="wb") as fw:
    df_vif_table_non_collin \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/_cols_coarse_binning_iv_sorted.xlsx".format(result_path), mode="wb") as fw:
    pd.Series(_cols_coarse_binning_iv_sorted, name="column_name") \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/cols_non_collin_0.xlsx".format(result_path), mode="wb") as fw:
    pd.Series(cols_non_collin_0, name="column_name") \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )
with open(file="{}/Proc06_Correlation/cols_non_collin.xlsx".format(result_path), mode="wb") as fw:
    pd.Series(cols_non_collin, name="column_name") \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

####################################################################################
func_correlation_matrix_hotmap(
    df=_df_features,
    figsize=12,
    plot_output_fn="{}/Proc06_Correlation/corr_hotmap.png".format(result_path),
)
func_correlation_matrix_hotmap(
    df=_df_features[["WOE_{}".format(s0) for s0 in cols_non_collin]],
    figsize=12,
    plot_output_fn="{}/Proc06_Correlation/corr_hotmap_cols_non_collin.png".format(result_path),
)

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc06_Correlation/df_corr_table.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_corr_table, file=fw)
with open(file="{}/Proc06_Correlation/df_vif_table.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_vif_table, file=fw)
with open(file="{}/Proc06_Correlation/df_corr_table_non_collin.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_corr_table_non_collin, file=fw)
with open(file="{}/Proc06_Correlation/df_vif_table_non_collin.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_vif_table_non_collin, file=fw)
with open(file="{}/Proc06_Correlation/_cols_coarse_binning_iv_sorted.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=_cols_coarse_binning_iv_sorted, file=fw)
with open(file="{}/Proc06_Correlation/cols_non_collin_0.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=cols_non_collin_0, file=fw)
with open(file="{}/Proc06_Correlation/cols_non_collin.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=cols_non_collin, file=fw)


####################################################################################
with open(file="{}/Proc06_Correlation/df_corr_table.pkl".format(result_path), mode="rb") as fr:
    df_corr_table = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/df_vif_table.pkl".format(result_path), mode="rb") as fr:
    df_vif_table = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/df_corr_table_non_collin.pkl".format(result_path), mode="rb") as fr:
    df_corr_table_non_collin = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/df_vif_table_non_collin.pkl".format(result_path), mode="rb") as fr:
    df_vif_table_non_collin = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/_cols_coarse_binning_iv_sorted.pkl".format(result_path), mode="rb") as fr:
    _cols_coarse_binning_iv_sorted = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/cols_non_collin_0.pkl".format(result_path), mode="rb") as fr:
    cols_non_collin_0 = pickle.load(file=fr)
with open(file="{}/Proc06_Correlation/cols_non_collin.pkl".format(result_path), mode="rb") as fr:
    cols_non_collin = pickle.load(file=fr)























#==============================================================================
# File: code09_Proc07_Model_LR_stepwise.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc07_Model_LR_stepwise
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc07_Model_LR_stepwise
## 模型训练（逐步回归筛选）


####################################################################
# Proc07_Model_LR_stepwise
# 模型训练（逐步回归筛选）


### 模型训练样本准备

####################################################################
# 模型训练样本准备

####################################################################
# 设置随机种子 oversample_random_seed
# 设置过采样处理的坏样本权重 oversample_bad_weight

# oversample_random_seed = random.randint(1, 100000)
# oversample_bad_weight = random.randint(25, 30)/100
oversample_random_seed = 12345678
oversample_bad_weight = 0.5
# oversample_bad_weight = 0.3

####################################################################
# 参与模型训练的指标清单 _cols_model_training

# _cols_model_training = [
#     s0 for s0 in df_wb.columns
#     if s0 in [
#         t for t in
#         list(dict_coarse_bin_mapping_categorical.keys())+list(dict_coarse_bin_mapping_numerical.keys())
#     ]
# ]
_cols_model_training = [
    re.sub("^WOE_", "", s0) for s0 in df_wb_coarse_woe.columns
    if re.sub("^WOE_", "", s0) in [
        t for t in
        # list(dict_coarse_bin_mapping_categorical.keys())+list(dict_coarse_bin_mapping_numerical.keys())
        list(dict_coarse_bin_mapping_numerical.keys())
    ]
]


####################################################################
# 模型训练样本
_df_wb_coarse_woe = pd.concat([
    df_wb_coarse_woe.query("flag==1"),
    df_wb_coarse_woe.query("flag==0"),
    # df_wb_coarse_woe.query("flag==0").sample(n=2000, random_state=oversample_random_seed),
], ignore_index=True).reset_index(drop=True)

# ####################################################################
# # 不设置过采样处理
# _df_model_training_sampling = _df_wb_coarse_woe \
#     [cols_base+[s0 for s0 in _df_wb_coarse_woe.columns if re.search("^WOE_", s0)!=None]] \
#     .query("data_role in ['01_train']") \
#     .reset_index(drop=True)

####################################################################
# 设置过采样处理
_df_model_training_sampling = func_oversample_stratify(
    in_df=_df_wb_coarse_woe[
            cols_base+["WOE_{}".format(s0) for s0 in _cols_model_training]
        ] \
        .query("data_role in ['01_train']"),
    # n=10000,
    # n=20000,
    # n=50000,
    n=100000,
    stratify_key="flag",
    group_weight={0: 100-oversample_bad_weight*100, 1: 0+oversample_bad_weight*100},
    random_seed=oversample_random_seed,
) \
#     .sample(10000, random_state=oversample_random_seed).reset_index(drop=True)

print(_df_model_training_sampling.shape[0])
print(_df_model_training_sampling["flag"].value_counts(dropna=False, normalize=False).to_dict())
print(_df_model_training_sampling["flag"].value_counts(dropna=False, normalize=True).to_dict())
print()




### 逐步回归

####################################################################
# 逐步回归
_time = time.time()

####################################################################
# v1
# model_cols_init = [
    
# ]

model_cols_init = cols_non_collin

model_cols_init = [
"三方数据源_bairong_14",
"三方数据源_bairong_15",
"三方数据源_hangliezhi_1",
"三方数据源_hengpu_4",
"三方数据源_ruizhi_6",
"三方数据源_xinyongsuanli_1",
"education",
"loan_overdue_last_months_l9m",
"query_tfau_mg",
"coaa_zbva_xavh_bbvf_n9",
"debts_sgac_mc",
"pboc_education_new",
"repay_dcal_tmd",
"loan_overdue_last_months_l12m",
"INQUIRY0047",
"cc_overdue_last_months_l12m",
"debts_sgad_mc",
"repay_dhaa_md",
"overdue_m1_cnt_l6m_all",
"coaa_zbvg_xawd_bbvd_n3",
"CARD0065",
"quad_p1",
"income_izai_tmf",
"digitalScore",
"pboc_sex_new",

]

####################################################################
_df_train_X = _df_model_training_sampling \
    [
        ["WOE_{}".format(s0) for s0 in model_cols_init] \
        # ["WOE_{}".format(s0) for s0 in list(dict_coarse_bin_mapping_numerical.keys())] \
        # ["WOE_{}".format(s0) for s0 in cols_non_collin]
    ].reset_index(drop=True)
_df_train_y = _df_model_training_sampling[["flag"]].reset_index(drop=True)


####################################################################
# stepwise计算
logistic_model_res, model_cols, df_t_stepwise_selection = \
    func_logistic_model_stepwise(
        df_X=_df_train_X,
        y=_df_train_y.iloc[:, 0],
        method="ncg",
        initial_cols=None,
        # initial_cols=["COAR_WOE_{}".format(s0) for s0 in model_cols_init],
        include_cols=None,
        exclude_cols=None,
        sle=0.000001, sls=0.0000001, with_intercept=True, verbose=1,
        # sle=0.05, sls=0.01, with_intercept=True, verbose=1,
        # sle=0.1, sls=0.1, with_intercept=True, verbose=1,
        max_step=200,
        # max_step=20,
        # check_coef_positive=False,
        check_coef_positive=True,
)

# logistic_model_res = func_logistic_model(
#     df_X=_df_train_X,
#     y=_df_train_y.iloc[:, 0],
#     method="ncg",
#     with_intercept=True,
#     disp=1,
# )
# model_cols = _df_train_X.columns.tolist()

t = logistic_model_res.summary2().tables
df_model_summary = t[0]
df_model_params = t[1]
df_model_params = df_model_params.reset_index().rename(columns={"index": "feature_name"})
df_model_params.insert(
    loc=0, column="column_name",
    value=[re.sub("^WOE_", "", s0) for s0 in df_model_params["feature_name"]],
)

################################
logistic_model_res_stepwise_0 = copy.deepcopy(logistic_model_res)
df_model_summary_stepwise_0 = df_model_summary.copy(deep=True)
df_model_params_stepwise_0 = df_model_params.copy(deep=True)

print(time.time()-_time)
print(len(model_cols))


####################################################################
logistic_model_res_stepwise_0.summary2()
# df_model_summary_stepwise_0
# df_model_params_stepwise_0
# df_t_stepwise_selection

# df_model_summary_stepwise_0.to_clipboard(index=False, header=False)
# df_model_params_stepwise_0.to_clipboard(index=True)
# df_t_stepwise_selection.to_clipboard(index=False)



####################################################################
# 剔除模型系数与pvalue不符合要求的指标
df_t_aft_stepwise_rm_step = []
step = 0
df_model_params = df_model_params_stepwise_0.copy(deep=True)
model_cols = df_model_params["feature_name"].tolist()
_l = len(model_cols)
while 1:
    
    step = step+1
    
    _mc = model_cols
    model_cols = df_model_params[(
        (
            (df_model_params["feature_name"]!="Intercept") &
            (df_model_params["Coef."]>0) &
            (df_model_params["P>|z|"].notna()) &
            (df_model_params["P>|z|"]<=0.01)
    #         (df_model_params["P>|z|"]<=0.1)
        )
    #     | (df_model_params.index.isin(["COAR_WOE_{}".format(s0) for s0 in model_cols_0]))
    )]["feature_name"].tolist()
    
    if _l==len(model_cols):
        break
    else:
        _l = len(model_cols)
    
    df_t_aft_stepwise_rm_step.append(OrderedDict({
        "step": step,
        "remian_cols": [s0.replace("WOE_", "") for s0 in model_cols if s0!="Intercept"],
        "remian_cols_cnt": len([s0.replace("WOE_", "") for s0 in model_cols if s0!="Intercept"]),
        "remove_cols": [s0.replace("WOE_", "") for s0 in _mc if s0 not in model_cols and s0!="Intercept"],
        "remove_cols_cnt": len([s0.replace("WOE_", "") for s0 in _mc if s0 not in model_cols and s0!="Intercept"]),
    }))
    
    logistic_model_res = func_logistic_model(
        df_X=_df_train_X[model_cols],
        y=_df_train_y.iloc[:, 0],
        method="ncg",
        with_intercept=True,
        disp=1,
    )
    
    t = logistic_model_res.summary2().tables
    df_model_summary = t[0]
    df_model_params = t[1]
    df_model_params = df_model_params.reset_index().rename(columns={"index": "feature_name"})
    df_model_params.insert(
        loc=0, column="column_name",
        value=[re.sub("^WOE_", "", s0) for s0 in df_model_params["feature_name"]],
    )
    
    print(len(model_cols))
    # print(logistic_model_res.summary2())

################################
logistic_model_res_stepwise = copy.deepcopy(logistic_model_res)
df_model_summary_stepwise = df_model_summary.copy(deep=True)
df_model_params_stepwise = df_model_params.copy(deep=True)

df_t_aft_stepwise_rm_step = pd.DataFrame(df_t_aft_stepwise_rm_step)


####################################################################
logistic_model_res_stepwise.summary2()
# df_model_summary_stepwise
# df_model_params_stepwise

# df_model_summary_stepwise.to_clipboard(index=False, header=False)
# df_model_params_stepwise.to_clipboard(index=True)



### （训练样本导出）

# ###################################################################################
# # pkl文件

# ####################################################################################
# with open(file="{}/Proc07_Model_LR_adj/_df_model_training_sampling.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=_df_model_training_sampling, file=fw)

# # ####################################################################################
# # with open(file="{}/Proc07_Model_LR_adj/_df_model_training_sampling.pkl".format(result_path), mode="rb") as fr:
# #     _df_model_training_sampling = pickle.load(file=fr)



### （模型结果导出）

# ####################################################################################
# df_model_summary_stepwise_0.to_clipboard(index=False, header=False)
# df_model_params_stepwise_0.to_clipboard(index=True)
# df_t_stepwise_selection.to_clipboard(index=False)

# ####################################################################################
# df_model_summary_stepwise.to_clipboard(index=False, header=False)
# df_model_params_stepwise.to_clipboard(index=True)

# ####################################################################################
# # to_excel
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise_0.xlsx".format(result_path), mode="wb") as fw:
#     df_model_summary_stepwise_0 \
#         .to_excel(
#             fw,
#             index=False,
#             header=False,
#             sheet_name="data",
#         )
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise_0.xlsx".format(result_path), mode="wb") as fw:
#     df_model_params_stepwise_0 \
#         .to_excel(
#             fw,
#             index=True,
#             sheet_name="data",
#         )
# with open(file="{}/Proc07_Model_LR_stepwise/df_t_stepwise_selection.xlsx".format(result_path), mode="wb") as fw:
#     df_t_stepwise_selection \
#         .to_excel(
#             fw,
#             index=False,
#             sheet_name="data",
#         )
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise.xlsx".format(result_path), mode="wb") as fw:
#     df_model_summary_stepwise \
#         .to_excel(
#             fw,
#             index=False,
#             header=False,
#             sheet_name="data",
#         )
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise.xlsx".format(result_path), mode="wb") as fw:
#     df_model_params_stepwise \
#         .to_excel(
#             fw,
#             index=True,
#             sheet_name="data",
#         )

# ###################################################################################
# # pkl文件

# ####################################################################################
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise_0.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_model_summary_stepwise_0, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise_0.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_model_params_stepwise_0, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/df_t_stepwise_selection.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_t_stepwise_selection, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_model_summary_stepwise, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=df_model_params_stepwise, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/logistic_model_res_stepwise_0.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=logistic_model_res_stepwise_0, file=fw)
# with open(file="{}/Proc07_Model_LR_stepwise/logistic_model_res_stepwise.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=logistic_model_res_stepwise, file=fw)


# ####################################################################################
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise_0.pkl".format(result_path), mode="rb") as fr:
#     df_model_summary_stepwise_0 = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise_0.pkl".format(result_path), mode="rb") as fr:
#     df_model_params_stepwise_0 = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/df_t_stepwise_selection.pkl".format(result_path), mode="rb") as fr:
#     df_t_stepwise_selection = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_summary_stepwise.pkl".format(result_path), mode="rb") as fr:
#     df_model_summary_stepwise = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/df_model_params_stepwise.pkl".format(result_path), mode="rb") as fr:
#     df_model_params_stepwise = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/logistic_model_res_stepwise_0.pkl".format(result_path), mode="rb") as fr:
#     logistic_model_res_stepwise_0 = pickle.load(file=fr)
# with open(file="{}/Proc07_Model_LR_stepwise/logistic_model_res_stepwise.pkl".format(result_path), mode="rb") as fr:
#     logistic_model_res_stepwise = pickle.load(file=fr)




























#==============================================================================
# File: code10_Proc07_Model_LR_adj.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc07_Model_LR_adj
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc07_Model_LR_adj
## 模型训练（手动调整）


####################################################################
# Proc07_Model_LR_adj
# 模型训练（手动调整）


### 模型训练样本准备

####################################################################
# 模型训练样本准备

####################################################################
# 设置随机种子 oversample_random_seed
# 设置过采样处理的坏样本权重 oversample_bad_weight

# oversample_random_seed = random.randint(1, 100000)
# oversample_bad_weight = random.randint(25, 30)/100
# oversample_random_seed = 12345678
# oversample_random_seed = 123

oversample_random_seed = 24

# oversample_bad_weight = 0.5
oversample_bad_weight = 0.06
# oversample_bad_weight = 0.3

####################################################################
# 参与模型训练的指标清单 _cols_model_training

# _cols_model_training = [
#     s0 for s0 in df_wb.columns
#     if s0 in [
#         t for t in
#         list(dict_coarse_bin_mapping_categorical.keys())+list(dict_coarse_bin_mapping_numerical.keys())
#     ]
# ]
_cols_model_training = [
    re.sub("^WOE_", "", s0) for s0 in df_wb_coarse_woe.columns
    if re.sub("^WOE_", "", s0) in [
        t for t in
        # list(dict_coarse_bin_mapping_categorical.keys())+
        list(dict_coarse_bin_mapping_numerical.keys())
    ]
]

####################################################################
# 模型训练样本
_df_wb_coarse_woe = pd.concat([
    df_wb_coarse_woe.query("flag==1"),
    df_wb_coarse_woe.query("flag==0"),
    # df_wb_coarse_woe.query("flag==0").sample(n=2000, random_state=oversample_random_seed),
], ignore_index=True).reset_index(drop=True)

# ####################################################################
# # 不设置过采样处理
# _df_model_training_sampling = _df_wb_coarse_woe \
#     [cols_base+[s0 for s0 in _df_wb_coarse_woe.columns if re.search("^WOE_", s0)!=None]] \
#     .query("data_role in ['01_train']") \
#     .reset_index(drop=True)


# _df_model_training_sampling = pd.concat([_df_model_training_sampling, \
#                                           _df_model_training_sampling], \
#                                         axis = 0)
# _df_model_training_sampling.shape

####################################################################
# 设置过采样处理
_df_model_training_sampling = func_oversample_stratify(
    in_df=_df_wb_coarse_woe[
            cols_base+["WOE_{}".format(s0) for s0 in _cols_model_training]
        ] \
        .query("data_role in ['01_train']"),
    # n=10000,
    # n=20000,
    # n=50000,
        
    n=100000,

    stratify_key="flag",
    group_weight={0: 100-oversample_bad_weight*100, 1: 0+oversample_bad_weight*100},
    random_seed=oversample_random_seed,
) \
#     .sample(10000, random_state=oversample_random_seed).reset_index(drop=True)

print(_df_model_training_sampling.shape[0])
print(_df_model_training_sampling["flag"].value_counts(dropna=False, normalize=False).to_dict())
print(_df_model_training_sampling["flag"].value_counts(dropna=False, normalize=True).to_dict())
print()




# for s0 in df_model_params_stepwise["column_name"].values[:]:
#     print('"{}",'.format(s0))

# # for s0 in df_model_params_stepwise["Coef."].values[:]:
# #     print('{:.6f},'.format(s0))




### 手动调整

# print("-"*40)
# for s0 in df_model_params_stepwise["column_name"].values[:]:
#     print('"{}",'.format(s0))


####################################################################
# 手动调整
_time = time.time()


####################################################################
# v3
model_cols_init = [

# V1
# "三方数据源_hengpu_4",
# "INQUIRY0047",
# "三方数据源_hangliezhi_1",
# "digitalScore",
# "CARD0065",
# "三方数据源_ruizhi_6",
# "query_tfau_mg",
# "repay_dcal_tmd",
# "pboc_sex_new",
# "pboc_education_new",
# "income_izai_tmf",
# "quad_p1",
# "debts_sgad_mc",
# "三方数据源_bairong_14",
# "repay_dhaa_md",
# "loan_overdue_last_months_l12m",
# "三方数据源_xinyongsuanli_1",
# "三方数据源_bairong_15",


# V2
# "三方数据源_hengpu_4",
# "INQUIRY0047",
# "三方数据源_hangliezhi_1",
# "digitalScore",
# "CARD0065",
# "三方数据源_ruizhi_6",
# "query_tfau_mg",
# "repay_dcal_tmd",
# "pboc_sex_new",
# "pboc_education_new",
# "income_izai_tmf",
# "quad_p1",
# "debts_sgad_mc",
# # "三方数据源_bairong_14",
# "repay_dhaa_md",
# "loan_overdue_last_months_l12m",
# "三方数据源_xinyongsuanli_1",
# # "三方数据源_bairong_15",

    
# V3
"三方数据源_hengpu_4",
"INQUIRY0047",
"三方数据源_hangliezhi_1",
"digitalScore",
"CARD0065",
"三方数据源_ruizhi_6",
"query_tfau_mg",


"debts_sgad_mc",
"repay_dcal_tmd",

"repay_dhaa_md",    
"pboc_education_new",
"income_izai_tmf",

"quad_p1",
"三方数据源_bairong_14",

# "loan_overdue_last_months_l12m",
# "三方数据源_xinyongsuanli_1",
 
    
]

####################################################################
_df_train_X = _df_model_training_sampling \
    [
        ["WOE_{}".format(s0) for s0 in model_cols_init]
    ].reset_index(drop=True)
_df_train_y = _df_model_training_sampling[["flag"]].reset_index(drop=True)


####################################################################
# 手动调整
logistic_model_res = func_logistic_model(
    df_X=_df_train_X,
    y=_df_train_y.iloc[:, 0],
    method="ncg",
    with_intercept=True,
    # with_intercept=False,
    disp=1,
)
model_cols = _df_train_X.columns.tolist()

t = logistic_model_res.summary2().tables
df_model_summary = t[0]
df_model_params = t[1]
df_model_params = df_model_params.reset_index().rename(columns={"index": "feature_name"})
df_model_params.insert(
    loc=0, column="column_name",
    value=[re.sub("^WOE_", "", s0) for s0 in df_model_params["feature_name"]],
)

print(time.time()-_time)
print(len(model_cols))


####################################################################
print(logistic_model_res.summary2())
# df_model_summary
# df_model_params

# df_model_summary.to_clipboard(index=False, header=False)
# df_model_params.to_clipboard(index=True)




####################################################################
# 计算相关性矩阵，WOE编码后的pearson相关系数
_df_features_lr_adj = df_wb_coarse_woe \
    .query("data_role in ['01_train', '02_test']") \
    [model_cols] \
    .reset_index(drop=True)

####################################################################
_, df_corr_table_lr_adj = func_colinearity_rt_col(
    in_df_features=_df_features_lr_adj,
    corr_threshold=1,
    corr_method="pearson",
    # corr_method="kendall",
    # corr_method="spearman"
)

df_corr_table_lr_adj.to_clipboard()

####################################################################
func_correlation_matrix_hotmap(
    df=_df_features_lr_adj,
    figsize=6,
    plot_output_fn=None,
)

####################################################################
df_vif_table_lr_adj = func_multicolinearity_vif(
    in_df_features=_df_features_lr_adj,
    with_constant=True,
)
df_vif_table_lr_adj.insert(
    loc=0, column="column_name",
    value=df_vif_table_lr_adj["feature_name"].apply(lambda s0: s0.replace("WOE_", ""))
)
df_vif_table_lr_adj = df_vif_table_lr_adj \
    .sort_values(by=["VIF"], ascending=[False]) \
    .reset_index(drop=True)
df_vif_table_lr_adj



## （训练样本导出）

###################################################################################
# pkl文件

# ####################################################################################
# with open(file="{}/Proc07_Model_LR_adj/_df_model_training_sampling.pkl".format(result_path), mode="wb") as fw:
#     pickle.dump(obj=_df_model_training_sampling, file=fw)

# # ####################################################################################
# # with open(file="{}/Proc07_Model_LR_adj/_df_model_training_sampling.pkl".format(result_path), mode="rb") as fr:
# #     _df_model_training_sampling = pickle.load(file=fr)





## （模型结果导出）

####################################################################################
df_model_summary.to_clipboard(index=False, header=False)
df_model_params.to_clipboard(index=True)

####################################################################################
# to_excel
with open(file="{}/Proc07_Model_LR_adj/df_model_summary.xlsx".format(result_path), mode="wb") as fw:
    df_model_summary \
        .to_excel(
            fw,
            index=False,
            header=False,
            sheet_name="data",
        )
with open(file="{}/Proc07_Model_LR_adj/df_model_params.xlsx".format(result_path), mode="wb") as fw:
    df_model_params \
        .to_excel(
            fw,
            index=True,
            sheet_name="data",
        )
with open(file="{}/Proc07_Model_LR_adj/df_corr_table_lr_adj.xlsx".format(result_path), mode="wb") as fw:
    df_corr_table_lr_adj \
        .to_excel(
            fw,
            index=True,
            sheet_name="data",
        )
with open(file="{}/Proc07_Model_LR_adj/df_vif_table_lr_adj.xlsx".format(result_path), mode="wb") as fw:
    df_vif_table_lr_adj \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc07_Model_LR_adj/df_model_summary.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_model_summary, file=fw)
with open(file="{}/Proc07_Model_LR_adj/df_model_params.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_model_params, file=fw)
with open(file="{}/Proc07_Model_LR_adj/logistic_model_res.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=logistic_model_res, file=fw)
with open(file="{}/Proc07_Model_LR_adj/df_corr_table_lr_adj.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_corr_table_lr_adj, file=fw)
with open(file="{}/Proc07_Model_LR_adj/df_vif_table_lr_adj.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_vif_table_lr_adj, file=fw)


####################################################################################
with open(file="{}/Proc07_Model_LR_adj/df_model_summary.pkl".format(result_path), mode="rb") as fr:
    df_model_summary = pickle.load(file=fr)
with open(file="{}/Proc07_Model_LR_adj/df_model_params.pkl".format(result_path), mode="rb") as fr:
    df_model_params = pickle.load(file=fr)
with open(file="{}/Proc07_Model_LR_adj/logistic_model_res.pkl".format(result_path), mode="rb") as fr:
    logistic_model_res = pickle.load(file=fr)
    model_cols = [s0 for s0 in df_model_params["feature_name"] if s0!="Intercept"]
with open(file="{}/Proc07_Model_LR_adj/df_corr_table_lr_adj.pkl".format(result_path), mode="rb") as fr:
    df_corr_table_lr_adj = pickle.load(file=fr)
with open(file="{}/Proc07_Model_LR_adj/df_vif_table_lr_adj.pkl".format(result_path), mode="rb") as fr:
    df_vif_table_lr_adj = pickle.load(file=fr)

####################################################################################
func_correlation_matrix_hotmap(
    df=_df_features_lr_adj,
    figsize=12,
    plot_output_fn="{}/Proc07_Model_LR_adj/corr_hotmap_lr_adj.png".format(result_path),
)















# # ####################################################################
# # # 计算模型结果
# # _df_model_prediction = df_wb_coarse_woe[cols_base+model_cols].reset_index(drop=True)

# # # 模型方法：sm.Logit
# # _df_model_prediction["_model_out"] = logistic_model_res.predict(
# #     exog=(
# #         sm.add_constant(_df_model_prediction[model_cols], has_constant="add")
# #         if "Intercept" in logistic_model_res.model.exog_names else
# #         _df_model_prediction[model_cols]
# #     )
# # )
# # _df_model_prediction["model_prob"] = _df_model_prediction["_model_out"]
# # _df_model_prediction["model_score"] = _func_model_prob2score(
# #     data_model_prob=_df_model_prediction["model_prob"],
# #     p_Odds=1/50,
# #     p_SCORE=750,
# #     p_PDO=25,
# # )

# # ####################################################################
# # df_t_model_evaluation = []

# # ####################################################################
# # # _ALL_
# # _ks, _ = func_calc_ks_cross(
# #     y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     y_score=_df_model_prediction["model_prob"],
# #     plot=True,
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, "00_ALL"),
# #     plot_figure_type=["distribution", "cum_distribution"],
# #     # plot_figure_type=["distribution"],
# #     # plot_figure_type=["cum_distribution"],
# #     plot_title_remark="DataSet: {}".format("_ALL_"),
# # )
# # _auc = func_calc_auc_roc(
# #     y_labels=_df_model_prediction["flag"],
# #     y_score=_df_model_prediction["model_prob"],
# #     y_score_ascending=True,
# #     plot=True,
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, "00_ALL"),
# #     plot_title_remark="DataSet: {}".format("_ALL_"),
# # )
# # _lift_table, _lift_top_result = func_calc_lift(
# #     y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     y_score=_df_model_prediction["model_prob"],
# #     bins_q=None,
# #     bucket_cnt=20,
# #     lift_calc_ascending=False, lift_top_threshold=0.5,
# #     plot=True,
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, "00_ALL"),
# #     plot_figure_type=["lift", "bad_rate"],
# #     # plot_figure_type=["lift"],
# #     # plot_figure_type=["bad_rate"],
# #     plot_title_remark="DataSet: {}".format("_ALL_"),
# # )
# # # _crosstab_model_porb_bin20 = _func_crosstab_plot_output(
# # #     data=_df_model_prediction["model_prob"],
# # #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
# # # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, "00_ALL"),
# # # )
# # _crosstab_model_porb_bin10 = _func_crosstab_plot_output(
# #     data=_df_model_prediction["model_prob"],
# #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, "00_ALL"),
# # )
# # # _crosstab_model_score_bin05 = _func_crosstab_plot_output(
# # #     data=_df_model_prediction["model_score"],
# # #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
# # # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, "00_ALL"),
# # # )
# # _crosstab_model_score_bin20 = _func_crosstab_plot_output(
# #     data=_df_model_prediction["model_score"],
# #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, "00_ALL"),
# # )

# # #############################
# # df_t_model_evaluation.append(
# #     dict(
# #         list({
# #             "data_role": "00_ALL",
# #             "cnt": _df_model_prediction.shape[0],
# #             "bad_cnt": _df_model_prediction.query("flag==1").shape[0],
# #             "bad_rate": _df_model_prediction.query("flag==1").shape[0]/_df_model_prediction.shape[0],
# #             "ks": _ks["gap"].max(),
# #             "auc": _auc,
# #             "_lift_table": _lift_table,
# # #             "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
# # #             "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
# # #             "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
# # #             "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
# #         }.items())+list(_lift_top_result.items())
# #     )
# # )

# # ####################################################################
# # # data_role
# # for _data_role in _df_model_prediction["data_role"].drop_duplicates().sort_values().tolist()[:]:
# #     _df = _df_model_prediction.query("data_role=='{}'".format(_data_role))
    
# #     #############################
# #     _ks, _ = func_calc_ks_cross(
# #         y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         y_score=_df["model_prob"],
# # #         plot=True,
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, _data_role),
# #         plot_figure_type=["distribution", "cum_distribution"],
# #         # plot_figure_type=["distribution"],
# #         # plot_figure_type=["cum_distribution"],
# #         plot_title_remark="DataSet: {}".format(_data_role),
# #     )
# #     _auc = func_calc_auc_roc(
# #         y_labels=_df["flag"],
# #         y_score=_df["model_prob"],
# #         y_score_ascending=True,
# # #         plot=True,
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, _data_role),
# #         plot_title_remark="DataSet: {}".format(_data_role),
# #     )
# #     _lift_table, _lift_top_result = func_calc_lift(
# #         y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         y_score=_df["model_prob"],
# #         bins_q=None,
# #         bucket_cnt=20,
# #         lift_calc_ascending=False, lift_top_threshold=0.5,
# # #         plot=True,
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, _data_role),
# #         plot_figure_type=["lift", "bad_rate"],
# #         # plot_figure_type=["lift"],
# #         # plot_figure_type=["bad_rate"],
# #         plot_title_remark="DataSet: {}".format(_data_role),
# #     )
# # #     _crosstab_model_porb_bin20 = _func_crosstab_plot_output(
# # #         data=_df["model_prob"],
# # #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, _data_role),
# # #     )
# # #     _crosstab_model_porb_bin10 = _func_crosstab_plot_output(
# # #         data=_df["model_prob"],
# # #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, _data_role),
# # #     )
# # #     _crosstab_model_score_bin05 = _func_crosstab_plot_output(
# # #         data=_df["model_score"],
# # #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, _data_role),
# # #     )
# # #     _crosstab_model_score_bin20 = _func_crosstab_plot_output(
# # #         data=_df["model_score"],
# # #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
# # #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, _data_role),
# # #     )
    
# #     #############################
# #     df_t_model_evaluation.append(
# #         dict(
# #             list({
# #                 "data_role": _data_role,
# #                 "cnt": _df.shape[0],
# #                 "bad_cnt": _df.query("flag==1").shape[0],
# #                 "bad_rate": _df.query("flag==1").shape[0]/_df.shape[0],
# #                 "ks": _ks["gap"].max(),
# #                 "auc": _auc,
# #                 "_lift_table": _lift_table,
# # #                 "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
# # #                 "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
# # #                 "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
# # #                 "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
# #             }.items())+list(_lift_top_result.items())
# #         )
# #     )

# # ####################################################################
# # df_t_model_evaluation = pd.DataFrame(df_t_model_evaluation)

# # df_t_model_evaluation












# # ################################################################################################
# # # 单指标调整（Numerical）
# # # 查看woe趋势图

# # _col = "sb_12m_zzs_qbxse_growth_rate_yoy"
# # # _col = random.choice(list(dict_coarse_bin_mapping_numerical.keys()))
# # # _col = random.choice(cols_v_num)
# # # _bin = [-inf, 0, 0.3, 1, inf]
# # _bin = dict_coarse_bin_mapping_numerical.get(_col)
# # _col, _bin = list(dict_coarse_bin_mapping_numerical.items())[-1]


# # ################################################################################################
# # _f_auto = True
# # _f_auto = False
# # if _f_auto:
# #     ######################################################
# #     _, _crosstab, _bin = func_auto_binning_continuous_v1(
# #         in_var=_df_wb_dev[_col],
# #         in_target=_df_wb_dev["target_label"],
# #         min_pct=0.05-0.0001, max_bins_cnt=20,
# #         # min_pct=0.05-0.0001, max_bins_cnt=6,
# #         # min_pct=0.05-0.0001, max_bins_cnt=5,
# #         # min_pct=0.05-0.0001, max_bins_cnt=4,
# #         # min_pct=0.05-0.0001, max_bins_cnt=3,
# #         # min_pct=0.1-0.0001, max_bins_cnt=9,
# #         # min_pct=0.01-0.0001, max_bins_cnt=50,
# #         method="01_equal_freq",
# #         # method="02_decision_tree",
# #         # method="03_chi2_comb",
# #         # method="09_best_ks",
# #         chi2_min_pvalue=0.1,
# #         chi2_min_cnt=2,
# #         with_lift_ks=False,
# #         # with_lift_ks=True,
# #         lift_calc_ascending=True,
# #         with_wls_adj_woe=True,
# #         woe_with_nan=True,
# #         # woe_with_nan=False,
# #     )
# #     # ######################################################
# #     # _data_converted, _bin, _ = func_binning_continuous_quantile_v1(
# #     #     in_data=_df_wb_dev[_col],
# #     #     # bins_q=[0.5, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99, 99.5],
# #     #     # bins_q=[1, 5, 10, 25, 50, 75, 90, 95, 99],
# #     #     # bins_q=[5, 10, 25, 50, 75, 90, 95],
# #     #     bins_q=[10, 25, 50, 75, 90],
# #     #     # bins_q=[25, 50, 75],
# #     #     out_type="01_info_cp",
# #     #     right_border=True, include_lowest=False,
# #     #     output_cp=True,
# #     # )
# #     # _crosstab = func_woe_report_v1(
# #     #     in_var=_data_converted,
# #     #     in_target=_df_wb_dev["target_label"],
# #     #     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.0001,
# #     #     # with_lift_ks=True,
# #     #     with_lift_ks=False,
# #     #     lift_calc_ascending=True,
# #     #     with_wls_adj_woe=True,
# #     #     woe_with_nan=True,
# #     #     # woe_with_nan=False,
# #     # )
# # else:
# #     ######################################################
# #     _crosstab = func_woe_report_v1(
# #         in_var=func_binning_continuous_v1(
# #             in_data=_df_wb_dev[_col],
# #             bins=_bin,
# #             right_border=True, include_lowest=True,
# #         ),
# #         in_target=_df_wb_dev["target_label"],
# #         with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.0001,
# #         # with_lift_ks=True,
# #         with_lift_ks=False,
# #         lift_calc_ascending=True,
# #         with_wls_adj_woe=True,
# #         woe_with_nan=True,
# #         # woe_with_nan=False,
# #     )


# # print(_col)
# # print(_bin)
# # print()
# # func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=False, with_wls_adj_woe=True)
# # # func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=True, with_wls_adj_woe=True)
# # # func_plot_woe(_crosstab, plot_badrate=False, with_nan_info=False, with_wls_adj_woe=False)

# # # _crosstab.to_clipboard(index=True)
# # _crosstab




























#==============================================================================
# File: code11_Proc08_Model_Evaluation.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc08_Model_Evaluation
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc08_Model_Evaluation
## 模型效果评估


####################################################################
# Proc08_Model_Evaluation
# 模型效果评估


####################################################################
def _func_model_prob2score(
        data_model_prob,
        p_Odds=1/50,
        p_SCORE=750,
        p_PDO=25,
    ):
    _score_wx_sum = (-1)*np.log(1/data_model_prob-1)
    _B = p_PDO/np.log(2)
    _A = p_SCORE+_B*np.log(p_Odds)
    rt = _A+_B*(-1)*_score_wx_sum
    return rt


####################################################################
# 计算模型 KS、AUC、LIFT_TABLE
def _func_crosstab_plot_output(
        data,
        target,
        bins,
        plot_output_fn=None,
    ):
    ###########################
    # 间隔
    _crosstab = func_woe_report_v1(
        in_var=func_binning_continuous_v1(
            in_data=data,
            bins=bins, right_border=True, include_lowest=True,
        ),
        in_target=target,
        with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
        with_lift_ks=True,
        with_wls_adj_woe=True,
    )
    func_plot_woe(
        _crosstab,
        plot_badrate=True,
        with_nan_info=False,
        with_wls_adj_woe=True,
        plot_output_fn=plot_output_fn,
    )
    _crosstab = _crosstab \
        .reset_index(drop=False).rename(columns={"index": "gb_idx"}) \
        [[
            "gb_idx", "0_good_#", "1_bad_#", "0_good_%", "1_bad_%",
            "WOE", "IV", "total", "total_pct", "bad_rate", "ks",
        ]] \
        .reset_index(drop=True)
    return _crosstab



####################################################################
# 计算模型结果
_df_model_prediction = df_wb_coarse_woe[cols_base+model_cols].reset_index(drop=True)

# 模型方法：sm.Logit
_df_model_prediction["_model_out"] = logistic_model_res.predict(
    exog=(
        sm.add_constant(_df_model_prediction[model_cols], has_constant="add")
        if "Intercept" in logistic_model_res.model.exog_names else
        _df_model_prediction[model_cols]
    )
)
_df_model_prediction["model_prob"] = _df_model_prediction["_model_out"]
_df_model_prediction["model_score"] = _func_model_prob2score(
    data_model_prob=_df_model_prediction["model_prob"],
    # p_Odds=1/50,
    p_Odds=1/20,
    p_SCORE=700,
    # p_PDO=25,
    p_PDO=60,
)



####################################################################
df_t_model_evaluation = []

####################################################################
# _ALL_
_ks, _ = func_calc_ks_cross(
    y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    y_score=_df_model_prediction["model_prob"],
    plot=True,
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, "00_ALL"),
    plot_figure_type=["distribution", "cum_distribution"],
    # plot_figure_type=["distribution"],
    # plot_figure_type=["cum_distribution"],
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
_auc = func_calc_auc_roc(
    y_labels=_df_model_prediction["flag"],
    y_score=_df_model_prediction["model_prob"],
    y_score_ascending=True,
    plot=True,
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, "00_ALL"),
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
_lift_table, _lift_top_result = func_calc_lift(
    y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    y_score=_df_model_prediction["model_prob"],
    bins_q=None,
    bucket_cnt=20,
    lift_calc_ascending=False, lift_top_threshold=0.5,
    plot=True,
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, "00_ALL"),
    plot_figure_type=["lift", "bad_rate"],
    # plot_figure_type=["lift"],
    # plot_figure_type=["bad_rate"],
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
_crosstab_model_porb_bin20 = _func_crosstab_plot_output(
    data=_df_model_prediction["model_prob"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, "00_ALL"),
)
_crosstab_model_porb_bin10 = _func_crosstab_plot_output(
    data=_df_model_prediction["model_prob"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, "00_ALL"),
)
_crosstab_model_score_bin05 = _func_crosstab_plot_output(
    data=_df_model_prediction["model_score"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, "00_ALL"),
)
_crosstab_model_score_bin20 = _func_crosstab_plot_output(
    data=_df_model_prediction["model_score"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
    plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, "00_ALL"),
)

#############################
df_t_model_evaluation.append(
    dict(
        list({
            "data_role": "00_ALL",
            "cnt": _df_model_prediction.shape[0],
            "bad_cnt": _df_model_prediction.query("flag==1").shape[0],
            "bad_rate": _df_model_prediction.query("flag==1").shape[0]/_df_model_prediction.shape[0],
            "ks": _ks["gap"].max(),
            "auc": _auc,
            "_lift_table": _lift_table,
            "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
            "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
            "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
            "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
        }.items())+list(_lift_top_result.items())
    )
)

####################################################################
# data_role
for _data_role in _df_model_prediction["data_role"].drop_duplicates().sort_values().tolist()[:]:
    _df = _df_model_prediction.query("data_role=='{}'".format(_data_role))
    
    #############################
    _ks, _ = func_calc_ks_cross(
        y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        y_score=_df["model_prob"],
        plot=True,
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, _data_role),
        plot_figure_type=["distribution", "cum_distribution"],
        # plot_figure_type=["distribution"],
        # plot_figure_type=["cum_distribution"],
        plot_title_remark="DataSet: {}".format(_data_role),
    )
    _auc = func_calc_auc_roc(
        y_labels=_df["flag"],
        y_score=_df["model_prob"],
        y_score_ascending=True,
        plot=True,
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, _data_role),
        plot_title_remark="DataSet: {}".format(_data_role),
    )
    _lift_table, _lift_top_result = func_calc_lift(
        y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        y_score=_df["model_prob"],
        bins_q=None,
        bucket_cnt=20,
        lift_calc_ascending=False, lift_top_threshold=0.5,
        plot=True,
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, _data_role),
        plot_figure_type=["lift", "bad_rate"],
        # plot_figure_type=["lift"],
        # plot_figure_type=["bad_rate"],
        plot_title_remark="DataSet: {}".format(_data_role),
    )
    _crosstab_model_porb_bin20 = _func_crosstab_plot_output(
        data=_df["model_prob"],
        target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, _data_role),
    )
    _crosstab_model_porb_bin10 = _func_crosstab_plot_output(
        data=_df["model_prob"],
        target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, _data_role),
    )
    _crosstab_model_score_bin05 = _func_crosstab_plot_output(
        data=_df["model_score"],
        target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, _data_role),
    )
    _crosstab_model_score_bin20 = _func_crosstab_plot_output(
        data=_df["model_score"],
        target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, _data_role),
    )
    
    #############################
    df_t_model_evaluation.append(
        dict(
            list({
                "data_role": _data_role,
                "cnt": _df.shape[0],
                "bad_cnt": _df.query("flag==1").shape[0],
                "bad_rate": _df.query("flag==1").shape[0]/_df.shape[0],
                "ks": _ks["gap"].max(),
                "auc": _auc,
                "_lift_table": _lift_table,
                "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
                "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
                "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
                "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
            }.items())+list(_lift_top_result.items())
        )
    )

####################################################################
df_t_model_evaluation = pd.DataFrame(df_t_model_evaluation)
print(df_t_model_evaluation[['data_role', 'cnt', 'bad_cnt', 'bad_rate', 'ks', 'auc']])

# ####################################################################
# # to_clipboard
# df_t_model_evaluation.to_clipboard(index=False)










####################################################################################
# to_excel
with open(file="{}/Proc08_Model_Evaluation/df_t_model_evaluation.xlsx".format(result_path), mode="wb") as fw:
    df_t_model_evaluation \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc08_Model_Evaluation/df_t_model_evaluation.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_model_evaluation, file=fw)

####################################################################################
with open(file="{}/Proc08_Model_Evaluation/df_t_model_evaluation.pkl".format(result_path), mode="rb") as fr:
    df_t_model_evaluation = pickle.load(file=fr)

df_t_model_evaluation.drop(
    labels=[s0 for s0 in df_t_model_evaluation.columns if re.search("^_", s0)!=None],
    axis=1,
)






####################################################################
# 计算模型 KS、AUC、LIFT_TABLE
# 图像输出

####################################################################
print("----------------------------")
print("_ALL_")
print("----------------------------")

# ks
ks, crossdens = func_calc_ks_cross(
    y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    y_score=_df_model_prediction["model_prob"],
    plot=True, plot_output_fn=None,
    plot_figure_type=["distribution", "cum_distribution"],
    # plot_figure_type=["distribution"],
    # plot_figure_type=["cum_distribution"],
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
print("{:6}: {:7.4f}".format("KS", ks["gap"].max()))

# auc
auc = func_calc_auc_roc(
    y_labels=_df_model_prediction["flag"],
    y_score=_df_model_prediction["model_prob"],
    y_score_ascending=True,
    plot=True, plot_output_fn=None,
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
print("{:6}: {:7.4f}".format("AUC", auc))

# lift_table
lift_table, lift_top_result = func_calc_lift(
    y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    y_score=_df_model_prediction["model_prob"],
    bins_q=None,
    bucket_cnt=20,
    lift_calc_ascending=False, lift_top_threshold=0.5,
    plot=True, plot_output_fn=None,
    plot_figure_type=["lift", "bad_rate"],
    # plot_figure_type=["lift"],
    # plot_figure_type=["bad_rate"],
    plot_title_remark="DataSet: {}".format("_ALL_"),
)
print("{:6}: {:}".format("LIFT", lift_top_result))

_crosstab_model_prob_bin20 = _func_crosstab_plot_output(
    data=_df_model_prediction["model_prob"],
    # data=_df_model_prediction["model_score"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
    plot_output_fn=None,
)
_crosstab_model_score_bin05 = _func_crosstab_plot_output(
    # data=_df_model_prediction["model_prob"],
    data=_df_model_prediction["model_score"],
    target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
    bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
    # bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
    plot_output_fn=None,
)






# ### （直接跑测模型效果，不设置导出结果到result_path）

# ####################################################################
# # 计算模型结果
# _df_model_prediction = _df_wb_coarse_woe[cols_base+model_cols].reset_index(drop=True)

# # _df_model_prediction = df_wb[df_wb["zcfzb_assets"].notna()][["uid", "observation_dt"]] \
# #     .merge(
# #         right=_df_model_prediction,
# #         how="left", left_on=["uid", "observation_dt"], right_on=["uid", "observation_dt"],
# #     ).reset_index(drop=True)


# # 模型方法：sm.Logit
# _df_model_prediction["_model_out"] = logistic_model_res.predict(
#     exog=(
#         sm.add_constant(_df_model_prediction[model_cols], has_constant="add")
#         if "Intercept" in logistic_model_res.model.exog_names else
#         _df_model_prediction[model_cols]
#     )
# )


# _df_model_prediction["model_prob"] = _df_model_prediction["_model_out"]
# _df_model_prediction["model_score"] = _func_model_prob2score(
#     data_model_prob=_df_model_prediction["model_prob"],
#     p_Odds=1/50,
#     p_SCORE=750,
#     p_PDO=25,
# )

# ####################################################################
# df_t_model_evaluation = []

# ####################################################################
# # _ALL_
# _ks, _ = func_calc_ks_cross(
#     y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     y_score=_df_model_prediction["model_prob"],
#     plot=True,
# #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, "00_ALL"),
#     plot_figure_type=["distribution", "cum_distribution"],
#     # plot_figure_type=["distribution"],
#     # plot_figure_type=["cum_distribution"],
#     plot_title_remark="DataSet: {}".format("_ALL_"),
# )
# _auc = func_calc_auc_roc(
#     y_labels=_df_model_prediction["flag"],
#     y_score=_df_model_prediction["model_prob"],
#     y_score_ascending=True,
#     plot=True,
# #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, "00_ALL"),
#     plot_title_remark="DataSet: {}".format("_ALL_"),
# )
# _lift_table, _lift_top_result = func_calc_lift(
#     y_labels=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     y_score=_df_model_prediction["model_prob"],
#     bins_q=None,
#     bucket_cnt=20,
#     lift_calc_ascending=False, lift_top_threshold=0.5,
#     plot=True,
# #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, "00_ALL"),
#     plot_figure_type=["lift", "bad_rate"],
#     # plot_figure_type=["lift"],
#     # plot_figure_type=["bad_rate"],
#     plot_title_remark="DataSet: {}".format("_ALL_"),
# )
# # _crosstab_model_porb_bin20 = _func_crosstab_plot_output(
# #     data=_df_model_prediction["model_prob"],
# #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, "00_ALL"),
# # )
# _crosstab_model_porb_bin10 = _func_crosstab_plot_output(
#     data=_df_model_prediction["model_prob"],
#     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
# #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, "00_ALL"),
# )
# # _crosstab_model_score_bin05 = _func_crosstab_plot_output(
# #     data=_df_model_prediction["model_score"],
# #     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
# # #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, "00_ALL"),
# # )
# _crosstab_model_score_bin20 = _func_crosstab_plot_output(
#     data=_df_model_prediction["model_score"],
#     target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+10, 10)]+[np.inf],
# #     plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, "00_ALL"),
# )

# #############################
# df_t_model_evaluation.append(
#     dict(
#         list({
#             "data_role": "00_ALL",
#             "cnt": _df_model_prediction.shape[0],
#             "bad_cnt": _df_model_prediction.query("flag==1").shape[0],
#             "bad_rate": _df_model_prediction.query("flag==1").shape[0]/_df_model_prediction.shape[0],
#             "ks": _ks["gap"].max(),
#             "auc": _auc,
#             "_lift_table": _lift_table,
# #             "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
# #             "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
# #             "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
# #             "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
#         }.items())+list(_lift_top_result.items())
#     )
# )

# ####################################################################
# # data_role
# for _data_role in _df_model_prediction["data_role"].drop_duplicates().sort_values().tolist()[:]:
#     _df = _df_model_prediction.query("data_role=='{}'".format(_data_role))
    
#     #############################
#     _ks, _ = func_calc_ks_cross(
#         y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#         y_score=_df["model_prob"],
# #         plot=True,
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, _data_role),
#         plot_figure_type=["distribution", "cum_distribution"],
#         # plot_figure_type=["distribution"],
#         # plot_figure_type=["cum_distribution"],
#         plot_title_remark="DataSet: {}".format(_data_role),
#     )
#     _auc = func_calc_auc_roc(
#         y_labels=_df["flag"],
#         y_score=_df["model_prob"],
#         y_score_ascending=True,
# #         plot=True,
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, _data_role),
#         plot_title_remark="DataSet: {}".format(_data_role),
#     )
#     _lift_table, _lift_top_result = func_calc_lift(
#         y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#         y_score=_df["model_prob"],
#         bins_q=None,
#         bucket_cnt=20,
#         lift_calc_ascending=False, lift_top_threshold=0.5,
# #         plot=True,
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_lift.png".format(result_path, _data_role),
#         plot_figure_type=["lift", "bad_rate"],
#         # plot_figure_type=["lift"],
#         # plot_figure_type=["bad_rate"],
#         plot_title_remark="DataSet: {}".format(_data_role),
#     )
# #     _crosstab_model_porb_bin20 = _func_crosstab_plot_output(
# #         data=_df["model_prob"],
# #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/20)]+[np.inf],
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin20.png".format(result_path, _data_role),
# #     )
# #     _crosstab_model_porb_bin10 = _func_crosstab_plot_output(
# #         data=_df["model_prob"],
# #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(0, 1, 1/10)]+[np.inf],
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_prob_bin10.png".format(result_path, _data_role),
# #     )
# #     _crosstab_model_score_bin05 = _func_crosstab_plot_output(
# #         data=_df["model_score"],
# #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+5, 5)]+[np.inf],
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin05.png".format(result_path, _data_role),
# #     )
# #     _crosstab_model_score_bin20 = _func_crosstab_plot_output(
# #         data=_df["model_score"],
# #         target=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# #         bins=[-np.inf]+[round(s0, 6) for s0 in np.arange(500, 800+20, 20)]+[np.inf],
# #         plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_model_score_bin20.png".format(result_path, _data_role),
# #     )
    
#     #############################
#     df_t_model_evaluation.append(
#         dict(
#             list({
#                 "data_role": _data_role,
#                 "cnt": _df.shape[0],
#                 "bad_cnt": _df.query("flag==1").shape[0],
#                 "bad_rate": _df.query("flag==1").shape[0]/_df.shape[0],
#                 "ks": _ks["gap"].max(),
#                 "auc": _auc,
#                 "_lift_table": _lift_table,
# #                 "_crosstab_model_porb_bin20": _crosstab_model_porb_bin20,
# #                 "_crosstab_model_porb_bin10": _crosstab_model_porb_bin10,
# #                 "_crosstab_model_score_bin05": _crosstab_model_score_bin05,
# #                 "_crosstab_model_score_bin20": _crosstab_model_score_bin20,
#             }.items())+list(_lift_top_result.items())
#         )
#     )

# ####################################################################
# df_t_model_evaluation = pd.DataFrame(df_t_model_evaluation)
# df_t_model_evaluation.drop(labels=["_lift_table"], axis=1).to_clipboard()
# df_t_model_evaluation


























# # # ####################################################################
# # # # 计算模型 KS、AUC、LIFT_TABLE
# # # # 图像输出

# # # ####################################################################
# # # # 01_train
# # # print("----------------------------")
# # # print("01_train")
# # # print("----------------------------")
# # # # ks
# # # ks, crossdens = func_calc_ks_cross(
# # #     y_labels=_df_model_prediction.query("data_role=='01_train'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='01_train'")["model_prob"],
# # #     plot=True, plot_output_fn=None,
# # #     # plot_figure_type=["distribution", "cum_distribution"],
# # #     # plot_figure_type=["distribution"],
# # #     plot_figure_type=["cum_distribution"],
# # #     plot_title_remark="DataSet: {}".format("01_train"),
# # # )
# # # print("{:4}: {:7.4f}".format("KS", ks["gap"].max()))
# # # # auc
# # # auc = func_calc_auc_roc(
# # #     y_labels=_df_model_prediction.query("data_role=='01_train'")["flag"],
# # #     y_score=_df_model_prediction.query("data_role=='01_train'")["model_prob"],
# # #     y_score_ascending=True,
# # #     plot=True,
# # #     plot_output_fn=None,
# # #     plot_title_remark="DataSet: {}".format("01_train"),
# # # )
# # # print("{:4}: {:7.4f}".format("AUC", auc))
# # # # lift_table
# # # lift_table, lift_top_result = func_calc_lift(
# # #     y_labels=_df_model_prediction.query("data_role=='01_train'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='01_train'")["model_prob"],
# # #     bins_q=None,
# # #     bucket_cnt=20,
# # #     lift_calc_ascending=False, lift_top_threshold=0.5,
# # #     plot=True, plot_output_fn=None,
# # #     plot_figure_type=["lift", "bad_rate"],
# # #     # plot_figure_type=["lift"],
# # #     # plot_figure_type=["bad_rate"],
# # #     plot_title_remark="DataSet: {}".format("01_train"),
# # # )
# # # print("{:6}: {:}".format("LIFT", lift_top_result))

# # # ####################################################################
# # # # 02_test
# # # print("----------------------------")
# # # print("02_test")
# # # print("----------------------------")
# # # # ks
# # # ks, crossdens = func_calc_ks_cross(
# # #     y_labels=_df_model_prediction.query("data_role=='02_test'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='02_test'")["model_prob"],
# # #     plot=True, plot_output_fn=None,
# # #     # plot_figure_type=["distribution", "cum_distribution"],
# # #     # plot_figure_type=["distribution"],
# # #     plot_figure_type=["cum_distribution"],
# # #     plot_title_remark="DataSet: {}".format("02_test"),
# # # )
# # # print("{:4}: {:7.4f}".format("KS", ks["gap"].max()))
# # # # auc
# # # auc = func_calc_auc_roc(
# # #     y_labels=_df_model_prediction.query("data_role=='02_test'")["flag"],
# # #     y_score=_df_model_prediction.query("data_role=='02_test'")["model_prob"],
# # #     y_score_ascending=True,
# # #     plot=True,
# # #     plot_output_fn=None,
# # #     plot_title_remark="DataSet: {}".format("02_test"),
# # # )
# # # print("{:4}: {:7.4f}".format("AUC", auc))
# # # # lift_table
# # # lift_table, lift_top_result = func_calc_lift(
# # #     y_labels=_df_model_prediction.query("data_role=='02_test'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='02_test'")["model_prob"],
# # #     bins_q=None,
# # #     bucket_cnt=20,
# # #     lift_calc_ascending=False, lift_top_threshold=0.5,
# # #     plot=True, plot_output_fn=None,
# # #     plot_figure_type=["lift", "bad_rate"],
# # #     # plot_figure_type=["lift"],
# # #     # plot_figure_type=["bad_rate"],
# # #     plot_title_remark="DataSet: {}".format("02_test"),
# # # )
# # # print("{:6}: {:}".format("LIFT", lift_top_result))

# # # ####################################################################
# # # # 03_oos
# # # print("----------------------------")
# # # print("03_oos")
# # # print("----------------------------")
# # # # ks
# # # ks, crossdens = func_calc_ks_cross(
# # #     y_labels=_df_model_prediction.query("data_role=='03_oos'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='03_oos'")["model_prob"],
# # #     plot=True, plot_output_fn=None,
# # #     # plot_figure_type=["distribution", "cum_distribution"],
# # #     # plot_figure_type=["distribution"],
# # #     plot_figure_type=["cum_distribution"],
# # #     plot_title_remark="DataSet: {}".format("03_oos"),
# # # )
# # # print("{:4}: {:7.4f}".format("KS", ks["gap"].max()))
# # # # auc
# # # auc = func_calc_auc_roc(
# # #     y_labels=_df_model_prediction.query("data_role=='03_oos'")["flag"],
# # #     y_score=_df_model_prediction.query("data_role=='03_oos'")["model_prob"],
# # #     y_score_ascending=True,
# # #     plot=True,
# # #     plot_output_fn=None,
# # #     plot_title_remark="DataSet: {}".format("03_oos"),
# # # )
# # # print("{:4}: {:7.4f}".format("AUC", auc))
# # # # lift_table
# # # lift_table, lift_top_result = func_calc_lift(
# # #     y_labels=_df_model_prediction.query("data_role=='03_oos'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='03_oos'")["model_prob"],
# # #     bins_q=None,
# # #     bucket_cnt=20,
# # #     lift_calc_ascending=False, lift_top_threshold=0.5,
# # #     plot=True, plot_output_fn=None,
# # #     plot_figure_type=["lift", "bad_rate"],
# # #     # plot_figure_type=["lift"],
# # #     # plot_figure_type=["bad_rate"],
# # #     plot_title_remark="DataSet: {}".format("03_oos"),
# # # )
# # # print("{:6}: {:}".format("LIFT", lift_top_result))

# # # ####################################################################
# # # # 04_oot
# # # print("----------------------------")
# # # print("04_oot")
# # # print("----------------------------")
# # # # ks
# # # ks, crossdens = func_calc_ks_cross(
# # #     y_labels=_df_model_prediction.query("data_role=='04_oot'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='04_oot'")["model_prob"],
# # #     plot=True, plot_output_fn=None,
# # #     # plot_figure_type=["distribution", "cum_distribution"],
# # #     # plot_figure_type=["distribution"],
# # #     plot_figure_type=["cum_distribution"],
# # #     plot_title_remark="DataSet: {}".format("04_oot"),
# # # )
# # # print("{:4}: {:7.4f}".format("KS", ks["gap"].max()))
# # # # auc
# # # auc = func_calc_auc_roc(
# # #     y_labels=_df_model_prediction.query("data_role=='04_oot'")["flag"],
# # #     y_score=_df_model_prediction.query("data_role=='04_oot'")["model_prob"],
# # #     y_score_ascending=True,
# # #     plot=True,
# # #     plot_output_fn=None,
# # #     plot_title_remark="DataSet: {}".format("04_oot"),
# # # )
# # # print("{:4}: {:7.4f}".format("AUC", auc))
# # # # lift_table
# # # lift_table, lift_top_result = func_calc_lift(
# # #     y_labels=_df_model_prediction.query("data_role=='04_oot'")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
# # #     y_score=_df_model_prediction.query("data_role=='04_oot'")["model_prob"],
# # #     bins_q=None,
# # #     bucket_cnt=20,
# # #     lift_calc_ascending=False, lift_top_threshold=0.5,
# # #     plot=True, plot_output_fn=None,
# # #     plot_figure_type=["lift", "bad_rate"],
# # #     # plot_figure_type=["lift"],
# # #     # plot_figure_type=["bad_rate"],
# # #     plot_title_remark="DataSet: {}".format("04_oot"),
# # # )
# # # print("{:6}: {:}".format("LIFT", lift_top_result))






# # 分数分布

# df_yx_rej = pd.merge(df_dpd[df_dpd['auth_status'] == '拒绝']
#          , data_all[['order_no', 
#                     "INQUIRY0047",
#                     "digitalScore",
#                     "CARD0065",
#                     "query_tfau_mg",
#                     "repay_dcal_tmd",
#                     "pboc_education_new",
#                     "income_izai_tmf",
#                     "quad_p1",
#                     "debts_sgad_mc",
#                     "repay_dhaa_md",
#                     "loan_overdue_last_months_l12m",
#                     ]]
#          , left_on = 'auth_order_no'
#          , right_on = 'order_no'
#          , how = 'inner'
#          )



# df_yx_rej['三方数据源_hengpu_4_sc'] = \
#     df_yx_rej['三方数据源_hengpu_4'].apply(lambda x: 39.68 if pd.isna(x)
#                                       else 87.08 if x<=25
#                                       else 48.87 if x<=30
#                                       else 30.54 if x<=35
#                                       else 29.55 if x<=40
#                                       else 25.05
#                                       )


# df_yx_rej['INQUIRY0047_sc'] = \
#     df_yx_rej['INQUIRY0047'].apply(lambda x: 16.50 if pd.isna(x)
#                                       else 51.88 if x<=0.35
#                                       else 29.79 if x<=0.50
#                                       else 25.55
#                                       )


# df_yx_rej['三方数据源_hangliezhi_1_sc'] = \
#     df_yx_rej['三方数据源_hangliezhi_1'].apply(lambda x: 31.94 if pd.isna(x)
#                                       else 32.47 if x<=495
#                                       else 39.83 if x<=515
#                                       else 59.66 if x<=535
#                                       else 91.33
#                                       )


# df_yx_rej['digitalScore_sc'] = \
#     df_yx_rej['digitalScore'].apply(lambda x: 30.80 if pd.isna(x)
#                                       else 22.75 if x<=800
#                                       else 45.45 if x<=825
#                                       else 51.05 if x<=870
#                                       else 55.97
#                                       )


# df_yx_rej['CARD0065_sc'] = \
#     df_yx_rej['CARD0065'].apply(lambda x: 33.09 if pd.isna(x)
#                                       else 43.99 if x<=0
#                                       else 6.43
#                                       )


# df_yx_rej['三方数据源_ruizhi_6_sc'] = \
#     df_yx_rej['三方数据源_ruizhi_6'].apply(lambda x: 39.68 if pd.isna(x)
#                                       else 28.67 if x<=720
#                                       else 32.58 if x<=765
#                                       else 40.38 if x<=800
#                                       else 45.49 if x<=840
#                                       else 58.63
#                                       )


# df_yx_rej['query_tfau_mg_sc'] = \
#     df_yx_rej['query_tfau_mg'].apply(lambda x: 14.01 if pd.isna(x)
#                                       else 31.55 if x<=1
#                                       else 38.63 if x<=10
#                                       else 48.16 if x<=30
#                                       else 51.41
#                                       )


# df_yx_rej['pboc_education_new_sc'] = \
#     df_yx_rej['pboc_education_new'].apply(lambda x: 33.49 if pd.isna(x)
#                                       else 54.09 if x<=5
#                                       else 38.90
#                                       )



# df_yx_rej['income_izai_tmf_sc'] = \
#     df_yx_rej['income_izai_tmf'].apply(lambda x: 37.97 if pd.isna(x)
#                                       else 34.20 if x<=2000
#                                       else 48.97 if x<=4000
#                                       else 76.60 if x<=6000
#                                       else 88.39
#                                       )


# df_yx_rej['quad_p1_sc'] = \
#     df_yx_rej['quad_p1'].apply(lambda x: 32.62 if pd.isna(x)
#                                       else 44.33 if x<=1
#                                       else 42.33 if x<=2
#                                       else 39.44 if x<=3
#                                       else 18.64
#                                       )

# df_yx_rej['三方数据源_bairong_14_sc'] = \
#     df_yx_rej['三方数据源_bairong_14'].apply(lambda x: 39.68 if pd.isna(x)
#                                       else 32.90 if x<=640
#                                       else 40.61 if x<=700
#                                       else 43.52 if x<=780
#                                       else 46.06
#                                       )


# df_yx_rej['loan_overdue_last_months_l12m_sc'] = \
#     df_yx_rej['loan_overdue_last_months_l12m'].apply(lambda x: 36.85 if pd.isna(x)
#                                       else 40.65 if x<=0
#                                       else 20.55
#                                       )
    
 
# df_yx_rej['三方数据源_xinyongsuanli_1_sc'] = \
#     df_yx_rej['三方数据源_xinyongsuanli_1'].apply(lambda x: 39.68 if pd.isna(x)
#                                       else 31.61 if x<=620
#                                       else 36.20 if x<=640
#                                       else 38.76 if x<=670
#                                       else 43.45 if x<=700
#                                       else 47.57
#                                       )    

  
# df_yx_rej['debts_sgad_mc_sc'] = \
#     df_yx_rej['debts_sgad_mc'].apply(lambda x: 34.92 if pd.isna(x)
#                                       else 30.77 if x<=1
#                                       else 43.19
#                                       )
   
# df_yx_rej['repay_dcal_tmd_sc'] = \
#     df_yx_rej['repay_dcal_tmd'].apply(lambda x: 43.48 if pd.isna(x)
#                                       else 14.77 if x<=30
#                                       else 26.97 if x<=50
#                                       else 44.09
#                                       )   


# df_yx_rej['repay_dhaa_md_sc'] = \
#     df_yx_rej['repay_dhaa_md'].apply(lambda x: 37.05 if pd.isna(x)
#                                       else 31.22 if x<=40
#                                       else 42.26 if x<=150
#                                       else 44.23
#                                       )   
    
    

# df_yx_rej['sum_sc'] = df_yx_rej[['三方数据源_hangliezhi_1_sc',
#  'digitalScore_sc',
#  'CARD0065_sc',
#  '三方数据源_ruizhi_6_sc',
#  'query_tfau_mg_sc',
#  'loan_overdue_last_months_l12m_sc',
#  '三方数据源_hengpu_4_sc',
#  'INQUIRY0047_sc',
#  'pboc_education_new_sc',
#  'quad_p1_sc',
#  '三方数据源_bairong_14_sc',
#  '三方数据源_xinyongsuanli_1_sc',
#  'debts_sgad_mc_sc',
#  'repay_dhaa_md_sc',
#  'income_izai_tmf_sc',
#  'repay_dcal_tmd_sc']].sum(1)

# df_yx_rej['flag'] = np.random.randint(0,2,size=len(df_yx_rej))    





# # 等距分箱
# _bin = [-inf]+list(np.arange(500, 900+20, 20))+[inf]
# # 有标签样本分布

# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=_df_model_prediction["model_score"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)



# # 拒绝样本分布
# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=df_yx_rej["sum_sc"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=df_yx_rej["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)





# # 等频分箱
# _bin = [-inf]+list(pd.qcut(_df_model_prediction["model_score"], 20, retbins = True)[1][1:])+[inf]

# # 有标签样本分布
# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=_df_model_prediction["model_score"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=_df_model_prediction["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)



# # 拒绝样本分布
# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=df_yx_rej["sum_sc"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=df_yx_rej["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)






# # train——等频分箱
# _bin = [-inf]+list(pd.qcut(_df_model_prediction[_df_model_prediction.data_role == '01_train']["model_score"], 15, retbins = True)[1][1:])+[inf]

# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=_df_model_prediction[_df_model_prediction.data_role == '01_train']["model_score"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=_df_model_prediction[_df_model_prediction.data_role == '01_train']["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)







# # oot——等频分箱
# _bin = [-inf]+list(pd.qcut(_df_model_prediction[_df_model_prediction.data_role == '04_oot']["model_score"], 15, retbins = True)[1][1:])+[inf]

# _crosstab = func_woe_report_v1(
#     in_var=func_binning_continuous_v1(
#         in_data=_df_model_prediction[_df_model_prediction.data_role == '04_oot']["model_score"],
#         # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
#         bins=_bin, right_border=True, include_lowest=True,
#     ),
#     in_target=_df_model_prediction[_df_model_prediction.data_role == '04_oot']["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
#     with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
#     with_lift_ks=True,
#     with_wls_adj_woe=True,
# )

# func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# # _crosstab
# print(_crosstab)









# 逐月ks
_tmp = _df_model_prediction.copy()
_tmp['create_month'] = _tmp.dt.apply(lambda x: x[:7])


for _create_month in _tmp["create_month"].drop_duplicates().sort_values().tolist()[:-1]:
    _df = _tmp.query("create_month=='{}'".format(_create_month))
    
    #############################
    _ks, _ = func_calc_ks_cross(
        y_labels=_df["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
        y_score=_df["model_prob"],
        plot=False,
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_ks.png".format(result_path, _create_month),
        plot_figure_type=["distribution", "cum_distribution"],
        # plot_figure_type=["distribution"],
        # plot_figure_type=["cum_distribution"],
        plot_title_remark="DataSet: {}".format(_create_month),
    )
    _auc = func_calc_auc_roc(
        y_labels=_df["flag"],
        y_score=_df["model_prob"],
        y_score_ascending=True,
        plot=False,
        plot_output_fn="{}/Proc08_Model_Evaluation/fig_evaluation_{}_auc.png".format(result_path, _create_month),
        plot_title_remark="DataSet: {}".format(_create_month),
    )
    print(_create_month, _ks, _auc)


#==============================================================================
# File: code12_Proc09_Scorecard_Trans.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc09_Scorecard_Trans
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc09_Scorecard_Trans
## 评分卡模型（评分转换）


####################################################################
# Proc09_Scorecard_Trans
# 评分卡模型（评分转换）


####################################################################
# _func_lr2scorecard
def _func_lr2scorecard(
        feature_names,
        coef,
        df_woe_report,
        p_Odds=1/50,
        p_SCORE=750,
        p_PDO=25,
        missing_fill_type="0_zero",
        num_woe_type="0_ori",
        ):
    
    ####################################################################
    _B = p_PDO/np.log(2)
    _A = p_SCORE+_B*np.log(p_Odds)
    
    ####################################################################
    # _df_t_model_coef
    _df_t_model_coef = pd.DataFrame(
        data=zip(feature_names, coef),
        columns=["feature_name", "Coef"],
    ).reset_index(drop=False).rename(columns={"index": "var_idx"})
    
    ####################################################################
    # _df_t_woe_report
    _df_t_woe_report = df_woe_report \
        [df_woe_report["feature_name"].isin(feature_names)] \
        [[
            "feature_name", "data_type", "gb_idx", "label", "WOE", "wls_adj_WOE",
            # "0_good_#", "1_bad_#", "bad_rate",
        ]] \
        .rename(columns={"WOE": "_WOE", "wls_adj_WOE": "_wls_adj_WOE"}) \
        .reset_index(drop=True)
    
    _df_t_woe_report["gb_idx"] = _df_t_woe_report["gb_idx"].apply(lambda s0: int(s0.split("_")[0]))
    _df_t_woe_report["_label"] = _df_t_woe_report["label"]
    _df_t_woe_report["_missing_tag"] = _df_t_woe_report["label"].apply(
        lambda s0: (
            "Y"
            if re.search("(NaN, NaN)|(NaN)", s0)!=None else
            ""
        ),
    )
    _df_t_woe_report["WOE"] = _df_t_woe_report.apply(
        lambda s0: (
            (
                (
                    0
                    if missing_fill_type=="0_zero" else
                    (0 if pd.isna(s0["_WOE"]) else s0["_WOE"])
                    if missing_fill_type=="1_woe" else
                    np.NaN
                )
                if s0["_missing_tag"]=="Y" else
                s0["_WOE"]
            )
            if s0["data_type"]=="Categorical" else

            (
                (
                    0
                    if missing_fill_type=="0_zero" else
                    (0 if pd.isna(s0["_WOE"]) else s0["_WOE"])
                    if missing_fill_type=="1_woe" else
                    np.NaN
                )
                if s0["_missing_tag"]=="Y" else
                (
                    s0["_WOE"]
                    if num_woe_type=="0_ori" else
                    s0["_wls_adj_WOE"]
                    if num_woe_type=="1_wls_adj" else
                    np.NaN
                )
            )
            if s0["data_type"]=="Numerical" else
            np.NaN
        ),
        axis=1,
    )
    
    ###### 230916添加 ######
    # 把下面几个分的空箱WOE替换成0
    _df_t_woe_report.loc[((_df_t_woe_report['_missing_tag'] == 'Y')&\
                          (_df_t_woe_report['feature_name'].isin([
                                                            '三方数据源_hengpu_4', 
                                                            '三方数据源_ruizhi_6', 
                                                            '三方数据源_bairong_14', 
                                                            '三方数据源_hangliezhi_1',
                                                            # 'INQUIRY0047',
                                                            'digitalScore', 
                                                            'CARD0065',
                                                           # 'query_tfau_mg', 
                                                           'pboc_education_new', 
                                                           'income_izai_tmf', 
                                                           'quad_p1',
                                                           'loan_overdue_last_months_l12m', 
                                                           'debts_sgad_mc', 
                                                           'repay_dcal_tmd',
                                                           'repay_dhaa_md'
                                                                  
                                                                  
                                                                  ]))),\
                          'WOE'] = 0
    

    ###### 230916添加 ######
    
    
    _df_t_woe_report["_eq_side"] = _df_t_woe_report.apply(
        lambda s0: (
            ""
            if s0["data_type"]=="Categorical" else
            (
                "1_left" if re.search("\[", s0["label"])!=None else
                "2_right" if re.search("\]", s0["label"])!=None else
                ""
            )
            if s0["data_type"]=="Numerical" else
            ""
        ),
        axis=1,
    )
    ####################################################################
    _cond = (
        _df_t_woe_report["feature_name"].isin(_df_t_woe_report.query("_missing_tag=='Y'")["feature_name"].unique())
    )
    _df = _df_t_woe_report[-_cond] \
        [["feature_name", "data_type"]].drop_duplicates() \
        .reset_index(drop=True)
    _df["gb_idx"] = 0
    _df["label"] = _df["data_type"].apply(
        lambda s0: (
            "NaN" if s0=="Categorical" else
            "(NaN, NaN]" if s0=="Numerical" else
            ""
        ),
    )
    _df["_label"] = _df["label"]
    _df["_missing_tag"] = "Y"
    _df["WOE"] = 0
    _m = _df_t_woe_report[_df_t_woe_report["feature_name"].isin(_df["feature_name"])][["feature_name", "_eq_side"]].drop_duplicates(subset=["feature_name"]).set_index(keys=["feature_name"])["_eq_side"].to_dict()
    _df["_eq_side"] = _df["feature_name"].apply(lambda s0: _m.get(s0))
    ####################################################################
    _df_t_woe_report = pd.concat([_df_t_woe_report, _df], ignore_index=True) \
        .sort_values(by=["data_type", "feature_name", "gb_idx"], ascending=[True, True, True]) \
        .reset_index(drop=True)
    
    ####################################################################
    # _df_t_scorecard
    _df_t_scorecard = pd.merge(
        left=_df_t_model_coef,
        right=_df_t_woe_report,
        how="left", left_on=["feature_name"], right_on=["feature_name"],
    )
    
    ####################################################################
    _df_t_scorecard.loc[_df_t_scorecard["feature_name"]=="Intercept", "gb_idx"] = 0
    _df_t_scorecard.loc[_df_t_scorecard["feature_name"]=="Intercept", "label"] = "_BASE_SCORE_"
    _df_t_scorecard.loc[_df_t_scorecard["feature_name"]=="Intercept", "WOE"] = np.NaN
    _df_t_scorecard.loc[_df_t_scorecard["feature_name"]=="Intercept", ["data_type", "_missing_tag", "_eq_side"]] = ""
    
    ####################################################################
    _df_t_scorecard["WOE_Coef"] = _df_t_scorecard["WOE"]*_df_t_scorecard["Coef"]
    _df_t_scorecard["Score_0"] = _df_t_scorecard.apply(
        lambda s0: (
            _A-_B*s0["Coef"]
            if s0["feature_name"]=="Intercept" else
            (-1)*_B*s0["WOE_Coef"]
        ),
        axis=1,
    )
    _base_score = (_df_t_scorecard.query("feature_name=='Intercept'")["Score_0"].tolist()+[0])[0]
    _df_t_scorecard["Score"] = _df_t_scorecard.apply(
        lambda s0: (
            0
            if s0["feature_name"]=="Intercept" else
            s0["Score_0"]+_base_score/len([s0 for s0 in feature_names if s0!="Intercept"])
        ),
        axis=1,
    )
    
    ####################################################################
    _df_t_scorecard["gb_idx"] = _df_t_scorecard["gb_idx"].astype(int)
    _df_t_scorecard["label"] = _df_t_scorecard.apply(
        lambda s0: ("_MISSING_" if s0["_missing_tag"]=="Y" else s0["label"]),
        axis=1,
    )
    _df_t_scorecard["_label_left"] = _df_t_scorecard.apply(
        lambda s0: (
            re.sub("[\(\[]", "", s0["label"].split(",")[0])
            if re.search("^[\(\[][^\)\]]{1,}[\)\]]$", s0["label"])!=None else
            s0["label"]
        ),
        axis=1,
    )
    _df_t_scorecard["_label_right"] = _df_t_scorecard.apply(
        lambda s0: (
            re.sub("[\)\] ]", "", s0["label"].split(",")[1])
            if re.search("^[\(\[][^\)\]]{1,}[\)\]]$", s0["label"])!=None else
            s0["label"]
        ),
        axis=1,
    )
    _df_t_scorecard["_p_Odds"] = p_Odds
    _df_t_scorecard["_p_SCORE"] = p_SCORE
    _df_t_scorecard["_p_PDO"] = p_PDO
    
    ####################################################################
    _df_t_scorecard = _df_t_scorecard[[
        "var_idx", "feature_name", "data_type", "gb_idx", "label", "Score_0", "Score",
        "_missing_tag", "_eq_side", "_label_left", "_label_right",
        "Coef", "WOE", "WOE_Coef",
        "_WOE", "_wls_adj_WOE", "_label", "_p_Odds", "_p_SCORE", "_p_PDO",
        ]] \
            .reset_index(drop=True)
    
    return _df_t_scorecard






### 构建评分卡参数表



# print("-"*40)
# for s0 in df_model_params["column_name"].values[:]:
#     print('"{}",'.format(s0))

# print()
# print("-"*40)
# for s0 in df_model_params["Coef."].values[:]:
#     print('{:.6f},'.format(s0))



####################################################################
# 构建评分卡参数表

####################################################################
scorecard_feature_names = list(df_model_params['column_name'])
scorecard_coef = list(df_model_params['Coef.'])

df_woe_report = df_t_features_coarse_woe_report.reset_index(drop=True)

####################################################################
# p_Odds = 1/50
p_Odds = 1/20
p_SCORE = 700
# p_PDO = 25
p_PDO = 60

####################################################################
# missing_fill_type="0_zero"
missing_fill_type="1_woe"
num_woe_type="0_ori"
# num_woe_type="1_wls_adj"

####################################################################
df_t_scorecard = _func_lr2scorecard(
    feature_names=scorecard_feature_names,
    coef=scorecard_coef,
    df_woe_report=df_woe_report,
    p_Odds=p_Odds,
    p_SCORE=p_SCORE,
    p_PDO=p_PDO,
    missing_fill_type=missing_fill_type,
    num_woe_type=num_woe_type,
)

####################################################################
df_t_scorecard_gb_var = df_t_scorecard.groupby(by=["var_idx", "feature_name", "data_type", "_eq_side", "Coef"]).apply(
    lambda s0: s0[["gb_idx", "label", "Score_0", "Score", "_missing_tag", "_label_left", "_label_right", "WOE", "WOE_Coef", "_WOE", "_wls_adj_WOE", "_label"]] \
        .apply(lambda s0: s0.to_dict(), axis=1).tolist(),
).rename("_gb_info").reset_index()






## （评分卡参数表导出）

####################################################################################
# to_clipboard
df_t_scorecard.to_clipboard(index=False)

####################################################################################
# to_excel
with open(file="{}/Proc09_Scorecard_Trans/df_t_scorecard.xlsx".format(result_path), mode="wb") as fw:
    df_t_scorecard \
        .to_excel(
            fw,
            index=False,
            sheet_name="data",
        )

###########################################################################
# json
with open(file="{}/Proc09_Scorecard_Trans/df_t_scorecard.json".format(result_path), mode="w", encoding="utf-8") as fw:
    json.dump(
        obj=df_t_scorecard.apply(lambda s0: s0.to_dict(), axis=1).tolist(),
        fp=fw,
        indent=4,
        ensure_ascii=False,
    )

###################################################################################
# pkl文件

####################################################################################
with open(file="{}/Proc09_Scorecard_Trans/df_t_scorecard.pkl".format(result_path), mode="wb") as fw:
    pickle.dump(obj=df_t_scorecard, file=fw)

####################################################################################
with open(file="{}/Proc09_Scorecard_Trans/df_t_scorecard.pkl".format(result_path), mode="rb") as fr:
    df_t_scorecard = pickle.load(file=fr)
    df_t_scorecard_gb_var = df_t_scorecard.groupby(by=["var_idx", "feature_name", "data_type", "_eq_side", "Coef"]).apply(
        lambda s0: s0[["gb_idx", "label", "Score_0", "Score", "_missing_tag", "_label_left", "_label_right", "WOE", "WOE_Coef", "_WOE", "_wls_adj_WOE", "_label"]] \
            .apply(lambda s0: s0.to_dict(), axis=1).tolist(),
    ).rename("_gb_info").reset_index()






### （评分卡参数表导入）

# ####################################################################################
# # to_json
# with open(file="{}/Proc09_Scorecard_Trans/df_t_scorecard.json".format(result_path), mode="r", encoding="utf-8") as fr:
#     df_t_scorecard = pd.DataFrame(json.load(fp=fr))
#     df_t_scorecard_gb_var = df_t_scorecard.groupby(by=["var_idx", "feature_name", "data_type", "_eq_side", "Coef"]).apply(
#         lambda s0: s0[["gb_idx", "label", "Score_0", "Score", "_missing_tag", "_label_left", "_label_right", "WOE", "WOE_Coef", "_WOE", "_wls_adj_WOE", "_label"]] \
#             .apply(lambda s0: s0.to_dict(), axis=1).tolist(),
#     ).rename("_gb_info").reset_index()


df_t_scorecard.head(10)

df_t_scorecard_gb_var.head()









































#==============================================================================
# File: code13_Proc10_Model_Apply_Scorecard.py
#==============================================================================

# -*- coding: utf-8 -*-
# 
#----------------------------------------------------------
# script name: Proc10_Model_Apply_Scorecard
#----------------------------------------------------------
# creator: luzhidong94
# create date: 2022-08-30
# update date: 2022-08-30
# version: 1.0
#----------------------------------------------------------
# 
# 
#----------------------------------------------------------


## Proc10_Model_Apply_Scorecard
## 评分卡模型跑批应用


####################################################################
# Proc10_Model_Apply_Scorecard
# 评分卡模型跑批应用


####################################################################
# 入模指标清单 cols_model
cols_model = [
    s0 for s0 in df_t_scorecard["feature_name"].drop_duplicates().tolist()
    if s0!="Intercept"
]

####################################################################
# 跑批binning、woe、score结果
def _func_model_apply_var_mapping(
        in_data,
        data_type,
        _eq_side,
        _gb_info,
        output_type="binning",
        ):
    
    ####################################################################
    def _func_value_mapping(_value, _data_type, _mapping, _eq_side):
        rt = np.NaN
        if _data_type=="Categorical":
            
            
            if pd.isna(_value):
                rt = _mapping["_MISSING_"]
            else:
                for _k, _v in [(_k, _v) for _k, _v in _mapping.items() if _k!="_MISSING_"][:]:
                    if _value in _k.split("/"):
                        rt = _v
                        break
        elif _data_type=="Numerical":
            if pd.isna(_value):
                rt = _mapping["_MISSING_"]
            else:
                for _k, _v in [(_k, _v) for _k, _v in _mapping.items() if _k!="_MISSING_"][:]:
                    if _eq_side=="1_left":
                        if _value<float(_k):
                            rt = _v
                            break
                    elif _eq_side=="2_right":
                        if _value<=float(_k):
                            rt = _v
                            break
                    else:
                        pass
        else:
            pass
        return rt
    
    ####################################################################
    _df_gb_info = pd.DataFrame(_gb_info).set_index(keys=["_label_right"])
    
    if output_type=="binning":
        _mapping = _df_gb_info.apply(
            lambda s0: "{:0{}d}_{}".format(s0["gb_idx"], int(np.log10(_df_gb_info.shape[0]-1))+1, s0["_label"]),
            axis=1,
        )
    elif output_type=="woe":
        _mapping = _df_gb_info["WOE"]
    elif output_type=="score":
        _mapping = _df_gb_info["Score"]
        # _mapping = _df_gb_info["Score_0"]
    else:
        pass
    
    ####################################################################
    _out = in_data.apply(
        lambda s0: _func_value_mapping(_value=s0, _data_type=data_type, _mapping=_mapping, _eq_side=_eq_side),
    )
    return _out
    

####################################################################
# df_wb_model_result
df_wb_model_result = df_wb[cols_base+cols_model].reset_index(drop=True)
for _feature_name, _data_type, _eq_side, _gb_info in \
    df_t_scorecard_gb_var[["feature_name", "data_type", "_eq_side", "_gb_info"]].values[:]:
    
    ####################################################################
    if _feature_name=="Intercept":
        pass
    else:
        df_wb_model_result["BIN_{}".format(_feature_name)] = \
            _func_model_apply_var_mapping(
                in_data=df_wb_model_result[_feature_name],
                data_type=_data_type,
                _eq_side=_eq_side,
                _gb_info=_gb_info,
                output_type="binning",
                # output_type="woe",
                # output_type="score",
            )
        df_wb_model_result["WOE_{}".format(_feature_name)] = \
            _func_model_apply_var_mapping(
                in_data=df_wb_model_result[_feature_name],
                data_type=_data_type,
                _eq_side=_eq_side,
                _gb_info=_gb_info,
                # output_type="binning",
                output_type="woe",
                # output_type="score",
            )
        df_wb_model_result["SCORE_{}".format(_feature_name)] = \
            _func_model_apply_var_mapping(
                in_data=df_wb_model_result[_feature_name],
                data_type=_data_type,
                _eq_side=_eq_side,
                _gb_info=_gb_info,
                # output_type="binning",
                # output_type="woe",
                output_type="score",
            )
    
####################################################################
# 跑批模型结果：

####################################################################
_coef_mapping = df_t_scorecard.set_index(keys=["feature_name"])["Coef"].drop_duplicates().to_dict()
_coef_list = [_coef_mapping.get(s0) for s0 in cols_model]
_coef_Intercept = (df_t_scorecard.query("feature_name=='Intercept'")["Coef"].tolist()+[0])[0]
df_wb_model_result["_wx_sum"] = df_wb_model_result[["WOE_{}".format(s0) for s0 in cols_model]].apply(
    lambda s0: sum([_coef_Intercept]+[_coef*_woe for _coef, _woe in zip(_coef_list, s0.tolist())]),
    axis=1,
)
# _base_score = (df_t_scorecard.query("feature_name=='Intercept'")["Score_0"].tolist()+[0])[0]

####################################################################
# 概率值（model_prob）
df_wb_model_result["model_prob"] = df_wb_model_result["_wx_sum"].apply(lambda s0: 1/(1+np.exp((-1)*s0)))

####################################################################
# 评分值（model_score）
df_wb_model_result["model_score"] = df_wb_model_result[["SCORE_{}".format(s0) for s0 in cols_model]].apply(
    lambda s0: sum(s0.tolist()),
    axis=1,
)
# df_wb_model_result["model_score"] = df_wb_model_result["_wx_sum"].apply(lambda s0: _A+_B*(-1)*s0)
# df_wb_model_result["model_score"] = df_wb_model_result[["SCORE_{}".format(s0) for s0 in cols_model]].apply(
#     lambda s0: _base_score+sum(s0.tolist()),
#     axis=1,
# )

####################################################################
df_wb_model_result = df_wb_model_result[
    cols_base+["model_prob", "model_score", "_wx_sum"]+
    [s0 for s0 in df_wb_model_result.columns if s0 not in cols_base+["model_prob", "model_score"]]
].reset_index(drop=True)

####################################################################
print(df_wb_model_result.shape)


df_wb_model_result.head()






## (概率值分布检查)

##########################################################################

# 间隔
_bin = [-inf, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, inf]
# _bin = [-inf, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, inf]

_crosstab = func_woe_report_v1(
    in_var=func_binning_continuous_v1(
        in_data=df_wb_model_result["model_prob"],
        bins=_bin, right_border=True, include_lowest=True,
    ),
    in_target=df_wb_model_result["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
    with_lift_ks=True,
    with_wls_adj_woe=True,
)

func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# _crosstab






## (评分分布检查)

##########################################################################

# # 等距25分间隔
# # _bin = [-inf, 500, 525, 550, 575, 600, 625, 650, 675, 700, 725, 750, 775, 800, inf]
# _bin = [-inf, 550, 575, 600, 625, 650, 675, 700, 725, 750, 775, 800, inf]

# # 等距10分间隔
# _bin = [-inf, 500, 510, 520, 530, 540, 550, 560, 570, 580, 590,
#               600, 610, 620, 630, 640, 650, 660, 670, 680, 690,
#               700, 710, 720, 730, 740, 750, 760, 770, 780, 790,
#         inf]

_bin = [-inf]+list(np.arange(500, 900+20, 20))+[inf]
# _bin = [-inf]+list(np.arange(500, 900+10, 10))+[inf]
# _bin = [-inf]+list(np.arange(500, 900+5, 5))+[inf]


_crosstab = func_woe_report_v1(
    in_var=func_binning_continuous_v1(
        in_data=df_wb_model_result["model_score"],
        # in_data=df_wb_model_result.query("data_role in ['04_oot']")["model_score"],
        bins=_bin, right_border=True, include_lowest=True,
    ),
    in_target=df_wb_model_result["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    # in_target=df_wb_model_result.query("data_role in ['04_oot']")["flag"].apply(lambda s0: ("0_good" if s0==0 else "1_bad")),
    with_total=True, good_label_val="0_good", bad_label_val="1_bad", floating_point=0.01,
    with_lift_ks=True,
    with_wls_adj_woe=True,
)

func_plot_woe(_crosstab, plot_badrate=True, with_nan_info=False, with_wls_adj_woe=True)
# _crosstab.to_clipboard()
# _crosstab
print(_crosstab)





































#==============================================================================
# File: config_loader.py
#==============================================================================

import json
import os
import logging
import pandas as pd
from datetime import datetime

logger = logging.getLogger(__name__)

# 动态获取项目根目录（用户主目录下的ml_monitor）
HOME = os.path.expanduser("~")
ROOT_DIR = os.path.join(HOME, "ml_monitor")

email_config = {
    'smtp_host': 'smtphz.qiye.163.com',
    'smtp_port': 465,
    'user': 'liaoxilin@hulianshuzhi.com',
    'password': 'Life2010.',
    'receivers': ["xiaohan@hulianshuzhi.com","wumin1@hulianshuzhi.com","wangmiao@hulianshuzhi.com","tanxing@hulianshuzhi.com","jileilei@hulianshuzhi.com","liyi@hulianshuzhi.com","heyijia@hulianshuzhi.com","youpengyu@hulianshuzhi.com","wangxincan@hulianshuzhi.com","fuzhehao@hulianshuzhi.com","liaoxilin@hulianshuzhi.com"]
}

def get_project_root() -> str:
    """
    获取项目根目录（scripts/ 的上级目录）
    若 model_monitor.py 在 scripts/ 下，根目录为 scripts/../
    """
    # 当前脚本（model_monitor.py）的绝对路径
    current_script_path = os.path.abspath(__file__)
    # scripts/ 目录路径（当前脚本的父目录）
    scripts_dir = os.path.dirname(current_script_path)
    # 项目根目录（scripts/ 的父目录）
    project_root = os.path.dirname(scripts_dir)
    return project_root


#==============================================================================
# File: data_utils.py
#==============================================================================

import toad
import numpy as np
import pandas as pd
import scorecardpy as sc
import datetime as dt
import pytz
from sklearn.preprocessing import OrdinalEncoder
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
from dateutil.parser import parse

numeric_cols = ['duration.in.month',
                'credit.amount',
                'age.in.years',
                'present.residence.since',
                'number.of.existing.credits.at.this.bank',
                'installment.rate.in.percentage.of.disposable.income',
                'number.of.people.being.liable.to.provide.maintenance.for']

category_cols = ['status.of.existing.checking.account', 'credit.history',
                 'savings.account.and.bonds', 'present.employment.since',
                 'personal.status.and.sex', 'other.debtors.or.guarantors',
                 'property', 'other.installment.plans', 'housing', 'job',
                 'telephone', 'foreign.worker', 'purpose']

x_cols = numeric_cols + category_cols

label = 'creditability'


def get_data():
    """
    导入原始数据集
    """
    german_credit_data = sc.germancredit()
    german_credit_data[label] = np.where(
        german_credit_data[label] == 'bad', 1, 0)
    # 设置随机数种子, 确保结果可复现
    np.random.seed(0)
    month_list = ['2020-01', '2020-02', '2020-03', '2020-04', '2020-05']
    # 随机分配月份
    german_credit_data['month'] = np.random.choice(
        month_list, len(german_credit_data))
    return german_credit_data


def get_all_x_y(transform_method='minmax'):
    """
    加载数据
    :param transform_method: 数据标准化方式
    """
    german_credit_data = sc.germancredit()
    # 类别型变量转化成数值型索引变量
    encoder = OrdinalEncoder()
    category_result = encoder.fit_transform(german_credit_data[category_cols])
    category_result = pd.DataFrame(data=category_result, columns=category_cols)
    numeric_result = german_credit_data[numeric_cols + [label]].copy()
    # 将标签creditability映射为数值
    numeric_result[label] = np.where(numeric_result[label] == 'bad', 1, 0)
    all_x_y = pd.merge(category_result, numeric_result, left_index=True, right_index=True)
    x_cols = [f for f in all_x_y.columns if f != label]
    # 数据标准化
    if transform_method == 'minmax':
        encoder = MinMaxScaler()
        all_x_y[x_cols] = encoder.fit_transform(all_x_y[x_cols])
    elif transform_method == 'standard':
        encoder = StandardScaler()
        all_x_y[x_cols] = encoder.fit_transform(all_x_y[x_cols])
    elif transform_method == 'origin':
        pass
    return all_x_y


def get_data_after_fs(empty=0.5, iv=0.02, corr=0.7):
    """
    加载特征选择后的数据
    :param empty: 缺失率阈值
    :param iv: iv阈值
    :param corr: 相关性阈值
    """
    all_x_y = get_all_x_y()
    selected_data, drop_lst = toad.selection.select(
        all_x_y, target=label, empty=0.5,
        iv=0.02, corr=0.7, return_drop=True)
    return selected_data


def get_x_y_split(test_rate=0.2, transform_method='minmax'):
    """
    划分训练集和测试集
    :param test_rate: 测试集样本占比
    :param transform_method: 数据标准化方式
    """
    german_credit_data = get_all_x_y(transform_method)
    y = german_credit_data.pop(label)
    x = german_credit_data
    x_train, x_valid, y_train, y_valid = train_test_split(
        x, y, test_size=test_rate, random_state=88)
    return x_train, x_valid, y_train, y_valid


def stamp_to_date(time_stamp, timezone=None):
    """
    时间戳转日期函数
    :param time_stamp:int，时间戳
    :param timezone:string，时区
    :return:datetime
    """
    try:
        if timezone is None:
            stamp_str = str(time_stamp)
            if len(stamp_str) >= 10:
                stamp_str = stamp_str[:10]
            else:
                stamp_str = stamp_str
            time_stamp = int(stamp_str)
            date = dt.datetime.fromtimestamp(time_stamp)
            return date
        else:
            stamp_str = str(time_stamp)
            if len(stamp_str) >= 10:
                stamp_str = stamp_str[:10]
            else:
                stamp_str = stamp_str
            time_stamp = int(stamp_str)
            tz = pytz.timezone(timezone)
            date = dt.datetime.fromtimestamp(time_stamp, tz).strftime('%Y-%m-%d %H:%M:%S')
            date = parse(date)
            return date
    except:
        return parse('2100-01-01')


def date_to_week(date):
    """
    日期转换为星期
    :param date:datetime，string
    :return:int
    """
    try:
        if isinstance(date, str):
            date = parse(date)
        if_weekend = date.weekday()
        return if_weekend
    except:
        return np.nan



#==============================================================================
# File: detector.py
#==============================================================================

#!/usr/bin/python

"""Command line tools for detecting csv data

Team: ESC

Examples:

    python detector.py -i xxx.csv -o report.csv

"""

import pandas as pd


def get_top_values(series, top=5, reverse=False):
    """Get top/bottom n values

    Args:
        series (Series): data series
        top (number): number of top/bottom n values
        reverse (bool): it will return bottom n values if True is given

    Returns:
        Series: Series of top/bottom n values and percentage. ['value:percent', None]
    """
    itype = 'top'
    counts = series.value_counts()
    counts = list(zip(counts.index, counts, counts.divide(series.size)))

    if reverse:
        counts.reverse()
        itype = 'bottom'

    template = "{0[0]}:{0[2]:.2%}"
    indexs = [itype + str(i + 1) for i in range(top)]
    values = [template.format(counts[i]) if i < len(counts) else None for i in range(top)]

    return pd.Series(values, index=indexs)


def get_describe(series, percentiles=[.25, .5, .75]):
    """Get describe of series

    Args:
        series (Series): data series
        percentiles: the percentiles to include in the output

    Returns:
        Series: the describe of data include mean, std, min, max and percentiles
    """
    d = series.describe(percentiles)
    return d.drop('count')


def count_blank(series, blanks=[None]):
    """Count number and percentage of blank values in series

    Args:
        series (Series): data series
        blanks (list): list of blank values

    Returns:
        number: number of blanks
        str: the percentage of blank values
    """
    # n = 0
    # counts = series.value_counts()
    # for blank in blanks:
    #     if blank in counts.keys():
    #         n += counts[blank]

    n = series.isnull().sum()

    return (n, "{0:.2%}".format(n / series.size), n / series.size)


def is_numeric(series):
    """Check if the series's type is numeric

    Args:
        series (Series): data series

    Returns:
        bool
    """
    return series.dtype.kind in 'ifc'


def detect(dataframe, dic_name=None):
    """ Detect data

    Args:
        dataframe (DataFrame): data that will be detected

    Returns:
        DataFrame: report of detecting
    """

    rows = []
    for name, series in dataframe.items():
        numeric_index = ['mean', 'std', 'min', '1%', '10%', '50%', '75%', '90%', '99%', 'max']
        discrete_index = ['top1', 'top2', 'top3', 'top4', 'top5', 'bottom5', 'bottom4', 'bottom3', 'bottom2', 'bottom1']

        details_index = [numeric_index[i] + '_or_' + discrete_index[i] for i in range(len(numeric_index))]
        details = []

        if is_numeric(series):
            desc = get_describe(
                series,
                percentiles=[.01, .1, .5, .75, .9, .99]
            )
            details = desc.tolist()
        else:
            top5 = get_top_values(series)
            bottom5 = get_top_values(series, reverse=True)
            details = top5.tolist() + bottom5[::-1].tolist()

        nblank, pblank, pblank_ = count_blank(series)

        ###add 2020/01/02 RyanZheng
        value_max_percent = get_max_percent(series)
        ###add 2020/01/02 RyanZheng

        row = pd.Series(
            index=['type', 'size', 'missing', 'missing_q', 'unique', 'value_max_percent'] + details_index,
            data=[series.dtype, series.size, pblank, pblank_, series.nunique(), value_max_percent] + details
        )

        row.name = name
        rows.append(row)

    # return pd.DataFrame(rows)

    ### add 2020/01/02 RyanZheng
    eda_df = pd.DataFrame(rows)
    if dic_name is not None and isinstance(dic_name, dict):
        # 增加一列中文名称列
        eda_df.insert(0, 'cn', eda_df.index.map(dic_name))
        # eda_df['cn'] = eda_df.index.map(dic_name)
    eda_df.index.name = 'var_name'
    eda_df = eda_df.reset_index()
    eda_df['type'] = eda_df['type'].astype(str)
    return eda_df
    ### add 2020/01/02 RyanZheng


###add 2020/01/02 RyanZheng
def get_max_percent(series):
    """
    获取变量中同一个值出现次数最多的该值的占比
    Args:
        series:

    Returns:

    """
    return max(series.value_counts(dropna=False) / len(series))
###add 2020/01/02 RyanZheng



#==============================================================================
# File: EDA分析.py
#==============================================================================


# coding: utf-8

# In[ ]:


# EDA分析

# 类别型变量的分布
def plot_cate_var(df,col_list,hspace=0.4,wspace=0.4,plt_size=None,plt_num=None,x=None,y=None):
    """
    df:数据集
    col_list:变量list集合
    hspace :子图之间的间隔(y轴方向)
    wspace :子图之间的间隔(x轴方向)
    plt_size :图纸的尺寸
    plt_num :子图的数量
    x :子图矩阵中一行子图的数量
    y :子图矩阵中一列子图的数量
    
    return :变量的分布图（柱状图形式）
    """
    plt.figure(figsize=plt_size)
    plt.subplots_adjust(hspace=hspace,wspace=wspace)
    plt.rcParams['font.sans-serif']=['Microsoft YaHei']
    plt.rcParams['axes.unicode_minus'] = False
    for i,col in zip(range(1,plt_num+1,1),col_list):
        plt.subplot(x,y,i)
        plt.title(col)
        sns.countplot(data=df,y=col)
        plt.ylabel('')
    return plt.show()


# 数值型变量的分布
def plot_num_col(df,col_list,hspace=0.4,wspace=0.4,plt_type=None,plt_size=None,plt_num=None,x=None,y=None):
    """
    df:数据集
    col_list:变量list集合
    hspace :子图之间的间隔(y轴方向)
    wspace :子图之间的间隔(x轴方向)
    plt_type: 选择直方图/箱线图
    plt_size :图纸的尺寸
    plt_num :子图的数量
    x :子图矩阵中一行子图的数量
    y :子图矩阵中一列子图的数量
    
    return :变量的分布图（箱线图/直方图）
    """
    plt.figure(figsize=plt_size)
    plt.subplots_adjust(hspace=hspace,wspace=wspace)
    if plt_type=='hist':
        for i,col in zip(range(1,plt_num+1,1),col_list):
            plt.subplot(x,y,i)
            plt.title(col)
            sns.distplot(df[col].dropna())
            plt.xlabel('')
    if plt_type=='box':
        for i,col in zip(range(1,plt_num+1,1),col_list):
            plt.subplot(x,y,i)
            plt.title(col)
            sns.boxplot(data=df,x=col)
            plt.xlabel('')
    return plt.show()


# 类别型变量的违约率分析
def plot_default_cate(df,col_list,target,hspace=0.4,wspace=0.4,plt_size=None,plt_num=None,x=None,y=None):
    """
    df:数据集
    col_list:变量list集合
    target ：目标变量的字段名
    hspace :子图之间的间隔(y轴方向)
    wspace :子图之间的间隔(x轴方向)
    plt_size :图纸的尺寸
    plt_num :子图的数量
    x :子图矩阵中一行子图的数量
    y :子图矩阵中一列子图的数量
    
    return :违约率分布图（柱状图形式）
    """
    all_bad = df[target].sum()
    total = df[target].count()
    all_default_rate = all_bad*1.0/total
    
    plt.figure(figsize=plt_size)
    plt.subplots_adjust(hspace=hspace,wspace=wspace)
    plt.rcParams['font.sans-serif']=['Microsoft YaHei']
    plt.rcParams['axes.unicode_minus'] = False
    for i,col in zip(range(1,plt_num+1,1),col_list):
        d1 = df.groupby(col)
        d2 = pd.DataFrame()
        d2['total'] = d1[target].count()
        d2['bad'] = d1[target].sum()
        d2['default_rate'] = d2['bad']/d2['total']
        d2 = d2.reset_index()
        plt.subplot(x,y,i)
        plt.title(col)
        plt.axvline(x=all_default_rate)
        sns.barplot(data=d2,y=col,x='default_rate')
        plt.ylabel('')
    return plt.show()


# 数值型变量的违约率分析
def plot_default_num(df,col_list,target,hspace=0.4,wspace=0.4,q=None,plt_size=None,plt_num=None,x=None,y=None):
    """
    df:数据集
    col_list:变量list集合
    target ：目标变量的字段名
    hspace :子图之间的间隔(y轴方向)
    wspace :子图之间的间隔(x轴方向)
    q :等深分箱的箱体个数
    plt_size :图纸的尺寸
    plt_num :子图的数量
    x :子图矩阵中一行子图的数量
    y :子图矩阵中一列子图的数量
    
    return :违约率分布图（折线图形式）
    """
    all_bad = df[target].sum()
    total = df[target].count()
    all_default_rate = all_bad*1.0/total 
    
    plt.figure(figsize=plt_size)
    plt.subplots_adjust(hspace=hspace,wspace=wspace)
    for i,col in zip(range(1,plt_num+1,1),col_list):
        bucket = pd.qcut(df[col],q=q,duplicates='drop')
        d1 = df.groupby(bucket)
        d2 = pd.DataFrame()
        d2['total'] = d1[target].count()
        d2['bad'] = d1[target].sum()
        d2['default_rate'] = d2['bad']/d2['total']
        d2 = d2.reset_index()
        plt.subplot(x,y,i)
        plt.title(col)
        plt.axhline(y=all_default_rate)
        sns.pointplot(data=d2,x=col,y='default_rate',color='hotpink')
        plt.xticks(rotation=60)
        plt.xlabel('')
    return plt.show()




#==============================================================================
# File: execute_custom_query.py
#==============================================================================

import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
from odps_utils import execute_odps_sql
from config_loader import  ROOT_DIR

logger = logging.getLogger(__name__)


# ================== 新增：读取自定义SQL功能 ==================

def read_custom_sql(file_path: str, date_str: str = None) -> str:
    """
    读取外部SQL文件，并可选地替换{date}等占位符
    支持动态插入日期（格式自动转换为ODPS分区格式）
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            sql = f.read().strip()
        
        # 如果提供了日期，替换 {date} 占位符为实际分区格式（YYYYMMDD）
        if date_str:
            dt = date_str.replace('-', '')
            sql = sql.replace('{date}', dt)
            logger.info(f"📅 已将 {{date}} 替换为 {dt}")
        
        if not sql:
            raise ValueError("SQL文件为空")
        
        logger.info(f"✅ 成功读取自定义SQL: {file_path}")
        return sql
    except Exception as e:
        logger.error(f"❌ 读取SQL文件失败: {e}")
        raise

def execute_custom_query(file_path: str, date_str: str = None) -> pd.DataFrame:
    """
    执行自定义SQL查询（主要面向临时分析）
    """
    try:
        sql = read_custom_sql(file_path, date_str)
        logger.info(f"🔍 正在执行自定义SQL:\n{sql[:200]}...")  # 打印前200字符预览
        result_df = execute_odps_sql(sql)
        logger.info(f"🎉 自定义查询执行成功，返回 {len(result_df)} 条记录")
        return result_df
    except Exception as e:
        logger.error(f"❌ 执行自定义SQL失败: {e}")
        raise




#==============================================================================
# File: feature_binning.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: feature_binning.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-20
'''

import warnings

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.tree import DecisionTreeClassifier, _tree

from .statistics import calc_iv_and_inflection_point, bin_badrate
from .utils import fillna, split_empty, split_points_to_bin, t_sum_np, t_min_np, \
    t_cols_sum_axis_0_np, t_cols_sum_axis_1_np, support_dataframe

warnings.filterwarnings('ignore')

from .logger_utils import Logger

log = Logger(level="info", name=__name__).logger

DEFAULT_BINS = 10
IS_EMPTY_BIN_DEFAULT_BINS = 9


def equal_freq_bin(feature, target=None, min_sample_rate=0.05,
                   n_bins=None,
                   q_cut_list=None, is_need_monotonic=True, is_empty_bin=True):
    """
    等频分箱

    Args:
        feature (array) : 某个x特征
        target (array) : 目标值y变量
        min_sample_rate (number) : 每个箱子的最小占比
        n_bins (int) : 需要分成几箱
        q_cut_list (array、list) : 百分比分割点列表
        is_need_monotonic (bool) : 是否强制单调
        is_empty_bin (bool) : 是否将空箱单独归为一个箱子

    Returns:
        array (array): 分割点
    """

    empty_mask = np.array([])
    if is_empty_bin:
        if target is None:
            feature, empty_mask = split_empty(feature, target)
        else:
            feature, target, empty_mask = split_empty(feature, target)
    else:
        feature = fillna(feature)

    if n_bins is None and q_cut_list is None and is_empty_bin and empty_mask.any():
        n_bins = IS_EMPTY_BIN_DEFAULT_BINS
    elif n_bins is None and q_cut_list is None:
        n_bins = DEFAULT_BINS

    if q_cut_list is None:
        q_cut_list = np.arange(0, 1, 1 / n_bins)

    is_monotonic = False
    while not is_monotonic:
        splits_tmp = np.quantile(feature, q_cut_list)
        splits = np.unique(splits_tmp)[1:]

        x_bins = split_points_to_bin(feature, splits)  # 返回0、1、2、3这样的数值分箱
        if target is None:
            _ = bin_badrate(x_bins)
            is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足
        else:
            bin_badrate_li, _ = bin_badrate(x_bins, target)
            is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足
            if is_need_monotonic:
                if is_monotonic:  # 满足最小分箱占比后，判断是否满足单调性
                    is_monotonic = pd.Series(bin_badrate_li).is_monotonic_decreasing or pd.Series(
                        bin_badrate_li).is_monotonic_increasing

        if n_bins <= 2:
            break

        n_bins = q_cut_list.size - 1  # 这种是从初始的n_bins(10)，一点一点的减
        q_cut_list = np.arange(0, 1, 1 / n_bins)

    if is_empty_bin and empty_mask.any():
        splits = np.append(splits, np.nan)

    return splits


def dt_bin(feature, target, min_sample_rate=0.05, n_bins=None,
           is_need_monotonic=True, is_empty_bin=True, **kwargs):
    """
    决策树分箱

    Args:
        feature (array) : 某个x特征
        target (array) : 目标值y变量
        min_sample_rate (number) : 每个箱子的最小占比
        n_bins (int) : 需要分成几箱
        is_need_monotonic (bool) : 是否强制单调
        is_empty_bin (bool) : 是否将空箱单独归为一个箱子
        **kwargs : 决策树的其它参数

    Returns:
        array (array): 分割点
    """

    empty_mask = np.array([])
    if is_empty_bin:
        feature, target, empty_mask = split_empty(feature, target)
    else:
        feature = fillna(feature)

    if n_bins is None and is_empty_bin and empty_mask.any():
        n_bins = IS_EMPTY_BIN_DEFAULT_BINS
    elif n_bins is None:
        n_bins = DEFAULT_BINS

    is_monotonic = False
    while not is_monotonic:

        # 决策树分箱逻辑
        '''
        1、初始n_b：10（point），splits：9（point）
        2、9+3=12（point） 》11（bin）
        3、11（bin）- 1 ---》10（bin）
        4、老版本有，新版本跳过
        5、10 - 1 = 9（point）
        '''
        tree = DecisionTreeClassifier(
            min_samples_leaf=min_sample_rate,
            max_leaf_nodes=n_bins,
            # 优先满足min_samples_leaf参数。在满足min_samples_leaf参数参数后，再考虑max_leaf_nodes。
            # 比如情况1：min_samples_leaf设置成0.05，max_leaf_nodes设置成20。满足0.05后，最大max_leaf_nodes只有10，那也就这样了
            # 比如情况2：min_samples_leaf设置成0.05，max_leaf_nodes设置成6。满足0.05后，最大max_leaf_nodes有10，那再考虑max_leaf_nodes，继续分到满足max_leaf_nodes=6停止
            # ps:min_samples_leaf=1表示没有限制
            **kwargs,
        )
        tree.fit(feature.reshape((-1, 1)), target)
        thresholds = tree.tree_.threshold
        thresholds = thresholds[thresholds != _tree.TREE_UNDEFINED]
        splits = np.sort(thresholds)
        # 决策树分箱逻辑

        is_monotonic = True
        if is_need_monotonic:
            x_bins = split_points_to_bin(feature, splits)  # 返回0、1、2、3这样的数值分箱

            bin_badrate_li, _ = bin_badrate(x_bins, target)

            # 不需要判断，tree里面已经判断了
            # is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足

            is_monotonic = pd.Series(bin_badrate_li).is_monotonic_decreasing or pd.Series(
                bin_badrate_li).is_monotonic_increasing

        n_bins = len(splits) + 1  # 初始n_bins为10，对应的splits为9。需要+1，在后面的n_bins -= 1后，n_bins才会是9

        if n_bins <= 2:
            break
        n_bins -= 1

    if is_empty_bin and empty_mask.any():
        splits = np.append(splits, np.nan)
    return splits


def kmeans_bin(feature, target=None, min_sample_rate=0.05,
               n_bins=None,
               random_state=1, is_need_monotonic=True, is_empty_bin=True):
    """
    kmeans聚类分箱

    Args:
        feature (array) : 某个x特征
        target (array) :  目标值y变量
        min_sample_rate (number) : 每个箱子的最小占比
        n_bins (int): 需要分成几箱
        random_state (int): kmeans模型中的随机种子
        is_need_monotonic (bool) : 是否强制单调
        is_empty_bin (bool) : 是否将空箱单独归为一个箱子

    Returns:
        array (array): 分割点
    """

    empty_mask = np.array([])
    if is_empty_bin:
        if target is None:
            feature, empty_mask = split_empty(feature, target)
        else:
            feature, target, empty_mask = split_empty(feature, target)
    else:
        feature = fillna(feature)

    if n_bins is None and is_empty_bin and empty_mask.any():
        n_bins = IS_EMPTY_BIN_DEFAULT_BINS
    elif n_bins is None:
        n_bins = DEFAULT_BINS

    is_monotonic = False
    while not is_monotonic:

        # kmeans 逻辑
        kmeans = KMeans(
            n_clusters=n_bins,
            random_state=random_state
        )
        kmeans.fit(feature.reshape((-1, 1)), target)

        centers = np.sort(kmeans.cluster_centers_.reshape(-1))

        l = len(centers) - 1
        splits = np.zeros(l)
        for i in range(l):
            splits[i] = (centers[i] + centers[i + 1]) / 2
        # kmeans 逻辑

        x_bins = split_points_to_bin(feature, splits)  # 返回0、1、2、3这样的数值分箱
        if target is None:
            _ = bin_badrate(x_bins)
            is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足
        else:
            bin_badrate_li, _ = bin_badrate(x_bins, target)
            is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足
            if is_need_monotonic:
                if is_monotonic:  # 满足最小分箱占比后，判断是否满足单调性
                    is_monotonic = pd.Series(bin_badrate_li).is_monotonic_decreasing or pd.Series(
                        bin_badrate_li).is_monotonic_increasing

        n_bins = len(splits) + 1  # 初始n_bins为10，对应的splits为9。需要+1，在后面的n_bins -= 1后，n_bins才会是9

        if n_bins <= 2:
            break
        n_bins -= 1

    if is_empty_bin and empty_mask.any():
        splits = np.append(splits, np.nan)

    return splits


def chi_bin(feature, target, balance=True, min_sample_rate=0.05, n_bins=None,
            is_need_monotonic=True, is_empty_bin=True, min_threshold=None):
    """
    卡方分箱

    Args:
        feature (array) : 某个x特征
        target (array) :  目标值y变量
        min_sample_rate (number) : 每个箱子的最小占比
        n_bins (int): 需要分成几箱
        is_need_monotonic (bool) : 是否强制单调
        is_empty_bin (bool) : 是否将空箱单独归为一个箱子
        min_threshold (number): 最小的卡方阀值

    Returns:
        array (array): 分割点
    """

    empty_mask = np.array([])
    if is_empty_bin:
        feature, target, empty_mask = split_empty(feature, target)
    else:
        feature = fillna(feature)

    if n_bins is None and min_threshold is None and is_empty_bin and empty_mask.any():
        n_bins = IS_EMPTY_BIN_DEFAULT_BINS
    elif n_bins is None and min_threshold is None:
        n_bins = DEFAULT_BINS

    if min_sample_rate and min_sample_rate < 1:
        min_sample_rate = len(feature) * min_sample_rate

    ###
    target_unique = np.unique(target)
    feature_unique = np.unique(feature)
    len_f = len(feature_unique)
    len_t = len(target_unique)

    grouped = np.zeros((len_f, len_t), dtype=float)

    for r in range(len_f):
        tmp = target[feature == feature_unique[r]]
        for c in range(len_t):
            grouped[r, c] = (tmp == target_unique[c]).sum()

    is_monotonic = False
    while not is_monotonic:

        # 卡方 逻辑
        while True:  # 内循环

            ###bmt
            # 判断卡方分箱是否同时满足最小分箱占比和箱子个数
            if len(grouped) <= n_bins and t_min_np(t_cols_sum_axis_1_np(grouped)) >= min_sample_rate:
                break
            ###bmt

            # 计算每一组的卡方
            l = len(grouped) - 1
            chi_list = np.zeros(l, dtype=float)
            chi_min = np.inf
            # chi_ix = []
            for i in range(l):
                chi = 0
                couple = grouped[i:i + 2, :]
                total = t_sum_np(couple)
                cols = t_cols_sum_axis_0_np(couple)
                # t_cols_sum_axis_1(couple)
                # rows = x
                rows = t_cols_sum_axis_1_np(couple)

                for j in range(couple.shape[0]):
                    for k in range(couple.shape[1]):
                        e = rows[j] * cols[k] / total
                        if e != 0:
                            chi += (couple[j, k] - e) ** 2 / e

                # 平衡卡方值
                if balance:
                    chi *= total

                chi_list[i] = chi

                if chi == chi_min:
                    chi_ix.append(i)
                    continue

                if chi < chi_min:
                    chi_min = chi
                    chi_ix = [i]

                # if chi < chi_min:
                #     chi_min = chi

            # 当最小值大于阈值时中断循环
            if min_threshold and chi_min > min_threshold:
                break

            # 获取最小卡方值的那组索引
            min_ix = np.array(chi_ix)
            # min_ix = np.where(chi_list == chi_min)[0]

            # 获取需要删除的索引
            drop_ix = min_ix + 1

            # 按索引合并
            retain_ix = min_ix[0]
            last_ix = retain_ix
            for ix in min_ix:
                # set a new group
                if ix - last_ix > 1:
                    retain_ix = ix

                # 将所有连续索引合并为一组
                for p in range(grouped.shape[1]):
                    grouped[retain_ix, p] = grouped[retain_ix, p] + grouped[ix + 1, p]

                last_ix = ix

            # 删除分组
            grouped = np.delete(grouped, drop_ix, axis=0)
            feature_unique = np.delete(feature_unique, drop_ix)

        # 卡方 逻辑

        splits = feature_unique[1:]

        is_monotonic = True
        if is_need_monotonic:
            x_bins = split_points_to_bin(feature, splits)  # 返回0、1、2、3这样的数值分箱
            bin_badrate_li, _ = bin_badrate(x_bins, target)

            # 不需要判断，内循环里面已经判断了
            # is_monotonic = True if _ >= min_sample_rate else False  # 判断最小分箱占比是否满足

            is_monotonic = pd.Series(bin_badrate_li).is_monotonic_decreasing or pd.Series(
                bin_badrate_li).is_monotonic_increasing

        n_bins = len(splits) + 1  # 初始n_bins为10，对应的splits为9。需要+1，在后面的n_bins -= 1后，n_bins才会是9
        if n_bins <= 2:
            break
        n_bins -= 1

    if is_empty_bin and empty_mask.any():
        splits = np.append(splits, np.nan)

    return splits


@support_dataframe(require_target=False)
def bin_method_run(feature, target=None, method='dt', return_bin=False, **kwargs):
    """
    对数据进行分箱
    Args:
        feature (array-like) : 某个x特征
        target (array-like) :  目标值y变量
        method (str): 分箱方法；'dt'、'chi'、'equal_freq'、'kmeans'四种供选择
        return_bin (bool): 是否返回分箱后的数据
        min_sample_rate (number) : 每个箱子的最小占比
        n_bins (int): 需要分成几箱
        is_need_monotonic (bool) : 是否强制单调
        is_empty_bin (bool) : 是否将空箱单独归为一个箱子
        min_threshold (number): 最小的卡方阀值

    Returns:
        array: 分割点
        array: 原始数据用分箱点替换后的数据

    """

    if method == 'dt':
        splits = dt_bin(feature, target, **kwargs)
    elif method == 'chi':
        splits = chi_bin(feature, target, **kwargs)
    elif method == 'equal_freq':
        splits = equal_freq_bin(feature, target, **kwargs)
    elif method == 'kmeans':
        splits = kmeans_bin(feature, target, **kwargs)
    else:
        splits = np.array([])

    ##返回splits
    if return_bin:
        bins = np.zeros(len(feature))
        if len(splits):
            if np.isnan(splits[-1]):
                mask = pd.isna(feature)
                bins[~mask] = split_points_to_bin(feature[~mask], splits[:-1])
                bins[mask] = len(splits)
            else:
                bins = split_points_to_bin(feature, splits)

        # return splits, pd.Series(bins, name=feature.name)
        return splits, bins

    return splits


def best_binning(df, x_list=[], target='target', **kwargs):
    """
    最优分箱
    Args:
        df (DataFrame) : 需要分箱的数据集
        x_list (list): 需要分箱的特征列表
        target (str): 目标变量
        **kwargs: 'dt'、'chi'、'equal_freq'、'kmeans'四种分箱方法的分箱参数

    Returns:

    """
    from .transformer import FeatureBin

    assert df[target].isin([0, 1]).all(), 'ERROR: :-) {} :-) 目标变量不是0/1值，请检查！！！'.format(target)
    iv_inflection_arr = []
    cutoff_dic = {}
    # for method in ['equal_freq', 'chi', 'dt', 'kmeans']:
    for method in ['equal_freq', 'chi', 'dt']:
        log.info("正在执行最优分箱之 [{}] ".format(method))

        fb = FeatureBin()
        fb.fit(df[x_list], df[target], method=method, **kwargs)
        cutoff_dic[method] = fb.splits_dict
        var_iv_inflection_df = calc_iv_and_inflection_point(fb.transform(df, labels=True)[x_list + [target]],
                                                            target=target, bin_func=method)
        iv_inflection_arr.append(var_iv_inflection_df.reset_index())

        log.info("执行最优分箱之 [{}] over!!!".format(method))

    # 分析获取最优分箱的结果
    iv_inflection_df = pd.concat(iv_inflection_arr, axis=0)
    best_binning_result = iv_inflection_df.groupby('var_name').apply(
        lambda x: x.sort_values(['inflection_point_num', 'IV', 'bin_count'], ascending=[True, False, False]).head(
            1)).set_index("var_name")

    # 找到各个变量最优分箱的分箱方法
    best_binning_func_index = best_binning_result['bin_func'].to_dict()
    best_cutoff = {k: cutoff_dic[v][k] for k, v in best_binning_func_index.items()}
    fb.manual_bin(best_cutoff)

    # 最优分箱，多返回一个数据集记录选择最优的过程
    best_binning_result = best_binning_result.reset_index('var_name')
    return fb, best_binning_result



#==============================================================================
# File: feature_selection.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: feature_selection.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-20
'''

import numpy as np
import pandas as pd
from joblib import Parallel, delayed
import shap
from tqdm import tqdm
from xgboost import XGBClassifier

from .metrics import psi, get_ks
from .statistics import calc_iv
from .transformer import FeatureBin
from .utils import unpack_tuple, select_features_dtypes, is_continuous


def select_features_by_miss(df, nan=None, threshold=0.9, include_cols=[], return_drop=False, only_return_drop=False,
                            if_select_flow=False):
    """
    通过缺失率筛选特征
    Args:
        df (DataFrame): 需要进行特征筛选的数据集
        nan (str, regex, list, dict, Series, int, float, or None): 要替换为空的具体值
        threshold (float): 缺失率筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if nan is not None:
        df = df.replace(nan, np.nan)

    if include_cols:
        cols = include_cols
    else:
        cols = list(df.columns)

    missing_series = df[cols].isnull().sum() / len(df)

    del_c = list(missing_series[missing_series > threshold].index)

    if if_select_flow:
        return (del_c, threshold,
                pd.DataFrame({'feature': cols, 'miss_rate': missing_series}))  # TODO 需要检查下cols和psi_series是不是一一对应的

    if only_return_drop:
        return del_c

    r = df.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    return unpack_tuple(res)


def select_features_by_concentrate(df, nan=None, threshold=0.9, include_cols=[], return_drop=False,
                                   only_return_drop=False, if_select_flow=False):
    """
    通过集中度筛选特征
    Args:
        df (DataFrame): 需要进行特征筛选的数据集
        nan (str, regex, list, dict, Series, int, float, or None): 要替换为空的具体值
        threshold (float): 集中度筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if nan is not None:
        df = df.replace(nan, np.nan)

    if include_cols:
        cols = include_cols
    else:
        cols = list(df.columns)
    del_c = []
    max_ratios_ls = []
    row_count = df.shape[0]
    for col in cols:
        max_ratios = max(df[col].value_counts() / row_count)  # 缺失的占比不会被放进来
        # max_ratios = max(df[col].value_counts(dropna=False, normalize=True))    #缺失的占比会被放进来

        max_ratios_ls.append(max_ratios)
        if max_ratios > threshold:
            del_c.append(col)

    if if_select_flow:
        return (del_c, threshold,
                pd.DataFrame(
                    {'feature': cols, 'concentration_rate': max_ratios_ls}))

    if only_return_drop:
        return del_c

    r = df.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    return unpack_tuple(res)


def select_features_by_psi(base, no_base, target='target', threshold=0.05, include_cols=[],
                           return_drop=False, only_return_drop=False, if_select_flow=False, feature_bin=None):
    """
    通过psi筛选特征
    Args:
        base (DataFrame): 基准数据集
        no_base (DataFrame): 非基准数据集
        target (str): 目标变量名称
        threshold (float): psi筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流
        feature_bin (autobmt.FeatureBin): 特征分箱对象

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if include_cols:
        cols = include_cols
    else:
        cols = list(base.columns)

    if feature_bin is None or feature_bin == False:
        log.info('未进行分箱计算的psi')
        psi_series = psi(no_base[cols], base[cols])
    else:
        to_bin_cols = []
        if isinstance(feature_bin, FeatureBin):
            for i in cols:
                if is_continuous(base[i]):
                    to_bin_cols.append(i)
            if feature_bin.splits_dict:
                to_bin_cols = list(set(to_bin_cols) - set(feature_bin.splits_dict.keys()))
        else:
            feature_bin = FeatureBin()
            for i in cols:
                if is_continuous(base[i]):
                    to_bin_cols.append(i)
        if to_bin_cols:
            feature_bin.fit(base[to_bin_cols], base[target], update=False, method='dt', is_need_monotonic=False)

        psi_series = psi(feature_bin.transform(no_base[cols]), feature_bin.transform(base[cols]))

    del_c = list(psi_series[psi_series > threshold].index)  # 大于psi阈值的变量删除

    if if_select_flow:
        return (
            del_c, threshold,
            pd.DataFrame({'feature': psi_series.index, 'PSI': psi_series}))

    if only_return_drop:
        return del_c

    r = base.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    return unpack_tuple(res)


def select_features_by_iv(df, target='target', threshold=0.02, cpu_cores=-1, include_cols=[], return_drop=False,
                          return_iv=False,
                          only_return_drop=False, if_select_flow=False, feature_bin=None):
    """
    通过iv筛选特征
    Args:
        df (DataFrame): 需要进行特征筛选的数据集
        target (str): 目标变量名称
        threshold (float): iv筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        return_iv (bool): 是否返回iv
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流
        feature_bin (autobmt.FeatureBin): 特征分箱对象

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if include_cols:
        cols = np.array(include_cols)
    else:
        cols = np.array(df.columns)

    iv_l = Parallel(n_jobs=cpu_cores, backend='threading')(
        delayed(calc_iv)(df[cols[i]], df[target], feature_bin=feature_bin) for i in range(len(cols))
    )
    iv = np.array(iv_l)

    drop_index = np.where(iv < threshold)

    del_c = cols[drop_index]

    if if_select_flow:
        return (del_c, threshold, pd.DataFrame({'feature': cols, 'IV': iv}))

    if only_return_drop:
        return del_c

    r = df.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    if return_iv:
        res += (pd.Series(iv, index=cols),)

    return unpack_tuple(res)


def select_features_by_iv_diff(dev, no_dev, target='target', threshold=2, include_cols=[], return_drop=False,
                               return_iv=False, only_return_drop=False, if_select_flow=False, feature_bin=None):
    """
    通过iv差值筛选特征
    Args:
        dev (DataFrame): 基准数据集
        no_dev (DataFrame): 非基准数据集
        target (str): 目标变量名称
        threshold (float): iv差值筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        return_iv (bool): 是否返回iv
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流
        feature_bin (autobmt.FeatureBin): 特征分箱对象

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if include_cols:
        cols = np.array(include_cols)
    else:
        cols = np.array(dev.columns)

    iv = np.zeros(len(cols))
    iv_no_dev = np.zeros(len(cols))
    if feature_bin is None:
        feature_bin = FeatureBin()
        for i in range(len(cols)):
            iv[i] = calc_iv(dev[cols[i]], dev[target], feature_bin=feature_bin)
            no_dev_feature = feature_bin.transform(no_dev[cols[i]])
            iv_no_dev[i] = calc_iv(no_dev_feature, no_dev[target], feature_bin=feature_bin)
    else:
        for i in range(len(cols)):
            iv[i] = calc_iv(dev[cols[i]], dev[target], feature_bin=feature_bin)
            no_dev_feature = feature_bin.transform(no_dev[cols[i]])
            iv_no_dev[i] = calc_iv(no_dev_feature, no_dev[target], feature_bin=feature_bin)

    # iv_diff = abs(iv - iv_no_dev) * 10
    iv_diff = (iv - iv_no_dev) * 10

    drop_index = np.where(iv_diff > threshold)  # IV差值大于2个点的变量剔除

    del_c = cols[drop_index]

    if if_select_flow:
        return (del_c, threshold,
                pd.DataFrame({'feature': cols, 'dev_nodev_iv_diff': iv_diff}))

    if only_return_drop:
        return del_c

    r = dev.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    if return_iv:
        res += (pd.Series(iv, index=cols),)

    return unpack_tuple(res)


def select_features_by_corr(df, target='target', by='IV', threshold=0.7, cpu_cores=-1, include_cols=[],
                            return_drop=False,
                            only_return_drop=False, if_select_flow=False, feature_bin=None):
    """
    通过通过相关性筛选特征
    Args:
        df (DataFrame): 需要进行特征筛选的数据集
        target (str): 目标变量名称
        by (str|array): 用于删除特征的特征权重
        threshold (float): iv筛选阀值
        include_cols (list): 需要进行筛选的特征
        return_drop (bool): 是否返回删除的特征列表
        only_return_drop (bool): 是否只返回删除的特征列表
        if_select_flow (bool): 是否是筛选工作流
        feature_bin (autobmt.FeatureBin): 特征分箱对象

    Returns:
        DataFrame: 筛选后的数据集
        list: 删除的特征列表
    """
    if include_cols:
        cols = include_cols
    else:
        cols = list(df.columns)

    if isinstance(by, pd.DataFrame):
        if by.shape[1] == 1:
            by = pd.Series(by.iloc[:, 0].values, index=by.index)
        else:
            by = pd.Series(by.iloc[:, 1].values, index=by.iloc[:, 0].values)

    if not isinstance(by, (str, pd.Series)):
        by = pd.Series(by, index=df.columns)

    # 计算iv
    if isinstance(by, str):
        df_corr = df[cols].corr().abs()
        # df_corr = df[cols].fillna(-999).corr().abs()
        ix, cn = np.where(np.triu(df_corr.values, 1) > threshold)  # ix是行，cn是列
        if len(ix):
            gt_thre = np.unique(np.concatenate((ix, cn)))
            gt_thre_cols = df_corr.index[gt_thre]
            iv_t = Parallel(n_jobs=cpu_cores, backend='threading')(
                delayed(calc_iv)(df[i], df[target], feature_bin=feature_bin, return_name=True, col_name=i) for i in
                gt_thre_cols
            )
            iv = dict(iv_t)

            by = pd.Series(iv, index=gt_thre_cols)

    # 给重要性排下序，倒序
    by = by[list(set(by.index) & set(cols))].sort_values(ascending=False)

    by.index = by.index.astype(type(list(df.columns)[0]))
    df_corr = df[list(by.index)].corr().abs()
    # df_corr = df[list(by.index)].fillna(-999).corr().abs()

    ix, cn = np.where(np.triu(df_corr.values, 1) > threshold)

    del_all = []

    if len(ix):

        for i in df_corr:

            if i not in del_all:
                # 找出与当前特征的相关性大于域值的特征
                del_tmp = list(df_corr[i][(df_corr[i] > threshold) & (df_corr[i] != 1)].index)

                # 比较当前特征与需要删除的特征的特征重要性
                if del_tmp:
                    by_tmp = by.loc[del_tmp]
                    del_l = list(by_tmp[by_tmp <= by.loc[i]].index)
                    del_all.extend(del_l)

    del_c = list(set(del_all))

    if if_select_flow:
        return (del_c, threshold, pd.DataFrame({'feature': df_corr.index}))

    if only_return_drop:
        return del_c

    r = df.drop(columns=del_c)

    res = (r,)
    if return_drop:
        res += (del_c,)

    return unpack_tuple(res)


from .logger_utils import Logger

log = Logger(level='info', name=__name__).logger


class ShapSelectFeature:
    def __init__(self, estimator, linear=False, estimator_is_fit_final=False):
        self.estimator = estimator
        self.linear = linear
        self.weight = None
        self.estimator_is_fit_final = estimator_is_fit_final

    def fit(self, X, y, exclude=None):
        '''

        Args:
            X:
            y:
            exclude:

        Returns:

        '''
        if exclude is not None:
            X = X.drop(columns=exclude)
        if not self.estimator_is_fit_final:
            self.estimator.fit(X, y)
        if self.linear:
            explainer = shap.LinearExplainer(self.estimator, X)
        else:
            estimator = self.estimator.get_booster()
            temp = estimator.save_raw()[4:]
            estimator.save_raw = lambda: temp
            explainer = shap.TreeExplainer(estimator)
        shap_values = explainer.shap_values(X)
        shap_abs = np.abs(shap_values)
        shap_importance_list = shap_abs.mean(0)
        self.weight = pd.DataFrame(shap_importance_list, index=X.columns, columns=['weight'])
        return self.weight


def feature_select(datasets, fea_names, target, feature_select_method='shap', method_threhold=0.001,
                   corr_threhold=0.8, psi_threhold=0.1, params={}):
    '''

    Args:
        datasets:
        fea_names:
        target:
        feature_select_method:
        method_threhold:
        corr_threhold:
        psi_threhold:
        params:

    Returns:

    '''
    dev_data = datasets['dev']
    nodev_data = datasets['nodev']

    params = {
        'learning_rate': params.get('learning_rate', 0.05),
        'n_estimators': params.get('n_estimators', 200),
        'max_depth': params.get('max_depth', 3),
        # 'min_child_weight': params.get('min_child_weight', max(round(len(dev_data) * 0.01), 50)),
        'min_child_weight': params.get('min_child_weight', 5),
        'subsample': params.get('subsample', 0.7),
        'colsample_bytree': params.get('colsample_bytree', 0.9),
        'colsample_bylevel': params.get('colsample_bylevel', 0.7),
        'gamma': params.get('gamma', 7),
        'reg_alpha': params.get('reg_alpha', 10),
        'reg_lambda': params.get('reg_lambda', 10)
    }

    xgb_clf = XGBClassifier(**params)
    xgb_clf.fit(dev_data[fea_names], dev_data[target])

    if feature_select_method == 'shap':
        shap_model = ShapSelectFeature(estimator=xgb_clf, estimator_is_fit_final=True)
        fea_weight = shap_model.fit(dev_data[fea_names], dev_data[target])
        fea_weight.sort_values(by='weight', inplace=True)
        fea_weight = fea_weight[fea_weight['weight'] >= method_threhold]
        log.info('Shap阈值: {}'.format(method_threhold))
        log.info('Shap剔除的变量个数: {}'.format(len(fea_names) - fea_weight.shape[0]))
        log.info('Shap保留的变量个数: {}'.format(fea_weight.shape[0]))
        fea_names = list(fea_weight.index)
        log.info('*' * 50 + 'Shap筛选变量' + '*' * 50)


    elif feature_select_method == 'feature_importance':
        fea_weight = pd.DataFrame(list(xgb_clf.get_booster().get_score(importance_type='gain').items()),
                                  columns=['fea_names', 'weight']
                                  ).sort_values('weight').set_index('fea_names')
        fea_weight = fea_weight[fea_weight['weight'] >= method_threhold]
        log.info('feature_importance阈值: {}'.format(method_threhold))
        log.info('feature_importance剔除的变量个数: {}'.format(len(fea_names) - fea_weight.shape[0]))
        fea_names = list(fea_weight.index)
        log.info('feature_importance保留的变量个数: {}'.format(fea_names))
        log.info('*' * 50 + 'feature_importance筛选变量' + '*' * 50)

    if corr_threhold:
        _, del_fea_list = select_features_by_corr(dev_data[fea_names], by=fea_weight, threshold=corr_threhold,
                                                  return_drop=True)
        log.info('相关性阈值: {}'.format(corr_threhold))
        log.info('相关性剔除的变量个数: {}'.format(len(del_fea_list)))
        fea_names = [i for i in fea_names if i not in del_fea_list]
        # fea_names = list(set(fea_names) - set(del_fea_list))
        log.info('相关性保留的变量个数: {}'.format(len(fea_names)))
        log.info('*' * 50 + '相关性筛选变量' + '*' * 50)

    if psi_threhold:
        psi_df = psi(dev_data[fea_names], nodev_data[fea_names]).sort_values(0)
        psi_df = psi_df.reset_index()
        psi_df = psi_df.rename(columns={'index': 'fea_names', 0: 'psi'})
        psi_list = psi_df[psi_df.psi < psi_threhold].fea_names.tolist()
        log.info('PSI阈值: {}'.format(psi_threhold))
        log.info('PSI剔除的变量个数: {}'.format(len(fea_names) - len(psi_list)))
        fea_names = [i for i in fea_names if i in psi_list]
        # fea_names = list(set(fea_names) and set(psi_list))
        log.info('PSI保留的变量个数: {}'.format(len(fea_names)))
        log.info('*' * 50 + 'PSI筛选变量' + '*' * 50)

    return fea_names


def stepwise_del_feature(model, datasets, fea_names, target, params={}):
    '''

    Args:
        datasets:
        fea_names:
        target:
        params:

    Returns:

    '''
    log.info("开始逐步删除变量")
    dev_data = datasets['dev']
    nodev_data = datasets['nodev']
    # stepwise_del_params = {
    #     'learning_rate': params.get('learning_rate', 0.05),
    #     'n_estimators': params.get('n_estimators', 200),
    #     'max_depth': params.get('max_depth', 3),
    #     'min_child_weight': params.get('min_child_weight', max(round(len(dev_data) * 0.01), 50)),
    #     'subsample': params.get('subsample', 0.7),
    #     'colsample_bytree': params.get('colsample_bytree', 0.9),
    #     'colsample_bylevel': params.get('colsample_bylevel', 0.7),
    #     'gamma': params.get('gamma', 7),
    #     'reg_alpha': params.get('reg_alpha', 10),
    #     'reg_lambda': params.get('reg_lambda', 10)
    # }

    # xgb_clf = XGBClassifier(**stepwise_del_params)
    model.fit(dev_data[fea_names], dev_data[target])

    pred_test = model.predict_proba(nodev_data[fea_names])[:, 1]
    pred_train = model.predict_proba(dev_data[fea_names])[:, 1]

    test_ks = get_ks(nodev_data[target], pred_test)
    train_ks = get_ks(dev_data[target], pred_train)
    log.info('test_ks is : {}'.format(test_ks))
    log.info('train_ks is : {}'.format(train_ks))

    train_number, oldks, del_list = 0, test_ks, list()
    log.info('train_number: {}, test_ks: {}'.format(train_number, test_ks))

    # while True:
    #     flag = True
    #     for fea_name in tqdm(fea_names):
    #         print('变量{}进行逐步：'.format(fea_name))
    #         names = [fea for fea in fea_names if fea_name != fea]
    #         print('变量names is：', names)
    #         xgb_clf.fit(dev_data[names], dev_data[target])
    #         train_number += 1
    #         pred_test = xgb_clf.predict_proba(nodev_data[names])[:, 1]
    #         test_ks = get_ks(nodev_data[target], pred_test)
    #         if test_ks >= oldks:
    #             oldks = test_ks
    #             flag = False
    #             del_list.append(fea_name)
    #             log.info(
    #                 '等于或优于之前结果 train_number: {}, test_ks: {} by feature: {}'.format(train_number, test_ks, fea_name))
    #             fea_names = names
    #     if flag:
    #         print('=====================又重新逐步==========')
    #         break
    #     log.info("结束逐步删除变量 train_number: %s, test_ks: %s del_list: %s" % (train_number, oldks, del_list))
    #     print('oldks is ：',oldks)
    #     print('fea_names is : ',fea_names)

    for fea_name in tqdm(fea_names):
        names = [fea for fea in fea_names if fea_name != fea]
        model.fit(dev_data[names], dev_data[target])
        train_number += 1
        pred_test = model.predict_proba(nodev_data[names])[:, 1]
        test_ks = get_ks(nodev_data[target], pred_test)
        if test_ks >= oldks:
            oldks = test_ks
            del_list.append(fea_name)
            log.info(
                '等于或优于之前结果 train_number: {}, test_ks: {} by feature: {}'.format(train_number, test_ks, fea_name))
            fea_names = names
    log.info("结束逐步删除变量 train_number: %s, test_ks: %s del_list: %s" % (train_number, oldks, del_list))

    ########################
    log.info('逐步剔除的变量个数: {}'.format(del_list))
    fea_names = [i for i in fea_names if i not in del_list]
    # fea_names = list(set(fea_names) - set(del_list))
    log.info('逐步保留的变量个数: {}'.format(len(fea_names)))
    log.info('*' * 50 + '逐步筛选变量' + '*' * 50)

    return del_list, fea_names


class FeatureSelection:
    def __init__(self, df, target='target', data_type='type',
                 exclude_columns=['key', 'target', 'apply_time', 'type'], params=None,
                 match_dict=None):
        """
        特征选择模块，初始化方法
        Args:
            df (DataFrame): 需要进行变量筛选的数据集
            target (str): 目标值y变量名称
            data_type (str): 数据集划分标示的名称【即划分train、test、oot的字段名称】
            exclude_columns (list): 需要排除的特征
            match_dict (DataFrame): 数据源特征字典
            params (dict): 筛选特征的方法字典，有'empty'、'const'、'psi'、'iv'、'iv_diff'、'corr' 6种筛选方法供选择。字典形如：
            {
                'empty': {'threshold': 0.9},    #若特征的缺失值大于0.9被删除
                'const': {'threshold': 0.95},   #若特征单个值的占比大于0.95被删除
                'psi': {'threshold': 0.05},  #若特征在train、test上psi值大于0.05被删除
                'iv': {'threshold': 0.02},  #若特征的iv值小于0.02被删除
                'iv_diff': {'threshold': 2},    #若特征在train、test上的iv差值乘10后，大于2，特征被删除
                'corr': {'threshold': 0.7}, #若两个特征相关性高于0.7时，iv值低的特征被删除
            }
        """
        self.df = df
        self.target = target
        self.data_type = data_type
        self.exclude_columns = exclude_columns + [self.target]
        self.match_dict = match_dict
        self.params = params
        self.check()
        self.features = [name for name in list(self.df.columns) if name not in self.exclude_columns]
        self.feature_dict = self.get_feature_dict()
        self.select_log_df = self.build_select_log_df()
        self.step_evaluate_log_df = []  # 记录每一步的评估结果

    @property
    def get_features(self):
        """返回当前数据集最新的特征"""
        return [name for name in list(self.df.columns) if name not in self.exclude_columns]

    @property
    def get_evaluate_df_log(self):
        """合并每一步的评估结果"""
        if len(self.step_evaluate_log_df) == 0:
            log.info("并未进行评估过!!!")
            return None
        else:
            evaluate_log_df = pd.concat(self.step_evaluate_log_df, axis=0).reset_index(drop=True)
            return evaluate_log_df

    def get_feature_dict(self):
        """通过数据源简称去数据字典中获取数据源的特征名称"""
        if self.match_dict is not None and isinstance(self.match_dict, pd.DataFrame):
            if self.match_dict.columns.isin(['feature', 'cn']).all():
                model_name_dict_df = pd.DataFrame({'feature': self.features})
                model_name_dict_df = model_name_dict_df.merge(self.match_dict[['feature', 'cn']], on='feature',
                                                              how='left')
                return model_name_dict_df
            else:
                raise KeyError("原始数据字典中没有feature或cn字段，请保证同时有feature字段和cn字段")
        else:
            model_name_dict_df = pd.DataFrame(
                {'feature': self.features, 'cn': ""})
        return model_name_dict_df

    def build_select_log_df(self):
        """返回特征字典，如果需要个性化修改，可以在此方法中修改"""
        return self.feature_dict

    def mapping_selection_func(self):
        """
        特征选择方法映射类
        如果需要增加新的特征选择方法，只需要增加这个字典即可
        注意:python3.6的字典默认会按照顺序进行遍历
        """
        return {
            "empty": select_features_by_miss,
            "const": select_features_by_concentrate,
            "psi": select_features_by_psi,
            "iv": select_features_by_iv,
            "iv_diff": select_features_by_iv_diff,
            "corr": select_features_by_corr,
        }

    def select(self):
        """
        执行定义的特征选择方法，返回筛选过后的数据集，剩余特征名称，以及筛选过程
        Returns:

        """
        log.info('开始执行特征选择模块... 数据集结构为[{}]'.format(self.df.shape))

        if self.params is None:  # 默认只进行3种方式进行变量筛选
            self.params = {
                'empty': {'threshold': 0.9},
                # 'const': {'threshold': 0.95},
                # 'psi': {'threshold': 0.05},
                'iv': {'threshold': 0.02},
                # 'iv_diff': {'threshold': 2},
                'corr': {'threshold': 0.7},
            }
            if self.data_type in self.df:
                self.params['psi'] = {'threshold': 0.05, 'target': self.target}
            log.info('未指定筛选方法的阈值，使用默认方法和阈值：{}'.format(self.params))

        if self.data_type in self.df:
            dev_data = self.df[self.df[self.data_type] == 'train']
            if 'oot' in np.unique(self.df[self.data_type]):
                nodev_data = self.df[self.df[self.data_type] == 'oot']
            else:
                nodev_data = self.df[self.df[self.data_type] == 'test']
        else:
            dev_data = self.df

        fb = FeatureBin()

        for k, v in self.params.items():
            v.update({'if_select_flow': True})
            v.update({'include_cols': self.get_features})
            if k in ['iv', 'iv_diff', 'psi', 'corr']:
                v.update({'feature_bin': fb})
                if 'target' not in v:
                    v.update({'target': self.target})
            # 执行具体的筛选方法
            if k in ['iv_diff', 'psi'] and self.data_type in self.df:
                del_c, th, fea_value_df = self.mapping_selection_func()[k](dev_data, nodev_data, **v)
                log.info('删除变量 ：{}'.format(del_c))
            else:
                del_c, th, fea_value_df = self.mapping_selection_func()[k](dev_data, **v)
                log.info('删除变量 ：{}'.format(del_c))
            if k == 'iv':
                # 将算好的iv值放进去
                self.params['corr']['by'] = fea_value_df

            self.df = self.df.drop(columns=del_c)

            self.select_log_df = self.select_log_df.merge(fea_value_df, on='feature', how='left')
            if k == 'iv':
                step_name = "{}_selection_feature_flag (<{})".format(k, th)
            else:
                step_name = "{}_selection_feature_flag (>{})".format(k, th)
            log.info("{} 方法剔除变量，阈值为{}，剔除的变量有 : {} 个".format(k, step_name, len(del_c)))
            self.select_log_df[step_name] = self.select_log_df['feature'].map(lambda x: 1 if x in del_c else 0)

        # 所有的特征选择方法，只要命中一个，即剔除该特征
        filter_condition_features = [name for name in list(self.select_log_df.columns) if '_feature_flag' in name]
        self.select_log_df['feature_filter_flag'] = self.select_log_df[filter_condition_features].sum(axis=1)
        log.info('特征选择执行完成... 数据集结构为[{}]'.format(self.df.shape))

        # return self.df, self.get_features, self.select_log_df, self.get_evaluate_df_log, fb
        return self.df, self.get_features, self.select_log_df, fb

    def check(self):
        """
        特征选择模块，前置检查，符合要求，则往下运行
        Returns:

        """
        log.info('开始进行前置检查')
        if self.df is None or not isinstance(self.df, pd.DataFrame):
            raise ValueError('数据集不能为空并且数据集必须是dataframe!!!')

        if self.data_type not in self.df:
            log.info('train、test数据集标识的字段名不存在！或未进行数据集的划分，筛选变量无法使用psi、iv_diff进行筛选，建议请将数据集划分为train、test!!!')
            if self.params is not None:
                if 'psi' in self.params:
                    del self.params['psi']
                if 'iv_diff' in self.params:
                    del self.params['iv_diff']
        else:
            data_type_ar = np.unique(self.df[self.data_type])
            if 'train' not in data_type_ar:
                raise KeyError("""没有开发样本，数据集标识字段{}没有`train`该取值!!!""".format(self.data_type))

            if 'test' not in data_type_ar:
                raise KeyError("""没有验证样本，数据集标识字段{}没有`test`该取值!!!""".format(self.data_type))

        if self.target is None:
            raise ValueError('数据集的目标变量名称不能为空!!!')

        if self.target not in self.df:
            raise KeyError('样本中没有目标变量y值!!!')

        if self.exclude_columns is None or self.target not in self.exclude_columns:
            raise ValueError('exclude_columns 不能为空，必须包含target字段!!!')
        n_cols, c_cols, d_cols = select_features_dtypes(self.df, exclude=self.exclude_columns)
        log.info('数值特征个数: {}'.format(len(n_cols)))
        log.info('字符特征个数: {}'.format(len(c_cols)))
        log.info('日期特征个数: {}'.format(len(d_cols)))
        if len(c_cols) > 0:
            log.info('数据集中包含有{}个字符特征,{}个日期特征'.format(len(c_cols), len(d_cols)))



#==============================================================================
# File: feature_selection_2_treemodel.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: feature_selection_2_treemodel.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-04-26
'''

import warnings

import numpy as np
import pandas as pd
import shap
from tqdm import tqdm
from xgboost import XGBClassifier

from .utils import get_ks

warnings.filterwarnings('ignore')

from .logger_utils import Logger

log = Logger(level='info', name=__name__).logger


class ShapSelectFeature:
    def __init__(self, estimator, linear=False, estimator_is_fit_final=False):
        self.estimator = estimator
        self.linear = linear
        self.weight = None
        self.estimator_is_fit_final = estimator_is_fit_final

    def fit(self, X, y, exclude=None):
        '''

        Args:
            X:
            y:
            exclude:

        Returns:

        '''
        if exclude is not None:
            X = X.drop(columns=exclude)
        if not self.estimator_is_fit_final:
            self.estimator.fit(X, y)
        if self.linear:
            explainer = shap.LinearExplainer(self.estimator, X)
        else:
            estimator = self.estimator.get_booster()
            temp = estimator.save_raw()[4:]
            estimator.save_raw = lambda: temp
            explainer = shap.TreeExplainer(estimator)
        shap_values = explainer.shap_values(X)
        shap_abs = np.abs(shap_values)
        shap_importance_list = shap_abs.mean(0)
        self.weight = pd.DataFrame(shap_importance_list, index=X.columns, columns=['weight'])
        return self.weight


def corr_select_feature(frame, by='auc', threshold=0.95, return_frame=False):
    '''

    Args:
        frame:
        by:
        threshold:
        return_frame:

    Returns:

    '''
    if not isinstance(by, (str, pd.Series)):

        if isinstance(by, pd.DataFrame):
            if by.shape[1] == 1:
                by = pd.Series(by.iloc[:, 0].values, index=by.index)
            else:
                by = pd.Series(by.iloc[:, 1].values, index=by.iloc[:, 0].values)
            # by = pd.Series(by.iloc[:, 1].values, index=frame.columns)
        else:
            by = pd.Series(by, index=frame.columns)

    # 给重要性排下序
    by.sort_values(ascending=False, inplace=True)
    # print('给重要性排下序：', by)

    # df = frame.copy()

    by.index = by.index.astype(type(list(frame.columns)[0]))
    df_corr = frame[list(by.index)].fillna(-999).corr().abs()  # 填充
    # df_corr = frame[list(by.index)].corr().abs()

    ix, cn = np.where(np.triu(df_corr.values, 1) > threshold)

    del_all = []

    if len(ix):

        for i in df_corr:

            if i not in del_all:
                # 找出与当前特征的相关性大于域值的特征
                del_tmp = list(df_corr[i][(df_corr[i] > threshold) & (df_corr[i] != 1)].index)

                # 比较当前特征与需要删除的特征的特征重要性
                if del_tmp:
                    by_tmp = by.loc[del_tmp]
                    del_l = list(by_tmp[by_tmp <= by.loc[i]].index)
                    del_all.extend(del_l)

    del_f = list(set(del_all))

    if return_frame:
        r = frame.drop(columns=del_f)
        return (del_f, r)

    return del_f


def psi(no_base, base, return_frame=False):
    '''
    psi计算
    Args:
        no_base:非基准数据集
        base:基准数据集
        return_frame:是否返回详细的psi数据集

    Returns:
        float或Series
    '''
    psi = list()
    frame = list()

    if isinstance(no_base, pd.DataFrame):
        for col in no_base:
            p, f = calc_psi(no_base[col], base[col])
            psi.append(p)
            frame.append(f)

        psi = pd.Series(psi, index=no_base.columns)

        frame = pd.concat(
            frame,
            keys=no_base.columns,
            names=['columns', 'id'],
        ).reset_index()
        frame = frame.drop(columns='id')
    else:
        psi, frame = calc_psi(no_base, base)

    res = (psi,)

    if return_frame:
        res += (frame,)

    return unpack_tuple(res)


def unpack_tuple(x):
    if len(x) == 1:
        return x[0]
    else:
        return x


def calc_psi(no_base, base):
    '''
    psi计算的具体逻辑
    Args:
        no_base: 非基准数据集
        base: 基准数据集

    Returns:
        float或DataFrame
    '''
    no_base_prop = pd.Series(no_base).value_counts(normalize=True, dropna=False)
    base_prop = pd.Series(base).value_counts(normalize=True, dropna=False)

    psi = np.sum((no_base_prop - base_prop) * np.log(no_base_prop / base_prop))

    frame = pd.DataFrame({
        'no_base': no_base_prop,
        'base': base_prop,
    })
    frame.index.name = 'value'

    return psi, frame.reset_index()


def feature_select(datasets, fea_names, target, feature_select_method='shap', method_threhold=0.001,
                   corr_threhold=0.8, psi_threhold=0.1, params={}):
    '''

    Args:
        datasets:
        fea_names:
        target:
        feature_select_method:
        method_threhold:
        corr_threhold:
        psi_threhold:
        params:

    Returns:

    '''
    dev_data = datasets['dev']
    nodev_data = datasets['nodev']

    params = {
        'learning_rate': params.get('learning_rate', 0.05),
        'n_estimators': params.get('n_estimators', 200),
        'max_depth': params.get('max_depth', 3),
        'min_child_weight': params.get('min_child_weight', 5),
        'subsample': params.get('subsample', 0.7),
        'colsample_bytree': params.get('colsample_bytree', 0.9),
        'colsample_bylevel': params.get('colsample_bylevel', 0.7),
        'gamma': params.get('gamma', 7),
        'reg_alpha': params.get('reg_alpha', 10),
        'reg_lambda': params.get('reg_lambda', 10)
    }

    xgb_clf = XGBClassifier(**params)
    xgb_clf.fit(dev_data[fea_names], dev_data[target])

    if feature_select_method == 'shap':
        shap_model = ShapSelectFeature(estimator=xgb_clf, estimator_is_fit_final=True)
        fea_weight = shap_model.fit(dev_data[fea_names], dev_data[target])
        fea_weight.sort_values(by='weight', inplace=True)
        fea_weight = fea_weight[fea_weight['weight'] >= method_threhold]
        log.info('Shap阈值: {}'.format(method_threhold))
        log.info('Shap剔除的变量个数: {}'.format(len(fea_names) - fea_weight.shape[0]))
        log.info('Shap保留的变量个数: {}'.format(fea_weight.shape[0]))
        fea_names = list(fea_weight.index)
        log.info('*' * 50 + 'Shap筛选变量' + '*' * 50)


    elif feature_select_method == 'feature_importance':
        fea_weight = pd.DataFrame(list(xgb_clf.get_booster().get_score(importance_type='gain').items()),
                                  columns=['fea_names', 'weight']
                                  ).sort_values('weight').set_index('fea_names')
        fea_weight = fea_weight[fea_weight['weight'] >= method_threhold]
        log.info('feature_importance阈值: {}'.format(method_threhold))
        log.info('feature_importance剔除的变量个数: {}'.format(len(fea_names) - fea_weight.shape[0]))
        fea_names = list(fea_weight.index)
        log.info('feature_importance保留的变量个数: {}'.format(fea_names))
        log.info('*' * 50 + 'feature_importance筛选变量' + '*' * 50)

    if corr_threhold:
        del_fea_list = corr_select_feature(dev_data[fea_names], by=fea_weight, threshold=0.8)
        log.info('相关性阈值: {}'.format(corr_threhold))
        log.info('相关性剔除的变量个数: {}'.format(len(del_fea_list)))
        fea_names = [i for i in fea_names if i not in del_fea_list]
        # fea_names = list(set(fea_names) - set(del_fea_list))
        log.info('相关性保留的变量个数: {}'.format(len(fea_names)))
        log.info('*' * 50 + '相关性筛选变量' + '*' * 50)

    if psi_threhold:
        psi_df = psi(dev_data[fea_names], nodev_data[fea_names]).sort_values(0)
        psi_df = psi_df.reset_index()
        psi_df = psi_df.rename(columns={'index': 'fea_names', 0: 'psi'})
        psi_list = psi_df[psi_df.psi < psi_threhold].fea_names.tolist()
        log.info('PSI阈值: {}'.format(psi_threhold))
        log.info('PSI剔除的变量个数: {}'.format(len(fea_names) - len(psi_list)))
        fea_names = [i for i in fea_names if i in psi_list]
        # fea_names = list(set(fea_names) and set(psi_list))
        log.info('PSI保留的变量个数: {}'.format(len(fea_names)))
        log.info('*' * 50 + 'PSI筛选变量' + '*' * 50)

    return fea_names


def stepwise_del_feature(datasets, fea_names, target, params={}):
    '''

    Args:
        datasets:
        fea_names:
        target:
        params:

    Returns:

    '''
    log.info("开始逐步删除变量")
    dev_data = datasets['dev']
    nodev_data = datasets['nodev']
    stepwise_del_params = {
        'learning_rate': params.get('learning_rate', 0.05),
        'n_estimators': params.get('n_estimators', 200),
        'max_depth': params.get('max_depth', 3),
        'min_child_weight': params.get('min_child_weight', 5),
        'subsample': params.get('subsample', 0.7),
        'colsample_bytree': params.get('colsample_bytree', 0.9),
        'colsample_bylevel': params.get('colsample_bylevel', 0.7),
        'gamma': params.get('gamma', 7),
        'reg_alpha': params.get('reg_alpha', 10),
        'reg_lambda': params.get('reg_lambda', 10)
    }

    xgb_clf = XGBClassifier(**stepwise_del_params)
    xgb_clf.fit(dev_data[fea_names], dev_data[target])

    pred_test = xgb_clf.predict_proba(nodev_data[fea_names])[:, 1]
    pred_train = xgb_clf.predict_proba(dev_data[fea_names])[:, 1]

    test_ks = get_ks(nodev_data[target], pred_test)
    train_ks = get_ks(dev_data[target], pred_train)
    log.info('test_ks is : {}'.format(test_ks))
    log.info('train_ks is : {}'.format(train_ks))

    train_number, oldks, del_list = 0, test_ks, list()
    log.info('train_number: {}, test_ks: {}'.format(train_number, test_ks))

    # while True:
    #     flag = True
    #     for fea_name in tqdm(fea_names):
    #         print('变量{}进行逐步：'.format(fea_name))
    #         names = [fea for fea in fea_names if fea_name != fea]
    #         print('变量names is：', names)
    #         xgb_clf.fit(dev_data[names], dev_data[target])
    #         train_number += 1
    #         pred_test = xgb_clf.predict_proba(nodev_data[names])[:, 1]
    #         test_ks = get_ks(nodev_data[target], pred_test)
    #         if test_ks >= oldks:
    #             oldks = test_ks
    #             flag = False
    #             del_list.append(fea_name)
    #             log.info(
    #                 '等于或优于之前结果 train_number: {}, test_ks: {} by feature: {}'.format(train_number, test_ks, fea_name))
    #             fea_names = names
    #     if flag:
    #         print('=====================又重新逐步==========')
    #         break
    #     log.info("结束逐步删除变量 train_number: %s, test_ks: %s del_list: %s" % (train_number, oldks, del_list))
    #     print('oldks is ：',oldks)
    #     print('fea_names is : ',fea_names)

    for fea_name in tqdm(fea_names):
        names = [fea for fea in fea_names if fea_name != fea]
        xgb_clf.fit(dev_data[names], dev_data[target])
        train_number += 1
        pred_test = xgb_clf.predict_proba(nodev_data[names])[:, 1]
        test_ks = get_ks(nodev_data[target], pred_test)
        if test_ks >= oldks:
            oldks = test_ks
            del_list.append(fea_name)
            log.info(
                '等于或优于之前结果 train_number: {}, test_ks: {} by feature: {}'.format(train_number, test_ks, fea_name))
            fea_names = names
    log.info("结束逐步删除变量 train_number: %s, test_ks: %s del_list: %s" % (train_number, oldks, del_list))

    ########################
    log.info('逐步剔除的变量个数: {}'.format(del_list))
    fea_names = [i for i in fea_names if i not in del_list]
    # fea_names = list(set(fea_names) - set(del_list))
    log.info('逐步保留的变量个数: {}'.format(len(fea_names)))
    log.info('*' * 50 + '逐步筛选变量' + '*' * 50)

    return del_list, fea_names



#==============================================================================
# File: function.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8




import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import os 

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)





# 获取当前目录下的CSV文件名
def get_filename(file_dir):
   #创建一个空列表，存储当前目录下的CSV文件全称
   file_name = []
   #将当前目录下的所有文件名称读取进来
   for root, dirs, files in os.walk(file_dir):
       # 判断是否为CSV文件，如果是则存储到列表中
       for file in files:
           if file[-4:] == '.csv':
               file_name.append(file)   
   
#     files = os.listdir()        
#     for j in files:
#     #判断是否为CSV文件，如果是则存储到列表中
#     if os.path.splitext(j)[1] == '.csv':
#         file_name.append(j)
       
   return file_name





def process_data(df_base, cols_left, df_three, cols_right, needcols=[], suffix=None):
    df1 = pd.merge(df_base[cols_left], df_three[cols_right], how='inner', on='id_no_des')
    df1['apply_date'] = pd.to_datetime(df1['apply_date'],format='%Y-%m-%d')
    df1['create_time'] = pd.to_datetime(df1['create_time'].str[0:10], format='%Y-%m-%d')
    df1['days'] = df1['apply_date'] - df1['create_time']
    df1['days'] = df1['days'].dt.days
    df1['order_no_is_equal'] = [*map(lambda t1,t2: 1 if t2==t1 else 0,df1['order_no_x'],df1['order_no_y'])]
    df1['order_no_is_equal'].value_counts(dropna=False)
    
    # 拆分数据
    print('-----------拆分数据------------------')
    df1_part1 = df1.query("order_no_is_equal==1")
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)
    # 去重
    df1_part1 = df1_part1.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)
    
    df1_part2 = df1.query("order_no_is_equal==0")
    df1_part2 = df1_part2.query("days<=30 & days>=0")
#     print(df1_part2['order_no_x'].nunique(), df1_part2.shape)
    # 去重
    df1_part2 = df1_part2.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print( df1_part2['order_no_x'].nunique(), df1_part2.shape)
    
    # 合并数据
    print('-----------合并数据------------------')
    df1_new = pd.concat([df1_part1, df1_part2], axis=0)
#     print( df1_new['order_no_x'].nunique(), df1_new.shape)
    # 去重
    df1_new = df1_new.sort_values(by=['order_no_x','order_no_is_equal','create_time'], ascending=False)
    df1_new = df1_new.drop_duplicates(subset=['order_no_x'],keep='first')
    print("去重后流水订单号数：", df1_new['order_no_x'].nunique())
    print("成功匹配的三方数据：", df1_new.shape)
    
    # 重新命名字段
    usecols = ['order_no_x','apply_date','order_no_y','create_time','order_no_is_equal'] + needcols
    df1_new = df1_new[usecols]

    cols = []
    for col in usecols[1:]:
        var = col +  '_' + suffix
        cols.append(var)
    
    cols = ['order_no'] + cols
    df1_new.columns = cols
    df1_new.to_csv(r'D:\liuyedao\转转渠道\mid_result\auth_{}_{}.csv'.format(suffix,str(datetime.today())[:10].replace('-','')), index=False)
    
    # 返回需要的数据
    return_keys = ['order_no'] + cols[5:]
    df2 = df1_new[return_keys]
    gc.collect()
    
    return df2
  





def score_distribute(data, col, target='target'):    
    total = data.groupby(col)[target].count()
    bad = data.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['bad_rate_cum'] = regroup['bad'].cumsum()/regroup['total'].cumsum()
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['good_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['total_pct_cum'] = regroup['total_pct'].cumsum()
    regroup['ks'] = regroup['bad_pct_cum'] - regroup['good_pct_cum']
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift_cum'] = regroup['bad_rate_cum']/data[target].mean()
    usecols = ['varsname','bins','bad','good','total','ks', 'bad_rate','bad_rate_cum','lift_cum','total_pct_cum'
            ,'total_pct', 'bad_pct','good_pct','bad_pct_cum','good_pct_cum']
    return_regroup = regroup[usecols]
    return_regroup = return_regroup.reset_index(drop=True)

    return return_regroup





def regroup(data_bins, col, target='target'):    
    total = data_bins.groupby(col)[target].count()
    bad = data_bins.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['ks_bins'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
    regroup['ks'] = regroup['ks_bins'].max()
    regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv'] = regroup['iv_bins'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift'] = regroup['bad_rate']/data_bins[target].mean()
    usecols = ['varsname','bins','iv','ks','bad','good','total','bad_rate','total_pct','lift'
            ,'bad_pct','good_pct','bad_pct_cum','goo_pct_cum','ks_bins','iv_bins']
    return_regroup = regroup[usecols]

    return return_regroup


# In[1]:


# 离散变量分组
def bins_group(data, flag, col):
    df = data[[flag, col]]
    df['bins'] = df[col]
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    
    return regroup

# 保留特殊值不参与分箱
def regroup_special_merge(regroup_special, regroup_normal):
    while regroup_special['bad'].min()==0:
        id_min = regroup_special['bad'].idxmin()
        regroup_special.loc[id_min,'bad'] = 1
    while regroup_special['good'].min()==0:
        id_min = regroup_special['good'].idxmin()
        regroup_special.loc[id_min,'good']=  1
    regroup_normal.index.name = regroup_special.index.name
    return_regroup = pd.concat([regroup_special,regroup_normal],axis=0)
    
    return return_regroup
    
# 删除当前索引值所在行的后一行
def DelIndexPlus1(np_regroup, index_value):
    np_regroup[index_value,1] = np_regroup[index_value,1] + np_regroup[index_value+1,1]
    np_regroup[index_value,2] = np_regroup[index_value,2] + np_regroup[index_value+1,2]
    np_regroup[index_value,0] = np_regroup[index_value+1,0]
    np_regroup = np.delete(np_regroup, index_value+1, axis=0)
    
    return np_regroup
 
# 删除当前索引值所在行
def DelIndex(np_regroup, index_value):
    np_regroup[index_value-1,1] = np_regroup[index_value,1] + np_regroup[index_value-1,1]
    np_regroup[index_value-1,2] = np_regroup[index_value,2] + np_regroup[index_value-1,2]
    np_regroup[index_value-1,0] = np_regroup[index_value,0]
    np_regroup = np.delete(np_regroup, index_value, axis=0)
    
    return np_regroup 
    
# 删除/合并客户数为0的箱子
def MergeZero(np_regroup):
    #合并好坏客户数连续都为0的箱子
    i = 0
    while i<=np_regroup.shape[0]-2:
        if (np_regroup[i,1]==0 and np_regroup[i+1,1]==0) or (np_regroup[i,2]==0 and np_regroup[i+1,2]==0):
            np_regroup = DelIndexPlus1(np_regroup,i)
        i = i+1
        
    #合并坏客户数为0的箱子
    while True:
        if all(np_regroup[:,1]>0) or np_regroup.shape[0]==2:
            break
        bad_zero_index = np.argwhere(np_regroup[:,1]==0)[0][0]
        if bad_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
        elif bad_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, bad_zero_index)
        else:
            if np_regroup[bad_zero_index-1,2]/np_regroup[bad_zero_index-1,1]>=np_regroup[bad_zero_index+1,2]/np_regroup[bad_zero_index+1,1]:
                np_regroup = DelIndex(np_regroup, bad_zero_index)
            else:
                np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
    #合并好客户数为0的箱子
    while True:
        if all(np_regroup[:,2]>0) or np_regroup.shape[0]==2:
            break
        good_zero_index = np.argwhere(np_regroup[:,2]==0)[0][0]
        if good_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
        elif good_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, good_zero_index)
        else:
            if np_regroup[good_zero_index-1,2]/np_regroup[good_zero_index-1,1]>=np_regroup[good_zero_index+1,2]/np_regroup[good_zero_index+1,1]:
                np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
            else:
                np_regroup = DelIndex(np_regroup, good_zero_index)
                
    return np_regroup
    
# 箱子的单调性
def MonTone(np_regroup):
    while True:
        if np_regroup.shape[0]==2:
            break
        GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
        GoodBadRateMonetone = [GoodBadRate[i]<GoodBadRate[i+1] for i in range(np_regroup.shape[0]-1)]
        #确定是否单调
        if_Montone = len(set(GoodBadRateMonetone))
        #判断跳出循环
        if if_Montone==1:
            break
        else:
            WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
            Montone_index = WoeDiffMin.index(min(WoeDiffMin))
            np_regroup = DelIndexPlus1(np_regroup, Montone_index)
            
    return np_regroup
    
#箱子最小占比
def MinPct(np_regroup):
    while True:
        bins_pct = [(np_regroup[i,1]+np_regroup[i,2])/np_regroup.sum() for i in range(np_regroup.shape[0])]
        min_pct = min(bins_pct)
        if min_pct>=0.02 or len(bins_pct)==2:
            break
        else:
            min_pct_index = bins_pct.index(min(bins_pct))
            if min_pct_index==0:
                np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
            elif min_pct_index == np_regroup.shape[0]-1:
                np_regroup = DelIndex(np_regroup, min_pct_index)
            else:
                GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
                WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
                if WoeDiffMin[min_pct_index-1]>=WoeDiffMin[min_pct_index]:
                    np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
                else:
                    np_regroup = DelIndex(np_regroup, min_pct_index)
    return np_regroup

    
# 连续变量分箱主函数
def ContinueVarBins(data, col, flag='target', cutbins=[]):
    df = data[[flag, col]]
    df = df[~df[col].isnull()].reset_index(drop=True)
    df['bins'] = pd.cut(df[col], cutbins, duplicates='drop', right=False, precision=4)
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup.drop(['total'], axis=1, inplace=True)
    np_regroup = np.array(regroup)
    np_regroup = MergeZero(np_regroup)
    np_regroup = MinPct(np_regroup)
    np_regroup = MonTone(np_regroup)
    regroup = pd.DataFrame(index=np.arange(np_regroup.shape[0]))
    regroup['bins'] = np_regroup[:,0]
    regroup['total'] = np_regroup[:,1] + np_regroup[:,2]
    regroup['bad'] = np_regroup[:,1]
    regroup['good'] = np_regroup[:,2]
    cutoffpoints = list(np_regroup[:,0])
    cutoffpoints = [float('-inf')] + cutoffpoints
    # 最大值分割点，转换最小值分割点
    df['bins_new'] = pd.cut(df[col], cutoffpoints, duplicates='drop', right=True, precision=4)
    tmp = pd.DataFrame()
    tmp['bins'] = df.groupby(['bins_new'])[col].min()
    cutoffpoints = list(tmp['bins'])  
    
    return cutoffpoints


def CalWoeIv(data_bins, col, target='target'):  
    total = data_bins.groupby(col)[target].count()
    bad = data_bins.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['ks_bins'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
    regroup['ks'] = regroup['ks_bins'].max()
    regroup['woe'] = np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv'] = regroup['iv_bins'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift'] = regroup['bad_rate']/data_bins[target].mean()
    usecols = ['varsname','bins','iv','ks','bad','good','total','bad_rate','total_pct','lift','woe']
    return_regroup = regroup[usecols]

    return return_regroup


# 变量间相关性
def CorrSelect(df, iv_df, exclude_list=[], threshold=0.7):
    X = [i for i in df.columns if i not in exclude_list]
    data = df[X]
    df_corr = data.corr()
    droped_list = []
    while True:
        dict_cols = dict(zip(range(df_corr.shape[1]), list(df_corr.columns)))
        np_corr = abs(np.array(df_corr))
        np.fill_diagonal(np_corr, 0)
        if np.amax(np_corr) < threshold:
            break
        index1, index2 = np.unravel_index(np_corr.argmax(), np_corr.shape)
        x1 = dict_cols[index1]
        x2 = dict_cols[index2]
        if iv_df.loc[x1,'iv']>=iv_df.loc[x2,'iv']:
            droped_list.append(x2)
            df_corr.drop(index=[x2], inplace=True)
            df_corr.drop(columns=[x2], inplace=True)
        else:
            droped_list.append(x1)
            df_corr.drop(index=[x1], inplace=True)
            df_corr.drop(columns=[x1], inplace=True)
    
    return droped_list
 

def sklearn_vif(exogs, data):
    from sklearn.linear_model import LinearRegression
    # initialize dictionaries
    vif_dict, tolerance_dict = {}, {}

    # form input data for each exogenous variable
    for exog in exogs:
        not_exog = [i for i in exogs if i != exog]
        X, y = data[not_exog], data[exog]

        # extract r-squared from the fit
        r_squared = LinearRegression().fit(X, y).score(X, y)

        # calculate VIF
        vif = 1/(1 - r_squared)
        vif_dict[exog] = vif

        # calculate tolerance
        tolerance = 1 - r_squared
        tolerance_dict[exog] = tolerance

    # return VIF DataFrame
    df_vif = pd.DataFrame({'VIF': vif_dict, 'Tolerance': tolerance_dict})

    return df_vif


# In[1]:


import matplotlib.pyplot as plt
def lr(X_train,X_test,y_train,y_test):
    from sklearn.linear_model import LogisticRegression
    from scipy import stats
    from toad.metrics import KS,AUC 
    # 模型训练和p_values查看 只能是线性回归
    clf = LogisticRegression(C=1e8).fit(X_train, y_train)
    params = np.append(clf.intercept_,clf.coef_)
    
    new_X_train = pd.DataFrame({"Constant":np.ones(len(X_train))}).join(X_train.reset_index(drop=True))
    predictions = clf.predict(X_train)
    MSE = (sum((y_train-predictions)**2))/(len(new_X_train)-len(new_X_train.columns))

    var_b = MSE*(np.linalg.inv(np.dot(new_X_train.T,new_X_train)).diagonal())
    sd_b = np.sqrt(var_b)
    ts_b = params/ sd_b

    p_values =[2*(1-stats.t.cdf(np.abs(i),(len(new_X_train)-1))) for i in ts_b]

    sd_b = np.round(sd_b,3)
    ts_b = np.round(ts_b,3)
    p_values = np.round(p_values,3)
    params = np.round(params,4)

    myDF3 = pd.DataFrame()
    myDF3["Vars"],myDF3["Coefficients"],myDF3["Standard Errors"],myDF3["t values"],myDF3["P values"] = [new_X_train.columns,params,sd_b,ts_b,p_values]
    print(myDF3)
    
    # 预测值
    pred_train = clf.predict_proba(X_train)[:,1]
    pred_test = clf.predict_proba(X_test)[:,1]
    
    # 训练集KS/AUC
    print('-------------训练集结果--------------------')
    print('train AUC: ', AUC(pred_train, y_train))
    print('train KS: ', KS(pred_train, y_train))
    
    # 测试集KS/AUC
    print('-------------测试集结果--------------------')
    print('test AUC: ', AUC(pred_test, y_test))
    print('test KS: ', KS(pred_test, y_test))
    
    print('-------------------------分割线--------------------------')
    # 模型评估
    train_AUC = roc_auc_score(y_train, pred_train)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,pred_train)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_AUC = roc_auc_score(y_test, pred_test)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,pred_test)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)

    return (myDF3, params, clf)

# 单次训练
def LR_model(X_train,X_test,y_train,y_test):
    clf = sm.Logit(y_train, sm.add_constant(X_train)).fit()
    print(clf.summary())
    # 模型评估
    train_y_pred = clf.predict(sm.add_constant(X_train))
    train_AUC = roc_auc_score(y_train, train_y_pred)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,train_y_pred)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_y_pred = clf.predict(sm.add_constant(X_test))
    test_AUC = roc_auc_score(y_test,test_y_pred)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,test_y_pred)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)
    
    return clf





def cal_auth(df_auth, target='target', channel=None):
    if channel:
        df_auth = df_auth.query("channel_id==@channel")
        
    total_lend = df_auth.groupby('bins')['order_no',target].count()
    tg_jj = df_auth.groupby(['bins', 'auth_status'])['order_no'].count().unstack()
    lend =  df_auth.groupby(['bins', target])['order_no'].count().unstack()
    result = pd.concat([total_lend, tg_jj, lend], axis=1)
    result.columns = ['授信申请','放款人数','通过人数','拒绝人数','好','灰','坏']
    
    result_sum = pd.DataFrame(result.sum(axis=0),columns=['total']).T
    
    for col in result.columns:
        result['{}占比'.format(col)] = result[col]/result[col].sum()
        result['{}累计占比'.format(col)] = result['{}占比'.format(col)].cumsum()
        
    result['坏客率'] = result['坏']/(result['好']+result['坏'])
    result['通过率'] = result['通过人数']/result['授信申请']
    result = pd.concat([result, result_sum],axis=0)
    
    cols = ['授信申请','授信申请占比','授信申请累计占比','通过人数','通过率','拒绝人数','拒绝人数占比','拒绝人数累计占比',
            '通过人数占比','通过人数累计占比','放款人数','放款人数占比','放款人数累计占比','好','好占比','好累计占比',
            '坏','坏占比','坏累计占比','坏客率','灰','灰占比','灰累计占比']
    result = result[cols]
    
    return result





# observe_date：观察点日期

# ---------------------逾期类变量------------------
def his_max_ovdue_day(data, observe_date):
    """
    历史最大逾期天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len (data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date] 
        if len (data) >0 : 
            # 计算每一期的历史最大逾期天数
            ovdue_day = max(data[['repay_date', 'period_settle_date']].apply(
                lambda x:(observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days, axis=1).max(),0)
        else: 
            #未到还款日期-9998 
            ovdue_day = -9998 
    else : 
        # 无借据-9999 
        ovdue_day = -9999 
    
    return ovdue_day
    
def cur_ovdue_day(data, observe_date): 
    """
    当前逾期天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len (data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date] 
        if len (data) >0 : 
            # 计算每一期的当前逾期天数，取max
            ovdue_day = data[['repay_date', 'period_settle_date']].apply(
                lambda x:(observe_date-x[0]).days+1 if x[1]> observe_date else 0, axis=1).max()
        else: 
            #未到还款日期-9998 
            ovdue_day = -9998 
    else : 
        # 无借据-9999 
        ovdue_day = -9999 
    
    return ovdue_day
 
def cur_ovdue_period(data, observe_date):
    """
    当前逾期期数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期逾期状态，逾期取1，不逾期取0，求sum
            ovdue_period = data[['repay_date','period_settle_date']].apply(
            lambda x: 1 if x[1]>observe_date else 0 ,axis=1).sum()
        else :
            # 未到还款日期
            ovdue_period = -9998
    else :
        # 无借据 -9999
        ovdue_period = -9999
    
    return ovdue_period
            
def cur_ovdue_loan(data, observe_date):
    """
    当前逾期借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期逾期状态，逾期取order_no，不逾期取np.nan，求借据的唯一数
            ovdue_loan = data[['repay_date','period_settle_date','order_no']].apply(
            lambda x: x[2] if x[1]>observe_date else np.nan ,axis=1).nunique()
        else :
            # 未到还款日期
            ovdue_loan = -9998
    else :
        # 无借据 -9999
        ovdue_loan = -9999
    
    return ovdue_loan
            
def cur_ovdue_prin(data, observe_date):
    """
    当前逾期本金
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期逾期状态，逾期取principal(本期应还本金)，不逾期取0，求sum
            cur_ovdue_prin = data[['repay_date','period_settle_date','principal']].apply(
            lambda x: x[2] if x[1]>observe_date else 0 ,axis=1).sum()
        else :
            # 未到还款日期
            cur_ovdue_prin = -9998
    else :
        # 无借据 -9999
        cur_ovdue_prin = -9999
    
    return cur_ovdue_prin
           
def maxovdue_nearly_30days(data, observe_date, ndays=30):
    """
    近N天最大逾期天数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天的最大逾期天数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，求Max
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date - data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return max(ovdue_days_list)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 -9999
        return -9999
           
def maxovdue_nearly_60days(data, observe_date, ndays=60):
    """
    近N天最大逾期天数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天的最大逾期天数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，求Max
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date - data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return max(ovdue_days_list)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 -9999
        return -9999
           
def maxovdue_nearly_90days(data, observe_date, ndays=90):
    """
    近N天最大逾期天数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天的最大逾期天数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，求Max
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date - data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return max(ovdue_days_list)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 -9999
        return -9999
           
def maxovdue_nearly_180days(data, observe_date, ndays=180):
    """
    近N天最大逾期天数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天的最大逾期天数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，求Max
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date - data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return max(ovdue_days_list)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 -9999
        return -9999

           
def ovdue1_cnt_nearly_30days(data, observe_date, ndays=30, ovdue_days=1):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue1_cnt_nearly_90days(data, observe_date, ndays=90, ovdue_days=1):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue1_cnt_nearly_180days(data, observe_date, ndays=180, ovdue_days=1):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue1_cnt_his(data, observe_date, ndays=720, ovdue_days=1):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
            
def ovdue3_cnt_nearly_30days(data, observe_date, ndays=30, ovdue_days=3):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue3_cnt_nearly_90days(data, observe_date, ndays=90, ovdue_days=3):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue3_cnt_nearly_180days(data, observe_date, ndays=180, ovdue_days=3):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue3_cnt_his(data, observe_date, ndays=720, ovdue_days=3):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue7_cnt_nearly_30days(data, observe_date, ndays=30, ovdue_days=7):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue7_cnt_nearly_90days(data, observe_date, ndays=90, ovdue_days=7):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue7_cnt_nearly_180days(data, observe_date, ndays=180, ovdue_days=7):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue7_cnt_his(data, observe_date, ndays=720, ovdue_days=7):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue30_cnt_nearly_30days(data, observe_date, ndays=30, ovdue_days=30):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue30_cnt_nearly_90days(data, observe_date, ndays=90, ovdue_days=30):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue30_cnt_nearly_180days(data, observe_date, ndays=180, ovdue_days=30):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
           
def ovdue30_cnt_his(data, observe_date, ndays=720, ovdue_days=30):
    """
    近N天最大逾期1/3/7/30+天次数：ndays 30/60/90/180
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) > 0 : 
        # 计算近N天逾期1/3/7/30+天次数
        ovdue_days_list = [] # 统计每一期的近N天的逾期天数，判断逾期天数>=ovdue_days的次数
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            for i in range(len(data)):
                if (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]>observe_date):
                    #实还日期>应还日期 且 实还日期>观察点日期，则逾期天数=观察日期-应还日期+1
                    ovdue_days_list.append((observe_date-data['repay_date'].iloc[i]).days + 1)
                elif (data['period_settle_date'].iloc[i]>data['repay_date'].iloc[i]
                    and data['period_settle_date'].iloc[i]<=observe_date
                    and data['period_settle_date'].iloc[i]>observe_date + relativedelta(days=-ndays)):
                    #实还日期>应还日期 且 实还日期<=观察点日期 且实还日期>观察日期-N，则逾期天数=实还日期-应还日期
                    ovdue_days_list.append((data['period_settle_date'].iloc[i]-data['repay_date'].iloc[i]).days)
                else:
                    ovdue_days_list.append(0)
                
            return np.sum(np.array(ovdue_days_list)>ovdue_days)
        else :
            # 未到还款日期,默认值-9998
            return -9998
    else :
        # 无借据 默认值-9999
        return -9999
       
        
def first_ovdue_periods(data, observe_date):
    """
    首次逾期的期数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # 取历史逾期>=1最早的repay_date对应的period
                first_ovdue_period = data[data['ever_ovdue_day']>=1].sort_values(by=['repay_date','period'])['period'].iloc[0]
            else:
                # 无逾期，默认值-1
                first_ovdue_period = -1
        else :
            # 未到还款日期
            first_ovdue_period = -9998
    else :
        # 无借据 -9999
        first_ovdue_period = -9999
    
    return first_ovdue_period
  
  
def first_ovdue_days(data, observe_date):
    """
    首次逾期的天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # 取历史逾期>=1最早的repay_date对应的ever_ovdue_day
                first_ovdue_day = data[data['ever_ovdue_day']>=1].sort_values(by=['repay_date','period'])['ever_ovdue_day'].iloc[0]
            else:
                # 无逾期，默认值-1
                first_ovdue_day = -1
        else :
            # 未到还款日期
            first_ovdue_day = -9998
    else :
        # 无借据 -9999
        first_ovdue_day = -9999
    
    return first_ovdue_day
  
  
def first_ovdue1_days_to_cur(data, observe_date, ovdue_days=1):
    """
    最早一次逾期1+/30+距今天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # observe_date-取历史逾期>=ovdue_days最早的repay_date的天数差
                first_ovdue_day_fromnow = (observe_date-data[data['ever_ovdue_day']>=ovdue_days]['repay_date'].min()).days
            else:
                # 无逾期，默认值-1
                first_ovdue_day_fromnow = -1
        else :
            # 未到还款日期
            first_ovdue_day_fromnow = -9998
    else :
        # 无借据 -9999
        first_ovdue_day_fromnow = -9999
    
    return first_ovdue_day_fromnow
  
  
def last_ovdue1_days_to_cur(data, observe_date, ovdue_days=1):
    """
    最近一次逾期1+/30+距今天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # observe_date-取历史逾期>=ovdue_days最近的repay_date的天数差
                last_ovdue_day_fromnow = (observe_date-data[data['ever_ovdue_day']>=ovdue_days]['repay_date'].max()).days
            else:
                # 无逾期，默认值-1
                last_ovdue_day_fromnow = -1
        else :
            # 未到还款日期
            last_ovdue_day_fromnow = -9998
    else :
        # 无借据 -9999
        last_ovdue_day_fromnow = -9999
    
    return last_ovdue_day_fromnow
  
def first_ovdue30_days_to_cur(data, observe_date, ovdue_days=30):
    """
    最早一次逾期1+/30+距今天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # observe_date-取历史逾期>=ovdue_days最早的repay_date的天数差
                first_ovdue_day_fromnow = (observe_date-data[data['ever_ovdue_day']>=ovdue_days]['repay_date'].min()).days
            else:
                # 无逾期，默认值-1
                first_ovdue_day_fromnow = -1
        else :
            # 未到还款日期
            first_ovdue_day_fromnow = -9998
    else :
        # 无借据 -9999
        first_ovdue_day_fromnow = -9999
    
    return first_ovdue_day_fromnow
  
  
def last_ovdue30_days_to_cur(data, observe_date, ovdue_days=30):
    """
    最近一次逾期1+/30+距今天数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 剔除未到还款日期的还款计划
        data = data[data['repay_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            data['ever_ovdue_day'] = data[['repay_date','period_settle_date']].apply(
            lambda x: (observe_date-x[0]).days+1 if x[1]>observe_date else (x[1]-x[0]).days ,axis=1)
            
            if len(data[data['ever_ovdue_day']>=1])>=1:
                # observe_date-取历史逾期>=ovdue_days最近的repay_date的天数差
                last_ovdue_day_fromnow = (observe_date-data[data['ever_ovdue_day']>=ovdue_days]['repay_date'].max()).days
            else:
                # 无逾期，默认值-1
                last_ovdue_day_fromnow = -1
        else :
            # 未到还款日期
            last_ovdue_day_fromnow = -9998
    else :
        # 无借据 -9999
        last_ovdue_day_fromnow = -9999
    
    return last_ovdue_day_fromnow
 
    
# ---------------------还款行为类变量------------------

def normal_repay_cnt_nearly_60days(data, observe_date, ndays=60):
    """
    近N天正常还款次数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        cnt = len(data[(data['period_settle_date']<=data['repay_date'])&(data['period_settle_date']<=observe_date)
        &(data['period_settle_date']>observe_date+relativedelta(days=-ndays))])
        
        return cnt

    else :
        # 无借据 -9999
        return -9999

def normal_repay_cnt_history(data, observe_date, ndays=720):
    """
    近N天正常还款次数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        cnt = len(data[(data['period_settle_date']<=data['repay_date'])&(data['period_settle_date']<=observe_date)
        &(data['period_settle_date']>observe_date+relativedelta(days=-ndays))])
        
        return cnt

    else :
        # 无借据 -9999
        return -9999
  
def cur_balance_observe(data, observe_date):
    """
    观察日期对应当前剩余本金
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 总放款金额- 截止观察时间已还本金
        current_balance = (data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
                           - data[data['period_settle_date']<=observe_date]['already_repaid_principal'].sum())
        
    else :
        # 无借据 -9999
        current_balance = -9999
        
    return current_balance
    
 
def cur_prin_repay_pct(data, observe_date):
    """
    当前实际还款率
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 总放款金额- 截止观察时间已还本金
        already_repaid_principal =  data[data['period_settle_date']<=observe_date]['already_repaid_principal'].sum()
        loan_amount = data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
        cur_prin_repay_pct = already_repaid_principal/loan_amount
    else :
        # 无借据 -9999
        cur_prin_repay_pct = -9999
        
    return cur_prin_repay_pct
    

def cur_nopay_period(data, observe_date):
    """
    当前剩余期数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        settle_period = len(data[data['period_settle_date']<=observe_date])
        # 总期数
        res = data.groupby(by=['order_no','total_periods'])['period'].max().reset_index()
        all_period = res[['total_periods','period']].max(axis=1).sum()
        
        # 剩余期数
        norepay_period = all_period - settle_period
    else :
        # 无借据 -9999
        norepay_period =  -9999
        
    return norepay_period
    

def cur_nopay_period_pct(data, observe_date):
    """
    当前剩余期数占总放款期数的占比
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        settle_period = len(data[data['period_settle_date']<=observe_date])
        # 总期数
        res = data.groupby(by=['order_no','total_periods'])['period'].max().reset_index()
        all_period = res[['total_periods','period']].max(axis=1).sum()
        
        # 剩余期数占比
        norepay_period_pct = (all_period - settle_period)/all_period
    else :
        # 无借据 -9999
        norepay_period_pct =  -9999
        
    return norepay_period_pct
    

def cur_nosettle_max_period(data, observe_date):
    """
    当前未结清借据最大期限
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 :
        
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 未结清
        if len(res[res['period_settle_date']>observe_date])>=1:
            cur_nosettle_max_period = res[res['period_settle_date']>observe_date][['total_periods','period']].max(axis=1).max(axis=0)
        else:
            # 无未结清, 默认值-1
            cur_nosettle_max_period = -1         
    else :
        # 无借据 -9999
        cur_nosettle_max_period =  -9999
        
    return cur_nosettle_max_period
    

def nearly_180days_advsettle_loan_pct(data, observe_date, ndays=180):
    """
    近N天内提前结清的借据数/总借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 :
        
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 实还日期<应还日期为提前结清，再增加近N天判断
        # 近N天提前结清借据数
        advsettle_loan1 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        advsettle_loan2 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        if advsettle_loan1==len(res):
            # 近N天前全部结清 默认值-1
            nearly_advsettle_loan_pct = -1
        else:
            nearly_advsettle_loan_pct = advsettle_loan2/len(res)
    else :
        # 无借据 -9999
        nearly_advsettle_loan_pct =  -9999
        
    return nearly_advsettle_loan_pct

def his_advsettle_loan_pct(data, observe_date, ndays=720):
    """
    近N天内提前结清的借据数/总借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 :
        
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 实还日期<应还日期为提前结清，再增加近N天判断
        # 近N天提前结清借据数
        advsettle_loan1 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        advsettle_loan2 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        if advsettle_loan1==len(res):
            # 近N天前全部结清 默认值-1
            nearly_advsettle_loan_pct = -1
        else:
            nearly_advsettle_loan_pct = advsettle_loan2/len(res)
    else :
        # 无借据 -9999
        nearly_advsettle_loan_pct =  -9999
        
    return nearly_advsettle_loan_pct


def nearly_180days_advsettle_loan(data, observe_date, ndays=180):
    """
    近N天内提前结清的借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 :
        # 提前结清
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 实还日期<应还日期为提前结清，再增加近N天判断
        # 近N天提前结清借据数
        advsettle_loan1 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        advsettle_loan2 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        if advsettle_loan1==len(res):
            # 近N天前全部结清 默认值-1
            nearly_advsettle_loan = -1
        else:
            nearly_advsettle_loan = advsettle_loan2
    else :
        # 无借据 -9999
        nearly_advsettle_loan =  -9999
        
    return nearly_advsettle_loan

def his_advsettle_loan(data, observe_date, ndays=720):
    """
    近N天内提前结清的借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 :
        # 提前结清
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 实还日期<应还日期为提前结清，再增加近N天判断
        # 近N天提前结清借据数
        advsettle_loan1 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']<=observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        advsettle_loan2 = len(set(
        list(res[(res['period_settle_date']<res['repay_date'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        
        +list(res[(res['total_periods']<res['period'])&(res['period_settle_date']<=observe_date)
        &(res['period_settle_date']>observe_date+relativedelta(days=-ndays))]['order_no'].unique())
        ))
        
        if advsettle_loan1==len(res):
            # 近N天前全部结清 默认值-1
            nearly_advsettle_loan = -1
        else:
            nearly_advsettle_loan = advsettle_loan2
    else :
        # 无借据 -9999
        nearly_advsettle_loan =  -9999
        
    return nearly_advsettle_loan
      

def payment_pct_t30(data, observe_date, ndays=30):
    """
    T-X实际还款率：分母为当前时点放款总金额
    """
    observe_date_new = observe_date + relativedelta(days=-ndays)
    
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 本期已还本金
        prin_payed = data[data['period_settle_date']<=observe_date_new]['already_repaid_principal'].sum()
        
        loan_amount = data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
        # 截止T-x观察时间已还本金/T总放款金额
        ever_prin_repay_pct = prin_payed/loan_amount
    else :
        # 无借据 -9999
        ever_prin_repay_pct = -9999
        
    return ever_prin_repay_pct
   
def last_repay_months_to_cur(data, observe_date):
    """
    最后一次还款距现在月份数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 选择已结清的的还款计划
        data = data[data['period_settle_date']<=observe_date]
        if len(data) > 0 : 
            # 计算每期的历史逾期天数
            max_period_settle_date = data['period_settle_date'].max()
            last_repay_months_fromnow = (12*(observe_date.year-max_period_settle_date.year)
                                            + observe_date.month-max_period_settle_date.month)
        else :
            # 无还款
            last_repay_months_fromnow = -1
    else :
        # 无借据 -9999
        last_repay_months_fromnow = -9999
    
    return last_repay_months_fromnow
  
    
def his_max_repay_bymonth(data, observe_date):
    """
    历史单月最大还款金额
    """
    data['settle_mth'] = data['period_settle_date'].apply(lambda x: int(str(x)[0:7].replace('-','')))
    
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    if len(data) >0 : 
        # 选择已结清的的还款计划
        data = data[data['period_settle_date']<=observe_date]
        if len(data) > 0 : 
            # 历史单月最大还款金额
            his_max_repay_bymonth = data.groupby(['settle_mth'])['already_repaid_principal'].sum().max()           
        else :
            # 无还款
            his_max_repay_bymonth = -1
    else :
        # 无借据 -9999
        his_max_repay_bymonth = -9999
    
    return his_max_repay_bymonth
  
      
def last_repay_bymonth(data, observe_date):
    """
    最后一次还款月还款金额
    """
    data['settle_mth'] = data['period_settle_date'].apply(lambda x: int(str(x)[0:7].replace('-','')))
    
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    if len(data) >0 : 

        # 选择已结清的的还款计划
        data = data[data['period_settle_date']<=observe_date]
        if len(data) > 0 :
            # 最近一次还款月的月还款金额
            last_repay_bymonth = data.groupby(['settle_mth'])['already_repaid_principal'].sum().sort_index().iloc[-1]        
        else :
            # 无还款
            last_repay_bymonth = -1
    else :
        # 无借据 -9999
        last_repay_bymonth = -9999
    
    return last_repay_bymonth
   
def payment75pct_t30(data, observe_date, ndays=30):
    """
    T-x实际还款率：分母为T-x时点放款总金额
    """
    observe_date_new = observe_date + relativedelta(days=-ndays)
    
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date_new]

    if len(data) >0 : 

        if data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()>0.1:
            # 已还本金
            prin_payed = data[data['period_settle_date']<=observe_date_new]['already_repaid_principal'].sum()
            
            loan_amount = data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
            # 截止T-x观察时间已还本金/T-x总放款金额
            ever_prin_repay_pct2 = prin_payed/loan_amount
            if ever_prin_repay_pct2>0.75:
                ever_prin_repay_pct2 = 1
            else:
                ever_prin_repay_pct2 = 0
        else:
            # 分母为0 默认值为-1
            ever_prin_repay_pct2 = -1
    else :
        # 无借据 -9999
        ever_prin_repay_pct2 = -9999
        
    return ever_prin_repay_pct2

    
def bal60_bal0_ratio(data, observe_date, ndays=30):
    """
    T-x贷款余额/T-0贷款余额
    """
    observe_date_new = observe_date + relativedelta(days=-ndays)
    
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        # 截止T-x观察时间贷款余额/T-0贷款余额
        balx = (data[data['lending_time']<=observe_date_new][['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
                -data[data['period_settle_date']<=observe_date_new]['already_repaid_principal'].sum())
        
        bal0 = (data[data['lending_time']<=observe_date][['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
                -data[data['period_settle_date']<=observe_date]['already_repaid_principal'].sum())
        
        if bal0>0.1:
            cur_bal_repay_pct = balx/bal0
        else:
            # 分母为0 默认值为-1
            cur_bal_repay_pct = -1
    else :
        # 无借据 -9999
        cur_bal_repay_pct = -9999
        
    return cur_bal_repay_pct
    
# ----------------借款行为--------------      
def his_loan_cnt(data, observe_date):
    """
    历史成功借据数
    """   
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        return data['order_no'].nunique()
    else :
        # 无借据 -9999
        return -9999

def first_loan_months_fromnow(data, observe_date):
    """
    首笔借据距今月份数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 首笔借据距今月份数
        max_lending_time = data['lending_time'].min()
        first_loan_months_fromnow = (12*(observe_date.year-max_lending_time.year)
                                    + (observe_date.month-max_lending_time.month))
    else :
        # 无借据 -9999
        first_loan_months_fromnow = -9999
    
    return first_loan_months_fromnow
  
def all_settle_months_fromnow(data, observe_date):
    """
    客户结清距今月份数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 最后一期还款数据
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        
        if len(res[res['period_settle_date']<=observe_date])==len(res):
            # 判断是否所有借据结清
            last_period_settle_date = res[res['period_settle_date']<=observe_date]['period_settle_date'].max()
            all_settle_months_fromnow = (12*(observe_date.year - last_period_settle_date.year)
                                        +(observe_date.month - last_period_settle_date.month))
        else:
            # 客户未结清默认值-1
            all_settle_months_fromnow = -1
    else :
        # 无借据 -9999
        all_settle_months_fromnow = -9999
    
    return all_settle_months_fromnow
    
def his_loan_period_avg(data, observe_date):
    """
    历史成功借据平均期限
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]
    
    if len(data) >0 : 
        # 按照实际最大期限求avg
        loan_period = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        loan_period = loan_period[['total_periods','period']].max(axis=1).to_list()
        
        his_loan_period_avg = sum(loan_period)/len(loan_period)
    else :
        # 无借据 -9999
        his_loan_period_avg = -9999
    
    return his_loan_period_avg
    
def his_loan_amount_sum(data, observe_date):
    """
    历史成功借据金额之和
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        # 历史成功借据金额之和
        his_loan_amount_sum = data[['order_no','loan_amount']].drop_duplicates()['loan_amount'].sum()
    else :
        # 无借据 -9999
        his_loan_amount_sum = -9999
    
    return his_loan_amount_sum
     
def cur_order_nosettle_cnt(data, observe_date):
    """
    在贷借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        # 最后一期还款数据
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 在贷借据数 = 所有借据数 - 结清借据数
        cur_order_nosettle_cnt =len(res) - len(res[res['period_settle_date']<=observe_date])
    else :
        # 无借据 -9999
        cur_order_nosettle_cnt = -9999
    
    return cur_order_nosettle_cnt
     
def his_loan_settle_cnt(data, observe_date):
    """
    历史结清借据数
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        # 最后一期还款数据
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 结清借据数
        his_loan_settle_cnt = len(res[res['period_settle_date']<=observe_date])
    else :
        # 无借据 -9999
        his_loan_settle_cnt = -9999
    
    return his_loan_settle_cnt
     
def his_loan_settle_amt(data, observe_date):
    """
    历史结清借据本金
    """
    # 剔除观察时间点后的借据
    data = data[data['lending_time']<=observe_date]

    if len(data) >0 : 
        # 最后一期还款数据
        res = data.sort_values(by=['order_no','total_periods','period']).drop_duplicates(['order_no'],keep='last')
        # 结清借据金额
        his_loan_settle_amt = res[res['period_settle_date']<=observe_date]['loan_amount'].sum()
    else :
        # 无借据 -9999
        his_loan_settle_amt = -9999
    
    return his_loan_settle_amt
 
# -------------授信-------------    
def observe_credit_amt(data, observe_date, ndays=0):
    """
    最新授信额度
    """
    # 剔除观察时间点后的授信申请信息
    data = data[data['apply_date']<=observe_date+relativedelta(days=-ndays)]

    if len(data) >0 : 
        # 剔除授信失败的数据
        data = data[data['auth_status']==6]
        if len(data)>0:
            observe_credit_amt = data.sort_values(by=['apply_date'])['auth_credit_amount'].iloc[-1]
        else:
            # 没有通过授信 默认值-1
            observe_credit_amt = -1
    else :
        # 无授信 -9999
        observe_credit_amt = -9999
    
    return observe_credit_amt
    
def his_max_credit_amt(data, observe_date):
    """
    历史最高授信额度
    """
    # # 剔除观察时间点后的授信申请信息
    data = data[data['apply_date']<=observe_date]

    if len(data) >0 : 
        # 剔除授信失败的数据
        data = data[data['auth_status']==6]
        if len(data)>0:
            his_max_credit_amt = data['auth_credit_amount'].max()
        else:
            # 没有通过授信 默认值-1
            his_max_credit_amt = -1
    else :  
        # 无授信 -9999
        his_max_credit_amt = -9999
    
    return his_max_credit_amt





def object_explore(df, col, tablename):
    df1 = df[col].value_counts(dropna=False,sort=False)
    df2 = df[col].value_counts(dropna=False,sort=False,normalize=True)
    data = pd.concat([df1, df2], axis=1)
    data = data.reset_index()
    data.columns = ['bins','count','pct']
    data['tablename'] = tablename
    data['varsname'] = col
    data = data[['tablename','varsname','bins','count','pct']]
    
    return data

def numeric_explore(df, col, tablename):
    zero_pct = df[df[col]==0].shape[0]/df.shape[0]
    df1 = pd.DataFrame(df[col]).describe()
    df_numeric = df1.T
    cols = list(df_numeric.columns)
    df_numeric['zero_pct'] = str(zero_pct*100)[0:5] + '%'
    df_numeric['tablename'] = tablename
    df_numeric['varsname'] = col
    df_numeric['total'] = df.shape[0]
    df_numeric = df_numeric[['tablename','varsname','total','zero_pct']+cols]
    
    return df_numeric

def datetime_explore(df, col, tablename):
    total = df.shape[0]
    df_datetime = pd.DataFrame(columns=['tablename','varsname', 'count','min_date','max_date','null','null_pct'],index=[col])
    df_datetime.loc[col,'tablename']=tablename
    df_datetime.loc[col,'varsname']=col
    df_datetime.loc[col,'count']=total
    df_datetime.loc[col,'min_date']=df[col].min()
    df_datetime.loc[col,'max_date']=df[col].max()
    df_datetime.loc[col,'null']=df[col].isnull().sum()
    df_datetime.loc[col,'null_pct']=str(df[col].isnull().sum()/total *100)[0:5] + '%'
    
    return df_datetime





def data_view(df_sheet, sheet):
    # 表的数据质量评估
    cols=['name','columns_num','records','is_key','key_name','key_records','key_unique','key_duplicate']
#              'is_natrual_key','natrual_key_name','natrual_key_records','natrual_key_unique','natrual_key_duplicate']
    table = pd.Series(index=cols)
    table['tablename'] = sheet
    table['columns_num'] = df_sheet.shape[1]
    table['records'] = df_sheet.shape[0]

    keys = [i for i in df_sheet.columns if df_sheet[i].nunique() == df_sheet.shape[0]]
    if len(keys)==0:
        key_var = np.nan
        table['is_key'] = 0
    else:
        if 'user_id' in keys:
            key_var = 'user_id'
        elif 'order_no' in keys:
            key_var = 'order_no'
        elif 'id_no' in keys:
            key_var = 'id_no'
        else:
            key_var = keys[0]
        table['is_key'] = 1
        table['key_name'] = key_var
        table['key_records'] = df_sheet[key_var].count()
        table['key_unique'] = df_sheet[key_var].nunique()
        table['key_duplicate'] = table['key_records'] - table['key_unique']

#     table['is_natrual_key'] = '1'
#     table['natrual_key_name'] = natrual_key_var
#     table['natrual_key_records'] = df_sheet[natrual_key_var].count()
#     table['natrual_key_unique'] = df_sheet[natrual_key_var].nunique()
#     table['natrual_key_duplicate'] = table['natrual_key_records'] - table['natrual_key_unique']   
    table = pd.DataFrame(table)
    table = table.T

    # 字段数据质量评估
    # 字段数据类型
    var_dt_types = df_sheet.dtypes   
    var_dt_types = pd.DataFrame({'keys':list(var_dt_types.index),'keys_type':list(var_dt_types.values)})
    var_dt_types['tablename'] = sheet

    # 数据集记录数
    var_dt_types['records'] = df_sheet.shape[0]
    # 字段缺失数
    var_dt_types_null = df_sheet.isnull().sum()
    var_dt_types_null = pd.DataFrame({'keys':list(var_dt_types_null.index),'null_records':list(var_dt_types_null.values)})
    # 合并表计算缺失率
    var_dt_types_step1 = pd.merge(var_dt_types, var_dt_types_null, how='left', on='keys')
    var_dt_types_step1['null_pct'] = var_dt_types_step1['null_records'] / var_dt_types_step1['records']
    var_dt_types_step1['null_pct'] = var_dt_types_step1['null_pct'].apply(lambda x: str(x * 100)[0:6] + '%')

    # 字段不同取值个数
    col_list_var = []
    for j in df_sheet.columns:
        tmp_var = []
        tmp_var.append(j)
        tmp_var.append(df_sheet[j].nunique())
        col_list_var.append(tmp_var)
    var_dt_types_unique = pd.DataFrame(col_list_var, columns=['keys', 'unique'])
    # 合并表
    var_dt_types_step2 = pd.merge(var_dt_types_step1, var_dt_types_unique, how='left', on='keys')
 
    # 返回值
    return table, var_dt_types_step2




#==============================================================================
# File: function_20231108.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import os 
import json

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)


# 获取当前目录下的文件名
def get_filename(file_dir, filetype):
   """
   file_dir: 字符串，文件路径
   filetype：字符串，文件格式如.csv，.txt, .xls
   """
   #创建一个空列表，存储当前目录下的CSV文件全称
   file_name = []
   #将当前目录下的所有文件名称读取进来
   for root, dirs, files in os.walk(file_dir):
       # 判断是否为CSV文件，如果是则存储到列表中
       for file in files:
           n = len(filetype)
           n = -1 * n
           if file[n:] == filetype:
               file_name.append(file)   
       
   return file_name

# 对不同数据源的相同格式的文件进行合并
filepath = r'\\192.168.100.120\d\juzi\0904'
def merge_file(data_source, file_names, path=filepath, usecols=None):
    """
    data_source:    数据源名称
    file_names:     文件名称列表
    usecols:        列表，需要读取的字段
    """
    data_list = {}
    for i, filename in enumerate(file_names):
        if data_source in filename:
            data_list[i] = pd.read_csv(r'{}\{}'.format(path, filename), usecols=usecols)   
    if data_list:
        merge_data = pd.concat(list(data_list.values()), axis=0)
        
        return merge_data
    else:
        print('----------{}:无数据-------------'.format(data_source_name))
        return None

        
        
        
# 样本匹配三方数据,适用于单个文件比较小的数据
# 匹配逻辑：通过身份证号关联，按照如下逻辑保留匹配的样本
# 首先，申请样本的订单号能够匹配三方数据的订单号，否则，取30天以内最近的一次查询
def process_data(df_base, cols_left, df_three, cols_right, id_no,
                needcols=[], suffix=None, if_need=True):
    """
    df_base:        样本基础表
    cols_left：     关联三方数据时，样本表必须要的字段
    df_three：      数据库中存在的三方历史数据源
    cols_right：    关联样本数据时，三方数据表必须要的字段
    id_no：         身份证号
    needcols：      返回匹配的三方数据时，需要保留的评分/变量字段
    suffix：        给每个三方数据源字段添加的后缀，用来区分不同数据来源
    """
    df1 = pd.merge(df_base[cols_left], df_three[cols_right], how='inner', on=id_no, suffix=['_base', '_three'])
    df1['apply_date'] = pd.to_datetime(df1['apply_date'],format='%Y-%m-%d')
    df1['create_time'] = pd.to_datetime(df1['create_time'].str[0:10], format='%Y-%m-%d')
    df1['days'] = (df1['apply_date'] - df1['create_time']).dt.days
    df1['order_no_is_equal'] = [*map(lambda t1,t2: 1 if t2==t1 else 0, df1['order_no_base'], df1['order_no_three'])]
    df1['order_no_is_equal'].value_counts(dropna=False)
    
    # 拆分数据
    print('-----------拆分数据------------------')
    
    df1_part1 = df1.query("order_no_is_equal==1")
    # 去重
    df1_part1 = df1_part1.sort_values(by=['order_no_base', 'create_time'], ascending=False)
    df1_part1 = df1_part1.drop_duplicates(subset=['order_no_base'], keep='first')

    df1_part2 = df1.query("order_no_is_equal==0")
    df1_part2 = df1_part2.query("days<=30 & days>=0")
    # 去重
    df1_part2 = df1_part2.sort_values(by=['order_no_base', 'create_time'], ascending=False).
    df1_part2 = df1_part2.drop_duplicates(subset=['order_no_base'],keep='first')

    
    # 合并数据
    print('-----------合并数据------------------')
    
    df1 = pd.concat([df1_part1, df1_part2], axis=0)
    # 去重
    df1 = df1.sort_values(by=['order_no_base', 'order_no_is_equal', 'create_time'], ascending=False)
    df1 = df1.drop_duplicates(subset=['order_no_base'], keep='first')

    
    # 重新命名字段
    usecols = ['order_no_base', 'apply_date', 'order_no_three', 'create_time', 'order_no_is_equal'] + needcols
    df1 = df1[usecols]
    df1.rename(columns={'order_no_base':'order_no'}, inplace=True)
    
    cols = []
    if suffix:  
        for col in usecols[1:]:
            var = col +  '_' + suffix
            cols.append(var)
    
        cols = ['order_no'] + cols
        df1.columns = cols

        # 返回需要的数据
        return_columns = ['order_no'] + [i + '_{}'.format(suffix) for i in needcols]
        df2 = df1[return_columns]
    
    if if_need and suffix:
        return df2
    else:
        return df1
  

# 样本匹配三方数据,适用于单个文件很大的数据,字段很多，如多头变量
# 匹配逻辑：通过身份证号关联，按照如下逻辑保留匹配的样本
# 首先，申请样本的订单号能够匹配三方数据的订单号，否则，取30天以内最近的一次查询
def chunk_process_data(filepath, usecols,
                       df_base, cols_left, cols_right, id_no, needcols,
                       suffix=None, if_need=False, engine='c', chunk_size=50000, is_json=False):
    """
    example:
    filepath = r'\\192.168.100.120\d\juzi\0907\{}'.format(filename)
    usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_089']
    cols_left = ['order_no','id_no_des','apply_date']
    cols_right = ['order_no','id_no_des','create_time'] + needcols
    needcols =  ['value_089']
    """                   
    df_threes = pd.read_csv(filepath, usecols=usecols, engine=engine, chunksize=chunk_size)
    result_dict = {}
    for i, df_three in enumerate(df_threes):
        df1 = process_data(df_base, cols_left, df_three, cols_right, id_no,
                           needcols=needcols, suffix=suffix, if_need=if_need)
        result_dict[i] = df1
        
        del df_three

    result_df = pd.concat(list(result_dict.values()), axis=0)

    sortbycols = ['order_no'] + [i + '_{}'.format(suffix) for i in ['order_no_is_equal', 'create_time']]
    result_df = result_df.sort_values(by=bycols, ascending=False).drop_duplicates(subset=['order_no'],keep='first')
    
    if is_json:
        col = needcols[0] + '_{}'.format(suffix)
        result_df[col] = result_df[col].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
        result_df = pd.concat([result_df.drop([col], axis=1), result_df[col].apply(pd.Series)], axis=1)   
    
    return result_df

# 评分分布情况-建模样本
def score_distribute(data, col, target='target'):    
    total = data.groupby(col)[target].count()
    bad = data.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['bad_rate_cum'] = regroup['bad'].cumsum()/regroup['total'].cumsum()
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['good_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['total_pct_cum'] = regroup['total_pct'].cumsum()
    regroup['ks'] = regroup['bad_pct_cum'] - regroup['good_pct_cum']
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift_cum'] = regroup['bad_rate_cum']/data[target].mean()
    usecols = ['varsname','bins','bad','good','total','ks', 'bad_rate','bad_rate_cum','lift_cum','total_pct_cum'
            ,'total_pct', 'bad_pct','good_pct','bad_pct_cum','good_pct_cum']
    return_regroup = regroup[usecols]
    return_regroup = return_regroup.reset_index(drop=True)

    return return_regroup


# 评分分布情况-全部样本
def cal_auth(df, target='target'):       
    total_lend = df.groupby('bins')['order_no',target].count()
    tg_jj = df.groupby(['bins', 'auth_status'])['order_no'].count().unstack()
    lend =  df.groupby(['bins', target])['order_no'].count().unstack()
    result = pd.concat([total_lend, tg_jj, lend], axis=1)
    result.columns = ['授信申请','放款人数','通过人数','拒绝人数','好','灰','坏']
    
    result_sum = pd.DataFrame(result.sum(axis=0),columns=['total']).T
    
    for col in result.columns:
        result['{}占比'.format(col)] = result[col]/result[col].sum()
        result['{}累计占比'.format(col)] = result['{}占比'.format(col)].cumsum()
        
    result['坏客率'] = result['坏']/(result['好']+result['坏'])
    result['通过率'] = result['通过人数']/result['授信申请']
    result = pd.concat([result, result_sum],axis=0)
    
    cols = ['授信申请','授信申请占比','授信申请累计占比','通过人数','通过率','拒绝人数','拒绝人数占比','拒绝人数累计占比',
            '通过人数占比','通过人数累计占比','放款人数','放款人数占比','放款人数累计占比','好','好占比','好累计占比',
            '坏','坏占比','坏累计占比','坏客率','灰','灰占比','灰累计占比']
    result = result[cols]
    
    return result





# 离散变量分组
def bins_group(data, flag, col):
    df = data[[flag, col]]
    df['bins'] = df[col]
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    
    return regroup

# 保留特殊值不参与分箱
def regroup_special_merge(regroup_special, regroup_normal):
    while regroup_special['bad'].min()==0:
        id_min = regroup_special['bad'].idxmin()
        regroup_special.loc[id_min,'bad'] = 1
    while regroup_special['good'].min()==0:
        id_min = regroup_special['good'].idxmin()
        regroup_special.loc[id_min,'good']=  1
    regroup_normal.index.name = regroup_special.index.name
    return_regroup = pd.concat([regroup_special,regroup_normal],axis=0)
    
    return return_regroup
    
# 删除当前索引值所在行的后一行
def DelIndexPlus1(np_regroup, index_value):
    np_regroup[index_value,1] = np_regroup[index_value,1] + np_regroup[index_value+1,1]
    np_regroup[index_value,2] = np_regroup[index_value,2] + np_regroup[index_value+1,2]
    np_regroup[index_value,0] = np_regroup[index_value+1,0]
    np_regroup = np.delete(np_regroup, index_value+1, axis=0)
    
    return np_regroup
 
# 删除当前索引值所在行
def DelIndex(np_regroup, index_value):
    np_regroup[index_value-1,1] = np_regroup[index_value,1] + np_regroup[index_value-1,1]
    np_regroup[index_value-1,2] = np_regroup[index_value,2] + np_regroup[index_value-1,2]
    np_regroup[index_value-1,0] = np_regroup[index_value,0]
    np_regroup = np.delete(np_regroup, index_value, axis=0)
    
    return np_regroup 
    
# 删除/合并客户数为0的箱子
def MergeZero(np_regroup):
    #合并好坏客户数连续都为0的箱子
    i = 0
    while i<=np_regroup.shape[0]-2:
        if (np_regroup[i,1]==0 and np_regroup[i+1,1]==0) or (np_regroup[i,2]==0 and np_regroup[i+1,2]==0):
            np_regroup = DelIndexPlus1(np_regroup,i)
        i = i+1
        
    #合并坏客户数为0的箱子
    while True:
        if all(np_regroup[:,1]>0) or np_regroup.shape[0]==2:
            break
        bad_zero_index = np.argwhere(np_regroup[:,1]==0)[0][0]
        if bad_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
        elif bad_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, bad_zero_index)
        else:
            if np_regroup[bad_zero_index-1,2]/np_regroup[bad_zero_index-1,1]>=np_regroup[bad_zero_index+1,2]/np_regroup[bad_zero_index+1,1]:
                np_regroup = DelIndex(np_regroup, bad_zero_index)
            else:
                np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
    #合并好客户数为0的箱子
    while True:
        if all(np_regroup[:,2]>0) or np_regroup.shape[0]==2:
            break
        good_zero_index = np.argwhere(np_regroup[:,2]==0)[0][0]
        if good_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
        elif good_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, good_zero_index)
        else:
            if np_regroup[good_zero_index-1,2]/np_regroup[good_zero_index-1,1]>=np_regroup[good_zero_index+1,2]/np_regroup[good_zero_index+1,1]:
                np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
            else:
                np_regroup = DelIndex(np_regroup, good_zero_index)
                
    return np_regroup
    
# 箱子的单调性
def MonTone(np_regroup):
    while True:
        if np_regroup.shape[0]==2:
            break
        GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
        GoodBadRateMonetone = [GoodBadRate[i]<GoodBadRate[i+1] for i in range(np_regroup.shape[0]-1)]
        #确定是否单调
        if_Montone = len(set(GoodBadRateMonetone))
        #判断跳出循环
        if if_Montone==1:
            break
        else:
            WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
            Montone_index = WoeDiffMin.index(min(WoeDiffMin))
            np_regroup = DelIndexPlus1(np_regroup, Montone_index)
            
    return np_regroup
    
#箱子最小占比
def MinPct(np_regroup):
    while True:
        bins_pct = [(np_regroup[i,1]+np_regroup[i,2])/np_regroup.sum() for i in range(np_regroup.shape[0])]
        min_pct = min(bins_pct)
        if min_pct>=0.02 or len(bins_pct)==2:
            break
        else:
            min_pct_index = bins_pct.index(min(bins_pct))
            if min_pct_index==0:
                np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
            elif min_pct_index == np_regroup.shape[0]-1:
                np_regroup = DelIndex(np_regroup, min_pct_index)
            else:
                GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
                WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
                if WoeDiffMin[min_pct_index-1]>=WoeDiffMin[min_pct_index]:
                    np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
                else:
                    np_regroup = DelIndex(np_regroup, min_pct_index)
    return np_regroup

    
# 连续变量分箱主函数
def ContinueVarBins(data, col, flag='target', cutbins=[]):
    df = data[[flag, col]]
    df = df[~df[col].isnull()].reset_index(drop=True)
    df['bins'] = pd.cut(df[col], cutbins, duplicates='drop', right=False, precision=4)
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup.drop(['total'], axis=1, inplace=True)
    np_regroup = np.array(regroup)
    np_regroup = MergeZero(np_regroup)
    np_regroup = MinPct(np_regroup)
    np_regroup = MonTone(np_regroup)
    regroup = pd.DataFrame(index=np.arange(np_regroup.shape[0]))
    regroup['bins'] = np_regroup[:,0]
    regroup['total'] = np_regroup[:,1] + np_regroup[:,2]
    regroup['bad'] = np_regroup[:,1]
    regroup['good'] = np_regroup[:,2]
    cutoffpoints = list(np_regroup[:,0])
    cutoffpoints = [float('-inf')] + cutoffpoints
    # 最大值分割点，转换最小值分割点
    df['bins_new'] = pd.cut(df[col], cutoffpoints, duplicates='drop', right=True, precision=4)
    tmp = pd.DataFrame()
    tmp['bins'] = df.groupby(['bins_new'])[col].min()
    cutoffpoints = list(tmp['bins'])  
    
    return cutoffpoints


def CalWoeIv(data_bins, col, target='target'):  
    total = data_bins.groupby(col)[target].count()
    bad = data_bins.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['ks_bins'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
    regroup['ks'] = regroup['ks_bins'].max()
    regroup['woe'] = np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv'] = regroup['iv_bins'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift'] = regroup['bad_rate']/data_bins[target].mean()
    usecols = ['varsname','bins','iv','ks','bad','good','total','bad_rate','total_pct','lift','woe']
    return_regroup = regroup[usecols]

    return return_regroup


# 变量间相关性
def CorrSelect(df, iv_df, exclude_list=[], threshold=0.7):
    X = [i for i in df.columns if i not in exclude_list]
    data = df[X]
    df_corr = data.corr()
    droped_list = []
    while True:
        dict_cols = dict(zip(range(df_corr.shape[1]), list(df_corr.columns)))
        np_corr = abs(np.array(df_corr))
        np.fill_diagonal(np_corr, 0)
        if np.amax(np_corr) < threshold:
            break
        index1, index2 = np.unravel_index(np_corr.argmax(), np_corr.shape)
        x1 = dict_cols[index1]
        x2 = dict_cols[index2]
        if iv_df.loc[x1,'iv']>=iv_df.loc[x2,'iv']:
            droped_list.append(x2)
            df_corr.drop(index=[x2], inplace=True)
            df_corr.drop(columns=[x2], inplace=True)
        else:
            droped_list.append(x1)
            df_corr.drop(index=[x1], inplace=True)
            df_corr.drop(columns=[x1], inplace=True)
    
    return droped_list
 

def sklearn_vif(exogs, data):
    from sklearn.linear_model import LinearRegression
    # initialize dictionaries
    vif_dict, tolerance_dict = {}, {}

    # form input data for each exogenous variable
    for exog in exogs:
        not_exog = [i for i in exogs if i != exog]
        X, y = data[not_exog], data[exog]

        # extract r-squared from the fit
        r_squared = LinearRegression().fit(X, y).score(X, y)

        # calculate VIF
        vif = 1/(1 - r_squared)
        vif_dict[exog] = vif

        # calculate tolerance
        tolerance = 1 - r_squared
        tolerance_dict[exog] = tolerance

    # return VIF DataFrame
    df_vif = pd.DataFrame({'VIF': vif_dict, 'Tolerance': tolerance_dict})

    return df_vif





import matplotlib.pyplot as plt
def lr(X_train,X_test,y_train,y_test):
    from sklearn.linear_model import LogisticRegression
    from scipy import stats
    from toad.metrics import KS,AUC 
    # 模型训练和p_values查看 只能是线性回归
    clf = LogisticRegression(C=1e8).fit(X_train, y_train)
    params = np.append(clf.intercept_,clf.coef_)
    
    new_X_train = pd.DataFrame({"Constant":np.ones(len(X_train))}).join(X_train.reset_index(drop=True))
    predictions = clf.predict(X_train)
    MSE = (sum((y_train-predictions)**2))/(len(new_X_train)-len(new_X_train.columns))

    var_b = MSE*(np.linalg.inv(np.dot(new_X_train.T,new_X_train)).diagonal())
    sd_b = np.sqrt(var_b)
    ts_b = params/ sd_b

    p_values =[2*(1-stats.t.cdf(np.abs(i),(len(new_X_train)-1))) for i in ts_b]

    sd_b = np.round(sd_b,3)
    ts_b = np.round(ts_b,3)
    p_values = np.round(p_values,3)
    params = np.round(params,4)

    myDF3 = pd.DataFrame()
    myDF3["Vars"],myDF3["Coefficients"],myDF3["Standard Errors"],myDF3["t values"],myDF3["P values"] = [new_X_train.columns,params,sd_b,ts_b,p_values]
    print(myDF3)
    
    # 预测值
    pred_train = clf.predict_proba(X_train)[:,1]
    pred_test = clf.predict_proba(X_test)[:,1]
    
    # 训练集KS/AUC
    print('-------------训练集结果--------------------')
    print('train AUC: ', AUC(pred_train, y_train))
    print('train KS: ', KS(pred_train, y_train))
    
    # 测试集KS/AUC
    print('-------------测试集结果--------------------')
    print('test AUC: ', AUC(pred_test, y_test))
    print('test KS: ', KS(pred_test, y_test))
    
    print('-------------------------分割线--------------------------')
    # 模型评估
    train_AUC = roc_auc_score(y_train, pred_train)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,pred_train)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_AUC = roc_auc_score(y_test, pred_test)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,pred_test)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)

    return (myDF3, params, clf)

# 单次训练
def LR_model(X_train,X_test,y_train,y_test):
    clf = sm.Logit(y_train, sm.add_constant(X_train)).fit()
    print(clf.summary())
    # 模型评估
    train_y_pred = clf.predict(sm.add_constant(X_train))
    train_AUC = roc_auc_score(y_train, train_y_pred)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,train_y_pred)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_y_pred = clf.predict(sm.add_constant(X_test))
    test_AUC = roc_auc_score(y_test,test_y_pred)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,test_y_pred)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)
    
    return clf









#==============================================================================
# File: function_三方数据 - 副本.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import os 

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)


# In[ ]:


# 获取当前目录下的CSV文件名
def get_filename(file_dir):
   #创建一个空列表，存储当前目录下的CSV文件全称
   file_name = []
   #将当前目录下的所有文件名称读取进来
   for root, dirs, files in os.walk(file_dir):
       # 判断是否为CSV文件，如果是则存储到列表中
       for file in files:
           if file[-4:] == '.csv':
               file_name.append(file)   
   
#     files = os.listdir()        
#     for j in files:
#     #判断是否为CSV文件，如果是则存储到列表中
#     if os.path.splitext(j)[1] == '.csv':
#         file_name.append(j)
       
   return file_name


# In[ ]:


# filepath = r'\\192.168.100.120\d\juzi\0904'
# def merge_csv_file(data_source_name, file_name_zz, path=r'\\192.168.100.120\d\juzi\0904', usecols=None):
#     """
#     data_source_name:数据源名称
#     file_name_zz:csv文件列表
#     """
#     data_list = {}
#     for channle in [167, 174, 206, 80004, 80005]:
#         for i, filename in enumerate(file_name_zz):
#             if '_{}_'.format(channle) in filename:
#                 channle_lenght = len(f'dwd.dwd_beforeloan_third_combine_id_{channle}_')
#                 if data_source_name == filename[channle_lenght:-11]: #and filename[-10:-4]<=str(202304)
#                     data_list[i] = pd.read_csv(r'{}\{}'.format(path, filename), usecols=usecols)   
#     if data_list:
#         merge_df = pd.concat(list(data_list.values()), axis=0)
#         return merge_df
#     else:
#         print('----------{}:无数据-------------'.format(data_source_name))
#         return None


# In[ ]:


def process_data(df_base, cols_left, df_three, cols_right, needcols=[], suffix=None):
    df1 = pd.merge(df_base[cols_left], df_three[cols_right], how='inner', on='id_no_des')
    df1['apply_date'] = pd.to_datetime(df1['apply_date'],format='%Y-%m-%d')
    df1['create_time'] = pd.to_datetime(df1['create_time'].str[0:10], format='%Y-%m-%d')
    df1['days'] = df1['apply_date'] - df1['create_time']
    df1['days'] = df1['days'].dt.days
    df1['order_no_is_equal'] = [*map(lambda t1,t2: 1 if t2==t1 else 0,df1['order_no_x'],df1['order_no_y'])]
    df1['order_no_is_equal'].value_counts(dropna=False)
    
    # 拆分数据
    print('-----------拆分数据------------------')
    df1_part1 = df1.query("order_no_is_equal==1")
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)
    # 去重
    df1_part1 = df1_part1.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)
    
    df1_part2 = df1.query("order_no_is_equal==0")
    df1_part2 = df1_part2.query("days<=30 & days>=0")
#     print(df1_part2['order_no_x'].nunique(), df1_part2.shape)
    # 去重
    df1_part2 = df1_part2.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print( df1_part2['order_no_x'].nunique(), df1_part2.shape)
    
    # 合并数据
    print('-----------合并数据------------------')
    df1_new = pd.concat([df1_part1, df1_part2], axis=0)
#     print( df1_new['order_no_x'].nunique(), df1_new.shape)
    # 去重
    df1_new = df1_new.sort_values(by=['order_no_x','order_no_is_equal','create_time'], ascending=False)
    df1_new = df1_new.drop_duplicates(subset=['order_no_x'],keep='first')
    print("去重后流水订单号数：", df1_new['order_no_x'].nunique())
    print("成功匹配的三方数据：", df1_new.shape)
    
    # 重新命名字段
    usecols = ['order_no_x','apply_date','order_no_y','create_time','order_no_is_equal'] + needcols
    df1_new = df1_new[usecols]

    cols = []
    for col in usecols[1:]:
        var = col +  '_' + suffix
        cols.append(var)
    
    cols = ['order_no'] + cols
    df1_new.columns = cols
    df1_new.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_{}_{}.csv'.format(suffix,str(datetime.today())[:10].replace('-','')), index=False)
    
    # 返回需要的数据
    return_keys = ['order_no'] + cols[5:]
    df2 = df1_new[return_keys]
    gc.collect()
    
    return df2
  


# In[ ]:


def process_data_bairong(df_base, cols_left, df_three, cols_right, needcols=[], suffix=None):
    df1 = pd.merge(df_base[cols_left], df_three[cols_right], how='inner', on='id_no_des')
    df1['apply_date'] = pd.to_datetime(df1['apply_date'],format='%Y-%m-%d')
    df1['create_time'] = pd.to_datetime(df1['create_time'].str[0:10], format='%Y-%m-%d')
    df1['days'] = df1['apply_date'] - df1['create_time']
    df1['days'] = df1['days'].dt.days
    df1['order_no_is_equal'] = [*map(lambda t1,t2: 1 if t2==t1 else 0,df1['order_no_x'],df1['order_no_y'])]
    df1['order_no_is_equal'].value_counts(dropna=False)

    # 拆分数据
    print('-----------拆分数据------------------')
    df1_part1 = df1.query("order_no_is_equal==1")
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)
    # 去重
    df1_part1 = df1_part1.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print(df1_part1['order_no_x'].nunique(), df1_part1.shape)

    df1_part2 = df1.query("order_no_is_equal==0")
    df1_part2 = df1_part2.query("days<=30 & days>=0")
#     print(df1_part2['order_no_x'].nunique(), df1_part2.shape)
    # 去重
    df1_part2 = df1_part2.sort_values(by=['order_no_x', 'create_time'], ascending=False).drop_duplicates(subset=['order_no_x'],keep='first')
#     print( df1_part2['order_no_x'].nunique(), df1_part2.shape)

    # 合并数据
    print('-----------合并数据------------------')
    df1_new = pd.concat([df1_part1, df1_part2], axis=0)
#     print( df1_new['order_no_x'].nunique(), df1_new.shape)
    # 去重
    df1_new = df1_new.sort_values(by=['order_no_x','order_no_is_equal','create_time'], ascending=False)
    df1_new = df1_new.drop_duplicates(subset=['order_no_x'],keep='first')
    print("去重后流水订单号数：", df1_new['order_no_x'].nunique())
    print("成功匹配的三方数据：", df1_new.shape)

    # 重新命名字段
    usecols = ['order_no_x','apply_date','order_no_y','create_time','order_no_is_equal'] + needcols
    df1_new = df1_new[usecols]

    cols = []
    for col in usecols[1:]:
        var = col +  '_' + suffix
        cols.append(var)

    cols = ['order_no'] + cols
    df1_new.columns = cols

    return df1_new


# In[ ]:


def chunk_process_data(filepath, usecols, df_base, cols_left, cols_right, needcols, suffix='bairong_1', engine='c', chunk_size=50000):
    df_chunks = pd.read_csv(filepath, usecols=usecols, engine=engine, chunksize=chunk_size)
    result_dict = {}
    for i, chunk in enumerate(df_chunks):
        df1_new = process_data_bairong(df_base, cols_left, chunk, cols_right, needcols=needcols, suffix=suffix)
        result_dict[i] = df1_new
        print('-----处理完分块数据：{}-------'.format(i))

        del chunk

    result_df = pd.concat(list(result_dict.values()), axis=0)

    bycols = ['order_no'] + [i + '_{}'.format(suffix) for i in ['order_no_is_equal', 'create_time']]
    result_df = result_df.sort_values(by=bycols, ascending=False).drop_duplicates(subset=['order_no'],keep='first')

    return result_df




#==============================================================================
# File: function_三方数据.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import os 
import json

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)


# 获取当前目录下的文件名
def get_filename(file_dir, filetype):
   """
   file_dir: 字符串，文件路径
   filetype：字符串，文件格式如.csv，.txt, .xls
   """
   #创建一个空列表，存储当前目录下的CSV文件全称
   file_name = []
   #将当前目录下的所有文件名称读取进来
   for root, dirs, files in os.walk(file_dir):
       # 判断是否为CSV文件，如果是则存储到列表中
       for file in files:
           n = len(filetype)
           n = -1 * n
           if file[n:] == filetype:
               file_name.append(file)   
       
   return file_name

# 对不同数据源的相同格式的文件进行合并
filepath = r'\\192.168.100.120\d\juzi\0904'
def merge_file(data_source, file_names, path=filepath, usecols=None):
    """
    data_source:    数据源名称
    file_names:     文件名称列表
    usecols:        列表，需要读取的字段
    """
    data_list = {}
    for i, filename in enumerate(file_names):
        if data_source in filename:
            data_list[i] = pd.read_csv(r'{}\{}'.format(path, filename), usecols=usecols)   
    if data_list:
        merge_data = pd.concat(list(data_list.values()), axis=0)
        
        return merge_data
    else:
        print('----------{}:无数据-------------'.format(data_source_name))
        return None

        
        
        
# 样本匹配三方数据,适用于单个文件比较小的数据
# 匹配逻辑：通过身份证号关联，按照如下逻辑保留匹配的样本
# 首先，申请样本的订单号能够匹配三方数据的订单号，否则，取30天以内最近的一次查询
def process_data(df_base, cols_left, df_three, cols_right, id_no,
                needcols=[], suffix=None, if_need=True):
    """
    df_base:        样本基础表
    cols_left：     关联三方数据时，样本表必须要的字段
    df_three：      数据库中存在的三方历史数据源
    cols_right：    关联样本数据时，三方数据表必须要的字段
    id_no：         身份证号
    needcols：      返回匹配的三方数据时，需要保留的评分/变量字段
    suffix：        给每个三方数据源字段添加的后缀，用来区分不同数据来源
    """
    df1 = pd.merge(df_base[cols_left], df_three[cols_right], how='inner', on=id_no, suffix=['_base', '_three'])
    df1['apply_date'] = pd.to_datetime(df1['apply_date'],format='%Y-%m-%d')
    df1['create_time'] = pd.to_datetime(df1['create_time'].str[0:10], format='%Y-%m-%d')
    df1['days'] = (df1['apply_date'] - df1['create_time']).dt.days
    df1['order_no_is_equal'] = [*map(lambda t1,t2: 1 if t2==t1 else 0, df1['order_no_base'], df1['order_no_three'])]
    df1['order_no_is_equal'].value_counts(dropna=False)
    
    # 拆分数据
    print('-----------拆分数据------------------')
    
    df1_part1 = df1.query("order_no_is_equal==1")
    # 去重
    df1_part1 = df1_part1.sort_values(by=['order_no_base', 'create_time'], ascending=False)
    df1_part1 = df1_part1.drop_duplicates(subset=['order_no_base'], keep='first')

    df1_part2 = df1.query("order_no_is_equal==0")
    df1_part2 = df1_part2.query("days<=30 & days>=0")
    # 去重
    df1_part2 = df1_part2.sort_values(by=['order_no_base', 'create_time'], ascending=False).
    df1_part2 = df1_part2.drop_duplicates(subset=['order_no_base'],keep='first')

    
    # 合并数据
    print('-----------合并数据------------------')
    
    df1 = pd.concat([df1_part1, df1_part2], axis=0)
    # 去重
    df1 = df1.sort_values(by=['order_no_base', 'order_no_is_equal', 'create_time'], ascending=False)
    df1 = df1.drop_duplicates(subset=['order_no_base'], keep='first')

    
    # 重新命名字段
    usecols = ['order_no_base', 'apply_date', 'order_no_three', 'create_time', 'order_no_is_equal'] + needcols
    df1 = df1[usecols]
    df1.rename(columns={'order_no_base':'order_no'}, inplace=True)
    
    cols = []
    if suffix:  
        for col in usecols[1:]:
            var = col +  '_' + suffix
            cols.append(var)
    
        cols = ['order_no'] + cols
        df1.columns = cols

        # 返回需要的数据
        return_columns = ['order_no'] + [i + '_{}'.format(suffix) for i in needcols]
        df2 = df1[return_columns]
    
    if if_need and suffix:
        return df2
    else:
        return df1
  

# 样本匹配三方数据,适用于单个文件很大的数据,字段很多，如多头变量
# 匹配逻辑：通过身份证号关联，按照如下逻辑保留匹配的样本
# 首先，申请样本的订单号能够匹配三方数据的订单号，否则，取30天以内最近的一次查询
def chunk_process_data(filepath, usecols,
                       df_base, cols_left, cols_right, id_no, needcols,
                       suffix=None, if_need=False, engine='c', chunk_size=50000, is_json=False):
    """
    example:
    filepath = r'\\192.168.100.120\d\juzi\0907\{}'.format(filename)
    usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_089']
    cols_left = ['order_no','id_no_des','apply_date']
    cols_right = ['order_no','id_no_des','create_time'] + needcols
    needcols =  ['value_089']
    """                   
    df_threes = pd.read_csv(filepath, usecols=usecols, engine=engine, chunksize=chunk_size)
    result_dict = {}
    for i, df_three in enumerate(df_threes):
        df1 = process_data(df_base, cols_left, df_three, cols_right, id_no,
                           needcols=needcols, suffix=suffix, if_need=if_need)
        result_dict[i] = df1
        
        del df_three

    result_df = pd.concat(list(result_dict.values()), axis=0)

    sortbycols = ['order_no'] + [i + '_{}'.format(suffix) for i in ['order_no_is_equal', 'create_time']]
    result_df = result_df.sort_values(by=bycols, ascending=False).drop_duplicates(subset=['order_no'],keep='first')
    
    if is_json:
        col = needcols[0] + '_{}'.format(suffix)
        result_df[col] = result_df[col].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
        result_df = pd.concat([result_df.drop([col], axis=1), result_df[col].apply(pd.Series)], axis=1)   
    
    return result_df

# 评分分布情况-建模样本
def score_distribute(data, col, target='target'):    
    total = data.groupby(col)[target].count()
    bad = data.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['bad_rate_cum'] = regroup['bad'].cumsum()/regroup['total'].cumsum()
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['good_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['total_pct_cum'] = regroup['total_pct'].cumsum()
    regroup['ks'] = regroup['bad_pct_cum'] - regroup['good_pct_cum']
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift_cum'] = regroup['bad_rate_cum']/data[target].mean()
    usecols = ['varsname','bins','bad','good','total','ks', 'bad_rate','bad_rate_cum','lift_cum','total_pct_cum'
            ,'total_pct', 'bad_pct','good_pct','bad_pct_cum','good_pct_cum']
    return_regroup = regroup[usecols]
    return_regroup = return_regroup.reset_index(drop=True)

    return return_regroup


# 评分分布情况-全部样本
def cal_auth(df, target='target'):       
    total_lend = df.groupby('bins')['order_no',target].count()
    tg_jj = df.groupby(['bins', 'auth_status'])['order_no'].count().unstack()
    lend =  df.groupby(['bins', target])['order_no'].count().unstack()
    result = pd.concat([total_lend, tg_jj, lend], axis=1)
    result.columns = ['授信申请','放款人数','通过人数','拒绝人数','好','灰','坏']
    
    result_sum = pd.DataFrame(result.sum(axis=0),columns=['total']).T
    
    for col in result.columns:
        result['{}占比'.format(col)] = result[col]/result[col].sum()
        result['{}累计占比'.format(col)] = result['{}占比'.format(col)].cumsum()
        
    result['坏客率'] = result['坏']/(result['好']+result['坏'])
    result['通过率'] = result['通过人数']/result['授信申请']
    result = pd.concat([result, result_sum],axis=0)
    
    cols = ['授信申请','授信申请占比','授信申请累计占比','通过人数','通过率','拒绝人数','拒绝人数占比','拒绝人数累计占比',
            '通过人数占比','通过人数累计占比','放款人数','放款人数占比','放款人数累计占比','好','好占比','好累计占比',
            '坏','坏占比','坏累计占比','坏客率','灰','灰占比','灰累计占比']
    result = result[cols]
    
    return result





# 离散变量分组
def bins_group(data, flag, col):
    df = data[[flag, col]]
    df['bins'] = df[col]
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    
    return regroup

# 保留特殊值不参与分箱
def regroup_special_merge(regroup_special, regroup_normal):
    while regroup_special['bad'].min()==0:
        id_min = regroup_special['bad'].idxmin()
        regroup_special.loc[id_min,'bad'] = 1
    while regroup_special['good'].min()==0:
        id_min = regroup_special['good'].idxmin()
        regroup_special.loc[id_min,'good']=  1
    regroup_normal.index.name = regroup_special.index.name
    return_regroup = pd.concat([regroup_special,regroup_normal],axis=0)
    
    return return_regroup
    
# 删除当前索引值所在行的后一行
def DelIndexPlus1(np_regroup, index_value):
    np_regroup[index_value,1] = np_regroup[index_value,1] + np_regroup[index_value+1,1]
    np_regroup[index_value,2] = np_regroup[index_value,2] + np_regroup[index_value+1,2]
    np_regroup[index_value,0] = np_regroup[index_value+1,0]
    np_regroup = np.delete(np_regroup, index_value+1, axis=0)
    
    return np_regroup
 
# 删除当前索引值所在行
def DelIndex(np_regroup, index_value):
    np_regroup[index_value-1,1] = np_regroup[index_value,1] + np_regroup[index_value-1,1]
    np_regroup[index_value-1,2] = np_regroup[index_value,2] + np_regroup[index_value-1,2]
    np_regroup[index_value-1,0] = np_regroup[index_value,0]
    np_regroup = np.delete(np_regroup, index_value, axis=0)
    
    return np_regroup 
    
# 删除/合并客户数为0的箱子
def MergeZero(np_regroup):
    #合并好坏客户数连续都为0的箱子
    i = 0
    while i<=np_regroup.shape[0]-2:
        if (np_regroup[i,1]==0 and np_regroup[i+1,1]==0) or (np_regroup[i,2]==0 and np_regroup[i+1,2]==0):
            np_regroup = DelIndexPlus1(np_regroup,i)
        i = i+1
        
    #合并坏客户数为0的箱子
    while True:
        if all(np_regroup[:,1]>0) or np_regroup.shape[0]==2:
            break
        bad_zero_index = np.argwhere(np_regroup[:,1]==0)[0][0]
        if bad_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
        elif bad_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, bad_zero_index)
        else:
            if np_regroup[bad_zero_index-1,2]/np_regroup[bad_zero_index-1,1]>=np_regroup[bad_zero_index+1,2]/np_regroup[bad_zero_index+1,1]:
                np_regroup = DelIndex(np_regroup, bad_zero_index)
            else:
                np_regroup = DelIndexPlus1(np_regroup, bad_zero_index)
    #合并好客户数为0的箱子
    while True:
        if all(np_regroup[:,2]>0) or np_regroup.shape[0]==2:
            break
        good_zero_index = np.argwhere(np_regroup[:,2]==0)[0][0]
        if good_zero_index==0:
            np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
        elif good_zero_index==np_regroup.shape[0]-1:
            np_regroup = DelIndex(np_regroup, good_zero_index)
        else:
            if np_regroup[good_zero_index-1,2]/np_regroup[good_zero_index-1,1]>=np_regroup[good_zero_index+1,2]/np_regroup[good_zero_index+1,1]:
                np_regroup = DelIndexPlus1(np_regroup, good_zero_index)
            else:
                np_regroup = DelIndex(np_regroup, good_zero_index)
                
    return np_regroup
    
# 箱子的单调性
def MonTone(np_regroup):
    while True:
        if np_regroup.shape[0]==2:
            break
        GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
        GoodBadRateMonetone = [GoodBadRate[i]<GoodBadRate[i+1] for i in range(np_regroup.shape[0]-1)]
        #确定是否单调
        if_Montone = len(set(GoodBadRateMonetone))
        #判断跳出循环
        if if_Montone==1:
            break
        else:
            WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
            Montone_index = WoeDiffMin.index(min(WoeDiffMin))
            np_regroup = DelIndexPlus1(np_regroup, Montone_index)
            
    return np_regroup
    
#箱子最小占比
def MinPct(np_regroup):
    while True:
        bins_pct = [(np_regroup[i,1]+np_regroup[i,2])/np_regroup.sum() for i in range(np_regroup.shape[0])]
        min_pct = min(bins_pct)
        if min_pct>=0.02 or len(bins_pct)==2:
            break
        else:
            min_pct_index = bins_pct.index(min(bins_pct))
            if min_pct_index==0:
                np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
            elif min_pct_index == np_regroup.shape[0]-1:
                np_regroup = DelIndex(np_regroup, min_pct_index)
            else:
                GoodBadRate = [np_regroup[i,2]/np_regroup[i,1] for i in range(np_regroup.shape[0])]
                WoeDiffMin = [abs(np.log(GoodBadRate[i]/GoodBadRate[i+1])) for i in range(np_regroup.shape[0]-1)]
                if WoeDiffMin[min_pct_index-1]>=WoeDiffMin[min_pct_index]:
                    np_regroup = DelIndexPlus1(np_regroup, min_pct_index)
                else:
                    np_regroup = DelIndex(np_regroup, min_pct_index)
    return np_regroup

    
# 连续变量分箱主函数
def ContinueVarBins(data, col, flag='target', cutbins=[]):
    df = data[[flag, col]]
    df = df[~df[col].isnull()].reset_index(drop=True)
    df['bins'] = pd.cut(df[col], cutbins, duplicates='drop', right=False, precision=4)
    regroup = pd.DataFrame()
    regroup['bins'] = df.groupby(['bins'])[col].max()
    regroup['total'] = df.groupby(['bins'])[flag].count()
    regroup['bad'] = df.groupby(['bins'])[flag].sum()
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup.drop(['total'], axis=1, inplace=True)
    np_regroup = np.array(regroup)
    np_regroup = MergeZero(np_regroup)
    np_regroup = MinPct(np_regroup)
    np_regroup = MonTone(np_regroup)
    regroup = pd.DataFrame(index=np.arange(np_regroup.shape[0]))
    regroup['bins'] = np_regroup[:,0]
    regroup['total'] = np_regroup[:,1] + np_regroup[:,2]
    regroup['bad'] = np_regroup[:,1]
    regroup['good'] = np_regroup[:,2]
    cutoffpoints = list(np_regroup[:,0])
    cutoffpoints = [float('-inf')] + cutoffpoints
    # 最大值分割点，转换最小值分割点
    df['bins_new'] = pd.cut(df[col], cutoffpoints, duplicates='drop', right=True, precision=4)
    tmp = pd.DataFrame()
    tmp['bins'] = df.groupby(['bins_new'])[col].min()
    cutoffpoints = list(tmp['bins'])  
    
    return cutoffpoints


def CalWoeIv(data_bins, col, target='target'):  
    total = data_bins.groupby(col)[target].count()
    bad = data_bins.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['ks_bins'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
    regroup['ks'] = regroup['ks_bins'].max()
    regroup['woe'] = np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv'] = regroup['iv_bins'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    regroup['lift'] = regroup['bad_rate']/data_bins[target].mean()
    usecols = ['varsname','bins','iv','ks','bad','good','total','bad_rate','total_pct','lift','woe']
    return_regroup = regroup[usecols]

    return return_regroup


# 变量间相关性
def CorrSelect(df, iv_df, exclude_list=[], threshold=0.7):
    X = [i for i in df.columns if i not in exclude_list]
    data = df[X]
    df_corr = data.corr()
    droped_list = []
    while True:
        dict_cols = dict(zip(range(df_corr.shape[1]), list(df_corr.columns)))
        np_corr = abs(np.array(df_corr))
        np.fill_diagonal(np_corr, 0)
        if np.amax(np_corr) < threshold:
            break
        index1, index2 = np.unravel_index(np_corr.argmax(), np_corr.shape)
        x1 = dict_cols[index1]
        x2 = dict_cols[index2]
        if iv_df.loc[x1,'iv']>=iv_df.loc[x2,'iv']:
            droped_list.append(x2)
            df_corr.drop(index=[x2], inplace=True)
            df_corr.drop(columns=[x2], inplace=True)
        else:
            droped_list.append(x1)
            df_corr.drop(index=[x1], inplace=True)
            df_corr.drop(columns=[x1], inplace=True)
    
    return droped_list
 

def sklearn_vif(exogs, data):
    from sklearn.linear_model import LinearRegression
    # initialize dictionaries
    vif_dict, tolerance_dict = {}, {}

    # form input data for each exogenous variable
    for exog in exogs:
        not_exog = [i for i in exogs if i != exog]
        X, y = data[not_exog], data[exog]

        # extract r-squared from the fit
        r_squared = LinearRegression().fit(X, y).score(X, y)

        # calculate VIF
        vif = 1/(1 - r_squared)
        vif_dict[exog] = vif

        # calculate tolerance
        tolerance = 1 - r_squared
        tolerance_dict[exog] = tolerance

    # return VIF DataFrame
    df_vif = pd.DataFrame({'VIF': vif_dict, 'Tolerance': tolerance_dict})

    return df_vif





import matplotlib.pyplot as plt
def lr(X_train,X_test,y_train,y_test):
    from sklearn.linear_model import LogisticRegression
    from scipy import stats
    from toad.metrics import KS,AUC 
    # 模型训练和p_values查看 只能是线性回归
    clf = LogisticRegression(C=1e8).fit(X_train, y_train)
    params = np.append(clf.intercept_,clf.coef_)
    
    new_X_train = pd.DataFrame({"Constant":np.ones(len(X_train))}).join(X_train.reset_index(drop=True))
    predictions = clf.predict(X_train)
    MSE = (sum((y_train-predictions)**2))/(len(new_X_train)-len(new_X_train.columns))

    var_b = MSE*(np.linalg.inv(np.dot(new_X_train.T,new_X_train)).diagonal())
    sd_b = np.sqrt(var_b)
    ts_b = params/ sd_b

    p_values =[2*(1-stats.t.cdf(np.abs(i),(len(new_X_train)-1))) for i in ts_b]

    sd_b = np.round(sd_b,3)
    ts_b = np.round(ts_b,3)
    p_values = np.round(p_values,3)
    params = np.round(params,4)

    myDF3 = pd.DataFrame()
    myDF3["Vars"],myDF3["Coefficients"],myDF3["Standard Errors"],myDF3["t values"],myDF3["P values"] = [new_X_train.columns,params,sd_b,ts_b,p_values]
    print(myDF3)
    
    # 预测值
    pred_train = clf.predict_proba(X_train)[:,1]
    pred_test = clf.predict_proba(X_test)[:,1]
    
    # 训练集KS/AUC
    print('-------------训练集结果--------------------')
    print('train AUC: ', AUC(pred_train, y_train))
    print('train KS: ', KS(pred_train, y_train))
    
    # 测试集KS/AUC
    print('-------------测试集结果--------------------')
    print('test AUC: ', AUC(pred_test, y_test))
    print('test KS: ', KS(pred_test, y_test))
    
    print('-------------------------分割线--------------------------')
    # 模型评估
    train_AUC = roc_auc_score(y_train, pred_train)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,pred_train)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_AUC = roc_auc_score(y_test, pred_test)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,pred_test)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)

    return (myDF3, params, clf)

# 单次训练
def LR_model(X_train,X_test,y_train,y_test):
    clf = sm.Logit(y_train, sm.add_constant(X_train)).fit()
    print(clf.summary())
    # 模型评估
    train_y_pred = clf.predict(sm.add_constant(X_train))
    train_AUC = roc_auc_score(y_train, train_y_pred)
    train_fpr, train_tpr, train_thresholds = roc_curve(y_train,train_y_pred)
    train_KS = max(train_tpr - train_fpr)
    print('TRAIN AUC: {}'.format(train_AUC))
    print('TRAIN KS: {}'.format(train_KS))
    
    test_y_pred = clf.predict(sm.add_constant(X_test))
    test_AUC = roc_auc_score(y_test,test_y_pred)
    test_fpr, test_tpr, test_thresholds = roc_curve(y_test,test_y_pred)
    test_KS = max(test_tpr - test_fpr)
    print('Test AUC: {}'.format(test_AUC))
    print('Test KS: {}'.format(test_KS))
    
    fig = plt.figure(figsize=(10,10))
    plt.plot(train_fpr,train_tpr,color='darkorange',lw=3,label='Train ROC curve (Area = %0.2f)'%train_AUC)
    plt.plot(test_fpr,test_tpr,color='navy',lw=3,label='Test ROC curve (Area = %0.2f)'%test_AUC)
    plt.plot([0,1],[0,1],color='gray',lw=1,linestyle='--')
    plt.xlim([0.0,1.0])
    plt.ylim([0.0,1.05])
    plt.xlabel('False Positive Rate',fontsize=16)
    plt.ylabel('True Positive Rate',fontsize=16)
    plt.title('Train&Test ROC curve',fontsize=25)
    plt.legend(loc='lower right',fontsize=20)
    
    return clf









#==============================================================================
# File: init_logger.py
#==============================================================================

import os
import logging
from datetime import datetime
from typing import Optional
from config_loader import ROOT_DIR,get_project_root


logger = logging.getLogger(__name__)

PROJECT_ROOT = ROOT_DIR or get_project_root()

def init_logger(
    task_name: str = None,
    log_dir_suffix: str = None,
    log_filename: Optional[str] = None,
    level: int = logging.INFO
) -> logging.Logger:
    """
    初始化通用日志器，支持不同任务使用独立 logger 实例
    
    Parameters:
    -----------
    task_name : str
        任务名称，作为 logger 名称和日志文件的一部分，例如 'KS', 'PSI', 'AUC', 'FeatureMonitor'
    log_dir_suffix : str
        日志子目录名，例如 "ModelKSMonitor", "ModelPSI"
    log_filename : str, optional
        日志文件名，若未指定则使用 task_name + 日期
    level : int
        日志级别
    """
    # 日志目录：PROJECT_ROOT/logs/log_dir_suffix/
    log_dir = os.path.join(PROJECT_ROOT, "logs", log_dir_suffix)
    os.makedirs(log_dir, exist_ok=True)
    
    # 日志文件名：默认为 taskname_日期.log
    if log_filename is None:
        log_filename = f"{task_name.lower()}_{datetime.now().strftime('%Y%m%d')}.log"
    log_file_path = os.path.join(log_dir, log_filename)
    
    # 使用 task_name 作为 logger 名称，实现多任务隔离
    logger = logging.getLogger(f"ModelMonitor.{task_name}")
    logger.setLevel(level)
    logger.propagate = False  # 防止重复输出
    
    # 清理旧 handlers（避免重复）
    if logger.handlers:
        logger.handlers.clear()
    
    # --- 文件处理器 ---
    file_handler = logging.FileHandler(log_file_path, encoding="utf-8")
    file_handler.setLevel(level)
    file_formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(module)s:%(funcName)s:%(lineno)d - %(message)s"
    )
    file_handler.setFormatter(file_formatter)
    
    # --- 控制台处理器 ---
    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)
    console_formatter = logging.Formatter(
        "%(asctime)s - %(levelname)s - %(message)s"
    )
    console_handler.setFormatter(console_formatter)
    
    # 添加处理器
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # 记录初始化信息
    logger.info(f"✅ 日志系统初始化完成")
    logger.info(f"📁 日志文件路径: {log_file_path}")
    logger.info(f"📦 任务名称: {task_name}")
    logger.info(f"🚀 项目根目录: {PROJECT_ROOT}")
    
    return logger



#==============================================================================
# File: logger_utils.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: logger_utils.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-06-06
'''

import logging


# 日志输出
class Logger():
    # 日志级别关系映射
    level_relations = {
        "debug": logging.DEBUG,
        "info": logging.INFO,
        "warning": logging.WARNING,
        "error": logging.ERROR,
        "critical": logging.CRITICAL
    }

    def __init__(self, level="info", name=None,
                 fmt="%(asctime)s - %(name)s[line:%(lineno)d] - %"
                     "(levelname)s: %(message)s"):
        logging.basicConfig(level=self.level_relations.get(level), format=fmt)
        self.logger = logging.getLogger(name)



#==============================================================================
# File: M1A0030.py
#==============================================================================

import pandas as pd
import numpy as np
import toad
import lightgbm as lgb
from sklearn.metrics import roc_curve, roc_auc_score
import joblib
import pickle
import time
from datetime import datetime, timedelta
import os 
import gc
import warnings
warnings.filterwarnings("ignore")

# 获取数据
def get_data(sql):
    from odps import ODPS
    import time
    from datetime import datetime
    import multiprocessing
    
    # 获取cpu核的数量
    n_process = multiprocessing.cpu_count()
    # 输入账号密码
    conn= ODPS(username='liaoxilin', password='j02vYCxx')

    print('开始跑数：' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    start = time.time()
    # 执行脚本
    instance = conn.execute_sql(sql)
    # 输出执行结果
    with instance.open_reader() as reader:
        print('-------数据开始转换为DataFrame--------')
        data = reader.to_pandas(n_process=n_process) # 多核处理，避免单核处理

    end = time.time()
    print('结束跑数：' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("运行事件：{}秒".format(end-start))   

    return data


# 获取今天的日期
today = datetime.today()
# 计算昨天的日期
yesterday = today - timedelta(days=1)
yesterday = yesterday.strftime('%Y-%m-%d')

model_code='M1A0030'
file_path = f'/data/home/liaoxilin/模型部署/{model_code}'
model_file = '授信全渠道子分融合模型三期标签v2_v2_20250331211709.pkl'
table_name1 = 'lxl_model01_M1A0030_vars'
table_name2 = 'lxl_model01_M1A0030_score'



sql = f"""
select t.*
from znzz_fintech_ads.{table_name1} as t 
where dt=date_sub(current_date(),1)
  and order_no is not null
"""
df1 = get_data(sql)


def load_model_from_pkl(path):
    """
    从路径path加载模型
    :param path: 保存的目标路径
    """
    
    with open(path, 'rb') as f:
        model = pickle.load(f)
    return model

def getScore(prob):
    score = 450+ 50*np.log(prob/(1-prob))
    return score

lgb_model = load_model_from_pkl(file_path + f'/{model_file}')
lgb_model = lgb.Booster(model_str=lgb_model._handle)
varsname = lgb_model.feature_name()

for col in varsname:
    if df1[col].dtype == 'object':
        print(f'########{col}########')
        df1[col] = pd.to_numeric(df1[col])

df1['good_score'] = lgb_model.predict(df1[varsname], num_iteration=lgb_model.best_iteration)
df1['bad_score'] = 1 - df1['good_score']
df1['standard_score'] =  df1['good_score'].apply(getScore)
print("-----完成打分-------")
usecols = ['order_no','apply_date','good_score','bad_score','standard_score']
df2 = df1[usecols]


csv_file_list = []
# chunk_size 是每个拆分后 DataFrame 的行数
chunk_size = 8000000
# 计算需要拆分成多少个文件
num_chunks = len(df2) // chunk_size + (1 if len(df2) % chunk_size else 0)
for i in range(num_chunks):
    # 计算当前块的开始和结束索引
    start = i * chunk_size
    end = min((i + 1) * chunk_size, len(df2))
    # 拆分得到当前块的 DataFrame
    df_chunk = df2.iloc[start:end]
    # 生成文件名并保存
    filename = f'{model_code}_part_{i+1}.csv'
    csv_file_list.append(filename)
    # 将当前块保存为 CSV 文件
    df_chunk.to_csv(file_path + '/' + filename, index=False, header=None,sep='|')
    print(f'*****Saved {file_path}/{filename}*****')   

time.sleep(60)

from hl_data_mc_upload_v2_0 import DataUploadMc

upload = DataUploadMc(username='liaoxilin',
                      password='j02vYCxx',
                      env='prd')

upload.upload_data_to_table(    
        ## 字段名称
        fields='{"order_no":"string","apply_date":"string","good_score":"double","bad_score":"double","standard_score":"double"}',
        ## 本地文件，注意：只写文件名即可，参数是 list 类型
        csv_filename_list=csv_file_list,
        ## 本地文件路径，注意：需要本地的绝对路径
        input_path=f'{file_path}',                    
        ## 上传的数据库
        database='znzz_fintech_ads',        
        ## 上传的表名
        table_name=f'{table_name2}',
        # 自定义分隔符
        delimiter='|',
        partition=f'dt={yesterday}'
       )  



#==============================================================================
# File: main.py
#==============================================================================

import pandas as pd
import numpy as np
import os
import sys
import logging
from datetime import datetime,timedelta

from odps_utils import get_odps_client, execute_odps_sql,load_pivot_config,generate_and_execute_wide_table_query
from config_loader import ROOT_DIR,get_project_root,email_config
from execute_custom_query import read_custom_sql, execute_custom_query
from init_logger import init_logger
from ModelMonitor import ModelMonitor
from notification import EnhancedEmailSender

# --------------------------
# 1. 配置日志
# --------------------------
# 初始化路径和日志
LOG_DIR = os.path.join(ROOT_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
log_file = os.path.join(LOG_DIR, f"model_monitor_{datetime.now().strftime('%Y%m%d')}.log")

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def channel_type(x):
    if x in (209, 210, 213, 229, 233, 235, 236, 237, 240, 241, 244, 248,249,252,254,256,258,270,272,273,274,277,226, 227, 231, 234, 245, 246, 247, 251,263,265,267):
        channel='1_金科渠道'
    elif x==1:
        channel='3_桔子商城'
    else:
        channel='2_api渠道'
    return channel

def channel_rate(x): #(227,213,231,233,240,245,241,246)
    if x in (209, 210, 213, 229, 233, 235, 236, 237, 240, 241, 244, 248,249,252,254,256,258,270,272,273,274,277,226, 227, 231, 234, 245, 246, 247, 251,263,265,267):
        if x == 227:
            channel='227渠道'
        elif x in (209, 210, 213, 229, 233, 235, 236, 237, 240, 241, 244, 248,249,252,254,256,258,270,272,273,274,277):
            channel='24利率'
        elif x in (226, 227, 231, 234, 245, 246, 247, 251,263,265,267):
            channel='36利率'
        else:
            channel=None
    else:
        channel=None

    return channel

# --------------------------
# 主流程：加载数据 → 初始化监控器 → 批量处理 → 输出结果
# --------------------------
def main():
    try:
        logger.info("=" * 50)
        logger.info(f"通过拒绝KS和PSI监控流程启动 | 跑数日期：{datetime.now().strftime('%Y-%m-%d')} ")
        logger.info("=" * 50)

        # --------------------------
        # 1. 配置日期参数
        # --------------------------
        # 获取今天的日期
        today = datetime.today()
        # 计算昨天的日期
        run_day = today - timedelta(days=1)
        current_date = run_day.strftime('%Y-%m-%d')
        # --------------------------
        # 2. 获取输入数据
        # --------------------------
        logger.info("\n【1/4】开始加载数据...")
        # 授信场景
        variable_codes = load_pivot_config('授信', current_date)
        df = generate_and_execute_wide_table_query(
                                                    biz_type='授信',
                                                    start_date='2025-01-01',
                                                    end_date=run_day,
                                                    variable_codes=variable_codes,
                                                    main_table="znzz_fintech_ads.apply_model01_scores_off",
                                                    join_table="znzz_fintech_ads.lxl_model_auth_tags",
                                                    date_col_main= "apply_time",
                                                    date_col_join="dt",
                                                    key_column="order_no",
                                                    metric_column="good_score",
                                                    config_current_dt=run_day
        )
        df['channel_types']=df['channel_id'].apply(channel_type)
        df['channel_rates']=df['channel_id'].apply(channel_rate)
        
        model_ids = [col.lower() for col in variable_codes]

        # --------------------------
        # 3. 实例化监控器
        # -------------------------- 
        logger.info("\n【2/4】初始化模型监控器...")
        monitor = ModelMonitor()
        
        logger.info("\n【3/4】开始批量处理模型...")
        result = monitor.process_batch(df, model_ids, stat_date=current_date)

        # --------------------------
        # 4. 输出处理结果
        # --------------------------
        logger.info(f"\n✅ 批量处理完成")
        logger.info(f"   成功：{len(result['success'])} 个模型")
        logger.info(f"   失败：{len(result['fail'])} 个模型")

        if result['success']:
            logger.info("\n✅ 成功模型及报告路径：")
            for item in result['success']:
                logger.info(f"   - {item['model_id']} → {item['report_path']}")

        if result['fail']:
            logger.info("\n❌ 失败详情：")
            for item in result['fail']:
                logger.info(f"   - {item['model_id']}: {item['error']}")                                           


        # --------------------------
        # 5. 汇总并保存所有报告
        # --------------------------
        logger.info("\n📥 正在加载所有生成的报告...") 
        report_dir = monitor.summary_excel_reports()

        # --------------------------
        # 6. 发送邮件email
        # --------------------------
        logger.info("\n📥 正在发送邮件...")
        sender = EnhancedEmailSender(email_config)
        attachment_paths = [report_dir]
        
        sender.send_generic_email(
            subject_template="通过拒绝KS和PSI的监控报告-{date}",
            body_template="""
            <h3>通过拒绝KS和PSI</h3>
            <p>日期: {date}</p>
            <p>业务场景: {biz_type}</p>
            <p>详情请见附件</p>
            """,
            body_vars={"date": current_date, "biz_type": "授信"},
            # 可以添加现有文件作为附件
            attachment_paths=attachment_paths,
            # 添加DataFrame作为附件（会自动转换为Excel并美化）
            df_attachments=None )  
                                           
        logger.info("\n" + "=" * 50)
        logger.info(f"流程结束 | 报告目录：{report_dir}")
        logger.info("=" * 50)
                                        
    except Exception as e:
        logger.error(f"\n 通过拒绝KS和PSI监控流程执行失败：{str(e)}", exc_info=True)
        raise 

# --------------------------
# 5. 入口函数（仅当直接运行 main.py 时执行）
# --------------------------
if __name__ == "__main__":
    main()


#==============================================================================
# File: metrics.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: metrics.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-11-05
'''

import numpy as np
import pandas as pd
from sklearn.metrics import roc_auc_score

from .utils import del_df, unpack_tuple


# auc
def get_auc(target, y_pred):
    """
    计算auc值
    Args:
        target (array-like): 目标变量列表
        y_pred (array-like): 模型预测的分数或概率列表

    Returns:
        float: auc值
    """
    if len(np.unique(target)) != 2:
        raise ValueError('the target is not 2 classier target')
    else:
        return roc_auc_score(target, y_pred)


# ks
def get_ks(target, y_pred):
    """
    计算ks值
    Args:
        target (array-like): 目标变量列表
        y_pred (array-like): 模型预测的分数或概率列表

    Returns:
        float: ks值
    """
    df = pd.DataFrame({
        'y_pred': y_pred,
        'target': target,
    })
    crossfreq = pd.crosstab(df['y_pred'], df['target'])
    crossdens = crossfreq.cumsum(axis=0) / crossfreq.sum()
    crossdens['ks'] = abs(crossdens[0] - crossdens[1])
    ks = max(crossdens['ks'])
    return ks


def get_auc_ks_psi(df, target='target', pred='p'):
    """
    计算auc、ks、psi
    Args:
        df:
        target:
        pred:

    Returns:

    """
    df_pred_dev = df[df['type'] == 'train']
    df_pred_nodev = df[df['type'] == 'test']

    # 计算auc、ks、psi
    test_ks = get_ks(df_pred_nodev[target], df_pred_nodev[pred])
    train_ks = get_ks(df_pred_dev[target], df_pred_dev[pred])
    test_auc = get_auc(df_pred_nodev[target], df_pred_nodev[pred])
    train_auc = get_auc(df_pred_dev[target], df_pred_dev[pred])

    q_cut_list = np.arange(0, 1, 1 / 20)
    bins = np.append(np.unique(np.quantile(df_pred_nodev[pred], q_cut_list)), df_pred_nodev[pred].max() + 0.1)
    df_pred_nodev['range'] = pd.cut(df_pred_nodev[pred], bins=bins, precision=0, right=False).astype(str)
    df_pred_dev['range'] = pd.cut(df_pred_dev[pred], bins=bins, precision=0, right=False).astype(str)
    nodev_psi = psi(df_pred_nodev['range'], df_pred_dev['range'])
    res_dict = {'dev_auc': train_auc, 'nodev_auc': test_auc, 'dev_ks': train_ks, 'nodev_ks': test_ks,
                'nodev_dev_psi': nodev_psi}
    del_df(df_pred_dev)
    del_df(df_pred_nodev)
    return pd.DataFrame([res_dict])


def psi(no_base, base, return_frame=False, featurebin=None):
    """
    计算psi值
    Args:
        no_base (DataFrame|array-like): 非基准数据集
        base (DataFrame|array-like):基准数据集
        return_frame (bool): 是否需要返回占比

    Returns:
        float|Series:psi值
    """

    if featurebin is not None:
        if isinstance(featurebin, (dict, list)):
            from .transformer import FeatureBin
            featurebin = FeatureBin().manual_bin(featurebin)

            no_base = featurebin.transform(no_base, labels=True)
        base = featurebin.transform(base, labels=True)

    psi = list()
    frame = list()

    if isinstance(no_base, pd.DataFrame):
        for col in no_base:
            p, f = calc_psi(no_base[col], base[col])
            psi.append(p)
            frame.append(f)

        psi = pd.Series(psi, index=no_base.columns, name='psi')

        frame = pd.concat(
            frame,
            keys=no_base.columns,
            names=['columns', 'id'],
        ).reset_index()
        frame = frame.drop(columns='id')
    else:
        psi, frame = calc_psi(no_base, base)

    res = (psi,)

    if return_frame:
        res += (frame,)

    return unpack_tuple(res)


def calc_psi(no_base, base):
    """
    psi计算的具体逻辑
    Args:
        no_base (array-like): 非基准数据集
        base (array-like): 基准数据集

    Returns:
        float,DataFrame : psi值，占比
    """
    no_base_prop = pd.Series(no_base).value_counts(normalize=True, dropna=False)
    base_prop = pd.Series(base).value_counts(normalize=True, dropna=False)

    psi = np.sum((no_base_prop - base_prop) * np.log(no_base_prop / base_prop))

    frame = pd.DataFrame({
        'no_base': no_base_prop,
        'base': base_prop,
    })
    frame.index.name = 'value'

    return psi, frame.reset_index()


def SSE(y_pred, y):
    """sse
    """
    return np.sum((y_pred - y) ** 2)


def MSE(y_pred, y):
    """mse
    """
    return np.mean((y_pred - y) ** 2)


def AIC(y_pred, y, k, llf=None):
    """AIC信息准则

    Args:
        y_pred (array-like)
        y (array-like)
        k (int): x的数量
        llf (float): 对数似然函数的值
    """
    if llf is None:
        llf = np.log(SSE(y_pred, y))

    return 2 * k - 2 * llf


def BIC(y_pred, y, k, llf=None):
    """贝叶斯信息准则

    Args:
        y_pred (array-like)
        y (array-like)
        k (int): x的数量
        llf (float): 对数似然函数的值
    """
    n = len(y)
    if llf is None:
        llf = np.log(SSE(y_pred, y))

    return np.log(n) * k - 2 * llf


def get_metrics_info(df, by=['apply_mon'], feature_type='td', target='target', data_type='type',
                     apply_time='apply_time'):
    if data_type not in df.columns:
        data_type = 'None'
        df[data_type] = 'None'
    auc_ks = df.groupby(by).apply(
        lambda df_tmp: pd.Series({
            'auc': get_auc(df_tmp[target], df_tmp[feature_type]),
            'ks': get_ks(df_tmp[target], df_tmp[feature_type]),
            '正样本': sum(df_tmp[target]),
            '总数': len(df_tmp),
            '正样本占比': np.mean(df_tmp[target]),
            'apply_time': f"{df_tmp[apply_time].min()}至{df_tmp[apply_time].max()}",
            'data_type': f"{list(set(df_tmp[data_type]))}",
        })
    )
    all_auc_ks = pd.DataFrame([{'all': 'all', 'auc': get_auc(df[target], df[feature_type]),
                                'ks': get_ks(df[target], df[feature_type]), '正样本': sum(df[target]), '总数': len(df),
                                '正样本占比': np.mean(df[target]),
                                'apply_time': f"{df[apply_time].min()}至{df[apply_time].max()}",
                                'data_type': f"{list(set(df[data_type]))}", }]).set_index('all')
    all_auc_ks.index.name = auc_ks.index.name
    res = auc_ks.append(all_auc_ks)
    res.index.name = feature_type
    return res


def psi_by_col(df, by_col='apply_mon'):
    by_col_v = sorted(list(set(df[by_col])))
    by_col_psi_lis = []
    for n, j in enumerate(by_col_v):
        by_col_d = df[df[by_col] == j]
        ###计算PSI
        by_col_psi = psi(by_col_d[['score']], df[['score']], )
        by_col_psi.name = f"{j}_PSI"
        by_col_psi_lis.append(by_col_psi)
    by_col_psi_df = pd.DataFrame(by_col_psi_lis).T
    by_col_psi_df['MaxPSI'] = by_col_psi_df.max(axis=1)
    return by_col_psi_df



#==============================================================================
# File: ModelMonitor.py
#==============================================================================

from pathlib import Path 
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import toad
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Union
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill)
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image
from matplotlib.font_manager import FontProperties

from config_loader import ROOT_DIR, get_project_root
from init_logger import init_logger

# 项目根目录（全局变量，后续路径基于此计算）
PROJECT_ROOT = ROOT_DIR

# --------------------------
# 1. 日志配置
# --------------------------
logger = init_logger(task_name='ModelMonitor', log_dir_suffix='ModelMonitor')


# --------------------------
# 2. 模型监控类
# --------------------------
class ModelMonitor:
    def __init__(
            self,
            default_bins: int = 10,
            default_binning_strategy: str = 'quantile',
            psi_default_threshold: float = 0.1,
            mean_diff_default_threshold: float = 0.001,
            custom_thresholds: Optional[Dict[str, Dict]] = None,
            report_dir: Optional[str] = None
    ):
        self._validate_binning_strategy(default_binning_strategy)
        self.default_binning_strategy = default_binning_strategy
        self.default_bins = default_bins
        self.psi_default_threshold = psi_default_threshold
        self.mean_diff_default_threshold = mean_diff_default_threshold
        self.custom_thresholds = custom_thresholds or {}

        if report_dir is None:
            self.report_dir = os.path.join(PROJECT_ROOT, "reports", "ModelMonitor")
        else:
            self.report_dir = os.path.join(report_dir, "ModelMonitor")
        os.makedirs(self.report_dir, exist_ok=True)

        self.current_baseline_month = None
        self.rolling_9m_months = []
        self.models_result = {}

        logger.info(
            f"ModelMonitor初始化完成 | "
            f"分箱数：{self.default_bins} | 分箱策略：{self.default_binning_strategy} | "
            f"报告目录：{self.report_dir} | 日志目录：{os.path.dirname(logger.handlers[0].baseFilename)}"
        )

    # ==========================
    # 公共接口方法（入口）
    # ==========================

    def process_single_model(self, df: pd.DataFrame, model_id: str, model_ids: List[str], stat_date: str) -> Dict:
        logger.info(f"开始处理模型：{model_id} | 统计日期：{stat_date}")
        cols = model_ids[:]
        data = df[[col for col in df.columns if col not in cols] + [model_id]]
        if data[model_id].dtype == 'object':
            data[model_id] = pd.to_numeric(data[model_id], errors='coerce')
        data = self._validate_input_data(data, model_id)
        data['month'] = data['apply_date'].dt.strftime('%Y%m')

        self._determine_baseline_month(data)
        baseline_start, baseline_end = self._get_month_date_range(self.current_baseline_month)
        baseline_data = data[(data['apply_date'] >= baseline_start) & (data['apply_date'] <= baseline_end)]
        if len(baseline_data) < 100:
            logger.warning(f"固定基期样本量仅{len(baseline_data)}（建议≥100），分布对比结果可能偏差")

        monthly_data = {}
        for month in self.rolling_9m_months:
            month_start, month_end = self._get_month_date_range(month)
            monthly_data[month] = data[(data['apply_date'] >= month_start) & (data['apply_date'] <= month_end)]
            if len(monthly_data[month]) == 0:
                logger.warning(f"月份[{month}]无数据，跳过")
                del monthly_data[month]

        groupby_fields_list = [['channel_types', 'month'],['channel_types', 'customer_tags', 'month']]

        ks_df = self._process_ks_results(data, model_id,'target_ar',groupby_fields_list)
        
        overall_dist = self._generate_monthly_distribution(data, model_id, stat_date)
        derived_tables = self._generate_derived_tables(overall_dist)
        # 通过客群
        pass_data = data[data["auth_status"]==6]
        pass_overall_dist = self._generate_monthly_distribution(pass_data, model_id, stat_date)
        pass_derived_tables = self._generate_derived_tables(pass_overall_dist)
        
        stability_df = self._calculate_monthly_stability(
            baseline_data=baseline_data,
            monthly_data=monthly_data,
            model_id=model_id,
            stat_date=stat_date
        )

        report_path = self._generate_excel_report(
            model_id=model_id,
            stat_date=stat_date,
            ks_df=ks_df,
            stability_df=stability_df,
            overall_dist=overall_dist,
            derived_tables=derived_tables
        )

        result = {
            'stability': stability_df,
            'overall_dist': overall_dist,
            'mean_comparison': derived_tables[0],
            'ratio_comparison': derived_tables[1],
            'count_comparison': derived_tables[2],
            'pass_overall_dist':pass_overall_dist,
            'pass_ratio_comparison': pass_derived_tables[1],
            'pass_count_comparison': pass_derived_tables[2],
            'ks_result': ks_df, 
            'report_path': report_path
        }
        self.models_result[model_id] = result
        return result

    def process_batch(self,data: pd.DataFrame,model_ids: List[str],stat_date: Optional[str] = None) -> Dict:
        stat_date = stat_date or datetime.now().strftime('%Y%m%d')
        validated_data = self._validate_input_data(data,model_ids)
        self.models_result = {}
        result = {'success': [], 'fail': []}

        logger.info(f"开始批量处理 | 模型数量：{len(model_ids)} | 统计日期：{stat_date}")
        models_columns = model_ids[:]
        for model_id in model_ids:
            try:
                # 过滤出该 model_id 的数据
                model_data = validated_data[[col for col in validated_data.columns if col not in models_columns] + [model_id]]
                if model_data.empty:
                    raise ValueError(f"模型 [{model_id}] 在输入数据中无对应记录")

                model_result = self.process_single_model(model_data, model_id, models_columns, stat_date)
                result['success'].append({
                    'model_id': model_id,
                    'report_path': model_result['report_path'],
                    'status': '完成'
                })
            except Exception as e:
                error_msg = str(e)
                logger.error(f"模型[{model_id}]处理失败：{error_msg}", exc_info=True)
                result['fail'].append({
                    'model_id': model_id,
                    'error': error_msg
                })

        total = len(model_ids)
        success_count = len(result['success'])
        fail_count = len(result['fail'])
        logger.info(
            f"批量处理完成 | 总模型数：{total} | 成功：{success_count} | 失败：{fail_count} | "
            f"报告目录：{self.report_dir}"
        )
        return result

    def get_model_result(self, model_id: str) -> Optional[Dict]:
        return self.models_result.get(model_id)

    def get_all_results(self) -> Dict:
        return self.models_result

    def export_all_reports(self):
        """
        读取每个模型对应的报告文件，返回一个字典：
        key: model_id
        value: 对应报告的DataFrame
        """
        all_results = self.get_all_results()
        reports_dfs = {}

        for model_id, result in all_results.items():
            report_path = result.get('report_path')
            if not report_path:
                logger.warning(f"模型 {model_id} 没有生成报告路径。")
                reports_dfs[model_id] = None
                continue

            path = Path(report_path)
            if not path.exists():
                logger.warning(f"报告文件不存在: {report_path}")
                reports_dfs[model_id] = None
                continue

            try:
                if path.suffix.lower() == '.csv':
                    df = pd.read_csv(path)
                elif path.suffix.lower() in ['.xlsx', '.xls']:
                    df = pd.read_excel(path, sheet_name=None)
                else:
                    logger.error(f"不支持的文件格式: {path.suffix}")
                    reports_dfs[model_id] = None
                    continue

                reports_dfs[model_id] = df
                logger.info(f"成功加载模型 {model_id} 的报告: {report_path}")

            except Exception as e:
                logger.error(f"读取模型 {model_id} 的报告时出错: {e}")
                reports_dfs[model_id] = None

        logger.info(f"共成功加载 {len([df for df in reports_dfs.values() if df is not None])} 个报告的DataFrame。")
        return reports_dfs

    def summary_excel_reports(self):
        """
        生成Excel格式的KS和PSI监控报告
        
        Returns:
            str: 生成的报告文件路径
        """
        logger.info("开始生成KS和PSI监控报告")
        
        try:
            # 验证必要的属性是否存在
            if not hasattr(self, 'models_result'):
                logger.error("对象缺少models_result属性")
                raise AttributeError("对象必须包含models_result属性")
                
            if not hasattr(self, 'report_dir'):
                logger.error("对象缺少report_dir属性")
                raise AttributeError("对象必须包含report_dir属性")
                
            # 验证报告目录是否存在，不存在则创建
            if not os.path.exists(self.report_dir):
                logger.warning(f"报告目录不存在，将创建: {self.report_dir}")
                try:
                    os.makedirs(self.report_dir, exist_ok=True)
                    logger.info(f"成功创建报告目录: {self.report_dir}")
                except OSError as e:
                    logger.error(f"创建报告目录失败: {str(e)}", exc_info=True)
                    raise
            
            result = self.models_result
            if not isinstance(result, dict):
                logger.error(f"models_result应为字典类型，实际为: {type(result)}")
                raise TypeError("models_result必须是字典类型")
                
            if not result:
                logger.warning("models_result为空，可能导致报告内容为空")
            
            model_columns = list(result.keys())
            logger.info(f"发现{len(model_columns)}个模型需要处理")
            
            psi_list = []
            ks_list = []
            ratio_comparison_list = []
            count_comparison_list = [] 
            pass_ratio_comparison_list = []
            pass_count_comparison_list = []
            
            for idx, model_id in enumerate(model_columns):
                try:
                    model_data = result[model_id]
                    # 验证模型数据是否包含必要的键
                    required_keys = ['stability', 'ks_result', 'ratio_comparison', 'count_comparison','pass_ratio_comparison','pass_count_comparison']
                    missing_keys = [key for key in required_keys if key not in model_data]
                    if missing_keys:
                        logger.warning(f"模型{model_id}缺少必要的键: {missing_keys}，将跳过该模型")
                        continue
                    
                    # 验证数据是否为DataFrame
                    if not isinstance(model_data['stability'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的stability不是DataFrame类型，将跳过")
                        continue
                        
                    if not isinstance(model_data['ks_result'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的ks_result不是DataFrame类型，将跳过")
                        continue
                        
                    if not isinstance(model_data['ratio_comparison'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的ratio_comparison不是DataFrame类型，将跳过")
                        continue
                        
                    if not isinstance(model_data['count_comparison'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的count_comparison不是DataFrame类型，将跳过")
                        continue
                        
                    if not isinstance(model_data['pass_ratio_comparison'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的pass_ratio_comparison不是DataFrame类型，将跳过")
                        continue  
                        
                    if not isinstance(model_data['pass_count_comparison'], pd.DataFrame):
                        logger.warning(f"模型{model_id}的pass_count_comparison不是DataFrame类型，将跳过")
                        continue
                        
                    psi_list.append(model_data['stability'])
                    ks_list.append(model_data['ks_result'])
                    ratio_comparison_list.append(model_data['ratio_comparison'])
                    count_comparison_list.append(model_data['count_comparison'])
                    pass_ratio_comparison_list.append(model_data['pass_ratio_comparison'])
                    pass_count_comparison_list.append(model_data['pass_count_comparison'])                    
                    logger.debug(f"已处理模型 {idx+1}/{len(model_columns)}: {model_id}")
                    
                except Exception as e:
                    logger.error(f"处理模型{model_id}时发生错误: {str(e)}", exc_info=True)
                    continue
            
            # 验证数据列表是否为空
            if not psi_list:
                logger.warning("psi_list为空，PSI相关表格将为空")
            if not ks_list:
                logger.warning("ks_list为空，KS相关表格将为空")
            if not ratio_comparison_list:
                logger.warning("ratio_comparison_list为空，各月人数占比表格将为空")
            if not count_comparison_list:
                logger.warning("count_comparison_list为空，各月人数表格将为空")
            if not pass_ratio_comparison_list:
                logger.warning("pass_ratio_comparison_list为空，通过客群各月人数占比表格将为空")
            if not pass_count_comparison_list:
                logger.warning("pass_count_comparison_list为空，通过客群各月人数表格将为空")
                
            # 合并数据
            logger.info("开始合并数据")
            df_psi = pd.concat(psi_list, axis=0, ignore_index=True) if psi_list else pd.DataFrame()
            df_ks = pd.concat(ks_list, axis=0, ignore_index=True) if ks_list else pd.DataFrame()
            df_ratio = pd.concat(ratio_comparison_list, axis=0, ignore_index=False) if ratio_comparison_list else pd.DataFrame()
            df_count = pd.concat(count_comparison_list, axis=0, ignore_index=False) if count_comparison_list else pd.DataFrame()
            df_ratio_pass = pd.concat(pass_ratio_comparison_list, axis=0, ignore_index=False) if pass_ratio_comparison_list else pd.DataFrame()
            df_count_pass = pd.concat(pass_count_comparison_list, axis=0, ignore_index=False) if pass_count_comparison_list else pd.DataFrame()            
            logger.info("数据合并完成")
            
            # 创建透视表
            logger.info("开始创建透视表")
            psi_pivot = pd.pivot_table(
                df_psi,
                index=['月份'],
                columns='model_id',
                values='PSI值',
                aggfunc='first',
                dropna=True
            ).reset_index() if not df_psi.empty else pd.DataFrame()
            
            ks_pivot = pd.pivot_table(
                df_ks,
                index=['channel', 'customer_tags', 'month','样本量','坏样本量','坏样本占比'],
                columns='model_id',
                values='KS',
                aggfunc='first',
                dropna=True
            ).reset_index() if not df_ks.empty else pd.DataFrame()
            ks_pivot=ks_pivot.sort_values(by=['customer_tags', 'channel', 'month'])
            logger.info("透视表创建完成")
            
            # 生成报告文件名和路径
            stat_date = datetime.now().strftime('%Y%m%d')
            report_filename = f"KS_PSI_监控报告_{stat_date}.xlsx"
            report_path = os.path.join(self.report_dir, report_filename)
            logger.info(f"报告将保存至: {report_path}")
            
            # 写入Excel文件
            logger.info("开始写入Excel文件")
            report_filename2 = f"KS_PSI_监控报告_detail_{stat_date}.xlsx"
            report_path2 = os.path.join(self.report_dir, report_filename2)
            with pd.ExcelWriter(report_path2, engine='openpyxl') as writer:
                ks_pivot.to_excel(writer, sheet_name='通过拒绝KS', index=False)
                psi_pivot.to_excel(writer, sheet_name='PSI', index=False)
                df_ratio.to_excel(writer, sheet_name='各月人数占比', index=False)
                df_count.to_excel(writer, sheet_name='各月人数', index=False)
                df_ratio_pass.to_excel(writer, sheet_name='通过客群各月人数占比', index=False)
                df_count_pass.to_excel(writer, sheet_name='通过客群各月人数', index=False)                
                df_ks.to_excel(writer, sheet_name='模型KS汇总表', index=False)
                df_psi.to_excel(writer, sheet_name='模型PSI汇总表', index=False)
            
            if 'stat_date' in df_ratio.columns:
                df_ratio.drop(columns=['stat_date'],inplace=True)
            if 'stat_date' in df_count.columns:
                df_count.drop(columns=['stat_date'],inplace=True)
            if 'stat_date' in df_ratio_pass.columns:
                df_ratio_pass.drop(columns=['stat_date'],inplace=True)
            if 'stat_date' in df_count_pass.columns:
                df_count_pass.drop(columns=['stat_date'],inplace=True)                
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                ks_pivot.to_excel(writer, sheet_name='通过拒绝KS', index=False)
                psi_pivot.to_excel(writer, sheet_name='PSI', index=False)
                df_ratio.to_excel(writer, sheet_name='各月人数占比', index=False)
                df_count.to_excel(writer, sheet_name='各月人数', index=False)   
                df_ratio_pass.to_excel(writer, sheet_name='通过客群各月人数占比', index=False)
                df_count_pass.to_excel(writer, sheet_name='通过客群各月人数', index=False)                   
            logger.info("Excel文件写入完成")
            
            # 美化Excel表格
            logger.info("开始美化Excel表格")
            try:
                wb = load_workbook(report_path)
                
                # 验证样式获取方法是否存在
                if not hasattr(self, '_get_excel_styles') or not callable(self._get_excel_styles):
                    logger.warning("对象缺少_get_excel_styles方法，将使用默认样式")
                    # 可以在这里定义默认样式作为备选
                    styles = {
                        'thin_border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                                           top=Side(style='thin'), bottom=Side(style='thin')),
                        'header_font': Font(bold=True),
                        'normal_font': Font(),
                        'header_fill': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    }
                else:
                    styles = self._get_excel_styles()
                
                # 验证列宽调整方法是否存在
                adjust_column_width = self._adjust_column_width if hasattr(self, '_adjust_column_width') and callable(self._adjust_column_width) else self._default_adjust_column_width
                
                for sheet_name in ['通过拒绝KS', 'PSI', '各月人数占比', '各月人数','通过客群各月人数占比','通过客群各月人数']:
                    if sheet_name not in wb.sheetnames:
                        logger.warning(f"工作表{sheet_name}不存在，将跳过美化")
                        continue
                        
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        for cell in row:
                            cell.border = styles.get('thin_border')
                            cell.font = styles.get('header_font') if cell.row == 1 else styles.get('normal_font')
                            if cell.row == 1:
                                cell.fill = styles.get('header_fill')
                    adjust_column_width(ws)
                    logger.debug(f"已美化工作表: {sheet_name}")
                
                wb.save(report_path)
                logger.info("Excel表格美化完成")
                
            except Exception as e:
                logger.error(f"美化Excel表格时发生错误: {str(e)}", exc_info=True)
                logger.warning("将返回未美化的报告文件")
            
            logger.info(f"报告生成完成 | 报告路径：{report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"生成报告时发生严重错误: {str(e)}", exc_info=True)
            raise  # 重新抛出异常，让调用者知道发生了错误

    # ==========================
    # 内部流程方法（按执行顺序）
    # ==========================

    def _validate_input_data(self, data: pd.DataFrame, model_id: Union[str, List[str]]) -> pd.DataFrame:
        required_cols = ["channel_id", "apply_date", "order_no", "target_ar", "customer_tags"]
        missing_cols = [col for col in required_cols if col not in data.columns]
        if missing_cols:
            raise ValueError(f"数据缺少必填列：{', '.join(missing_cols)}")
        if not isinstance(model_id, list):
            model_id = [model_id]
        if not pd.api.types.is_datetime64_any_dtype(data['apply_date']):
            data['apply_date'] = pd.to_datetime(data['apply_date'], errors='raise')
            logger.debug("自动将'apply_date'列转换为datetime格式")

        dup_mask = data.duplicated(subset=['order_no', 'apply_date'])
        if dup_mask.any():
            raise ValueError(f"存在{dup_mask.sum()}条重复数据（UUID+model_id+日期需唯一）")
      
        out_of_range = 0
        for col in model_id:
            if col not in data.columns:
                raise ValueError(f"Column '{col}' not found in data.")
            out_of_range += ((data[col] < 0) | (data[col] > 1)).sum()
        
        if out_of_range > 0:
            logger.warning(f"发现{out_of_range}条分数超出[0,1]范围，可能影响分箱结果")

        return data

    def _determine_baseline_month(self, data: pd.DataFrame) -> str:
        data['month'] = data['apply_date'].dt.strftime('%Y%m')
        data_months = set(data['month'].unique())
        max_date = data['apply_date'].max()
        self.rolling_9m_months = self._get_rolling_9m_months(max_date)
        logger.info(f"近9个月范围：{self.rolling_9m_months}")

        for month in self.rolling_9m_months:
            if month in data_months:
                self.current_baseline_month = month
                logger.info(f"确定固定基期月份：{month}（用于分布对比）")
                return month

        raise ValueError(f"近9个月（{self.rolling_9m_months}）无有效数据，无法确定固定基期")

    def _get_rolling_9m_months(self, end_date: datetime) -> List[str]:
        months = []
        for i in range(9):
            year = end_date.year
            month = end_date.month - i
            if month <= 0:
                year -= 1
                month += 12
            months.append(f"{year}{month:02d}")
        return sorted(months)

    def _get_month_date_range(self, month_tag: str) -> Tuple[datetime, datetime]:
        year = int(month_tag[:4])
        month = int(month_tag[4:])
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end = datetime(year, month + 1, 1) - timedelta(days=1)
        return start, end


    def _validate_binning_strategy(self, strategy: str):
        valid_strategies = ['quantile', 'step']
        if strategy not in valid_strategies:
            raise ValueError(f"无效分箱策略：{strategy}（支持：{valid_strategies}）")

    def _generate_monthly_distribution(self, df: pd.DataFrame, model_id: str, stat_date: str) -> pd.DataFrame:
        data = df.copy()
        c = toad.transform.Combiner()
        c.fit(data[[model_id,'target_ar']], y='target_ar', method=self.default_binning_strategy, n_bins=self.default_bins, empty_separate=True) 
        data['分箱区间'] = c.transform(data[model_id], labels=True)
        data['分箱区间'] = data['分箱区间'].astype(str)

        data['model_id'] = model_id
        data['stat_date'] = stat_date

        overall_dist = data.groupby(
            ['model_id', 'stat_date', 'month', '分箱区间'],
            observed=True
        ).agg(
            人数=('order_no', 'count'),
            分数均值=(model_id, 'mean')
        ).reset_index()

        total_by_month = overall_dist.groupby('month')['人数'].transform('sum')
        overall_dist['占比'] = (overall_dist['人数'] / total_by_month).round(4)
        overall_dist = overall_dist[overall_dist['人数'] > 0].reset_index(drop=True)

        for month in overall_dist['month'].unique():
            month_dist = overall_dist[overall_dist['month'] == month]
            self._validate_bin_ratio(month_dist, model_id, month)

        return overall_dist[['model_id', 'stat_date', 'month', '分箱区间', '人数', '分数均值', '占比']]

    def _validate_bin_ratio(self, dist_df: pd.DataFrame, model_id: str, month: str):
        min_ratio = 0.05
        max_ratio = 0.20
        invalid_bins = dist_df[(dist_df['占比'] < min_ratio) | (dist_df['占比'] > max_ratio)]
        
        if not invalid_bins.empty:
            warn_msg = f"模型[{model_id}]月份[{month}]存在占比异常分箱：\n"
            for _, row in invalid_bins.iterrows():
                warn_msg += f"- 分箱{row['分箱区间']}：占比{row['占比']:.4f}（需{min_ratio}-{max_ratio}）\n"
            logger.warning(warn_msg)

    def _generate_derived_tables(self, overall_dist_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        overall_dist = overall_dist_df.copy()
        for col in overall_dist.columns:
            if pd.api.types.is_categorical_dtype(overall_dist[col]):
                overall_dist[col] = overall_dist[col].astype(str)
                logger.debug(f"将列[{col}]从 Categorical 转为字符串类型")

        mean_pivot = pd.pivot_table(
            overall_dist,
            index=['model_id', 'stat_date',  '分箱区间'],
            columns='month',
            values='分数均值',
            aggfunc='first',
            dropna=True
        ).reset_index()
        mean_pivot.columns = [
            col if col in ['model_id', 'stat_date',  '分箱区间'] 
            else f"{col}的分数均值" 
            for col in mean_pivot.columns
        ]

        ratio_pivot = pd.pivot_table(
            overall_dist,
            index=['model_id', 'stat_date',  '分箱区间'],
            columns='month',
            values='占比',
            aggfunc='first',
            dropna=True
        ).reset_index()
        ratio_pivot.columns = [
            col if col in ['model_id', 'stat_date',  '分箱区间'] 
            else f"{col}的人数占比" 
            for col in ratio_pivot.columns
        ]

        count_pivot = pd.pivot_table(
            overall_dist,
            index=['model_id', 'stat_date',  '分箱区间'],
            columns='month',
            values='人数',
            aggfunc='first',
            dropna=True
        ).reset_index()
        count_pivot.columns = [
            col if col in ['model_id', 'stat_date',  '分箱区间'] 
            else f"{col}的人数" 
            for col in count_pivot.columns
        ]

        return mean_pivot.fillna(0), ratio_pivot.fillna(0), count_pivot.fillna(0)

    def _calculate_psi(self, df_expected: pd.DataFrame, df_actual: pd.DataFrame, model_id:str) -> float:
        expected = df_expected.dropna()
        # expected = df_expected.copy()
        actual = df_actual.dropna()
        # actual = df_actual.copy()
        if expected.empty or actual.empty:
            logger.warning(f"警告: 模型 {model_id} 的 PSI 输入数据为空")
            return np.nan
        # 检查数据是否为空或全为 NaN
        if expected[model_id].empty or actual[model_id].empty:
            logger.warning(f"警告: 模型 {model_id} 的 PSI 输入数据为空")
            return np.nan        
        if len(expected[model_id]) == 0 or len(actual[model_id]) == 0:
            logger.warning(f"警告: 模型 {model_id} 的 PSI 输入数据全为 NaN")
            return np.nan
        
        try:
            # 初始化分箱器
            c = toad.transform.Combiner()
            c.fit(
                expected, 
                y='target_ar',
                method=self.default_binning_strategy, 
                n_bins=self.default_bins, 
                empty_separate=True
            )
            
            # 计算 PSI
            psi_value = toad.metrics.PSI(expected[model_id], actual[model_id], combiner=c)
            return round(psi_value, 4)
        
        except Exception as e:
            logger.error(f"模型 {model_id} 的 PSI 计算失败: {str(e)}")
            return np.nan

    def _calculate_ks_by_group(
        self,
        data: pd.DataFrame,
        model_id: str,
        y_label: str = 'target_ar',
        group_cols: Optional[List[str]] = None,
    ) -> pd.DataFrame:

        # 1. 检查必要字段是否存在
        required_cols = [model_id, y_label]
        if group_cols:
            required_cols.extend(group_cols)
        
        missing_cols = [col for col in required_cols if col not in data.columns]
        if missing_cols:
            logger.error(
                f"模型 {model_id} 计算 KS 失败：数据中缺少必要字段 {missing_cols}。 "
                f"数据包含的字段: {list(data.columns)}"
            )
            if group_cols:
                empty_df = pd.DataFrame(columns=['model_id'] + group_cols + ['样本量',  '坏样本量', '坏样本占比', 'KS'])
            else:
                empty_df = pd.DataFrame(columns=['model_id', '样本量',  '坏样本量', '坏样本占比', 'KS'])
            return empty_df

        # 2. 提取数据副本（避免修改原始数据）
        df = data[required_cols].copy()
        df = df[df[y_label]>=0].reset_index(drop=True)

        # 3. 阶段1：在 dropna 前统计原始样本量、坏样本量、坏样本占比（按分组）
        if not group_cols:
            df['_group_'] = 'overall'
            group_cols_actual = ['_group_']
        else:
            group_cols_actual = group_cols[:]

        # 统计原始样本量（含缺失值）
        sample_size_df = df.groupby(group_cols_actual).size().reset_index(name='样本量')

        # 统计原始坏样本量（y_label=1 的样本数）
        df['is_bad'] = (df[y_label] > 0).astype(int)  # 规范化 Y 标签
        bad_sample_df = df.groupby(group_cols_actual)['is_bad'].sum().reset_index(name='坏样本量')

        # 合并样本量和坏样本量统计
        stage1_df = pd.merge(
            sample_size_df,
            bad_sample_df,
            on=group_cols_actual,
            how='left'
        )
        stage1_df['坏样本占比'] = round(stage1_df['坏样本量'] / stage1_df['样本量'],4)
        stage1_df.insert(loc=0, column='model_id', value=model_id, allow_duplicates=False)

        # 4. 阶段2：dropna 后仅计算 KS
        df_clean = df.dropna(subset=[model_id, y_label] + (group_cols if group_cols else []))
        
        # 初始化 KS 结果存储
        ks_results = []

        # 按分组计算 KS
        for group_key, group in df_clean.groupby(group_cols_actual):
            score = group[model_id]
            y = (group[y_label] > 0).astype(int)  # 规范化 Y 标签

            # 计算 KS
            if len(y) < 2 or y.nunique() < 2:
                ks = np.nan
                logger.warning(f"数据缺失无法计算KS{len(score)}，{len(y)}")
            else:
                try:
                    ks = toad.metrics.KS(score, y)
                    ks = round(ks, 4)
                except Exception as e:
                    logger.warning(f"KS 计算失败（模型={model_id}, 分组={group_key}）: {e}")
                    ks = np.nan

            # 构造结果行
            result_row = {
                'model_id': model_id,
                'KS': ks
            }

            # 处理 group_key（可能是单个值或 tuple）
            if isinstance(group_key, tuple):
                for col, val in zip(group_cols_actual, group_key):
                    result_row[col] = val
            else:
                result_row[group_cols_actual[0]] = group_key

            ks_results.append(result_row)

        # 合并两部分结果
        ks_result_df = pd.DataFrame(ks_results)
        final_df = pd.merge(
            stage1_df,
            ks_result_df,
            on=['model_id'] + group_cols_actual,
            how='left'
        )

        return final_df

    def _process_ks_results(self, df, model_id,y_label,groupby_fields_list):
        # 定义所有可能的维度及其重命名映射
        ALL_DIMENSIONS = ['channel', 'customer_tags', 'month']
        RENAME_MAPPING = {
            'channel_types': 'channel',
            'channel_rates': 'channel'
        }
        data = df.copy()
        ks_results = []
         
        for fields in groupby_fields_list:
            # 按当前分组字段分组，并计算 KS
            result_df = self._calculate_ks_by_group(data=data,model_id=model_id,y_label=y_label,group_cols=fields)
            
            # 重命名列（将channel_types和channel_rates统一为channel）
            renamed_fields = [RENAME_MAPPING.get(field, field) for field in fields]
            rename_dict = dict(zip(fields, renamed_fields))
            result_df = result_df.rename(columns=rename_dict)
            
            # 找出缺失的维度并添加，填充"全部"
            missing_dimensions = [dim for dim in ALL_DIMENSIONS if dim not in renamed_fields]
            for dim in missing_dimensions:
                result_df[dim] = "0_全部"
            
            # 按指定顺序排列列
            result_df = result_df[ALL_DIMENSIONS + [col for col in result_df.columns if col not in ALL_DIMENSIONS]]
            
            ks_results.append(result_df)
         
        # 合并所有分组结果
        final_ks = pd.concat(ks_results, axis=0).reset_index(drop=True)
        final_ks = final_ks.sort_values(by=ALL_DIMENSIONS)
        return final_ks
            
    def _get_model_thresholds(self, model_id: str) -> Tuple[float, float]:
        if model_id in self.custom_thresholds:
            thresholds = self.custom_thresholds[model_id]
            return (
                thresholds.get('psi', self.psi_default_threshold),
                thresholds.get('mean_diff', self.mean_diff_default_threshold)
            )
        return self.psi_default_threshold, self.mean_diff_default_threshold

    def _calculate_monthly_stability(self, baseline_data: pd.DataFrame, monthly_data: Dict[str, pd.DataFrame],
                                    model_id: str, stat_date: str) -> pd.DataFrame:
        baseline_scores = baseline_data[model_id]
        baseline_mean = round(baseline_scores.mean(), 4)
        psi_thresh, mean_thresh = self._get_model_thresholds(model_id)
        stability_records = []

        sorted_months = sorted(monthly_data.keys())
        month_mean_map = {
            month: round(data[model_id].mean(), 4) 
            for month, data in monthly_data.items()
        }
                
        for i, current_month in enumerate(sorted_months):
            current_data = monthly_data[current_month]
            current_scores = current_data[[model_id, 'target_ar']]
            current_mean = month_mean_map[current_month]

            if i == 0:
                psi_value = None
                psi_status = "无环比数据"
                psi_compare_month = "无"
            else:
                last_month = sorted_months[i-1]
                last_month_data = monthly_data[last_month]
                last_month_scores = last_month_data[[model_id, 'target_ar']]
                psi_value = self._calculate_psi(last_month_scores, current_scores, model_id)
                if psi_value < 0:
                    psi_status = "数据异常"
                else:
                    psi_status = "预警" if psi_value > psi_thresh else "正常"
                psi_compare_month = last_month

            if i == 0:
                mean_diff_abs = None
                mean_compare_month = "无"
                mean_status = "无环比数据"
            else:
                last_month = sorted_months[i-1]
                last_month_mean = month_mean_map[last_month]
                mean_diff_abs = round(abs(current_mean - last_month_mean), 4)
                mean_compare_month = last_month
                mean_status = "预警" if mean_diff_abs > mean_thresh else "正常"

            stability_records.append({
                'model_id': model_id,
                'stat_date': stat_date,
                '月份': current_month,
                '固定基期月份': self.current_baseline_month,
                'PSI对比月份': psi_compare_month,
                '均值对比月份': mean_compare_month,
                '固定基期样本量': len(baseline_scores),
                '当前月份样本量': len(current_data),
                '固定基期均值': baseline_mean,
                '当前月份均值': current_mean,
                '上月均值': month_mean_map.get(sorted_months[i-1]) if i > 0 else None,
                '均值差异绝对值': mean_diff_abs,
                'PSI值': psi_value,
                'PSI阈值': psi_thresh,
                '均值阈值': mean_thresh,
                'PSI状态': psi_status,
                '均值状态': mean_status
            })

        return pd.DataFrame(stability_records)

    def _generate_excel_report(self, model_id: str, stat_date: str, ks_df: pd.DataFrame,
                              stability_df: pd.DataFrame, overall_dist: pd.DataFrame,
                              derived_tables: Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]
                              ) -> str:
        report_filename = f"{model_id}_监控报告_{stat_date}_固定基期{self.current_baseline_month}.xlsx"
        report_path = os.path.join(self.report_dir, report_filename)

        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            ks_df.to_excel(writer, sheet_name='KS', index=False)
            stability_df.to_excel(writer, sheet_name='PSI', index=False)
            derived_tables[1].to_excel(writer, sheet_name='各月人数占比', index=False)
            derived_tables[2].to_excel(writer, sheet_name='各月人数', index=False)

        wb = load_workbook(report_path)
        styles = self._get_excel_styles()

        ws_stability = wb['PSI']
        for row in ws_stability.iter_rows(min_row=1, max_row=ws_stability.max_row):
            for cell in row:
                cell.border = styles['thin_border']
                cell.font = styles['header_font'] if cell.row == 1 else styles['normal_font']
                if cell.row == 1:
                    cell.fill = styles['header_fill']
        self._adjust_column_width(ws_stability)

        for sheet_name in ['KS', '各月人数占比', '各月人数']:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.border = styles['thin_border']
                    cell.font = styles['header_font'] if cell.row == 1 else styles['normal_font']
                    if cell.row == 1:
                        cell.fill = styles['header_fill']
            self._adjust_column_width(ws)

        ws_catalog = wb.create_sheet('目录', 0)
        catalog_data = [
            ['模型监控报告'],
            [''],
            ['Sheet名称', '说明'],
            ['KS', 'KS分组'],
            ['PSI', 'PSI环比'],
            ['各月人数占比', '分箱-月份人数占比对比'],
            ['各月人数', '分箱-月份人数对比']
        ]
        for row_idx, row_data in enumerate(catalog_data, 1):
            for col_idx, val in enumerate(row_data, 1):
                ws_catalog.cell(row=row_idx, column=col_idx, value=val)
        ws_catalog['A1'].font = Font(bold=True, size=14)
        for row in ws_catalog.iter_rows(min_row=3, max_row=ws_catalog.max_row):
            row[0].font = Font(bold=True)

        self._adjust_column_width(ws_catalog)

        wb.save(report_path)
        logger.info(f"报告生成完成 | 报告路径：{report_path}")
        return report_path

    def _get_excel_styles(self) -> Dict:
        return {
            'thin_border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
            'header_font': Font(bold=True, size=10, color='FFFFFF'),
            'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'warning_font': Font(color='FF0000', bold=True),
            'baseline_font': Font(color='FF6B00', bold=True),
            'normal_font': Font(size=9)
        }

    def _adjust_column_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


#==============================================================================
# File: notification.py
#==============================================================================

import pandas as pd
import smtplib
import os
import tempfile
from typing import Dict, List, Optional
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.header import Header
from email.utils import formatdate
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging

# 配置日志
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

class EnhancedEmailSender:
    """增强版邮件发送类，支持通用邮件发送和Excel附件美化"""
    
    def __init__(self, email_config: Dict[str, str]):
        """
        初始化邮件发送器
        
        参数:
            email_config: 邮件配置字典，包含smtp_host, smtp_port, user, pass, receivers等
        """
        self.email_config = email_config
        # 解析配置
        self.smtp_host = email_config.get("smtp_host")
        self.smtp_port = int(email_config.get("smtp_port", 465))
        self.sender = email_config.get("user")
        self.password = email_config.get("password")
        self.receivers = email_config.get("receivers","")
        
        # 校验基本配置
        self._validate_config()
    
    def _validate_config(self) -> bool:
        """验证邮件配置是否完整"""
        if not all([self.smtp_host, self.sender, self.password, self.receivers]):
            logger.warning("⚠️ 邮件配置不完整，缺少必要参数")
            return False
        return True
    
    def _get_excel_styles(self) -> Dict:
        """获取Excel美化所需的样式定义"""
        return {
            'thin_border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
            'header_font': Font(bold=True, size=10, color='FFFFFF'),
            'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'warning_font': Font(color='FF0000', bold=True),
            'baseline_font': Font(color='FF6B00', bold=True),
            'normal_font': Font(size=9)
        }
    
    def _adjust_column_width(self, ws):
        """自动调整Excel列宽"""
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            # 限制最大宽度，避免过宽
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    
    def beautify_excel(self, file_path: str) -> None:
        """
        美化Excel文件格式
        
        参数:
            file_path: Excel文件路径
        """
        if not os.path.exists(file_path):
            logger.warning(f"⚠️ 要美化的Excel文件不存在: {file_path}")
            return
            
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb.active
            styles = self._get_excel_styles()
            
            # 美化表头
            for cell in ws[1]:
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = styles['thin_border']
            
            # 美化内容单元格
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = styles['normal_font']
                    cell.border = styles['thin_border']
                    # 对数字单元格进行右对齐
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
            
            # 自动调整列宽
            self._adjust_column_width(ws)
            
            # 保存修改
            wb.save(file_path)
            logger.info(f"✨ Excel文件已美化: {file_path}")
            
        except Exception as e:
            logger.error(f"❌ 美化Excel文件失败: {str(e)}", exc_info=True)
    
    def df_to_beautified_excel(self, df: pd.DataFrame, file_name: str = "data.xlsx") -> str:
        """
        将DataFrame转换为已美化的Excel文件
        
        参数:
            df: 要转换的DataFrame
            file_name: 输出文件名
            
        返回:
            生成的Excel文件路径
        """
        try:
            # 创建临时文件或指定路径
            if os.path.dirname(file_name):
                # 确保目录存在
                os.makedirs(os.path.dirname(file_name), exist_ok=True)
                file_path = file_name
            else:
                # 使用临时文件
                temp_dir = tempfile.gettempdir()
                file_path = os.path.join(temp_dir, file_name)
            
            # 保存DataFrame到Excel
            df.to_excel(file_path, index=False, engine='openpyxl')
            
            # 美化Excel
            self.beautify_excel(file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ 将DataFrame转换为Excel失败: {str(e)}", exc_info=True)
            raise
    
    def send_generic_email(
        self,
        subject_template: str,
        body_template: str,
        body_vars: Dict[str, str],
        attachment_paths: Optional[List[str]] = None,
        df_attachments: Optional[List[Dict]] = None
    ) -> None:
        """
        通用邮件发送函数（支持动态主题/正文/附件，包括DataFrame转换的Excel）
        
        参数:
            subject_template: 邮件主题模板（含占位符）
            body_template: HTML正文模板（含占位符）
            body_vars: 正文变量字典（填充模板占位符）
            attachment_paths: 现有附件路径列表
            df_attachments: DataFrame附件列表，每个元素为字典，包含'df'和'file_name'键
        """
        # 校验配置
        if not self._validate_config():
            return
            
        # 处理DataFrame附件，转换为美化后的Excel
        temp_files = []
        try:
            # 确保附件列表存在
            all_attachments = attachment_paths.copy() if attachment_paths else []
            
            # 处理DataFrame附件
            if df_attachments:
                for df_info in df_attachments:
                    df = df_info.get('df')
                    file_name = df_info.get('file_name', 'data.xlsx')
                    
                    if isinstance(df, pd.DataFrame) and df.shape[0] > 0:
                        # 转换并美化DataFrame为Excel
                        excel_path = self.df_to_beautified_excel(df, file_name)
                        all_attachments.append(excel_path)
                        temp_files.append(excel_path)  # 记录临时文件以便后续清理
                    else:
                        logger.warning(f"⚠️ 无效的DataFrame附件，跳过: {file_name}")
            
            # 构建邮件
            msg = MIMEMultipart()
            msg["From"] = f"<{self.sender}>"
            msg["To"] = ", ".join(self.receivers)
            msg["Subject"] = Header(subject_template.format(** body_vars), "utf-8")
            msg["Date"] = formatdate(localtime=True)
            
            # 渲染HTML正文
            body = body_template.format(**body_vars)
            msg.attach(MIMEText(body, "html", "utf-8"))
            
            # 添加附件
            for path in all_attachments:
                if not os.path.exists(path):
                    logger.warning(f"⚠️ 附件不存在: {path}，跳过")
                    continue
                try:
                    with open(path, "rb") as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(path))
                    part["Content-Disposition"] = f'attachment; filename="{os.path.basename(path)}"'
                    msg.attach(part)
                    logger.info(f"📎 成功添加附件: {os.path.basename(path)}")
                except Exception as e:
                    logger.error(f"❌ 附件添加失败: {path} - {str(e)}")
            
            # 发送邮件（SSL加密）
            with smtplib.SMTP_SSL(self.smtp_host, self.smtp_port) as server:
                server.login(self.sender, self.password)
                server.sendmail(self.sender, self.receivers, msg.as_string())
            logger.info(f"📧 邮件已发送至: {self.receivers}")
            
        except Exception as e:
            logger.error(f"❌ 邮件发送失败: {str(e)}", exc_info=True)
            raise
        finally:
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        logger.info(f"🗑️ 已清理临时文件: {temp_file}")
                except Exception as e:
                    logger.warning(f"⚠️ 清理临时文件失败: {temp_file} - {str(e)}")
  


#==============================================================================
# File: odps_utils.py
#==============================================================================

from odps import ODPS
import pandas as pd
import logging
import time
import os
from multiprocessing import cpu_count
from typing import Dict, List, Optional, Tuple, Union


logger = logging.getLogger(__name__)
_odps_client = None  # 单例客户端

def get_odps_client():
    """初始化并获取ODPS客户端（单例模式）"""
    global _odps_client
    if _odps_client is None:
        try:
            # 从环境变量读取配置
            USERNAME = 'liaoxilin'
            PASSWORD = 'j02vYCxx'

            if not all([USERNAME, PASSWORD]):
                raise ValueError("请设置USERNAME和PASSWORD环境变量")

            _odps_client = ODPS(username=USERNAME, password=PASSWORD)
            logger.info("✅ ODPS客户端初始化成功")
        except Exception as e:
            logger.error(f"❌ ODPS初始化失败: {str(e)}")
            raise
    return _odps_client

def execute_odps_sql(sql: str) -> pd.DataFrame:
    """执行ODPS SQL并返回DataFrame"""
    client = get_odps_client()
    logger.info(f"📌 执行SQL: {sql[:100]}...")  # 打印SQL前100字符
    start = time.time()

    try:
        instance = client.execute_sql(sql)
        while not instance.is_terminated():
            time.sleep(1)  # 等待SQL执行完成

        if instance.is_successful():
            with instance.open_reader() as reader:
                # 根据CPU核心数设置并行数
                n_process = min(8, max(1, cpu_count() - 1))
                df = reader.to_pandas(n_process=n_process)
                logger.info(f"📥 加载 {len(df)} 行数据，耗时 {time.time()-start:.2f}s")
                return df
        else:
            raise RuntimeError(f"SQL执行失败: {instance.get_logview_address()}")
    except Exception as e:
        logger.error(f"❌ SQL执行失败: {str(e)}")
        raise
        

def load_pivot_config(
    biz_type: str,
    current_dt: str = None,
    config_table: str = "znzz_fintech_ads.lxl_model_pivot_config"
) -> List[str]:
    """
    从配置表加载需要 pivot 的 variable_code 列表
    """
    if current_dt is None:
        current_dt = datetime.now().strftime("%Y-%m-%d")

    sql = f"""
    SELECT variable_code
    FROM {config_table}
    WHERE dt = '{current_dt}'
      AND biz_type = '{biz_type}'
      AND is_active = 1
    """

    logger.info(f"📌 正在加载配置: biz_type={biz_type}, date={current_dt}")
    df = execute_odps_sql(sql)
    
    if df.empty:
        logger.warning(f"⚠️ 未找到 biz_type='{biz_type}' 的有效配置")
        return []
    
    var_list = df['variable_code'].tolist()
    logger.info(f"✅ 加载到 {len(var_list)} 个 variable_code: {var_list[:5]}{'...' if len(var_list) > 5 else ''}")
    return var_list

def generate_and_execute_wide_table_query(
    biz_type: str,
    start_date: str,
    end_date: str,
    variable_codes: List[str] = None,
    dimension_cols: List[str] = None,
    main_table: str = "znzz_fintech_ads.apply_model01_scores_off",
    join_table: str = "znzz_fintech_ads.lxl_model_auth_tags",
    date_col_main: str = "apply_time",
    date_col_join: str = "dt",
    key_column: str = "order_no",
    metric_column: str = "good_score",
    config_current_dt: str = None
) -> pd.DataFrame:
    """
    主函数：先在 t2 中 pivot，再与 t3 关联（推荐方式）
    """
    if dimension_cols is None:
        dimension_cols = [
            "id_no_des", "mobile_des", "channel_id", "apply_date",
            "auth_status", "target_ar", "customer_tags", "umeng_flag"
        ]

    # Step 1: 获取需要 pivot 的 variable_code
    if not variable_codes:
        variable_codes = load_pivot_config(biz_type, current_dt=config_current_dt)
    if not variable_codes:
        raise ValueError(f"未加载到 {biz_type} 的 variable_code 配置")

    # Step 2: 构建 t2 的 pivot 字段
    pivot_clauses = []
    for var in variable_codes:
        safe_col = var.replace('-', '_').replace('.', '_').replace('%', '_pct_')
        pivot_clauses.append(
            f"MAX(CASE WHEN variable_code = '{var}' THEN {metric_column} END) AS `{safe_col}`"
        )
    pivot_select = ",\n        ".join(pivot_clauses)

    # Step 3: 构建优化后的 SQL —— 先 pivot 再 join
    sql = f"""
    SELECT
        t2_pivot.`{key_column}`,
        {', '.join(f't2_pivot.`{col}`' for col in variable_codes)},  -- pivot 后的 score 列
        {', '.join(f't3.{col}' for col in dimension_cols)}            -- t3 的维度列
    FROM (
        -- ✅ Step A: 在 t2 中完成横转列（按 order_no 聚合）
        SELECT
            `{key_column}`,
            {pivot_select}
        FROM {main_table}
        WHERE `{date_col_main}` >= '{start_date}'
          AND `{date_col_main}` <= '{end_date}'
          AND variable_code IN ({','.join(f"'{v}'" for v in variable_codes)})
        GROUP BY `{key_column}`
    ) t2_pivot
    RIGHT JOIN (
        -- Step B: 过滤维度表 t3
        SELECT 
            `{key_column}`,
            {', '.join(f'`{col}`' for col in dimension_cols)}
        FROM {join_table}
        WHERE `{date_col_join}` >= '{start_date}'
          AND `{date_col_join}` <= '{end_date}'
          AND channel_id != 1
    ) t3 ON t2_pivot.`{key_column}` = t3.`{key_column}`
    """

    # Step 4: 执行 SQL
    logger.info("🔄 开始执行优化版宽表查询（先 pivot 再 join）...")
    df_wide = execute_odps_sql(sql)
    return df_wide    


#==============================================================================
# File: over_sample_under_sample.py
#==============================================================================

from sklearn.linear_model import LogisticRegression  
lr_model = LogisticRegression(class_weight={0:4, 1:1})  
lr_model.fit(x,y) 

from sklearn.linear_model import LogisticRegression  
lr_model = LogisticRegression(class_weight="balanced")  
lr_model.fit(x,y)  

#生成数据集  
from sklearn.datasets import make_classification  
from collections import Counter  
X, y = make_classification(n_samples=5000, n_features=2,
                                n_informative=2, n_redundant=0, 
                                n_repeated=0, n_classes=3, 
                                n_clusters_per_class=1, 
                                weights=[0.01, 0.05, 0.94], 
                                class_sep=0.8, random_state=0)
#随机过采样  
from imblearn.over_sampling import RandomOverSampler  
X_resampled, y_resampled = RandomOverSampler().fit_sample(X, y)  
#SMOTE过采样及其变体  
from imblearn.over_sampling import SMOTE  
X_resampled_smote, y_resampled_smote = SMOTE().fit_sample(X, y)  
X_resampled, y_resampled = SMOTE(kind='borderline1').fit_sample(X, y)  
X_resampled, y_resampled = SMOTE(kind='borderline2').fit_sample(X, y)  
#ADASYN过采样  
from imblearn.over_sampling import ADASYN  
X_resampled_adasyn, y_resampled_adasyn = ADASYN().fit_sample(X, y)  
#随机欠采样  
from imblearn.under_sampling import RandomUnderSampler  
X_resampled, y_resampled = RandomUnderSampler().fit_sample(X, y)  
#基于k-means聚类的欠采样  
from imblearn.under_sampling import ClusterCentroids  
X_resampled, y_resampled = ClusterCentroids().fit_sample(X, y)  
#基于最近邻算法的欠采样  
from imblearn.under_sampling import RepeatedEditedNearestNeighbours  
X_resampled, y_resampled = 
     RepeatedEditedNearestNeighbours().fit_sample(X, y)  
#在数据上运用一种分类器
#然后将概率低于阈值的样本剔除掉
#从而实现欠采样  
from sklearn.linear_model import LogisticRegression  
from imblearn.under_sampling import InstanceHardnessThreshold  
lr_underS = InstanceHardnessThreshold(
               random_state=0, estimator=LogisticRegression())
X_resampled, y_resampled = lr_underS.fit_sample(X, y) 


#==============================================================================
# File: plot.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: plot.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-10-18
'''
import os
import re

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns
import six
from matplotlib import gridspec
from openpyxl.drawing.image import Image
from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_curve, roc_auc_score, precision_recall_curve

from .statistics import calc_var_summary
from .utils import save_json


def ellipsis_fun(x, decimal=2, ellipsis=16):
    if re.search(r'\[(.*?) ~ (.*?)\)', str(x)):
        le = x.split('~')[0]
        ri = x.split('~')[1]
        left = re.match(r'\d+.\[(.*?) ', le)
        right = re.match(r' (.*?)\)', ri)
        if left:
            tmp = round(float(left.groups()[0]), decimal)
            l = f"{le.split('[')[0]}[{tmp} "
            le = l
        if right:
            tmp = round(float(right.groups()[0]), decimal)
            r = f" {tmp})"
            ri = r

        return f"{le} ~ {ri}"
    else:
        return str(x)[:ellipsis] + '..'


def plot_var_bin_summary(frame, cols, target='target', by='type', file_path=None, sheet_name='plot_var_bin_summary',
                         need_bin=False, decimal=2, ellipsis=16, **kwargs):
    """
    画变量分箱图
    Args:
        frame (DataFrame):
        cols (str|list): 需要画图的变量
        target (str): 目标变量
        by (str): 分开统计的列名
        file_path (str): 保存路径
        sheet_name (str):保存excel的sheet名称
        need_bin (bool):是否要进行分箱

    Returns:

    """

    if isinstance(cols, str):
        cols = [cols]

    if isinstance(frame, pd.DataFrame):
        summary_df_dict = {}
        if by in frame:
            for name, df in frame.groupby(by):
                summary_df_dict[name] = calc_var_summary(df, include_cols=cols, target=target, need_bin=need_bin,
                                                         **kwargs)
        else:
            summary_df_dict['all'] = calc_var_summary(frame, include_cols=cols, target=target, need_bin=need_bin,
                                                      **kwargs)
    else:
        summary_df_dict = frame
    summary_num = len(summary_df_dict)

    save_jpg_path = None
    if file_path is not None:
        if '.xlsx' in file_path:
            fp_sp = file_path.split('.xlsx')
            if '' == fp_sp[1]:
                save_jpg_path = fp_sp[0] + '_var_jpg'
                os.makedirs(save_jpg_path, exist_ok=True)

            else:
                save_jpg_path = os.path.join(file_path, 'var_jpg')
                os.makedirs(save_jpg_path, exist_ok=True)
                file_path = os.path.join(file_path, 'plot_var_bin_summary.xlsx')
        else:
            save_jpg_path = os.path.join(file_path, 'var_jpg')
            os.makedirs(save_jpg_path, exist_ok=True)
            file_path = os.path.join(file_path, 'plot_var_bin_summary.xlsx')

    for col in cols:
        # 做图
        # plt.figure(figsize=(15, 4))
        # gs = gridspec.GridSpec(1, 2)
        # gs = gridspec.GridSpec(1, 3)

        gs_num = len(summary_df_dict)

        if gs_num == 1:
            plt.figure(figsize=(gs_num * 5, 4), dpi=400)
        else:
            plt.figure(figsize=(gs_num * 5, 4))

        gs = gridspec.GridSpec(1, gs_num)

        for i, k in enumerate(summary_df_dict):

            df = summary_df_dict[k]

            if i == 0:
                ax1 = plt.subplot(gs[0, 0])
            else:
                ax1 = plt.subplot(gs[0, i], sharey=ax1)

            df = df[df['var_name'] == col]
            df['range'] = df['range'].map(lambda x: ellipsis_fun(x, decimal, ellipsis))

            x_point = np.arange(len(df))
            y_point = df['total_pct']

            ax1.bar(x_point, y_point, color='Orange', alpha=0.4, width=0.5, label='PctTotal')
            x_labels = list(df['range'])
            plt.xticks(np.arange(len(df)), x_labels, fontsize=10, rotation=45)

            for x, y in zip(x_point, y_point):
                ax1.text(x + 0.05, y + 0.01, str(round(y * 100, 2)) + '%', ha='center', va='bottom', fontsize=12)
            ax1.set_ylabel('total_pct', fontsize=12)

            ax1.set_ylim([0, max(y_point) + ((max(y_point) - min(y_point)) / len(y_point))])
            bottom, top = ax1.get_ylim()
            ax2 = ax1.twinx()
            ax2.plot(x_point, df['positive_rate'], '-ro', color='red')

            for x, y in zip(x_point, df['positive_rate']):
                ax2.text(x + 0.05, y, str(round(y * 100, 2)) + '%', ha='center', va='bottom', fontsize=12, color='r')
            ax2.set_ylabel('positive_rate', fontsize=12)
            # ax2.set_ylim([0, max(df['positive_rate']) + 0.01])
            ax2.set_ylim(
                [0, max(df['positive_rate']) + ((max(df['positive_rate']) - min(df['positive_rate'])) / len(df))])

            plt.title('{}:{}\nIV: {:.5f}'.format(k, col, df['IV'].iloc[0]), loc='right', fontsize='small')
            # 将数据详情表添加
            tmp_df = df[['range', 'woe', 'iv', 'total']]
            round_cols = ['woe', 'iv']
            tmp_df[round_cols] = tmp_df[round_cols].applymap(lambda v: round(v, 4) if pd.notnull(v) else '')
            mpl_table = plt.table(cellText=tmp_df.values, colLabels=tmp_df.columns,
                                  colWidths=[0.2, 0.1, 0.1, 0.1],
                                  loc='top')  # loc='top'将详情放到顶部

            mpl_table.auto_set_font_size(False)
            mpl_table.set_fontsize(5)

            header_color = 'darkorange'
            row_colors = ['bisque', 'w']
            header_columns = 0

            for k, cell in six.iteritems(mpl_table._cells):
                cell.set_edgecolor('w')
                if k[0] == 0 or k[1] < header_columns:
                    cell.set_text_props(weight='bold', color='w')
                    cell.set_facecolor(header_color)
                else:
                    cell.set_facecolor(row_colors[k[0] % len(row_colors)])

        plt.tight_layout()

        if save_jpg_path is not None:  # 判断是否需要保存
            plt.savefig(os.path.join(save_jpg_path, '{}.png'.format(col)), dpi=300, bbox_inches='tight')  # dpi控制清晰度
        # plt.show()

    # if save_jpg_path is not None:
    #     workbook = xlsxwriter.Workbook(file_path)
    #     worksheet = workbook.add_worksheet(sheet_name)
    #
    #     for i, jpg_name in enumerate(cols):
    #         worksheet.insert_image('A{}'.format(i * 29 + 3), os.path.join(save_jpg_path, '{}.jpg'.format(jpg_name),
    #                                {'x_scale': 1.5, 'y_scale': 1.5})
    #
    #     workbook.close()
    if save_jpg_path is not None:
        try:
            wb = openpyxl.load_workbook(file_path)
        except:
            wb = openpyxl.Workbook()
        sh = wb.create_sheet(sheet_name)
        for i, jpg_name in enumerate(cols):
            img = Image(os.path.join(save_jpg_path, '{}.png'.format(jpg_name)))
            newsize = (summary_num * 700, 600)
            img.width, img.height = newsize  # 设置图片的宽和高
            sh.add_image(img, 'A{}'.format(i * 35 + 3))
        wb.save(file_path)


def plot_describe(df, by_col='p', title='model distribution'):
    try:
        sns.displot(df[by_col], color='#ff8080', bins=20,
                    kde_kws={"lw": 2.5, 'linestyle': '--'})
    except RuntimeError as rte:
        if str(rte).startswith("Selected KDE bandwidth is 0. Cannot estiamte density."):
            sns.displot(df[by_col], color='#ff8080',
                        bins=20, kde_kws={"lw": 2.5, 'linestyle': '--', 'bw': 1})
        else:
            raise rte
    tem = df[by_col].describe().reset_index()
    table_ = plt.table(cellText=[[round(x, 4)] for x in tem[by_col].tolist()],
                       colWidths=[0.1] * 1, rowLabels=tem['index'].tolist(), loc='right')
    table_.set_fontsize(15)
    table_.scale(1.9, 2.265)
    plt.title(f'{title} by {by_col}')


def get_optimal_cutoff(fpr_recall, tpr_precision, threshold, is_f1=False):
    if is_f1:
        youdenJ_f1score = (2 * tpr_precision * fpr_recall) / (tpr_precision + fpr_recall)
    else:
        youdenJ_f1score = tpr_precision - fpr_recall
    point_index = np.argmax(youdenJ_f1score)
    optimal_threshold = threshold[point_index]
    point = [fpr_recall[point_index], tpr_precision[point_index]]
    return optimal_threshold, point


def plot_ks(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    ks = max(tpr - fpr)  ###计算ks的值

    plt.figure(figsize=(6, 6))
    x = np.arange(len(thresholds)) / len(thresholds)
    plt.plot(x, tpr, lw=1)
    plt.plot(x, fpr, lw=1)
    plt.plot(x, tpr - fpr, lw=1, linestyle='--', label='KS curve (KS = %0.4f)' % ks)

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    optimal_th_index = np.where(thresholds == optimal_th)
    plt.plot(optimal_th_index[0][0] / len(thresholds), ks, marker='o', color='r')
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, (float('%.4f' % optimal_point[0]),
                                                            float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Thresholds Index')
    plt.ylabel('TPR FPR KS')
    name = '{} KS Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'KS': ks, 'KS最大值-threshold': optimal_th}


def plot_roc(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    auc_value = roc_auc_score(y_true, y_pred)  ###计算auc的值

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(fpr, tpr, color='darkorange',
             lw=lw, label='ROC curve (AUC = %0.4f)' % auc_value)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('False Positive Rate(FPR)')
    plt.ylabel('True Positive Rate(TPR)')
    name = '{} ROC Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'AUC': auc_value}


def plot_pr(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score
    max_f1score = max(f1score)

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(recall, precision, color='darkorange',
             lw=lw, label='PR curve (F1score = %0.4f)' % max_f1score)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=recall, tpr_precision=precision, threshold=thresholds,
                                                   is_f1=True)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    name = '{} PR Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'F1_Score最大值': max_f1score, 'F1_Score最大值-threshold': optimal_th, '模型拐点': optimal_th, '阀值': optimal_th,
            'Precision': optimal_point[1], 'Recall': optimal_point[0], 'F1_Score': max_f1score}


def plot_pr_f1(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    thresholds = np.insert(thresholds, 0, 0, axis=None)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score

    x = np.arange(len(thresholds)) / len(thresholds)

    pr_f1_dict = {'Precision': precision, 'Recall': recall, 'F1_score': f1score}

    for i in pr_f1_dict:
        plt.figure(figsize=(6, 6))

        plt.plot(x, pr_f1_dict[i], lw=1)

        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.0])
        plt.xlabel('Thresholds Index')
        plt.ylabel('{}'.format(i))
        name = '{} {} Curve'.format(title, i)
        plt.title(name)
        plt.savefig(output_path + name, bbox_inches='tight')
        plt.show()

    return {'Thresholds': list(thresholds), '模型召回率': list(recall), '模型精准率': list(precision),
            '模型F1-score': list(f1score)}


def calc_celue_cm(df: pd.DataFrame, target='target', to_bin_col='p'):
    q_cut_list = np.arange(0, 1, 1 / 10) + 0.1
    confusion_matrix_df = pd.DataFrame()
    for i in q_cut_list:
        df['pred_label'] = np.where(df[to_bin_col] >= i, 1, 0)
        tmp_list = []
        tmp_list.append(i)

        tn, fp, fn, tp = confusion_matrix(np.array(df[target]), np.array(df['pred_label'])).ravel()

        tmp_list.extend([tp, fp, tn, fn])

        confusion_matrix_df = confusion_matrix_df.append(pd.DataFrame(tmp_list).T)

    # confusion_matrix_df.columns = ['阈值', 'TP', 'FP', 'TN', 'FN']
    confusion_matrix_df.columns = ['阈值', '实际正样本-预测为正样本', '实际负样本-预测为正样本', '实际负样本-预测为负样本', '实际正样本-预测为负样本']
    confusion_matrix_df.set_index('阈值', inplace=True)
    confusion_matrix_df['sum'] = confusion_matrix_df.apply(lambda x: x.sum(), axis=1)

    # return confusion_matrix_df
    return confusion_matrix_df.to_dict()


def calc_plot_metrics(df: pd.DataFrame, to_bin_col='p', target='target', curve_save_path=''):
    data = {k: v for k, v in df.groupby('type')}
    data.update({'all': df})

    for data_type, type_df in data.items():
        res_save_path = os.path.join(curve_save_path, data_type)
        os.makedirs(res_save_path, exist_ok=True)

        res_dict = {}
        res_dict.update(plot_roc(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_ks(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_pr(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(calc_celue_cm(type_df, target, to_bin_col))
        res_dict.update(plot_pr_f1(type_df[target], type_df[to_bin_col], res_save_path, data_type))

        ###相关指标保存json格式
        save_json(res_dict, os.path.join(res_save_path, '{}_res_json.json'.format(data_type)))


if __name__ == "__main__":
    ######读取数据
    data_path = 'TD47p25combine_td_to_report_data.csv'
    df = pd.read_csv(data_path)
    df = df[df['label'].notnull()]

    print(len(df))

    ######结果保存路径
    curve_save_path = '../examples/curve_result'
    calc_plot_metrics(df=df, to_bin_col='td', target='label', curve_save_path=curve_save_path)



#==============================================================================
# File: plot_metrics.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: plot_metrics.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2022-08-26
'''

import json
import os

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_curve, roc_auc_score, precision_recall_curve


def get_optimal_cutoff(fpr_recall, tpr_precision, threshold, is_f1=False):
    if is_f1:
        youdenJ_f1score = (2 * tpr_precision * fpr_recall) / (tpr_precision + fpr_recall)
    else:
        youdenJ_f1score = tpr_precision - fpr_recall
    point_index = np.argmax(youdenJ_f1score)
    optimal_threshold = threshold[point_index]
    point = [fpr_recall[point_index], tpr_precision[point_index]]
    return optimal_threshold, point


def plot_ks(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    ks = max(tpr - fpr)  ###计算ks的值

    plt.figure(figsize=(6, 6))
    x = np.arange(len(thresholds)) / len(thresholds)
    plt.plot(x, tpr, lw=1)
    plt.plot(x, fpr, lw=1)
    plt.plot(x, tpr - fpr, lw=1, linestyle='--', label='KS curve (KS = %0.4f)' % ks)

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    optimal_th_index = np.where(thresholds == optimal_th)
    plt.plot(optimal_th_index[0][0] / len(thresholds), ks, marker='o', color='r')
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, (float('%.4f' % optimal_point[0]),
                                                            float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Thresholds Index')
    plt.ylabel('TPR FPR KS')
    name = '{} KS Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'KS': ks, 'KS最大值-threshold': optimal_th}


def plot_roc(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    auc_value = roc_auc_score(y_true, y_pred)  ###计算auc的值

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(fpr, tpr, color='darkorange',
             lw=lw, label='ROC curve (AUC = %0.4f)' % auc_value)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('False Positive Rate(FPR)')
    plt.ylabel('True Positive Rate(TPR)')
    name = '{} ROC Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'AUC': auc_value}


def plot_pr(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score
    max_f1score = max(f1score)

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(recall, precision, color='darkorange',
             lw=lw, label='PR curve (F1score = %0.4f)' % max_f1score)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=recall, tpr_precision=precision, threshold=thresholds,
                                                   is_f1=True)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    name = '{} PR Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'F1_Score最大值': max_f1score, 'F1_Score最大值-threshold': optimal_th, '模型拐点': optimal_th, '阀值': optimal_th,
            'Precision': optimal_point[1], 'Recall': optimal_point[0], 'F1_Score': max_f1score}


def plot_pr_f1(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    thresholds = np.insert(thresholds, 0, 0, axis=None)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score

    x = np.arange(len(thresholds)) / len(thresholds)

    pr_f1_dict = {'Precision': precision, 'Recall': recall, 'F1_score': f1score}

    for i in pr_f1_dict:
        plt.figure(figsize=(6, 6))

        plt.plot(x, pr_f1_dict[i], lw=1)

        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.0])
        plt.xlabel('Thresholds Index')
        plt.ylabel('{}'.format(i))
        name = '{} {} Curve'.format(title, i)
        plt.title(name)
        plt.savefig(output_path + name, bbox_inches='tight')
        plt.show()

    return {'Thresholds': list(thresholds), '模型召回率': list(recall), '模型精准率': list(precision), '模型F1-score': list(f1score)}


def calc_celue_cm(df: pd.DataFrame, target='target', to_bin_col='p'):
    q_cut_list = np.arange(0, 1, 1 / 10) + 0.1
    confusion_matrix_df = pd.DataFrame()
    for i in q_cut_list:
        df['pred_label'] = np.where(df[to_bin_col] >= i, 1, 0)
        tmp_list = []
        tmp_list.append(i)

        tn, fp, fn, tp = confusion_matrix(np.array(df[target]), np.array(df['pred_label'])).ravel()

        tmp_list.extend([tp, fp, tn, fn])

        confusion_matrix_df = confusion_matrix_df.append(pd.DataFrame(tmp_list).T)

    # confusion_matrix_df.columns = ['阈值', 'TP', 'FP', 'TN', 'FN']
    confusion_matrix_df.columns = ['阈值', '实际正样本-预测为正样本', '实际负样本-预测为正样本', '实际负样本-预测为负样本', '实际正样本-预测为负样本']
    confusion_matrix_df.set_index('阈值', inplace=True)
    confusion_matrix_df['sum'] = confusion_matrix_df.apply(lambda x: x.sum(), axis=1)

    # return confusion_matrix_df
    return confusion_matrix_df.to_dict()


def save_json(res_dict, file, indent=4):
    if isinstance(file, str):
        of = open(file, 'w')

    with of as f:
        json.dump(res_dict, f, ensure_ascii=False, indent=indent)


def load_json(file):
    if isinstance(file, str):
        of = open(file, 'r')

    with of as f:
        res_dict = json.load(f)

    return res_dict


def calc_plot_metrics(df: pd.DataFrame, to_bin_col='p', target='target', curve_save_path=''):
    data = {k: v for k, v in df.groupby('type')}
    data.update({'all': df})

    for data_type, type_df in data.items():
        res_save_path = os.path.join(curve_save_path, data_type)
        os.makedirs(res_save_path, exist_ok=True)

        res_dict = {}
        res_dict.update(plot_roc(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_ks(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_pr(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(calc_celue_cm(type_df, target, to_bin_col))
        res_dict.update(plot_pr_f1(type_df[target], type_df[to_bin_col], res_save_path, data_type))

        ###相关指标保存json格式
        save_json(res_dict, os.path.join(res_save_path, '{}_res_json.json'.format(data_type)))


if __name__ == "__main__":
    ######读取数据
    data_path = 'TD47p25combine_td_to_report_data.csv'
    df = pd.read_csv(data_path)
    df = df[df['label'].notnull()]

    print(len(df))

    ######结果保存路径
    curve_save_path = 'curve_result'
    calc_plot_metrics(df=df, to_bin_col='td', target='label', curve_save_path=curve_save_path)



#==============================================================================
# File: report2excel.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: report2excel.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-20
'''

import re
import string
import warnings

import xlsxwriter

from .logger_utils import Logger

warnings.filterwarnings('ignore')

log = Logger(level="info", name=__name__).logger


class Report2Excel:
    """
    统计信息写入到excel
    """

    def __init__(self, file_name, workbook=None, current_row=8, row_spaces=3, column_space=2):
        self.file_name = file_name
        self.workbook = self.create_workbook(workbook)
        self.define_global_format()
        self.row_spaces = row_spaces
        self.column_space = column_space

    def create_workbook(self, workbook):
        """类初始化的时候,必须创建或者返回一个workbook对象"""
        if isinstance(workbook, xlsxwriter.workbook.Workbook):
            return workbook
        else:
            workbook = xlsxwriter.Workbook(self.file_name, {'nan_inf_to_errors': True})
        return workbook

    def close_workbook(self):
        """关闭workbook对象,开始写入到excel文件中"""
        self.workbook.close()

    def df_to_excel(self, evaluate_df, sheet_name):
        # excel相关格式
        worksheet = self.workbook.add_worksheet(sheet_name)

        def write_df_to_excel_by_row(df, worksheet, start_row, start_column, str_formats):
            """按列写入df,因为要单独给每一列一种格式"""
            for col_num, column_name in enumerate(df.columns.values):
                try:
                    worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                           str_formats)
                except Exception as e:
                    t = list(map(str, df[column_name].values.tolist()))
                    worksheet.write_column(start_row, start_column + col_num, t,
                                           str_formats)

        # 写入标题
        for col_num, value in enumerate(evaluate_df.columns.values):
            worksheet.write(0, col_num, value, self.header_format)

        # 写入数据和插入图像
        start_row, start_column = 1, 0
        write_df_to_excel_by_row(evaluate_df, worksheet, start_row, start_column, self.str_formats)
        worksheet.set_column(0, 1, 18)
        # workbook.close()

    def define_global_format(self):
        """定义Excel全局格式"""
        self.merge_format = self.workbook.add_format({  # 大标题的格式
            'bold': True,
            'font_name': '微软雅黑',
            'font_size': 20,
            'border': False,  # 边框线
            'align': 'center',  # 水平居中
            'valign': 'vcenter',  # 垂直居中
            'fg_color': '#ffffff',  # 颜色填充
        })
        self.title_format = self.workbook.add_format({  # 标题的格式
            'bold': True,
            'font_name': '微软雅黑',
            'font_size': 15,
            'border': False,  # 边框线
            'valign': 'top',
            # 'fg_color': '#ddebf7'
        })
        self.sub_title_format = self.workbook.add_format({  # 子标题的格式
            'bold': True,
            'font_name': '微软雅黑',
            'font_size': 12,
            'border': False,  # 边框线
            'valign': 'top',
            # 'fg_color': '#ddebf7'
        })
        self.content_format = self.workbook.add_format({  # 文字的格式
            'bold': False,
            'font_name': '微软雅黑',
            'font_size': 10,
            'border': False,  # 边框线
            'valign': 'top',
        })
        self.content_dict_format = self.workbook.add_format({  # 文字的格式
            'bold': True,
            'font_name': '微软雅黑',
            'font_size': 10,
            'border': False,  # 边框线
            'valign': 'top',
            'font_color': '#275b8e',
        })
        self.ps_format = self.workbook.add_format({  # 注释的格式
            'bold': False,
            'font_name': '微软雅黑',
            'font_size': 10,
            'font_color': 'red',
            'border': False,  # 边框线
            'valign': 'top'
        })
        self.header_format = self.workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'center',
            'font_name': 'Consolas',
            'font_size': 11,
            'border': 1,
            'bottom': 2,
        })
        ### 数字格式
        self.left_formats = self.workbook.add_format({'font_name': 'Consolas', 'align': 'left'})
        self.integer_formats = self.workbook.add_format(
            {'font_name': 'Consolas', 'num_format': '#,##0', 'align': 'left'})  # 整数型
        self.float_formats = self.workbook.add_format(
            {'font_name': 'Consolas', 'num_format': '#,##0.00', 'align': 'left'})  # 浮点型保留2位
        self.percentage_formats = self.workbook.add_format(
            {'font_name': 'Consolas', 'num_format': '0.00%', 'align': 'left'})  # 百分数保留2位
        self.varname_formats = self.workbook.add_format(
            {'color': 'blue', 'underline': True, 'font_name': 'Consolas', 'font_size': 11, 'align': 'left'})
        self.str_formats = self.workbook.add_format(
            {'font_name': 'Consolas', 'font_size': 11, 'align': 'left'})
        self.float_formats4 = self.workbook.add_format(
            {'num_format': '#,##0.0000', 'font_name': 'Consolas', 'font_size': 11, 'align': 'left'})  # 浮点型保留4位
        self.font_formats = self.workbook.add_format({'font_name': 'Consolas', 'font_size': 11, 'align': 'left'})  # 其他
        self.float_formats_new = self.workbook.add_format(
            {'font_name': 'Consolas', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)',
             'align': 'left'})  # 百分数保留2位


# Standard Library


# Third Party Stuff


EXCEL_RANGE_PATTERN = re.compile(r'([a-zA-Z]+)([\d]+):([a-zA-Z]+)([\d]+)')

XLSXWRITER_FORMAT_PROPERTIES = (
    'font_name',
    'font_size',
    'font_color',
    'bold',
    'italic',
    'underline',
    'font_strikeout',
    'font_script',
    'num_format',
    'locked',
    'hidden',
    'text_h_align',
    'text_v_align',
    'rotation',
    'text_wrap',
    'text_justlast',
    # 'center_across',
    'indent',
    'shrink',
    'pattern',
    'bg_color',
    'fg_color',
    'bottom',
    'top',
    'left',
    'right',
    'bottom_color',
    'top_color',
    'left_color',
    'right_color',
)


def duplicate_xlsxwriter_format_object(workbook, old_format):
    properties = {}
    if old_format is not None:
        for property_name in XLSXWRITER_FORMAT_PROPERTIES:
            properties[property_name] = getattr(old_format, property_name)

    return workbook.add_format(properties)


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def excel_range_string_to_indices(range_string):
    try:
        first_col_name, first_row, last_col_name, last_row = EXCEL_RANGE_PATTERN.findall(
            range_string)[0]
    except IndexError:
        raise ValueError("Invalid range string.")

    first_col_index = col2num(first_col_name) - 1
    first_row_index = int(first_row) - 1
    last_col_index = col2num(last_col_name) - 1
    last_row_index = int(last_row) - 1

    return (
        first_col_index,
        first_row_index,
        last_col_index,
        last_row_index
    )


def apply_border_to_cell(workbook, worksheet, row_index, col_index, format_properties):
    try:
        cell = worksheet.table[row_index][col_index]
        new_format = duplicate_xlsxwriter_format_object(workbook, cell.format)

        # Convert properties in the constructor to method calls.
        for key, value in format_properties.items():
            getattr(new_format, 'set_' + key)(value)

        # Update cell object
        worksheet.table[row_index][col_index] = cell = cell._replace(format=new_format)
    except KeyError:
        format = workbook.add_format(format_properties)
        worksheet.write(row_index, col_index, None, format)


def apply_outer_border_to_range(workbook, worksheet, options=None):
    options = options or {}

    border_style = options.get("border_style", 1)
    range_string = options.get("range_string", None)

    if range_string is not None:
        first_col_index, first_row_index, last_col_index, last_row_index = excel_range_string_to_indices(
            range_string)
    else:
        first_col_index = options.get("first_col_index", None)
        last_col_index = options.get("last_col_index", None)
        first_row_index = options.get("first_row_index", None)
        last_row_index = options.get("last_row_index", None)

        all_are_none = all(map(lambda x: x is None, [
            first_col_index,
            last_col_index,
            first_row_index,
            last_row_index,
        ]))

        if all_are_none:
            raise Exception("You need to specify the range")

    for row_index in range(first_row_index, last_row_index + 1):
        left_border = {
            "left": border_style,
        }
        right_border = {
            "right": border_style,
        }

        apply_border_to_cell(workbook, worksheet, row_index, first_col_index, left_border)
        apply_border_to_cell(workbook, worksheet, row_index, last_col_index, right_border)

    for col_index in range(first_col_index, last_col_index + 1):
        top_border = {
            "top": border_style,
        }

        bottom_border = {
            "bottom": border_style,
        }

        apply_border_to_cell(workbook, worksheet, first_row_index, col_index, top_border)
        apply_border_to_cell(workbook, worksheet, last_row_index, col_index, bottom_border)

    top_left_border = {
        "top": border_style,
        "left": border_style,
    }
    apply_border_to_cell(workbook, worksheet, first_row_index, first_col_index, top_left_border)

    top_right_border = {
        "top": border_style,
        "right": border_style,
    }
    apply_border_to_cell(workbook, worksheet, first_row_index, last_col_index, top_right_border)

    bottom_left_border = {
        "bottom": border_style,
        "left": border_style,
    }
    apply_border_to_cell(workbook, worksheet, last_row_index, first_col_index, bottom_left_border)

    bottom_right_border = {
        "bottom": border_style,
        "right": border_style,
    }
    apply_border_to_cell(workbook, worksheet, last_row_index, last_col_index, bottom_right_border)


def var_summary_to_excelold(var_summary_df, workbook=None, sheet_name=None):
    # excel相关格式
    # workbook = xlsxwriter.Workbook(output, {'nan_inf_to_errors': True})
    # worksheet = workbook.add_worksheet('分箱详情')
    worksheet = workbook.add_worksheet(sheet_name)
    # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'center',
        'font_name': 'Consolas',
        'font_size': 11,
        'border': 1,
        'bottom': 2,
    })
    varname_formats = workbook.add_format(
        {'color': 'blue', 'underline': True, 'font_name': 'Consolas', 'font_size': 11, })
    percentage_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, })  # 百分数保留2位
    pct_bad_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, 'font_color': '#980101'})  # 百分数保留2位
    pct_good_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, 'font_color': '#0050aa'})  # 百分数保留2位
    float_formats4 = workbook.add_format(
        {'num_format': '#,##0.0000', 'font_name': 'Consolas', 'font_size': 11, })  # 浮点型保留4位
    integer_formats = workbook.add_format(
        {'num_format': '#,##0', 'font_name': 'Consolas', 'font_size': 11, })  # 整数型
    font_formats = workbook.add_format({'font_name': 'Consolas', 'font_size': 11, })  # 整数型
    row_formats = workbook.add_format({'bottom': 1})  # 下边框线

    def colnum_string(n):
        """将数字n转成excel的列符号"""
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def write_df_to_excel_by_row(index, feature, df, workbook, worksheet, start_row, start_column):
        """按列写入df,因为要单独给每一列一种格式"""
        for col_num, column_name in enumerate(df.columns.values):
            if column_name in ['positive_rate', 'Pct_Bin']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       percentage_formats)
            elif column_name in ['Pct_Bad']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       pct_bad_formats)
            elif column_name in ['Pct_Good']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       pct_good_formats)
            elif column_name in ['var_name']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       varname_formats)
            elif column_name in ['iv', 'IV']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       float_formats4)
            elif column_name in ['Total', 'Bad', 'Good']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       integer_formats)
            elif column_name in ['woe']:
                # 输入到Excel中的时候，将WoE的值放大100倍，方便观察
                # worksheet.write_column(start_row, start_column + col_num,
                #                        # (df[column_name] * 100).astype(int).values.tolist(),
                #                        integer_formats)
                worksheet.write_column(start_row, start_column + col_num,
                                       df[column_name].values.tolist(),
                                       float_formats4)
            else:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       font_formats)

        # 给最后一行加上下框线,用于区分变量
        end_row, end_column = start_row + df.shape[0], df.shape[1]
        worksheet.conditional_format('A{end_row}:{end_column}{end_row}' \
                                     .format(end_row=end_row, end_column=colnum_string(end_column)), \
                                     {'type': 'no_blanks', 'format': row_formats})

        # 给BadRate列加上条件格式
        badrate_index = colnum_string(start_column + df.columns.get_loc('Bad_Rate') + 1)
        worksheet.conditional_format(
            '{c}{start_row}:{c}{end_row}'.format(start_row=start_row, end_row=end_row, c=badrate_index),
            {'type': 'data_bar', 'bar_color': '#f0c4c4', 'bar_solid': True})

        # 给每个分箱变量,绘制badrate曲线和woe分布图
        # badrate
        graph = workbook.add_chart({'type': 'line'})
        range_column_index = df.columns.get_loc('range') + start_column
        badrate_column_index = df.columns.get_loc('Bad_Rate') + start_column

        range_categories = [worksheet.get_name(), start_row, range_column_index, end_row - 1,
                            range_column_index]
        range_value = [worksheet.get_name(), start_row, badrate_column_index, end_row - 1,
                       badrate_column_index]
        graph.add_series({'name': 'BadRate', 'categories': range_categories, 'values': range_value,
                          'marker': {'type': 'circle'}, 'data_labels': {'value': True}})
        graph.set_size({'width': 450, 'height': 200})
        graph.set_title(
            {'name': 'Bad Rate - {}'.format(feature), 'overlay': False,
             'name_font': {'name': '微软雅黑', 'size': 9, 'bold': True}})
        graph.set_x_axis({'line': {'none': True}})
        graph.set_y_axis({'line': {'none': True}})
        graph.set_legend({'none': True})  # 设置图例
        worksheet.insert_chart(start_row + index * 7, end_column + 1, graph)

        # WoE
        column_graph = workbook.add_chart({'type': 'column'})
        range_column_index = df.columns.get_loc('range') + start_column
        badrate_column_index = df.columns.get_loc('woe') + start_column

        range_categories = [worksheet.get_name(), start_row, range_column_index, end_row - 1,
                            range_column_index]
        range_value = [worksheet.get_name(), start_row, badrate_column_index, end_row - 1,
                       badrate_column_index]
        column_graph.add_series({'name': 'woe', 'categories': range_categories, 'values': range_value,
                                 'marker': {'type': 'circle'}, 'data_labels': {'value': True}})
        column_graph.set_size({'width': 450, 'height': 200})  # TODO 大小改成和单元格的大小一致,以单元格的大小为单位
        column_graph.set_title(
            {'name': 'WoE - {}'.format(feature), 'overlay': False,
             'name_font': {'name': '微软雅黑', 'size': 9, 'bold': True}})
        column_graph.set_x_axis({'line': {'none': True}, 'label_position': 'low'})
        column_graph.set_y_axis({'line': {'none': True}})
        column_graph.set_legend({'none': True})  # 设置图例
        worksheet.insert_chart(start_row + index * 7, end_column + 9, column_graph)

        # 给变量名称插入跳转超链接
        for i in range(1, df.shape[0] + 1):
            worksheet.write_url('B{}'.format(start_row + i),
                                'internal:{}!M{}:AC{}'.format(worksheet.get_name(), start_row + index * 7 + 1,
                                                              start_row + index * 7 + 10), string=feature,
                                tip='Jump to charts')

    # 写入标题
    for col_num, value in enumerate(var_summary_df.columns.values):
        worksheet.write(0, col_num, value, header_format)

    # 写入数据和插入图像
    start_row, start_column = 1, 0
    index = 0
    for name, single_df in var_summary_df.groupby('var_name'):
        write_df_to_excel_by_row(index, name, single_df, workbook, worksheet, start_row, start_column)
        start_row = single_df.shape[0] + start_row
        index += 1

    # 冻结窗格
    worksheet.freeze_panes(1, 2)
    worksheet.set_column(0, 1, 18)

    # # 保存为excel
    # workbook.close()


def var_summary_to_excel(var_summary_df, workbook=None, sheet_name=None):
    # excel相关格式
    # workbook = xlsxwriter.Workbook(output, {'nan_inf_to_errors': True})
    # worksheet = workbook.add_worksheet('分箱详情')
    worksheet = workbook.add_worksheet(sheet_name)
    # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'center',
        'font_name': 'Consolas',
        'font_size': 11,
        'border': 1,
        'bottom': 2,
    })
    varname_formats = workbook.add_format(
        {'color': 'blue', 'underline': True, 'font_name': 'Consolas', 'font_size': 11, })
    percentage_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, })  # 百分数保留2位
    positive_pct_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, 'font_color': '#980101'})  # 百分数保留2位
    negative_pct_formats = workbook.add_format(
        {'num_format': '0.00%', 'font_name': 'Consolas', 'font_size': 11, 'font_color': '#0050aa'})  # 百分数保留2位
    float_formats4 = workbook.add_format(
        {'num_format': '#,##0.0000', 'font_name': 'Consolas', 'font_size': 11, })  # 浮点型保留4位
    integer_formats = workbook.add_format(
        {'num_format': '#,##0', 'font_name': 'Consolas', 'font_size': 11, })  # 整数型
    font_formats = workbook.add_format({'font_name': 'Consolas', 'font_size': 11, })  # 整数型
    row_formats = workbook.add_format({'bottom': 1})  # 下边框线

    def colnum_string(n):
        """将数字n转成excel的列符号"""
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def write_df_to_excel_by_row(index, feature, df, workbook, worksheet, start_row, start_column):
        """按列写入df,因为要单独给每一列一种格式"""
        for col_num, column_name in enumerate(df.columns.values):
            if column_name in ['positive_rate', 'total_pct']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       percentage_formats)
            elif column_name in ['positive_pct']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       positive_pct_formats)
            elif column_name in ['negative_pct']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       negative_pct_formats)
            elif column_name in ['var_name']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       varname_formats)
            elif column_name in ['iv', 'IV']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       float_formats4)
            elif column_name in ['total', 'positive', 'negative']:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       integer_formats)
            elif column_name in ['woe']:
                # 输入到Excel中的时候，将WoE的值放大100倍，方便观察
                # worksheet.write_column(start_row, start_column + col_num,
                #                        # (df[column_name] * 100).astype(int).values.tolist(),
                #                        integer_formats)
                worksheet.write_column(start_row, start_column + col_num,
                                       df[column_name].values.tolist(),
                                       float_formats4)
            else:
                worksheet.write_column(start_row, start_column + col_num, df[column_name].values.tolist(),
                                       font_formats)

        # 给最后一行加上下框线,用于区分变量
        end_row, end_column = start_row + df.shape[0], df.shape[1]
        worksheet.conditional_format('A{end_row}:{end_column}{end_row}' \
                                     .format(end_row=end_row, end_column=colnum_string(end_column)), \
                                     {'type': 'no_blanks', 'format': row_formats})

        # 给BadRate列加上条件格式
        badrate_index = colnum_string(start_column + df.columns.get_loc('positive_rate') + 1)
        worksheet.conditional_format(
            '{c}{start_row}:{c}{end_row}'.format(start_row=start_row, end_row=end_row, c=badrate_index),
            {'type': 'data_bar', 'bar_color': '#f0c4c4', 'bar_solid': True})

        # 给每个分箱变量,绘制badrate曲线和woe分布图
        # badrate
        graph = workbook.add_chart({'type': 'line'})
        range_column_index = df.columns.get_loc('range') + start_column
        badrate_column_index = df.columns.get_loc('positive_rate') + start_column

        range_categories = [worksheet.get_name(), start_row, range_column_index, end_row - 1,
                            range_column_index]
        range_value = [worksheet.get_name(), start_row, badrate_column_index, end_row - 1,
                       badrate_column_index]
        graph.add_series({'name': 'positive_rate', 'categories': range_categories, 'values': range_value,
                          'marker': {'type': 'circle'}, 'data_labels': {'value': True}})
        graph.set_size({'width': 450, 'height': 200})
        graph.set_title(
            {'name': 'positive_rate - {}'.format(feature), 'overlay': False,
             'name_font': {'name': '微软雅黑', 'size': 9, 'bold': True}})
        graph.set_x_axis({'line': {'none': True}})
        graph.set_y_axis({'line': {'none': True}})
        graph.set_legend({'none': True})  # 设置图例
        worksheet.insert_chart(start_row + index * 7, end_column + 1, graph)

        # WoE
        column_graph = workbook.add_chart({'type': 'column'})
        range_column_index = df.columns.get_loc('range') + start_column
        badrate_column_index = df.columns.get_loc('woe') + start_column

        range_categories = [worksheet.get_name(), start_row, range_column_index, end_row - 1,
                            range_column_index]
        range_value = [worksheet.get_name(), start_row, badrate_column_index, end_row - 1,
                       badrate_column_index]
        column_graph.add_series({'name': 'woe', 'categories': range_categories, 'values': range_value,
                                 'marker': {'type': 'circle'}, 'data_labels': {'value': True}})
        column_graph.set_size({'width': 450, 'height': 200})  # TODO 大小改成和单元格的大小一致,以单元格的大小为单位
        column_graph.set_title(
            {'name': 'WoE - {}'.format(feature), 'overlay': False,
             'name_font': {'name': '微软雅黑', 'size': 9, 'bold': True}})
        column_graph.set_x_axis({'line': {'none': True}, 'label_position': 'low'})
        column_graph.set_y_axis({'line': {'none': True}})
        column_graph.set_legend({'none': True})  # 设置图例
        worksheet.insert_chart(start_row + index * 7, end_column + 9, column_graph)

        # 给变量名称插入跳转超链接
        for i in range(1, df.shape[0] + 1):
            worksheet.write_url('B{}'.format(start_row + i),
                                'internal:{}!M{}:AC{}'.format(worksheet.get_name(), start_row + index * 7 + 1,
                                                              start_row + index * 7 + 10), string=feature,
                                tip='Jump to charts')

    # 写入标题
    for col_num, value in enumerate(var_summary_df.columns.values):
        worksheet.write(0, col_num, value, header_format)

    # 写入数据和插入图像
    start_row, start_column = 1, 0
    index = 0
    for name, single_df in var_summary_df.groupby('var_name'):
        write_df_to_excel_by_row(index, name, single_df, workbook, worksheet, start_row, start_column)
        start_row = single_df.shape[0] + start_row
        index += 1

    # 冻结窗格
    worksheet.freeze_panes(1, 2)
    worksheet.set_column(0, 1, 18)

    # # 保存为excel
    # workbook.close()


if __name__ == "__main__":
    import pandas as pd
    import numpy as np
    import os

    # 创建一个数据
    df = pd.DataFrame(np.random.randn(182, 9), columns=list('ABCDEFGHI'))
    column_list = df.columns
    # 使用XlsxWriter引擎创建一个pandas Excel writer。
    writer = pd.ExcelWriter(os.path.join('..','tests','test_report2excel.xlsx'), engine='xlsxwriter')

    df.to_excel(writer, index=False)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    ###########################################
    # worksheet.set_landscape()
    # worksheet.set_paper(8)
    # worksheet.set_margins(0.787402, 0.787402, 0.5, 0.787402)

    apply_outer_border_to_range(
        workbook,
        worksheet,
        {
            "range_string": "C10:M20",
            "border_style": 5,
        },
    )

    # 关闭workbook
    workbook.close()



#==============================================================================
# File: scorecard.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: scorecard.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-20
'''

import re
from copy import deepcopy

import numpy as np
import pandas as pd
from sklearn.base import BaseEstimator
from sklearn.linear_model import LogisticRegression

from .transformer import FeatureBin, WoeTransformer
from .utils import to_ndarray, save_json, load_json

RE_NUM = r'-?\d+(.\d+)?'
RE_SEP = r'[~-]'
RE_BEGIN = r'(-inf|{num})'.format(num=RE_NUM)
RE_END = r'(inf|{num})'.format(num=RE_NUM)
RE_RANGE = r'\[{begin}\s*{sep}\s*{end}\)'.format(
    begin=RE_BEGIN,
    end=RE_END,
    sep=RE_SEP,
)

NUMBER_EXP = re.compile(RE_RANGE)

NUMBER_EMPTY = -9999999
NUMBER_INF = 1e10
FACTOR_EMPTY = 'MISSING'
FACTOR_UNKNOWN = 'UNKNOWN'
ELSE_GROUP = 'else'


class ScoreCard(BaseEstimator):
    def __init__(self, pdo=50, rate=2, odds=15, base_score=600,
                 card={}, combiner={}, transer=None, AB={}, **kwargs):
        """

        Args:
            pdo (int):point double odds;;;当odds增加一倍，评分增加的分数
            rate (int):
            odds (int): odds at base point;;;基准分值对应的odds
            base_score (int): base point;;;基准分数
            card (dict): 评分卡
            combiner (autobmt.FeatureBin): 分箱规则
            transer (autobmt.WoeTransformer): 变量分箱对应的woe值
            **kwargs:
        """
        ##实际意义为当比率为1/15，输出基准评分600，当比率为基准比率2倍时，1/7.5，基准分下降50分，为550
        self.pdo = pdo  # point double odds;;;当odds增加一倍，评分增加的分数
        self.rate = rate  #
        self.odds = odds  # odds at base point;;;基准分值对应的odds
        self.base_score = base_score  # base point;;;基准分数
        self.AB = AB    #自定义的大A，大B

        if AB:
            self.factor = self.AB['B']
            self.offset = self.AB['A']
        else:
            self.factor = pdo / np.log(rate)  # 大B;;;B=72.13475204444818
            self.offset = base_score - (pdo / np.log(rate)) * np.log(odds)  # 大A;;;A=404.65547021957406

        self._combiner = combiner
        self.transer = transer
        self.model = LogisticRegression(**kwargs)

        self._feature_names = None

        self.card = card
        if card:
            self.load(card)

    def __len__(self):
        return len(self.card.keys())

    def __contains__(self, key):
        return key in self.card

    def __getitem__(self, key):
        return self.card[key]

    def __setitem__(self, key, value):
        self.card[key] = value

    def __iter__(self):
        return iter(self.card)

    @property
    def coef_(self):
        """ 逻辑回归模型系数
        """
        return self.model.coef_[0]

    @property
    def intercept_(self):
        """ 逻辑回归模型截距
        """
        return self.model.intercept_[0]

    @property
    def n_features_(self):
        """ 变量个数
        """
        return (self.coef_ != 0).sum()

    @property
    def features_(self):
        """ 变量列表
        """
        if not self._feature_names:
            self._feature_names = list(self.card.keys())

        return self._feature_names

    @property
    def combiner(self):
        if not self._combiner:
            # 如果不存在，则生成新的分箱器
            rules = {}
            for key in self.card:
                rules[key] = self.card[key]['bins']

                self._combiner = FeatureBin().manual_bin(rules)

        return self._combiner

    def fit(self, X, y):
        """
        Args:
            X (2D DataFrame): 变量
            Y (array-like): 目标变量列表
        """
        self._feature_names = X.columns.tolist()

        for f in self.features_:
            if f not in self.transer:
                raise Exception('column \'{f}\' is not in transer'.format(f=f))

        self.model.fit(X, y)
        self.card = self._generate_rules()

        # keep sub_score-median of each feature, as `base_effect` for reason-calculation
        sub_score = self.woe_to_score(X)
        # self.base_effect = pd.Series(
        #     np.median(sub_score, axis=0),
        #     index=self.features_
        # )

        return self

    def predict(self, X, **kwargs):
        """预测分数
        Args:
            X (2D-DataFrame|dict): 需要去预测的变量
            return_sub (Bool): 是否需要返回特征中每个箱子的得分
            default (str|number): 未知特征的默认分数，' min '(默认)，' max '

        Returns:
            array-like: 预测的分数
            DataFrame|dict: 每个特征对应的分数
        """
        if isinstance(X, pd.DataFrame):
            X = X[self.features_]

        bins = self.combiner.transform(X)
        res = self.bin_to_score(bins, **kwargs)
        return res

    def bin_to_score(self, bins, return_sub=False, default='min'):
        """
        通过分箱值直接算分
        Args:
            bins (2D-DataFrame|dict): 使用分箱值替换后的变量
            return_sub (bool): 是否需要返回特征中每个箱子的得分
            default (bool): 未知特征的默认分数，' min '(默认)，' max '

        Returns:

        """
        score = 0
        res = bins.copy()
        for col, rule in self.card.items():
            s_map = rule['scores']
            b = bins[col]

            # set default value for empty group
            default_value = default
            if default == 'min':
                default_value = np.min(s_map)
            elif default == 'max':
                default_value = np.max(s_map)
            elif isinstance(default, str):
                raise ValueError(f'default `{default}` is not valid, only support `min`, `max` or number')

            # append default value to the end of score map
            s_map = np.append(s_map, default_value)

            # # set default group to min score
            # if np.isscalar(b):
            #     b = np.argmin(s_map) if b == self.EMPTY_BIN else b
            # else:
            #     b[b == self.EMPTY_BIN] = np.argmin(s_map)

            # replace score
            res[col] = s_map[b]
            score += s_map[b]

        if return_sub:
            return score, res
        else:
            return score

    def predict_proba(self, X):
        """
        预测概率
        Args:
            X (2D array-like): 需要去预测的变量

        Returns:
            2d array: 预测的概率值（包括正样本和负样本的概率值）
        """
        proba = self.score_to_proba(self.predict(X))
        return np.stack((1 - proba, proba), axis=1)

    def _generate_rules(self):
        if not self._check_rules(self.combiner, self.transer):
            raise Exception('generate failed')

        rules = {}

        for idx, key in enumerate(self.features_):
            weight = self.coef_[idx]

            if weight == 0:
                continue

            # woe = self.transer[key]['woe']
            woe = list(self.transer[key].values())

            rules[key] = {
                'bins': self.combiner[key],
                'woes': woe,
                'weight': weight,
                'scores': self.woe_to_score(woe, weight=weight),
            }

        return rules

    def _check_rules(self, combiner, transer):
        for col in self.features_:
            if col not in combiner:
                raise Exception('column \'{col}\' is not in combiner'.format(col=col))

            if col not in transer:
                raise Exception('column \'{col}\' is not in transer'.format(col=col))

            l_c = len(combiner.export(bin_format=False)[col])
            # l_t = len(transer[col]['woe'])
            l_t = len(transer[col].values())

            if l_c == 0:
                continue

            if np.issubdtype(combiner[col].dtype, np.number):
                if l_c != l_t - 1:
                    raise Exception(
                        'column \'{col}\' is not matched, assert {l_t} bins but given {l_c}'.format(col=col, l_t=l_t,
                                                                                                    l_c=l_c + 1))
            else:
                if l_c != l_t:
                    raise Exception(
                        'column \'{col}\' is not matched, assert {l_t} bins but given {l_c}'.format(col=col, l_t=l_t,
                                                                                                    l_c=l_c))

        return True

    def proba_to_score(self, prob):
        """概率转分

        odds = (1 - prob) / prob    #good:bad
        score = factor * log(odds) + offset

        odds = prob / (1 - prob)    #bad:good
        score = offset - factor * log(odds)

        log(odds) = intercept+woe1*coef1+woe2*coef2+woe3*coef3
        """
        # return self.factor * np.log((1 - prob) / prob) + self.offset
        return self.offset - self.factor * np.log(prob / (1 - prob))

    def woe_sum_to_score(self, woe, weight=None):
        """通过woe计算分

        odds = (1 - prob) / prob    #good:bad
        score = factor * log(odds) + offset

        odds = prob / (1 - prob)    #bad:good
        score = offset - factor * log(odds)

        log(odds) = intercept+woe1*coef1+woe2*coef2+woe3*coef3
        """
        woe = to_ndarray(woe)

        if weight is None:
            weight = self.coef_

        z_cols = weight * woe
        z = self.intercept_ + np.sum(z_cols, axis=1)

        return self.offset - self.factor * z

    def score_to_proba(self, score):
        """分转概率

        Returns:
            array-like|float: 正样本的概率【即：1的概率】
        """
        return 1 / (1 + np.exp((score - self.offset) / self.factor))

    def woe_to_score(self, woe, weight=None):
        """通过woe计算分
        score = A - Blog(odds) = A - B( β0 + β1x1 + … βnxn) = (A - Bβ0) - Bβ1 x1 - … Bβn xn = sum((A - Bβ0)/n - Bβ1 x1 - … Bβn xn)
        """
        woe = to_ndarray(woe)

        if weight is None:
            weight = self.coef_

        b = (self.offset - self.factor * self.intercept_) / self.n_features_  # (A - Bβ0)/n
        s = -self.factor * weight * woe  # -B*βn*xn

        # drop score whose weight is 0
        mask = 1
        if isinstance(weight, np.ndarray):
            mask = (weight != 0).astype(int)

        return (s + b) * mask

    def export(self, to_dataframe=False, to_json=None, to_csv=None, decimal=2):
        """生成一个评分卡对象

        Args:
            to_dataframe (bool): 生成的评分卡是1个DataFrame
            to_json (str|IOBase): 生成的评分卡写出json文件
            to_csv (filepath|IOBase): 生成的评分卡写出csv文件

        Returns:
            dict: 评分卡
        """
        card = dict()
        combiner = self.combiner.export(bin_format=True, index=False)

        for col, rule in self.card.items():
            s_map = rule['scores']
            bins = combiner[col]
            woe_map = np.zeros(len(bins))
            if 'woes' in rule:
                woe_map = rule['woes']
            weight = np.nan
            if 'weight' in rule:
                weight = rule['weight']
            card[col] = dict()

            for i in range(len(bins)):
                # card[col][bins[i]] = round(s_map[i], decimal)
                card[col][bins[i]] = [round(s_map[i], decimal), woe_map[i], weight]

        if to_json is not None:
            save_json(card, to_json)

        if to_dataframe or to_csv is not None:
            rows = list()
            for feature in card:
                for value, score in card[feature].items():
                    rows.append({
                        'feature': feature,
                        'value': value,
                        'score': score[0],
                        'woe': score[1],
                        'weight': score[2],
                    })

            card = pd.DataFrame(rows)

        if to_csv is not None:
            return card.to_csv(to_csv)

        return card

    def _is_numeric(self, bins):
        m = NUMBER_EXP.match(bins[0])

        return m is not None

    def _numeric_parser(self, bins):
        l = list()

        for item in bins:

            # if re.compile('{}.nan'.format(RE_NUM)).match(item):
            if item == 'nan':
                l.append(np.nan)
                continue

            m = NUMBER_EXP.match(item)
            split = m.group(3)

            if split == 'inf':
                # split = np.inf
                continue

            split = float(split)

            l.append(split)

        return np.array(l)

    def parse_bins(self, bins):
        """解析格式化的分箱值
        """
        if self._is_numeric(bins):
            return self._numeric_parser(bins)

        l = list()

        for item in bins:
            if item == ELSE_GROUP:
                l.append(item)
            else:
                l.append(item.split(','))

        return np.array(l, dtype=object)

    def _parse_rule(self, rule):
        bins = self.parse_bins(list(rule.keys()))
        v = list(rule.values())
        if isinstance(v[0], list):
            scores = np.array([i[0] for i in v])
        else:
            scores = np.array(v)

        return {
            'bins': bins,
            'scores': scores,
        }

    def load(self, card=None):
        """
        加载评分卡
        Args:
            card: 从dict或json文件加载评分卡

        Returns:

        """
        card = deepcopy(card)
        if not isinstance(card, dict):
            card = load_json(card)

        for feature in card:
            card[feature] = self._parse_rule(card[feature])

        self.card = card
        self._combiner = {}

        return self



#==============================================================================
# File: statistics.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: statistics.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-03
'''

import re

import numpy as np
import pandas as pd
import statsmodels.api as sm
from joblib import Parallel, delayed

import autobmt
from .metrics import psi, get_auc, get_ks
from .utils import is_continuous, to_ndarray, np_count, support_dataframe, split_points_to_bin


# TODO compare_inflection_point和calc_iv_and_inflection_point考虑要合并

def compare_inflection_point(df_summary):
    """
    比较数据集的拐点趋势
    切记，数据集使用type区分，因此一定要包含type列，数据集是分好箱的数据集，不是原始数据集而是转woe后的
    因为需要转换计算badrate
    Args:
        df_summary (DataFrame):分好箱的数据集，可能包含train/test/oot

    Returns:
        DataFrame: 拐点信息
    """

    def __calc_group(data):
        range = data.range.tolist()
        nan_bin = '{}.nan'.format(len(range) - 1)
        if nan_bin in range:
            data = data[~data['range'].isin([nan_bin])]
        badrate = data.positive_rate
        is_monotonic = badrate.is_monotonic_decreasing or badrate.is_monotonic_increasing

        inflection_point_index, inflection_shape = get_inflection_point_index(badrate)
        return pd.Series({'is_monotonic': is_monotonic, 'bin_count': len(range),
                          'inflection_point_num': len(inflection_point_index),
                          'inflection_point_index': inflection_point_index,
                          'inflection_shape': inflection_shape})

    data_inflection_df = df_summary.groupby('var_name').apply(__calc_group)
    return data_inflection_df


def merge_rows_one_row_df(df, name="", stepname=None):
    """将一个多行的dataframe合并成只有一行的dataframe"""
    tmp_arr = []
    for i in range(df.shape[0]):
        tmp = df.iloc[i, :].add_prefix("{}_{}".format(df.index[i], name))
        tmp_arr.append(tmp)

    result_df = pd.DataFrame(pd.concat(tmp_arr, axis=0)).T
    if stepname is not None:  # 合并成一行后,增加一列标识，用于和别的评估进行区分
        result_df['stepname'] = stepname
    return result_df


def calc_iv_and_inflection_point(df, target='target', bin_func=None, bin_format={}):
    """
    计算iv，拐点
    Args:
        df (DataFrame):分好箱转换后的数据集
        target (str):目标变量
        bin_func (str): 分箱方法
        bin_format (dict): 格式化好的分箱点

    Returns:
        DataFrame
    """

    def __calc_group(data, bin_func):
        range = data.range.tolist()
        nan_bin = '{}.nan'.format(len(range) - 1)
        if nan_bin in range:
            data = data[~data['range'].isin([nan_bin])]
        bin_badrate = data.positive_rate
        inflection_point_index, inflection_shape = get_inflection_point_index(bin_badrate)
        return pd.Series({'IV': data.IV.get(0), 'bin_count': len(range),
                          'inflection_point_num': len(inflection_point_index),
                          'inflection_point_index': inflection_point_index,
                          'inflection_shape': inflection_shape, 'bin_func': bin_func})

    summary = Parallel(n_jobs=-1)(
        delayed(calc_bin_summary)(df[[col, target]], bin_col=col, bin=False, target=target, is_sort=False,
                                  bin_format=bin_format) for col in df.columns if col not in [target])
    var_summary = pd.concat(summary, axis=0)

    var_iv_inflection_df = var_summary.groupby('var_name').apply(__calc_group, bin_func)
    return var_iv_inflection_df


def calc_var_summary(df, bin_format={}, include_cols=[], target='target', need_bin=True, **kwargs):
    """
    计算所有变量的详情
    Args:
        df (DataFrame): 含有目标变量及分箱后的数据集
        bin_format (dict): 格式化好的分箱点
        include_cols (list): 需要统计的特征
        target (str): 目标值变量名称
        need_bin (bool): 是否需要分箱

    Returns:
        DataFrame: 变量分箱详情
    """

    if include_cols:
        cols = np.array(include_cols)
    else:
        cols = np.array(df.columns)

    if not need_bin:
        kwargs = {'bin': False}
    summary = Parallel(n_jobs=-1)(
        delayed(calc_bin_summary)(df[[col, target]], bin_col=col, target=target, is_sort=False, bin_format=bin_format,
                                  **kwargs) for col in cols if col != target)
    var_summary = pd.concat(summary, axis=0)
    return var_summary


def calc_bin_summary(df, bin_col='score', bin=10, target='target', is_sort=True, method='equal_freq',
                     is_need_monotonic=False, bin_format={}, **kwargs):
    """
    变量分箱详情
    Args:
        df (DataFrame): 含目标变量的数据集
        bin_col (str): 需要计算的列名
        bin (int): 分几箱
        target (str):目标变量列名
        method (str):分箱方法；'dt'、'chi'、'equal_freq'、'kmeans'四种供选择
        is_sort (bool):是否需要排序，默认True，倒序排序
        **kwargs:

    Returns:
        DataFrame: 变量分箱详情
    """

    # negative:good，positive:bad
    def __calc_group(data_, var_name):
        """获取分组的人数,好坏人数"""
        count = len(data_)
        bad_num = data_.y.sum()
        good_num = count - bad_num
        return pd.Series(
            {'var_name': var_name, 'min': min(data_.x), 'max': max(data_.x), 'positive': bad_num, 'negative': good_num,
             'total': count})

    data = pd.DataFrame({'x': df[bin_col], 'y': df[target]})
    total = len(data)
    positive_count = data.y.sum()
    negative_count = total - positive_count

    data['range'] = 0
    if bin is False:
        data['range'] = data.x
    elif isinstance(bin, (list, np.ndarray, pd.Series)):
        if len(bin) < len(data.x):
            bin = split_points_to_bin(data.x, bin)

        data['range'] = bin
    elif isinstance(bin, int):
        fb = autobmt.FeatureBin()
        fb.fit(data.x, data.y, n_bins=bin, method=method, is_need_monotonic=is_need_monotonic, **kwargs)
        data['range'] = fb.transform(data.x)
        bin_format = fb.export()
        bin_format[bin_col] = bin_format.pop('x')

    bin_g = data.groupby(data['range'], dropna=False).apply(__calc_group, var_name=bin_col)

    if is_sort:
        bin_g.sort_values(by='min', ascending=False, inplace=True)  # 正常

    bin_g['positive_rate'] = bin_g['positive'] / bin_g['total']  # bad_rate,,,区间_正样本比率
    bin_g['negative_rate'] = bin_g['negative'] / bin_g['total']  # 区间_负样本比率

    bin_g['odds'] = bin_g['positive'] / bin_g['negative']

    # bin_g['positive_pct'] = bin_g['positive'] / positive_count  # 区间正样本占比
    bin_g['positive_pct'] = bin_g['positive'].map(
        lambda x: 1 / positive_count if x == 0 else x / positive_count)  # 区间正样本占比
    # bin_g['negative_pct'] = bin_g['negative'] / negative_count  # 区间负样本/总体负样本
    bin_g['negative_pct'] = bin_g['negative'].map(
        lambda x: 1 / negative_count if x == 0 else x / negative_count)  # 区间负样本/总体负样本
    bin_g['total_pct'] = bin_g['total'] / total  # 区间总人数/总体总人数

    bin_g['cum_negative_pct'] = bin_g['negative_pct'].cumsum()  # 累计负样本人数占比

    cum_positive = bin_g['positive'].cumsum()  # 累计正样本人数
    cum_total = bin_g['total'].cumsum()  # 累计总人数

    bin_g['cum_total_pct'] = cum_total / total  # 累计通过人数占比，累计捕获率，agg2['cum_total_prop'] = cum_total / all_total
    bin_g['cum_positive_pct'] = bin_g['positive_pct'].cumsum()  # 查全率,,,累计正样本人数占比
    bin_g['cum_positive_rate'] = cum_positive / cum_total  # 查准率,,,累计捕获的样本中正样本的占比

    bin_g['ks'] = bin_g['cum_positive_pct'] - bin_g['cum_negative_pct']  # 累计正样本人数占比/累计负样本人数占比

    bin_g['lift'] = bin_g['positive_rate'] / (positive_count / total)
    if bin_g['ks'].sum() < 0:
        bin_g['ks'] = -bin_g['ks']
        bin_g['cum_negative_pct'] = bin_g.loc[::-1, 'negative_pct'].cumsum()[::-1]  # 累计负样本人数占比
        cum_positive_rev = bin_g.loc[::-1, 'positive'].cumsum()[::-1]  # 累计正样本人数
        cum_total_rev = bin_g.loc[::-1, 'total'].cumsum()[::-1]  # 累计总人数
        bin_g['cum_total_pct'] = cum_total_rev / total  # 累计通过人数占比，累计捕获率，agg2['cum_total_prop'] = cum_total / all_total
        bin_g['cum_positive_pct'] = bin_g.loc[::-1, 'positive_pct'].cumsum()[::-1]  # 查全率,,,累计正样本人数占比
        bin_g['cum_positive_rate'] = cum_positive_rev / cum_total_rev  # 查准率,,,累计捕获的样本中正样本的占比

    bin_g['cum_lift'] = bin_g['cum_positive_pct'] / bin_g['cum_total_pct']

    bin_g['woe'] = bin_g.apply(lambda x: np.log(x['positive_pct'] / x['negative_pct']), axis=1)
    bin_g['iv'] = (bin_g['positive_pct'] - bin_g['negative_pct']) * bin_g.woe
    bin_g['IV'] = bin_g.iv.sum()

    bin_g.index.name = 'range'
    bin_g = bin_g.reset_index()

    if bin_col in bin_format:
        range_format = {int(re.match(r"^(\d+)\.", i).group(1)): i for i in bin_format[bin_col]}
        bin_g['range_num'] = bin_g['range']
        bin_g['range'] = bin_g['range'].map(range_format)
    else:
        bin_g['range_num'] = bin_g['range'].fillna(len(bin_g))

    return bin_g


def calc_woe_iv(df, col_name='default_name', bin_format={}, target='target'):
    """
    已对齐IV的计算方式
    计算单变量详情，woe,iv值
    Args:
        df: 含有目标变量及分箱后的数据集
        col_name: 单变量名称
        bin_format: 格式化好的分箱点
        target: 目标值变量名称

    Returns:

    """

    def __calc_group(data_, var_name):
        """获取分组的人数,好坏人数"""
        count = len(data_)
        bad_num = data_.Y.sum()
        good_num = count - bad_num

        return pd.Series({'var_name': var_name, 'Total': count, 'Bad': bad_num, 'Good': good_num})

    X, Y = df[col_name], df[target]

    data = pd.DataFrame({'X': X, 'Y': Y})

    bin_g = data.groupby(data['X'], dropna=False).apply(__calc_group, var_name=col_name)
    total = data.Y.count()
    bad_count = (data.Y == 1).sum()
    good_count = (data.Y == 0).sum()
    bin_g['Bad_Rate'] = bin_g['Bad'] / bin_g['Total']  # bad_rate
    # bin_g['Pct_Bad'] = bin_g['Bad'] / bad_count  # bad_人数占比
    # bin_g['Pct_Good'] = bin_g['Good'] / good_count  # good_人数占比
    bin_g['Pct_Bad'] = bin_g['Bad'].map(lambda x: 1 / bad_count if x == 0 else x / bad_count)  # bad_人数占比
    bin_g['Pct_Good'] = bin_g['Good'].map(lambda x: 1 / good_count if x == 0 else x / good_count)  # good_人数占比
    bin_g['Pct_Bin'] = bin_g['Total'] / total  # 总人数占比
    # bin_g['累计坏人数'] = bin_g['Good'].cumsum()
    # bin_g['累计好人数'] = bin_g['Good'].cumsum()
    # bin_g['累计坏人数占比'] = bin_g['Good'].cumsum() / bad_count
    # bin_g['累计好人数占比'] = bin_g['Good'].cumsum() / good_count
    # bin_g['woe'] = bin_g.apply(
    #     lambda x: 0.0 if x['Pct_Good'] == 0 or x['Pct_Bad'] == 0 else round(np.log(x['Pct_Bad'] / x['Pct_Good']),
    #                                                                         5),
    #     axis=1)
    bin_g['woe'] = bin_g.apply(lambda x: np.log(x['Pct_Bad'] / x['Pct_Good']), axis=1)
    # bin_g['ks'] = abs(bin_g['累计坏人数占比'] - bin_g['累计好人数占比'])
    bin_g['iv'] = (bin_g['Pct_Bad'] - bin_g['Pct_Good']) * bin_g.woe
    bin_g['IV'] = bin_g.iv.replace({np.inf: 0, -np.inf: 0}).sum()
    bin_g.index.name = 'range'
    bin_g = bin_g.reset_index()
    if col_name in bin_format:
        range_format = {int(re.match(r"^(\d+)\.", i).group(1)): i for i in bin_format[col_name]}
        bin_g['range_num'] = bin_g['range']
        bin_g['range'] = bin_g['range'].map(range_format)
    else:
        bin_g['range_num'] = bin_g['range'].fillna(len(bin_g))

    # 每个分箱字段之间加上一个空行
    # bin_g = bin_g.append(pd.Series(),ignore_index=True)
    return bin_g


def get_inflection_point_index(arr):
    """
    返回一个数组的拐点索引,以及单调情况
    Args:
        arr (array) : 数组

        + 单调递增
        - 单调递减
        u u形曲线
        ~u 倒u形曲线
        ~ 不单调
    Returns:
        array: 拐点的位置
        str: 单调标志
    """
    diff_arr = np.diff(arr).tolist()
    # 返回是单调递增(+)，还是单调递减(-)，还是不单调(~)
    monotonic_flag = check_monotonic(diff_arr)
    index_arr = []  # 记录拐点的位置
    for i in range(0, len(diff_arr) - 1):
        if np.signbit(diff_arr[i]) != np.signbit(diff_arr[i + 1]):
            index_arr.append(i + 1)
    if len(index_arr) == 1:
        monotonic_flag = "~U" if arr[1] - arr[0] > 0 else "U"

    return index_arr, monotonic_flag


def check_monotonic(arr):
    """判断是单调递增还是单调递减,先调用arr = np.diff(arr)"""
    count = np.sum(np.array(arr) > 0)
    if count == len(arr):
        return "+"  # 单调递增
    elif count == 0:
        return "-"  # 单调递减
    else:
        return "~"  # 不单调


def get_vif(frame):
    """
    计算VIF
    Args:
        frame:  (ndarray|DataFrame)

    Returns:
        Series
    """
    index = None
    if isinstance(frame, pd.DataFrame):
        index = frame.columns
        frame = frame.values

    l = frame.shape[1]
    vif = np.zeros(l)

    for i in range(l):
        X = frame[:, np.arange(l) != i]
        y = frame[:, i]

        model = sm.OLS(y, X)
        r2 = model.fit().rsquared_adj

        vif[i] = 1 / (1 - r2)

    return pd.Series(vif, index=index, name='vif')


def WOE(y_prob, n_prob):
    """计算woe值

    Args:
        y_prob: 正样本在整个正样本中的占比
        n_prob: 负样本在整个负样本中的占比

    Returns:
        number: woe 值
    """
    return np.log(y_prob / n_prob)


def _IV(feature, target):
    """计算IV值

    Args:
        feature (array-like)
        target (array-like)

    Returns:
        number: IV值
    """
    feature = to_ndarray(feature, dtype='str')
    target = to_ndarray(target)

    value = 0

    for v in np.unique(feature):
        y_prob, n_prob = probability(target, mask=(feature == v))
        value += (y_prob - n_prob) * WOE(y_prob, n_prob)

    return value


@support_dataframe()
def calc_iv(feature, target, feature_bin=None, return_name=False, col_name='feature', **kwargs):
    """计算1个特征的IV值

    Args:
        feature (array-like)
        target (array-like)
        n_bins (int): 需要分几箱
        method (str): 分箱方法；'dt'、'chi'、'equal_freq'、'kmeans'四种供选择， 默认 'dt'
        **kwargs (): bin_method_run分箱函数的其它参数
    """

    ###TODO: 考虑增加是否单调的参数

    if is_continuous(feature):

        if feature_bin is not None:
            if hasattr(feature, 'name') and feature.name in feature_bin.splits_dict:
                feature = feature_bin.transform(feature)
            else:
                feature = feature_bin.fit_transform(feature, target, method='dt', is_need_monotonic=False, **kwargs)
        else:
            if 'return_bin' in kwargs: del kwargs['return_bin']
            s, feature = autobmt.bin_method_run(feature, target, return_bin=True, is_need_monotonic=False, **kwargs)
    if return_name:
        return col_name, _IV(feature, target)
    else:
        return _IV(feature, target)


def bin_badrate(feature, target=None):
    """
    计算badrate【即正样本占比】
    Args:
        feature:
        target:

    Returns:

    """
    badrate_list = []
    bin_rate_list = []
    total = len(feature)
    uni_fea = np.sort(np.unique(feature))
    # uni_fea = np.unique(feature)
    if target is None:
        for value in uni_fea:
            # mask = (feature == value)
            mask = feature == value
            bin_rate_list.append(np.sum(mask) / total)
        return min(bin_rate_list)
    else:
        for value in uni_fea:
            # mask = (feature == value)
            mask = feature == value
            bin_target = target[mask]
            bin_badrate = np.sum(bin_target) / len(bin_target)
            bin_rate_list.append(len(bin_target) / total)
            # bin_rate_list.append(np.sum(mask) / total)
            badrate_list.append(bin_badrate)
        return badrate_list, min(bin_rate_list)


def bin_badratebase(feature, target):
    badrate_list = []
    bin_rate_list = []
    total = len(feature)
    uni_fea = np.sort(np.unique(feature))
    # uni_fea = np.unique(feature)
    for value in uni_fea:
        # mask = (feature == value)
        mask = feature == value
        bin_target = target[mask]
        bin_badrate = np.sum(bin_target) / len(bin_target)
        bin_rate_list.append(len(bin_target) / total)
        badrate_list.append(bin_badrate)

    return badrate_list, min(bin_rate_list)


def probability(target, mask=None):
    """计算目标变量占比
    """
    if mask is None:
        return 1, 1

    counts_0 = np_count(target, 0, default=1)
    counts_1 = np_count(target, 1, default=1)

    sub_target = target[mask]

    sub_0 = np_count(sub_target, 0, default=1)
    sub_1 = np_count(sub_target, 1, default=1)

    y_prob = sub_1 / counts_1
    n_prob = sub_0 / counts_0

    return y_prob, n_prob


def get_iv_psi(df, feature_list=[], target='target', by_col='apply_mon', only_psi=True):
    """
    计算iv、psi
    Args:
        df (DataFrame): 原始数据集
        feature_list (list): 需要计算iv、psi的变量
        target (str): 目标变量名称
        by_col (str): 根据哪个变量分组计算iv、psi
        only_psi (bool): 是否只计算psi

    Returns:
        DataFrame: 变量的psi、iv
    """
    fb = autobmt.FeatureBin()
    fb.fit(df[feature_list], df[target], method='equal_freq', is_need_monotonic=False)
    # fb.fit(df[feature_list], df[target], method='dt')
    dev = fb.transform(df)
    by_col_v = sorted(list(set(df[by_col])))

    month_IV = pd.DataFrame()
    month_PSI_lis = []
    for n, j in enumerate(by_col_v):
        by_col_d = dev[dev[by_col] == j]
        if not only_psi:
            ###计算IV
            iv = {}
            iv_all = {}
            for i in feature_list:
                iv[i] = calc_iv(by_col_d[i], target=by_col_d[target], feature_bin=fb)
                if n == 0:
                    iv_all[i] = calc_iv(dev[i], target=dev[target], feature_bin=fb)
            if iv_all:
                month_IV = pd.concat([month_IV, pd.DataFrame([iv_all]).T.rename(columns={0: 'IV'})], axis=1)
            month_IV = pd.concat([month_IV, pd.DataFrame([iv]).T.rename(columns={0: f"{j}_IV"})], axis=1)
        ###计算PSI
        by_col_psi = psi(by_col_d[feature_list], dev[feature_list])
        by_col_psi.name = f"{j}_PSI"
        month_PSI_lis.append(by_col_psi)
    month_PSI = pd.DataFrame(month_PSI_lis).T
    month_PSI['MaxPSI'] = month_PSI.max(axis=1)

    ###iv趋势
    if not month_IV.empty:
        s_col_n = [f"{i}_IV" for i in by_col_v]
        for i in feature_list:
            month_IV.loc[i, 'IV趋势'] = get_inflection_point_index(month_IV.loc[i, s_col_n])[1]

    res = pd.concat([month_IV, month_PSI], axis=1)
    return res.sort_values(by='MaxPSI', ascending=False, )


from IPython.display import display
from pandas._libs import lib
from pandas.core.dtypes.generic import ABCSeries
from pandas.core.groupby import Grouper
from pandas.core.indexes.api import Index
import warnings

is_scalar = lib.is_scalar

warnings.filterwarnings('ignore')


def _convert_by(by):
    if by is None:
        by = []
    elif (
        is_scalar(by)
        or isinstance(by, (np.ndarray, Index, ABCSeries, Grouper))
        or hasattr(by, "__call__")
    ):
        by = [by]
    else:
        by = list(by)
    return by


# # 正负样本在不同数据集不同数据集_不同设备的匹配率===一体

def get_pivot_table(df, index_col_name=None, columns_name=None, target='target'):
    index_col_name = _convert_by(index_col_name)
    if target not in df:
        raise KeyError('数据中目标变量y值名称错误！！！')
    df['dfid'] = 1

    zb = pd.pivot_table(df, values='dfid', index=index_col_name, columns=[target], aggfunc='count',
                        margins=True,
                        margins_name='合计')
    zb['正样本占比'] = zb[1] / zb['合计']
    # sb_os = zb

    concat_lis = []
    if columns_name:
        for n, i in enumerate(columns_name):
            sb = pd.pivot_table(df, values='dfid', index=index_col_name, columns=[i], aggfunc='count',
                                margins=True,
                                margins_name='合计')
            sb.drop(columns='合计', inplace=True)
            concat_lis.append(sb)

    concat_lis.append(zb)
    sb_os = pd.concat(concat_lis, axis=1)
    return sb_os


def get_pivot_table_posnegasam(df, index_col_name=None, columns_name=None, target='target'):
    index_col_name = _convert_by(index_col_name)
    columns_name = _convert_by(columns_name)
    if target not in df:
        raise KeyError('数据中目标变量y值名称错误！！！')
    df['dfid'] = 1
    concat_lis = []
    for n, i in enumerate(columns_name):
        sb = pd.pivot_table(df, values='dfid', index=index_col_name,
                            columns=[i, target],
                            aggfunc='count', margins=True, margins_name='合计')

        if n != len(columns_name) - 1:
            sb.drop(columns='合计', inplace=True)

        concat_lis.append(sb)

        zb = pd.pivot_table(df, values=target, index=index_col_name, columns=[i],
                            aggfunc=['mean'],
                            margins=True, margins_name='合计')

        if n != len(columns_name) - 1:
            zb.drop(columns=('mean', '合计'), inplace=True)

        zb.columns = zb.columns.swaplevel(0, 1)
        zb.columns = pd.MultiIndex.from_frame(pd.DataFrame(zb.columns.tolist()).replace({'mean': '正样本占比'}))

        concat_lis.append(zb)

    sb_os = pd.concat(concat_lis, axis=1)
    sb_os.columns = pd.MultiIndex.from_frame(pd.DataFrame(list(sb_os.columns)).replace({'': '样本个数'}))
    sb_os.columns.names = ['', '']
    return sb_os


def get_mr(original_df, match_df, index_col_name=None, columns_name=None, target='target'):
    sb_os = get_pivot_table(match_df, index_col_name, columns_name, target)

    ##############分割线
    ys_sb_os = get_pivot_table(original_df, index_col_name, columns_name, target)

    ###匹配率
    pp_ratio = sb_os / ys_sb_os
    del pp_ratio['正样本占比']

    index = [tuple(['匹配样本数', i]) for i in list(sb_os.columns)]
    sb_os.columns = pd.MultiIndex.from_tuples(index)

    index = [tuple(['原始样本数', i]) for i in list(ys_sb_os.columns)]
    ys_sb_os.columns = pd.MultiIndex.from_tuples(index)

    index = [tuple(['匹配率', i]) for i in list(pp_ratio.columns)]
    pp_ratio.columns = pd.MultiIndex.from_tuples(index)
    # pp_ratio.columns = pp_ratio.columns.map(lambda x: str(x) + '_匹配率')

    tmp = pd.merge(ys_sb_os, sb_os, how='left', left_index=True, right_index=True, suffixes=('_原始', '_匹配'))

    res = pd.merge(pp_ratio, tmp, how='left', left_index=True, right_index=True)

    return res


def get_posnegasam_mr(original_df, match_df, index_col_name=None, columns_name=None, target='target'):
    if not columns_name:
        return pd.DataFrame()
        # raise KeyError('columns_name参数是空的！！！调用get_mr函数即可')

    sb_os = get_pivot_table_posnegasam(match_df, index_col_name, columns_name, target)

    #################分割线

    ys_sb_os = get_pivot_table_posnegasam(original_df, index_col_name, columns_name, target)

    ###匹配率
    pp_ratio = sb_os / ys_sb_os
    pp_ratio.drop('正样本占比', axis=1, level=1, inplace=True)

    index = [tuple(['匹配样本数'] + list(i)) for i in list(sb_os.columns)]
    sb_os.columns = pd.MultiIndex.from_tuples(index)

    index = [tuple(['原始样本数'] + list(i)) for i in list(ys_sb_os.columns)]
    ys_sb_os.columns = pd.MultiIndex.from_tuples(index)

    index = [tuple(['匹配率'] + list(i)) for i in list(pp_ratio.columns)]
    pp_ratio.columns = pd.MultiIndex.from_tuples(index)
    pp_ratio.columns = pd.MultiIndex.from_frame(pd.DataFrame(list(pp_ratio.columns)).replace({'样本个数': '总匹配率'}))
    pp_ratio.columns.names = ['', '', '']

    tmp = pd.merge(ys_sb_os, sb_os, how='left', left_index=True, right_index=True)
    res = pd.merge(pp_ratio, tmp, how='left', left_index=True, right_index=True)
    return res


def get_model_auc_ks(match_df, index_col_name=None, columns_name=None, target='target', pred='p'):
    groupby_list = _convert_by(index_col_name)
    columns_name = _convert_by(columns_name)
    if target not in match_df:
        raise KeyError('数据中目标变量y值名称错误！！！')
    if pred not in match_df:
        raise KeyError('数据中模型预测的概率值名称错误！！！')

    auc_lis = []
    ks_lis = []
    if columns_name:
        for i in columns_name:
            # auc
            try:
                type_na_device_type_auc = match_df[match_df[target].notnull()].groupby(
                    groupby_list + [i]).apply(
                    lambda tmp: pd.Series(
                        {'auc': get_auc(tmp[target], tmp[pred])}))
            except:
                type_na_device_type_auc = match_df[match_df[target].notnull()].groupby(
                    groupby_list + [i]).apply(
                    lambda tmp: pd.Series({'auc': 0}))
            auc_lis.append(type_na_device_type_auc.unstack())

            # ks
            try:
                type_na_device_type_ks = match_df[match_df[target].notnull()].groupby(
                    groupby_list + [i]).apply(
                    lambda tmp: pd.Series({'ks': get_ks(tmp[target], tmp[pred])}))
            except:
                type_na_device_type_ks = match_df[match_df[target].notnull()].groupby(
                    groupby_list + [i]).apply(
                    lambda tmp: pd.Series({'ks': 0}))

            ks_lis.append(type_na_device_type_ks.unstack())

    ###auc
    try:
        type_na_auc = match_df[match_df[target].notnull()].groupby(groupby_list).apply(
            lambda tmp: pd.Series({'auc': get_auc(tmp[target], tmp[pred])}))
        index = pd.MultiIndex.from_tuples([('auc', 'all')])
    except:
        type_na_auc = match_df[match_df[target].notnull()].groupby(groupby_list).apply(
            lambda tmp: pd.Series({'auc': 0}))
        index = pd.MultiIndex.from_tuples([('auc', 'all')])
    type_na_auc.columns = index

    auc_lis.append(type_na_auc)

    datatype_didtype_auc = pd.concat(auc_lis, axis=1)

    ###ks
    try:
        type_na_ks = match_df[match_df[target].notnull()].groupby(groupby_list).apply(
            lambda tmp: pd.Series({'ks': get_ks(tmp[target], tmp[pred])}))
        index = pd.MultiIndex.from_tuples([('ks', 'all')])
    except:
        type_na_ks = match_df[match_df[target].notnull()].groupby(groupby_list).apply(
            lambda tmp: pd.Series({'ks': 0}))
        index = pd.MultiIndex.from_tuples([('ks', 'all')])
    type_na_ks.columns = index

    ks_lis.append(type_na_ks)

    datatype_didtype_ks = pd.concat(ks_lis, axis=1)

    datatype_didtype_auc_ks = pd.concat([datatype_didtype_auc, datatype_didtype_ks], axis=1)
    return datatype_didtype_auc_ks


class StatisticsMrAucKs:

    def __init__(self, conf_dict={}, data_path=''):
        self.cust_id = conf_dict.get('cust_id', 'device_id')  # 主键
        self.target_na = conf_dict.get('target_na', 'target')  # 目标变量列名
        self.device_type_na = conf_dict.get('device_type_na', 'device_type')  # 设备列名
        self.year_month_na = conf_dict.get('year_month_na', 'apply_month')
        self.date_na = conf_dict.get('date_na', 'apply_time')  # 时间列名
        self.type_na = conf_dict.get('type_na', 'type')  # 数据集列名
        self.model_pred_res = conf_dict.get('model_pred_res', 'p')  # 模型预测值

        ###设备号列名
        self.oaid_col_na = conf_dict.get('oaid_col_na', 'oaid_md5')
        self.imei_col_na = conf_dict.get('imei_col_na', 'imei_md5')
        self.idfa_col_na = conf_dict.get('idfa_col_na', 'idfa_md5')

        ###设备号的具体取值
        self.oaid_value = conf_dict.get('oaid_value', 'oaid')
        self.imei_value = conf_dict.get('imei_value', 'imei')
        self.idfa_value = conf_dict.get('idfa_value', 'idfa')

        self.data_path = data_path

    def statistics_model_mr_auc_ks(self, by_cols=['device_type', 'os', 'media']):
        model_res = pd.read_csv(self.data_path)

        model_res[self.year_month_na] = model_res[self.date_na].map(lambda x: x[:7])

        model_res[self.device_type_na] = model_res[self.device_type_na].replace(
            {0: self.oaid_value, 1: self.imei_value, 2: self.idfa_value})

        p_null = model_res[model_res[self.model_pred_res].isnull()]

        if len(p_null[p_null[self.device_type_na].isnull()]) > 0:
            print("device_type有为空！！！")
            mdn_oaid = p_null[p_null[self.oaid_col_na].notnull()].rename(columns={self.oaid_col_na: 'device_id'})
            mdn_imei = p_null[p_null[self.imei_col_na].notnull()].rename(columns={self.imei_col_na: 'device_id'})
            mdn_idfa = p_null[p_null[self.idfa_col_na].notnull()].rename(columns={self.idfa_col_na: 'device_id'})

            mdn_oaid[self.device_type_na] = 0  # 0,oaid
            mdn_imei[self.device_type_na] = 1  # 1,imei
            mdn_idfa[self.device_type_na] = 2  # 2,idfa
            p_null_did = mdn_oaid.append(mdn_imei).append(mdn_idfa)

            p_null_did = p_null_did.sort_values([self.device_type_na]).drop_duplicates([self.cust_id],
                                                                                       keep='first')  # first是取0(oaid)、1(imei)、2(idfa)
            del p_null_did['device_id']

            p_null_did[self.device_type_na] = p_null_did[self.device_type_na].replace(
                {0: self.oaid_value, 1: self.imei_value, 2: self.idfa_value})

            model_res = model_res[model_res[self.model_pred_res].notnull()].append(p_null_did)

        model_res['os'] = model_res[self.device_type_na].replace(
            {self.oaid_value: 'android', self.imei_value: 'android', self.idfa_value: 'ios'})

        print('总：', model_res.shape)
        print('没有p值的：', p_null.shape)
        p_notnull = model_res[model_res[self.model_pred_res].notnull()]
        print('有p值的：', p_notnull.shape)

        file_name = self.data_path.split('.csv')[0] + '_模型mr_auc_ks_res——new.xlsx'

        writer = pd.ExcelWriter(file_name)

        ###不同数据集_不同设备的匹配率

        res = get_mr(model_res, p_notnull, self.type_na, by_cols)
        print('1.1、不同集匹配率')
        display(res)

        start_row = 1
        res.to_excel(writer, sheet_name='1、匹配详情', startrow=start_row)  # 指定从哪一行开始写入数据，默认值为0
        # 修改下一次开始写入数据的行位置
        start_row = start_row + res.shape[0] + 4

        ###不同月份_不同设备的匹配率

        res = get_mr(model_res, p_notnull, self.year_month_na, by_cols)
        print('1.2、不同月份匹配率')
        display(res)

        res.to_excel(writer, sheet_name='1、匹配详情', startrow=start_row)
        start_row = start_row + res.shape[0] + 5

        ###正负样本不同数据集_不同设备的匹配率

        res = get_posnegasam_mr(model_res, p_notnull, self.type_na, by_cols)
        print('1.3、不同集、不同设备匹配率')
        display(res)

        res.to_excel(writer, sheet_name='1、匹配详情', startrow=start_row)
        start_row = start_row + res.shape[0] + 5

        res = get_posnegasam_mr(model_res, p_notnull, self.year_month_na, by_cols)
        print('1.4、不同月份、不同设备匹配率')
        display(res)

        res.to_excel(writer, sheet_name='1、匹配详情', startrow=start_row)

        ### 效果统计

        ### 不同集

        res = get_model_auc_ks(p_notnull, self.type_na, by_cols)
        print('2.1、不同集效果')
        display(res)

        start_row = 1
        res.to_excel(writer, sheet_name='2、模型效果详情', startrow=start_row)
        start_row = start_row + res.shape[0] + 4

        ### 不同月份

        res = get_model_auc_ks(p_notnull, self.year_month_na, by_cols)
        print('2.2、不同月份效果')
        display(res)

        res.to_excel(writer, sheet_name='2、模型效果详情', startrow=start_row)
        start_row = start_row + res.shape[0] + 4

        ### 不同集上不同月份

        res = get_model_auc_ks(p_notnull, [self.type_na, self.year_month_na], by_cols)
        print('2.3、不同集、月份效果')
        display(res)

        res.to_excel(writer, sheet_name='2、模型效果详情', startrow=start_row)

        # # 不同集上不同月份

        # # 效果统计结束

        writer.save()
        writer.close()



#==============================================================================
# File: stepwise.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: stepwise.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-03-06
'''
import warnings

import numpy as np
import pandas as pd
import statsmodels.api as sm
from scipy import stats
from sklearn.linear_model import LogisticRegression

from .logger_utils import Logger
from .metrics import get_auc, get_ks, AIC, BIC, MSE
from .utils import split_target, unpack_tuple, step_evaluate_models, \
    get_max_corr_feature, model_predict_evaluate

INTERCEPT_COLS = 'intercept'

warnings.filterwarnings(action='ignore')
log = Logger(level="info", name=__name__).logger


class StatsModel:
    def __init__(self, estimator='ols', criterion='aic', intercept=False):
        if isinstance(estimator, str):
            Est = self.get_estimator(estimator)
            estimator = Est(fit_intercept=intercept, )

        self.estimator = estimator
        self.intercept = intercept
        self.criterion = criterion

    def get_estimator(self, name):
        from sklearn.linear_model import (
            LinearRegression,
            LogisticRegression,
            Lasso,
            Ridge,
        )

        ests = {
            'ols': LinearRegression,
            'lr': LogisticRegression,
            'lasso': Lasso,
            'ridge': Ridge,
        }

        if name in ests:
            return ests[name]

        raise Exception('estimator {name} is not supported'.format(name=name))

    def stats(self, X, y):
        """
        """
        X = X.copy()

        if isinstance(X, pd.Series):
            X = X.to_frame()

        self.estimator.fit(X, y)

        if hasattr(self.estimator, 'predict_proba'):
            pre = self.estimator.predict_proba(X)[:, 1]
        else:
            pre = self.estimator.predict(X)

        coef = self.estimator.coef_.reshape(-1)

        if self.intercept:
            coef = np.append(coef, self.estimator.intercept_)
            X[INTERCEPT_COLS] = np.ones(X.shape[0])

        n, k = X.shape

        t_value = self.t_value(pre, y, X, coef)
        p_value = self.p_value(t_value, n)
        c = self.get_criterion(pre, y, k)

        return {
            't_value': pd.Series(t_value, index=X.columns),
            'p_value': pd.Series(p_value, index=X.columns),
            'criterion': c
        }

    def get_criterion(self, pre, y, k):
        if self.criterion == 'aic':
            llf = self.loglikelihood(pre, y, k)
            return AIC(pre, y, k, llf=llf)

        if self.criterion == 'bic':
            llf = self.loglikelihood(pre, y, k)
            return BIC(pre, y, k, llf=llf)

        if self.criterion == 'ks':
            return get_ks(y, pre)

        if self.criterion == 'auc':
            return get_auc(y, pre)

    def t_value(self, pre, y, X, coef):
        n, k = X.shape
        mse = sum((y - pre) ** 2) / float(n - k)
        nx = np.dot(X.T, X)

        if np.linalg.det(nx) == 0:
            return np.nan

        std_e = np.sqrt(mse * (np.linalg.inv(nx).diagonal()))
        return coef / std_e

    def p_value(self, t, n):
        return stats.t.sf(np.abs(t), n - 1) * 2

    def loglikelihood(self, pre, y, k):
        n = len(y)
        mse = MSE(pre, y)
        return (-n / 2) * np.log(2 * np.pi * mse * np.e)


def stepwise(frame, target='target', estimator='ols', direction='both', criterion='aic',
             p_enter=0.01, p_remove=0.01, p_value_enter=0.2, intercept=False,
             max_iter=None, return_drop=False, exclude=None):
    """
    逐步回归选择特征
    Args:
        frame (DataFrame): 用于训练模型的数据集
        target (str): 目标变量名称
        estimator (str): 用于统计的模型
        direction (str): 前向逐步还是后向逐步, 支持“forward”、“backward”和“both”，建议“both”
        criterion (str): 统计模型的信息准则, 支持“aic”、“bic”
        p_enter (float): 阈值将在“forward”和“both”中使用，用于保留特征
        p_remove (float): 阈值将在“backward”中使用，用于剔除特征
        intercept (bool): 是否需要截距项
        p_value_enter (float): 阈值将在“both”中使用，用于剔除特征
        max_iter (int): 最大迭代次数
        return_drop (bool): 是否需要返回删除的特征
        exclude (array-like): 不参与特征筛选的特征列表

    Returns:
        DataFrame: 筛选后的数据集
        array: 删除的特征列表
    """
    df, y = split_target(frame, target)

    if exclude is not None:
        df = df.drop(columns=list(set(exclude) & set(df.columns)))

    drop_list = []
    remaining = df.columns.tolist()

    selected = []

    sm = StatsModel(estimator=estimator, criterion=criterion, intercept=intercept)

    order = -1 if criterion in ['aic', 'bic'] else 1

    best_score = -np.inf * order

    iter = -1
    while remaining:
        iter += 1
        if max_iter and iter > max_iter:
            break

        l = len(remaining)
        test_score = np.zeros(l)
        test_res = np.empty(l, dtype=object)

        if direction == 'backward':
            for i in range(l):
                test_res[i] = sm.stats(
                    df[remaining[:i] + remaining[i + 1:]],
                    y,
                )
                test_score[i] = test_res[i]['criterion']

            curr_ix = np.argmax(test_score * order)
            curr_score = test_score[curr_ix]

            if (curr_score - best_score) * order < p_remove:
                break

            name = remaining.pop(curr_ix)
            drop_list.append(name)

            best_score = curr_score

        # forward and both
        else:
            for i in range(l):
                test_res[i] = sm.stats(
                    df[selected + [remaining[i]]],
                    y,
                )
                test_score[i] = test_res[i]['criterion']

            curr_ix = np.argmax(test_score * order)
            curr_score = test_score[curr_ix]

            name = remaining.pop(curr_ix)
            if (curr_score - best_score) * order < p_enter:
                drop_list.append(name)

                # early stop
                if selected:
                    drop_list += remaining
                    break

                continue

            selected.append(name)
            best_score = curr_score

            if direction == 'both':
                p_values = test_res[curr_ix]['p_value']
                drop_names = p_values[p_values > p_value_enter].index

                for name in drop_names:
                    selected.remove(name)
                    drop_list.append(name)

    r = frame.drop(columns=drop_list)

    res = (r,)
    if return_drop:
        res += (drop_list,)

    return unpack_tuple(res)


class StepWise:
    def __init__(self, df, target='target', features=[], exclude_columns=None, iv_psi_df=None, var_bin_woe={},
                 match_dict={},
                 object='test_ks', features_corr_df=None,
                 include_columns=[], p_threshold=0.05, vif_threshold=5, max_varcnt=10, A=404.65547022, B=72.1347520444,
                 is_return_var=False):
        self.df = df
        self.target = target
        self.features = features
        self.exclude_columns = exclude_columns
        self.iv_psi_df = iv_psi_df
        self.var_bin_woe = var_bin_woe
        self.match_dict = match_dict  # 特征的中文字典名称
        self.object = object
        self.features_corr_df = features_corr_df  # 变量相关性替换列表
        self.include_columns = include_columns
        self.p_threshold = p_threshold
        self.vif_threshold = vif_threshold
        self.A = A
        self.B = B
        self.max_varcnt = len(self.features) if max_varcnt > len(self.features) else max_varcnt
        self.step_evaluate_log_df = []
        self.predict_data = None
        self.is_return_var = is_return_var

    @property
    def get_evaluate_df_log(self):
        """合并每一步的评估结果"""
        if len(self.step_evaluate_log_df) == 0:
            log.info("并未进行评估过!!!")
            return None
        else:
            evaluate_log_df = pd.concat(self.step_evaluate_log_df, axis=0).reset_index(drop=True)
            return evaluate_log_df

    @property
    def get_output_data(self):
        """获取预测的结果"""
        if self.predict_data is not None and isinstance(self.predict_data, pd.DataFrame):
            if self.is_return_var:
                for k, v in self.model.params.to_dict().items():
                    if k == 'const':
                        self.predict_data['const'] = round(self.A - self.B * v)
                    else:
                        self.predict_data["{}_fscore".format(k)] = self.predict_data[k].map(
                            lambda x: round(-(self.B * v * x)))

                # dump_to_pkl(score_card_structure,'./score_card_structure.pkl')
                in_model_features = [i for i in self.predict_data.columns.tolist() if '_fscore' in i] + ['const']
                self.predict_data['woe2score'] = self.predict_data[in_model_features].sum(axis=1).map(
                    lambda x: round(x))
                return self.predict_data
            else:
                return self.predict_data
        else:
            ValueError("并没有进行预测，请先执行stepwise_apply()!")

    def run(self):
        """
        新版双向stepwise方法
        Returns:

        """
        # 执行了最优分箱后auc,ks
        step_name = "best_binning"
        step_evaluate_df, evaluate_log = step_evaluate_models(self.df, self.features, self.target, stepname=step_name)
        self.step_evaluate_log_df.append(step_evaluate_df)
        log.info("{},evaluate --- {}".format(step_name, evaluate_log))

        log.info("开始进行Stepwise,变量个数为:{}".format(len(self.features)))
        # 前置检查
        self.check()
        log.info("Stepwise前置检查完成")
        record_features_log = []  # 记录整个stepwise的变量

        initial_list = self.include_columns.copy()  # 保留的特征
        train_data = self.df[self.df['type'] == 'train']
        test_data = self.df[self.df['type'] == 'test']
        if self.iv_psi_df is not None and isinstance(self.iv_psi_df, pd.DataFrame):
            self.features = self.iv_psi_df.sort_values(by='IV', ascending=False).index.to_list()
        high_iv_name = self.features[0]
        initial_list.append(high_iv_name)
        self.features.remove(high_iv_name)
        flag_features = initial_list.copy()  # 记录特征
        # 最大iv的目标值，作为基准
        base_object_value, _ = get_object_value_by_train_lr(train_data[initial_list], train_data[self.target],
                                                            test_data[initial_list], test_data[self.target])
        record_features_log.append(("round-0", ",".join(initial_list), base_object_value, 0))
        log.info("最高iv变量{}的基础目标值为{}".format(high_iv_name, base_object_value))
        step = 1
        while True:
            result_dict = {}
            # 前向选择过程
            # 遍历判断添加一个变量后，目标值是否有提升
            for name in self.features:
                if name not in flag_features:
                    object_value, _ = get_object_value_by_train_lr(train_data[initial_list + [name]],
                                                                   train_data[self.target],
                                                                   test_data[initial_list + [name]],
                                                                   test_data[self.target])
                    result_dict[name] = object_value
            max_key, max_value = get_max_value_in_dic(result_dict)  # 返回最优的特征和object值
            # 若有提升，则加入到候选变量中
            round_history_arr = [("rount-{}".format(step), ",".join(initial_list + [k]), v, v - base_object_value) for
                                 k, v in
                                 result_dict.items()]
            record_features_log.extend(round_history_arr)

            if max_value > base_object_value:
                initial_list.append(max_key)  # 加入到候选变量
                flag_features.append(max_key)  # 记录该特征已经选择过
                base_object_value = max_value  # 更新当前最优object值
                # 后向判断过程
                df_pvalue_vif, initial_list = calculate_features_p_value_vif(train_data, initial_list, self.target)
                filter_df = df_pvalue_vif[(df_pvalue_vif['pvalue'] > self.p_threshold) | (
                    df_pvalue_vif['vif'] > self.vif_threshold)]
                if filter_df.shape[0] > 0:
                    drop_col = filter_df.index.tolist()
                    # 删除后向剔除的特征
                    # TODO
                    initial_list = [i for i in initial_list if i not in drop_col]
                    drop_log = ("rount-{}-drop".format(step), ",".join(drop_col), 0, 0)  # 被剔除变量
                    record_features_log.append(drop_log)
                tmp = ("rount-{}-best".format(step), ",".join(initial_list), base_object_value, 0)  # 这一轮的最优变量
                record_features_log.append(tmp)

            step += 1
            # 若满足下列条件，则跳出循环
            if len(initial_list) >= self.max_varcnt or max_value - base_object_value < 0:
                break
        log.info("Stepwise目标选择完成,剩余变量个数为:{}".format(len(initial_list)))
        # 构造stepwise的过程输出和模型结果
        result_df, evaluate_df = self.output(train_data, initial_list, self.target)
        log_df = pd.DataFrame(record_features_log, columns=['round-n', 'features', 'object_value', 'diff'])
        step_name = "toad_stepwise"
        step_evaluate_df, evaluate_log = step_evaluate_models(self.df, initial_list, self.target, stepname=step_name)
        self.step_evaluate_log_df.append(step_evaluate_df)
        log.info("{},evaluate --- {}".format(step_name, evaluate_log))
        return initial_list, result_df, log_df, evaluate_df, self.get_evaluate_df_log

    def stepwise_apply(self):
        """
        基于pvalue、aic、bid的Stepwise方法
        Returns:

        """
        # 执行了最优分箱后auc,ks
        step_name = "best_binning"
        step_evaluate_df, evaluate_log = step_evaluate_models(self.df, self.features, self.target, stepname=step_name)
        self.step_evaluate_log_df.append(step_evaluate_df)
        log.info("{},evaluate --- {}".format(step_name, evaluate_log))

        train_data = self.df[self.df['type'] == 'train']
        if self.features_corr_df is not None and isinstance(self.features_corr_df, pd.DataFrame):
            max_corr_df = self.features_corr_df  # 如果相关性df传入了，则用传入的
        else:  # 否则用woe转换后的
            max_corr_df = get_max_corr_feature(train_data, self.features)  # 计算每个变量相关性最高的变量
        tmp_df = train_data[self.features + [self.target]]
        final_data = stepwise(tmp_df, target=self.target, estimator='ols', direction='both', criterion='aic',
                              p_enter=0.05, p_remove=0.05, )

        selected_features = [name for name in final_data.columns.to_list() if name not in self.exclude_columns]
        log.info("Stepwise目标选择完成,剩余变量个数为:{}".format(len(selected_features)))
        log.info("Stepwise保留变量为:{}".format(",".join(selected_features)))
        result_df, evaluate_df = self.output(train_data, selected_features, self.target)  # 训练模型
        step_name = "toad_stepwise"
        step_evaluate_df, evaluate_log = step_evaluate_models(self.df, selected_features, self.target,
                                                              stepname=step_name)
        self.step_evaluate_log_df.append(step_evaluate_df)
        log.info("{},evaluate --- {}".format(step_name, evaluate_log))

        # 关联上相关性最强的特征,方便替换
        result_df = result_df.merge(max_corr_df, how='left', left_index=True, right_index=True)
        if len(self.match_dict) > 0:
            result_df.insert(8, 'corr_name_cn', result_df['corr_name'].map(self.match_dict))
        result_df = result_df.reset_index()

        ##TODO 考虑在result_df进行psi、vif、相关性的过滤。result_df是10.model_stepwise

        return selected_features, result_df, evaluate_df, self.get_evaluate_df_log

    def output(self, train_data, features, target):
        '''

        Args:
            train_data:
            features:
            target:

        Returns:

        '''
        if features is None:
            raise ValueError("没有入模变量")

        # 训练模型,计算pvalue,vif
        df_pvalue_vif_coef, model, features = calculate_features_p_value_vif(train_data, features, target,
                                                                             need_coef=True,
                                                                             need_model=True)
        self.model = model
        if len(self.match_dict) > 0:
            df_pvalue_vif_coef.insert(0, 'cn', df_pvalue_vif_coef.index.map(self.match_dict))
        # 评估模型
        evaluate_df, self.predict_data = model_predict_evaluate(model, self.df, features, self.target, self.A, self.B,
                                                                self.exclude_columns, self.is_return_var)
        if self.iv_psi_df is not None:
            df_iv_psi = self.iv_psi_df[self.iv_psi_df.index.isin(features)]
            result_df = df_pvalue_vif_coef.merge(df_iv_psi, how='left', left_index=True, right_index=True)
            return result_df, evaluate_df
        else:
            return df_pvalue_vif_coef, evaluate_df

    def get_scorecard_structure(self):
        score_card_structure = {}
        for k, v in self.model.params.to_dict().items():
            # scorecard
            if k == 'const':
                score_card_structure['const'] = ['-', round(self.A - self.B * v)]
            else:
                one_feature_binning = self.var_bin_woe[k]
                one_feature_dict = {}
                for k1, v1 in one_feature_binning.items():
                    one_feature_dict[k1] = [float(v1), round(-(self.B * v * float(v1)))]  # v1是woe值
                score_card_structure[k] = one_feature_dict
        return score_card_structure

    def check(self):
        """
        特征选择模块，前置检查,符合要求，则往下运行
        Returns:

        """
        if len(self.features) <= 0 or self.features is None:
            raise ValueError("入模型变量不能为空，请检查!!!")
        if self.df is None or not isinstance(self.df, pd.DataFrame):
            raise ValueError("数据集不能为空并且数据集必须是dataframe!!!")
        if self.target is None:
            raise ValueError("数据集的目标变量名称不能为空!!!")
        if self.exclude_columns is None or "type" not in self.exclude_columns or self.target not in self.exclude_columns:
            raise ValueError("exclude 不能为空，必须包含target,type字段!!!")

    @staticmethod
    def scorecard_to_excel(scorecard, workbook, sheet_name):
        # excel相关格式
        # workbook = xlsxwriter.Workbook(output, {'nan_inf_to_errors': True})
        # worksheet = workbook.add_worksheet('分箱详情')
        worksheet = workbook.add_worksheet(sheet_name)
        # Add a header format.
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'center',
            'font_name': 'Consolas',
            'font_size': 11,
            'border': 1,
            'bottom': 2,
        })
        font_formats = workbook.add_format({'font_name': 'Consolas', 'font_size': 11, })  # 其他

        def write_to_excel_by_dict(scorecard, worksheet, start_row, start_column):
            for featurename, interval_value in scorecard.items():
                if isinstance(interval_value, dict):
                    worksheet.write(start_row, start_column, featurename, font_formats)
                    for interval, value in interval_value.items():
                        woe, score = value[0], value[1]
                        worksheet.write(start_row, start_column + 1, interval, font_formats)
                        worksheet.write(start_row, start_column + 2, woe, font_formats)
                        worksheet.write(start_row, start_column + 3, score, font_formats)
                        start_row += 1
                elif isinstance(interval_value, list):
                    nothing, score = interval_value[0], interval_value[1]
                    worksheet.write(start_row, start_column, featurename, font_formats)
                    worksheet.write(start_row, start_column + 1, nothing, font_formats)
                    worksheet.write(start_row, start_column + 2, nothing, font_formats)
                    worksheet.write(start_row, start_column + 3, score, font_formats)
                    start_row += 1

        # 写入标题
        for col_num, value in enumerate(['特征名称', '特征区间', '特征区间WOE值', '特征区间得分']):
            worksheet.write(0, col_num, value, header_format)

        # 写入数据
        start_row, start_column = 1, 0
        write_to_excel_by_dict(scorecard, worksheet, start_row, start_column)
        worksheet.set_column(0, 1, 20)


def get_object_value_by_train_lr(X_train, y_train, X_test, y_test, object_func='default'):
    """
    计算目标值
    Args:
        X_train:
        y_train:
        X_test:
        y_test:
        object_func:

    Returns:

    """

    lr = LogisticRegression().fit(X_train, y_train)
    y_test_pred = lr.predict_proba(X_test)[:, 1]
    test_ks = get_ks(y_test, y_test_pred)
    test_auc = get_auc(y_test, y_test_pred)

    if object_func == "test_ks":
        return test_ks
    elif object_func == "test_auc":
        return test_auc
    elif object_func == "default":
        return test_ks, test_auc


def get_max_value_in_dic(dic):
    """
    返回字典中最大value的k,v
    Args:
        dic:

    Returns:

    """
    import operator
    key = max(dic.items(), key=operator.itemgetter(1))[0]
    return key, dic[key]


def calculate_features_p_value_vif(df, features, target, just_pvalue=False, need_coef=False, need_model=False):
    """
    计算每个变量的p-value和vif和coef
    Args:
        df:
        features:
        target:
        just_pvalue:
        need_coef:
        need_model:

    Returns:

    """
    import autobmt

    index = 1
    while True:  # 循环的目的是保证入模变量的系数都为整
        model = sm.Logit(df[target], sm.add_constant(df[features])).fit()
        ignore_features = [k for k, v in model.params.to_dict().items() if k != 'const' and v < 0]
        if len(ignore_features) == 0:
            break
        features = [i for i in features if i not in ignore_features]
        index += 1
    df_pvalue = pd.DataFrame(model.pvalues, columns=['pvalue'])
    if just_pvalue:
        return df_pvalue
    df_coef = pd.DataFrame(model.params, columns=['coef'])
    df_vif = pd.Series(index=features)
    for col in features:
        df_vif[col] = pd.Series(autobmt.get_vif(df[list(set(features) - set([col]))], df[col]))
    df_vif = pd.DataFrame(df_vif, columns=['vif'])
    if need_coef:
        df_pvalue_vif_coef = df_coef.merge(df_pvalue, how='left', left_index=True, right_index=True) \
            .merge(df_vif, how='left', left_index=True, right_index=True)
        if need_model:
            return df_pvalue_vif_coef, model, features
        else:
            return df_pvalue_vif_coef, features
    else:
        df_pvalue_vif = df_pvalue.merge(df_vif, how='inner', left_index=True, right_index=True)
        return df_pvalue_vif, features



#==============================================================================
# File: text_utils.py
#==============================================================================

# -*- coding: utf-8 -*-

import os
import random
import jieba
import pandas as pd

# 读取停用词
stopwords = pd.read_csv("data/text_data/stopwords.txt", index_col=False, quoting=3, sep="\t", names=['stopword'],
                        encoding='utf-8')
stopwords = stopwords['stopword'].values


def cut_words(line, words_min=2):
    line_segments = jieba.lcut(line)
    line_segments = filter(lambda x: len(x) >= words_min, line_segments)
    line_segments = filter(lambda x: x not in stopwords, line_segments)
    return list(line_segments)


def load_corpus():
    """
    加载语料库：取自搜狗新闻语料库(https://www.sogou.com/labs/resource/cs.php)
    :return: sentences 语料库
    """
    # 取样后的文本存储
    df_entertainment = pd.read_csv(os.path.join('data/text_data/entertainment_news.csv'))
    df_sports = pd.read_csv(os.path.join('data/text_data/sports_news.csv'))

    entertainment = df_entertainment.content.values.tolist()
    sports = df_sports.content.values.tolist()
    content_file = {'entertainment': entertainment, 'sports': sports}

    return content_file


def sentences_prepare():
    """
    语料库预处理（无标签）
    """
    sentences = []
    content_file = load_corpus()
    for category in content_file.keys():
        for line in content_file[category]:
            try:
                words_list = cut_words(line)
                sentences.append(" ".join(words_list))
            except Exception as e:
                sentences.append("")
                print(e)
                continue
    random.seed(1)
    random.shuffle(sentences)
    return sentences


def sentences_prepare_with_y():
    """
    语料库预处理（含标签）
    """
    sentences = []
    content_file = load_corpus()
    for category in content_file.keys():
        for line in content_file[category]:
            try:
                words_list = cut_words(line)
                sentences.append("__label__" + str(category) + " , " + " ".join(words_list))
            except Exception as e:
                sentences.append("")
                print(line)
                continue
    random.seed(1)
    random.shuffle(sentences)
    return sentences


def sentences_prepare_x_y():
    """
    语料库预处理（语料和标签分别输出）
    """
    cate_dic = {'entertainment': 0, 'sports': 1}
    content_file = load_corpus()
    # 生成训练数据
    sentences = []
    y = []

    for category in content_file.keys():
        # 文本预处理
        for line in content_file[category]:
            try:
                words_list = cut_words(line)
                sentences.append(" ".join(words_list))
                y.append(str(cate_dic.get(category)))
            except Exception as e:
                print(line)
                continue
    sentences_df = pd.DataFrame({'sentences': sentences, 'target': y})
    sentences_df = sentences_df.sample(frac=1, random_state=1)
    return sentences_df.sentences.tolist(), sentences_df.target.tolist()



#==============================================================================
# File: time_utils.py
#==============================================================================

# -*- coding: utf-8 -*-

import time
import pytz
import numpy as np
import datetime as dt
from dateutil.parser import parse


def stamp_to_date(time_stamp, timezone=None):
    """
    时间戳转日期函数
    :param time_stamp:int，时间戳
    :param timezone:string，时区
    :return: datetime
    """
    try:
        if timezone is None:
            stamp_str = str(time_stamp)
            if len(stamp_str) >= 10:
                stamp_str = stamp_str[:10]
            else:
                stamp_str = stamp_str
            time_stamp = int(stamp_str)
            date = dt.datetime.fromtimestamp(time_stamp)
            return date
        else:
            stamp_str = str(time_stamp)
            if len(stamp_str) >= 10:
                stamp_str = stamp_str[:10]
            else:
                stamp_str = stamp_str
            time_stamp = int(stamp_str)
            tz = pytz.timezone(timezone)
            date = dt.datetime.fromtimestamp(time_stamp, tz).strftime('%Y-%m-%d %H:%M:%S')
            date = parse(date)
            return date
    except:
        return parse('2100-01-01')


def date_to_stamp(date_time):
    """
    将日期转换为时间戳
    :param date_time: string，datetime
    :return: int
    """
    try:
        if isinstance(date_time, str):
            date_time = parse(date_time)
        return int(time.mktime(date_time.timetuple()))
    except:
        return int(631123200)


def date_to_week(date):
    '''
    日期转换为星期
    :param date:datetime，string
    :return: int
    '''
    try:
        if isinstance(date, str):
            date = parse(date)
        if_weekend = date.weekday()
        return if_weekend
    except:
        return np.nan



#==============================================================================
# File: transformer.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: transformer.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-03-31
'''

import copy
import math
from functools import wraps

import numpy as np
import pandas as pd
from joblib import Parallel, delayed
from sklearn.base import TransformerMixin

from .statistics import probability, WOE
from .utils import split_points_to_bin, FILLNA, save_json, split_target

DEFAULT_NAME = 'default_name'
EMPTY_BIN = -1
ELSE_GROUP = 'else'


def df_exclude_cols(func):
    @wraps(func)
    def exclude_cols(self, X, y, **kwargs):
        exclude = kwargs.get('exclude', None)
        if exclude is not None and isinstance(X, pd.DataFrame):
            X = X.drop(columns=exclude)
            del kwargs['exclude']

        return func(self, X, y, **kwargs)

    return exclude_cols


def df_select_dtypes(func):
    @wraps(func)
    def select_dtypes(self, X, y, **kwargs):
        select_dtypes = kwargs.get('select_dtypes', None)
        if select_dtypes is not None and isinstance(X, pd.DataFrame):
            X = X.select_dtypes(include=select_dtypes)
            del kwargs['select_dtypes']

        return func(self, X, y, **kwargs)

    return select_dtypes


def _check_duplicated_keys(X):
    if isinstance(X, pd.DataFrame) and X.columns.has_duplicates:
        keys = X.columns[X.columns.duplicated()].values
        raise Exception("X has duplicate keys `{keys}`".format(keys=str(keys)))

    return True


class FeatureBin(TransformerMixin):
    """
    分箱类
    """

    def __init__(self, n_jobs=-1):
        self.splits_dict = dict()
        self.n_jobs = n_jobs

    def __len__(self):
        return len(self.splits_dict.keys())

    def __contains__(self, key):
        return key in self.splits_dict

    def __getitem__(self, key):
        return self.splits_dict[key]

    def __setitem__(self, key, value):
        self.splits_dict[key] = value

    def __iter__(self):
        return iter(self.splits_dict)

    @df_exclude_cols
    @df_select_dtypes
    def fit(self, X, y, update = False, **kwargs):
        """
        分箱
        Args:
            X (DataFrame|array-like): 要分箱的X
            y (str|array-like): 目标变量
            min_sample_rate (number) : 每个箱子的最小占比，默认0.05
            n_bins (int): 需要分成几箱，默认10
            is_need_monotonic (bool) : 是否强制单调，默认True，强制单调
            is_empty_bin (bool) : 是否将空箱单独归为一个箱子，默认True，空箱单独归1箱
            min_threshold (number): 最小的卡方阀值
            exclude (str|array-like): 排除的特征，该特征将不参与分箱
            select_dtypes (str|numpy.dtypes): `'object'`, `'number'` 等. 只有选定的数据类型才会被分箱

        """

        # assert y.isin([0, 1]).all(), 'ERROR: :-) :-) 目标变量不是0/1值，请检查！！！'

        if not isinstance(X, pd.DataFrame):
            fea_name, splits = self._fit(X, y, **kwargs)
            self.splits_dict[fea_name] = splits
            return self

        if isinstance(y, str):
            # y = X.pop(y)
            X, y = split_target(X, y)

        _check_duplicated_keys(X)

        data = Parallel(n_jobs=self.n_jobs)(
            delayed(self._fit)(X[col], y, **kwargs) for col in X)  # 批量处理

        if update:
            self.splits_dict.update(dict(data))
        else:
            self.splits_dict = dict(data)

        return self

    def _fit(self, X, y, method='chi', is_empty_bin=True, **kwargs):  # method='dt'修改为method='chi'
        """
        分箱
        Args:
            X (DataFrame|array-like): 要分箱的X
            y (str|array-like): 目标变量
            min_sample_rate (number) : 每个箱子的最小占比
            n_bins (int): 需要分成几箱
            is_need_monotonic (bool) : 是否强制单调
            is_empty_bin (bool) : 是否将空箱单独归为一个箱子
            min_threshold (number): 最小的卡方阀值

        Returns:
            str : 分箱变量名
            array : 分割点

        """

        from .feature_binning import bin_method_run

        fea_name = DEFAULT_NAME
        if hasattr(X, 'name'):
            fea_name = X.name

        # 判断是否连续型变量，如果不是，将其原始值替换为0、1、2、3这样有序的连续性变量，序是按原始值所对应的woe值大小来给予的
        unique_X_val = None
        if not np.issubdtype(X.dtype, np.number):
            transer = WoeTransformer()
            # if X.dtype.type is np.object_:
            #     X = X.astype(np.str)
            empty_mask = pd.isna(X).any()
            if empty_mask:
                X = X.astype(np.str)
            woe = transer.fit_transform(X, y)
            # 获取变量的唯一值，及其所在该变量中的索引；unique_X_val=['A' 'B' 'C' 'D' 'E' 'F' 'G'] unique_X_index=[25  0  2  9  1  6 18]
            unique_X_val, unique_X_index = np.unique(X, return_index=True)
            # 通过原始值所在的索引将其原始对应的woe值取出。unique_woe=[-0.10178269 -0.44183275  0.22730944  0.15707894  0.50662326 -0.27946387 -0.10178269]
            unique_woe = woe[unique_X_index]
            # argsort函数是将woe值从小到大排序，然后将排序后的woe值所对应的原始woe值所在的索引输出；woe_argsort_index=[1 5 0 6 3 2 4]
            woe_argsort_index = np.argsort(unique_woe)
            # 变量唯一值按woe值从小到大的顺序调整变量唯一值的位置，变成有顺序的；unique_X_val=['B' 'F' 'A' 'G' 'D' 'C' 'E']
            unique_X_val = unique_X_val[woe_argsort_index]
            # 将原始X根据unique_X_val=['B' 'F' 'A' 'G' 'D' 'C' 'E']的有序顺序替换为0、1、2、3
            # unique_X_val=['B' 'F' 'G' 'nan' 'D' 'C' 'E']
            if empty_mask and is_empty_bin:
                unique_X_val = unique_X_val[np.isin(unique_X_val, 'nan', invert=True)]
            X = self._raw_category_x_to_bin(X, unique_X_val, is_empty_bin=is_empty_bin)

        splits = bin_method_run(X, y, method, is_empty_bin=is_empty_bin, **kwargs)

        # 如果不是连续型变量，X原始值被0、1、2、3替换了，自然出来的splits也是数值，需要将splits中的数值从unique_X_val=['B' 'F' 'A' 'G' 'D' 'C' 'E']进行还原
        splits = self._restore_category_splits(splits, unique_X_val)

        return fea_name, splits

    def transform(self, X, bins_dict={}, **kwargs):
        """
        原始数据根据分割点变换原始X
        Args:
            X (DataFrame|array-like): 需要转换的原始X
            bins_dict: 分箱字典, 形如: {'D157': [-999, 1.0, 2.0, 3.0, 5.0, inf]}
            **kwargs:

        Returns:

        """

        if not isinstance(bins_dict, dict):
            assert '请传入类似 {\'D157\': [-999, 1.0, 2.0, 3.0, 5.0, inf]}'

        if not bins_dict:
            bins_dict = self.splits_dict
        else:
            bins_dict = {k: np.array(v) for k, v in bins_dict.items()}

        if getattr(X, 'ndim', 1) == 1:

            if hasattr(X, 'name'):  # pd.Series
                if X.name in bins_dict:
                    fea_name, bins = self._transform(X, bins_dict.get(X.name), **kwargs)
                    return bins
                else:
                    return X

            if len(bins_dict) == 1:
                if DEFAULT_NAME in bins_dict:
                    fea_name, bins = self._transform(X, bins_dict.get(DEFAULT_NAME), **kwargs)
                    return bins
                else:
                    return X

        # X.reset_index(inplace=True)

        _check_duplicated_keys(X)

        ###并行处理
        data_with_bins = Parallel(n_jobs=self.n_jobs)(
            delayed(self._transform)(X[col], bins, **kwargs) for col, bins in bins_dict.items() if col in X)
        ###并行处理

        if isinstance(X, dict):
            return dict(data_with_bins)
        else:
            bin_df = pd.DataFrame(dict(data_with_bins), index=X.index)
            X_cols = list(X.columns)
            no_bin_cols = list(set(X_cols) - set(bin_df.columns))
            bin_df[no_bin_cols] = X[no_bin_cols]
            # X.set_index('index', inplace=True)

            # return bin_df.set_index('index')
            return bin_df[X_cols]

    def _transform(self, X, splits, labels=False):
        """
        原始数据根据分割点变换原始X
        Args:
            X (DataFrame|array-like): 需要转换的原始X
            splits (array) : 分割点
            labels (bool) : 转换后的X是否需要标签，默认False，不需要

        Returns:

        """

        fea_name = DEFAULT_NAME
        if hasattr(X, 'name'):
            fea_name = X.name

        if splits.ndim > 1 or not np.issubdtype(splits.dtype, np.number):
            empty_mask = pd.isna(X).any()
            if empty_mask:
                X = X.astype(np.str)
            bins = self._raw_category_x_to_bin(X, splits)

        else:
            bins = np.zeros(len(X), dtype=int)

            if len(splits):  # TODO 需要看下splits为什么会有空
                if np.isnan(splits[-1]):
                    mask = pd.isna(X)
                    bins[~mask] = split_points_to_bin(X[~mask], splits[:-1])
                    bins[mask] = len(splits)
                else:
                    bins = split_points_to_bin(X, splits)

        if labels:
            splits_format = self.splits_point_format(
                splits,
                index=True)  ## ['0.[-inf, 0.7470300495624542)' '1.[0.7470300495624542, inf)' '2.nan']  ['0.B,F,G' '1.D,C,E' '2.nan']
            mask = (bins == EMPTY_BIN)
            bins = splits_format[bins]
            bins[mask] = FILLNA

        return fea_name, bins
        # row = pd.Series({fea_name:bins})
        # row.index = X.index
        # return row

    def _raw_category_x_to_bin(self, X, unique_val, is_empty_bin=False):
        """
        原始变量进行转换
        Args:
            X (array-like): 需要转换的X
            unique_val (array-like): 分割点
            is_empty_bin (bool): 是否有空箱

        Returns:

        """
        if is_empty_bin:
            bins = np.full(len(X), np.nan)
        else:
            bins = np.full(len(X), EMPTY_BIN)
            # bins = np.full(len(X), len(unique_val) - 1)

        for i in range(len(unique_val)):
            val = unique_val[i]
            if isinstance(val, str) and val == ELSE_GROUP:
                bins[bins == EMPTY_BIN] = i
            else:
                bins[np.isin(X, val)] = i
        return bins

    def _restore_category_splits(self, splits, x_val):
        """
        将分割点复原回原始值
        Args:
            splits (array-like): 分割点
            x_val (array-like): 原始分割点

        Returns:
            array: 原回原始值分割点
        """
        if x_val is None:
            return splits

        empty_mask = np.isnan(splits).any()
        if empty_mask:
            splits = splits[~np.isnan(splits)]

        if isinstance(x_val, np.ndarray):
            x_val = x_val.tolist()

        restore_category_splits = []
        start = 0
        for i in splits:
            index = math.ceil(i)
            restore_category_splits.append(x_val[start:index])
            start = index

        restore_category_splits.append(x_val[start:])

        if empty_mask:
            restore_category_splits.append(['nan'])

        return np.array(restore_category_splits)

    def splits_point_format(self, splits, index=False, ellipsis=None, decimal=None):
        """
        将分割点格式化，形如：[0.[4 ~ 7), 1.[7 ~ 10)]
        Args:
            splits (array-like): 分割点
            index (bool): 是否需要下标，0.[4 ~ 7)中的0
            ellipsis:

        Returns:
            array: 格式化后的分割点
        """
        ## 数值型：splits=[0.45343156        nan]，类别型：splits=[list(['B', 'F', 'G']) list(['D', 'C', 'E']) list(['nan'])]
        l = list()

        if not np.issubdtype(splits.dtype, np.number):
            # for i in splits:
            #     l.append(','.join(i))
            for i in splits:
                if isinstance(i, str) and i == ELSE_GROUP:
                    l.append(i)
                else:
                    label = ','.join(i)
                    if ellipsis is not None:
                        label = label[:ellipsis] + '..' if len(label) > ellipsis else label
                    l.append(label)

        else:
            is_empty_split = len(splits) > 0 and np.isnan(splits[-1])  # TODO 需要看下splits为什么会有空
            if is_empty_split:
                splits = splits[:-1]

            splits_ = [-np.inf] + splits.tolist() + [np.inf]
            for i in range(len(splits_) - 1):
                l.append('['.format(i) + str(splits_[i]) + ' ~ ' + str(splits_[i + 1]) + ')')

            if is_empty_split:
                l.append('nan')

        if index:
            l = ["{}.{}".format(i, v) for i, v in enumerate(l)]

        return np.array(l)

    def manual_bin(self, manual_set_dict):
        """
        手动分箱
        Args:
            manual_set_dict (dict|array-like): map结构的分箱点，形如: {'D157': [1.0, 2.0, 3.0, 5.0]}

        Returns:

        """
        if not isinstance(manual_set_dict, dict):
            manual_set_dict = {
                DEFAULT_NAME: manual_set_dict,
            }

        assert isinstance(manual_set_dict, dict), '请传入类似 {\'D157\': [1.0, 2.0, 3.0, 5.0]}'

        for i in manual_set_dict:
            self.splits_dict[i] = np.array(manual_set_dict[i])
        return self

    def export(self, to_dataframe=False, to_json=None, to_csv=None, bin_format=True, index=True):
        """
        导出规则到dict或json或csv文件
        Args:
            to_dataframe (bool): 是否导出成pd.DataFrame形式
            to_json (str): 保存成json的路径
            to_csv (str): 保存成csv的路径
            bin_format (bool): 是否将分割点格式化
            index (bool): 分割点格式化是否需要下标

        Returns:
            dict: 分割点规则字典

        """

        splits = copy.deepcopy(self.splits_dict)
        if bin_format:
            # splits = {k: self.splits_point_format(v, index=False).tolist() for k, v in splits.items()}
            splits = {k: list(self.splits_point_format(v, index=index)) for k, v in splits.items()}
        else:
            splits = {k: v.tolist() for k, v in splits.items()}
        if to_json is not None:
            save_json(splits, to_json)

        if to_dataframe or to_csv is not None:

            row = []
            for var_name in splits:
                for bin in splits[var_name]:
                    row.append({
                        'feature': var_name,
                        'bins': bin
                    })

            splits = pd.DataFrame(row)

        if to_csv is not None:
            splits.to_csv(to_csv, index=False)

        return splits


class WoeTransformer(TransformerMixin):
    """
    woe转换类
    """

    def __init__(self, n_jobs=-1):
        self.fea_woe_dict = dict()
        self.n_jobs = n_jobs

    def __len__(self):
        return len(self.fea_woe_dict.keys())

    def __contains__(self, key):
        return key in self.fea_woe_dict

    def __getitem__(self, key):
        return self.fea_woe_dict[key]

    def __setitem__(self, key, value):
        self.fea_woe_dict[key] = value

    def __iter__(self):
        return iter(self.fea_woe_dict)

    @df_exclude_cols
    @df_select_dtypes
    def fit(self, X, y, update = False):
        """
        woe转换
        Args:
            X (DataFrame|array-like): 需要转换woe的X
            y (str|array-like): 目标变量
            exclude (str|array-like): 排除的变量，该变量将不参与woe计算
            select_dtypes (str|numpy.dtypes): `'object'`, `'number'` 等. 只有选定的数据类型才会被计算

        Returns:

        """
        if not isinstance(X, pd.DataFrame):
            fea_name, value_woe = self._fit_woe(X, y)
            self.fea_woe_dict[fea_name] = value_woe
            return self

        if isinstance(y, str):
            # y = X.pop(y)
            X, y = split_target(X, y)

        _check_duplicated_keys(X)

        data = Parallel(n_jobs=self.n_jobs)(
            delayed(self._fit_woe)(X[col], y) for col in X)  # 批量处理

        if update:
            self.fea_woe_dict.update(dict(data))
        else:
            self.fea_woe_dict = dict(data)

        return self

    def _fit_woe(self, X, y):
        """
        woe转换
        Args:
            X (DataFrame|array-like): 需要计算woe的X
            y (str|array-like): 目标变量

        Returns:
            str : 变量名
            array : 计算出来的woe值

        """
        fea_name = DEFAULT_NAME
        if hasattr(X, 'name'):
            fea_name = X.name
        X = np.copy(X)  # Series, array 转 np.array
        if X.dtype.type is np.object_:
            X = X.astype(np.str)
        unique_val = np.unique(X)

        # # TODO 如果X是连续性变量，且有空。计算出来的woe不正确
        # if X.dtype.type is np.object_:
        #     X = X.astype(np.str)
        #     unique_val = np.unique(X)
        # else:
        #     unique_val = [int(i) for i in np.unique(X)]

        value_woe = dict()
        for val in unique_val:
            y_prob, n_prob = probability(y, mask=(X == val))  #
            value_woe[val] = WOE(y_prob, n_prob)

        return fea_name, value_woe

    def transform(self, X, fea_woe_dict={}, **kwargs):
        """
        将原始值用woe值替换
        Args:
            X (DataFrame|array-like): 需要转换woe的X
            fea_woe_dict (dict): 变量和woe值的字典，形如：{'D157': {0: -0.46554351769099783, 1: -0.10263802400162944}}
            **kwargs:

        Returns:
            DataFrame: 转换woe后的X

        """

        if not isinstance(fea_woe_dict, dict):
            assert """请传入类似 {'D157': {'0.[-inf, 1.5)': -0.46554351769099783, '1.[1.5, 2.5)': -0.10263802400162944, '2.[2.5, 3.5)': 0.9591358533174893, '3.[3.5, 4.5)': 1.115806812932841, '4.[4.5, 7.5)': 1.1319717497861965, '5.[7.5, inf)': 2.369093204627806, '6.nan': -1.2516811662966312}}"""

        if not fea_woe_dict:
            fea_woe_dict = self.fea_woe_dict

        if getattr(X, 'ndim', 1) == 1:

            if hasattr(X, 'name'):  # pd.Series
                if X.name in fea_woe_dict:
                    fea_name, woe = self._transform(X, fea_woe_dict.get(X.name), fea_name=X.name)
                    return woe
                else:
                    return X

            if len(fea_woe_dict) == 1:
                if DEFAULT_NAME in fea_woe_dict:
                    fea_name, woe = self._transform(X, fea_woe_dict.get(DEFAULT_NAME))
                    return woe
                else:
                    return X

        # X.reset_index(inplace=True)

        _check_duplicated_keys(X)

        ###并行处理
        data_with_bins = Parallel(n_jobs=self.n_jobs)(
            delayed(self._transform)(X[col], woe_dict, fea_name=col, **kwargs) for col, woe_dict in fea_woe_dict.items()
            if col in X)
        ###并行处理

        if isinstance(X, dict):
            return dict(data_with_bins)
        else:
            bin_df = pd.DataFrame(dict(data_with_bins), index=X.index)
            X_cols = list(X.columns)
            no_bin_cols = list(set(X_cols) - set(bin_df.columns))
            bin_df[no_bin_cols] = X[no_bin_cols]
            # X.set_index('index', inplace=True)

            # return bin_df.set_index('index')
            return bin_df[X_cols]

    def _transform(self, X, woe_dict={}, fea_name=DEFAULT_NAME, other='min'):
        """

        Args:
            X (DataFrame|array-like): 需要转换woe的X
            woe_dict (dict): 变量和woe值的字典，形如：{0: -0.46554351769099783, 1: -0.10263802400162944}
            fea_name (str): 变量名
            other (str): 未来出现的新值给对应的woe值

        Returns:

        """

        try:
            woe = np.zeros(len(X))
        except:
            woe = np.zeros(np.array(X).size)

        if other == 'min':
            other = np.min(list(woe_dict.values()))
        elif other == 'max':
            other = np.max(list(woe_dict.values()))

        woe[np.isin(X, list(woe_dict.keys()), invert=True)] = other

        for k, v in woe_dict.items():
            woe[X == k] = v

        return fea_name, woe

    def load(self, manual_set_dict):
        """
        自定的woe值
        Args:
            manual_set_dict (dict): map结构的woe值，形如: {'D157': {0: -5.545177444479562, 1: 5.497168225293202}}

        Returns:

        """

        assert isinstance(manual_set_dict, dict), '请传入类似 {\'D157\': {0: -5.545177444479562, 1: 5.497168225293202}}'

        for i in manual_set_dict:
            self.fea_woe_dict[i] = manual_set_dict[i]
        return self

    def export(self, to_dataframe=False, to_json=None, to_csv=None, var_bin_woe={}):
        """
        导出规则到dict或json或csv文件
        Args:
            to_dataframe (bool): 是否导出成pd.DataFrame形式
            to_json (str): 保存成json的路径
            to_csv (str): 保存成csv的路径
            var_bin_woe (dict): {'D157': {0: -5.545177444479562, 1: 5.497168225293202}}

        Returns:
            dict: 分割点规则字典

        """

        if var_bin_woe:
            fea_bin_woe = var_bin_woe
        else:
            fea_bin_woe = copy.deepcopy(self.fea_woe_dict)
            #fea_bin_woe = {k: {int(i): j for i, j in v.items()} for k, v in fea_bin_woe.items()}
        if to_json is not None:
            save_json(fea_bin_woe, to_json)

        if to_dataframe or to_csv is not None:
            row = list()
            for var_name in fea_bin_woe:
                for bin, woe in fea_bin_woe[var_name].items():
                    row.append({
                        'feature': var_name,
                        'bins': bin,
                        'woe': woe
                    })

            fea_bin_woe = pd.DataFrame(row)

        if to_csv is not None:
            fea_bin_woe.to_csv(to_csv, index=False)

        return fea_bin_woe



#==============================================================================
# File: utils.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: utils.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-05-20
'''

import gc
import json
import math
import pickle
import re
import sys
import warnings

import numpy as np
import pandas as pd
import statsmodels
import statsmodels.api as sm
# from catboost import CatBoostRegressor
# from lightgbm import LGBMRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_curve, accuracy_score, r2_score, recall_score, precision_score, f1_score
from sklearn.model_selection import train_test_split
# from xgboost.sklearn import XGBRegressor
from functools import wraps

import autobmt

warnings.filterwarnings('ignore')

from .logger_utils import Logger

log = Logger(level="info", name=__name__).logger

FILLNA = -999
CONTINUOUS_NUM = 10


def support_dataframe(require_target=True):
    """用于支持dataframe的装饰器
    """

    def sup_df(fn):
        @wraps(fn)
        def func(frame, *args, **kwargs):
            if not isinstance(frame, pd.DataFrame):
                return fn(frame, *args, **kwargs)

            frame = frame.copy()
            if require_target and isinstance(args[0], str):
                target = frame.pop(args[0])
                args = (target,) + args[1:]
            elif 'target' in kwargs and isinstance(kwargs['target'], str):
                kwargs['target'] = frame.pop(kwargs['target'])

            if 'return_bin' not in kwargs:
                kwargs['return_bin'] = True

            res = dict()
            for col in frame:
                r = fn(frame[col], *args, **kwargs)

                if not isinstance(r, np.ndarray):
                    if isinstance(r, tuple):
                        r = r[1]
                    else:
                        r = [r]
                res[col] = r
            return pd.DataFrame(res)

        return func

    return sup_df


def get_splitted_data(df_selected, target, selected_features):
    X = {}
    y = {}

    X['all'] = df_selected[selected_features]
    y['all'] = df_selected[target]

    for name, df in df_selected.groupby('type'):
        X[name] = df[selected_features]
        y[name] = df[target]

    if not X.__contains__('oot'):
        X['oot'] = None
        y['oot'] = None

    return X['all'], y['all'], X['train'], y['train'], X['test'], y['test'], X['oot'], y['oot']


# def to_score(x, A=404.65547022, B=72.1347520444):
#     result = round(A - B * math.log(x / (1 - x)))
#
#     if result < 0:
#         result = 0
#     if result > 1200:
#         result = 1200
#     result = 1200 - result
#     return result

def to_score(x, A=404.65547021957406, B=72.13475204444818, positive_corr=False):
    """
    概率值转分
    Args:
        x (float): 模型预测的概率值
        base_score=600
        odds=15
        pdo=50
        rate=2
        #实际意义为当比率为1/15，输出基准评分600，当比率为基准比率2倍时，1/7.5，基准分下降50分，为550
        A (float): 评分卡offset；；；offset = base_score - (pdo / np.log(rate)) * np.log(odds)
        B (float): 评分卡factor；；；factor = pdo / np.log(rate)
        positive_corr: 分数与模型预测的概率值是否正相关。默认False，负相关，即概率约高，分数越低

    Returns:
        score (float): 标准评分
    """
    if x <= 0:
        x = 0.000001
    elif x >= 1:
        x = 0.999999
    result = round(A - B * math.log(x / (1 - x)))

    if positive_corr:
        if result < 0:
            result = 0
        if result > 1200:
            result = 1200
        result = 1200 - result
        return result
    else:
        if result < 300:
            result = 300
        if result > 900:
            result = 900
        return result


def score2p(x, A=404.65547022, B=72.1347520444):
    """
    分转概率
    Args:
        x (float): 标准评分
        pdo=50;rate=2;odds=15;base_score=600
        A (float): 评分卡offset；；；offset = base_score - (pdo / np.log(rate)) * np.log(odds)
        B (float): 评分卡factor；；；factor = pdo / np.log(rate)

    Returns:
        p (float): 概率值
    """
    return 1 / (1 + np.exp((x - A) / B))


def train_test_split_(df_src, target='target', test_size=0.3):
    """
    样本切分函数.先按target分类，每类单独切成train/test，再按train/test合并，
    使得train/test的badrate能高度一致
    Args:
        df_src:
        target:
        test_size:

    Returns:

    """

    l = [[], [], [], []]
    for target_value, X in df_src.groupby(target):

        X[target] = target_value

        row = train_test_split(X.drop(labels=target, axis=1), X[target], test_size=test_size, random_state=1024)

        for i in range(0, 4):
            l[i].append(row[i])

    list_df = []
    for i in range(0, 4):
        list_df.append(pd.concat(l[i]))

    return tuple(list_df)


def split_data_type(df, key_col='id', target='target', apply_time='apply_time', test_size=0.3):
    if df[target].isin([0, 1]).all():
        log.info('样本y值在0，1')
    else:
        log.info('\033[0;31m样本y值不在0，1之间，请检查！！！\033[0m')

    assert df[target].isin([0, 1]).all()

    log.info('样本情况：', df.shape)
    df.drop_duplicates(subset=key_col, inplace=True)
    log.info('分布情况：', df.groupby(target)[key_col].count().sort_index())
    log.info('样本drop_duplicates情况：', df.shape)

    # ---------查看各月badrate---------------------
    df['apply_month'] = df[apply_time].map(lambda s: s[:7])
    log.info(df.groupby('apply_month').describe()[target])
    del df['apply_month']

    # ---------样本划分----------------------------
    ##需要oot
    # df_selected = df_id #can filter records here
    # # df_oot = df_selected[df_selected['apply_time']>= '2019-04-01']
    # # X_train = df_selected[df_selected['apply_time']<= '2019-01-31']
    # # X_test = df_selected[(df_selected['apply_time']> '2019-01-31') & (df_selected['apply_time']< '2019-04-01')]

    # df_oot = df_selected[df_selected['apply_time']>= '2019-03-01']
    # X_train = df_selected[df_selected['apply_time']<= '2018-12-31']
    # X_test = df_selected[(df_selected['apply_time']> '2018-12-31') & (df_selected['apply_time']< '2019-03-01')]

    # #X_train, X_test, y_train, y_test = geo_train_test_split(df_not_oot,label=label)

    # df_id.loc[df_oot.index,'type'] = 'oot'
    ##需要oot

    # 不需要oot的时候运行下面这一行代码
    X_train, X_test, y_train, y_test = train_test_split_(df, target=target, test_size=test_size)
    # X_train, X_test, y_train, y_test = train_test_split(df_id.drop(columns=target), df_id[target], test_size=test_size,
    #                                                     random_state=123)
    # 不需要oot的时候运行下面这一行代码

    df.loc[X_train.index, 'type'] = 'train'
    df.loc[X_test.index, 'type'] = 'test'

    log.info(df.groupby('type').describe()[target])

    # ----------输出---------------------------------
    # df_id.to_csv(data_dir + '{}_split.csv'.format(client_batch), index=False)
    return df


def select_features_dtypes(df, exclude=None):
    """
    根据数据集，筛选出数据类型
    Args:
        df: 数据集
        exclude: 排除不需要参与筛选的列

    Returns:

    """
    if exclude is not None:
        df = df.drop(columns=exclude)
    # 筛选出数值类型列
    numeric_list = df.select_dtypes([np.number]).columns.tolist()

    no_numeric_df = df.select_dtypes(include=['object'])
    # 将object类型的列尝试转成时间类型
    dates_objs_df = no_numeric_df.apply(pd.to_datetime, errors='ignore')
    # 筛选出字符类型列
    objs_list = dates_objs_df.select_dtypes(include=['object']).columns.tolist()
    # 筛选出时间类型列
    # dates_df = list(set(dates_objs_df.columns) - set(objs_df.columns))
    date_list = dates_objs_df.select_dtypes(include=['datetime64']).columns.tolist()

    assert len(numeric_list) + len(objs_list) + len(date_list) == df.shape[1]

    return numeric_list, objs_list, date_list


def filter_miss(df, miss_threshold=0.9):
    """

    Args:
        df (DataFrame): 用于训练模型的数据集
        miss_threshold: 缺失率大于等于该阈值的变量剔除

    Returns:

    """
    names_list = []
    for name, series in df.items():
        n = series.isnull().sum()
        miss_q = n / series.size
        if miss_q < miss_threshold:
            names_list.append(name)
    return names_list


###################
def step_evaluate_models(df, features, target, stepname="", is_turner=False):
    """
    用lr,xgb,评估train/test/oot数据集的auc,ks
    Args:
        df: 数据集，包含y,type,features
        features: 入模特征
        target: 目标值
        stepname: 标识是在哪一步进行评估的
        is_turner: 是否需要进行调参

    Returns:

    """
    X_train = df[df['type'] == 'train'][features]
    y_train = df[df['type'] == 'train'][target]
    data = df[['type', target]]

    # xgb默认参数
    xgb_params = {"base_score": 0.5, "booster": "gbtree", "colsample_bylevel": 1, "colsample_bytree": 0.8, "gamma": 3,
                  "learning_rate": 0.1, "max_delta_step": 0, "max_depth": 6, "min_child_weight": 50,
                  "n_estimators": 200, "n_jobs": -1, "objective": "binary:logistic", "random_state": 0,
                  "reg_alpha": 5, "reg_lambda": 5, "scale_pos_weight": 1,
                  "subsample": 0.8}
    lightgbm_params = {'boosting_type': 'gbdt', 'num_threads': 20,
                       'min_child_weight': 50, 'max_depth': 6,
                       'colsample_bytree': 0.8, 'subsample': 0.8,
                       'num_iterations': 200, 'learning_rate': 0.1, 'verbose': -1
                       }
    rf_params = {'max_depth': 6,
                 'n_estimators': 200,
                 'min_samples_leaf': 60, 'n_jobs': -1, 'min_samples_split': 60,
                 'verbose': 0
                 }
    catboost_params = {'depth': 6, 'l2_leaf_reg': 3,
                       'n_estimators': 200, 'learning_rate': 0.1,
                       'subsample': 0.8
                       }
    if is_turner:
        # 需要调参走这个逻辑
        # models = {
        #     "lr": LogisticRegression().fit(X_train, y_train),
        #     "rf": rf_turner(X_train, y_train),
        #     "xgb": xgb_turner(X_train, y_train, X_test, y_test)[1],
        #     # "lightgbm": lightgbm_turner(X_train, y_train, X_test, y_test)[1],
        #     # "catboost": catboost_turner(X_train, y_train, X_test, y_test)[1]
        # }

        pass
    else:
        # 使用默认参数进行训练评估
        models = {
            "lr": LogisticRegression().fit(X_train, y_train),
            "rf": RandomForestRegressor(**rf_params).fit(X_train, y_train),
            # "xgb": XGBRegressor(**xgb_params).fit(X_train, y_train),
            # "lightgbm": LGBMRegressor(**lightgbm_params).fit(X_train, y_train),
            # "catboost": CatBoostRegressor(**catboost_params, verbose=False).fit(X_train, y_train)
        }

    result = []
    for name, model in models.items():
        # model = model.fit(X_train, y_train)
        if isinstance(model, LogisticRegression):
            data['prob'] = model.predict_proba(df[features])[:, 1]
        elif isinstance(model, RandomForestRegressor):
            data['prob'] = model.predict(df[features])
        # elif isinstance(model, XGBRegressor):
        #     data['prob'] = model.predict(df[features])
        # elif isinstance(model, LGBMRegressor):
        #     data['prob'] = model.predict(df[features])
        # elif isinstance(model, CatBoostRegressor):
        #     data['prob'] = model.predict(df[features])

        df_splitted_type_auc_ks = data.groupby('type').apply(
            lambda df_: pd.Series({'{}_auc'.format(name): autobmt.get_auc(df_[target], df_['prob']),
                                   '{}_ks'.format(name): autobmt.get_ks(df_[target], df_['prob'])}))
        result.append(df_splitted_type_auc_ks)

    evaluate_df = pd.concat(result, axis=1)
    one_step_df = merge_rows_one_row_df(evaluate_df)
    one_step_df['feature_num'] = len(features)  # 加上特征个数
    one_step_df['stepname'] = stepname  # 用于标识是哪一步计算的auc和ks
    xgb_evaluate_log = str(one_step_df.filter(like='xgb').applymap(lambda x: round(x, 4)).iloc[0].to_dict())
    del_df(data)
    del_df(X_train)
    del_df(y_train)
    return one_step_df, xgb_evaluate_log


def model_predict_evaluate(model, df, features, target, A=404.65547022, B=72.1347520444, exclude_cols=None,
                           is_return_var=False):
    """
    1. 利用已经训练好的model对象，对数据集进行预测
    2. 数据集中需要包含有一列type，标识train/test/oot
    3. 返回各个数据集的auc,ks
    4. 返回type,真实y值，预测的概率值，预测的标准分(供后续输入到模型报告中生成外部报告)
    LogisticRegression().fit()
    model = sm.Logit(train_data[target], sm.add_constant(train_data[features])).fit()
    XGBRegressor().fit()
    Args:
        model: 已经训练好的模型对象，支持lr,xgb
        df: 用于训练模型的数据集
        features: 入模变量
        target: 目标值
        A: 评分卡大A
        B: 评分卡大B
        exclude_cols: 样本中的原始字段（非x，y）

    Returns:

    """
    assert 'type' in df.columns.tolist()
    data = df.copy()
    if isinstance(model, LogisticRegression):
        data['p'] = model.pa(data[features])[:, 1]
    elif isinstance(model, statsmodels.discrete.discrete_model.BinaryResultsWrapper):
        log.info("mode is : statsmodels.discrete.discrete_model.BinaryResultsWrapper")
        data['p'] = model.predict(sm.add_constant(data[features]))
    # elif isinstance(model, XGBRegressor):
    #     data['p'] = model.predict(data[features])
    # elif isinstance(model, LGBMRegressor):
    #     data['p'] = model.predict(data[features])
    # elif isinstance(model, CatBoostRegressor):
    #     data['p'] = model.predict(data[features])
    elif isinstance(model, RandomForestRegressor):
        data['p'] = model.predict(data[features])
    data['score'] = data['p'].map(lambda x: to_score(x, A, B))
    evaluate_df = data.groupby('type').apply(
        lambda df: pd.Series({'auc': autobmt.get_auc(df[target], df['p']),
                              'ks': autobmt.get_ks(np.array(df[target]), df['p'])}))

    if is_return_var:
        return evaluate_df, data[exclude_cols + ['p', 'score'] + features]
    else:
        return evaluate_df, data[exclude_cols + ['p', 'score']]


def del_df(df):
    """
    清空一个dataframe
    Args:
        df (DataFrame): 用于训练模型的数据集

    Returns:

    """
    # df.drop(df.index, inplace=True)
    del df
    gc.collect()


def merge_rows_one_row_df(df, name="", stepname=None):
    """将一个多行的dataframe合并成只有一行的dataframe"""
    tmp_arr = []
    for i in range(df.shape[0]):
        tmp = df.iloc[i, :].add_prefix("{}_{}".format(df.index[i], name))
        tmp_arr.append(tmp)

    result_df = pd.DataFrame(pd.concat(tmp_arr, axis=0)).T
    if stepname is not None:  # 合并成一行后,增加一列标识，用于和别的评估进行区分
        result_df['stepname'] = stepname
    return result_df


def get_max_corr_feature(df, features):
    """返回每个变量与其相关性最高的变量以及相关性"""
    corr_df = df.loc[:, features].corr()
    corr_value_series = corr_df.apply(lambda x: x.nlargest(2)[1]).rename("corr_value")
    corr_name_series = corr_df.apply(lambda x: x.nlargest(2).index[1]).rename("corr_name")

    max_corr_df = pd.concat([corr_name_series, corr_value_series], axis=1)
    return max_corr_df


def dump_to_pkl(contents, path):
    pickle.dump(contents, open(path, "wb"))


def load_from_pkl(path):
    return pickle.load(open(path, 'rb'))


def read_sql_string_from_file(path):
    with open(path, 'r', encoding='utf-8') as fb:
        sql = fb.read()
        return sql


def fea_woe_dict_format(fea_woe_dict, splits_dict):
    for j in fea_woe_dict:
        range_format = {int(re.match(r"^(\d+)\.", i).group(1)): i for i in splits_dict[j]}
        fea_woe_dict[j] = {range_format[k]: v for k, v in fea_woe_dict[j].items()}

    return fea_woe_dict


###################

def save_json(res_dict, file, indent=4):

    try:
        if isinstance(file, str):
            file = open(file, 'w')
        json.dump(res_dict, file, ensure_ascii=False, indent=indent)
        file.close()
    except:
        if isinstance(file, str):
            file = open(file, 'r+')
        file.seek(0)
        file.truncate()
        res_dict_ = {}
        for key, value in res_dict.items():
            res_dict_[key] = {int(k): v for k, v in value.items()}
        json.dump(res_dict_, file, ensure_ascii=False, indent=indent)
        file.close()


def load_json(file):
    """
    读取json文件
    """
    if isinstance(file, str):
        file = open(file, 'r')

    with file as f:
        res_dict = json.load(f)

    return res_dict


def is_continuous(series):
    series = to_ndarray(series)
    if not np.issubdtype(series.dtype, np.number):
        return False

    n = len(np.unique(series))
    return n > CONTINUOUS_NUM or n / series.size > 0.5
    # return n / series.size > 0.5


def to_ndarray(s, dtype=None):
    """
    """
    if isinstance(s, np.ndarray):
        arr = np.copy(s)
    elif isinstance(s, pd.core.base.PandasObject):
        arr = np.copy(s.values)
    else:
        arr = np.array(s)

    if dtype is not None:
        arr = arr.astype(dtype)
    # covert object type to str
    elif arr.dtype.type is np.object_:
        arr = arr.astype(np.str)

    return arr


def fillna(feature, fillna_va=FILLNA):
    # 复制array 或者 将pandas.core.series.Series变成array
    copy_fea = np.copy(feature)

    mask = pd.isna(copy_fea)

    copy_fea[mask] = fillna_va

    return copy_fea


def split_empty(feature, y=None, return_mask=True):
    copy_fea = np.copy(feature)
    mask = pd.isna(copy_fea)

    copy_fea = copy_fea[~mask]
    if y is not None:
        copy_y = np.copy(y)
        copy_y = copy_y[~mask]
        return copy_fea, copy_y, mask
    return copy_fea, mask


def split_points_to_bin(feature, splits):
    """split points to bin feature
    """
    # log.info("def split_points_to_bin(feature, splits):")
    # log.info(splits)
    feature = fillna(feature)
    return np.digitize(feature, splits)


def np_count(arr, value, default=None):
    c = (arr == value).sum()

    if default is not None and c == 0:
        return default

    return c


def unpack_tuple(x):
    if len(x) == 1:
        return x[0]
    else:
        return x


def split_target(frame, target):
    """
    """
    if isinstance(target, str):
        f = frame.drop(columns=target)
        t = frame[target]
    else:
        f = frame.copy(deep=False)
        t = target

    return f, t


##################g
# corr
def get_corr(df):
    return df.corr()


# accuracy
def get_accuracy(y_true, y_pred):
    return accuracy_score(y_true, y_pred)


# precision
def get_precision(y_true, y_pred):
    if len(np.unique(y_true)) == 2:
        return precision_score(y_true, y_pred)
    else:
        return precision_score(y_true, y_pred, average='macro')


# recall
def get_recall(y_true, y_pred):
    if len(np.unique(y_true)) == 2:
        return recall_score(y_true, y_pred)
    else:
        return precision_score(y_true, y_pred, average='macro')


# f1
def get_f1(y_true, y_pred):
    if len(np.unique(y_true)) == 2:
        return f1_score(y_true, y_pred)
    else:
        return f1_score(y_true, y_pred, average='macro')


# r2
def r2(preds, target):
    return r2_score(target, preds)


def get_best_threshold(y_true, y_pred):
    fpr, tpr, threshold = roc_curve(y_true, y_pred)
    ks = list(tpr - fpr)
    thresh = threshold[ks.index(max(ks))]
    return thresh


def get_bad_rate(df):
    return df.sum() / df.count()


def sigmoid(z):
    return 1 / (1 + np.exp(-z))


##################g


##################考虑用太极实现
def t_cols_sum_axis_1_np(arr):
    res = np.zeros(arr.shape[0], dtype=float)

    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            res[i] += arr[i, j]

    return res


def t_cols_sum_axis_0_np(arr):
    res = np.zeros(arr.shape[1], dtype=float)

    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            res[j] += arr[i, j]

    return res


def t_min_np(arr):
    res = np.inf
    for i in range(arr.shape[0]):
        if res > arr[i]:
            res = arr[i]

    return res


def t_sum_np(arr):
    res = 0
    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            res += arr[i, j]

    return res

##################考虑用太极实现



#==============================================================================
# File: XGB全量.py
#==============================================================================


"""
模型建立
@author: kantt
"""
#%%
import pandas as pd
import numpy as np
import toad
import os 
import re
from datetime import datetime
from sklearn.model_selection import train_test_split
#%% 数据导入与处理
data_tmp1 = pd.read_csv(r'C:\Users\ruizhi\Desktop\kantt\1.base_data\valid_cust_first.csv') 

#列名筛选
col_list=data_tmp1.columns
drop_col=['Unnamed: 0.1','loan_amount_y','Unnamed: 0','order_no_y','channel_id_y','apply_date_y','apply_time','id_no_des_y','channel_id','apply_date_real']
#          ,'value_012_pd_time','model_score_01br_score','model_score_01_xysl_1.0','model_score_01_zr_tdzx']
data_tmp1.drop(columns=drop_col,inplace=True)
data_tmp1=data_tmp1.rename(columns={"model_score_01_xysl_3.0":"model_score_01_xysl_3"})
#剔除灰样本确认模型样本
data = data_tmp1[data_tmp1['Firs6ever30'].isin([0.,2.])]
#    channel_id_x  Firs6ever30  user_id
# 0           167          0.0    61704
# 1           167          2.0     5994
# 2           174          0.0    24740
# 3           174          2.0     2879

cal1=data.groupby(['apply_month','Firs6ever30','channel_id_x'])['user_id'].count().unstack().reset_index()

print('数据大小：', data.shape) #(95317, 143)

data['target']=data['Firs6ever30']/2

cal_tmp=data.groupby(['total_periods','Firs6ever30'])['user_id'].count().unstack().reset_index()

#%%  #STEP1. - 数据集切分  训练集测试集
model = data
#[data.apply_month.isin(['2022-05','2022-06','2022-07','2022-08','2022-09'])]
#oot = data[data.apply_month.isin(['2022-10','2022-11','2022-12'])]

allFeatures = list(model.columns.drop(['id_no_des_x','order_no_x','apply_date_x','order_status','loan_period','channel_id_x','loan_amount_x','lending_time','loan_rate',
'total_periods','Firs3ever15','Firs3ever30','Firs6ever15','Firs6ever30','apply_month','create_time',
'channel_name','cert_type','auth_status','auth_credit_amount','cust_type','close_reason','time_diff']))

X1 = model[model.channel_id_x==167][allFeatures]
Y1 = model[model.channel_id_x==167]['target']
X2 = model[model.channel_id_x==174][allFeatures]
Y2 = model[model.channel_id_x==174]['target']

X_train1, X_test1, y_train1, y_test1 = train_test_split(X1 ,Y1, test_size=0.3, random_state=88, stratify=Y1)
X_train2, X_test2, y_train2, y_test2 = train_test_split(X2 ,Y2, test_size=0.3, random_state=88, stratify=Y2)

X_train=pd.concat([X_train1,X_train2],axis=0)
X_test=pd.concat([X_test1,X_test2],axis=0)
y_train=pd.concat([y_train1,y_train2],axis=0)
y_test=pd.concat([y_test1,y_test2],axis=0)

model.loc[X_train.index,'sample_set']='train'
model.loc[X_test.index,'sample_set']='test'

dt_train=model[model.sample_set=='train'] #59094
dt_test=model[model.sample_set=='test'] #25327


cal2=dt_train.groupby(['apply_month','Firs6ever30','channel_id_x'])['user_id'].count().unstack().reset_index()
cal3=dt_test.groupby(['apply_month','Firs6ever30','channel_id_x'])['user_id'].count().unstack().reset_index()


#%%#评分转换
def creditCards(paramsEst,
                bin_woe_map_df,
                basepoints, 
                baseodds, 
                PDO,
                odds_new,
                odds_old,
                has_intercept_score):
    
    """
    output credit card for each var in model
    --------------------------------------------
    ParamsEst: pandas Series, 模型的参数估计结果，index为变量名, value为变量系数估计值
    bin_woe_map_df：变量分bin和woe的对应表
    basepoints: 标准odds，即baseodds时的标准评分
    baseodds: 标准odds，指定的标准odds
    PDO: odds增加翻倍时，评分变化的分数，风险模型，若要1的概率越高分数越低，PDO取负值；若要1的概率越高分数越高，PDO取正值
    odds_new：抽样建模样本的odds：1的数量/0的数量
    odds_old：原始未抽样样本的odds：1的数量/0的数量
    has_intercept_score: 是否含截距项分数
    -------------------------------------------
    Return
    creditCard: 评分卡结果，pandas dataframe
    """

    # 计算A&B
    beta = PDO/np.log(2)
    alpha = basepoints + beta*np.log(baseodds)
    
    #alpha, beta = _score_cal(basepoints, baseodds, PDO)
    odds_ratio = odds_new/odds_old
    
    # 计算截距项基础分
    if has_intercept_score:
        points_0 = round(alpha - beta * paramsEst['const'] - beta * np.log(odds_ratio))
    else:
        points_0 = alpha - beta * paramsEst['const'] - beta * np.log(odds_ratio)
    # 变量个数
    num_vars = len(paramsEst)-1
      
    print('标准odds: ' + str(baseodds))
    print('标准odds时的标准评分: ' + str(basepoints))
    print('odds翻倍时评分变化PDO：' + str(PDO))
    print('评分计算公式截距系数alpha = ' + str(alpha))
    print('评分计算公式斜率系数beta = ' + str(beta))
    print('建模样本的总体odds：' + str(odds_new))
    print('原始样本的总体odds：' + str(odds_old))
    print('模型变量个数: ' + str(num_vars))
    print('若评分卡包含截距项评分，截距项评分未: ' + str(round(points_0)))
    print('若评分卡不包含截距项评分，每个变量分摊截距评分or变量缺失默认评分: ' + str(round(points_0/num_vars)))
  
    # woe
    var_list = list(paramsEst.index)[1:]
    woe_maps_dict = {}
    for var in var_list:
        var_bin_woe_df = bin_woe_map_df[bin_woe_map_df['var']==var]
        var_bin_woe_df_dict = var_bin_woe_df.pivot(index='var', columns='bin', values='woe').to_dict('index')
        woe_maps_dict[var] = var_bin_woe_df_dict[var]
        
    # 根据各段woe，计算相应得分
    points = pd.DataFrame()
    for k in woe_maps_dict.keys():
        d = pd.DataFrame(woe_maps_dict[k], index=[k]).T
        if has_intercept_score:
            d['points'] = (-beta * d.loc[:, k] * paramsEst[k]).round()
        else:
            d['points'] = (points_0*1.0/num_vars - beta * d.loc[:, k] * paramsEst[k]).round()
        d = d.rename(columns={k: 'var_woe'})
        # range
        bin_map = bin_woe_map_df[bin_woe_map_df['var']==k]
        bin_map.index = bin_map['bin']
        bin_map.index.name = None
        bin_map = bin_map[['bin', 'woe']]
        bin_map['var'] = k
        d = d.merge(bin_map, left_index=True, right_index=True)
        # 构造新的index
        n = len(d)
        ind_0 = []
        i = 0
        while i < n:
            ind_0.append(k)
            i += 1
        d.index = [ind_0,list(d.index)]
        points = pd.concat([points, d], axis=0)
    points_df = points[['var', 'bin',  'points', 'var_woe']]    

    #输出评分卡
    if has_intercept_score:
        points_0_df = pd.DataFrame([['basePoints', '0', points_0, 1]], 
                                   columns = ['var', 'bin', 'points', 'var_woe'])
        points_0_df.index=[['basePoints'], ['0']]
        credit_card = pd.concat([points_0_df, points_df], axis=0)
        credit_card.index.names = ["varname", "binCode"]
    else:
        credit_card = points_df
        credit_card.index.names = ["varname", "binCode"]
    return credit_card
#%%
baseodds = 35
basepoints = 700
PDO = 60  
odds_old = 60510/6211
odds_new = 60510/6211
 
params = {'const':-2.2765 
          ,'crd_loan_gap_b': 0.6768
          ,'utl_b':0.9296
          ,'model_score_01_b':0.5923
          ,'model_score_01_x_tianchuang_b':0.7544
          ,'model_score_01_y_tianchuang_b':0.6475
          ,'model_score_01_xysl_3_b':0.8891
          } 

paramsEst = pd.Series(params)
paramsEst.index = [k.replace("_b", "") for k in paramsEst.index]
bin_woe_map_df=pd.read_excel(r'C:\Users\ruizhi\Desktop\kantt\0.cal_summary\bin_woe_map.xlsx')

credit_card_df = creditCards(paramsEst = paramsEst,     # 模型参数估计，Series，index为模型变量（不带后缀"_WOE"）
                            bin_woe_map_df = bin_woe_map_df,   # 变量分bin和woe的对应表
                            basepoints = basepoints,    # # 标准odds，1的占比：0的占比（1的数量：0的数量）
                            baseodds = baseodds,  # 标准odds时的标准评分
                            PDO= PDO,   # odds增加一倍时，分数变化
                            odds_new = odds_new,   # 抽样建模样本的odds：1的数量/0的数量
                            odds_old = odds_old,    # 原始未抽样样本的odds：1的数量/0的数量
                            has_intercept_score = False   # 评分卡是否包含截距项
                            )

#%%
#XGB算法

from matplotlib import pyplot as plt
import toad
import xgboost as xgb
from xgboost import plot_importance
from sklearn import metrics
from sklearn.model_selection import train_test_split
from datetime import datetime
import os 
import warnings
#%%
train_selected2, dropped = toad.selection.select(dt_train[allFeatures], target='target', empty=0.9, iv=0.01, corr=0.7, return_drop=True)

train_selected2.columns
print(train_selected2.shape)#(66721, 29)
train_selected2 = train_selected2.drop(['model_score_01_xysl_1.0','model_score_01br_score','model_score_01_zr_tdzx'],axis=1)
train_selected2=train_selected2.fillna(-9999)

test_selected2=dt_test[train_selected2.columns]
test_selected2=test_selected2.fillna(-9999)

#%%
xgb_model = xgb.XGBClassifier(learning_rate=0.2,
                              n_estimators = 100,
                              objective = "binary:logistic",
                              max_depth = 3,
                              n_jobs = -1,
                              min_child_weight = 1,
                              subsample=1,
                              nthread = 1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_lambda = 300
                              )

# 对训练集进行预测
cols=['model_score_01', 'model_score_01_rong360',
'model_score_01_tengxun', 
'model_score_01_xysl_3',
'model_score_01_fulin', 
'model_score_01_x_tianchuang', 'model_score_01_y_tianchuang',
'value_015_bairong', 'value_021_bairong', 'value_024_bairong',
'value_044_bairong', 'value_069_bairong', 'value_072_bairong',
'value_078_bairong', 'value_084_bairong', 'value_086_bairong',
'value_087_bairong', 'value_088_bairong', 'value_092_bairong',
'value_093_bairong', 'value_094_bairong', 'value_097_bairong',
'crd_loan_gap', 'utl']

xgb_model.fit(train_selected2[cols], train_selected2['target'])
y_pred = xgb_model.predict_proba(train_selected2[cols])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(train_selected2['target'], y_pred, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train KS: ', ks)
print('train AUC: ', roc_auc)
# train KS:  0.2641149621764085
# train AUC:  0.6829567377979494
train KS:  0.23552038925506297
train AUC:  0.6645428392022608
#%%
# 对测试集进行预测
y_pred_test = xgb_model.predict_proba(test_selected2[cols])[:,1]
fpr_test, tpr_test, thresholds_oot = metrics.roc_curve(test_selected2['target'], y_pred_test, pos_label=1)
roc_auc_test = metrics.auc(fpr_test, tpr_test)
ks_test = max(tpr_test-fpr_test)
print('test KS: ', ks_test)
print('test AUC: ', roc_auc_test)
# test KS:  0.17261682070252077
# test AUC:  0.620139007143893
test KS:  0.1730567631165908
test AUC:  0.6198366589939891
#%%
plot_importance(xgb_model,importance_type='gain')
xgb_model.get_booster().get_score(importance_type='gain')

# {'model_score_01': 5.048643112182617,
#  'model_score_01_rong360': 2.9522573947906494,
#  'model_score_01_tengxun': 4.439605236053467,
#  'model_score_01_xysl_3': 13.00546932220459,
#  'model_score_01_fulin': 1.538245439529419,
#  'model_score_01_x_tianchuang': 9.136465072631836,
#  'model_score_01_y_tianchuang': 5.9535136222839355,
#  'value_015_bairong': 3.471034288406372,
#  'value_021_bairong': 3.8786838054656982,
#  'value_024_bairong': 2.9369187355041504,
#  'value_044_bairong': 4.079094409942627,
#  'value_069_bairong': 3.250666379928589,
#  'value_072_bairong': 4.707425117492676,
#  'value_078_bairong': 2.669053792953491,
#  'value_084_bairong': 2.86083984375,
#  'value_086_bairong': 4.049018859863281,
#  'value_087_bairong': 1.911482572555542,
#  'value_088_bairong': 3.1695919036865234,
#  'value_092_bairong': 6.797971725463867,
#  'value_093_bairong': 5.965978145599365,
#  'value_094_bairong': 1.895758032798767,
#  'value_097_bairong': 5.28358268737793,
#  'crd_loan_gap': 5.403532981872559,
#  'utl': 14.493250846862793}
#%%
plt.plot(fpr, tpr, color='darkorange', lw=2, label='trian ROC curve (area = %0.2f)' % roc_auc)
plt.plot(fpr_test, tpr_test, color='blue', lw=2, label='test ROC curve (area = %0.2f)' % roc_auc_test)

plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="best")
plt.show()
#%% 参数调优1
from sklearn.model_selection import GridSearchCV
param_test1 = {'max_depth':[2,3,4,5,6],'n_estimators':[100, 200, 300, 400, 500, 600]}

gsearch = GridSearchCV(
    estimator = xgb_model,
    param_grid=param_test1, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch.fit(train_selected2[cols], train_selected2['target'])

print('gsearch1.best_params_', gsearch.best_params_)
print('gsearch1.best_score_', gsearch.best_score_)
# gsearch1.best_params_ {'max_depth': 3, 'n_estimators': 100}
# gsearch1.best_score_ 0.6243731784342935
#%% 参数调优2
xgb_model_2 = xgb.XGBClassifier(learning_rate=0.2,
                              n_estimators = 100,
                              objective = "binary:logistic",
                              max_depth = 3,
                              n_jobs = -1,
                              min_child_weight = 1,
                              subsample=1,
                              nthread = 1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_lambda = 300
                              )

param_test2 = {'learning_rate':[i/20.0 for i in range(1,20)]}
gsearch2 = GridSearchCV(
    estimator = xgb_model_2,
    param_grid=param_test2, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch2.fit(train_selected2[cols], train_selected2['target'])

print('gsearch2.best_params_', gsearch2.best_params_)
print('gsearch2.best_score_', gsearch2.best_score_)
# gsearch2.best_params_ {'learning_rate': 0.2}
# gsearch2.best_score_ 0.6243731784342935
#%% 参数调优3
xgb_model_3 = xgb.XGBClassifier(learning_rate=0.2,
                              n_estimators = 100,
                              objective = "binary:logistic",
                              max_depth = 3,
                              n_jobs = -1,
                              min_child_weight = 1,
                              subsample=1,
                              nthread = 1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_lambda = 300
)

param_test3 = {'min_child_weight':[i for i in range(1,6,1)]}

gsearch3 = GridSearchCV(
    estimator = xgb_model_3,
    param_grid=param_test3, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch3.fit(train_selected2[cols], train_selected2['target'])

print('gsearch3.best_params_', gsearch3.best_params_)
print('gsearch3.best_score_', gsearch3.best_score_)

# gsearch3.best_params_ {'min_child_weight': 1}
# gsearch3.best_score_ 0.6243731784342935
#%%
# 对训练集进行预测
pred_train = xgb_model.predict_proba(train_selected2[cols])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(train_selected2['target'], pred_train, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train KS: ', ks)
print('train AUC: ', roc_auc)

# 对测试集进行预测
pred_test = xgb_model.predict_proba(test_selected2[cols])[:,1]
fpr_test, tpr_test, thresholds_test = metrics.roc_curve(test_selected2['target'], pred_test, pos_label=1)
roc_auc_test = metrics.auc(fpr_test, tpr_test)
ks_test = max(tpr_test-fpr_test)
print('oot KS: ', ks_test)
print('oot AUC: ', roc_auc_test)

train KS:  0.23552038925506297
train AUC:  0.6645428392022608
oot KS:  0.1730567631165908
oot AUC:  0.6198366589939891
#%%
train_selected2=pd.concat([train_selected2,dt_train[['channel_id_x']]],axis=1)
test_selected2=pd.concat([test_selected2,dt_test[['channel_id_x']]],axis=1)
#%%
train_selected2_167 = train_selected2.query("channel_id_x == 167")
oot_selected_167 = test_selected2.query("channel_id_x == 167")

train_selected2_174 = train_selected2.query("channel_id_x == 174")
oot_selected_174 = test_selected2.query("channel_id_x == 174")
# 167对训练集进行预测
pred_train_167 = xgb_model.predict_proba(train_selected2_167[cols])[:,1]
fpr_167, tpr_167, thresholds_167 = metrics.roc_curve(train_selected2_167['target'], pred_train_167, pos_label=1)
roc_auc_167 = metrics.auc(fpr_167, tpr_167)
ks_167 = max(tpr_167-fpr_167)
print('渠道167的train KS: ', ks_167)
print('渠道167的train AUC: ', roc_auc_167)

# 167对测试集进行预测
pred_oot_167 = xgb_model.predict_proba(oot_selected_167[cols])[:,1]
fpr_167_oot, tpr_167_oot, thresholds_oot_167 = metrics.roc_curve(oot_selected_167['target'], pred_oot_167, pos_label=1)
roc_auc_oot_167 = metrics.auc(fpr_167_oot, tpr_167_oot)
ks_oot_167 = max(tpr_167_oot-fpr_167_oot)
print('渠道167的oot KS: ', ks_oot_167)
print('渠道167的oot AUC: ', roc_auc_oot_167)
渠道167的train KS:  0.23638257164100757
渠道167的train AUC:  0.6651819928212883
渠道167的oot KS:  0.17244125326998305
渠道167的oot AUC:  0.6189879360337953

# 174对训练集进行预测
pred_train_174 = xgb_model.predict_proba(train_selected2_174[cols])[:,1]
fpr_174, tpr_174, thresholds_174 = metrics.roc_curve(train_selected2_174['target'], pred_train_174, pos_label=1)
roc_auc_174 = metrics.auc(fpr_174, tpr_174)
ks_174 = max(tpr_174-fpr_174)
print('渠道174的train KS: ', ks_174)
print('渠道174的train AUC: ', roc_auc_174)

# 174对测试集进行预测
pred_oot_174 = xgb_model.predict_proba(oot_selected_174[cols])[:,1]
fpr_174_oot, tpr_174_oot, thresholds_oot_174 = metrics.roc_curve(oot_selected_174['target'], pred_oot_174, pos_label=1)
roc_auc_oot_174 = metrics.auc(fpr_174_oot, tpr_174_oot)
ks_oot_174 = max(tpr_174_oot-fpr_174_oot)
print('渠道174的oot KS: ', ks_oot_174)
print('渠道174的oot AUC: ', roc_auc_oot_174)

渠道174的train KS:  0.22589700700113513
渠道174的train AUC:  0.6581301401287318
渠道174的oot KS:  0.17141418904757622
渠道174的oot AUC:  0.6145730411090152
#%%
# 业务效果
pred_data = toad.metrics.KS_bucket(pred_train, train_selected2['target'], bucket=10, method='step')
pred_data_test = toad.metrics.KS_bucket(pred_test, test_selected2['target'], bucket=10, method='step')

# 167业务效果
pred_data_167 = toad.metrics.KS_bucket(pred_train_167, train_selected2_167['target'], bucket=10, method='step')
pred_data_oot_167 = toad.metrics.KS_bucket(pred_oot_167, oot_selected_167['target'], bucket=10, method='step')

# 174业务效果
pred_data_174 = toad.metrics.KS_bucket(pred_train_174, train_selected2_174['target'], bucket=10, method='step')
pred_data_oot_174 = toad.metrics.KS_bucket(pred_oot_174, oot_selected_174['target'], bucket=10, method='step')
#%%
train_selected2_167['prob'] = pred_train_167
oot_selected_167['prob'] = pred_oot_167

train_selected2_174['prob'] = pred_train_174
oot_selected_174['prob'] = pred_oot_174

train_selected2['prob'] = pred_train
test_selected2['prob'] = pred_test

#%%
def Prob2Score(prob, basePoint, PDO):
    # 将概率转化成分数且为正整数
    y = np.log(prob / (1 - prob))
    y2 = basePoint + PDO / np.log(2) * (-y)
    score = y2.astype("int")
    return score
#%%

train_selected2['score2'] = train_selected2['prob'].apply(lambda x:Prob2Score(x, 700, 60))
test_selected2['score2'] = test_selected2['prob'].apply(lambda x:Prob2Score(x, 700, 60))
train_selected2_174['score2'] = train_selected2_174['prob'].apply(lambda x:Prob2Score(x, 700, 60))
oot_selected_174['score2'] = oot_selected_174['prob'].apply(lambda x:Prob2Score(x, 700, 60))
train_selected2_167['score2'] = train_selected2_167['prob'].apply(lambda x:Prob2Score(x, 700, 60))
oot_selected_167['score2'] = oot_selected_167['prob'].apply(lambda x:Prob2Score(x, 700, 60))
#%%
train_selected2['score_band2'] =pd.cut(train_selected2['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])
test_selected2['score_band2'] =pd.cut(test_selected2['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])


train_selected2_167['score_band2'] =pd.cut(train_selected2_167['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])
oot_selected_167['score_band2'] =pd.cut(oot_selected_167['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])
train_selected2_174['score_band2'] =pd.cut(train_selected2_174['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])
oot_selected_174['score_band2'] =pd.cut(oot_selected_174['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])
#%%
cal_score1=train_selected2_167.groupby('score_band2')['target'].count().reset_index()
cal_score2=oot_selected_167.groupby('score_band2')['target'].count().reset_index()

cal_score3=train_selected2_167.groupby(['score_band2','target'])['model_score_01_tengxun'].count().unstack().reset_index()
cal_score4=oot_selected_167.groupby(['score_band2','target'])['user_id'].count().unstack().reset_index()


cal_score1=train_selected2_174.groupby('score_band2')['target'].count().reset_index()
cal_score2=oot_selected_174.groupby('score_band2')['target'].count().reset_index()

cal_score3=train_selected2_174.groupby(['score_band2','target'])['model_score_01_tengxun'].count().unstack().reset_index()
cal_score4=oot_selected_174.groupby(['score_band2','target'])['user_id'].count().unstack().reset_index()



cal_score1=train_selected2.groupby('score_band2')['target'].count().reset_index()
cal_score2=test_selected2.groupby('score_band2')['target'].count().reset_index()

cal_score3=train_selected2.groupby(['score_band2','target'])['model_score_01_tengxun'].count().unstack().reset_index()
cal_score4=test_selected2.groupby(['score_band2','target'])['user_id'].count().unstack().reset_index()
#%%
train_final_score=pd.DataFrame()
test_final_score=pd.DataFrame()
train_final_score=pd.concat([train_adj[['score_band1','target']],train_selected2[['channel_id_x','score_band2']]],axis=1)
test_final_score=pd.concat([test_adj[['score_band1','target']],test_selected2[['channel_id_x','score_band2']]],axis=1)

score_dis_train=train_final_score.groupby(['score_band1','score_band2'])['target'].count().unstack().reset_index()
score_dis_train_bad=train_final_score.groupby(['score_band1','score_band2'])['target'].sum().unstack().reset_index()


score_dis_test=test_final_score.groupby(['score_band1','score_band2'])['target'].count().unstack().reset_index()
score_dis_test_bad=test_final_score.groupby(['score_band1','score_band2'])['target'].sum().unstack().reset_index()

#%%
#evaluate_dt_origin=pd.read_csv(r'C:\Users\ruizhi\Desktop\kantt\1.base_data\apply_2means_target.csv')
evaluate_dt_origin=pd.read_csv(r'C:\Users\ruizhi\Desktop\kantt\1.base_data\apply_2means_target_all.csv')
evaluate_dt_origin=evaluate_dt_origin.rename(columns={"model_score_01_xysl_3.0":"model_score_01_xysl_3"})

evaluate_dt=evaluate_dt_origin[evaluate_dt_origin.apply_month.isin(['2022-05', '2022-06', '2022-07','2022-08', '2022-09', '2022-10', '2022-11', '2022-12'])]
evaluate_dt['target']=evaluate_dt_origin['Firs6ever30']/2
use_col_keep=['user_id','target','crd_loan_gap','model_score_01','model_score_01_x_tianchuang','model_score_01_xysl_3',
              'model_score_01_y_tianchuang','utl']
evaluate_dt=evaluate_dt[use_col_keep].fillna(-9999)
evaluate_dt['crd_loan_gap_b']=pd.cut(evaluate_dt.crd_loan_gap,[-10000,0,7,np.inf])
evaluate_dt['model_score_01_b']=pd.cut(evaluate_dt.model_score_01,[-10000,0,690,720,740,770,850])
evaluate_dt['model_score_01_x_tianchuang_b']=pd.cut(evaluate_dt.model_score_01_x_tianchuang,[-10000,0,575,645,850])
evaluate_dt['model_score_01_xysl_3_b']=pd.cut(evaluate_dt['model_score_01_xysl_3'],[-10000,0,600,620,645,670,700,np.inf])
evaluate_dt['model_score_01_y_tianchuang_b']=pd.cut(evaluate_dt.model_score_01_y_tianchuang,[-10000,0,580,850])
evaluate_dt['utl_b']=pd.cut(evaluate_dt.utl,[0,0.5,0.9,np.inf])

evaluate_dt['WEIGHT']=1

evaluate_dt.crd_loan_gap_b.value_counts()
#%%
dt_eva_woe = transer.transform(c.transform(evaluate_dt[model_name1]))

dt_eva_woe['lr_prob']=model_1.predict(dt_eva_woe[model_name1])

dt_eva_woe['score1']=p_to_score(dt_eva_woe['lr_prob'],pdo=60,base=700,odds=1)

dt_eva_woe['score_band1']=pd.cut(dt_eva_woe.score1,[-np.inf,855,870,880,890,900,910,920,935,955,np.inf])

eva_1=pd.concat([evaluate_dt_origin,dt_eva_woe],axis=1)

eva_1['suc']=np.where((eva_1.order_status==6)&(~eva_1.lending_time.isna()),1,0)

cal_dis3=eva_1[eva_1.channel_id_x==167].groupby(['score_band1','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis4=eva_1[(eva_1.suc==1)&(eva_1.channel_id_x==167)].groupby(['score_band1','Firs6ever30'])['order_no_x'].count().unstack().reset_index()

cal_dis5=eva_1[eva_1.channel_id_x==174].groupby(['score_band1','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis6=eva_1[(eva_1.suc==1)&(eva_1.channel_id_x==174)].groupby(['score_band1','Firs6ever30'])['order_no_x'].count().unstack().reset_index()

cal_dis1=eva_1.groupby(['score_band1','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis2=eva_1[(eva_1.suc==1)].groupby(['score_band1','Firs6ever30'])['order_no_x'].count().unstack().reset_index()
#%%
check1=eva_1.tail(20)
eva_1.apply_month.unique()

#%%
#%%
#evaluate_dt_origin=pd.read_csv(r'C:\Users\ruizhi\Desktop\kantt\1.base_data\apply_2means_target.csv')
evaluate_dt_origin=pd.read_csv(r'C:\Users\ruizhi\Desktop\kantt\1.base_data\apply_2means_target_all.csv')
evaluate_dt_origin=evaluate_dt_origin.rename(columns={"model_score_01_xysl_3.0":"model_score_01_xysl_3"})

evaluate_dt=evaluate_dt_origin[evaluate_dt_origin.apply_month.isin(['2022-05', '2022-06', '2022-07','2022-08', '2022-09', '2022-10', '2022-11', '2022-12'])]
evaluate_dt['target']=evaluate_dt_origin['Firs6ever30']/2

cols=['model_score_01', 'model_score_01_rong360','model_score_01_tengxun', 'model_score_01_xysl_3',
   'model_score_01_fulin', 'model_score_01_x_tianchuang','model_score_01_y_tianchuang', 'crd_loan_gap', 'utl']
# 对训练集进行预测
pred_all = xgb_model.predict_proba(evaluate_dt[cols])[:,1]

evaluate_dt['prob2'] = pred_all

#%%
def Prob2Score(prob, basePoint, PDO):
    # 将概率转化成分数且为正整数
    y = np.log(prob / (1 - prob))
    y2 = basePoint + PDO / np.log(2) * (-y)
    score = y2.astype("int")
    return score

evaluate_dt['score2'] = evaluate_dt['prob2'].apply(lambda x:Prob2Score(x, 700, 60))

#%%
evaluate_dt['score_band2'] =pd.cut(evaluate_dt['score2'],[-np.inf,850,870,880,890,905,915,925,940,960,np.inf])

check1=evaluate_dt.head(5)
evaluate_dt['suc']=np.where((evaluate_dt.order_status==6)&(~evaluate_dt.lending_time.isna()),1,0)

cal_dis1=evaluate_dt.groupby(['score_band2','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis2=evaluate_dt[(evaluate_dt.suc==1)].groupby(['score_band2','Firs6ever30'])['order_no_x'].count().unstack().reset_index()

cal_dis3=evaluate_dt[evaluate_dt.channel_id_x==167].groupby(['score_band2','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis4=evaluate_dt[(evaluate_dt.suc==1)&(evaluate_dt.channel_id_x==167)].groupby(['score_band2','Firs6ever30'])['order_no_x'].count().unstack().reset_index()


cal_dis5=evaluate_dt[evaluate_dt.channel_id_x==174].groupby(['score_band2','suc'])['order_no_x'].count().unstack().reset_index()
cal_dis6=evaluate_dt[(evaluate_dt.suc==1)&(evaluate_dt.channel_id_x==174)].groupby(['score_band2','Firs6ever30'])['order_no_x'].count().unstack().reset_index()
#%%
check1=eva_1.tail(20)
eva_1.apply_month.unique()


#==============================================================================
# File: __init__.py
#==============================================================================

#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: __init__.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2022-07-09
'''

from .feature_binning import bin_method_run, chi_bin, dt_bin, equal_freq_bin, kmeans_bin, best_binning
from .detector import detect
from .feature_selection import FeatureSelection, feature_select, stepwise_del_feature
from .report2excel import Report2Excel, var_summary_to_excel
from .statistics import calc_bin_summary, calc_var_summary, compare_inflection_point, get_vif
from .stepwise import stepwise, StatsModel
from .transformer import WoeTransformer, FeatureBin
from .metrics import psi, get_auc_ks_psi, get_auc, get_ks
from .utils import del_df, dump_to_pkl, load_from_pkl, fea_woe_dict_format, to_score
from .logger_utils import Logger
from .scorecard import ScoreCard
from .bayes_opt_tuner import classifiers_model_auto_tune_params
from .plot import plot_var_bin_summary

__version__ = "0.2.0"
VERSION = __version__



#==============================================================================
# File: 三方数据匹配-提现层 (1).py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8




import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import gc

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)





# 运行函数脚本
get_ipython().run_line_magic('run', 'function.ipynb')


# ## 读取三方数据并合并




# 获取167文件下的所有csv文件
file_dir = r'\\192.168.100.120\d\juzi\0904'
file_name_167 = get_filename(file_dir)
len(file_name_167)





len('dwd_beforeloan_third_combine_id_167_')





file_name_167[0][40:-11]





data_source_name = []
for channle in [167, 174, 206, 80004, 80005]:
    print(f'----------------{channle}-------------')
    for i, iterm in enumerate(file_name_167):
        if '_{}_'.format(channle) in iterm:
            channle_lenght = len(f'dwd.dwd_beforeloan_third_combine_id_{channle}_')
            data_source_name.append(iterm[channle_lenght:-11])





len(data_source_name)





data_source_name = list(set(data_source_name))





print(len(data_source_name))
print(data_source_name)





filepath = r'\\192.168.100.120\d\juzi\0904'
def merge_csv_file(data_source_name, file_name_zz, path=r'\\192.168.100.120\d\juzi\0904'):
    """
    data_source_name:数据源名称
    file_name_zz:csv文件列表
    """
    data_list = {}
    for channle in [167, 174, 206, 80004, 80005]:
        for i, filename in enumerate(file_name_zz):
            if '_{}_'.format(channle) in filename:
                channle_lenght = len(f'dwd.dwd_beforeloan_third_combine_id_{channle}_')
                if data_source_name == filename[channle_lenght:-11]: #and filename[-10:-4]<=str(202304)
                    data_list[i] = pd.read_csv(r'{}\{}'.format(path, filename))   
    if data_list:
        merge_df = pd.concat(list(data_list.values()), axis=0)
        return merge_df
    else:
        print('----------{}:无数据-------------'.format(data_source_name))
        return None


# ## 三方数据匹配基础表




# 训练集
df_smaples = pd.read_csv(r'\\192.168.100.120\d\liuyedao\B卡开发\mid_result\B卡_order_target_train_oot_sample_20230914.csv')
df_smaples.shape





# 提现表
usecols = ['order_no', 'user_id','id_no_des', 'channel_id', 'order_status','apply_date','apply_time']
df_order_167 = pd.read_csv(r'\\192.168.100.120\d\juzi\0711\167\dwd_beforeloan_order_examine_fd_167.csv',usecols=usecols)
df_order_other = pd.read_csv(r'\\192.168.100.120\d\juzi\0711\other\dwd_beforeloan_order_examine_fd_other.csv',usecols=usecols)
df_order_80005 = pd.read_csv(r'\\192.168.100.120\d\juzi\0809\dwd_beforeloan_order_examine_fd_80005.csv',usecols=usecols)





df_order_other = df_order_other[df_order_other["channel_id"].isin([167, 174, 206, 80004, 80005])]
df_order = pd.concat([df_order_167, df_order_other, df_order_80005], axis=0)
df_order.shape





print(df_order['order_no'].nunique(),df_order.shape)





df_order_base = pd.merge(df_order, df_smaples[['order_no', 'smaple_set']], how='inner', on='order_no')
print(df_order_base['order_no'].nunique(),df_order_base.shape)





df_order_base["channel_id"].value_counts(dropna=False)





df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\base_data\df_order_base_20230914.csv', index=False)





df_order_base.info()


# # 三方数据匹配




file_name_zz = file_name_167[:]
print(file_name_zz[0:2])
print(len(file_name_zz))


# ### baihang_1




ds = 'baihang_1'
df_baihang_1 = merge_csv_file('baihang_1', file_name_zz)
df_baihang_1.dropna(how='all', axis=1, inplace=True)
df_baihang_1.shape





df_baihang_1['return_massage'].value_counts(dropna=False)





df_baihang_1 = df_baihang_1[df_baihang_1['return_massage']=='请求成功']
print(df_baihang_1['order_no'].nunique(), df_baihang_1.shape)





df_baihang_1['create_time'].str[0:7].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_1, cols_right, needcols=needcols, suffix='baihang_1')
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_4




ds = 'baihang_4'
df_baihang_4 = merge_csv_file(ds, file_name_zz)
df_baihang_4.dropna(how='all', axis=1, inplace=True)
df_baihang_4.shape





df_baihang_4['return_massage'].value_counts(dropna=False)





df_baihang_4 = df_baihang_4[df_baihang_4['return_massage']=='查询成功']
df_baihang_4.shape





df_baihang_4['create_time'].max()





df_baihang_4['channel_id'].value_counts()





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_6




ds = 'baihang_6'
df_baihang_6 = merge_csv_file(ds, file_name_zz)
df_baihang_6.dropna(how='all', axis=1, inplace=True)
df_baihang_6.shape





df_baihang_6.info()





df_baihang_6['return_massage'].value_counts(dropna=False)





df_baihang_6 = df_baihang_6[df_baihang_6['return_massage']=='查询成功']
df_baihang_6.shape





df_baihang_6['create_time'].max()





df_baihang_6.columns[df_baihang_6.columns.str.contains('value_')][13:]





needcols = df_baihang_6.columns[df_baihang_6.columns.str.contains('value_')][13:].to_list()
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_8




ds = 'baihang_8'
df_baihang_8 = merge_csv_file(ds, file_name_zz)
df_baihang_8.dropna(how='all', axis=1, inplace=True)
df_baihang_8.shape





df_baihang_8['return_massage'].value_counts(dropna=False)





df_baihang_8 = df_baihang_8[df_baihang_8['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_baihang_8.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### rong360_4




ds = 'rong360_4'
df_rong360 = merge_csv_file(ds, file_name_zz)
df_rong360.dropna(how='all', axis=1, inplace=True)
df_rong360.shape





df_rong360['return_massage'].value_counts()





df_rong360 = df_rong360[df_rong360['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_rong360.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_rong360, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ronghuijinke_2




ds = 'ronghuijinke_2'
df_ronghuijinke_2 = merge_csv_file(ds, file_name_zz)
df_ronghuijinke_2.dropna(how='all', axis=1, inplace=True)
df_ronghuijinke_2.shape





df_ronghuijinke_2['return_massage'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ronghuijinke_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ronghuijinke_3




ds = 'ronghuijinke_3'
df_ronghuijinke_3 = merge_csv_file(ds, file_name_zz)
df_ronghuijinke_3.dropna(how='all', axis=1, inplace=True)
df_ronghuijinke_3.shape





df_ronghuijinke_3['return_massage'].value_counts(dropna=False)





df_ronghuijinke_3 = df_ronghuijinke_3[df_ronghuijinke_3['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_ronghuijinke_3.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ronghuijinke_3, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tengxun_1




ds = 'tengxun_1'
df_tengxun_1 = merge_csv_file(ds, file_name_zz)
df_tengxun_1.dropna(how='all', axis=1, inplace=True)
df_tengxun_1.shape





print(df_tengxun_1['order_no'].nunique(), df_tengxun_1.shape)





df_tengxun_1.groupby(['return_massage','value_005'],dropna=False)['order_no'].count().unstack()





df_tengxun_1['value_005'].value_counts(dropna=False)





df_tengxun_1['return_massage'].value_counts(dropna=False)





df_tengxun_1 = df_tengxun_1.query("value_005==0.0")
df_tengxun_1.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tengxun_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### xinyongsuanli_1




# 重跑数据20239/9 create_time需要重新处理
ds = 'xinyongsuanli_1'
df_xinyongsuanli_1 = merge_csv_file(ds, file_name_zz)
df_xinyongsuanli_1.dropna(how='all', axis=1, inplace=True)
df_xinyongsuanli_1.shape





df_xinyongsuanli_1.groupby(['return_massage','value_003'],dropna=False)['order_no'].count().unstack()





df_xinyongsuanli_1 = df_xinyongsuanli_1[df_xinyongsuanli_1['return_massage'].isin(['请求成功','处理成功'])]





df_xinyongsuanli_1['value_002'].value_counts(dropna=False)





df_xinyongsuanli_1.reset_index(drop=True,inplace=True)

tmp = pd.get_dummies(df_xinyongsuanli_1['value_002'])
df_xinyongsuanli_1 = pd.concat([df_xinyongsuanli_1, tmp.mul(df_xinyongsuanli_1['model_score_01'], axis=0)], axis=1)





df_xinyongsuanli_1.rename(columns={1.0:'model_score_01_1', 2.0:'model_score_01_2', 3.0:'model_score_01_3', 4.0:'model_score_01_4'},inplace=True)
df_xinyongsuanli_1['create_time'] = df_xinyongsuanli_1['create_time'].str[0:10]
df_xinyongsuanli_1['create_time'].head()





df_xinyongsuanli_1 = df_xinyongsuanli_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_1','model_score_01_2','model_score_01_3','model_score_01_4'].max()
df_xinyongsuanli_1 = df_xinyongsuanli_1.reset_index()





df_xinyongsuanli_1.info()





needcols = ['model_score_01_1','model_score_01_2','model_score_01_3','model_score_01_4']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_xinyongsuanli_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





del df_xinyongsuanli_1
gc.collect()


# ### bairong_1




file_name_bairong = get_filename(r'\\192.168.100.120\d\juzi\0907')
file_name_bairong


# #### 不含json串的字段处理




usecols1 = ['value_00' + str(i) for i in range(1,10)]
usecols2 = ['value_0'  + str(i) for i in range(10,100) if i!=89]
usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id'] + usecols1 + usecols2





df_bairong_1_list = {}
for filename in file_name_bairong[0:-4]:
    tmp_name = filename[-11:-4]
    tmp_df = pd.read_csv(r'\\192.168.100.120\d\juzi\0907\{}'.format(filename), usecols=usecols)
    tmp_df = tmp_df[tmp_df['id_no_des'].isin(df_order_base['id_no_des'])]
    df_bairong_1_list[tmp_name] = tmp_df
    
#     needcols = usecols1 + usecols2
#     cols_left = ['order_no','id_no_des','apply_date']
#     cols_right = ['order_no','id_no_des','create_time'] + needcols

#     # 返回匹配的三方数据
#     df_bairong_1_list[tmp_name] = process_data(df_order_base, cols_left, df_bairong_1_tmp, cols_right, needcols=needcols, suffix='bairong_1')






# 运行函数脚本
get_ipython().run_line_magic('run', 'function.ipynb')





needcols =  usecols1 + usecols2
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

for tmp_key, tmp_df in df_bairong_1_list.items():
    print(tmp_key)
    # 返回匹配的三方数据
    df_three = process_data_bairong(df_order_base, cols_left, tmp_df, cols_right, needcols=needcols, suffix='bairong_1')
    df_bairong_1_list[tmp_key] = df_three





df_bairong_1_no_json = pd.concat(list(df_bairong_1_list.values()), axis=0)





df_bairong_1_no_json.info()
df_bairong_1_no_json.head()





df_bairong_1_no_json = df_bairong_1_no_json.sort_values(by=['order_no','order_no_is_equal_bairong_1','create_time_bairong_1'], ascending=False)
df_bairong_1_no_json = df_bairong_1_no_json.drop_duplicates(subset=['order_no'],keep='first')
df_bairong_1_no_json.info()





df_bairong_1_no_json.shape





df_bairong_1_no_json.to_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_{}_nojson_{}.csv'.format('bairong_1',str(datetime.today())[:10].replace('-','')), index=False)





df_bairong_1_no_json.head()





print(df_bairong_1_no_json.columns[df_bairong_1_no_json.columns.str.contains('value')].to_list())





needcols =  df_bairong_1_no_json.columns[df_bairong_1_no_json.columns.str.contains('value')].to_list()
cols_right = ['order_no'] + needcols

# 匹配关联
df_order_base = pd.merge(df_order_base, df_bairong_1_no_json[cols_right], how='left',on='order_no')
print(df_order_base.shape)


# #### 只含json串字段处理




# 运行函数脚本
get_ipython().run_line_magic('run', 'function.ipynb')





import time
import json

start_time = time.time()
print(start_time)

usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_089']
needcols =  ['value_089']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

df_bairong_1_list = {}
for filename in file_name_bairong[0:-4]:
    filepath = r'\\192.168.100.120\d\juzi\0907\{}'.format(filename)
    tmp_name = filename[-11:-4]
    print(tmp_name)
    
    tmp_df = chunk_process_data(filepath, usecols, df_order_base, cols_left, cols_right, needcols, suffix='bairong_1', chunk_size=50000)
    tmp_df = tmp_df.reset_index()
    tmp_df['value_089_bairong_1'] = tmp_df['value_089_bairong_1'].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
    tmp_df = pd.concat([tmp_df.drop(['value_089_bairong_1'], axis=1), tmp_df['value_089_bairong_1'].apply(pd.Series)], axis=1)
    df_bairong_1_list[tmp_name] = tmp_df
    
    del tmp_df
    gc.collect()
    end_time = time.time()
    print(end_time)
    total_time = end_time - start_time
    print(total_time/60)
    






df_bairong_1_json = pd.concat(list(df_bairong_1_list.values()), axis=0)





df_bairong_1_json.info()
df_bairong_1_json.head()





df_bairong_1_json = df_bairong_1_json.sort_values(by=['order_no','order_no_is_equal_bairong_1','create_time_bairong_1'], ascending=False).drop_duplicates(subset=['order_no'],keep='first')
df_bairong_1_json.to_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_bairong_1_json_20230914.csv',index=False)





print(df_bairong_1_json.order_no.nunique(), df_bairong_1_json.shape)





needcols =  df_bairong_1_json.columns[df_bairong_1_json.columns.str.contains('als|bank')].to_list()
cols_right = ['order_no'] + needcols

# 匹配关联
df_order_base = pd.merge(df_order_base, df_bairong_1_json[cols_right], how='left',on='order_no')
print(df_order_base.shape)





# ds = 'bairong_1'
# df_bairong_1 = merge_csv_file(ds, file_name_zz)
# df_bairong_1.dropna(how='all', axis=1, inplace=True)
# df_bairong_1.shape


# ### bairong_8




ds = 'bairong_8'
df_bairong_8 = merge_csv_file(ds, file_name_zz)
df_bairong_8.dropna(how='all', axis=1, inplace=True)
df_bairong_8.shape





df_bairong_8['return_massage'].value_counts(dropna=False)





df_bairong_8 = df_bairong_8[df_bairong_8['return_massage']=='请求成功']





df_bairong_8['value_001'].value_counts(dropna=False)





df_bairong_8['value_002'].value_counts(dropna=False)





df_bairong_8.reset_index(drop=True,inplace=True)

tmp = pd.get_dummies(df_bairong_8['value_002'])
df_bairong_8 = pd.concat([df_bairong_8, tmp.mul(df_bairong_8['model_score_01'], axis=0)], axis=1)
df_bairong_8.rename(columns={'ScoreCust2':'model_score_01_2', 'ScoreCust3':'model_score_01_3','ScoreCust7':'model_score_01_7', 'ScoreCust8':'model_score_01_8'},inplace=True)

df_bairong_8['create_time'] = df_bairong_8['create_time'].str[0:10]
df_bairong_8.head()





df_bairong_8 = df_bairong_8.groupby(['order_no','id_no_des','create_time'])['model_score_01_2' ,'model_score_01_3','model_score_01_7' ,'model_score_01_8'].max()
df_bairong_8 = df_bairong_8.reset_index()





needcols = ['model_score_01_2' ,'model_score_01_3','model_score_01_7' ,'model_score_01_8']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_12




ds = 'bairong_12'
df_bairong_12 = merge_csv_file(ds, file_name_zz)
df_bairong_12.dropna(how='all', axis=1, inplace=True)
df_bairong_12.shape





df_bairong_12['value_001'].value_counts(dropna=False)





df_bairong_12 = df_bairong_12[df_bairong_12['value_001']==0.0]
df_bairong_12.shape





print(df_bairong_12.columns[df_bairong_12.columns.str.contains("value_")].to_list()[4:])





needcols = df_bairong_12.columns[df_bairong_12.columns.str.contains("value_")].to_list()[4:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_12, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_13




ds = 'bairong_13'
df_bairong_13 = merge_csv_file(ds, file_name_zz)
df_bairong_13.dropna(how='all', axis=1, inplace=True)
df_bairong_13.shape





df_bairong_13['return_massage'].value_counts(dropna=False)





df_bairong_13['value_012'].value_counts(dropna=False)





df_bairong_13['value_011'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_13, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_14




ds = 'bairong_14'
df_bairong_14 = merge_csv_file(ds, file_name_zz)
df_bairong_14.dropna(how='all',axis=1, inplace=True)
df_bairong_14.shape





df_bairong_14['return_massage'].value_counts(dropna=False)





df_bairong_14['value_012'].value_counts(dropna=False)





df_bairong_14['value_011'].value_counts(dropna=False)





needcols = list(df_bairong_14.select_dtypes(include='number').columns)
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_14, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_15




ds = 'bairong_15'
df_bairong_15 = merge_csv_file(ds, file_name_zz)
df_bairong_15.dropna(how='all', axis=1, inplace=True)
df_bairong_15.shape





df_bairong_15['return_massage'].value_counts(dropna=False)





df_bairong_15['value_012'].value_counts(dropna=False)





df_bairong_15['value_011'].value_counts(dropna=False)





print(list(df_bairong_15.select_dtypes(include='number').columns))





needcols = ['model_score_01', 'value_014', 'value_016', 'value_017', 'value_018', 'value_020', 'value_022', 'value_023']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_15, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_16




ds = 'bairong_16'
df_bairong_16 = merge_csv_file(ds, file_name_zz)
df_bairong_16.dropna(how='all', axis=1, inplace=True)
df_bairong_16.shape





df_bairong_16['return_massage'].value_counts(dropna=False)





df_bairong_16['value_012'].value_counts(dropna=False)





df_bairong_16 = df_bairong_16[df_bairong_16['value_012']==0.0]
df_bairong_16.shape





print(df_bairong_14.columns[df_bairong_14.columns.str.contains('value_')].to_list()[11:19])





needcols = df_bairong_14.columns[df_bairong_14.columns.str.contains('value_')].to_list()[11:19]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_16, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





df_order_base.info()


# ### bairong_tag1




ds = 'bairong_tag1'
df_bairong_tag1 = merge_csv_file(ds, file_name_zz)
df_bairong_tag1.dropna(how='all',axis=1, inplace=True)
df_bairong_tag1.shape


# ### bairong_tag2




ds = 'bairong_tag2'
df_bairong_tag2 = merge_csv_file(ds, file_name_zz)
df_bairong_tag2.dropna(how='all',axis=1, inplace=True)
df_bairong_tag2.shape


# ### tianchuang_1




ds = 'tianchuang_1'
df_tianchuang_1 = merge_csv_file(ds, file_name_zz)
df_tianchuang_1.dropna(how='all', axis=1, inplace=True)
df_tianchuang_1.shape





df_tianchuang_1['return_massage'].value_counts(dropna=False)





df_tianchuang_1['value_004'].value_counts(dropna=False)





df_tianchuang_1 = df_tianchuang_1[df_tianchuang_1['return_massage']=='请求成功']
print(df_tianchuang_1['order_no'].nunique(), df_tianchuang_1.shape)





df_tianchuang_1['value_001'].value_counts(dropna=False)





tmp = pd.get_dummies(df_tianchuang_1['value_001'])
df_tianchuang_1 = pd.concat([df_tianchuang_1, tmp.mul(df_tianchuang_1['model_score_01'], axis=0)], axis=1)





df_tianchuang_1.rename(columns={'LBMQ150101':'model_score_01_q', 'LBMR150101':'model_score_01_r'},inplace=True)
df_tianchuang_1['create_time'] = df_tianchuang_1['create_time'].str[0:10]
df_tianchuang_1['create_time'].head()





df_tianchuang_1 = df_tianchuang_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_q','model_score_01_r'].max()
df_tianchuang_1 = df_tianchuang_1.reset_index()





df_tianchuang_1.info()





needcols = ['model_score_01_q', 'model_score_01_r']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





del df_tianchuang_1,df_bairong_8
gc.collect()


# ### tianchuang_4




ds = 'tianchuang_4'
df_tianchuang_4 = merge_csv_file(ds, file_name_zz)
df_tianchuang_4.dropna(how='all', axis=1, inplace=True)
df_tianchuang_4.shape





df_tianchuang_4['value_011'].value_counts(dropna=False)





df_tianchuang_4['value_012'].value_counts(dropna=False)





df_tianchuang_4 = df_tianchuang_4[df_tianchuang_4['value_011']=='响应成功']
df_tianchuang_4.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tianchuang_7




ds = 'tianchuang_7'
df_tianchuang_7 = merge_csv_file(ds, file_name_zz)
df_tianchuang_7.dropna(how='all', axis=1, inplace=True)
df_tianchuang_7.shape





df_tianchuang_7['return_massage'].value_counts(dropna=False)





df_tianchuang_7['value_011'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_7, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tianchuang_8




ds = 'tianchuang_8'
df_tianchuang_8 = merge_csv_file(ds, file_name_zz)
df_tianchuang_8.dropna(how='all', axis=1, inplace=True)
df_tianchuang_8.shape





df_tianchuang_8['return_massage'].value_counts(dropna=False)





df_tianchuang_8['value_011'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### fulin_1




ds = 'fulin_1'
df_fulin_1 = merge_csv_file(ds, file_name_zz)
df_fulin_1.dropna(how='all', axis=1, inplace=True)
df_fulin_1.shape





df_fulin_1['return_massage'].value_counts(dropna=False).head()





df_fulin_1['return_code'].value_counts(dropna=False)





df_fulin_1['create_time'].min()





df_fulin_1 = df_fulin_1[df_fulin_1['return_massage'].isin(['success','查询成功'])]





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_fulin_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tongcheng_tag




ds = 'tongcheng_tag'
df_tongcheng_tag = merge_csv_file(ds, file_name_zz)
df_tongcheng_tag.dropna(how='all', axis=1, inplace=True)
df_tongcheng_tag.shape


# ### tongliantong_3




ds = 'tongliantong_3'
df_tongliantong_3 = merge_csv_file(ds, file_name_zz)
df_tongliantong_3.dropna(how='all', axis=1, inplace=True)
df_tongliantong_3.shape


# ### tongdun_2




ds = 'tongdun_2'
df_tongdun_2 = merge_csv_file(ds, file_name_zz)
df_tongdun_2.dropna(how='all', axis=1, inplace=True)
df_tongdun_2.shape





df_tongdun_2['value_005'].value_counts(dropna=False)





df_tongdun_2 = df_tongdun_2[df_tongdun_2['value_005'].isin([0.0])]





df_tongdun_2['value_001'].value_counts(dropna=False)





df_tongdun_2['value_002'].value_counts(dropna=False)





tmp = pd.get_dummies(df_tongdun_2['value_002'])
df_tongdun_2 = pd.concat([df_tongdun_2, tmp.mul(df_tongdun_2['model_score_01'], axis=0)], axis=1)
df_tongdun_2.rename(columns={'27984162d86f3baf':'model_score_01_zr', '3a9de0313a0c65ee':'model_score_01_zx'},inplace=True)
df_tongdun_2['create_time'] = df_tongdun_2['create_time'].str[0:10]
df_tongdun_2['create_time'].head()





df_tongdun_2 = df_tongdun_2.groupby(['order_no','id_no_des','create_time'])['model_score_01_zr','model_score_01_zx'].max()
df_tongdun_2 = df_tongdun_2.reset_index()





needcols = ['model_score_01_zr', 'model_score_01_zx']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tongdun_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





del df_tongdun_2
gc.collect()


# ### talkingdata_7




ds = 'talkingdata_7'
df_talkingdata_7 = merge_csv_file(ds, file_name_zz)
df_talkingdata_7.dropna(how='all',axis=1, inplace=True)
df_talkingdata_7.shape





df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_base_three_data_nopudao_3.csv',index=False)





df_order_base = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_base_three_data_nopudao_3.csv')
df_order_base.info()


# ### pudao_3




file_name_pudao_3 = get_filename(r'\\192.168.100.120\d\juzi\0912')
file_name_pudao_3





import time
import json

start_time = time.time()
print(start_time)

usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_012']
needcols =  ['value_012']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

df_pudao_3_list = {}
for filename in file_name_pudao_3[0:-4]:
    filepath = r'\\192.168.100.120\d\juzi\0912\{}'.format(filename)
    tmp_name = filename[-11:-4]
    print(tmp_name)
    
    tmp_df = chunk_process_data(filepath, usecols, df_order_base, cols_left,
                                cols_right, needcols, suffix='pudao_3', chunk_size=50000)
    tmp_df = tmp_df.reset_index()
    tmp_df['value_012_pudao_3'] = tmp_df['value_012_pudao_3'].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
    tmp_df = pd.concat([tmp_df.drop(['value_012_pudao_3'], axis=1), tmp_df['value_012_pudao_3'].apply(pd.Series)], axis=1)
    df_pudao_3_list[tmp_name] = tmp_df
    
    del tmp_df
    gc.collect()
    end_time = time.time()
    print(end_time)
    total_time = end_time - start_time
    print(total_time/60)
    





df_pudao_3 = pd.concat(list(df_pudao_3_list.values()), axis=0)
df_pudao_3 = df_pudao_3.sort_values(by=['order_no','order_no_is_equal_pudao_3','create_time_pudao_3'], ascending=False).drop_duplicates(subset=['order_no'],keep='first')
df_pudao_3.to_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_pudao_3.csv',index=False)





df_pudao_3.info()
df_pudao_3.head()





# 匹配关联
df_pudao_3 = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_pudao_3.csv')
needcols =  df_pudao_3.columns[df_pudao_3.columns.str.contains('ppdi|bank')].to_list()
cols_right = ['order_no'] + needcols

df_order_base = pd.merge(df_order_base, df_pudao_3[cols_right], how='left',on='order_no')
print(df_order_base.shape)





df_order_base.info()





ds = 'pudao_3'
df_pudao_3 = merge_csv_file(ds, file_name_zz)
df_pudao_3.dropna(how='all', axis=1, inplace=True)
df_pudao_3.shape





df_pudao_3['value_011'].value_counts(dropna=False)





df_pudao_3 = df_pudao_3[df_pudao_3['value_011']==0.0]
df_pudao_3 = df_pudao_3.reset_index(drop=True)





df_pudao_3['create_time'].max()





df_pudao_3[df_pudao_3['create_time']<'2023-04-01'].shape





df_pudao_3 = df_pudao_3[df_pudao_3['create_time']<'2023-05-01']
df_pudao_3 = df_pudao_3.reset_index(drop=True)





df_pudao_3.info(show_counts=True)





df_pudao_3.head()





import json
import ast
del df_pudao_3
gc.collect()
# df_pudao_3['value_012'] = df_pudao_3['value_012'].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
# [json.loads(x) if pd.notnull(x) else {} for x in df_pudao_3['value_012'].to_list()]
# df_pudao_3['value_012'] = df_pudao_3['value_012'].apply(lambda x: ast.literal_eval(x) if pd.notnull(x) else {})





df_pudao_3.info()





df_pudao_3 = pd.concat([df_pudao_3.drop(['value_012'], axis=1), df_pudao_3['value_012'].apply(pd.Series)], axis=1)
df_pudao_3.shape





df_pudao_3.head(1)





len(list(df_pudao_3.columns[df_pudao_3.columns.str.contains("ppdi|bank")]))





needcols = list(df_pudao_3.columns[df_pudao_3.columns.str.contains("ppdi|bank")])
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_3, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_4




ds = 'pudao_4'
df_pudao_4 = merge_csv_file(ds, file_name_zz)
df_pudao_4.dropna(how='all', axis=1, inplace=True)
df_pudao_4.shape





df_pudao_4['value_011'].value_counts(dropna=False)





df_pudao_4['value_010'].value_counts(dropna=False)





df_pudao_4['create_time'].min()





# df_pudao_4 = df_pudao_4[df_pudao_4['value_011']==0.0]





print(list(df_pudao_4.columns[df_pudao_4.columns.str.contains("value")])[9:])





needcols = list(df_pudao_4.columns[df_pudao_4.columns.str.contains("value")])[9:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_5




ds = 'pudao_5'
df_pudao_5 = merge_csv_file(ds, file_name_zz)
df_pudao_5.dropna(how='all', axis=1, inplace=True)
df_pudao_5.shape





df_pudao_5['value_014'].value_counts(dropna=False)





df_pudao_5['value_013'].value_counts(dropna=False)





df_pudao_5 = df_pudao_5[df_pudao_5['value_014']==0.0]





needcols = ['value_012']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_5, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_8




ds = 'pudao_8'
df_pudao_8 = merge_csv_file(ds, file_name_zz)
df_pudao_8.dropna(how='all', axis=1, inplace=True)
df_pudao_8.shape





df_pudao_8['value_012'].value_counts(dropna=False)





df_pudao_8['value_011'].value_counts(dropna=False)





df_pudao_8 = df_pudao_8[df_pudao_8['value_012']==0.0]
df_pudao_8.shape





df_pudao_8.model_score_01.head()





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_11




ds = 'pudao_11'
df_pudao_11 = merge_csv_file(ds, file_name_zz)
df_pudao_11.dropna(how='all', axis=1, inplace=True)
df_pudao_11.shape





df_pudao_11['return_massage'].value_counts(dropna=False)





df_pudao_11['value_011'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_11, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_12




ds = 'pudao_12'
df_pudao_12 = merge_csv_file(ds, file_name_zz)
df_pudao_12.dropna(how='all', axis=1, inplace=True)
df_pudao_12.shape





df_pudao_12['value_012'].value_counts(dropna=False)





df_pudao_12['return_massage'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_12, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_13 




ds = 'pudao_13'
df_pudao_13 = merge_csv_file(ds, file_name_zz)
df_pudao_13.dropna(how='all', axis=1, inplace=True)
df_pudao_13.shape





df_pudao_13['value_012'].value_counts(dropna=False)





df_pudao_13['return_massage'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_13, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_14 




ds = 'pudao_14'
df_pudao_14 = merge_csv_file(ds, file_name_zz)
df_pudao_14.dropna(how='all', axis=1, inplace=True)
df_pudao_14.shape


# ### pudao_15




ds = 'pudao_15'
df_pudao_15 = merge_csv_file(ds, file_name_zz)
df_pudao_15.dropna(how='all', axis=1, inplace=True)
df_pudao_15.shape





df_pudao_15['return_massage'].value_counts(dropna=False)





df_pudao_15['value_012'].value_counts(dropna=False)





df_pudao_15['value_011'].value_counts(dropna=False)





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_15, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_16




ds = 'pudao_16'
df_pudao_16 = merge_csv_file(ds, file_name_zz)
df_pudao_16.dropna(how='all', axis=1, inplace=True)
df_pudao_16.shape





df_pudao_16['return_massage'].value_counts(dropna=False)





df_pudao_16['value_012'].value_counts(dropna=False)





df_pudao_16['value_011'].value_counts(dropna=False)





df_pudao_16 = df_pudao_16[df_pudao_16['return_massage']=='调用成功']
df_pudao_16.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_16, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





df_pudao_16['create_time'].min()


# ### ruizhi_4




ds = 'ruizhi_4'
df_ruizhi_4 = merge_csv_file(ds, file_name_zz)
df_ruizhi_4.dropna(how='all', axis=1, inplace=True)
df_ruizhi_4.shape





df_ruizhi_4['return_massage'].value_counts(dropna=False)





df_ruizhi_4['value_011'].value_counts(dropna=False)





df_ruizhi_4['value_012'].value_counts(dropna=False)





df_ruizhi_4 = df_ruizhi_4[df_ruizhi_4['return_massage']=='通过']
df_ruizhi_4.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ruizhi_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ruizhi_6




ds = 'ruizhi_6'
df_ruizhi_6 = merge_csv_file(ds, file_name_zz)
df_ruizhi_6.dropna(how='all', axis=1, inplace=True)
df_ruizhi_6.shape





df_ruizhi_6['return_massage'].value_counts(dropna=False)





df_ruizhi_6['value_012'].value_counts(dropna=False)





df_ruizhi_6['create_time'].min()





df_ruizhi_6 = df_ruizhi_6[df_ruizhi_6['return_massage']=='调用成功']
df_ruizhi_6.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ruizhi_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ruizhi_5




ds = 'ruizhi_5'
df_ruizhi_5 = merge_csv_file(ds, file_name_zz)
df_ruizhi_5.dropna(how='all', axis=1, inplace=True)
df_ruizhi_5.shape





df_ruizhi_5['create_time'].min()


# ### bileizhen_1




ds = 'bileizhen_1'
df_bileizhen_1 = merge_csv_file(ds, file_name_zz)
df_bileizhen_1.dropna(how='all', axis=1, inplace=True)
df_bileizhen_1.shape





df_bileizhen_1['return_massage'].value_counts(dropna=False)





df_bileizhen_1['value_002'].value_counts(dropna=False)





df_bileizhen_1['value_001'].value_counts(dropna=False)





df_bileizhen_1 = df_bileizhen_1.query("value_002==0.0")
df_bileizhen_1.shape





df_bileizhen_1['create_time'].min()





tmp = pd.get_dummies(df_bileizhen_1['value_001'])
df_bileizhen_1 = pd.concat([df_bileizhen_1, tmp.mul(df_bileizhen_1['model_score_01'], axis=0)], axis=1)
df_bileizhen_1.rename(columns={1.0:'model_score_01_1', 2.0:'model_score_01_2', 3.0:'model_score_01_3'},inplace=True)





df_bileizhen_1['create_time'] = df_bileizhen_1['create_time'].str[0:10]
df_bileizhen_1['create_time'].head()

df_bileizhen_1 = df_bileizhen_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_1','model_score_01_2','model_score_01_3'].max()
df_bileizhen_1 = df_bileizhen_1.reset_index()
df_bileizhen_1.shape





df_bileizhen_1.info()





needcols = ['model_score_01_1', 'model_score_01_2', 'model_score_01_3']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bileizhen_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





df_order_base.columns[df_order_base.columns.str.contains('bileizhen')]





del df_bileizhen_1
gc.collect()


# ### duxiaoman_1




ds = 'duxiaoman_1'
df_duxiaoman_1 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_1.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_1.shape





df_duxiaoman_1['create_time'].min()





df_duxiaoman_1['return_massage'].value_counts(dropna=False)





df_duxiaoman_1['value_003'].value_counts(dropna=False)





df_duxiaoman_1['value_001'].value_counts(dropna=False)





df_duxiaoman_1['value_002'].value_counts(dropna=False)





df_duxiaoman_1 = df_duxiaoman_1[df_duxiaoman_1['value_003']==0.0]





df_duxiaoman_1['value_002'].value_counts(dropna=False)





df_duxiaoman_1.groupby(['value_001','value_002'],dropna=False)['order_no'].count().unstack()





df_duxiaoman_1.reset_index(drop=True, inplace=True)
tmp = pd.get_dummies(df_duxiaoman_1['value_002'])
df_duxiaoman_1 = pd.concat([df_duxiaoman_1, tmp.mul(df_duxiaoman_1['model_score_01'], axis=0)], axis=1)
df_duxiaoman_1.rename(columns={'dxm_zzyzv1':'model_score_01_v1', 'dxm_zzyzv2':'model_score_01_v2'},inplace=True)
df_duxiaoman_1['create_time'] = df_duxiaoman_1['create_time'].str[0:10]
df_duxiaoman_1['create_time'].head()

df_duxiaoman_1 = df_duxiaoman_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_v1','model_score_01_v2'].max()
df_duxiaoman_1 = df_duxiaoman_1.reset_index()





needcols = ['model_score_01_v1','model_score_01_v2']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_duxiaoman_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





del df_duxiaoman_1
gc.collect()


# ### duxiaoman_5




ds = 'duxiaoman_5'
df_duxiaoman_5 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_5.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_5.shape


# ### duxiaoman_6




ds = 'duxiaoman_6'
df_duxiaoman_6 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_6.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_6.shape





df_duxiaoman_6['return_massage'].value_counts(dropna=False)





df_duxiaoman_6['value_017'].value_counts(dropna=False)





df_duxiaoman_6['create_time'].min()





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_duxiaoman_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# # hangliezhi_1




ds = 'hangliezhi_1'
df_hangliezhi_1 = merge_csv_file(ds, file_name_zz)
df_hangliezhi_1.dropna(how='all', axis=1, inplace=True)
df_hangliezhi_1.shape





df_hangliezhi_1['return_massage'].value_counts(dropna=False)





df_hangliezhi_1 = df_hangliezhi_1[df_hangliezhi_1['return_massage']=='查询成功']
df_hangliezhi_1.shape





df_hangliezhi_1['create_time'].min()





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hangliezhi_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### hengpu_4




ds = 'hengpu_4'
df_hengpu_4 = merge_csv_file(ds, file_name_zz)
df_hengpu_4.dropna(how='all', axis=1, inplace=True)
df_hengpu_4.shape





df_hengpu_4['value_012'].value_counts(dropna=False)





df_hengpu_4['return_massage'].value_counts(dropna=False)





df_hengpu_4['value_011'].value_counts(dropna=False)





df_hengpu_4 = df_hengpu_4[df_hengpu_4['value_012']==0.0]
df_hengpu_4.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols
# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hengpu_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### huanbei_tag




ds = 'huanbei_tag'
df_huanbei_tag = merge_csv_file(ds, file_name_zz)
df_huanbei_tag.dropna(how='all', axis=1, inplace=True)
df_huanbei_tag.shape


# ### hulian_4




ds = 'hulian_4'
df_hulian_4 = merge_csv_file(ds, file_name_zz)
df_hulian_4.dropna(how='all', axis=1, inplace=True)
df_hulian_4.shape





df_hulian_4['return_massage'].value_counts(dropna=False)





df_hulian_4['value_001'].value_counts(dropna=False)





print(df_hulian_4['order_no'].nunique(), df_hulian_4.shape)





tmp = pd.get_dummies(df_hulian_4['value_001'])
df_hulian_4 = pd.concat([df_hulian_4[['order_no','id_no_des','create_time']], tmp.mul(df_hulian_4['model_score_01'], axis=0)], axis=1)





df_hulian_4.info(show_counts=True)





df_hulian_4.rename(columns={100:'model_score_01_100', 101:'model_score_01_101', 103:'model_score_01_103'
                           ,105:'model_score_01_105', 107:'model_score_01_107', 121:'model_score_01_121'
                           ,122:'model_score_01_122', 124:'model_score_01_124'} ,inplace=True)





df_hulian_4['create_time'] = df_hulian_4['create_time'].str[0:10]
df_hulian_4['create_time'].head(2)





df_hulian_4 = df_hulian_4.groupby(['order_no','id_no_des','create_time'])[df_hulian_4.columns[3:].to_list()].max()
df_hulian_4 = df_hulian_4.reset_index()
df_hulian_4.shape





needcols = df_hulian_4.columns.to_list()[3:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hulian_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





df_order_base.columns[df_order_base.columns.str.contains('hulian_4')]





del df_hulian_4
gc.collect()


# ### hulian_5 




df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\df_order_base_three_data_nohulian5.csv',index=False)





gc.collect()





ds = 'hulian_5'
df_hulian_5 = merge_csv_file(ds, file_name_zz)
df_hulian_5.dropna(how='all', axis=1, inplace=True)
df_hulian_5.shape





df_hulian_5['return_massage'].value_counts(dropna=False)





df_hulian_5['value_002'].value_counts(dropna=False)





df_hulian_5['value_003'].value_counts(dropna=False)





df_hulian_5['value_004'].value_counts(dropna=False)





print(df_hulian_5['order_no'].nunique(), df_hulian_5.shape)





df_hulian_5['create_time'].min()





df_hulian_5 = df_hulian_5[df_hulian_5['return_massage']=='操作成功'].query("create_time<'2023-06-01'")





df_hulian_5['create_time'].max()





df_order_base.apply_date.max()





tmp = pd.get_dummies(df_hulian_5['value_002'].apply(lambda x: int(x)))
df_hulian_5 = pd.concat([df_hulian_5[['order_no','id_no_des','create_time']], tmp.mul(df_hulian_5['model_score_01'], axis=0)], axis=1)





df_hulian_5.info(show_counts=True)





df_hulian_5.rename(columns={100:'model_score_01_100', 101:'model_score_01_101', 103:'model_score_01_103'
                           ,105:'model_score_01_105', 107:'model_score_01_107', 121:'model_score_01_121'
                           ,122:'model_score_01_122', 124:'model_score_01_124', 125:'model_score_01_125'
                           ,127:'model_score_01_127'} ,inplace=True)





df_hulian_5['create_time'] = df_hulian_5['create_time'].str[0:10]
df_hulian_5['create_time'].head(2)





df_hulian_5 = df_hulian_5.groupby(['order_no','id_no_des','create_time'])[df_hulian_5.columns[3:].to_list()].max()
df_hulian_5 = df_hulian_5.reset_index()
df_hulian_5.shape





needcols = df_hulian_5.columns.to_list()[3:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hulian_5, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)





del df_hulian_5
gc.collect()





df_order_base = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\df_order_base_three_data_nohulian5.csv')





print(df_order_base.shape)





# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
df_order_base.info()
print(df_order_base.shape)





df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\df_order_base_three_data_20230915.csv',index=False)


# ### moxingfen_27 




ds = 'moxingfen_27'
df_moxingfen_27 = merge_csv_file(ds, file_name_zz)
df_moxingfen_27.dropna(how='all', axis=1, inplace=True)
df_moxingfen_27.shape





df_moxingfen_27.create_time.max()


# ### moxingfen_7




ds = 'moxingfen_7'
df_moxingfen_7 = merge_csv_file(ds, file_name_zz)
df_moxingfen_7.dropna(how='all', axis=1, inplace=True)
df_moxingfen_7.shape





print(df_moxingfen_7['create_time'].min(), df_moxingfen_7['create_time'].max())





tmp = df_moxingfen_7.columns[df_moxingfen_7.columns.str.contains('value')].to_list()
print(tmp[0:15] , tmp[23:])





needcols = ['model_score_01'] + tmp[0:15] + tmp[23:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_moxingfen_7, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### my_1




ds = 'my_1'
df_my_1 = merge_csv_file(ds, file_name_zz)
df_my_1.dropna(how='all', axis=1, inplace=True)
df_my_1.shape





df_my_1['return_massage'].value_counts(dropna=False)





df_my_1['value_011'].value_counts(dropna=False)





df_my_1['value_012'].value_counts(dropna=False)





df_my_1 = df_my_1.query("value_012==0.0")





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_my_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### aliyun_2




ds = 'aliyun_2'
df_aliyun_2 = merge_csv_file(ds, file_name_zz)
df_aliyun_2.dropna(how='all', axis=1, inplace=True)
df_aliyun_2.shape





df_aliyun_2['value_003'].value_counts(dropna=False)





df_aliyun_2['return_massage'].value_counts(dropna=False)





df_aliyun_2['value_001'].value_counts(dropna=False)





df_aliyun_2 = df_aliyun_2[df_aliyun_2['return_massage'].isin(['调用成功,有效数据', '请求成功','OK'])]
df_aliyun_2.shape





needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_aliyun_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### 94ai_1




ds = '94ai_1'
df_94ai_1 = merge_csv_file(ds, file_name_zz)
df_94ai_1.dropna(how='all', axis=1, inplace=True)
df_94ai_1.shape





df_94ai_1.create_time.min()


# ### face_1




ds = 'face_1'
df_face_1 = merge_csv_file(ds, file_name_zz)
df_face_1.dropna(how='all', axis=1, inplace=True)
df_face_1.shape





df_face_1.info()
df_face_1.head()





df_face_1['return_massage'].value_counts(dropna=False)





df_face_1['value_011'].value_counts(dropna=False)





df_face_1['value_010'].value_counts(dropna=False)





df_face_1.create_time.min()


# ### vivo_1




ds = 'vivo_1'
df_vivo_1 = merge_csv_file(ds, file_name_zz)
df_vivo_1.dropna(how='all', axis=1, inplace=True)
df_vivo_1.shape


# ### vivo_tag1




ds = 'vivo_tag1'
df_vivo_tag1 = merge_csv_file(ds, file_name_zz)
df_vivo_tag1.dropna(how='all', axis=1, inplace=True)
df_vivo_tag1.shape


# ### vivo_tag2




ds = 'vivo_tag2'
df_vivo_tag2 = merge_csv_file(ds, file_name_zz)
df_vivo_tag2.dropna(how='all', axis=1, inplace=True)
df_vivo_tag2.shape


# ### xl_1




ds = 'xl_1'
df_xl_1 = merge_csv_file(ds, file_name_zz)
df_xl_1.dropna(how='all', axis=1, inplace=True)
df_xl_1.shape


# ### yinlian_1




ds = 'yinlian_1'
df_yinlian_1 = merge_csv_file(ds, file_name_zz)
df_yinlian_1.dropna(how='all', axis=1, inplace=True)
df_yinlian_1.shape





print(df_yinlian_1.create_time.min(), df_yinlian_1.create_time.max())


# ### zhixin_1




ds = 'zhixin_1'
df_zhixin_1 = merge_csv_file(ds, file_name_zz)
df_zhixin_1.dropna(how='all', axis=1, inplace=True)
df_zhixin_1.shape


# ## 保存三方匹配数据




df_order_base.info()





# 1. 无pudao_3,bairong_1,含9月8日跑的数据，第二次重启后跑数
df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v1.csv', index=False)





df_order_base.shape

# 2. 无pudao_3,bairong_1,含9月8日重跑的数据和第二次重启后跑数
df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v2.csv', index=False)





# 增加9月8日匹配的数据

## 获取9月8日匹配的数据
files_0908 = get_filename(r'C:\Users\ruizhi\Documents\lxl\mid_result')
files_0908 = [i for i in files_0908 if i[-12:-4]=='20230908']
print(len(files_0908))
print(files_0908)





tmp_df_order_base = df_order_base[['order_no']]
for file in files_0908:
    tmp_df = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\{}'.format(file))
    tmp_df_order_base = pd.merge(tmp_df_order_base, tmp_df, how='left', on='order_no')





tmp_df_order_base.info()





tmp_df_order_base.columns[tmp_df_order_base.columns.str.contains('value|model|score')]





tmp_df_order_base = tmp_df_order_base[['order_no']+tmp_df_order_base.columns[tmp_df_order_base.columns.str.contains('value|model|score')].to_list()]
tmp_df_order_base.info()





tmp_df_order_base.select_dtypes(include='object')['model_score_01_baihang_8'].value_counts()





df_order_base_0908 = pd.merge(df_order_base, tmp_df_order_base, how='inner', on='order_no')
df_order_base_0908.info()





df_order_base_0908.dropna(how='all', axis=1, inplace=True)
df_order_base_0908.shape





# 3. 无pudao_3,bairong_1
df_order_base_0908.to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v3.csv', index=False)





df_order_base['order_status'].value_counts()





# 4. 无pudao_3,部分bairong_1
df_order_base_0908.to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v4.csv', index=False)





# 5. 无pudao_3,
df_order_base.to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v5.csv', index=False)





df_order_base['smaple_set'].value_counts()





df_order_base.query("smaple_set=='train'").to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v5_train.csv', index=False)
df_order_base.query("smaple_set=='oot'").to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v5_oot.csv', index=False)


# ## 三方数据匹配率




# 读取数据
df_order_base = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v4.csv')
df_bairong_1_json = pd.read_csv(r'C:\Users\ruizhi\Documents\lxl\mid_result\order_bairong_1_json_20230909.csv')





df_order_base.info()





df_bairong_1_json.info()





print(df_order_base.shape, df_bairong_1_json.shape)
df_bairong_1_json.drop(['apply_date_bairong_1','order_no_y_bairong_1','create_time_bairong_1','order_no_is_equal_bairong_1'],axis=1,inplace=True)
df_order_base = pd.merge(df_order_base, df_bairong_1_json, how='left', on='order_no')
print(df_order_base.shape)





df_order_base.info()





df_order_base.select_dtypes(include='object').columns





df_order_base[['value_015_moxingfen_7', 'swift_number', 'orderNo', 'name', 'mobileEncrypt', 'idCardEncrypt']].head()





df_order_base['value_015_moxingfen_7'].value_counts(dropna=False)





df_order_base.drop(['value_015_moxingfen_7', 'swift_number', 'orderNo', 'name', 'mobileEncrypt', 'idCardEncrypt'],axis=1,inplace=True)
df_order_base.shape





# 需要不计算匹配率的字段
df_order_base.columns[0:8]





# 按渠道统计各字段的匹配情况
df_order_base['channel_id'].value_counts()





import toad 

to_drop = list(df_order_base.columns[0:8])
to_keep = list(df_order_base.columns[8:])
tmp_total = toad.detect(df_order_base.drop(to_drop, axis=1))





df_order_base.insert(7, 'apply_month', df_order_base['apply_date'].str[0:7])





to_drop = list(df_order_base.columns[0:9])
print(to_drop)
to_keep = list(df_order_base.columns[9:])





# 按申请年月统计匹配情况
year_month = df_order_base.groupby(by=['apply_month'])[to_keep].apply(lambda x: x.isna().sum()/x.shape[0])
year_month = year_month.T
year_month['avg_missing_rate'] = year_month.mean(axis=1)
year_month['avg_std_rate'] = year_month.std(axis=1)





# 按数据集类型统计匹配情况
smaple_set = df_order_base.groupby(by=['smaple_set'])[to_keep].apply(lambda x: x.isna().sum()/x.shape[0])
smaple_set = smaple_set.T
smaple_set['avg_missing_rate'] = year_month.mean(axis=1)
smaple_set['avg_std_rate'] = year_month.std(axis=1)





# 保存统计的数据
writer=pd.ExcelWriter(r'C:\Users\ruizhi\Documents\lxl\统计结果\B卡_三方数据匹配_'+str(datetime.today())[:10].replace('-','')+'.xlsx')
tmp_total.to_excel(writer,sheet_name='总体')

year_month.to_excel(writer,sheet_name='总体_年月')
smaple_set.to_excel(writer,sheet_name='总体_样本类型')

writer.save()



# 需要计算匹配率的字段
cols = ['order_no']+list(df_auth_base.columns[df_auth_base.columns.str.contains('order_no_y|model_score|value')])
cols


# In[13]:


# 按渠道统计各字段的匹配情况
channel = df_auth_base.groupby(by=['channel_id','auth_status'])[cols].count().unstack()
channel = channel.append(pd.DataFrame(channel.sum(axis=0)).T)
channel  


# In[14]:


# 通过与不通过的进行加总
for col in cols:
    channel.loc[:,(col, 0)] = channel.loc[:,(col, 6)] + channel.loc[:,(col, 7)]


# In[18]:


# 计算匹配率
channel_pct = channel.copy()
for col in cols:
    channel_pct.loc[:,(col, 0)] = channel_pct.loc[:,(col, 0)] / channel.loc[:,('order_no', 0)]
    channel_pct.loc[:,(col, 6)] = channel_pct.loc[:,(col, 6)] / channel.loc[:,('order_no', 6)]
    channel_pct.loc[:,(col, 7)] = channel_pct.loc[:,(col, 7)] / channel.loc[:,('order_no', 7)]
channel_pct


# In[46]:


df_auth_base['apply_month'] = df_auth_base['apply_date'].str[0:7]


# In[47]:


# 按申请年月统计匹配情况
year_month = df_auth_base.groupby(by=['apply_month','auth_status'])[cols].count().unstack()
year_month = year_month.append(pd.DataFrame(year_month.sum(axis=0)).T)
year_month


# In[48]:


# 通过与不通过的进行加总
for col in cols:
    year_month.loc[:,(col, 0)] = year_month.loc[:,(col, 6)] + year_month.loc[:,(col, 7)]


# In[49]:


# 计算匹配率
year_month_pct = year_month.copy()
for col in cols:
    year_month_pct.loc[:,(col, 0)] = year_month_pct.loc[:,(col, 0)] / year_month.loc[:,('order_no', 0)]
    year_month_pct.loc[:,(col, 6)] = year_month_pct.loc[:,(col, 6)] / year_month.loc[:,('order_no', 6)]
    year_month_pct.loc[:,(col, 7)] = year_month_pct.loc[:,(col, 7)] / year_month.loc[:,('order_no', 7)]
year_month_pct


# In[20]:


# 保存统计的数据
writer=pd.ExcelWriter(r"d:\liuyedao\result\三方数据匹配_授信_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
channel.to_excel(writer,sheet_name='渠道')
channel_pct.to_excel(writer,sheet_name='渠道比例')

# year_month.to_excel(writer,sheet_name='申请年月')
# year_month_pct.to_excel(writer,sheet_name='申请年月比例')

writer.save()




#==============================================================================
# File: 三方数据匹配-提现层.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[2]:


import numpy as np
import pandas as pd
from datetime import datetime
import re
from IPython.core.interactiveshell import InteractiveShell
import warnings
import gc

warnings.filterwarnings("ignore")
InteractiveShell.ast_node_interactivity = 'all'
pd.set_option('display.max_row',None)
pd.set_option('display.width',1000)


# In[3]:


# 运行函数脚本
get_ipython().run_line_magic('run', 'function_三方数据.ipynb')


# ## 读取三方数据并合并

# In[3]:


# 获取167文件下的所有csv文件
file_dir = r'\\192.168.100.120\d\juzi\0904'
file_name = get_filename(file_dir)
len(file_name)


# In[4]:


len('dwd_beforeloan_third_combine_id_167_')


# In[5]:


file_name[0][40:-11]


# In[6]:


data_source_name = []
for channle in [174, 80004, 80005]:
# for channle in [167, 174, 206, 80004, 80005]:
    print(f'----------------{channle}-------------')
    for i, iterm in enumerate(file_name):
        if '_{}_'.format(channle) in iterm:
            channle_lenght = len(f'dwd.dwd_beforeloan_third_combine_id_{channle}_')
            data_source_name.append(iterm[channle_lenght:-11])


# In[7]:


len(data_source_name)


# In[8]:


data_source_name = list(set(data_source_name))


# In[9]:


print(len(data_source_name))
print(data_source_name)


# In[10]:


filepath = r'\\192.168.100.120\d\juzi\0904'
def merge_csv_file(data_source_name, file_name_zz, path=r'\\192.168.100.120\d\juzi\0904'):
    """
    data_source_name:数据源名称
    file_name_zz:csv文件列表
    """
    data_list = {}
    for channle in [174, 80004, 80005]:
        for i, filename in enumerate(file_name_zz):
            if '_{}_'.format(channle) in filename:
                channle_lenght = len(f'dwd.dwd_beforeloan_third_combine_id_{channle}_')
                if data_source_name == filename[channle_lenght:-11]: #and filename[-10:-4]<=str(202304)
                    data_list[i] = pd.read_csv(r'{}\{}'.format(path, filename))   
    if data_list:
        merge_df = pd.concat(list(data_list.values()), axis=0)
        return merge_df
    else:
        print('----------{}:无数据-------------'.format(data_source_name))
        return None


# ## 三方数据匹配基础表

# In[13]:


# 训练集
df_smaples = pd.read_csv(r'\\192.168.100.120\d\liuyedao\B卡开发\mid_result\B卡_order_target_mob3_train_oot_20230927.csv')
df_smaples.shape


# In[14]:


df_smaples.info()


# In[ ]:


# # 提现表
# usecols = ['order_no', 'user_id','id_no_des', 'channel_id', 'order_status','apply_date','apply_time']
# df_order_167 = pd.read_csv(r'\\192.168.100.120\d\juzi\0711\167\dwd_beforeloan_order_examine_fd_167.csv',usecols=usecols)
# df_order_other = pd.read_csv(r'\\192.168.100.120\d\juzi\0711\other\dwd_beforeloan_order_examine_fd_other.csv',usecols=usecols)
# df_order_80005 = pd.read_csv(r'\\192.168.100.120\d\juzi\0809\dwd_beforeloan_order_examine_fd_80005.csv',usecols=usecols)


# In[ ]:


# df_order_other = df_order_other[df_order_other["channel_id"].isin([167, 174, 206, 80004, 80005])]
# df_order = pd.concat([df_order_167, df_order_other, df_order_80005], axis=0)
# df_order.shape


# In[ ]:


# print(df_order['order_no'].nunique(),df_order.shape)


# In[15]:


# df_order_base = pd.merge(df_order, df_smaples[['order_no','lending_time','lending_month','target']], how='inner', on='order_no')
df_order_base = df_smaples[['order_no','user_id','id_no_des','channel_id','lending_time','lending_month','target']]
print(df_order_base['order_no'].nunique(),df_order_base.shape)


# In[16]:


df_order_base["channel_id"].value_counts(dropna=False)


# In[ ]:


# df_order_base.to_csv(r'D:\liuyedao\B卡开发\mid_result\df_order_base_20230919.csv', index=False)


# In[21]:


df_order_base.info()
df_order_base.head()


# In[22]:


df_order_base.rename(columns={'lending_time':'apply_date'}, inplace=True)


# # 三方数据匹配

# In[23]:


file_name_zz = file_name[:]
print(file_name_zz[0:2])
print(len(file_name_zz))


# ### baihang_1

# In[18]:


ds = 'baihang_1'
df_baihang_1 = merge_csv_file('baihang_1', file_name_zz)
df_baihang_1.dropna(how='all', axis=1, inplace=True)
df_baihang_1.shape


# In[ ]:


df_baihang_1['return_massage'].value_counts(dropna=False)


# In[19]:


df_baihang_1 = df_baihang_1[df_baihang_1['return_massage']=='请求成功']
print(df_baihang_1['order_no'].nunique(), df_baihang_1.shape)


# In[ ]:


df_baihang_1['create_time'].str[0:7].value_counts(dropna=False)


# In[25]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_1, cols_right, needcols=needcols, suffix='baihang_1')
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_4

# In[26]:


ds = 'baihang_4'
df_baihang_4 = merge_csv_file(ds, file_name_zz)
df_baihang_4.dropna(how='all', axis=1, inplace=True)
df_baihang_4.shape


# In[27]:


df_baihang_4['return_massage'].value_counts(dropna=False)


# In[28]:


df_baihang_4 = df_baihang_4[df_baihang_4['return_massage']=='查询成功']
df_baihang_4.shape


# In[ ]:


df_baihang_4['create_time'].max()


# In[ ]:


df_baihang_4['channel_id'].value_counts()


# In[29]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_6

# In[30]:


ds = 'baihang_6'
df_baihang_6 = merge_csv_file(ds, file_name_zz)
df_baihang_6.dropna(how='all', axis=1, inplace=True)
df_baihang_6.shape


# In[ ]:


df_baihang_6.info()


# In[31]:


df_baihang_6['return_massage'].value_counts(dropna=False)


# In[32]:


df_baihang_6 = df_baihang_6[df_baihang_6['return_massage']=='查询成功']
df_baihang_6.shape


# In[ ]:


df_baihang_6['create_time'].max()


# In[33]:


df_baihang_6.columns[df_baihang_6.columns.str.contains('value_')][13:]


# In[34]:


needcols = df_baihang_6.columns[df_baihang_6.columns.str.contains('value_')][13:].to_list()
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### baihang_8

# In[35]:


ds = 'baihang_8'
df_baihang_8 = merge_csv_file(ds, file_name_zz)
df_baihang_8.dropna(how='all', axis=1, inplace=True)
df_baihang_8.shape


# In[36]:


df_baihang_8['return_massage'].value_counts(dropna=False)


# In[37]:


df_baihang_8 = df_baihang_8[df_baihang_8['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_baihang_8.shape


# In[38]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_baihang_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### rong360_4

# In[39]:


ds = 'rong360_4'
df_rong360 = merge_csv_file(ds, file_name_zz)
df_rong360.dropna(how='all', axis=1, inplace=True)
df_rong360.shape


# In[40]:


df_rong360['return_massage'].value_counts()


# In[41]:


df_rong360 = df_rong360[df_rong360['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_rong360.shape


# In[42]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_rong360, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ronghuijinke_2

# In[43]:


ds = 'ronghuijinke_2'
df_ronghuijinke_2 = merge_csv_file(ds, file_name_zz)
df_ronghuijinke_2.dropna(how='all', axis=1, inplace=True)
df_ronghuijinke_2.shape


# In[44]:


df_ronghuijinke_2['return_massage'].value_counts(dropna=False)


# In[45]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ronghuijinke_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ronghuijinke_3

# In[46]:


ds = 'ronghuijinke_3'
df_ronghuijinke_3 = merge_csv_file(ds, file_name_zz)
df_ronghuijinke_3.dropna(how='all', axis=1, inplace=True)
df_ronghuijinke_3.shape


# In[47]:


df_ronghuijinke_3['return_massage'].value_counts(dropna=False)


# In[48]:


df_ronghuijinke_3 = df_ronghuijinke_3[df_ronghuijinke_3['return_massage'].isin(['请求成功','查询成功', 'OK'])]
df_ronghuijinke_3.shape


# In[49]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ronghuijinke_3, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tengxun_1

# In[50]:


ds = 'tengxun_1'
df_tengxun_1 = merge_csv_file(ds, file_name_zz)
df_tengxun_1.dropna(how='all', axis=1, inplace=True)
df_tengxun_1.shape


# In[ ]:


print(df_tengxun_1['order_no'].nunique(), df_tengxun_1.shape)


# In[ ]:


df_tengxun_1.groupby(['return_massage','value_005'],dropna=False)['order_no'].count().unstack()


# In[51]:


df_tengxun_1['value_005'].value_counts(dropna=False)


# In[52]:


df_tengxun_1['return_massage'].value_counts(dropna=False)


# In[53]:


df_tengxun_1 = df_tengxun_1.query("value_005==0.0")
df_tengxun_1.shape


# In[54]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tengxun_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### xinyongsuanli_1

# In[55]:


ds = 'xinyongsuanli_1'
df_xinyongsuanli_1 = merge_csv_file(ds, file_name_zz)
df_xinyongsuanli_1.dropna(how='all', axis=1, inplace=True)
df_xinyongsuanli_1.shape


# In[56]:


df_xinyongsuanli_1.groupby(['return_massage','value_003'],dropna=False)['order_no'].count().unstack()


# In[57]:


df_xinyongsuanli_1 = df_xinyongsuanli_1[df_xinyongsuanli_1['return_massage'].isin(['请求成功','处理成功'])]


# In[58]:


df_xinyongsuanli_1['value_002'].value_counts(dropna=False)


# In[59]:


df_xinyongsuanli_1.reset_index(drop=True,inplace=True)

tmp = pd.get_dummies(df_xinyongsuanli_1['value_002'])
df_xinyongsuanli_1 = pd.concat([df_xinyongsuanli_1, tmp.mul(df_xinyongsuanli_1['model_score_01'], axis=0)], axis=1)


# In[60]:


df_xinyongsuanli_1.rename(columns={1.0:'model_score_01_1', 2.0:'model_score_01_2', 3.0:'model_score_01_3', 4.0:'model_score_01_4'},inplace=True)
df_xinyongsuanli_1['create_time'] = df_xinyongsuanli_1['create_time'].str[0:10]
df_xinyongsuanli_1['create_time'].head()


# In[61]:


df_xinyongsuanli_1 = df_xinyongsuanli_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_1','model_score_01_2','model_score_01_3','model_score_01_4'].max()
df_xinyongsuanli_1 = df_xinyongsuanli_1.reset_index()


# In[ ]:


df_xinyongsuanli_1.info()


# In[62]:


needcols = ['model_score_01_1','model_score_01_2','model_score_01_3','model_score_01_4']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_xinyongsuanli_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[63]:


del df_xinyongsuanli_1
gc.collect()


# ### bairong_1

# In[268]:


file_name_bairong = get_filename(r'\\192.168.100.120\d\juzi\0907')
file_name_bairong


# #### 不含json串的字段处理

# In[269]:


usecols1 = ['value_00' + str(i) for i in range(1,10)]
usecols2 = ['value_0'  + str(i) for i in range(10,100) if i!=89]
usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id'] + usecols1 + usecols2


# In[270]:


needcols =  usecols1 + usecols2
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

df_bairong_1_list = {}
# for filename in file_name_bairong:
#     tmp_name = filename[-11:-4]
#     tmp_df = pd.read_csv(r'\\192.168.100.120\d\juzi\0907\{}'.format(filename), usecols=usecols)
#     tmp_df = tmp_df[tmp_df['id_no_des'].isin(df_order_base['id_no_des'])]
#     df_bairong_1_list[tmp_name] = tmp_df

#     # 返回匹配的三方数据
#     df_three = process_data_bairong(df_order_base, cols_left, tmp_df, cols_right, needcols=needcols, suffix='bairong_1')
#     df_bairong_1_list[tmp_key] = df_three    
# for filename in file_name_bairong:
#     filepath = r'\\192.168.100.120\d\juzi\0907\{}'.format(filename)
#     tmp_name = filename[-11:-4]
#     print(tmp_name)
    
#     tmp_df = chunk_process_data(filepath, usecols, df_order_base, cols_left,
#                                 cols_right, needcols, suffix='bairong_1', chunk_size=50000)
#     tmp_df = tmp_df.reset_index()
#     df_bairong_1_list[tmp_name] = tmp_df
    
#     del tmp_df
#     gc.collect()
#     end_time = time.time()
#     print(end_time)
#     total_time = end_time - start_time
#     print(total_time/60)
   


# In[271]:


needcols =  usecols1 + usecols2
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

for tmp_key, tmp_df in df_bairong_1_list.items():
    print(tmp_key)
    # 返回匹配的三方数据
    df_three = process_data_bairong(df_order_base, cols_left, tmp_df, cols_right, needcols=needcols, suffix='bairong_1')
    df_bairong_1_list[tmp_key] = df_three


# In[272]:


df_bairong_1_no_json = pd.concat(list(df_bairong_1_list.values()), axis=0)


# In[273]:


df_bairong_1_no_json.info()
df_bairong_1_no_json.head()


# In[274]:


df_bairong_1_no_json = df_bairong_1_no_json.sort_values(by=['order_no','order_no_is_equal_bairong_1','create_time_bairong_1'], ascending=False)
df_bairong_1_no_json = df_bairong_1_no_json.drop_duplicates(subset=['order_no'],keep='first')
# df_bairong_1_no_json.info()


# In[275]:


df_bairong_1_no_json.shape


# In[276]:


df_bairong_1_no_json.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_{}_nojson_{}.csv'.format('bairong_1',str(datetime.today())[:10].replace('-','')), index=False)


# In[277]:


df_bairong_1_no_json.head(2)


# In[278]:


print(df_bairong_1_no_json.columns[df_bairong_1_no_json.columns.str.contains('value')].to_list())


# In[279]:


needcols =  df_bairong_1_no_json.columns[df_bairong_1_no_json.columns.str.contains('value')].to_list()
cols_right = ['order_no'] + needcols

# 匹配关联
df_order_base = pd.merge(df_order_base, df_bairong_1_no_json[cols_right], how='left',on='order_no')
print(df_order_base.shape)


# #### 只含json串字段处理

# In[280]:


import time
import json

start_time = time.time()
print(start_time)

usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_089']
needcols =  ['value_089']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

df_bairong_1_list = {}
for filename in file_name_bairong:
    filepath = r'\\192.168.100.120\d\juzi\0907\{}'.format(filename)
    tmp_name = filename[-11:-4]
    print(tmp_name)
    
    tmp_df = chunk_process_data(filepath, usecols, df_order_base, cols_left, cols_right, needcols, suffix='bairong_1', chunk_size=50000)
    tmp_df = tmp_df.reset_index()
    tmp_df['value_089_bairong_1'] = tmp_df['value_089_bairong_1'].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
    tmp_df = pd.concat([tmp_df.drop(['value_089_bairong_1'], axis=1), tmp_df['value_089_bairong_1'].apply(pd.Series)], axis=1)
    df_bairong_1_list[tmp_name] = tmp_df
    
    del tmp_df
    gc.collect()
    end_time = time.time()
    print(end_time)
    total_time = end_time - start_time
    print(total_time/60)


# In[281]:


df_order_base.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_no_bairong_1_json_20230927.csv',index=False)


# In[282]:


df_bairong_1_json = pd.concat(list(df_bairong_1_list.values()), axis=0)


# In[283]:


df_bairong_1_json.info()
df_bairong_1_json.head()


# In[285]:


# df_bairong_1_json = df_bairong_1_json.sort_values(by=['order_no','order_no_is_equal_bairong_1','create_time_bairong_1'], ascending=False).drop_duplicates(subset=['order_no'],keep='first')
df_bairong_1_json.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_bairong_1_json_20230927.csv',index=False)


# In[286]:


print(df_bairong_1_json.order_no.nunique(), df_bairong_1_json.shape)


# In[287]:


df_bairong_1_json.head()


# In[289]:


print(df_bairong_1_json.columns[6:-2].to_list())


# In[290]:


print(len(df_bairong_1_json.columns[6:-2].to_list()))


# In[288]:


needcols =  df_bairong_1_json.columns[df_bairong_1_json.columns.str.contains('als|bank')].to_list()
print(len(needcols))


# In[291]:


needcols =  df_bairong_1_json.columns[6:-2].to_list()
print(len(needcols))
cols_right = ['order_no'] + needcols

# 匹配关联
df_order_base = pd.merge(df_order_base, df_bairong_1_json[cols_right], how='left',on='order_no')
print(df_order_base.shape)


# In[292]:


df_order_base.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_20230927.csv',index=False)


# In[ ]:


# ds = 'bairong_1'
# df_bairong_1 = merge_csv_file(ds, file_name_zz)
# df_bairong_1.dropna(how='all', axis=1, inplace=True)
# df_bairong_1.shape


# ### bairong_8

# In[234]:


ds = 'bairong_8'
df_bairong_8 = merge_csv_file(ds, file_name_zz)
df_bairong_8.dropna(how='all', axis=1, inplace=True)
df_bairong_8.shape


# In[235]:


df_bairong_8['return_massage'].value_counts(dropna=False)


# In[236]:


df_bairong_8 = df_bairong_8[df_bairong_8['return_massage']=='请求成功']


# In[237]:


df_bairong_8['value_001'].value_counts(dropna=False)


# In[238]:


df_bairong_8['value_002'].value_counts(dropna=False)


# In[239]:


df_bairong_8.reset_index(drop=True,inplace=True)

tmp = pd.get_dummies(df_bairong_8['value_002'])
df_bairong_8 = pd.concat([df_bairong_8, tmp.mul(df_bairong_8['model_score_01'], axis=0)], axis=1)
df_bairong_8.rename(columns={'ScoreCust2':'model_score_01_2', 'ScoreCust3':'model_score_01_3','ScoreCust7':'model_score_01_7', 'ScoreCust8':'model_score_01_8'},inplace=True)

df_bairong_8['create_time'] = df_bairong_8['create_time'].str[0:10]
df_bairong_8.head()


# In[240]:


df_bairong_8 = df_bairong_8.groupby(['order_no','id_no_des','create_time'])['model_score_01_2' ,'model_score_01_3','model_score_01_7' ,'model_score_01_8'].max()
df_bairong_8 = df_bairong_8.reset_index()


# In[241]:


needcols = ['model_score_01_2' ,'model_score_01_3','model_score_01_7' ,'model_score_01_8']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_12

# In[242]:


ds = 'bairong_12'
df_bairong_12 = merge_csv_file(ds, file_name_zz)
df_bairong_12.dropna(how='all', axis=1, inplace=True)
df_bairong_12.shape


# In[243]:


df_bairong_12['value_001'].value_counts(dropna=False)


# In[244]:


df_bairong_12 = df_bairong_12[df_bairong_12['value_001']==0.0]
df_bairong_12.shape


# In[245]:


print(df_bairong_12.columns[df_bairong_12.columns.str.contains("value_")].to_list()[4:])


# In[246]:


needcols = df_bairong_12.columns[df_bairong_12.columns.str.contains("value_")].to_list()[4:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_12, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_13

# In[247]:


ds = 'bairong_13'
df_bairong_13 = merge_csv_file(ds, file_name_zz)
df_bairong_13.dropna(how='all', axis=1, inplace=True)
df_bairong_13.shape


# In[248]:


df_bairong_13['return_massage'].value_counts(dropna=False)


# In[ ]:


df_bairong_13['value_012'].value_counts(dropna=False)


# In[ ]:


df_bairong_13['value_011'].value_counts(dropna=False)


# In[249]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_13, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_14

# In[250]:


ds = 'bairong_14'
df_bairong_14 = merge_csv_file(ds, file_name_zz)
df_bairong_14.dropna(how='all',axis=1, inplace=True)
df_bairong_14.shape


# In[251]:


df_bairong_14['return_massage'].value_counts(dropna=False)


# In[252]:


df_bairong_14['value_012'].value_counts(dropna=False)


# In[253]:


df_bairong_14['value_011'].value_counts(dropna=False)


# In[254]:


needcols = list(df_bairong_14.select_dtypes(include='number').columns)
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_14, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_15

# In[255]:


ds = 'bairong_15'
df_bairong_15 = merge_csv_file(ds, file_name_zz)
df_bairong_15.dropna(how='all', axis=1, inplace=True)
df_bairong_15.shape


# In[256]:


df_bairong_15['return_massage'].value_counts(dropna=False)


# In[257]:


df_bairong_15['value_012'].value_counts(dropna=False)


# In[258]:


df_bairong_15['value_011'].value_counts(dropna=False)


# In[ ]:


print(list(df_bairong_15.select_dtypes(include='number').columns))


# In[259]:


needcols = ['model_score_01', 'value_014', 'value_016', 'value_017', 'value_018', 'value_020', 'value_022', 'value_023']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_15, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_16

# In[260]:


ds = 'bairong_16'
df_bairong_16 = merge_csv_file(ds, file_name_zz)
df_bairong_16.dropna(how='all', axis=1, inplace=True)
df_bairong_16.shape


# In[261]:


df_bairong_16['return_massage'].value_counts(dropna=False)


# In[262]:


df_bairong_16['value_012'].value_counts(dropna=False)


# In[263]:


df_bairong_16 = df_bairong_16[df_bairong_16['value_012']==0.0]
df_bairong_16.shape


# In[264]:


print(df_bairong_14.columns[df_bairong_14.columns.str.contains('value_')].to_list()[11:19])


# In[265]:


needcols = df_bairong_14.columns[df_bairong_14.columns.str.contains('value_')].to_list()[11:19]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bairong_16, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### bairong_tag1

# In[266]:


ds = 'bairong_tag1'
df_bairong_tag1 = merge_csv_file(ds, file_name_zz)
df_bairong_tag1.dropna(how='all',axis=1, inplace=True)
df_bairong_tag1.shape


# ### bairong_tag2

# In[267]:


ds = 'bairong_tag2'
df_bairong_tag2 = merge_csv_file(ds, file_name_zz)
df_bairong_tag2.dropna(how='all',axis=1, inplace=True)
df_bairong_tag2.shape


# ### tianchuang_1

# In[64]:


ds = 'tianchuang_1'
df_tianchuang_1 = merge_csv_file(ds, file_name_zz)
df_tianchuang_1.dropna(how='all', axis=1, inplace=True)
df_tianchuang_1.shape


# In[65]:


df_tianchuang_1['return_massage'].value_counts(dropna=False)


# In[66]:


df_tianchuang_1['value_004'].value_counts(dropna=False)


# In[67]:


df_tianchuang_1 = df_tianchuang_1[df_tianchuang_1['return_massage']=='请求成功']
print(df_tianchuang_1['order_no'].nunique(), df_tianchuang_1.shape)


# In[68]:


df_tianchuang_1['value_001'].value_counts(dropna=False)


# In[69]:


tmp = pd.get_dummies(df_tianchuang_1['value_001'])
df_tianchuang_1 = pd.concat([df_tianchuang_1, tmp.mul(df_tianchuang_1['model_score_01'], axis=0)], axis=1)


# In[70]:


df_tianchuang_1.rename(columns={'LBMQ150101':'model_score_01_q', 'LBMR150101':'model_score_01_r'},inplace=True)
df_tianchuang_1['create_time'] = df_tianchuang_1['create_time'].str[0:10]
df_tianchuang_1['create_time'].head()


# In[71]:


df_tianchuang_1 = df_tianchuang_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_q','model_score_01_r'].max()
df_tianchuang_1 = df_tianchuang_1.reset_index()


# In[ ]:


df_tianchuang_1.info()


# In[72]:


needcols = ['model_score_01_q', 'model_score_01_r']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[73]:


del df_tianchuang_1
gc.collect()


# ### tianchuang_4

# In[74]:


ds = 'tianchuang_4'
df_tianchuang_4 = merge_csv_file(ds, file_name_zz)
df_tianchuang_4.dropna(how='all', axis=1, inplace=True)
df_tianchuang_4.shape


# In[75]:


df_tianchuang_4['value_011'].value_counts(dropna=False)


# In[ ]:


df_tianchuang_4['value_012'].value_counts(dropna=False)


# In[76]:


df_tianchuang_4 = df_tianchuang_4[df_tianchuang_4['value_011']=='响应成功']
df_tianchuang_4.shape


# In[77]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tianchuang_7

# In[78]:


ds = 'tianchuang_7'
df_tianchuang_7 = merge_csv_file(ds, file_name_zz)
df_tianchuang_7.dropna(how='all', axis=1, inplace=True)
df_tianchuang_7.shape


# In[79]:


df_tianchuang_7['return_massage'].value_counts(dropna=False)


# In[ ]:


df_tianchuang_7['value_011'].value_counts(dropna=False)


# In[80]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_7, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tianchuang_8

# In[81]:


ds = 'tianchuang_8'
df_tianchuang_8 = merge_csv_file(ds, file_name_zz)
df_tianchuang_8.dropna(how='all', axis=1, inplace=True)
df_tianchuang_8.shape


# In[82]:


df_tianchuang_8['return_massage'].value_counts(dropna=False)


# In[83]:


df_tianchuang_8['value_011'].value_counts(dropna=False)


# In[84]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tianchuang_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### fulin_1

# In[85]:


ds = 'fulin_1'
df_fulin_1 = merge_csv_file(ds, file_name_zz)
df_fulin_1.dropna(how='all', axis=1, inplace=True)
df_fulin_1.shape


# In[ ]:


df_fulin_1['return_massage'].value_counts(dropna=False).head()


# In[ ]:


df_fulin_1['return_code'].value_counts(dropna=False)


# In[ ]:


df_fulin_1['create_time'].min()


# In[86]:


df_fulin_1 = df_fulin_1[df_fulin_1['return_massage'].isin(['success','查询成功'])]


# In[87]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_fulin_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### tongcheng_tag

# In[88]:


ds = 'tongcheng_tag'
df_tongcheng_tag = merge_csv_file(ds, file_name_zz)
df_tongcheng_tag.dropna(how='all', axis=1, inplace=True)
df_tongcheng_tag.shape


# ### tongliantong_3

# In[89]:


ds = 'tongliantong_3'
df_tongliantong_3 = merge_csv_file(ds, file_name_zz)
df_tongliantong_3.dropna(how='all', axis=1, inplace=True)
df_tongliantong_3.shape


# ### tongdun_2

# In[90]:


ds = 'tongdun_2'
df_tongdun_2 = merge_csv_file(ds, file_name_zz)
df_tongdun_2.dropna(how='all', axis=1, inplace=True)
df_tongdun_2.shape


# In[91]:


df_tongdun_2['value_005'].value_counts(dropna=False)


# In[92]:


df_tongdun_2 = df_tongdun_2[df_tongdun_2['value_005'].isin([0.0])]


# In[93]:


df_tongdun_2['value_001'].value_counts(dropna=False)


# In[94]:


df_tongdun_2['value_002'].value_counts(dropna=False)


# In[95]:


tmp = pd.get_dummies(df_tongdun_2['value_002'])
df_tongdun_2 = pd.concat([df_tongdun_2, tmp.mul(df_tongdun_2['model_score_01'], axis=0)], axis=1)
df_tongdun_2.rename(columns={'27984162d86f3baf':'model_score_01_zr', '3a9de0313a0c65ee':'model_score_01_zx'},inplace=True)
df_tongdun_2['create_time'] = df_tongdun_2['create_time'].str[0:10]
df_tongdun_2['create_time'].head()


# In[96]:


df_tongdun_2 = df_tongdun_2.groupby(['order_no','id_no_des','create_time'])['model_score_01_zr','model_score_01_zx'].max()
df_tongdun_2 = df_tongdun_2.reset_index()


# In[97]:


needcols = ['model_score_01_zr', 'model_score_01_zx']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_tongdun_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[98]:


del df_tongdun_2
gc.collect()


# ### talkingdata_7

# In[99]:


ds = 'talkingdata_7'
df_talkingdata_7 = merge_csv_file(ds, file_name_zz)
df_talkingdata_7.dropna(how='all',axis=1, inplace=True)
df_talkingdata_7.shape


# ### pudao_3

# In[185]:


file_name_pudao_3 = get_filename(r'\\192.168.100.120\d\juzi\0912')
file_name_pudao_3


# In[187]:


import time
import json

start_time = time.time()
print(start_time)

usecols = ['order_no', 'id_no_des', 'user_id', 'create_time','channel_id','value_012']
needcols =  ['value_012']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# df_pudao_3_list = {}
for filename in file_name_pudao_3:
    filepath = r'\\192.168.100.120\d\juzi\0912\{}'.format(filename)
    tmp_name = filename[-11:-4]
    print(tmp_name)
    
    tmp_df = chunk_process_data(filepath, usecols, df_order_base, cols_left,
                                cols_right, needcols, suffix='pudao_3', chunk_size=50000)
    tmp_df = tmp_df.reset_index()
    tmp_df['value_012_pudao_3'] = tmp_df['value_012_pudao_3'].apply(lambda x: json.loads(x) if pd.notnull(x) else {})
    tmp_df = pd.concat([tmp_df.drop(['value_012_pudao_3'], axis=1), tmp_df['value_012_pudao_3'].apply(pd.Series)], axis=1)
    df_pudao_3_list[tmp_name] = tmp_df
    
    del tmp_df
    gc.collect()
    end_time = time.time()
    print(end_time)
    total_time = end_time - start_time
    print(total_time/60)
    


# In[188]:


list(df_pudao_3_list.keys())


# In[189]:


df_pudao_3 = pd.concat(list(df_pudao_3_list.values()), axis=0)
df_pudao_3 = df_pudao_3.sort_values(by=['order_no','order_no_is_equal_pudao_3','create_time_pudao_3'], ascending=False).drop_duplicates(subset=['order_no'],keep='first')


# In[190]:


df_pudao_3.info()
df_pudao_3.head()


# In[192]:


df_pudao_3.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_pudao_3_20230927.csv',index=False)


# In[193]:


len(list(df_pudao_3.columns[df_pudao_3.columns.str.contains("ppdi|bank")]))


# In[194]:


# 匹配关联
needcols =  df_pudao_3.columns[df_pudao_3.columns.str.contains('ppdi|bank')].to_list()
print(len(needcols))
cols_right = ['order_no'] + needcols

df_order_base = pd.merge(df_order_base, df_pudao_3[cols_right], how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_4

# In[195]:


ds = 'pudao_4'
df_pudao_4 = merge_csv_file(ds, file_name_zz)
df_pudao_4.dropna(how='all', axis=1, inplace=True)
df_pudao_4.shape


# In[198]:


df_pudao_4['create_time'].min()


# In[200]:


df_pudao_4.info()
df_pudao_4.head()


# In[ ]:


# df_pudao_4 = df_pudao_4[df_pudao_4['value_011']==0.0]


# In[ ]:


print(list(df_pudao_4.columns[df_pudao_4.columns.str.contains("value")])[9:])


# In[ ]:


needcols = list(df_pudao_4.columns[df_pudao_4.columns.str.contains("value")])[9:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[201]:


del df_pudao_3, df_pudao_4
gc.collect()


# ### pudao_5

# In[202]:


ds = 'pudao_5'
df_pudao_5 = merge_csv_file(ds, file_name_zz)
df_pudao_5.dropna(how='all', axis=1, inplace=True)
df_pudao_5.shape


# In[203]:


df_pudao_5['value_014'].value_counts(dropna=False)


# In[204]:


df_pudao_5['value_013'].value_counts(dropna=False)


# In[205]:


df_pudao_5 = df_pudao_5[df_pudao_5['value_014']==0.0]


# In[206]:


needcols = ['value_012']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_5, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_8

# In[207]:


ds = 'pudao_8'
df_pudao_8 = merge_csv_file(ds, file_name_zz)
df_pudao_8.dropna(how='all', axis=1, inplace=True)
df_pudao_8.shape


# In[ ]:


df_pudao_8['value_012'].value_counts(dropna=False)


# In[ ]:


df_pudao_8['value_011'].value_counts(dropna=False)


# In[ ]:


df_pudao_8 = df_pudao_8[df_pudao_8['value_012']==0.0]
df_pudao_8.shape


# In[ ]:


df_pudao_8.model_score_01.head()


# In[ ]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_8, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_11

# In[208]:


ds = 'pudao_11'
df_pudao_11 = merge_csv_file(ds, file_name_zz)
df_pudao_11.dropna(how='all', axis=1, inplace=True)
df_pudao_11.shape


# In[209]:


df_pudao_11['return_massage'].value_counts(dropna=False)


# In[210]:


df_pudao_11['value_011'].value_counts(dropna=False)


# In[211]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_11, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_12

# In[212]:


ds = 'pudao_12'
df_pudao_12 = merge_csv_file(ds, file_name_zz)
df_pudao_12.dropna(how='all', axis=1, inplace=True)
df_pudao_12.shape


# In[213]:


df_pudao_12['value_012'].value_counts(dropna=False)


# In[214]:


df_pudao_12['return_massage'].value_counts(dropna=False)


# In[215]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_12, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_13 

# In[216]:


ds = 'pudao_13'
df_pudao_13 = merge_csv_file(ds, file_name_zz)
df_pudao_13.dropna(how='all', axis=1, inplace=True)
df_pudao_13.shape


# In[217]:


df_pudao_13['value_012'].value_counts(dropna=False)


# In[218]:


df_pudao_13['return_massage'].value_counts(dropna=False)


# In[219]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_13, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_14 

# In[220]:


ds = 'pudao_14'
df_pudao_14 = merge_csv_file(ds, file_name_zz)
df_pudao_14.dropna(how='all', axis=1, inplace=True)
df_pudao_14.shape


# ### pudao_15

# In[221]:


ds = 'pudao_15'
df_pudao_15 = merge_csv_file(ds, file_name_zz)
df_pudao_15.dropna(how='all', axis=1, inplace=True)
df_pudao_15.shape


# In[222]:


df_pudao_15['return_massage'].value_counts(dropna=False)


# In[223]:


df_pudao_15['value_012'].value_counts(dropna=False)


# In[224]:


df_pudao_15['value_011'].value_counts(dropna=False)


# In[225]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_15, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### pudao_16

# In[226]:


ds = 'pudao_16'
df_pudao_16 = merge_csv_file(ds, file_name_zz)
df_pudao_16.dropna(how='all', axis=1, inplace=True)
df_pudao_16.shape


# In[227]:


df_pudao_16['return_massage'].value_counts(dropna=False)


# In[228]:


df_pudao_16['value_012'].value_counts(dropna=False)


# In[229]:


df_pudao_16['value_011'].value_counts(dropna=False)


# In[230]:


df_pudao_16 = df_pudao_16[df_pudao_16['return_massage']=='调用成功']
df_pudao_16.shape


# In[231]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_pudao_16, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[232]:


df_pudao_16['create_time'].min()


# In[233]:


df_order_base.to_csv(r'D:\liuyedao\B卡开发\三方数据匹配\df_order_base_three_data_nobairong_20230927.csv',index=False)


# ### ruizhi_4

# In[100]:


ds = 'ruizhi_4'
df_ruizhi_4 = merge_csv_file(ds, file_name_zz)
df_ruizhi_4.dropna(how='all', axis=1, inplace=True)
df_ruizhi_4.shape


# In[101]:


df_ruizhi_4['return_massage'].value_counts(dropna=False)


# In[ ]:


df_ruizhi_4['value_011'].value_counts(dropna=False)


# In[ ]:


df_ruizhi_4['value_012'].value_counts(dropna=False)


# In[102]:


df_ruizhi_4 = df_ruizhi_4[df_ruizhi_4['return_massage']=='通过']
df_ruizhi_4.shape


# In[103]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ruizhi_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ruizhi_6

# In[104]:


ds = 'ruizhi_6'
df_ruizhi_6 = merge_csv_file(ds, file_name_zz)
df_ruizhi_6.dropna(how='all', axis=1, inplace=True)
df_ruizhi_6.shape


# In[105]:


df_ruizhi_6['return_massage'].value_counts(dropna=False)


# In[ ]:


df_ruizhi_6['value_012'].value_counts(dropna=False)


# In[ ]:


df_ruizhi_6['create_time'].min()


# In[106]:


df_ruizhi_6 = df_ruizhi_6[df_ruizhi_6['return_massage']=='调用成功']
df_ruizhi_6.shape


# In[107]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_ruizhi_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### ruizhi_5

# In[108]:


ds = 'ruizhi_5'
df_ruizhi_5 = merge_csv_file(ds, file_name_zz)
df_ruizhi_5.dropna(how='all', axis=1, inplace=True)
df_ruizhi_5.shape


# In[109]:


df_ruizhi_5['create_time'].min()


# ### bileizhen_1

# In[110]:


ds = 'bileizhen_1'
df_bileizhen_1 = merge_csv_file(ds, file_name_zz)
df_bileizhen_1.dropna(how='all', axis=1, inplace=True)
df_bileizhen_1.shape


# In[111]:


df_bileizhen_1['return_massage'].value_counts(dropna=False)


# In[112]:


df_bileizhen_1['value_002'].value_counts(dropna=False)


# In[ ]:


df_bileizhen_1['value_001'].value_counts(dropna=False)


# In[113]:


df_bileizhen_1 = df_bileizhen_1.query("value_002==0.0")
df_bileizhen_1.shape


# In[114]:


df_bileizhen_1['create_time'].min()


# In[115]:


tmp = pd.get_dummies(df_bileizhen_1['value_001'])
df_bileizhen_1 = pd.concat([df_bileizhen_1, tmp.mul(df_bileizhen_1['model_score_01'], axis=0)], axis=1)
df_bileizhen_1.rename(columns={1.0:'model_score_01_1', 2.0:'model_score_01_2', 3.0:'model_score_01_3'},inplace=True)


# In[116]:


df_bileizhen_1['create_time'] = df_bileizhen_1['create_time'].str[0:10]
df_bileizhen_1['create_time'].head()

df_bileizhen_1 = df_bileizhen_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_1','model_score_01_2','model_score_01_3'].max()
df_bileizhen_1 = df_bileizhen_1.reset_index()
df_bileizhen_1.shape


# In[ ]:


df_bileizhen_1.info()


# In[117]:


needcols = ['model_score_01_1', 'model_score_01_2', 'model_score_01_3']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_bileizhen_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[118]:


del df_bileizhen_1
gc.collect()


# ### duxiaoman_1

# In[119]:


ds = 'duxiaoman_1'
df_duxiaoman_1 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_1.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_1.shape


# In[ ]:


df_duxiaoman_1['create_time'].min()


# In[120]:


df_duxiaoman_1['return_massage'].value_counts(dropna=False)


# In[121]:


df_duxiaoman_1['value_003'].value_counts(dropna=False)


# In[ ]:


df_duxiaoman_1['value_001'].value_counts(dropna=False)


# In[ ]:


df_duxiaoman_1['value_002'].value_counts(dropna=False)


# In[122]:


df_duxiaoman_1 = df_duxiaoman_1[df_duxiaoman_1['value_003']==0.0]


# In[123]:


df_duxiaoman_1['value_002'].value_counts(dropna=False)


# In[ ]:


df_duxiaoman_1.groupby(['value_001','value_002'],dropna=False)['order_no'].count().unstack()


# In[124]:


df_duxiaoman_1.reset_index(drop=True, inplace=True)
tmp = pd.get_dummies(df_duxiaoman_1['value_002'])
df_duxiaoman_1 = pd.concat([df_duxiaoman_1, tmp.mul(df_duxiaoman_1['model_score_01'], axis=0)], axis=1)
df_duxiaoman_1.rename(columns={'dxm_zzyzv1':'model_score_01_v1', 'dxm_zzyzv2':'model_score_01_v2'},inplace=True)
df_duxiaoman_1['create_time'] = df_duxiaoman_1['create_time'].str[0:10]
df_duxiaoman_1['create_time'].head()

df_duxiaoman_1 = df_duxiaoman_1.groupby(['order_no','id_no_des','create_time'])['model_score_01_v1','model_score_01_v2'].max()
df_duxiaoman_1 = df_duxiaoman_1.reset_index()


# In[125]:


needcols = ['model_score_01_v1','model_score_01_v2']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_duxiaoman_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[126]:


del df_duxiaoman_1
gc.collect()


# ### duxiaoman_5

# In[127]:


ds = 'duxiaoman_5'
df_duxiaoman_5 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_5.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_5.shape


# ### duxiaoman_6

# In[128]:


ds = 'duxiaoman_6'
df_duxiaoman_6 = merge_csv_file(ds, file_name_zz)
df_duxiaoman_6.dropna(how='all', axis=1, inplace=True)
df_duxiaoman_6.shape


# In[129]:


df_duxiaoman_6['return_massage'].value_counts(dropna=False)


# In[ ]:


df_duxiaoman_6['value_017'].value_counts(dropna=False)


# In[ ]:


df_duxiaoman_6['create_time'].min()


# In[130]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_duxiaoman_6, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# # hangliezhi_1

# In[131]:


ds = 'hangliezhi_1'
df_hangliezhi_1 = merge_csv_file(ds, file_name_zz)
df_hangliezhi_1.dropna(how='all', axis=1, inplace=True)
df_hangliezhi_1.shape


# In[132]:


df_hangliezhi_1['return_massage'].value_counts(dropna=False)


# In[133]:


df_hangliezhi_1 = df_hangliezhi_1[df_hangliezhi_1['return_massage']=='查询成功']
df_hangliezhi_1.shape


# In[134]:


df_hangliezhi_1['create_time'].min()


# In[135]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hangliezhi_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### hengpu_4

# In[136]:


ds = 'hengpu_4'
df_hengpu_4 = merge_csv_file(ds, file_name_zz)
df_hengpu_4.dropna(how='all', axis=1, inplace=True)
df_hengpu_4.shape


# In[137]:


df_hengpu_4['value_012'].value_counts(dropna=False)


# In[138]:


df_hengpu_4['return_massage'].value_counts(dropna=False)


# In[ ]:


df_hengpu_4['value_011'].value_counts(dropna=False)


# In[139]:


df_hengpu_4 = df_hengpu_4[df_hengpu_4['value_012']==0.0]
df_hengpu_4.shape


# In[140]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols
# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hengpu_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### huanbei_tag

# In[141]:


ds = 'huanbei_tag'
df_huanbei_tag = merge_csv_file(ds, file_name_zz)
df_huanbei_tag.dropna(how='all', axis=1, inplace=True)
df_huanbei_tag.shape


# ### hulian_4

# In[144]:


ds = 'hulian_4'
df_hulian_4 = merge_csv_file(ds, file_name_zz)
# df_hulian_4.dropna(how='all', axis=1, inplace=True)
# df_hulian_4.shape
# len(df_hulian_4)


# In[ ]:


df_hulian_4['return_massage'].value_counts(dropna=False)


# In[ ]:


df_hulian_4['value_001'].value_counts(dropna=False)


# In[ ]:


print(df_hulian_4['order_no'].nunique(), df_hulian_4.shape)


# In[ ]:


tmp = pd.get_dummies(df_hulian_4['value_001'])
df_hulian_4 = pd.concat([df_hulian_4[['order_no','id_no_des','create_time']], tmp.mul(df_hulian_4['model_score_01'], axis=0)], axis=1)


# In[ ]:


df_hulian_4.info(show_counts=True)


# In[ ]:


df_hulian_4.rename(columns={100:'model_score_01_100', 101:'model_score_01_101', 103:'model_score_01_103'
                           ,105:'model_score_01_105', 107:'model_score_01_107', 121:'model_score_01_121'
                           ,122:'model_score_01_122', 124:'model_score_01_124'} ,inplace=True)


# In[ ]:


df_hulian_4['create_time'] = df_hulian_4['create_time'].str[0:10]
df_hulian_4['create_time'].head(2)


# In[ ]:


df_hulian_4 = df_hulian_4.groupby(['order_no','id_no_des','create_time'])[df_hulian_4.columns[3:].to_list()].max()
df_hulian_4 = df_hulian_4.reset_index()
df_hulian_4.shape


# In[ ]:


needcols = df_hulian_4.columns.to_list()[3:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hulian_4, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[ ]:


df_order_base.columns[df_order_base.columns.str.contains('hulian_4')]


# In[ ]:


del df_hulian_4
gc.collect()


# ### hulian_5 

# In[145]:


ds = 'hulian_5'
df_hulian_5 = merge_csv_file(ds, file_name_zz)
df_hulian_5.dropna(how='all', axis=1, inplace=True)
df_hulian_5.shape


# In[146]:


df_hulian_5['return_massage'].value_counts(dropna=False)


# In[ ]:


df_hulian_5['value_002'].value_counts(dropna=False)


# In[ ]:


df_hulian_5['value_003'].value_counts(dropna=False)


# In[ ]:


df_hulian_5['value_004'].value_counts(dropna=False)


# In[147]:


print(df_hulian_5['order_no'].nunique(), df_hulian_5.shape)


# In[148]:


df_hulian_5['create_time'].min()


# In[149]:


df_hulian_5 = df_hulian_5[df_hulian_5['return_massage']=='操作成功']


# In[ ]:


df_hulian_5['create_time'].max()


# In[150]:


tmp = pd.get_dummies(df_hulian_5['value_002'].fillna(-999).apply(lambda x: int(x)))
df_hulian_5 = pd.concat([df_hulian_5[['order_no','id_no_des','create_time']], tmp.mul(df_hulian_5['model_score_01'], axis=0)], axis=1)


# In[151]:


df_hulian_5.info(show_counts=True)


# In[152]:


df_hulian_5.rename(columns={100:'model_score_01_100', 101:'model_score_01_101', 103:'model_score_01_103'
                           ,105:'model_score_01_105', 107:'model_score_01_107', 121:'model_score_01_121'
                           ,122:'model_score_01_122', 124:'model_score_01_124', 125:'model_score_01_125'
                           ,127:'model_score_01_127', 130:'model_score_01_130', 131:'model_score_01_131'
                           ,132:'model_score_01_132', 133:'model_score_01_133', 136:'model_score_01_136'
                           ,137:'model_score_01_137'
                           } ,inplace=True)


# In[153]:


df_hulian_5['create_time'] = df_hulian_5['create_time'].str[0:10]
df_hulian_5['create_time'].head(2)
df_hulian_5.info()


# In[154]:


df_hulian_5 = df_hulian_5.groupby(['order_no','id_no_des','create_time'])[df_hulian_5.columns[4:].to_list()].max()
df_hulian_5 = df_hulian_5.reset_index()
df_hulian_5.shape


# In[155]:


needcols = df_hulian_5.columns.to_list()[4:]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_hulian_5, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# In[156]:


del df_hulian_5
gc.collect()


# ### moxingfen_27 

# In[157]:


ds = 'moxingfen_27'
df_moxingfen_27 = merge_csv_file(ds, file_name_zz)
df_moxingfen_27.dropna(how='all', axis=1, inplace=True)
df_moxingfen_27.shape


# In[ ]:


df_moxingfen_27.create_time.max()


# ### moxingfen_7

# In[158]:


ds = 'moxingfen_7'
df_moxingfen_7 = merge_csv_file(ds, file_name_zz)
df_moxingfen_7.dropna(how='all', axis=1, inplace=True)
df_moxingfen_7.shape


# In[159]:


print(df_moxingfen_7['create_time'].min(), df_moxingfen_7['create_time'].max())


# In[160]:


tmp = df_moxingfen_7.columns[df_moxingfen_7.columns.str.contains('value')].to_list()
print(tmp[0:14] , tmp[23:])


# In[161]:


df_moxingfen_7.head()


# In[162]:


needcols = ['model_score_01'] + tmp[0:14]
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_moxingfen_7, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### my_1

# In[163]:


ds = 'my_1'
df_my_1 = merge_csv_file(ds, file_name_zz)
df_my_1.dropna(how='all', axis=1, inplace=True)
df_my_1.shape


# In[164]:


df_my_1['return_massage'].value_counts(dropna=False)


# In[ ]:


df_my_1['value_011'].value_counts(dropna=False)


# In[165]:


df_my_1['value_012'].value_counts(dropna=False)


# In[166]:


df_my_1 = df_my_1.query("value_012==0.0")


# In[167]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_my_1, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### aliyun_2

# In[168]:


ds = 'aliyun_2'
df_aliyun_2 = merge_csv_file(ds, file_name_zz)
df_aliyun_2.dropna(how='all', axis=1, inplace=True)
df_aliyun_2.shape


# In[169]:


df_aliyun_2['value_003'].value_counts(dropna=False)


# In[170]:


df_aliyun_2['return_massage'].value_counts(dropna=False)


# In[171]:


df_aliyun_2['value_001'].value_counts(dropna=False)


# In[172]:


df_aliyun_2 = df_aliyun_2[df_aliyun_2['return_massage'].isin(['调用成功,有效数据', '请求成功','OK'])]
df_aliyun_2.shape


# In[173]:


needcols = ['model_score_01']
cols_left = ['order_no','id_no_des','apply_date']
cols_right = ['order_no','id_no_des','create_time'] + needcols

# 返回匹配的三方数据
df_three = process_data(df_order_base, cols_left, df_aliyun_2, cols_right, needcols=needcols, suffix=ds)
print(df_three.shape)
# 匹配关联
df_order_base = pd.merge(df_order_base, df_three, how='left',on='order_no')
print(df_order_base.shape)


# ### 94ai_1

# In[174]:


ds = '94ai_1'
df_94ai_1 = merge_csv_file(ds, file_name_zz)
df_94ai_1.dropna(how='all', axis=1, inplace=True)
df_94ai_1.shape


# In[175]:


df_94ai_1.create_time.min()


# ### face_1

# In[176]:


ds = 'face_1'
df_face_1 = merge_csv_file(ds, file_name_zz)
df_face_1.dropna(how='all', axis=1, inplace=True)
df_face_1.shape


# In[ ]:


df_face_1.info()
df_face_1.head()


# In[ ]:


df_face_1['return_massage'].value_counts(dropna=False)


# In[ ]:


df_face_1['value_011'].value_counts(dropna=False)


# In[ ]:


df_face_1['value_010'].value_counts(dropna=False)


# In[177]:


df_face_1.create_time.min()


# ### vivo_1

# In[178]:


ds = 'vivo_1'
df_vivo_1 = merge_csv_file(ds, file_name_zz)
df_vivo_1.dropna(how='all', axis=1, inplace=True)
df_vivo_1.shape


# ### vivo_tag1

# In[179]:


ds = 'vivo_tag1'
df_vivo_tag1 = merge_csv_file(ds, file_name_zz)
df_vivo_tag1.dropna(how='all', axis=1, inplace=True)
df_vivo_tag1.shape


# ### vivo_tag2

# In[180]:


ds = 'vivo_tag2'
df_vivo_tag2 = merge_csv_file(ds, file_name_zz)
df_vivo_tag2.dropna(how='all', axis=1, inplace=True)
df_vivo_tag2.shape


# ### xl_1

# In[181]:


ds = 'xl_1'
df_xl_1 = merge_csv_file(ds, file_name_zz)
df_xl_1.dropna(how='all', axis=1, inplace=True)
df_xl_1.shape


# ### yinlian_1

# In[182]:


ds = 'yinlian_1'
df_yinlian_1 = merge_csv_file(ds, file_name_zz)
df_yinlian_1.dropna(how='all', axis=1, inplace=True)
df_yinlian_1.shape


# In[183]:


print(df_yinlian_1.create_time.min(), df_yinlian_1.create_time.max())


# ### zhixin_1

# In[184]:


ds = 'zhixin_1'
df_zhixin_1 = merge_csv_file(ds, file_name_zz)
df_zhixin_1.dropna(how='all', axis=1, inplace=True)
df_zhixin_1.shape


# ## 保存三方匹配数据

# In[ ]:


df_order_base.shape


# In[ ]:


df_order_base.query("smaple_set=='train'").to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v5_train.csv', index=False)
df_order_base.query("smaple_set=='oot'").to_csv(r'C:\Users\ruizhi\Documents\lxl\result_output_data\df_order_base_three_data_v5_oot.csv', index=False)


# ## 三方数据匹配率

# In[4]:


# 读取数据
df_order_base = pd.read_csv(r'D:\liuyedao\B卡开发\三方数据匹配\order_20230927.csv')


# In[5]:


df_order_base.info()


# In[6]:


to_drop = list(df_order_base.select_dtypes(include='object').columns)
print(to_drop)


# In[7]:


df_order_base.drop(['operationType', 'swift_number', 'name', 'mobileEncrypt', 'orderNo', 'idCardEncrypt'],axis=1,inplace=True)


# In[8]:


# 需要不计算匹配率的字段
df_order_base.columns[0:6]


# In[9]:


# 按渠道统计各字段的匹配情况
df_order_base['channel_id'].value_counts()


# In[10]:


df_order_base_copy = df_order_base.copy()


# In[11]:


import toad 

df_order_base = df_order_base[df_order_base.channel_id.isin([174])]
to_drop = list(df_order_base.columns[0:6])
to_keep = list(df_order_base.columns[6:])
tmp_total = toad.detect(df_order_base.drop(to_drop, axis=1))


# In[12]:


# 按申请年月统计匹配情况
year_month = df_order_base.groupby(by=['lending_month'])[to_keep].apply(lambda x: x.isna().sum()/x.shape[0])
year_month = year_month.T
year_month['avg_missing_rate'] = year_month.mean(axis=1)
year_month['avg_std_rate'] = year_month.std(axis=1)


# In[ ]:


# # 按数据集类型统计匹配情况
# smaple_set = df_order_base.groupby(by=['smaple_set'])[to_keep].apply(lambda x: x.isna().sum()/x.shape[0])
# smaple_set = smaple_set.T
# smaple_set['avg_missing_rate'] = smaple_set.mean(axis=1)
# smaple_set['avg_std_rate'] = smaple_set.std(axis=1)


# In[ ]:


# 按申请渠道统计匹配情况
channel = df_order_base.groupby(by=['channel_id'])[to_keep].apply(lambda x: x.isna().sum()/x.shape[0])
channel = channel.T
channel['avg_missing_rate'] = channel.mean(axis=1)
channel['avg_std_rate'] = channel.std(axis=1)


# In[13]:


# 保存统计的数据
writer=pd.ExcelWriter(r'D:\liuyedao\B卡开发\三方数据匹配\B卡_三方数据匹配_缺失率_20231008.xlsx')
tmp_total.to_excel(writer,sheet_name='总体')

year_month.to_excel(writer,sheet_name='总体_年月')
# channel.to_excel(writer,sheet_name='总体_渠道')
# smaple_set.to_excel(writer,sheet_name='总体_样本类型')

writer.save()


# In[ ]:







#==============================================================================
# File: 三方数据评分卡建模-LR-final.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[13]:


import pandas as pd
import numpy as np
import toad
import os 
from datetime import datetime
from sklearn.model_selection import train_test_split


# In[14]:


os.getcwd()


# # 1.读取数据集

# In[15]:


# data = pd.read_csv(r'd:\liuyedao\model_result\model_data.csv') 
# data = pd.read_csv(r'd:\liuyedao\model_result\model_data_20230721.csv') 
data = pd.read_csv(r'd:\liuyedao\model_result\auth_model_data_channel_20230726.csv') 
print('数据大小：', data.shape)
data.head()


# In[16]:


data.info()


# ## 1.1标签处理

# In[17]:


data['Firs6ever30'].value_counts(dropna=False)


# In[18]:


data_gray = data[data['Firs6ever30'].isin([1.0])]
data_gray.shape


# In[19]:


data = data[data['Firs6ever30'].isin([0.0, 2.0])]
data.info()


# In[20]:


data['Firs6ever30'] = data['Firs6ever30']/2
data['Firs6ever30'].value_counts(dropna=False)


# In[21]:


tmp = data.groupby(['year_month'])['Firs6ever30'].agg({'count','sum','mean'})
tmp = pd.concat([tmp, pd.DataFrame(tmp.sum(axis=0),columns=['总计']).T], axis=0)
tmp.loc['总计', 'mean'] = tmp.loc['总计', 'sum'] /tmp.loc['总计', 'count']
tmp.columns = ['total', 'bad_rate', 'bad']
tmp


# In[22]:


xx = data.groupby(['channel_id'])['Firs6ever30'].agg({'count','sum','mean'})
xx = pd.concat([xx, pd.DataFrame(xx.sum(axis=0),columns=['总计']).T], axis=0)
xx.loc['总计', 'mean'] = xx.loc['总计', 'sum'] /xx.loc['总计', 'count']
xx.columns = ['total', 'bad_rate', 'bad']
xx


# In[23]:


data = data.reset_index(drop=True)
data.info()


# In[24]:


data_copy = data.copy()


# In[25]:


to_drop = list(data.columns[data.columns.str.contains('bairong')])+['apply_date_cash','lending_time','Firs3ever15','Firs3ever30','Firs6ever15','auth_credit_amount']
print(to_drop)


# In[26]:


data.rename(columns={'Firs6ever30':'target'},inplace=True)
data.drop(to_drop, axis=1,inplace=True)
data.info()


# ## 1.2拆分数据

# In[27]:


data = data.reset_index(drop=True)
data.info()


# In[28]:


to_drop = list(data.columns)[0:6]
print(to_drop)


# In[29]:


df_var = data.drop(['target'], axis=1)
df_target = data['target']

X = np.array(df_var)
y = np.array(df_target)


# In[30]:


X_train, X_test, y_train, y_test = train_test_split(df_var ,df_target, test_size=0.3, random_state=22, stratify=y)


# In[31]:


train = pd.merge(X_train, y_train, how='inner', left_index=True, right_index=True)
print(train.shape, X_train.shape, y_train.shape)


# In[32]:


oot = pd.merge(X_test, y_test, how='inner', left_index=True, right_index=True)
print(oot.shape, X_test.shape, y_test.shape)


# In[33]:


print(train['target'].value_counts())
print(oot['target'].value_counts())


# In[ ]:


train['type'] = 1
oot['type'] = 2


# In[ ]:


train.to_csv(r'd:\liuyedao\model_result\{}_model_data_channel_train_{}.csv'.format('auth', str(datetime.today())[:10].replace('-','')))


# In[ ]:


oot.to_csv(r'd:\liuyedao\model_result\{}_model_data_channel_oot_{}.csv'.format('auth', str(datetime.today())[:10].replace('-','')))


# # 2.数据探索分析

# In[34]:


train.info()
train.shape


# In[35]:


oot.info()
oot.shape


# In[27]:


train_df_explore = toad.detect(train.drop(to_drop,axis=1))
oot_df_explore = toad.detect(oot.drop(to_drop,axis=1))

train_df_iv = toad.quality(train.drop(to_drop,axis=1),'target',iv_only=True)
oot_df_iv = toad.quality(oot.drop(to_drop,axis=1),'target',iv_only=True)

# writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_数据探索性分析_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
# train_df_explore.to_excel(writer,sheet_name='train_df_explore')
# oot_df_explore.to_excel(writer,sheet_name='oot_df_explore')
# train_df_iv.to_excel(writer,sheet_name='train_df_iv')
# oot_df_iv.to_excel(writer,sheet_name='oot_df_iv')
# writer.save()


# # 3.特征筛选

# In[36]:


train_selected, dropped = toad.selection.select(train, target='target', empty=0.9, iv=0.01, corr=0.8, return_drop=True, exclude=to_drop)
print(dropped)
print(train_selected.shape)


# In[38]:


# for i in dropped['empty']:
#     print(i)
#     train_df_explore.loc[i, 'is_empty'] = 1
# for i in dropped['iv']:
#     print(i)
#     train_df_explore.loc[i, 'is_iv'] = 1


# In[39]:


train_selected.info()
train_selected.head()


# # 4. 变量分箱

# In[40]:


train_selected.min()


# In[92]:


train_selected.describe().T['count']/train_selected.shape[0] *100


# In[41]:


import warnings

warnings.filterwarnings("ignore")


# In[41]:


# for col in train_selected.columns[6:-1]:
#     print(col)
#     train_selected[col].fillna(-999,inplace=True)

# train_selected['model_score_01_x_tianchuang'].fillna(-999,inplace=True)
# train_selected['model_score_01_zr_tongdun'].fillna(train_selected['model_score_01_zr_tongdun'].mean(),inplace=True)
# train_selected['model_score_01_fulin'].fillna(train_selected['model_score_01_fulin'].mean(),inplace=True)
# train_selected['model_score_01_moxingfen_14'].fillna(-999,inplace=True)
# train_selected['value_012_pudao_6'].fillna(-999,inplace=True)
# train_selected['model_score_01_baihang'].fillna(-999,inplace=True)
# train_selected['model_score_01_rong360'].fillna(train_selected['model_score_01_rong360'].mean(),inplace=True)
# train_selected['model_score_01_tengxun'].fillna(train_selected['model_score_01_tengxun'].mean(),inplace=True)
# train_selected['model_score_01_xysl_1'].fillna(train_selected['model_score_01_xysl_1'].mean(),inplace=True)
# train_selected['model_score_01_xysl_3'].fillna(train_selected['model_score_01_xysl_3'].mean(),inplace=True)


# In[42]:


oot_selected = oot[list(train_selected.columns)]
oot_selected.info()


# In[43]:


# for col in oot_selected.columns[6:-1]:
#     print(col)
#     oot_selected[col].fillna(-999,inplace=True)

# # oot_selected['model_score_01_x_tianchuang'].fillna(-999,inplace=True)
# # oot_selected['model_score_01_zr_tongdun'].fillna(train_selected['model_score_01_zr_tongdun'].mean(),inplace=True)
# # oot_selected['model_score_01_fulin'].fillna(train_selected['model_score_01_fulin'].mean(),inplace=True)
# # oot_selected['model_score_01_moxingfen_14'].fillna(-999,inplace=True)
# # oot_selected['value_012_pudao_6'].fillna(-999,inplace=True)
# # oot_selected['model_score_01_baihang'].fillna(-999,inplace=True)
# # oot_selected['model_score_01_rong360'].fillna(train_selected['model_score_01_rong360'].mean(),inplace=True)
# # oot_selected['model_score_01_tengxun'].fillna(train_selected['model_score_01_tengxun'].mean(),inplace=True)
# # oot_selected['model_score_01_xysl_1'].fillna(train_selected['model_score_01_xysl_1'].mean(),inplace=True)
# # oot_selected['model_score_01_xysl_3'].fillna(train_selected['model_score_01_xysl_3'].mean(),inplace=True)


# In[43]:


# 第一次分箱
c = toad.transform.Combiner()
c.fit(train_selected.drop(to_drop,axis=1), y='target', method='dt', min_samples=1, n_bins=20, empty_separate=True) 
bins_result = c.export()


# In[54]:


def regroup(data_bins, col, target='target'):    
    total = data_bins.groupby(col)[target].count()
    bad = data_bins.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
    regroup['good_pct'] = regroup['good']/regroup['good'].sum()
    regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
    regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
    regroup['ks'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
    regroup['ks_max'] = regroup['ks'].max()
    regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
    regroup['iv'] = regroup['iv_bins'].sum()
    regroup['varsname'] = col

    return regroup


# In[118]:


train_selected_bins = c.transform(train_selected, labels=True)
train_selected_bins.head()


# In[119]:


oot_selected_bins = c.transform(oot_selected, labels=True)
oot_selected_bins.head()


# In[120]:


df_result = pd.DataFrame()
for col in train_selected_bins.columns[6:-1]:
    print('------------变量：{}-----------'.format(col))
    tmp = regroup(train_selected_bins, col, target='target')
    df_result = pd.concat([df_result, tmp], axis=0)


# In[121]:


df_result_oot = pd.DataFrame()
for col in oot_selected_bins.columns[6:-1]:
    print('------------变量：{}-----------'.format(col))
    tmp = regroup(oot_selected_bins, col, target='target')
    df_result_oot = pd.concat([df_result_oot, tmp], axis=0)


# In[122]:


df_result_oot = df_result_oot.reset_index()
df_result = df_result.reset_index()
xx = pd.merge(df_result, df_result_oot, how='left', on=['index', 'varsname'])


# In[123]:


xx.to_excel(r'd:\liuyedao\mid_result\auth_建模_分箱结果_df_result_14_{}.xlsx'.format(str(datetime.today())[:10].replace('-','')))


# In[44]:


# 调整分箱:空值单独一箱
for col in train_selected.columns[6:-1]:
    print(col)
    train_selected[col].fillna(-999,inplace=True)
for col in oot_selected.columns[6:-1]:
    print(col)
    oot_selected[col].fillna(-999,inplace=True)

adj_bins={'model_score_01_x_tianchuang': [-0.5, 555.5, 599.5],
 'model_score_01_zr_tongdun': [-0.5, 629.5, 655.5, 673.5, 749.5, 811.5, 850.5],
 'model_score_01_fulin': [-0.5, 655.5, 730.5, 804.5],
 'model_score_01_baihang': [-0.5, 699.5, 719.5, 734.5, 756.5, 774.5],
 'model_score_01_xysl_1': [-0.5, 497.5, 521.5, 574.5],
 'model_score_01_xysl_3': [-0.5, 569.5, 612.5, 635.5, 660.5, 680.5, 699.5, 723.5]}

# 更新分箱
c.update(adj_bins)


# In[ ]:


# writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auht_建模_分箱结果_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
# for col in list(train_selected_bins.drop(to_drop, axis=1).columns):
#     print(col)
#     regroup_df = regroup(train_selected_bins, col, target='target')
#     regroup_df.to_excel(writer, sheet_name=str(col))

# writer.save()
# # result.to_excel(r'd:\liuyedao\mid_result\{}_建模_bins_{}.xlsx'.format('auth', str(datetime.today())[:10].replace('-','')))


# In[ ]:


# # 业务解释调整分箱
# adj_bins = {
#  'model_score_01_x_tianchuang': [491.0, 596.0],
#  'model_score_01_zr_tongdun': [712.0, 832.0],
#  'model_score_01_fulin': [672.0, 739.0],
#  'model_score_01_xysl_3': [613.0, 661.0, 696.0],
# }
# c.update(adj_bins)


# In[ ]:


# c.export()
# c.load(dict)
# c.transform(dataframe, labels=False)


# ### 观察分箱并调整

# In[124]:


from toad.plot import bin_plot
from toad.plot import badrate_plot

# to_drop = ['target','type']
data_new = pd.concat([train_selected, oot_selected],axis=0)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[0]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[1]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[2]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[3]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[4]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[5]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[6]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[7]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[8]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[9]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot_selected[[col, 'target']], labels=True), x=col, target='target')
# 查看时间内分箱稳定性
badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[10]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# # from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# # badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[11]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# # from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# # badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(data_new[[col, 'target','type']], labels=True), x='type', target='target', by=col)


# In[ ]:


# # 调整分箱
# adj_bins = {'value_097_bairong': [-0.5, 3.0, 4,0, 5.0, 6.0, 9.0]}
# c.update(adj_bins)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[12]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# # bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[14]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# # bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[15]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# # bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[ ]:


# # 静态观察
# col = train_selected.drop(to_drop,axis=1).columns[16]
# print('========{}：变量的分箱图========='.format(col))
# bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
# # bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')

# from toad.plot import badrate_plot
# # 看时间内的分箱稳定性
# badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# # badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# # 5.WOE转换

# In[45]:


train_selected.info()


# In[46]:


exclude = to_drop + ['model_score_01_moxingfen_14','value_012_pudao_6','model_score_01_rong360', 'model_score_01_tengxun','model_score_01_zr_tongdun']+ ['target']
print(exclude)


# In[47]:


transer = toad.transform.WOETransformer()
train_woe = transer.fit_transform(c.transform(train_selected), train_selected['target'], exclude=exclude)


# In[333]:


train_woe.info()
train_woe.head()


# In[48]:


oot_selected = oot_selected[list(train_woe.columns)]
oot_woe = transer.transform(c.transform(oot_selected))


# In[335]:


oot_woe.info()
oot_woe.head()


# # 6.逐步回归

# In[49]:


exclude = to_drop + ['model_score_01_moxingfen_14','value_012_pudao_6','model_score_01_rong360', 'model_score_01_tengxun','model_score_01_zr_tongdun']
train_selected_woe, dropped = toad.selection.select(train_woe, target='target', empty=0.9, iv=0.01, corr=0.7, return_drop=True, exclude=exclude)
print(dropped)
print(train_selected_woe.shape)


# In[50]:


exclude =  ['model_score_01_moxingfen_14','value_012_pudao_6','model_score_01_rong360', 'model_score_01_tengxun','model_score_01_zr_tongdun']
final_data = toad.selection.stepwise(train_selected_woe.drop(exclude,axis=1), target='target', estimator='ols', direction='both', criterion='aic', exclude=to_drop)
final_oot = oot_woe[list(final_data.columns)]


# In[51]:


final_data.info()


# In[52]:


final_oot.info()


# In[53]:


cols = list(final_data.drop(to_drop + ['target'], axis=1).columns)
cols


# In[54]:


print(toad.metrics.PSI(final_data[cols], final_oot[cols]))


# # 7.建模和评估

# In[55]:


from sklearn.linear_model import LogisticRegression

lr = LogisticRegression()
lr.fit(final_data[cols], final_data['target'])


# In[343]:


lr.coef_


# In[56]:


lr.coef_


# In[57]:


lr.feature_names_in_


# In[58]:


# 预测训练/oot
pred_train = lr.predict_proba(final_data[cols])[:,1]
pred_oot = lr.predict_proba(final_oot[cols])[:,1]
# KS/AUC
from toad.metrics import KS,AUC

print('train KS: ', KS(pred_train, final_data['target']))
print('train AUC: ', AUC(pred_train, final_data['target']))


# In[59]:


# KS/AUC
from toad.metrics import KS,AUC

print('train KS: ', KS(pred_train, final_data['target']))
print('train AUC: ', AUC(pred_train, final_data['target']))

print('-------------oot结果--------------------')
print('oot KS: ', KS(pred_oot, final_oot['target']))
print('oot AUC: ', AUC(pred_oot, final_oot['target']))


# In[346]:


# KS/AUC
from toad.metrics import KS,AUC

print('train KS: ', KS(pred_train, final_data['target']))
print('train AUC: ', AUC(pred_train, final_data['target']))

print('-------------oot结果--------------------')
print('oot KS: ', KS(pred_oot, final_oot['target']))
print('oot AUC: ', AUC(pred_oot, final_oot['target']))


# In[347]:


# PSI
print(toad.metrics.PSI(pred_train, pred_oot))


# In[349]:


from sklearn.metrics import roc_auc_score
from sklearn.metrics import auc
from sklearn.metrics import roc_curve
from matplotlib import pyplot as plt
fpr, tpr, thresholds = roc_curve( final_data['target'],pred_train)
roc_auc = auc(fpr, tpr)
ks = abs(fpr - tpr).max()

fpr_oot, tpr_oot, thresholds_oot = roc_curve( final_oot['target'],pred_oot)
roc_auc_oot = auc(fpr_oot, tpr_oot)
ks_oot = abs(fpr_oot - tpr_oot).max()

print(roc_auc, ks, roc_auc_oot, ks_oot)

plt.plot(fpr, tpr, color='darkorange', lw=2, label='trian ROC curve (area = %0.2f)' % roc_auc)
plt.plot(fpr_oot, tpr_oot, color='blue', lw=2, label='test ROC curve (area = %0.2f)' % roc_auc_oot)
plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="best")
plt.show()


# # 9.转换评分

# In[352]:


card = toad.ScoreCard(combiner=c,
                      transer = transer,
                      base_odds=35,
                      base_score=700,
                      pdo=60,
                      rate=2
                     )


# In[353]:


card.fit(final_data[cols], final_data['target'])


# In[356]:


card1 = card.export(to_frame=True)
card1.to_excel(r'd:\liuyedao\mid_result\auth_建模_评分卡_3_{}.xlsx'.format(str(datetime.today())[:10].replace('-','')))


# # 10.分渠道看模型效果

# In[376]:


final_data_167 = final_data.query("channel_id == 167")
final_oot_167 = final_oot.query("channel_id == 167")


final_data_174 = final_data.query("channel_id == 174")
final_oot_174 = final_oot.query("channel_id == 174")

# 预测训练/oot
pred_train_167 = lr.predict_proba(final_data_167[cols])[:,1]
pred_oot_167 = lr.predict_proba(final_oot_167[cols])[:,1]

pred_train_174 = lr.predict_proba(final_data_174[cols])[:,1]
pred_oot_174 = lr.predict_proba(final_oot_174[cols])[:,1]

# KS/AUC
from toad.metrics import KS,AUC

print('train KS: ', KS(pred_train_167, final_data_167['target']))
print('train AUC: ', AUC(pred_train_167, final_data_167['target']))

print('-------------oot结果--------------------')
print('oot KS: ', KS(pred_oot_167, final_oot_167['target']))
print('oot AUC: ', AUC(pred_oot_167, final_oot_167['target']))

print('------------------------------------------')
print('train KS: ', KS(pred_train_174, final_data_174['target']))
print('train AUC: ', AUC(pred_train_174, final_data_174['target']))

print('-------------oot结果--------------------')
print('oot KS: ', KS(pred_oot_174, final_oot_174['target']))
print('oot AUC: ', AUC(pred_oot_174, final_oot_174['target']))



# In[377]:


def regroup(data, col, target='target'):
    total = data.groupby(col)[target].count()
    bad = data.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    cols = ['varsname','bins','bad','good','total','bad_rate','total_pct']
    regroup = regroup[cols]
    return regroup


# In[378]:


# 业务效果
pred_data = toad.metrics.KS_bucket(pred_train, final_data['target'], bucket=10, method='step')
pred_data


# In[379]:


pred_data['min_score'] = round(card.proba_to_score(pred_data['min']),0)
cut_bins = sorted(list(pred_data['min_score'].unique()),reverse=False)
print(cut_bins)


# In[380]:


cut_bins = [float('-inf')] + cut_bins[:-1] + [float('inf')]
print(cut_bins)


# In[381]:


train_all = pd.DataFrame({'score':card.proba_to_score(pred_train), 'target':np.array(final_data['target'])}).round()
train_all['bins'] = pd.cut(train_all['score'], bins=cut_bins, right=False)
print(train_all['bins'].unique())

oot_all = pd.DataFrame({'score':card.proba_to_score(pred_oot), 'target':np.array(final_oot['target'])}).round()
oot_all['bins'] = pd.cut(oot_all['score'], bins=cut_bins, right=False)
print(oot_all['bins'].unique())


# In[382]:


train_all_167 = pd.DataFrame({'score':card.proba_to_score(pred_train_167), 'target':np.array(final_data_167['target'])}).round()
train_all_167['bins'] = pd.cut(train_all_167['score'], bins=cut_bins, right=False)

oot_all_167 = pd.DataFrame({'score':card.proba_to_score(pred_oot_167), 'target':np.array(final_oot_167['target'])}).round()
oot_all_167['bins'] = pd.cut(oot_all_167['score'], bins=cut_bins, right=False)


# In[383]:


train_all_174 = pd.DataFrame({'score':card.proba_to_score(pred_train_174), 'target':np.array(final_data_174['target'])}).round()
train_all_174['bins'] = pd.cut(train_all_174['score'], bins=cut_bins, right=False)

oot_all_174 = pd.DataFrame({'score':card.proba_to_score(pred_oot_174), 'target':np.array(final_oot_174['target'])}).round()
oot_all_174['bins'] = pd.cut(oot_all_174['score'], bins=cut_bins, right=False)


# In[384]:


train_all_regroup = regroup(train_all, 'bins')
oot_all_regroup = regroup(oot_all, 'bins')

train_all_167_regroup = regroup(train_all_167, 'bins')
oot_all_167_regroup = regroup(oot_all_167, 'bins')

train_all_174_regroup = regroup(train_all_174, 'bins')
oot_all_174_regroup = regroup(oot_all_174, 'bins')


# In[385]:


writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_业务效果_4_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
train_all_regroup.to_excel(writer,sheet_name='train')
oot_all_regroup.to_excel(writer,sheet_name='test')

train_all_167_regroup.to_excel(writer,sheet_name='train_167')
oot_all_167_regroup.to_excel(writer,sheet_name='test_167')

train_all_174_regroup.to_excel(writer,sheet_name='train_174')
oot_all_174_regroup.to_excel(writer,sheet_name='test_174')
writer.save()


# In[387]:


def cal_psi(exp, act):
    psi = []
    for i in range(len(exp)):
        psi_i = (act[i] - exp[i])*np.log(act[i]/exp[i])
        psi.append(psi_i)
    return sum(psi)


# In[388]:


print(cal_psi(train_all_regroup['total_pct'], oot_all_regroup['total_pct'])) 
print(cal_psi(train_all_167_regroup['total_pct'], oot_all_167_regroup['total_pct'])) 
print(cal_psi(train_all_174_regroup['total_pct'], oot_all_174_regroup['total_pct'])) 


# In[198]:


# 全部授信申请客户
df1_auth = pd.read_csv(r'd:\liuyedao\model_result\model_data_20230728_bqc.csv')
df1_auth.info(show_counts=True)


# In[271]:


# 去重数据
data = df1_auth.sort_values(by=['user_id','auth_status','apply_date_auth'], ascending=[True,True,False]).drop_duplicates(subset=['user_id'],keep='first')
print(data['user_id'].nunique(), data.shape)


# In[267]:


# xx = data.query("auth_status==7 & Firs6ever30==Firs6ever30")
# xx['Firs6ever30'].value_counts()


# In[274]:


index = list(data.query("auth_status==7 & Firs6ever30==Firs6ever30").index)
data.drop(index=index,axis=0,inplace=True)
data.shape


# In[389]:


usecols_=list(data.columns[0:7])+['model_score_01_x_tianchuang','model_score_01_xysl_3','model_score_01_baihang','Firs6ever30']
print(usecols_)


# In[390]:


df2_auth = data[usecols_]
df2_auth.reset_index(drop=True, inplace=True)
df2_auth.info(show_counts=True)


# In[391]:


for col in ['model_score_01_x_tianchuang','model_score_01_xysl_3','model_score_01_baihang']:
    df2_auth[col].fillna(-999, inplace=True)
df2_auth.info(show_counts=True)


# In[395]:


df2_auth['score'] = card.predict(df2_auth).round()
df2_auth.info(show_counts=True)


# In[396]:


df2_auth['score'].head()


# In[397]:


print(cut_bins)


# In[398]:


df2_auth['bins'] = pd.cut(df2_auth['score'], bins=cut_bins, right=False)
df2_auth['bins'].unique()


# In[399]:


def cal_auth(df_auth, channel=None):
    if channel:
        df_auth = df_auth.query("channel_id==@channel")
        
    total_lend = df_auth.groupby('bins')['order_no','Firs6ever30'].count()
    tg_jj = df_auth.groupby(['bins', 'auth_status'])['order_no'].count().unstack()
    lend =  df_auth.groupby(['bins', 'Firs6ever30'])['order_no'].count().unstack()
    result = pd.concat([total_lend, tg_jj, lend], axis=1)
    result.columns = ['授信申请','放款人数','通过人数','拒绝人数','好','灰','坏']
    
    result_sum = pd.DataFrame(result.sum(axis=0),columns=['total']).T
    
    for col in result.columns:
        result['{}占比'.format(col)] = result[col]/result[col].sum()
        result['{}累计占比'.format(col)] = result['{}占比'.format(col)].cumsum()
        
    result['坏客率'] = result['坏']/(result['好']+result['坏'])
    result['通过率'] = result['通过人数']/result['授信申请']
    result = pd.concat([result, result_sum],axis=0)
    
    cols = ['授信申请','授信申请占比','授信申请累计占比','通过人数','通过率','拒绝人数','拒绝人数占比','拒绝人数累计占比',
            '通过人数占比','通过人数累计占比','放款人数','放款人数占比','放款人数累计占比','好','好占比','好累计占比',
            '坏','坏占比','坏累计占比','坏客率','灰','灰占比','灰累计占比']
    result = result[cols]
    
    return result


# In[400]:


result_auth = cal_auth(df2_auth, channel=None)
result_auth_167 = cal_auth(df2_auth, channel=167)
result_auth_174 = cal_auth(df2_auth, channel=174)


# In[401]:


result_auth


# In[402]:


writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_业务效果_授信申请_{}_v{}.xlsx".format(str(datetime.today())[:10].replace('-',''),4))
result_auth.to_excel(writer,sheet_name='all')
result_auth_167.to_excel(writer,sheet_name='167')
result_auth_174.to_excel(writer,sheet_name='1174')
writer.save()


# In[286]:


# result_auth.to_excel(r"d:\liuyedao\mid_result\auth_建模_评分卡_业务效果_授信申请_20230728_v2.xlsx")


# In[403]:


dx = data[['model_score_01_x_tianchuang','model_score_01_xysl_3','model_score_01_baihang']]
null_rows_df = data.loc[dx.isnull().all(axis=1),:][usecols_]
null_rows_df.info(show_counts=True)


# In[404]:


null_rows_df['auth_status'].value_counts()


# In[405]:


df2_auth.info(show_counts=True)


# In[406]:


df2_auth.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_LR_score_20230807.csv",index=False)


# # oot验证

# In[60]:


df3_auth = pd.read_csv(r'd:\liuyedao\model_result\oot_data_20230815.csv')
df3_auth.info()


# In[8]:


def cal_score_tc(x):
    score_list = []
    if x<555.5:
        score=188
    elif x>=555.5 and x<599.5:
        score=192
    elif x>=555.5 and x<599.5:
        score=192
    elif x>=599.5:
        score=212
    else:
        score=187
    return score
        
def cal_score_bh(x):      
    if x<699.5:
        score=182
    elif x>=699.5 and x<719.5:
        score=188
    elif x>=719.5 and x<734.5:
        score=193
    elif x>=734.5 and x<756.5:
        score=204
    elif x>=756.5 and x<774.5:
        score=211
    elif x>=774.5:
        score=218
    else:
        score=199
    return score

def cal_score_xysl(x):  
    if x<569.5:
        score=153
    elif x>=569.5 and x<612.5:
        score=171
    elif x>=612.5 and x<635.5:
        score=185
    elif x>=635.5 and x<660.5:
        score=196
    elif x>=660.5 and x<680.5:
        score=206
    elif x>=680.5 and x<699.5:
        score=217
    elif x>=699.5 and x<723.5:
        score=234
    elif x>=723.5:
        score=273
    else:
        score=159
    
    return score
        


# In[9]:


usecols = ['order_no','user_id','id_no_des','channel_id','apply_date_auth','apply_time','auth_status',
           'model_score_01_x_tianchuang','model_score_01_xysl_3','model_score_01_baihang','Firs6ever30','score_lr']

df3_auth['model_score_01_x_tianchuang_score'] = df3_auth['model_score_01_x_tianchuang'].apply(lambda x: cal_score_tc(x))
df3_auth['model_score_01_baihang_score'] = df3_auth['model_score_01_baihang'].apply(lambda x: cal_score_bh(x))
df3_auth['model_score_01_xysl_3_score'] = df3_auth['model_score_01_xysl_3'].apply(lambda x: cal_score_xysl(x))
df3_auth['score_lr'] = df3_auth[['model_score_01_x_tianchuang_score', 'model_score_01_baihang_score','model_score_01_xysl_3_score']].sum(axis=1)

df4_auth = df3_auth[usecols]


# In[11]:


df4_auth.info()
df4_auth.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_LR_score_20230815_oot.csv",index=False)


# In[12]:


df4_auth.info(show_counts=True)


# In[62]:


df3_auth['apply_month'] = df3_auth['apply_date_auth'].str[0:7]
df3_auth['apply_month'].value_counts()


# In[63]:


df3_auth['Firs6ever30'].value_counts()


# In[64]:


list(lr.feature_names_in_)


# In[66]:


auth_data_mob6.groupby(['apply_month', 'Firs6ever30'])['order_no'].count().unstack()


# In[68]:


# 对训练集进行预测
auth_data_mob6 = df3_auth[df3_auth['Firs6ever30'].isin([0.0, 2.0])][list(lr.feature_names_in_) + ['Firs6ever30','apply_month','order_no']]
# auth_data_mob6.groupby(['apply_month', 'Firs6ever30'])['order_no'].count().unstack()
auth_data_mob6 = auth_data_mob6[auth_data_mob6['apply_month']=='2023-01']
auth_data_mob6['target'] = auth_data_mob6['Firs6ever30'] /2.0
# auth_data_mob6.info()
auth_data_mob6 = transer.transform(c.transform(auth_data_mob6.fillna(-999)))
auth_data_mob6['prob'] = lr.predict_proba(auth_data_mob6[list(lr.feature_names_in_)])[:,1]

from toad.metrics import KS,AUC
print('train AUC: ', AUC(auth_data_mob6['prob'], auth_data_mob6['target']))
print('train KS: ', KS(auth_data_mob6['prob'], auth_data_mob6['target']))


# In[74]:


# 对训练集进行预测
auth_data_mob3 = df3_auth[df3_auth['Firs3ever30'].isin([0.0, 2.0])][list(lr.feature_names_in_) + ['Firs3ever30','apply_month','order_no']]
# auth_data_mob3.groupby(['apply_month', 'Firs6ever30'])['order_no'].count().unstack()
auth_data_mob3 = auth_data_mob3[auth_data_mob3['apply_month']!='2023-05']
auth_data_mob3['target'] = auth_data_mob3['Firs3ever30'] /2.0
# auth_data_mob3.info()
auth_data_mob3 = transer.transform(c.transform(auth_data_mob3.fillna(-999)))
auth_data_mob3['prob'] = lr.predict_proba(auth_data_mob3[list(lr.feature_names_in_)])[:,1]



# In[ ]:


from toad.metrics import KS,AUC
print('train AUC: ', AUC(auth_data_mob3['prob'], auth_data_mob3['target']))
print('train KS: ', KS(auth_data_mob3['prob'], auth_data_mob3['target']))


# In[76]:


auth_data_mob3.info()


# In[75]:


auth_data_mob3.groupby(['apply_month', 'Firs3ever30'])['order_no'].count().unstack()


# In[78]:


for i in list(auth_data_mob3['apply_month'].unique()):
    print(i)
    tmp = auth_data_mob3[auth_data_mob3['apply_month']==i]
    print('train AUC: ', AUC(tmp['prob'], tmp['target']))
    print('train KS: ', KS(tmp['prob'], tmp['target']))
    print('-----------------------------------')




#==============================================================================
# File: 三方数据评分卡建模-xgboost-final.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
import toad
import xgboost as xgb
from xgboost import plot_importance
from sklearn import metrics
from sklearn.model_selection import train_test_split
from datetime import datetime
import os 
import warnings
warnings.filterwarnings("ignore")


# In[3]:


os.getcwd()


# # 1.读取数据集

# In[4]:


data = pd.read_csv(r'd:\liuyedao\model_result\model_data_20230728_bqc.csv')
print('数据大小：', data.shape)
data.head(10)


# In[6]:


# 去重数据
data = data.sort_values(by=['user_id','auth_status','apply_date_auth'], ascending=[True,True,False]).drop_duplicates(subset=['user_id'],keep='first')
print(data['user_id'].nunique(), data.shape)


# In[7]:


index = list(data.query("auth_status==7 & Firs6ever30==Firs6ever30").index)
data.drop(index=index,axis=0,inplace=True)
data.reset_index(drop=True, inplace=True)
data.shape


# In[9]:


data['auth_status'].value_counts()


# In[10]:


data['channel_id'].value_counts()


# In[17]:


data_copy = data.copy()


# In[18]:


data = data[data['auth_status']==6]
data.info(show_counts=True)


# In[19]:


data['channel_id'].value_counts()


# In[20]:


data['Firs6ever30'].value_counts(dropna=False)


# In[21]:


data = data[data['Firs6ever30'].isin([0.0, 2.0])]
data.info()


# In[22]:


data['Firs6ever30'] = data['Firs6ever30']/2
data['Firs6ever30'].value_counts(dropna=False)


# In[23]:


data['Firs6ever30'].mean()


# In[24]:


data['year_month'] = data['apply_date_auth'].str[0:7]


# In[25]:


data['year_month'].value_counts(dropna=False)


# In[29]:


tmp = data.groupby(['year_month'])['Firs6ever30'].agg({'count','sum','mean'})
tmp = pd.concat([tmp, pd.DataFrame(tmp.sum(axis=0),columns=['总计']).T], axis=0)
tmp.loc['总计', 'mean'] = tmp.loc['总计', 'sum'] /tmp.loc['总计', 'count']
tmp


# In[30]:


xx = data.groupby(['channel_id'])['Firs6ever30'].agg({'count','sum','mean'})
xx = pd.concat([xx, pd.DataFrame(xx.sum(axis=0),columns=['总计']).T], axis=0)
xx.loc['总计', 'mean'] = xx.loc['总计', 'sum'] /xx.loc['总计', 'count']
xx


# In[31]:


data = data.reset_index(drop=True)
data.info()


# In[32]:


data_copy_loan= data.copy()


# In[33]:


data.rename(columns={'Firs6ever30':'target'},inplace=True)
data.drop(['apply_date_cash','lending_time','Firs3ever15','Firs3ever30','Firs6ever15','year_month','auth_credit_amount'],axis=1,inplace=True)
data.info()


# ## 拆分数据集

# In[37]:


to_drop = list(data.columns)[0:7]
print(to_drop)


# In[38]:


df_var = data.drop(['target'], axis=1)
df_target = data['target']

X = np.array(df_var)
y = np.array(df_target)


# In[40]:


X_train, X_test, y_train, y_test = train_test_split(df_var ,df_target, test_size=0.3, random_state=22, stratify=y)
print(X_train.shape, X_test.shape, y_train.shape, y_test.shape)


# In[41]:


train = pd.merge(X_train, y_train, how='inner', left_index=True, right_index=True)
print(train.shape)


# In[42]:


oot = pd.merge(X_test, y_test, how='inner', left_index=True, right_index=True)
print(oot.shape)


# In[45]:


train.to_csv(r'd:\liuyedao\model_result\{}_model_data_channel_xgb_train_{}.csv'.format('auth', str(datetime.today())[:10].replace('-','')))
oot.to_csv(r'd:\liuyedao\model_result\{}_model_data_channel_xgb_oot_{}.csv'.format('auth', str(datetime.today())[:10].replace('-','')))


# In[3]:


train = pd.read_csv(r'd:\liuyedao\model_result\auth_model_data_channel_xgb_train_20230728.csv')
oot = pd.read_csv(r'd:\liuyedao\model_result\auth_model_data_channel_xgb_oot_20230728.csv')


# # 2.数据探索分析

# In[4]:


train.info()
train.shape


# In[5]:


oot.info()
oot.shape


# In[6]:


to_drop = ['order_no', 'user_id', 'id_no_des', 'channel_id', 'apply_date_auth', 'apply_time', 'auth_status']
cols = ['model_score_01_x_tianchuang', 'model_score_01_fulin', 'model_score_01_baihang', 'model_score_01_rong360', 'model_score_01_tengxun', 'model_score_01_xysl_3']
train_selected = train[to_drop+cols+['target']]
oot_selected = oot[to_drop+cols+['target']]


# In[47]:


train_df_explore = toad.detect(train.drop(to_drop,axis=1))
oot_df_explore = toad.detect(oot.drop(to_drop,axis=1))

train_df_iv = toad.quality(train.drop(to_drop,axis=1),'target',iv_only=True)
oot_df_iv = toad.quality(oot.drop(to_drop,axis=1),'target',iv_only=True)

writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_数据探索性分析_xgb_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
train_df_explore.to_excel(writer,sheet_name='train_df_explore')
oot_df_explore.to_excel(writer,sheet_name='oot_df_explore')
train_df_iv.to_excel(writer,sheet_name='train_df_iv')
oot_df_iv.to_excel(writer,sheet_name='oot_df_iv')
writer.save()


# # 3.特征筛选

# In[189]:


train_selected, dropped = toad.selection.select(train, target='target', empty=0.9, iv=0.001, corr=0.70, return_drop=True, exclude=to_drop)
print(dropped)
print(train_selected.shape)


# In[190]:


train_selected = train_selected.drop(['value_012_pudao_6','model_score_01_moxingfen_14','model_score_01_zr_tongdun','model_score_01_xysl_1'],axis=1)
train_selected.info()
train_selected.head()


# In[191]:


oot_selected =oot[train_selected.columns]
oot_selected.info()


# In[192]:


train_selected.columns[:-1]


# # 4.模型训练和评估

# In[195]:


cols = list(train_selected.columns[7:-1])
print(cols)


# In[204]:


xgb_model = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 500,
                              max_depth = 3,
                              min_child_weight = 1,
                              gamma=0,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)

# 对训练集训练模型
# xgb_model.fit(train_selected[cols], train_selected['target'], early_stopping_rounds=1,eval_metric='auc',
#               eval_set=[(oot_selected[cols], oot_selected['target'])])

xgb_model.fit(train_selected[cols], train_selected['target'])


# In[205]:


plot_importance(xgb_model, importance_type='gain')


# In[206]:


xgb_model.get_booster().get_score(importance_type='gain')


# In[207]:


cols = list(train_selected.columns[7:-1])
print(cols)


# In[208]:


# 对训练集进行预测
y_pred = xgb_model.predict_proba(train_selected[cols])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(train_selected['target'], y_pred, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train KS: ', ks)
print('train AUC: ', roc_auc)

# 对测试集进行预测
y_pred_oot = xgb_model.predict_proba(oot_selected[cols])[:,1]
fpr_oot, tpr_oot, thresholds_oot = metrics.roc_curve(oot_selected['target'], y_pred_oot, pos_label=1)
roc_auc_oot = metrics.auc(fpr_oot, tpr_oot)
ks_oot = max(tpr_oot-fpr_oot)
print('oot KS: ', ks_oot)
print('oot AUC: ', roc_auc_oot)


# In[209]:


plt.plot(fpr, tpr, color='darkorange', lw=2, label='trian ROC curve (area = %0.2f)' % roc_auc)
plt.plot(fpr_oot, tpr_oot, color='blue', lw=2, label='test ROC curve (area = %0.2f)' % roc_auc_oot)
plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="best")
plt.show()


# In[210]:


from sklearn.model_selection import GridSearchCV


# In[211]:


param_test1 = {'max_depth':[3,4,5,6,7,8], 'min_child_weight':[i for i in range(1,6,1)]}

gsearch = GridSearchCV(
    estimator = xgb_model,
    param_grid=param_test1, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch.fit(train_selected[cols], train_selected['target'])

print('gsearch1.best_params_', gsearch.best_params_)
print('gsearch1.best_score_', gsearch.best_score_)


# In[212]:


xgb_model_2 = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 500,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)

param_test2 = {'learning_rate':[i/20.0 for i in range(1,20)]}
gsearch2 = GridSearchCV(
    estimator = xgb_model_2,
    param_grid=param_test2, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch2.fit(train_selected[cols], train_selected['target'])

print('gsearch2.best_params_', gsearch2.best_params_)
print('gsearch2.best_score_', gsearch2.best_score_)


# In[214]:


xgb_model_3 = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 500,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)
param_test3 = {'n_estimators':[100, 200, 300, 400, 500, 600]}

gsearch3 = GridSearchCV(
    estimator = xgb_model_3,
    param_grid=param_test3, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch3.fit(train_selected[cols], train_selected['target'])

print('gsearch3.best_params_', gsearch3.best_params_)
print('gsearch3.best_score_', gsearch3.best_score_)


# In[215]:


xgb_model_4 = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 200,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)
param_test4 = {'gamma':[i/10.0 for i in range(10)]}

gsearch4 = GridSearchCV(
    estimator = xgb_model_4,
    param_grid=param_test4, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch4.fit(train_selected[cols], train_selected['target'])

print('gsearch4.best_params_', gsearch4.best_params_)
print('gsearch4.best_score_', gsearch4.best_score_)


# In[216]:


xgb_model_5 = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 200,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0.7,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)
param_test5 = {'reg_alpha':[0, 0.0001, 0.001, 0.1, 1, 100]}

gsearch5 = GridSearchCV(
    estimator = xgb_model_5,
    param_grid=param_test5, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch5.fit(train_selected[cols], train_selected['target'])

print('gsearch5.best_params_', gsearch5.best_params_)
print('gsearch5.best_score_', gsearch5.best_score_)


# In[217]:


xgb_model_6 = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 200,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0.7,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 1
)
param_test6 = {'reg_lambda':[0, 0.001, 0.1, 1, 100]}

gsearch6 = GridSearchCV(
    estimator = xgb_model_6,
    param_grid=param_test6, 
    scoring='roc_auc', 
    n_jobs=-1, 
    cv=5)
gsearch6.fit(train_selected[cols], train_selected['target'])

print('gsearch6.best_params_', gsearch6.best_params_)
print('gsearch6.best_score_', gsearch6.best_score_)


# In[218]:


xgb_model = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 200,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0.7,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 100
)

# 对训练集训练模型
xgb_model.fit(train_selected[cols], train_selected['target'])


# In[15]:


# 保存模型
import pickle
# save
# pickle.dump(xgb_model, open(r"D:\liuyedao\model_result\auth_xgb_model.pkl", "wb"))


# In[16]:


# load
xgb_model_pkl = pickle.load(open(r"D:\liuyedao\model_result\auth_xgb_model.pkl", "rb"))


# In[17]:


plot_importance(xgb_model_pkl, importance_type='gain')


# In[7]:


from sklearn2pmml.pipeline import PMMLPipeline
from sklearn2pmml import sklearn2pmml


# In[8]:


cols


# In[9]:


xgb_model = xgb.XGBClassifier(booster='gbtree',
                              learning_rate=0.05,
                              n_estimators = 200,
                              max_depth = 3,
                              min_child_weight = 2,
                              gamma=0.7,
                              subsample=1,
                              colsample_bytree=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              n_jobs = -1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_alpha = 0,
                              reg_lambda = 100
)


# In[10]:


pipline_xgb_model = PMMLPipeline([("classfier",xgb_model)])
pipline_xgb_model.fit(train_selected[cols], train_selected['target'])


# In[13]:


sklearn2pmml(pipline_xgb_model, r'D:\liuyedao\model_result\xgb_model_20230810.pmml', with_repr=True)


# In[219]:


# 对训练集进行预测
pred_train = xgb_model.predict_proba(train_selected[cols])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(train_selected['target'], pred_train, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train KS: ', ks)
print('train AUC: ', roc_auc)

# 对测试集进行预测
pred_oot = xgb_model.predict_proba(oot_selected[cols])[:,1]
fpr_oot, tpr_oot, thresholds_oot = metrics.roc_curve(oot_selected['target'], pred_oot, pos_label=1)
roc_auc_oot = metrics.auc(fpr_oot, tpr_oot)
ks_oot = max(tpr_oot-fpr_oot)
print('oot KS: ', ks_oot)
print('oot AUC: ', roc_auc_oot)


# In[249]:


xgb_model.get_booster().get_score(importance_type='gain')


# In[250]:


plot_importance(xgb_model, importance_type='gain')


# In[220]:


train_selected_167 = train_selected.query("channel_id == 167")
oot_selected_167 = oot_selected.query("channel_id == 167")


train_selected_174 = train_selected.query("channel_id == 174")
oot_selected_174 = oot_selected.query("channel_id == 174")


# In[221]:


# 167对训练集进行预测
pred_train_167 = xgb_model.predict_proba(train_selected_167[cols])[:,1]
fpr_167, tpr_167, thresholds_167 = metrics.roc_curve(train_selected_167['target'], pred_train_167, pos_label=1)
roc_auc_167 = metrics.auc(fpr_167, tpr_167)
ks_167 = max(tpr_167-fpr_167)
print('渠道167的train KS: ', ks_167)
print('渠道167的train AUC: ', roc_auc_167)

# 167对测试集进行预测
pred_oot_167 = xgb_model.predict_proba(oot_selected_167[cols])[:,1]
fpr_167_oot, tpr_167_oot, thresholds_oot_167 = metrics.roc_curve(oot_selected_167['target'], pred_oot_167, pos_label=1)
roc_auc_oot_167 = metrics.auc(fpr_167_oot, tpr_167_oot)
ks_oot_167 = max(tpr_167_oot-fpr_167_oot)
print('渠道167的oot KS: ', ks_oot_167)
print('渠道167的oot AUC: ', roc_auc_oot_167)


# 174对训练集进行预测
pred_train_174 = xgb_model.predict_proba(train_selected_174[cols])[:,1]
fpr_174, tpr_174, thresholds_174 = metrics.roc_curve(train_selected_174['target'], pred_train_174, pos_label=1)
roc_auc_174 = metrics.auc(fpr_174, tpr_174)
ks_174 = max(tpr_174-fpr_174)
print('渠道174的train KS: ', ks_174)
print('渠道174的train AUC: ', roc_auc_174)

# 174对测试集进行预测
pred_oot_174 = xgb_model.predict_proba(oot_selected_174[cols])[:,1]
fpr_174_oot, tpr_174_oot, thresholds_oot_174 = metrics.roc_curve(oot_selected_174['target'], pred_oot_174, pos_label=1)
roc_auc_oot_174 = metrics.auc(fpr_174_oot, tpr_174_oot)
ks_oot_174 = max(tpr_174_oot-fpr_174_oot)
print('渠道174的oot KS: ', ks_oot_174)
print('渠道174的oot AUC: ', roc_auc_oot_174)


# In[222]:


# 业务效果
pred_data = toad.metrics.KS_bucket(pred_train, train_selected['target'], bucket=10, method='step')
pred_data


# In[223]:


train_selected_167['prob'] = pred_train_167
oot_selected_167['prob'] = pred_oot_167

train_selected_174['prob'] = pred_train_174
oot_selected_174['prob'] = pred_oot_174

train_selected['prob'] = pred_train
oot_selected['prob'] = pred_oot


# In[224]:


def Prob2Score(prob, base_odds=35, base_score=700, pdo=60, rate=2) :
    # 将概率转化成分数且为正整数
    y = np.log((1 - prob) / prob)
    factor = pdo/np.log(rate)
    offset = base_score - factor * np.log(base_odds)
    score = offset +  factor * (y)
    
    return score


# In[225]:


train_selected['score'] = train_selected['prob'].apply(lambda x:Prob2Score(x))
oot_selected['score'] = oot_selected['prob'].apply(lambda x:Prob2Score(x))

train_selected_167['score'] = train_selected_167['prob'].apply(lambda x:Prob2Score(x))
oot_selected_167['score'] = oot_selected_167['prob'].apply(lambda x:Prob2Score(x))

train_selected_174['score'] = train_selected_174['prob'].apply(lambda x:Prob2Score(x))
oot_selected_174['score'] = oot_selected_174['prob'].apply(lambda x:Prob2Score(x))


# In[226]:


[round(Prob2Score(x),0) for x in pred_data['min']]


# In[232]:


cut_bins = [float('-inf'), 529.0, 542.0, 553.0, 564.0, 577.0, 594.0, 611.0, 633.0, 660.0, float('inf')]

train_selected['bins'] = pd.cut(train_selected['score'], bins=cut_bins, include_lowest=True, right=False)
oot_selected['bins'] = pd.cut(oot_selected['score'], bins=cut_bins, include_lowest=True, right=False)

train_selected_167['bins'] = pd.cut(train_selected_167['score'], bins=cut_bins, include_lowest=True, right=False)
oot_selected_167['bins'] = pd.cut(oot_selected_167['score'], bins=cut_bins, include_lowest=True, right=False)

train_selected_174['bins'] = pd.cut(train_selected_174['score'], bins=cut_bins, include_lowest=True, right=False)
oot_selected_174['bins'] = pd.cut(oot_selected_174['score'], bins=cut_bins, include_lowest=True, right=False)


# In[233]:


def regroup(data, col, target='target'):
    total = data.groupby(col)[target].count()
    bad = data.groupby(col)[target].sum()
    regroup = pd.concat([total, bad],axis=1)
    regroup.columns = ['total', 'bad']
    regroup['good'] = regroup['total'] - regroup['bad']
    regroup['bad_rate'] = regroup['bad']/regroup['total']
    regroup['total_pct'] = regroup['total']/regroup['total'].sum()
    regroup['varsname'] = col
    regroup['bins'] = regroup.index
    cols = ['varsname','bins','bad','good','total','bad_rate','total_pct']
    regroup = regroup[cols]
    return regroup


# In[234]:


train_selected_regroup = regroup(train_selected, 'bins')
oot_selected_regroup = regroup(oot_selected, 'bins')
train_selected_167_regroup = regroup(train_selected_167, 'bins')
oot_selected_167_regroup = regroup(oot_selected_167, 'bins')
train_selected_174_regroup = regroup(train_selected_174, 'bins')
oot_selected_174_regroup = regroup(oot_selected_174, 'bins')


# In[235]:


writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_业务效果_3_xgb_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
train_selected_regroup.to_excel(writer,sheet_name='train')
oot_selected_regroup.to_excel(writer,sheet_name='test')
train_selected_167_regroup.to_excel(writer,sheet_name='train_167')
oot_selected_167_regroup.to_excel(writer,sheet_name='test_167')
train_selected_174_regroup.to_excel(writer,sheet_name='train_174')
oot_selected_174_regroup.to_excel(writer,sheet_name='test_174')
writer.save()


# In[247]:


def cal_psi(exp, act):
    psi = []
    for i in range(len(exp)):
        psi_i = (act[i] - exp[i])*np.log(act[i]/exp[i])
        psi.append(psi_i)
    return sum(psi)


# In[248]:


print(cal_psi(train_selected_regroup['total_pct'], oot_selected_regroup['total_pct'])) 
print(cal_psi(train_selected_167_regroup['total_pct'], oot_selected_167_regroup['total_pct'])) 
print(cal_psi(train_selected_174_regroup['total_pct'], oot_selected_174_regroup['total_pct'])) 


# In[236]:


def cal_auth(df_auth, channel=None):
    if channel:
        df_auth = df_auth.query("channel_id==@channel")
        
    total_lend = df_auth.groupby('bins')['order_no','Firs6ever30'].count()
    tg_jj = df_auth.groupby(['bins', 'auth_status'])['order_no'].count().unstack()
    lend =  df_auth.groupby(['bins', 'Firs6ever30'])['order_no'].count().unstack()
    result = pd.concat([total_lend, tg_jj, lend], axis=1)
    result.columns = ['授信申请','放款人数','通过人数','拒绝人数','好','灰','坏']
    
    result_sum = pd.DataFrame(result.sum(axis=0),columns=['total']).T
    
    for col in result.columns:
        result['{}占比'.format(col)] = result[col]/result[col].sum()
        result['{}累计占比'.format(col)] = result['{}占比'.format(col)].cumsum()
        
    result['坏客率'] = result['坏']/(result['好']+result['坏'])
    result['通过率'] = result['通过人数']/result['授信申请']
    result = pd.concat([result, result_sum],axis=0)
    
    cols = ['授信申请','授信申请占比','授信申请累计占比','通过人数','通过率','拒绝人数','拒绝人数占比','拒绝人数累计占比',
            '通过人数占比','通过人数累计占比','放款人数','放款人数占比','放款人数累计占比','好','好占比','好累计占比',
            '坏','坏占比','坏累计占比','坏客率','灰','灰占比','灰累计占比']
    result = result[cols]
    
    return result


# In[238]:


auth_data = data_copy[list(train_selected.columns[:-4])+['Firs6ever30']]


# In[239]:


auth_data.columns


# In[240]:


cols = list(auth_data.columns[7:-1])
print(cols)


# In[241]:


cut_bins = [float('-inf'), 529.0, 542.0, 553.0, 564.0, 577.0, 594.0, 611.0, 633.0, 660.0, float('inf')]

auth_data['prob'] = xgb_model.predict_proba(auth_data[cols])[:,1]
auth_data['score'] = auth_data['prob'].apply(lambda x:Prob2Score(x))
auth_data['bins'] = pd.cut(auth_data['score'], bins=cut_bins, include_lowest=True, right=False)


# In[242]:


result_auth = cal_auth(auth_data, channel=None)
result_auth_167 = cal_auth(auth_data, channel=167)
result_auth_174 = cal_auth(auth_data, channel=174)


# In[243]:


print(toad.metrics.PSI(pred_train, pred_oot))
print(toad.metrics.PSI(pred_train_167, pred_oot_167))
print(toad.metrics.PSI(pred_train_174, pred_oot_174))


# In[244]:


writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_业务效果_授信申请_xgb_{}_v{}.xlsx".format(str(datetime.today())[:10].replace('-',''),2))
result_auth.to_excel(writer,sheet_name='all')
result_auth_167.to_excel(writer,sheet_name='167')
result_auth_174.to_excel(writer,sheet_name='174')
writer.save()


# In[245]:


auth_data.info(show_counts=True)


# In[246]:


auth_data.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_score_20230807.csv",index=False)


# # oot数据

# In[4]:


from pypmml import Model


# In[35]:


auth_data = pd.read_csv(r'd:\liuyedao\model_result\oot_data_20230815.csv')
auth_data.info()


# In[36]:


import pickle
# load
xgb_model_pkl = pickle.load(open(r"D:\liuyedao\model_result\auth_xgb_model.pkl", "rb"))
xgb_model_pkl.get_booster().get_score(importance_type='gain')


# In[37]:


model_cols = list(xgb_model_pkl.get_booster().get_score(importance_type='gain').keys())
model_cols


# In[12]:


def Prob2Score(prob, base_odds=35, base_score=700, pdo=60, rate=2) :
    # 将概率转化成分数且为正整数
    y = np.log((1 - prob) / prob)
    factor = pdo/np.log(rate)
    offset = base_score - factor * np.log(base_odds)
    score = offset +  factor * (y)
    
    return score


auth_data['score_xgb'] = auth_data['prob'].apply(lambda x:Prob2Score(x))


# In[22]:


usecols = ['order_no','user_id','id_no_des','channel_id','apply_date_auth','apply_time','auth_status'] + cols + ['Firs6ever30', 'score_xgb']
usecols


# In[17]:


auth_data = auth_data[usecols]
auth_data.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_score_20230815_oot.csv",index=False)


# In[18]:


auth_data.info(show_counts=True)


# In[38]:


auth_data['apply_month'] = auth_data['apply_date_auth'].str[0:7]
auth_data['apply_month'].value_counts()


# In[26]:


auth_data['Firs6ever30'].value_counts()


# In[50]:


# 对训练集进行预测
auth_data_mob6 = auth_data[auth_data['Firs6ever30'].isin([0.0, 2.0])][model_cols + ['Firs6ever30','apply_month','order_no']]
auth_data_mob6.groupby(['apply_month', 'Firs6ever30'])['order_no'].count().unstack()


# In[51]:


auth_data_mob6 = auth_data_mob6[auth_data_mob6['apply_month']=='2023-01']
auth_data_mob6['target'] = auth_data_mob6['Firs6ever30'] /2.0
auth_data_mob6.info()


# In[52]:


auth_data_mob6.groupby(['apply_month', 'Firs6ever30'])['order_no'].count().unstack()


# In[53]:


model_cols


# In[54]:


auth_data_mob6['prob'] = xgb_model_pkl.predict_proba(auth_data_mob6[model_cols])[:,1]

fpr, tpr, thresholds = metrics.roc_curve(auth_data_mob6['target'], auth_data_mob6['prob'], pos_label=1)
auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train AUC: ', auc)
print('train KS: ', ks)


# In[55]:


auth_data_mob3 = auth_data[auth_data['Firs3ever30'].isin([0.0, 2.0])][model_cols + ['Firs3ever30','apply_month','order_no']]
auth_data_mob3.groupby(['apply_month', 'Firs3ever30'])['order_no'].count().unstack()


# In[56]:


auth_data_mob3 = auth_data_mob3[auth_data_mob3['apply_month']!='2023-05']
auth_data_mob3['target'] = auth_data_mob3['Firs3ever30'] /2.0
auth_data_mob3.info()


# In[57]:


auth_data_mob3.groupby(['apply_month', 'Firs3ever30'])['order_no'].count().unstack()


# In[58]:


# 对训练集进行预测
auth_data_mob3['prob'] = xgb_model_pkl.predict_proba(auth_data_mob3[model_cols])[:,1]

fpr, tpr, thresholds = metrics.roc_curve(auth_data_mob3['target'], auth_data_mob3['prob'], pos_label=1)
auc = metrics.auc(fpr, tpr)
ks = max(tpr-fpr)
print('train AUC: ', auc)
print('train KS: ', ks)


# In[59]:


for i in list(auth_data_mob3['apply_month'].unique()):
    print(i)
    tmp = auth_data_mob3[auth_data_mob3['apply_month']==i]
    tmp['prob'] = xgb_model_pkl.predict_proba(tmp[model_cols])[:,1]
    fpr, tpr, thresholds = metrics.roc_curve(tmp['target'], tmp['prob'], pos_label=1)
    auc = metrics.auc(fpr, tpr)
    ks = max(tpr-fpr)
    print('train AUC: ', auc)
    print('train KS: ', ks)
    print('-----------------------------------')




#==============================================================================
# File: 三方数据评分卡建模-分渠道.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[3]:


import pandas as pd
import numpy as np
import toad
import os 
from datetime import datetime


# In[4]:


os.getcwd()


# # 1.读取数据集

# In[5]:


# data = pd.read_csv(r'd:\liuyedao\model_result\model_data.csv') 
# data = pd.read_csv(r'd:\liuyedao\model_result\model_data_20230721.csv') 
data = pd.read_csv(r'd:\liuyedao\model_result\{}_model_data_channel_{}.csv'.format('auth', str(datetime.today())[:10].replace('-','')))
print('数据大小：', data.shape)
data.head(10)


# In[6]:


data.info()


# In[9]:


# data['Firs6ever30'].value_counts(dropna=False)


# In[10]:


# data = data[data['Firs6ever30'].isin([0.0, 2.0])]
# data.info()


# In[7]:


data['Firs6ever30'] = data['Firs6ever30']/2
data['Firs6ever30'].value_counts(dropna=False)


# In[8]:


data['Firs6ever30'].mean()


# In[9]:


data['year_month'].value_counts(dropna=False)


# In[10]:


print('month: ', data.year_month.nunique())


# In[11]:


data = data.reset_index(drop=True)
data.info()


# In[12]:


data_copy = data.copy()


# In[13]:


data.rename(columns={'Firs6ever30':'target'},inplace=True)
data.drop(['apply_date_cash','lending_time','Firs3ever15','Firs3ever30','Firs6ever15'],axis=1,inplace=True)
data.info()


# In[14]:


data.drop(['auth_credit_amount'],axis=1,inplace=True)


# In[115]:


# train = data[(data["apply_date_auth"]>='2022-05-01') & (data["apply_date_auth"]<'2022-12-01')]
# oot = data[data["apply_date_auth"]>='2022-12-01']
# print('train size: ', train.shape, '\noot size: ', oot.shape)


# # 2.数据探索分析

# In[18]:


df_explore = pd.DataFrame()
for channel, tmp_df in data.groupby('channel_id'):
    tmp = toad.detect(tmp_df)
    tmp['varsname'] = channel
    df_explore = pd.concat([df_explore, tmp], axis=0)
    


# In[19]:


df_explore.shape


# In[11]:


to_drop = list(data.columns)[0:7]
print(to_drop)


# In[22]:


df_iv = pd.DataFrame()
for channel, tmp_df in data.groupby('channel_id'):
    tmp_iv = toad.quality(tmp_df.drop(to_drop, axis=1),'target',iv_only=True)
    tmp_iv['varsname'] = channel
    df_iv = pd.concat([df_iv, tmp_iv], axis=0)
print(df_iv.shape)


# In[35]:


df_explore_iv = pd.DataFrame()
for channel, tmp_df in data.groupby('channel_id'):
    tmp = toad.detect(tmp_df)
    tmp_iv = toad.quality(tmp_df.drop(to_drop, axis=1),'target',iv_only=True)
    tmp = pd.merge(tmp, tmp_iv, left_index=True, right_index=True)
    tmp['varsname'] = channel
    df_explore_iv = pd.concat([df_explore_iv, tmp], axis=0)
df_explore_iv = df_explore_iv.reset_index()
df_explore_iv.to_excel(r'D:\liuyedao\mid_result\数据探索性分析.xlsx')


# In[44]:


from datetime import datetime
writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\数据探索性分析_建模_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
df_explore_iv.to_excel(writer,sheet_name='df_explore_iv')
df_explore.to_excel(writer,sheet_name='df_explore')
df_iv.to_excel(writer,sheet_name='df_iv')
writer.save()


# # 3.特征筛选

# In[31]:


vars_list = []
for channel, tmp_df in data.groupby('channel_id'):
    print('-----------渠道：{}-------------------'.format(channel))
    tmp = toad.selection.select(tmp_df, target='target', empty=0.9, iv=0.01, corr=0.75, return_drop=True, exclude=to_drop)
    var = 'data_selected_{}'.format(channel)
    globals()[var] = tmp[0]
    dropped = tmp[1]
    vars_list.append(globals()[var])
    print(dropped)
    print(globals()[var].shape)


# In[32]:


for var in vars_list:
    var.info()


# In[ ]:


data_selected_167.info()


# # 4. 变量分箱

# In[36]:


data_selected_167.drop(to_drop,axis=1).describe().T


# In[39]:


data_selected_167[data_selected_167['model_score_01_moxingfen_14']<=0]['model_score_01_moxingfen_14'].value_counts()


# In[40]:


data_selected_167[data_selected_167['model_score_01_rong360']<=0]['model_score_01_rong360'].value_counts()


# In[46]:


data_selected_167[data_selected_167['model_score_01_x_tianchuang']<=0]['model_score_01_x_tianchuang'].value_counts()


# In[47]:


for col in data_selected_167.drop(to_drop,axis=1).columns:
    data_selected_167[col].fillna(-999.0, inplace=True)


# In[48]:


data_selected_167.drop(to_drop,axis=1).describe().T


# In[49]:


data_selected_174.drop(to_drop,axis=1).describe().T


# In[50]:


for col in data_selected_174.drop(to_drop,axis=1).columns:
    data_selected_174[col].fillna(-999.0, inplace=True)


# In[51]:


data_selected_206.drop(to_drop,axis=1).describe().T


# In[52]:


for col in data_selected_206.drop(to_drop,axis=1).columns:
    data_selected_206[col].fillna(-999.0, inplace=True)


# In[53]:


data_selected_777.drop(to_drop,axis=1).describe().T


# In[54]:


for col in data_selected_777.drop(to_drop,axis=1).columns:
    data_selected_777[col].fillna(-999.0, inplace=True)


# In[55]:


# # 开始分箱
# c = toad.transform.Combiner()
# c.fit(train_selected.drop(to_drop,axis=1), y='target', method='chi', min_samples=0.05, n_bins=None, empty_separate=False) 
# bins_result = c.export()
# bins_result


# In[56]:


# for col in train_selected.drop(to_drop,axis=1).columns[:-1]:
#     print(col,':  ', c.export()[col])


# In[57]:


# 开始分箱
combiner_list = []
bins_list = []
for channel, tmp_df in zip([167, 174, 206, 777], vars_list):
    var = 'combiner_{}'.format(channel)
    globals()[var] = toad.transform.Combiner()
    globals()[var].fit(tmp_df.drop(to_drop,axis=1), y='target', method='chi', min_samples=0.05, n_bins=None, empty_separate=False) 
    combiner_list.append(globals()[var])
    
    var_bins = 'bins_{}'.format(channel)
    globals()[var_bins] = globals()[var].export()
    bins_list.append(globals()[var_bins])


# In[66]:


i = 0
for channel, tmp_df in zip([167, 174, 206, 777], vars_list):
    print('---------------渠道：{}---------------'.format(channel))
    for col in tmp_df.drop(to_drop,axis=1).columns[:-1]:
        print(col,':  ', bins_list[i][col])
    i = i + 1


# In[87]:


def regroup(data_bins, to_drop, target='target'):    
    result = pd.DataFrame()
    for col in data_bins.drop(to_drop, axis=1).columns[:-1]:
        print(col)
        total = data_bins.groupby(col)[target].count()
        bad = data_bins.groupby(col)[target].sum()
        regroup = pd.concat([total, bad],axis=1)
        regroup.columns = ['total', 'bad']
        regroup['good'] = regroup['total'] - regroup['bad']
        regroup['bad_rate'] = regroup['bad']/regroup['total']
        regroup['total_pct'] = regroup['total']/regroup['total'].sum()
        regroup['bad_pct'] = regroup['bad']/regroup['bad'].sum()
        regroup['good_pct'] = regroup['good']/regroup['good'].sum()
        regroup['bad_pct_cum'] = regroup['bad_pct'].cumsum()
        regroup['goo_pct_cum'] = regroup['good_pct'].cumsum()
        regroup['ks'] = regroup['bad_pct_cum'] - regroup['goo_pct_cum']
        regroup['ks_max'] = regroup['ks'].max()
        regroup['iv_bins'] = (regroup['bad_pct']-regroup['good_pct']) * np.log(regroup['bad_pct']/regroup['good_pct'])
        regroup['iv'] = regroup['iv_bins'].sum()
        regroup['varsname'] = col
        result = pd.concat([result, regroup], axis=0)
    
    return result


# In[67]:


# 分箱的分割点
bins_167


# In[71]:


# 调整分箱
adj_bins = {'model_score_01_x_tianchuang': [0.0, 591.0, 681.0], 'model_score_01_baihang': [488.0, 700.0, 720.0, 764.0]}
combiner_167.update(adj_bins)


# In[89]:


# 二次调整分箱
adj_bins = {'model_score_01_x_tianchuang': [0.0, 591.0], 
            'model_score_01_zr_tongdun': [630.0, 718.0, 784.0, 834.0]}
combiner_167.update(adj_bins)


# In[90]:


# 重新查看分箱的分割点
combiner_167.export()


# In[91]:


# 分箱转化
data_selected_bins_167 = combiner_167.transform(data_selected_167, labels=True)
data_selected_bins_167.head()


# In[74]:


data_selected_bins = []
data_selected_bins.append(data_selected_bins_167)


# In[78]:


df_result_167 = regroup(data_selected_bins_167, to_drop, target='target')
df_result_167.head(3)


# In[79]:


# 分箱的分割点
bins_174


# In[81]:


# 调整分箱
adj_bins = {'model_score_01_x_tianchuang': [0.0, 513.0, 549.0, 588.0, 656.0], 
            'model_score_01_zr_tongdun': [0.0, 657.0, 752.0, 814.0],
            'model_score_01_moxingfen_14': [0.0, 558.0],
            'value_012_pudao_6': [1.0, 4.0],
            'model_score_01_xysl_3': [0.0, 571.0, 634.0, 676.0, 698.0]}
combiner_174.update(adj_bins)

# 重新查看分箱的分割点
combiner_174.export()

# 分箱转化
data_selected_bins_174 = combiner_174.transform(data_selected_174, labels=True)
data_selected_bins.append(data_selected_bins_174)

# 计算iv ks
df_result_174 = regroup(data_selected_bins_174, to_drop, target='target')
df_result_174.head(3)


# In[92]:


# 二次调整分箱
adj_bins = {'model_score_01_x_tianchuang': [0.0, 588.0, 656.0]}
combiner_174.update(adj_bins)
# 分箱转化
data_selected_bins_174 = combiner_174.transform(data_selected_174, labels=True)
data_selected_bins.append(data_selected_bins_174)


# In[82]:


# 分箱的分割点
bins_206


# In[ ]:





# In[83]:


# 调整分箱
adj_bins = {'model_score_01_x_tianchuang': [0.0, 436.0, 565.0, 672.0], 
            'model_score_01_pudao_8': [0.0, 586.0]}
combiner_206.update(adj_bins)

# 重新查看分箱的分割点
combiner_206.export()

# 分箱转化
data_selected_bins_206 = combiner_206.transform(data_selected_206, labels=True)
data_selected_bins.append(data_selected_bins_206)

# 计算iv ks
df_result_206 = regroup(data_selected_bins_206, to_drop, target='target')
df_result_206.head(3)


# In[84]:


# 分箱的分割点
bins_777


# In[85]:


# 分箱转化
data_selected_bins_777 = combiner_777.transform(data_selected_777, labels=True)
data_selected_bins.append(data_selected_bins_777)

# 计算iv ks
df_result_777 = regroup(data_selected_bins_777, to_drop, target='target')
df_result_777.head(3)


# In[86]:


from datetime import datetime
writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\分箱结果_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
df_result_167.to_excel(writer,sheet_name='167')
df_result_174.to_excel(writer,sheet_name='174')
df_result_206.to_excel(writer,sheet_name='206')
df_result_777.to_excel(writer,sheet_name='777')
writer.save()


# In[88]:


writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\分箱结果_v2_"+str(datetime.today())[:10].replace('-','')+'.xlsx')

for i,j in zip([167, 174, 206, 777], data_selected_bins):
    print(i)
    result = regroup(j, to_drop, target='target')
    result.to_excel(writer,sheet_name=str(i))

writer.save()    


# In[ ]:


# c.export()
# c.load(dict)
# c.transform(dataframe, labels=False)


# ### 观察分箱并调整

# In[138]:


from toad.plot import bin_plot
# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[0]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[155]:


# 调整分箱
col = train_selected.drop(to_drop,axis=1).columns[0]
adj_bins = {col: [550, 675]}

c.update(adj_bins)
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[143]:


from toad.plot import badrate_plot
# 查看时间内和跨时间分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[0]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(data[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[144]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[1]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[145]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[1]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[146]:


from toad.plot import bin_plot
# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[2]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[148]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[2]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[147]:


from toad.plot import bin_plot
# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[3]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[150]:


# 调整分箱
col = train_selected.drop(to_drop,axis=1).columns[3]
adj_bins = {col: [638, 723, 757]}

c.update(adj_bins)
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[151]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[3]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[152]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[4]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[154]:


# 调整分箱
col = train_selected.drop(to_drop,axis=1).columns[4]
adj_bins = {col: [0.01428]}

c.update(adj_bins)
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[156]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[4]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[157]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[5]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[158]:


# 调整分箱
col = train_selected.drop(to_drop,axis=1).columns[5]
adj_bins = {col: [21]}

c.update(adj_bins)
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[159]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[5]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[160]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[6]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[161]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[6]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# In[162]:


# 静态观察
col = train_selected.drop(to_drop,axis=1).columns[7]
print('========{}：变量的分箱图========='.format(col))
bin_plot(c.transform(train_selected[[col, 'target']], labels=True), x=col, target='target')
bin_plot(c.transform(oot[[col, 'target']], labels=True), x=col, target='target')


# In[163]:


from toad.plot import badrate_plot
# 看时间内的分箱稳定性
col = train_selected.drop(to_drop,axis=1).columns[7]
badrate_plot(c.transform(train_selected[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)
# badrate_plot(c.transform(oot[[col, 'target','year_month']], labels=True), x='year_month', target='target', by=col)


# # 5.WOE转换

# In[94]:


transer_174 = toad.transform.WOETransformer()
data_selected_174_woe = transer_174.fit_transform(combiner_174.transform(data_selected_174), data_selected_174['target'], exclude=to_drop+['target'])


# In[111]:


data_selected_174_woe_selected, dropped = toad.selection.select(data_selected_174_woe, target='target', empty=0.9, iv=0.01, corr=0.75, return_drop=True, exclude=to_drop)
print(dropped)
print(data_selected_174_woe_selected.shape)


# In[112]:


data_selected_174_woe_selected.info()


# In[113]:


transer_167 = toad.transform.WOETransformer()
data_selected_167_woe = transer_167.fit_transform(combiner_167.transform(data_selected_167), data_selected_167['target'], exclude=to_drop+['target'])


# In[114]:


data_selected_167_woe_selected, dropped_167 = toad.selection.select(data_selected_167_woe, target='target', empty=0.9, iv=0.01, corr=0.75, return_drop=True, exclude=to_drop)
print(dropped_167)
print(data_selected_167_woe_selected.shape)


# In[116]:


data_selected_167_woe_selected.info()


# # 6.逐步回归

# In[117]:


final_data_174 = toad.selection.stepwise(data_selected_174_woe_selected, target='target', estimator='ols', direction='both', criterion='aic', exclude=to_drop)
final_data_174.info()


# In[119]:


cols_174 = list(final_data_174.drop(to_drop+['target'], axis=1).columns)
cols_174


# In[120]:


final_data_167 = toad.selection.stepwise(data_selected_167_woe_selected, target='target', estimator='ols', direction='both', criterion='aic', exclude=to_drop)
final_data_167.info()


# In[124]:


cols_167 = list(final_data_167.drop(to_drop+['target'], axis=1).columns)
cols_167


# # 7.建模和评估

# In[125]:


from sklearn.linear_model import LogisticRegression


# In[126]:


lr_174 = LogisticRegression()
lr_174.fit(final_data_174[cols_174], final_data_174['target'])


# In[127]:


lr_167 = LogisticRegression()
lr_167.fit(final_data_167[cols_167], final_data_167['target'])


# In[129]:


# 预测训练和隔月的oot
pred_train_174 = lr_174.predict_proba(final_data_174[cols_174])[:,1]
pred_train_167 = lr_167.predict_proba(final_data_167[cols_167])[:,1]
# pred_oot_may = lr.predict_proba(final_oot.loc[final_oot.year_month=='2022-11', col])[:,1]
# pred_oot_june = lr.predict_proba(final_oot.loc[final_oot.year_month=='2022-12', cols])[:,1]


# In[131]:


# KS/AUC
from toad.metrics import KS,AUC

print('渠道174的train KS: ', KS(pred_train_174, final_data_174['target']))
print('渠道174的train AUC: ', AUC(pred_train_174, final_data_174['target']))

print('渠道167的train KS: ', KS(pred_train_167, final_data_167['target']))
print('渠道167的train AUC: ', AUC(pred_train_167, final_data_167['target']))
# print('-------------oot结果--------------------')
# print('5月 KS', KS(pred_oot_may, final_oot.loc[final_oot.year_month=='2022-11', 'target']))
# print('6月 KS', KS(pred_oot_june, final_oot.loc[final_oot.year_month=='2022-12', 'target']))


# In[181]:


# PSI
# print(toad.metrics.PSI(pred_train, pred_oot_may))
# print(toad.metrics.PSI(pred_train, pred_oot_june))


# In[132]:


# 业务效果
KS_bucket_174 = toad.metrics.KS_bucket(pred_train_174, final_data_174['target'], bucket=10, method='quantile')
KS_bucket_167 = toad.metrics.KS_bucket(pred_train_167, final_data_167['target'], bucket=10, method='quantile')


# # 9.转换评分

# In[134]:


card_174 = toad.ScoreCard(combiner=combiner_174,
                      transer = transer_174,
#                       class_weight = 'balanced',
#                       C=0.1,
                      base_odds=20,
                      base_score=750,
                      pdo=60,
                      rate=2
                     )
card_174.fit(final_data_174[cols_174], final_data_174['target'])


# In[135]:


vars_card_174 = card_174.export(to_frame=True)
vars_card_174.head()


# In[ ]:





# In[136]:


data_selected_174['score'] = pd.DataFrame(card_174.predict(data_selected_174), index=data_selected_174.index,columns=['score'])
data_selected_174['bins'] = pd.cut(data_selected_174['score'], 10)
data_selected_174[['score','bins']].info()
data_selected_174[['score','bins']].head()


# In[139]:


card_167 = toad.ScoreCard(combiner=combiner_167,
                      transer = transer_167,
#                       class_weight = 'balanced',
#                       C=0.1,
                      base_odds=20,
                      base_score=750,
                      pdo=60,
                      rate=2
                     )
card_167.fit(final_data_167[cols_167], final_data_167['target'])


# In[140]:


vars_card_167 = card_167.export(to_frame=True)
vars_card_167.head()


# In[141]:


data_selected_167['score'] = pd.DataFrame(card_167.predict(data_selected_167), index=data_selected_167.index,columns=['score'])
data_selected_167['bins'] = pd.cut(data_selected_167['score'], 10)
data_selected_167[['score','bins']].info()
data_selected_167[['score','bins']].head()


# In[152]:


from datetime import datetime
writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\评分卡_"+str(datetime.today())[:10].replace('-','')+'.xlsx')
vars_card_167.to_excel(writer,sheet_name='变量得分167')
vars_card_174.to_excel(writer,sheet_name='变量得分174')
KS_bucket_167.to_excel(writer,sheet_name='业务效果167')
KS_bucket_174.to_excel(writer,sheet_name='业务效果174')
writer.save()


# # xgboost评分卡

# In[142]:


import xgboost as xgb
from sklearn import metrics


# In[147]:


xgb_model_174 = xgb.XGBClassifier(learning_rate=0.05,
                              n_estimators = 500,
                              max_depth = 3,
                              n_jobs = -1,
                              min_child_weight = 1,
                              subsample=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_lambda = 300
)


# In[148]:


xgb_model_174.fit(final_data_174[cols_174], final_data_174['target'])


# In[150]:


y_pred_174 = xgb_model_174.predict_proba(final_data_174[cols_174])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(final_data_174['target'], y_pred_174, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = abs(max(tpr-fpr))
print('渠道174的train KS: ', ks)
print('渠道174的train AUC: ', roc_auc)

# print('渠道167的train KS: ', KS(pred_train_167, final_data_167['target']))
# print('渠道167的train AUC: ', AUC(pred_train_167, final_data_167['target']))


# In[151]:


xgb_model_167 = xgb.XGBClassifier(learning_rate=0.05,
                              n_estimators = 500,
                              max_depth = 3,
                              n_jobs = -1,
                              min_child_weight = 1,
                              subsample=1,
                              objective = "binary:logistic",
                              nthread = 1,
                              random_state = 1,
                              scale_pos_weight = 1,
                              reg_lambda = 300
)


xgb_model_167.fit(final_data_167[cols_167], final_data_167['target'])

y_pred_167 = xgb_model_167.predict_proba(final_data_167[cols_167])[:,1]
fpr, tpr, thresholds = metrics.roc_curve(final_data_167['target'], y_pred_167, pos_label=1)
roc_auc = metrics.auc(fpr, tpr)
ks = abs(max(tpr-fpr))
print('渠道167的train KS: ', ks)
print('渠道167的train AUC: ', roc_auc)


# In[162]:


print(cols_167)
xgb_model_167.feature_importances_


# In[163]:


print(cols_174)
xgb_model_174.feature_importances_




#==============================================================================
# File: 临时数据需求-Copy1.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np


# In[2]:


df_lr = pd.read_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_LR_score_20230807.csv")
df_xgb = pd.read_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_score_20230807.csv")


# In[3]:


df_lr.info(show_counts=True)


# In[4]:


df_xgb.info(show_counts=True)


# In[5]:


cols = list(df_xgb.columns[0:7])
print(cols)


# In[6]:


df = pd.merge(df_xgb, df_lr[['order_no', 'score', 'bins']], how='inner', on=['order_no'])
df.info(show_counts=True)
df.head()


# In[7]:


df.rename(columns={'score_x':'score_xgb','bins_x':'bins_xgb','score_y':'score_lr','bins_y':'bins_lr','Firs6ever30':'target','apply_date_auth':'apply_date'}, inplace=True)


# In[8]:


df.info(show_counts=True)
df.head()


# In[9]:


df['target'] = df['target']/2


# In[10]:


df['target'].value_counts(dropna=False)


# In[11]:


df['score_xgb'] = df['score_xgb'].round()
df['score_xgb'].head()


# In[60]:


# df['score_lr'] = df['score_lr'].round()
# df['score_lr'].head()


# In[61]:


# cut_bins = [float('-inf'), 529.0, 542.0, 553.0, 564.0, 577.0, 594.0, 611.0, 633.0, 660.0, float('inf')]
# df['bins_xgb'] = pd.cut(df['score_xgb'], cut_bins)
# print(df['bins_xgb'].unique())


# In[12]:


tmp = df[df['target'].isin([0.0, 1.0])].groupby(['bins_lr','bins_xgb'])['order_no'].count().unstack()
tmp['total'] = tmp.sum(axis=1)
tmp.loc['total',:] = tmp.sum(axis=0)
tmp


# In[13]:


tmp_bad = df[df['target'].isin([1.0])].groupby(['bins_lr','bins_xgb'])['order_no'].count().unstack()
tmp_bad['total'] = tmp_bad.sum(axis=1)
tmp_bad.loc['total',:] = tmp_bad.sum(axis=0)
tmp_bad


# In[14]:


from datetime import datetime
writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_lr_xgb_矩阵_{}_v{}.xlsx".format(str(datetime.today())[:10].replace('-',''),1))
tmp_bad.to_excel(writer,sheet_name='bad')
tmp.to_excel(writer,sheet_name='total')
writer.save()


# In[15]:


df.drop(['prob'],axis=1,inplace=True)


# In[16]:


df.drop(['target','bins_xgb', 'bins_lr'],axis=1,inplace=True)


# In[17]:


df.info(show_counts=True)


# In[18]:


df.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_lr_score_v2_20230807.csv",index=False)


# In[ ]:


df.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_lr_score_v2_20230807.csv",index=False)




#==============================================================================
# File: 临时数据需求-oot.py
#==============================================================================

#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np


# In[3]:


df_xgb = pd.read_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_score_20230815_oot.csv")
df_lr = pd.read_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_LR_score_20230815_oot.csv")


# In[4]:


df_lr.info(show_counts=True)


# In[5]:


df_xgb.info(show_counts=True)


# In[7]:


cols = list(df_lr.columns[0:7])
print(cols)


# In[9]:


cols = list(df_xgb.columns[0:7])
print(cols)


# In[10]:


df = pd.merge(df_lr, df_xgb[['order_no']+ list(df_xgb.columns[7:])], how='inner', on=['order_no'])
df.info(show_counts=True)
df.head()


# In[11]:


df['Firs6ever30'].value_counts(dropna=False)


# In[12]:


df['Firs6ever30'] = df['Firs6ever30']/2


# In[13]:


df['Firs6ever30'].value_counts(dropna=False)


# In[14]:


df['score_xgb'] = df['score_xgb'].round()
df['score_xgb'].head()


# In[15]:


df.rename(columns={'Firs6ever30':'target'},inplace=True)


# In[60]:


# df['score_lr'] = df['score_lr'].round()
# df['score_lr'].head()


# In[61]:


# cut_bins = [float('-inf'), 529.0, 542.0, 553.0, 564.0, 577.0, 594.0, 611.0, 633.0, 660.0, float('inf')]
# df['bins_xgb'] = pd.cut(df['score_xgb'], cut_bins)
# print(df['bins_xgb'].unique())


# In[12]:


tmp = df[df['target'].isin([0.0, 1.0])].groupby(['bins_lr','bins_xgb'])['order_no'].count().unstack()
tmp['total'] = tmp.sum(axis=1)
tmp.loc['total',:] = tmp.sum(axis=0)
tmp


# In[13]:


tmp_bad = df[df['target'].isin([1.0])].groupby(['bins_lr','bins_xgb'])['order_no'].count().unstack()
tmp_bad['total'] = tmp_bad.sum(axis=1)
tmp_bad.loc['total',:] = tmp_bad.sum(axis=0)
tmp_bad


# In[14]:


from datetime import datetime
writer=pd.ExcelWriter(r"d:\liuyedao\mid_result\auth_建模_评分卡_lr_xgb_矩阵_{}_v{}.xlsx".format(str(datetime.today())[:10].replace('-',''),1))
tmp_bad.to_excel(writer,sheet_name='bad')
tmp.to_excel(writer,sheet_name='total')
writer.save()


# In[15]:


df.drop(['prob'],axis=1,inplace=True)


# In[16]:


df.info(show_counts=True)


# In[17]:


df.drop(['target'],axis=1,inplace=True)


# In[18]:


df.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_lr_score_v2_20230815_oot.csv",index=False)


# In[ ]:


df.to_csv(r"d:\liuyedao\mid_result\auth_建模_评分卡_xgb_lr_score_v2_20230807.csv",index=False)




#==============================================================================
# End of batch 3
#==============================================================================
